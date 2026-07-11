import os
import sys
import re
import json
import urllib.request
import threading
import platform
import getpass
import uuid
import traceback
import importlib
import subprocess
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.policy import SMTP
from email import message_from_bytes
from pathlib import Path
from tkinter import messagebox, filedialog
import tkinter as tk

# ------------------------------------------------------------------
# Dependency check
# ------------------------------------------------------------------
REQUIRED_PACKAGES = {
    "customtkinter": "customtkinter",
    "openpyxl":      "openpyxl",
    "keyring":       "keyring",
    "PIL":           "Pillow",
}

def _check_and_install_dependencies():
    missing = []
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing.append((module_name, pip_name))
    if not missing:
        return True
    missing_names = [f"  - {pip_name}" for _, pip_name in missing]
    msg = ("The following packages are required but not installed:\n\n"
           + "\n".join(missing_names)
           + "\n\nWould you like to install them now via pip?")
    root = tk.Tk()
    root.withdraw()
    result = messagebox.askyesno("Missing Dependencies", msg)
    if not result:
        root.destroy()
        return False
    pip_args = [sys.executable, "-m", "pip", "install"] + [pn for _, pn in missing]
    try:
        result = subprocess.run(pip_args, capture_output=True, text=True)
        if result.returncode != 0:
            messagebox.showerror("Installation Failed",
                f"Could not install packages:\n\n{result.stderr[:500]}")
            root.destroy()
            return False
    except Exception as e:
        messagebox.showerror("Installation Failed", f"Error running pip:\n{e}")
        root.destroy()
        return False
    root.destroy()
    for module_name, _ in missing:
        try:
            importlib.import_module(module_name)
        except ImportError:
            messagebox.showerror("Still Missing",
                f"Package '{module_name}' still could not be imported\n"
                f"after installation. Please install it manually.")
            return False
    return True

if not _check_and_install_dependencies():
    sys.exit(1)

import keyring
import openpyxl
from PIL import Image
import customtkinter as ctk
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# ==============================================================================
# PLATFORM-SPECIFIC FONT HELPER
# ==============================================================================
def get_platform_font():
    """Return the best modern font for the current platform."""
    system = platform.system()
    if system == "Windows":
        return "Segoe UI"
    elif system == "Darwin":  # macOS
        return "SF Pro Display"
    else:  # Linux and others
        return "Arial"

MODERN_FONT = get_platform_font()

# ==============================================================================
# CONFIGURATION & PRESETS
# ==============================================================================
SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587
KEYRING_SERVICE_NAME = "MBE_Automation_CBY"
SENDER_EMAIL = "cby@mbe.ky"

HELPFUL_LINKS_PRESETS = {
    "--- Select a Preset Link (Optional) ---": {"receiver_text": "", "url": ""},
    "How To Retrieve Amazon Invoice": {
        "receiver_text": "How To Retrieve Amazon Invoice", 
        "url": "https://www.youtube.com/watch?v=uL7opF8dvm4"
    },
    "How To Retrieve eBay Invoice": {
        "receiver_text": "How To Retrieve eBay Invoice", 
        "url": "https://www.youtube.com/watch?v=J1z-QHLwye0"
    },
    "How To Retrieve Proper Shein Invoice": {
        "receiver_text": "How To Retrieve Proper Shein Invoice", 
        "url": "https://www.youtube.com/watch?v=NHOp0SwDVBs"
    }
}

# ==============================================================================
# REMOTE SUPPORT: BUG REPORTING TO DISCORD
# ==============================================================================
APP_NAME = "Invoice Request Console"
APP_VERSION = "2.0.3"
DEVELOPER_NAME = "Atlas Ramoon"
BUG_REPORT_WEBHOOK_URL = (
    "https://discord.com/api/webhooks/1524620703259951104/"
    "fqpIEBXVWsKHy7f1iZ9xoryCpidmjPYIDuITfcwMOjBfMyS2HtJNWpVbfOetapl8vw9O"
)

# Portal for remote support — downloaded from GitHub on demand.
PORTAL_URL = (
    "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/"
    "consoles/Python%20Portal%20for%20Atlas.pyw"
)


def _summon_portal(parent_root):
    """Download the Python Portal for Atlas to a user-chosen folder."""
    # Step 1: Confirm with explanation
    confirm = messagebox.askyesno(
        "Open a Portal for Atlas?",
        "This will open a remote IT support portal that lets Atlas\n"
        "diagnose and fix issues on your machine from afar.\n\n"
        "Peace of mind:\n"
        "Atlas cannot see your screen or control your mouse.\n"
        "He can only carry out file-level tasks such as reading\n"
        "nearby files, adding or replacing files, and running\n"
        "Python scripts you send.\n\n"
        "You'll choose where the problem is, then a small portal\n"
        "file will be saved there for you to open.\n\n"
        "Atlas will be notified that you've opened it.\n"
        "When the issue is resolved, you can close and delete it.\n\n"
        "Continue?",
        parent=parent_root)
    if not confirm:
        return

    # Step 2: Choose folder
    folder = filedialog.askdirectory(
        title="Where is the problem located? Choose a folder:",
        parent=parent_root)
    if not folder:
        return

    # Step 3: Download fresh portal with cache-busting (no raw CDN stale content)
    import time as _time
    is_mac = platform.system() == "Darwin"
    ext = ".py" if is_mac else ".pyw"
    dest = os.path.join(folder, f"Python Portal for Atlas{ext}")
    try:
        busted_url = f"{PORTAL_URL}?t={int(_time.time())}"
        req = urllib.request.Request(
            busted_url,
            headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}",
                     "Cache-Control": "no-cache"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
        with open(dest, "wb") as f:
            f.write(data)
    except Exception as e:
        messagebox.showerror("Download Failed",
            f"Could not download the portal:\n\n{e}\n\n"
            "Please check your internet connection and try again.",
            parent=parent_root)
        return

    # Step 4: Launch it automatically
    try:
        if is_mac:
            # On Mac, use python3 explicitly and avoid Windows-only flags
            subprocess.Popen(
                ["python3", dest, "--color=#e0a8e0"],
                start_new_session=True,
            )
        else:
            subprocess.Popen(
                [sys.executable, dest, "--color=#e0a8e0"],
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
    except Exception as e:
        messagebox.showerror(
            "Could Not Launch",
            f"The portal was saved to:\n\n{dest}\n\n"
            f"But it could not be launched automatically:\n{e}\n\n"
            f"Please open it manually.",
            parent=parent_root)
        return

    messagebox.showinfo(
        "Portal Opened",
        "The portal is now opening.\n\n"
        "Leave it running and let Atlas know it's open.",
        parent=parent_root)

UPDATE_MANIFEST_URL = (
    "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/"
    "manifests/invoice-request-console.json"
)
SCRIPT_PATH = Path(__file__).resolve()

def _post_to_discord(content):
    """POST a text message to the Discord webhook. Returns (ok, error_msg)."""
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    payload = json.dumps({"content": content[:1900]}).encode("utf-8")
    try:
        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=payload,
            headers={"Content-Type": "application/json",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)

def _generate_case_number():
    year = datetime.now().year
    tag = uuid.uuid4().hex[:4].upper()
    return f"CASE-{year}-{tag}"

def _post_bug_report_with_files(description, case_number, file_paths,
                                reporter_email="", category="Bug Fix"):
    """POST a bug report to Discord with file attachments in ONE message.
    file_paths: list of file paths to attach (can be empty).
    The running script is always included so the developer gets the exact version.
    Returns (ok, error_msg)."""
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"

    content = (
        f"**Bug Report - {APP_NAME}**\n"
        f"**Case:** {case_number}\n"
        f"**Category:** {category}\n"
        f"**Version:** {APP_VERSION}\n"
        f"**From:** {user}@{host}\n"
        + (f"**Email:** {reporter_email}\n" if reporter_email else "")
        + f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"**Details:**\n{description}"
    )

    if not file_paths:
        return _post_to_discord(content)

    # Multipart: text + files in a single Discord message
    try:
        boundary = f"----WebKitFormBoundary{uuid.uuid4().hex[:16]}"
        payload_json = json.dumps({"content": content[:1900]})

        body = b""
        body += f"--{boundary}\r\n".encode()
        body += b'Content-Disposition: form-data; name="payload_json"\r\n'
        body += b"Content-Type: application/json\r\n\r\n"
        body += payload_json.encode() + b"\r\n"

        for i, fp in enumerate(file_paths):
            p = Path(fp)
            file_data = p.read_bytes()
            body += f"--{boundary}\r\n".encode()
            body += f'Content-Disposition: form-data; name="files[{i}]"; filename="{p.name}"\r\n'.encode()
            body += b"Content-Type: application/octet-stream\r\n\r\n"
            body += file_data + b"\r\n"

        body += f"--{boundary}--\r\n".encode()

        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)

def _send_auto_error_report(title, detail):
    """Send an auto error report with traceback to Discord."""
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    header = (
        f"**Auto Error Report - {APP_NAME}**\n"
        f"**Title:** {title}\n"
        f"**From:** {user}@{host}\n"
        f"**Version:** {APP_VERSION}\n"
        f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    full = header + f"**Details:**\n{detail}"
    LIMIT = 1900
    if len(full) <= LIMIT:
        return _post_to_discord(full)
    parts = []
    while full:
        if len(full) <= LIMIT:
            parts.append(full)
            break
        cut = full.rfind("\n", 0, LIMIT)
        if cut < LIMIT // 2:
            cut = LIMIT
        parts.append(full[:cut])
        full = full[cut:].lstrip("\n")
    ok_all = True
    err_all = None
    for i, part in enumerate(parts):
        prefix = f"[{i+1}/{len(parts)}] " if len(parts) > 1 else ""
        ok, err = _post_to_discord(prefix + part)
        if not ok:
            ok_all = False
            err_all = err
    return ok_all, err_all

def _version_tuple(v):
    out = []
    for p in str(v).split("."):
        try:
            out.append(int(p))
        except ValueError:
            out.append(0)
    return tuple(out)

def _http_get(url, timeout=10):
    req = urllib.request.Request(
        url, headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8")

def _check_for_update():
    try:
        data = json.loads(_http_get(UPDATE_MANIFEST_URL))
        remote = data.get("version", "")
        if remote and _version_tuple(remote) > _version_tuple(APP_VERSION):
            return data
    except Exception:
        pass
    return None

def _download_and_apply_update(new_url):
    try:
        new_text = _http_get(new_url, timeout=30)
        tmp = SCRIPT_PATH.with_name(SCRIPT_PATH.name + ".new")
        tmp.write_text(new_text, encoding="utf-8")
        os.replace(str(tmp), str(SCRIPT_PATH))
        return True, None
    except Exception as e:
        return False, str(e)

def _post_update_applied(old_ver, new_ver):
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    content = (
        f"**Update Applied - {APP_NAME}**\n"
        f"**Updated:** v{old_ver} -> v{new_ver}\n"
        f"**By:** {user}@{host}\n"
        f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    return _post_to_discord(content)

# ==============================================================================
# UTILITY HELPERS (PARSING & VALIDATION ENGINE)
# ==============================================================================
def parse_and_clean_emails(email_string):
    if not email_string or not email_string.strip():
        return []
    raw_list = re.split(r'[;,]', email_string)
    return [email.strip() for email in raw_list if email.strip()]

def validate_email_list(email_list):
    email_regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    for email in email_list:
        if not re.match(email_regex, email):
            return False, email
    return True, None

def normalize_reason(reason):
    if not reason:
        return ""
    normalized = str(reason).strip()
    normalized = normalized.lower()
    normalized = re.sub(r'\s+', ' ', normalized)
    normalized = normalized.rstrip('.')
    return normalized

# ==============================================================================
# PASSWORD SECURITY ENGINE (KEYRING)
# ==============================================================================
def get_secure_password():
    pwd = keyring.get_password(KEYRING_SERVICE_NAME, SENDER_EMAIL)
    if not pwd:
        dialog = ctk.CTk()
        dialog.title("First-Time O365 Setup")
        dialog.geometry("400x230")
        dialog.resizable(False, False)
        
        sw, sh = dialog.winfo_screenwidth(), dialog.winfo_screenheight()
        dialog.geometry(f"400x230+{int((sw-400)/2)}+{int((sh-230)/2)}")
        
        ctk.CTkLabel(dialog, text=f"Enter Office 365 App Password for:\n{SENDER_EMAIL}", 
                     font=(MODERN_FONT, 13, "bold")).pack(pady=(20, 5))
        ctk.CTkLabel(dialog, text="Do not use your main login password.\nGenerate an 'App Password' inside your Microsoft Security settings.", 
                     font=(MODERN_FONT, 11), text_color="#aaa").pack(pady=(0, 10))
                     
        pwd_entry = ctk.CTkEntry(dialog, width=280, show="*")
        pwd_entry.pack(pady=5)
        pwd_entry.focus()
        
        user_pwd = [""]
        def save_and_close():
            user_pwd[0] = pwd_entry.get().strip()
            dialog.destroy()
            
        ctk.CTkButton(dialog, text="Securely Save on this PC", command=save_and_close, 
                      fg_color="#28a745", hover_color="#218838").pack(pady=15)
        dialog.mainloop()
        
        pwd = user_pwd[0]
        if pwd:
            keyring.set_password(KEYRING_SERVICE_NAME, SENDER_EMAIL, pwd)
        else:
            messagebox.showerror("Error", "Password input is mandatory to authenticate.")
            sys.exit(1)
    return pwd

# ==============================================================================
# EMAIL PARSING ENGINE (REDIRECTED TO SYSTEM FILES)
# ==============================================================================
def get_template_data(template_file):
    if not os.path.exists(template_file):
        return "<html><body><h1>Missing Template</h1>{REASON_FOR_REQUEST}</body></html>", []
        
    with open(template_file, "rb") as f:
        msg = message_from_bytes(f.read(), policy=SMTP)
    html_content = ""
    images = []
    img_counter = 0
    
    system_files_dir = "System Files"
    if not os.path.exists(system_files_dir):
        os.makedirs(system_files_dir)
        
    for part in msg.walk():
        if part.get_content_type() == "text/html":
            html_content = part.get_payload(decode=True).decode(errors="ignore")
        elif part.get_content_type().startswith("image/"):
            img_data = part.get_payload(decode=True)
            raw_cid = part.get("Content-ID", "")
            cid = raw_cid.strip("<>") if raw_cid else f"image_{img_counter}"
            
            img_path = os.path.join(system_files_dir, f"temp_img_{img_counter}.png")
            
            with open(img_path, "wb") as f_img: 
                f_img.write(img_data)
            images.append({"path": img_path, "cid": cid, "content_type": part.get_content_type()})
            img_counter += 1
            
    return html_content, images

# ==============================================================================
# SMTP DELIVERY ENGINE
# ==============================================================================
def send_headless_smtp(to_emails, bcc_emails, subject, body, images, password):
    responsive_style = """
    <style>
        html, body { width: 100% !important; margin: 0 !important; padding: 0 !important; -webkit-text-size-adjust: 100%; -ms-text-size-adjust: 100%; }
        body, table, td { font-family: Arial, sans-serif; box-sizing: border-box; }
        table { border-collapse: collapse; margin: 0 auto !important; }
        img { display: block !important; width: 100% !important; max-width: 100% !important; height: auto !important; margin: 0 auto !important; border: 0; }
        
        .container-table { width: 100% !important; max-width: 100% !important; }
    </style>
    """
    
    if "</head>" in body:
        body = body.replace("</head>", responsive_style + "</head>")
    else:
        body = responsive_style + body

    wrapper_start = """<table width="100%" cellpadding="0" cellspacing="0" border="0" style="width:100% !important; table-layout:fixed; background-color: #ffffff;">
        <tr>
            <td align="center" style="padding: 0;">
                <table class="container-table" width="100%" cellpadding="0" cellspacing="0" border="0" style="width:100% !important; max-width:100% !important; margin:0 auto; text-align:left;">
                    <tr>
                        <td style="padding: 0;">"""
                        
    wrapper_end = """</td>
                    </tr>
                </table>
                </td>
        </tr>
    </table>"""

    if "<body>" in body:
        body = body.replace("<body>", "<body>" + wrapper_start)
    elif "<body " in body:
        body = re.sub(r'(<body[^>]*>)', r'\1' + wrapper_start, body)
    else:
        body = wrapper_start + body

    if "</body>" in body:
        body = body.replace("</body>", wrapper_end + "</body>")
    else:
        body = body + wrapper_end

    msg = MIMEMultipart("related")
    msg["From"] = SENDER_EMAIL
    
    msg["To"] = ", ".join(to_emails)
    if bcc_emails:
        msg["BCC"] = ", ".join(bcc_emails)
        
    msg["Subject"] = subject

    msg_alternative = MIMEMultipart("alternative")
    msg.attach(msg_alternative)
    msg_html = MIMEText(body, "html", "utf-8")
    msg_alternative.attach(msg_html)

    for img in images:
        if os.path.exists(img["path"]):
            with open(img["path"], "rb") as f:
                mime_img = MIMEImage(f.read(), _subtype=img["content_type"].split("/")[-1])
            mime_img.add_header("Content-ID", f"<{img['cid']}>")
            mime_img.add_header("Content-Disposition", "inline", filename=os.path.basename(img["path"]))
            msg.attach(mime_img)

    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(SENDER_EMAIL, password)
    
    all_recipients = list(set(to_emails + bcc_emails))
        
    server.sendmail(SENDER_EMAIL, all_recipients, msg.as_string())
    server.quit()

# ==============================================================================
# MANIFEST READER — parses a Manifest and Customs .xlsx file
# Reads "Email Automation Output" for data rows and " COLS" for the manifest date.
# ==============================================================================
def read_manifest(file_path):
    """Read a manifest .xlsx and return (grouped_rows, manifest_date_str).

    grouped_rows: list of dicts with keys To Email Address, CBY Number (list),
                  Package Number (list), Package Description (list),
                  Reason for Request (list with original text).
    manifest_date_str: formatted as MM-DD-YYYY from the COLS sheet, or "".
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return [], ""

    manifest_date_str = ""

    # --- Extract manifest date from the COLS sheet ---
    cols_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().upper() == "COLS":
            cols_sheet_name = name
            break
    if cols_sheet_name:
        try:
            cols_ws = wb[cols_sheet_name]
            date_val = cols_ws["E2"].value
            if isinstance(date_val, datetime):
                manifest_date_str = date_val.strftime("%m-%d-%Y")
            elif date_val:
                manifest_date_str = str(date_val).strip()
        except Exception:
            pass

    # --- Read data from the Email Automation Output sheet ---
    email_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().upper() == "EMAIL AUTOMATION OUTPUT":
            email_sheet_name = name
            break
    if not email_sheet_name:
        return [], manifest_date_str

    ws = wb[email_sheet_name]
    headers = [cell.value for cell in ws[1]]
    raw_rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]

    if not raw_rows:
        return [], manifest_date_str

    # Group rows by (To Email Address, normalized Reason)
    grouped_dict = {}
    for row in raw_rows:
        row_dict = dict(zip(headers, row))
        to_email = str(row_dict.get("To Email Address") or "").strip()
        raw_reason = str(row_dict.get("Reason for Request") or "").strip()

        if not to_email or not raw_reason:
            continue

        email_key = to_email.lower()
        normalized_reason = normalize_reason(raw_reason)
        composite_key = (email_key, normalized_reason)

        if composite_key not in grouped_dict:
            grouped_dict[composite_key] = {
                "To Email Address": to_email,
                "CBY Number": [],
                "Package Number": [],
                "Package Description": [],
                "Reason for Request": [raw_reason]
            }

        for key in ["CBY Number", "Package Number", "Package Description"]:
            val = str(row_dict.get(key) or "").strip()
            if key == "CBY Number":
                if val and val not in grouped_dict[composite_key][key]:
                    grouped_dict[composite_key][key].append(val)
            else:
                grouped_dict[composite_key][key].append(val)

    return list(grouped_dict.values()), manifest_date_str

# ==============================================================================
# STANDALONE & QUEUE INTERACTIVE REVIEW ENGINE
# ==============================================================================
def process_queue():
    o365_password = get_secure_password()
    html_template, images = get_template_data(os.path.join("System Files", "template.eml"))

    # Start in standalone mode — user can load a manifest via the button.
    grouped_rows = [{}]
    manifest_date_extracted = ""
    is_standalone_mode = True
    manifest_loaded = [False]  # mutable flag for nested functions

    approved_emails = []
    current_index = [0]

    queue_timer_id = [None]
    is_in_queued_state = [False]

    root = ctk.CTk()
    root.title(f"{APP_NAME} v{APP_VERSION}")
    root.geometry("540x760")
    root.resizable(True, True)
    root.configure(fg_color="#1a1a1a")

    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    root.geometry(f"540x760+{int((sw-540)/2)}+{int((sh-760)/2)}")

    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.pack(pady=(20, 10), padx=(5, 30))
    
    # Load and display MBE logo
    logo_path = os.path.join("System Files", "mbe-logo.png")
    if os.path.exists(logo_path):
        try:
            logo_image = Image.open(logo_path)
            logo_image = logo_image.resize((150, 90), Image.Resampling.LANCZOS)
            logo_photo = ctk.CTkImage(logo_image, size=(150, 90))
            logo_label = ctk.CTkLabel(header_frame, image=logo_photo, text="")
            logo_label.pack(side="left", padx=(0, 12))
        except Exception:
            pass  # If logo fails to load, continue without it
    
    # Create a frame for the title and subtitle on the right
    text_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
    text_frame.pack(side="left", fill="both", expand=True)
    
    title_lbl = ctk.CTkLabel(text_frame, text="Urgent Invoice Request", font=(MODERN_FONT, 15, "bold"))
    title_lbl.pack(pady=(10, 0), anchor="w")
    
    subtitle_lbl = ctk.CTkLabel(text_frame, text="Confirm Details Before Approving", font=(MODERN_FONT, 14, "bold"), text_color="#dc3545")
    subtitle_lbl.pack(pady=(0, 2), anchor="w")

    # ---- Manifest file picker button (Excel green, compact, under logo) ----
    EXCEL_GREEN = "#217346"
    EXCEL_GREEN_HOVER = "#1a5c38"

    def on_pick_manifest():
        nonlocal is_standalone_mode, manifest_date_extracted
        file_path = filedialog.askopenfilename(
            title="Select Manifest and Customs Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not file_path:
            return
        try:
            new_grouped, new_date = read_manifest(file_path)
        except Exception as e:
            messagebox.showerror("Manifest Error", f"Could not read the manifest file:\n{e}")
            return
        if not new_grouped:
            messagebox.showinfo(
                "No Data",
                "No email request rows were found in the 'Email Automation Output' sheet.\n"
                "Make sure the manifest has been processed and the output sheet is populated.",
            )
            return
        nonlocal grouped_rows
        grouped_rows[:] = new_grouped
        manifest_date_extracted = new_date
        is_standalone_mode = False
        manifest_loaded[0] = True
        current_index[0] = 0
        manifest_btn.configure(text="\U0001f4c2 Loaded", state="disabled",
                               fg_color="#3a3f44", hover_color="#3a3f44")
        # Show the progress label now that we're in queue mode
        progress_lbl.pack(pady=(0, 5))
        load_current_record()

    manifest_btn = ctk.CTkButton(
        root,
        text="\U0001f4c2 Pull Drafts From Manifest",
        command=on_pick_manifest,
        fg_color=EXCEL_GREEN,
        hover_color=EXCEL_GREEN_HOVER,
        width=260,
        height=28,
        corner_radius=6,
        font=(MODERN_FONT, 11, "bold"),
    )
    manifest_btn.pack(anchor="e", padx=30, pady=(0, 4))

    entries = {}
    
    def create_modern_row(label_text):
        row_frame = ctk.CTkFrame(root, fg_color="transparent")
        row_frame.pack(fill="x", padx=30, pady=5)
        lbl = ctk.CTkLabel(row_frame, text=label_text, width=110, anchor="w", font=(MODERN_FONT, 12))
        lbl.pack(side="left")
        entry = ctk.CTkEntry(row_frame, width=310, height=28, corner_radius=6, border_width=1, border_color="#3a3f44", fg_color="#2a2a2a")
        entry.pack(side="right", fill="x", expand=True)
        entries[label_text] = entry

    create_modern_row("Customer Email:")
    create_modern_row("BCC:")
    create_modern_row("CBY #:")
    create_modern_row("Manifest Date:")

    # ==============================================================================
    # EDITABLE PACKAGE TABLE (Pkg # and Description only — CBY is a single field above)
    # ==============================================================================
    pkg_table_frame = ctk.CTkFrame(root, fg_color="transparent")
    pkg_table_frame.pack(fill="x", expand=False, padx=30, pady=(5, 0))

    pkg_table_header_frame = ctk.CTkFrame(pkg_table_frame, fg_color="transparent")
    pkg_table_header_frame.pack(fill="x")

    # Column headers + "+" button on the same line
    ctk.CTkLabel(pkg_table_header_frame, text="Pkg #", width=120, anchor="w", font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=(0, 4))
    ctk.CTkLabel(pkg_table_header_frame, text="Description", anchor="w", font=(MODERN_FONT, 11, "bold")).pack(side="left")
    add_row_btn = ctk.CTkButton(pkg_table_header_frame, text="+", width=30, height=22, corner_radius=4, fg_color="#3a3f44", hover_color="#4a5054", font=(MODERN_FONT, 12, "bold"), command=lambda: add_package_row())
    add_row_btn.pack(side="right")

    # Fixed-height container that clips the scrollable area
    pkg_rows_container = ctk.CTkFrame(pkg_table_frame, fg_color="transparent", height=125)
    pkg_rows_container.pack(fill="x", expand=False, pady=(2, 0))
    pkg_rows_container.pack_propagate(False)  # Prevent children from expanding this frame

    # Scrollable frame inside the fixed container
    pkg_rows_scroll = ctk.CTkScrollableFrame(pkg_rows_container, fg_color="transparent")
    pkg_rows_scroll.pack(fill="both", expand=True)

    pkg_rows = []  # list of dicts: {"pkg": CTkEntry, "desc": CTkEntry, "frame": CTkFrame}

    def add_package_row(pkg_val="", desc_val=""):
        """Add one editable row to the package table."""
        row_frame = ctk.CTkFrame(pkg_rows_scroll, fg_color="transparent")
        row_frame.pack(fill="x", pady=2)

        pkg_entry = ctk.CTkEntry(row_frame, width=120, height=26, corner_radius=4, border_width=1, border_color="#3a3f44", fg_color="#2a2a2a", font=(MODERN_FONT, 11))
        pkg_entry.pack(side="left", padx=(0, 4))
        pkg_entry.insert(0, pkg_val)

        desc_entry = ctk.CTkEntry(row_frame, height=26, corner_radius=4, border_width=1, border_color="#3a3f44", fg_color="#2a2a2a", font=(MODERN_FONT, 11))
        desc_entry.pack(side="left", fill="x", expand=True)
        desc_entry.insert(0, desc_val)

        row_data = {"pkg": pkg_entry, "desc": desc_entry, "frame": row_frame}
        pkg_rows.append(row_data)
        return row_data

    # Start with 4 empty rows
    for _ in range(4):
        add_package_row()

    def get_package_table_data():
        """Collect non-empty rows from the package table, injecting the single CBY # into each."""
        cby_number = entries["CBY #:"].get().strip()
        rows = []
        for r in pkg_rows:
            pkg = r["pkg"].get().strip()
            desc = r["desc"].get().strip()
            if pkg or desc:
                rows.append({"cby": cby_number, "pkg": pkg, "desc": desc})
        return rows

    def clear_package_table():
        """Remove all rows and re-create 4 empty ones."""
        for r in pkg_rows:
            r["frame"].destroy()
        pkg_rows.clear()
        for _ in range(4):
            add_package_row()

    def populate_package_table(pkg_list, desc_list):
        """Populate table from grouped Excel data."""
        for r in pkg_rows:
            r["frame"].destroy()
        pkg_rows.clear()

        max_len = max(len(pkg_list), len(desc_list), 4)
        for i in range(max_len):
            pkg_val = pkg_list[i] if i < len(pkg_list) else ""
            desc_val = desc_list[i] if i < len(desc_list) else ""
            add_package_row(pkg_val, desc_val)
        # Ensure minimum 4 rows
        while len(pkg_rows) < 4:
            add_package_row()

    dropdown_frame = ctk.CTkFrame(root, fg_color="transparent")
    dropdown_frame.pack(fill="x", padx=30, pady=5)
    ctk.CTkLabel(dropdown_frame, text="Helpful Link:", width=110, anchor="w", font=(MODERN_FONT, 12)).pack(side="left")
    
    selected_link = ctk.StringVar()
    dropdown = ctk.CTkComboBox(dropdown_frame, variable=selected_link, values=list(HELPFUL_LINKS_PRESETS.keys()), width=310, height=28, corner_radius=6, border_width=1, border_color="#3a3f44", fg_color="#2a2a2a", dropdown_fg_color="#2a2a2a", dropdown_hover_color="#3a3f44", dropdown_text_color="white")
    dropdown.pack(side="right", fill="x", expand=True)

    # ==============================================================================
    # COMMAND FUNCTIONS
    # ==============================================================================
    DEFAULT_WHITE = "#ffffff"

    def reset_button_ui():
        is_in_queued_state[0] = False
        default_approve_text = "SEND EMAIL" if is_standalone_mode else "APPROVE"
        approve_btn.configure(text=default_approve_text, fg_color="#28a745", hover_color="#218838", state="normal")
        if not is_standalone_mode:
            skip_btn.configure(text="SKIP", fg_color="#dc3545", hover_color="#c82333")
        else:
            skip_btn.configure(text="CLOSE", fg_color="#dc3545", hover_color="#c82333")

    def load_current_record():
        idx = current_index[0]
        if idx >= len(grouped_rows):
            root.destroy()
            return
            
        reset_button_ui()
            
        if is_standalone_mode:
            root.title(f"{APP_NAME} v{APP_VERSION} — Standalone Mode")
        else:
            root.title(f"{APP_NAME} v{APP_VERSION} — Draft {idx + 1} of {len(grouped_rows)}")
            progress_lbl.configure(text=f"Queue Progress: Draft {idx + 1} of {len(grouped_rows)}")
            
        row_data = grouped_rows[idx] if not is_standalone_mode else {}
        
        for lbl_name, widget in entries.items():
            widget.delete(0, "end")
            
        if not is_standalone_mode:
            entries["Customer Email:"].insert(0, row_data.get("To Email Address", ""))
            entries["BCC:"].insert(0, "")
            # CBY Number: use first from the grouped list (they share the same CBY for this email)
            cby_list = row_data.get("CBY Number", [])
            entries["CBY #:"].insert(0, cby_list[0] if cby_list else "")
            entries["Manifest Date:"].insert(0, manifest_date_extracted)
            
            # Populate package table from grouped data (Pkg # and Description only)
            populate_package_table(
                row_data.get("Package Number", []),
                row_data.get("Package Description", [])
            )
            
            dropdown.set(list(HELPFUL_LINKS_PRESETS.keys())[0])
            reason_text.delete("1.0", "end")
            reason_text.configure(text_color=DEFAULT_WHITE)
            reason_text.insert("1.0", row_data.get("Reason for Request", [""])[0])
        else:
            entries["Manifest Date:"].insert(0, manifest_date_extracted)
            clear_package_table()
            dropdown.set(list(HELPFUL_LINKS_PRESETS.keys())[0])
            reason_text.configure(text_color=DEFAULT_WHITE)

    def reset_standalone_form():
        """Reset the form to a clean blank state for the next standalone email."""
        # Clear all form fields except Manifest Date
        entries["Customer Email:"].delete(0, "end")
        entries["BCC:"].delete(0, "end")
        entries["CBY #:"].delete(0, "end")
        
        # Reset package table to 4 empty rows
        clear_package_table()
        
        # Reset Reason textbox
        reason_text.delete("1.0", "end")
        reason_text.configure(text_color=DEFAULT_WHITE)
        
        # Reset Helpful Link dropdown to default
        dropdown.set(list(HELPFUL_LINKS_PRESETS.keys())[0])
        
        # Reset button UI
        reset_button_ui()
        
        # Focus on first input field (Customer Email:)
        entries["Customer Email:"].focus()

    def build_package_html_table(packages, manifest_date):
        """Build a horizontal HTML table for package details, wrapped in navy slate card style.
        
        CBY # and Manifest Date appear as bold footer lines below the table.
        Table columns are just Pkg # and Description to save horizontal space.
        """
        # NAVY SLATE CARD style (consistent with ocean alerts V2)
        card_bg = "#f8f9fa"
        card_border_color = "#003366"
        header_bg = "#003366"
        header_text_color = "#ffffff"
        header_border = "#002244"
        even_row_bg = "#ffffff"
        odd_row_bg = "#f8f9fa"
        cell_text = "#000000"
        cell_border = "#dee2e6"

        header_style = f"background-color: {header_bg}; color: {header_text_color}; padding: 8px 12px; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; text-align: left; border: 1px solid {header_border};"
        cell_style_base = f"padding: 8px 12px; font-family: Arial, sans-serif; font-size: 14px; color: {cell_text}; text-align: left; border: 1px solid {cell_border};"

        # Get CBY # from the first package row (all rows share the same CBY)
        cby_number = packages[0].get("cby", "") if packages else ""

        rows_html = ""
        for i, pkg in enumerate(packages):
            bg = even_row_bg if i % 2 == 0 else odd_row_bg
            rows_html += f"""<tr style="background-color: {bg};">
                        <td style="{cell_style_base}">{pkg.get('pkg', '')}</td>
                        <td style="{cell_style_base}">{pkg.get('desc', '')}</td>
                    </tr>"""

        table_html = f"""
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color: {card_bg}; border-left: 5px solid {card_border_color}; margin-top: 10px; margin-bottom: 10px;">
            <tr>
                <td style="padding: 14px; font-family: Arial, sans-serif;">
                    <p style="margin: 0 0 10px 0; font-family: Arial, sans-serif; font-size: 15px; font-weight: bold; color: {card_border_color};"><span>CBY # {cby_number}</span><span style="margin-left: 30px;">Manifest Date: {manifest_date}</span></p>
                    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse: collapse; border: 1px solid {cell_border};">
                        <tr>
                            <th style="{header_style}">Pkg #</th>
                            <th style="{header_style}">Description</th>
                        </tr>
                        {rows_html}
                    </table>
                </td>
            </tr>
        </table>
        """
        return table_html

    def send_single_email(record):
        """Send a single email immediately (for standalone mode)."""
        body = html_template
        
        packages = record.get("Packages", [])
        manifest_date = record.get("Manifest Date (MM-DD)", "")
        package_table_html = build_package_html_table(packages, manifest_date)
        
        body = body.replace("{PACKAGE_DETAILS}", package_table_html)
    
        raw_reason = str(record["Reason for Request"]).replace("\n", "<br>")
        link_obj = record["Link Object"]
        if link_obj["url"]:
            raw_reason += f"<br><br>This video may be helpful: <a href='{link_obj['url']}' style='color: #003366; font-weight: bold; text-decoration: underline;'>{link_obj['receiver_text']}</a>"
        
        highlighted_reason = f"""
        <table width="100%" cellpadding="0" cellspacing="0" border="0">
            <tr>
                <td style="background-color: #fff3cd; padding: 12px; border: 1px solid #ffeeba; border-left: 5px solid #d39e00; border-radius: 4px; font-family: Arial, sans-serif; font-size: 14px; color: #333333; text-align: left;">
                    {raw_reason}
                </td>
            </tr>
        </table>
        """
        body = body.replace("{REASON_FOR_REQUEST}", highlighted_reason)
        
        # Build subject: use first CBY# from packages for subject line
        first_cby = packages[0]["cby"] if packages else "N/A"
        subject = f"Urgent Attention Required: Customs Processing - CBY#{first_cby} / Package#{record['Subject Package Number']} ({manifest_date})"
        
        send_headless_smtp(
            to_emails=record["To Emails"],
            bcc_emails=record["BCC Emails"],
            subject=subject,
            body=body,
            images=images,
            password=o365_password
        )

    def commit_approved_record(record):
        if is_standalone_mode:
            # Standalone mode: send email immediately in background thread, then reset form
            def send_and_reset():
                try:
                    send_single_email(record)
                    # Update UI on main thread after successful send
                    root.after(0, lambda: on_send_success())
                except Exception as e:
                    # Update UI on main thread after error
                    root.after(0, lambda: on_send_error(str(e)))
            
            threading.Thread(target=send_and_reset, daemon=True).start()
        else:
            # Queue mode: add to queue and proceed to next record
            approved_emails.append(record)
            current_index[0] += 1
            load_current_record()

    def on_send_success():
        """Called after successful email send in standalone mode."""
        approve_btn.configure(text="Email Sent", fg_color="#28a745", hover_color="#218838", state="normal")
        skip_btn.configure(text="CLOSE", fg_color="#dc3545", hover_color="#c82333")
        is_in_queued_state[0] = False
        
        # Brief delay to show "Email Sent" before resetting
        root.after(1000, reset_standalone_form)

    def on_send_error(error_msg):
        """Called after failed email send in standalone mode."""
        messagebox.showerror("Delivery Error", f"Failed to send email:\n{error_msg}")
        reset_button_ui()

    def on_approve():
        if is_in_queued_state[0]:
            return
            
        raw_to = entries["Customer Email:"].get()
        raw_bcc = entries["BCC:"].get()
        
        to_list = parse_and_clean_emails(raw_to)
        bcc_list = parse_and_clean_emails(raw_bcc)
        
        if not to_list:
            messagebox.showwarning("Missing Input", "Please provide at least one recipient email address ('Customer Email:') before sending.")
            return
            
        to_valid, failed_to_email = validate_email_list(to_list)
        if not to_valid:
            messagebox.showerror("Invalid Recipient Format", f"'{failed_to_email}' is not a valid email address structure.\n\nPlease fix it to prevent O365 server blocks.")
            return

        bcc_valid, failed_bcc_email = validate_email_list(bcc_list)
        if not bcc_valid:
            messagebox.showerror("Invalid BCC Format", f"BCC address '{failed_bcc_email}' is malformed.\n\nPlease clear or correct it before processing.")
            return
            
        # Collect package rows from the editable table
        package_rows = get_package_table_data()
        
        # Build subject line package number (first pkg# with "+" if multiple)
        pkg_numbers = [r["pkg"] for r in package_rows if r["pkg"]]
        if len(pkg_numbers) > 1:
            subject_pkg = f"{pkg_numbers[0]}+"
        elif len(pkg_numbers) == 1:
            subject_pkg = pkg_numbers[0]
        else:
            subject_pkg = "N/A"

        record = {
            "To Emails": to_list,
            "BCC Emails": bcc_list,
            "Packages": package_rows,
            "Subject Package Number": subject_pkg,
            "Manifest Date (MM-DD)": entries["Manifest Date:"].get().strip(),
            "Reason for Request": reason_text.get("1.0", "end").strip(),
            "Link Object": HELPFUL_LINKS_PRESETS[selected_link.get()]
        }

        if is_standalone_mode:
            # Standalone mode: send immediately without queuing
            is_in_queued_state[0] = True
            approve_btn.configure(text="SENDING...", fg_color="#555b5e", hover_color="#555b5e", state="disabled")
            skip_btn.configure(text="STOP", fg_color="#d63031", hover_color="#ff7675")
            # Send immediately without delay
            commit_approved_record(record)
        else:
            # Queue mode: queue with delay
            is_in_queued_state[0] = True
            approve_btn.configure(text="QUEUING...", fg_color="#555b5e", hover_color="#555b5e", state="disabled")
            skip_btn.configure(text="STOP", fg_color="#d63031", hover_color="#ff7675")
            timer_id = root.after(1500, lambda: commit_approved_record(record))
            queue_timer_id[0] = timer_id

    def on_skip():
        if is_in_queued_state[0] and queue_timer_id[0] is not None:
            root.after_cancel(queue_timer_id[0])
            queue_timer_id[0] = None
            reset_button_ui()            
        else:
            if is_standalone_mode:
                root.destroy()
            else:
                current_index[0] += 1
                load_current_record()

    # ==============================================================================
    # REASON & BUTTONS UI LAYOUT
    # ==============================================================================
    reason_frame = ctk.CTkFrame(root, fg_color="transparent")
    reason_frame.pack(fill="both", expand=True, padx=30, pady=(4, 0))
    
    reason_header = ctk.CTkFrame(reason_frame, fg_color="transparent")
    reason_header.pack(fill="x", pady=(0, 2))

    ctk.CTkLabel(reason_header, text="Reason:", font=(MODERN_FONT, 12)).pack(side="left", anchor="nw")
    
    reason_text = ctk.CTkTextbox(reason_frame, height=60, corner_radius=6, border_width=1, border_color="#3a3f44", fg_color="#2a2a2a", text_color=DEFAULT_WHITE)
    reason_text.pack(fill="both", expand=True, pady=(2, 0))
    reason_text._textbox.configure(wrap="word")

    btn_frame = ctk.CTkFrame(root, fg_color="transparent")
    btn_frame.pack(fill="x", padx=30, pady=(10, 15))

    approve_btn_text = "SEND EMAIL" if is_standalone_mode else "APPROVE"
    skip_btn_text = "CLOSE" if is_standalone_mode else "SKIP"

    skip_btn = ctk.CTkButton(btn_frame, text=skip_btn_text, command=on_skip, fg_color="#dc3545", hover_color="#c82333", width=140, height=32, font=(MODERN_FONT, 12, "bold"), corner_radius=6)
    skip_btn.pack(side="right", padx=(10, 0))

    approve_btn = ctk.CTkButton(btn_frame, text=approve_btn_text, command=on_approve, fg_color="#28a745", hover_color="#218838", width=160, height=32, font=(MODERN_FONT, 12, "bold"), corner_radius=6)
    approve_btn.pack(side="right")
    
    # ==============================================================================
    # FOOTER UI
    # ==============================================================================
    divider = ctk.CTkFrame(root, height=1, fg_color="#3e454a")
    divider.pack(fill="x", padx=30, pady=(5, 10))

    progress_lbl = ctk.CTkLabel(root, text="", font=(MODERN_FONT, 12, "bold"), text_color="#aaa")
    if not is_standalone_mode:
        progress_lbl.pack(pady=(0, 5))

    # ---- Footer row: "designed by" on the left, bug icon on the right ----
    footer_row = ctk.CTkFrame(root, fg_color="transparent")
    footer_row.pack(fill="x", padx=30, pady=(0, 8))

    ctk.CTkLabel(footer_row, text="Program designed by Atlas Ramoon",
                 font=(MODERN_FONT, 10), text_color="#555555").pack(side="left", pady=(0, 0))

    # ---- Bug report icon (bottom-right) ----
    _tooltip_win = [None]
    _pending_update = [None]
    _support_tooltip = ["Report a Bug"]

    def _show_tooltip(widget, text):
        _hide_tooltip()
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + 18
        line_count = text.count("\n") + 1
        est_height = line_count * 16 + 16
        screen_h = widget.winfo_screenheight()
        if y + est_height > screen_h - 20:
            y = widget.winfo_rooty() - est_height - 8
        tw = tk.Toplevel(root)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=text, justify="left",
                         bg="#1a1a2e", fg="#e8e8e8", relief="solid",
                         bd=1, padx=10, pady=8,
                         font=(MODERN_FONT, 10), wraplength=320)
        label.pack()
        _tooltip_win[0] = tw

    def _hide_tooltip():
        if _tooltip_win[0] is not None:
            try:
                _tooltip_win[0].destroy()
            except Exception:
                pass
            _tooltip_win[0] = None

    def _refresh_support_icon():
        upd = _pending_update[0]
        if upd:
            ver = upd.get("version", "?")
            bug_btn.configure(text="Apply Fixes", fg_color="#b8860b",
                              hover_color="#daa520")
            _support_tooltip[0] = f"Apply Fixes — v{ver} available"
        else:
            bug_btn.configure(text="\U0001f41e", fg_color="#1a1a1a",
                              hover_color="#2a2a3e")
            _support_tooltip[0] = "Report a Bug"

    def _on_support_click():
        _hide_tooltip()
        if _pending_update[0]:
            _apply_fixes_dialog()
        else:
            _report_bug_dialog()

    def _apply_fixes_dialog():
        upd = _pending_update[0]
        if not upd:
            return
        dlg = ctk.CTkToplevel(root)
        dlg.title("Apply Fixes")
        dlg.configure(fg_color="#1a1a1a")
        dlg.transient(root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 380
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ver = upd.get("version", "?")
        ctk.CTkLabel(dlg, text=f"Fixes Available  (v{ver})",
                     font=(MODERN_FONT, 15, "bold"), text_color="#ffffff").pack(
            anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg, text=f"From {DEVELOPER_NAME}. Click Apply Fixes to download,\n"
                     "then choose how to restart.",
                     font=(MODERN_FONT, 11), text_color="#aaa",
                     anchor="w").pack(anchor="w", padx=16, pady=(0, 8))

        box = ctk.CTkTextbox(dlg, height=170, fg_color="#2a2a2a", border_color="#3a3f44",
                             border_width=1, corner_radius=4, text_color="#ffffff",
                             font=(MODERN_FONT, 11), wrap="word")
        box.pack(fill="both", expand=True, padx=16)
        box.insert("0.0", upd.get("changelog", "(no description provided)"))
        box.configure(state="disabled")

        def _apply():
            apply_btn.configure(state="disabled", text="Applying...")
            def _worker():
                ok, err = _download_and_apply_update(upd.get("url", ""))
                root.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                old_ver = APP_VERSION
                new_ver = upd.get("version", "?")
                threading.Thread(
                    target=lambda: _post_update_applied(old_ver, new_ver),
                    daemon=True).start()
                _pending_update[0] = None
                _refresh_support_icon()
                dlg.destroy()
                if messagebox.askyesno("Update Applied",
                    f"Updated to version {new_ver}.\n\n"
                    "The app needs to restart to apply the changes.\n"
                    "Restart now?"):
                    _restart_app()
            else:
                apply_btn.configure(state="normal", text="Apply Fixes")
                messagebox.showerror("Update Failed",
                    f"Could not apply the update:\n{err}\n\n"
                    "Check your internet connection and try again.")

        apply_btn = ctk.CTkButton(btns, text="Apply Fixes", command=_apply,
                                  fg_color="#b8860b", hover_color="#daa520",
                                  width=130, height=30, corner_radius=5,
                                  font=(MODERN_FONT, 11, "bold"))
        apply_btn.pack(side="left")
        ctk.CTkButton(btns, text="Later", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    def _restart_app():
        try:
            import subprocess
            subprocess.Popen([sys.executable, str(SCRIPT_PATH)])
        except Exception:
            pass
        try:
            root.destroy()
        except Exception:
            pass
        sys.exit(0)

    def _report_bug_dialog():
        _hide_tooltip()
        dlg = ctk.CTkToplevel(root)
        dlg.title("Report a Bug")
        dlg.configure(fg_color="#1a1a1a")
        dlg.transient(root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 620
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        ctk.CTkLabel(dlg, text="Report a Bug", font=(MODERN_FONT, 15, "bold"),
                     text_color="#ffffff").pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Describe the problem in as much detail as you can — what you\n"
                 f"did, what happened, and what you expected. This goes directly\n"
                 f"to the developer ({DEVELOPER_NAME}).",
            font=(MODERN_FONT, 11), text_color="#aaa",
            anchor="w", justify="left").pack(anchor="w", padx=16, pady=(0, 6))

        # ---- Check for Updates bar ----
        update_bar = ctk.CTkFrame(dlg, fg_color="transparent")
        update_bar.pack(fill="x", padx=16, pady=(0, 6))

        check_btn = ctk.CTkButton(
            update_bar, text="Check for Updates",
            command=lambda: None,
            fg_color="#3b82f6", hover_color="#2563eb",
            width=140, height=26, corner_radius=4,
            font=(MODERN_FONT, 10, "bold"))
        check_btn.pack(side="left")

        update_status = ctk.CTkLabel(
            update_bar, text="",
            font=(MODERN_FONT, 10), text_color="#aaa",
            anchor="w")
        update_status.pack(side="left", padx=(10, 0))

        _spin = {"frame": 0, "running": False}

        def _spin_step():
            if not _spin["running"]:
                return
            _spin["frame"] = (_spin["frame"] + 1) % 8
            dots = "." * ((_spin["frame"] % 4) + 1)
            update_status.configure(text=f"Checking{dots}")
            dlg.after(200, _spin_step)

        def _do_check():
            _spin["running"] = True
            check_btn.configure(state="disabled", text="Checking...")
            _spin_step()
            def _worker():
                result = _check_for_update()
                dlg.after(0, lambda: _check_done(result))
            threading.Thread(target=_worker, daemon=True).start()

        def _check_done(result):
            _spin["running"] = False
            check_btn.configure(state="normal", text="Check for Updates")
            if result:
                ver = result.get("version", "?")
                _pending_update[0] = result
                update_status.configure(
                    text=f"v{ver} available!", text_color="#ffffff")
                check_btn.configure(
                    text=f"Update Now — v{ver}",
                    command=_apply_fixes_dialog,
                    fg_color="#b8860b", hover_color="#daa520",
                    width=180)
                _refresh_support_icon()
            else:
                update_status.configure(
                    text=f"You're up to date (v{APP_VERSION})",
                    text_color="#aaa")

        check_btn.configure(command=_do_check)

        # Category selector
        cat_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        cat_frame.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(cat_frame, text="Category:",
                     font=(MODERN_FONT, 11, "bold"), text_color="#ffffff",
                     anchor="w").pack(anchor="w", pady=(0, 2))
        cat_var = ctk.StringVar(value="Bug Fix")
        cat_row = ctk.CTkFrame(cat_frame, fg_color="transparent")
        cat_row.pack(fill="x")
        for cat in ("Bug Fix", "Feature Request", "Environmental Change", "Other"):
            ctk.CTkRadioButton(cat_row, text=cat, variable=cat_var, value=cat,
                               font=(MODERN_FONT, 10), text_color="#ffffff",
                               fg_color="#3b82f6", hover_color="#2563eb",
                               border_color="#3a3f44").pack(side="left", padx=(0, 12))

        box = ctk.CTkTextbox(dlg, height=140, fg_color="#2a2a2a", border_color="#3a3f44",
                             border_width=1, corner_radius=4, text_color="#ffffff",
                             font=(MODERN_FONT, 11))
        box.pack(fill="both", expand=True, padx=16)
        box.focus_set()

        # Optional email field
        email_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        email_frame.pack(fill="x", padx=16, pady=(4, 0))
        ctk.CTkLabel(email_frame, text="Your email (optional — for updates on this report)",
                     font=(MODERN_FONT, 10), text_color="#aaa",
                     anchor="w").pack(anchor="w")
        email_var = ctk.StringVar(value="")
        email_entry = ctk.CTkEntry(email_frame, textvariable=email_var, height=28,
                                   fg_color="#2a2a2a", border_color="#3a3f44", border_width=1,
                                   corner_radius=4, text_color="#ffffff",
                                   font=(MODERN_FONT, 11))
        email_entry.pack(fill="x", pady=(1, 0))

        # "What to expect" info box
        info = ctk.CTkFrame(dlg, fg_color="#1a2a3a", corner_radius=6)
        info.pack(fill="x", padx=16, pady=(6, 8))
        ctk.CTkLabel(info, text="What to expect",
                     font=(MODERN_FONT, 11, "bold"), text_color="#88ccff",
                     anchor="w").pack(anchor="w", padx=10, pady=(8, 2))
        ctk.CTkLabel(info,
            text=f"\u2022 {DEVELOPER_NAME} will review your report within 24-48 hours\n"
                 f"\u2022 When a fix is ready, click 'Check for Updates' above\n"
                 f"   or relaunch the app \u2014 the bug icon will show 'Apply Fixes'\n"
                 f"\u2022 {DEVELOPER_NAME} will coordinate with management if the fix\n"
                 f"   requires substantial work",
            font=(MODERN_FONT, 10), text_color="#aaa",
            anchor="w", justify="left").pack(anchor="w", padx=10, pady=(0, 8))

        # Buttons
        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        # Remote support portal icon (bottom-right corner, in button row)
        _portal_canvas = tk.Canvas(btns, width=24, height=24,
                                    bg="#1a1a1a", highlightthickness=0)
        _portal_canvas.pack(side="right")
        _portal_canvas.create_oval(2, 2, 22, 22, outline="#e0a8e0", width=2)
        _portal_canvas.create_oval(7, 7, 17, 17, fill="#e0a8e0", outline="")
        _portal_canvas.configure(cursor="hand2")
        _portal_canvas.bind("<Button-1>", lambda e: (dlg.grab_release(), _summon_portal(dlg)))

        def _next():
            desc = box.get("0.0", "end").strip()
            if not desc:
                messagebox.showwarning("Empty", "Please describe the bug first.")
                return
            email = email_var.get().strip()
            category = cat_var.get()
            dlg.destroy()
            _show_attach_files_dialog(desc, email, category)

        ctk.CTkButton(btns, text="Next", command=_next,
                      fg_color="#28a745", hover_color="#218838", width=110,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        ctk.CTkButton(btns, text="Cancel", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=110,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    def _show_attach_files_dialog(description, reporter_email="", category="Bug Fix"):
        """After the bug description, offer to attach sample files.
        The running .pyw script is always attached automatically so the
        developer gets the user's exact version for testing."""
        case_num = _generate_case_number()
        dlg = ctk.CTkToplevel(root)
        dlg.title(f"Attach Files? (Case {case_num})")
        dlg.configure(fg_color="#1a1a1a")
        dlg.transient(root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 420
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(dlg, text="Attach Sample Files?",
                     font=(MODERN_FONT, 15, "bold"), text_color="#ffffff").pack(
            anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Based on your description, {DEVELOPER_NAME} may need sample\n"
                 f"files to reproduce the issue in a test environment.\n\n"
                 f"The script itself is always attached automatically\n"
                 f"so {DEVELOPER_NAME} gets your exact version for testing.\n\n"
                 f"If you'd like, attach additional Excel or CSV files below.\n"
                 f"This is completely optional and files are sent privately\n"
                 f"alongside your bug report.",
            font=(MODERN_FONT, 11), text_color="#aaa",
            anchor="w", justify="left").pack(anchor="w", padx=16, pady=(0, 8))

        # File list area
        file_list_frame = ctk.CTkFrame(dlg, fg_color="#2a2a2a", corner_radius=4,
                                       border_width=1, border_color="#3a3f44")
        file_list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        file_list_label = ctk.CTkLabel(file_list_frame, text="No files selected",
                                       font=(MODERN_FONT, 11), text_color="#aaa",
                                       anchor="w", justify="left")
        file_list_label.pack(anchor="w", padx=10, pady=10)

        selected_files = []

        def _pick_files():
            paths = filedialog.askopenfilenames(
                title="Select files to attach",
                filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv"),
                           ("All files", "*.*")])
            if paths:
                selected_files.clear()
                selected_files.extend(paths)
                names = [Path(p).name for p in paths]
                display = "\n".join(names[:6])
                if len(names) > 6:
                    display += f"\n...and {len(names)-6} more"
                file_list_label.configure(text=display, text_color="#ffffff")

        def _submit():
            submit_btn.configure(state="disabled", text="Sending...")
            skip_btn.configure(state="disabled")

            def _worker():
                # Always attach the running script itself
                all_files = [str(SCRIPT_PATH)] + list(selected_files)
                ok, err = _post_bug_report_with_files(
                    description, case_num, all_files, reporter_email, category)
                root.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                dlg.destroy()
                n_files = len(selected_files)
                if n_files:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted with {n_files} file(s).\n\n"
                        f"Your report and files have been sent to {DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' here.")
                else:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted.\n\n"
                        f"Your report has been sent to {DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' here.")
            else:
                submit_btn.configure(state="normal", text="Submit Report")
                skip_btn.configure(state="normal")
                messagebox.showerror("Could Not Send",
                    f"The report could not be sent:\n{err}\n\n"
                    "Check your internet connection and try again.")

        ctk.CTkButton(btns, text="Browse Files", command=_pick_files,
                      fg_color="#3b82f6", hover_color="#2563eb", width=120,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        submit_btn = ctk.CTkButton(btns, text="Submit Report", command=_submit,
                                   fg_color="#28a745", hover_color="#218838", width=130,
                                   height=30, corner_radius=5,
                                   font=(MODERN_FONT, 11, "bold"))
        submit_btn.pack(side="left", padx=(8, 0))
        skip_btn = ctk.CTkButton(btns, text="Skip & Send", command=_submit,
                                 fg_color="#667788", hover_color="#556677", width=110,
                                 height=30, corner_radius=5,
                                 font=(MODERN_FONT, 11))
        skip_btn.pack(side="left", padx=(8, 0))

    # The bug icon button — placed in footer_row at the far right
    bug_btn = ctk.CTkButton(
        footer_row, text="\U0001f41e", width=34, height=28,
        fg_color="#1a1a1a", hover_color="#2a2a3e", corner_radius=6,
        font=("Segoe UI Emoji", 15), command=_on_support_click)
    bug_btn.pack(side="right")
    bug_btn.bind("<Enter>", lambda e: _show_tooltip(bug_btn, _support_tooltip[0]))
    bug_btn.bind("<Leave>", lambda e: _hide_tooltip())

    load_current_record()
    root.mainloop()

    # ==============================================================================
    # MASS DELIVERY PIPELINE EXECUTION
    # ==============================================================================
    if approved_emails:
        success_count = 0
        try:
            for data in approved_emails:
                body = html_template
                
                packages = data.get("Packages", [])
                manifest_date = data.get("Manifest Date (MM-DD)", "")
                package_table_html = build_package_html_table(packages, manifest_date)
                
                body = body.replace("{PACKAGE_DETAILS}", package_table_html)
        
                raw_reason = str(data["Reason for Request"] if isinstance(data["Reason for Request"], str) else data["Reason for Request"][0] if data["Reason for Request"] else "").replace("\n", "<br>")
                link_obj = data["Link Object"]
                if link_obj["url"]:
                    raw_reason += f"<br><br>This video may be helpful: <a href='{link_obj['url']}' style='color: #003366; font-weight: bold; text-decoration: underline;'>{link_obj['receiver_text']}</a>"
                
                highlighted_reason = f"""
                <table width="100%" cellpadding="0" cellspacing="0" border="0">
                    <tr>
                        <td style="background-color: #fff3cd; padding: 12px; border: 1px solid #ffeeba; border-left: 5px solid #d39e00; border-radius: 4px; font-family: Arial, sans-serif; font-size: 14px; color: #333333; text-align: left;">
                            {raw_reason}
                        </td>
                    </tr>
                </table>
                """
                body = body.replace("{REASON_FOR_REQUEST}", highlighted_reason)
                
                # Build subject: use first CBY# from packages for subject line
                first_cby = packages[0]["cby"] if packages else "N/A"
                subject = f"Urgent Attention Required: Customs Processing - CBY#{first_cby} / Package#{data['Subject Package Number']} ({manifest_date})"
                
                send_headless_smtp(
                    to_emails=data["To Emails"],
                    bcc_emails=data["BCC Emails"],
                    subject=subject,
                    body=body,
                    images=images,
                    password=o365_password
                )
                success_count += 1
                
            messagebox.showinfo("Success", f"Successfully dispatched {success_count} invoice request email(s).")
        except Exception as e:
            messagebox.showerror("Delivery Error", f"An error occurred during mass mailing execution:\n{str(e)}")

if __name__ == "__main__":
    process_queue()