# -*- coding: utf-8 -*-
"""
Unified XML Declaration Console
Launches a small picker window -> opens Air or Ocean console.
Both modes share the same UI framework; colours, header fields,
XML structure and Excel parsing differ by mode.
"""
import os
import sys
import re
import platform
import warnings
import json
import getpass
import urllib.request
import uuid
import subprocess
import traceback
from datetime import datetime
from collections import Counter
import io
import base64
import xml.etree.ElementTree as ET
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
from difflib import get_close_matches
import tkinter as tk
from tkinter import messagebox, filedialog
import threading
import tkinter.ttk as ttk
import time
import importlib

# ------------------------------------------------------------------
# Dependency check
# ------------------------------------------------------------------
REQUIRED_PACKAGES = {
    "customtkinter": "customtkinter",
    "PIL":           "Pillow",
}

def _check_and_install_dependencies():
    missing = []
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing.append((module_name, pip_name))
    if not missing:
        return True
    missing_names = [f"  - {pip_name}" for _, pip_name in missing]
    msg = ("The following packages are required but not installed:\n\n"
           + "\n".join(missing_names)
           + "\n\nWould you like to install them now via pip?")
    root = tk.Tk()
    root.withdraw()
    result = messagebox.askyesno("Missing Dependencies", msg)
    if not result:
        root.destroy()
        return False
    pip_args = [sys.executable, "-m", "pip", "install"] + [pn for _, pn in missing]
    try:
        result = subprocess.run(pip_args, capture_output=True, text=True)
        if result.returncode != 0:
            messagebox.showerror("Installation Failed",
                f"Could not install packages:\n\n{result.stderr[:500]}")
            root.destroy()
            return False
    except Exception as e:
        messagebox.showerror("Installation Failed", f"Error running pip:\n{e}")
        root.destroy()
        return False
    root.destroy()
    for module_name, _ in missing:
        try:
            importlib.import_module(module_name)
        except ImportError:
            messagebox.showerror("Still Missing",
                f"Package '{module_name}' still could not be imported\n"
                f"after installation. Please install it manually.")
            return False
    return True

if not _check_and_install_dependencies():
    sys.exit(1)

from PIL import Image
import customtkinter as ctk

# ==============================================================================
# RIGHT-CLICK CONTEXT MENU for CTkEntry fields (Cut / Copy / Paste / Select All)
# Monkey-patched globally so every entry in the app gets it automatically.
# ==============================================================================
def _ctk_entry_context_menu(event):
    """Show a Cut/Copy/Paste/Select All menu on right-click."""
    widget = event.widget
    # Walk up to find the underlying tkinter Entry inside the CTkEntry frame
    entry_widget = widget
    if not isinstance(entry_widget, tk.Entry):
        # CTkEntry wraps a tk.Entry — search children
        for child in widget.winfo_children():
            if isinstance(child, tk.Entry):
                entry_widget = child
                break
    if not isinstance(entry_widget, tk.Entry):
        return
    # Detect the background color from the widget hierarchy so the menu
    # blends with whatever console/window the entry lives in.
    bg_color = "#1a1a2e"  # sensible default
    w = entry_widget
    for _ in range(10):
        try:
            c = w.cget("bg")
            if c and str(c) not in ("", "SystemButtonFace"):
                bg_color = str(c)
                break
        except Exception:
            pass
        try:
            fg = w.cget("fg_color")
            if fg and str(fg) not in ("", "SystemButtonFace"):
                bg_color = str(fg)
                break
        except Exception:
            pass
        w = w.master
        if w is None:
            break
    _menu_font = ("Segoe UI", 11) if platform.system() == "Windows" else ("Helvetica", 13)
    try:
        menu = tk.Menu(entry_widget, tearoff=0, bg=bg_color, fg="#e8e8e8",
                       activebackground="#2e8b57", activeforeground="#ffffff",
                       borderwidth=1, relief="solid", activeborderwidth=0,
                       font=_menu_font)
    except Exception:
        menu = tk.Menu(entry_widget, tearoff=0, bg=bg_color, fg="#e8e8e8",
                       activebackground="#2e8b57", activeforeground="#ffffff",
                       font=_menu_font)
    menu.add_command(label="Cut",
                     command=lambda: entry_widget.event_generate("<<Cut>>"))
    menu.add_command(label="Copy",
                     command=lambda: entry_widget.event_generate("<<Copy>>"))
    menu.add_command(label="Paste",
                     command=lambda: entry_widget.event_generate("<<Paste>>"))
    menu.add_separator()
    menu.add_command(label="Select All",
                     command=lambda: entry_widget.select_range(0, "end"))
    try:
        menu.tk_popup(event.x_root, event.y_root)
    finally:
        menu.grab_release()

# Bind right-click on the CTkEntry class itself
_orig_ctk_entry_init = ctk.CTkEntry.__init__
def _patched_ctk_entry_init(self, *args, **kwargs):
    _orig_ctk_entry_init(self, *args, **kwargs)
    self.bind("<Button-3>", _ctk_entry_context_menu, add="+")
ctk.CTkEntry.__init__ = _patched_ctk_entry_init

# Selenium for COLS upload (imported lazily so app doesn't crash if not installed)
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.edge.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from natsort import natsorted
    _SELENIUM_AVAILABLE = True
except ImportError:
    _SELENIUM_AVAILABLE = False

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ==============================================================================
# EMBEDDED MBE LOGO (base64 PNG – no external file needed)
# ==============================================================================
MBE_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAABLgAAALaCAYAAAAySeo9AAAAAXNSR0IArs4c6QAAAARnQU1BAACx"
    "jwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAFOQSURBVHhe7d0L0CVVYSfwM+/3A7emEt3ZBYpF"
    "sTABMcTJIhlZUSDBAAsuJCADArsYLdlCYyglOxg1SgSzlo9VQMH3A1gkgI+IATQmUCzMsobFlaBO"
    "YGVZNg6Iw4wyw+yc5tyx+eZ+39xH9719+v5+VV1fn7739u139/1/p08HAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADIxK/0FoOF27JR6ASB7s3ZKvQAwNCcVgEwIuABokzoD"
    "runOmUI1gPZygAfIhIALgDapOmzq9zwp7AJoFwd1gEwIuABok6oCpmHPj4IugHZwMAfIhIALgDap"
    "Ilia7tw43bj7fT8A+XAgB8jEdBflAJCjYUOlqefFfsbX7Zw67PQAMF6z018AAIAsDBNuRfH9USoW"
    "uoVeAOTDfykAMuHCG4A2mRow9Wrq+bAznn7Ok+Xvnm58AORFDS4AACBLg4ZR5VBLoAXQDgIuAAAg"
    "C70GU/G1sjT4WcrjKptuOADNJuACAABaLeVchTToWaYbDkA+BFwAAEBWBFIATCXgAgAAGq/uWweF"
    "ZgB5E3ABAAATodeQrO4wDYDqCbgAAIBWiQFVN+nlghpbAO0i4AIAACaKcAugfQRcAABAq8QAqyMN"
    "KqRBwi2AFhJwAQAArVUOtKbepghAewi4AACAiTFdyFUerpYXQH4EXAAAQOMNUxNLYAXQfgIuAABg"
    "ovQbkAHQfAIuAAAgC1XW4ip/vtw/9X0A5EHABQAAZKkcTPWiW3jV7zgAaCb/nQDIhAtwANpkmJpS"
    "U8+Jg46rqvEAMH5qcAEAAFmZGkRNDap6IdwCaBcHcYBMDHLxDgBNVUWg1O3cuKfxDvIZAJrPgRwg"
    "E90uyAEgV1WFSsOeH4VbAO3gYA6QCQEXAG1SdbDU73lSsAXQLg7qAJkQcAHQJnUFTDOdL4VaAO3l"
    "AA+QCQEXAG0ibAKgSp6iCAAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAA"
    "ZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwA"
    "AAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDW"
    "BFwAAAAAZG1W+gtAw+3YKfVCz+68887wJ3/yJ+Ghhx4Kc+fODStXrgyzZ88Oy5cvDwsWLAiLFi0K"
    "S5cuLV5btmxZmDdv3rNei13sj6/F96xYsaJ4z5IlS3b7PEA/Zu2UegFgaE4qAJkQcNGPe++9twi2"
    "rrvuujSkfjH86oRn8+fPD4sXL+4annVeW7hwYdGVw7M5c+YU5U54FvvjsBjMdV7rfA7Im4ALgCo5"
    "qQBkQsBFLzZu3Bguuuii8JnPfCZs27YtDW2nGIDFsKxbeBa72B+HxWCsHJ51aqXFWmjxc+XwrBOs"
    "dV6Ln4/hW/wLVEvABUCVnFQAMiHgYiaPPvpoePe73x0+/OEPtz7YGpcYdMXgqxOedWqnxYCsc9tm"
    "Jzwrvzb1ltBOeBYDtXJAV/58HBZDN2gzARcAVXJSAciEgItufvrTn4ZLL7206DZv3pyG0hax5tnU"
    "Wmn9hmedz5fDs/j5OO74N47PbZ+Mg4ALgCo5qQBkQsBF2datW4vaWhdffHFRewuqEAOyqeHZ1IAt"
    "Duvc9jlTeFaulRZDtAMPPDB9CzxDwAVAlZxUADIh4CKKtx9eeeWV4U//9E+LJyNCLo444ojwwQ9+"
    "UNDFLgIuAKrkpAKQCQEXX/3qV8Mf/dEfFU9IhBzF2l1vetObwvr16zXcj4ALgEo5qQBkQsA1ub77"
    "3e+G888/P9x8881pCORt1apVRbtxp512Wgw50lAmjYALgCrNTn8BgIZ5+OGHw9lnnx0OOeQQ4Rat"
    "EtuNO/3008Phhx8eNmzYkIYCAAzOf00AMqEG1+T4+c9/Hi655JLwnve8x5MRGUq8JTA29B51GoSP"
    "Og3GR52nMkaxcfmo08B8FF+L74k6T2uMOg3LR52nN0axgflYMac8jtgofWyAPio/sTGOI05L5zUm"
    "ixpcAFTJSQUgEwKuyRDb2TrvvPPC/fffn4bQFJ0nDEadICjq9JdDnM4TBKNyEFQOk+J742fib/wY"
    "CkXlcZQDqXKwVA6TOuOIuk0HNJmAC4AqOakAZELA1W7xdsQ3vOEN4brrrktD2qvfWkXdQpxyraLy"
    "OMphUnl85XF0ahiVg6ByIFUOk8rjAKol4AKgSk4qAJkQcLVTXK1XXHFF8XTExx9/PA3tXQxiuoU4"
    "e6pVVA5xyqFQtyAodt1qGE03jnINo3iLWrxVrTwOgEjABUCVnFQAMiHgaqdYc+uOO+7YYzhVrrFU"
    "bsMIIFcCLgCq5KQCkAkBFwBtIuACoEqz018AAAAAyJKACwAAAICsCbgAAAAAyJqACwAAAICsCbgA"
    "AAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICs"
    "CbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAA"
    "AICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICszUp/AWi4HTulXir05JNPhle96lXhBz/4"
    "QZg/f36YPXt2WL58efHaokWLwoIFC541bMmSJWHevHlh7ty5YenSpcWwZcuWhTlz5hTD4+vlYXGc"
    "ixcvLoatWLEizJo1a9d4o5UrVxZ/O8Pi6/F9AG2383jntwgAlXFSAciEgKt627dvD8cff3y48cYb"
    "05BmiaHX1NAthmVxWAzPYogW9RuwxSAtjjuaKWDrhHndhgEMS8AFQJWcVAAyIeCq3utf//rw0Y9+"
    "NJXoR6zBNlOYFmu3xfeUQ7d+A7bOsG416MoBX2dY+buA5hNwAVAlJxWATAi4qnXxxReHCy64IJVo"
    "mxh0xYCtHLp1wrRuw7oFbAsXLiy6XgO28rBO6FcO8+I44riAZwi4AKiSkwpAJgRc1fniF78YTjnl"
    "lFSC0SrXfusEbJ1bT6Opw8oBWyeI6xamlWuwzTSsE9xFnVp15WEwKgIuAKrkpAKQCQFXNe64445w"
    "xBFHhC1btqQhQFmnBttMAVu3Wm3DPHihXNNtplp1tIuAC4AqOakAZELANbyNGzeGQw89NDz66KNp"
    "CJCTGHpNDd06QVwMz2KIFlURsHWGlQO2OK74+qpVq4oywxFwAVAlJxWATAi4hvPTn/40/Ot//a/D"
    "vffem4YA9C+GXjfccEN4xStekYYwKAEXAFXS0ikArbd9+/aizS3hFjCseHvziSeeWNzuDAA0h4AL"
    "gNZ761vfGr761a+mEsBwHn/88XDUUUcJuQCgQVQLBsiEWxQH8+lPfzqcfvrpqQRQnb322ivcdttt"
    "4dd+7dfSEPrhFkUAquSkApAJAVf/7rrrrnD44Yd7YiJQm9jg/De/+U0h1wAEXABUyUkFIBMCrv48"
    "8sgj4Td+4zfCQw89lIYAues8RTHqPD1xan98T3wCYrRw4cKii8pPSiz3l5+qWH7qYrm//ITGbv3x"
    "yYr77LNPMYzeCbgAqJKTCkAmBFy9e+qpp8IRRxwRvvOd76QhwNy5c3cFNuVQp9xfDm9mz54dli9f"
    "vlt/zCRWrFixW38Ug56OOLyTX5T74/jj98zUH6czBkxT+2mXnduE3yIAVMZJBSATAq7evfnNbw7v"
    "f//7UwmqMV2QM13/dIHN0qVLi7ApKg8v1zyarkZSueZRefh0NZXK4RU0jYALgCo5qQBkQsDVm2uv"
    "vTacdNJJqURTxMClE9iUw5vpgpxeAptyf/l2snJ/uUZSL/3x9/Z0QRZQLQEXAFVyUgHIhIBrzzZu"
    "3BgOOuig4hH+bTLdrWXl28bKIU0vt5ZF09U86vd2snJ/uUZSuR9gKgEXAFVyUgHIhIBrZtu2bQsv"
    "f/nLu7a71UuoU75tbLrApqpby8r909VUKtdCAmgjARcAVXJSAciEgGtmsWH5zZs3p5JbywCaTsAF"
    "QJWcVAAyIeACoE0EXABUaXb6CwAAAABZEnABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3AB"
    "AAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZ"
    "E3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAA"
    "AABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkLVZ6S8ADbdjp9TbSieffHK44447wvz588PixYuLYXPm"
    "zAnLli0r+qOVK1emvhCWL18eZs9+5v80S5cuDXPnzi36Fy5cWHTRvHnzwpIlS3brn2m8cXh8PYrv"
    "j5+LFi1aFBYsWFD0l6cRgMHM2in1AsDQnFQAMtHmgOuuu+4Kv/Ebv5FKeYnBWic4i4FbDN46ysFZ"
    "OYSL4VgMyaLpgrNyCDfTeHsJ5MrhXpyGOC3R1PECjJKAC4AqOakAZKLNAddrXvOacM0116QS47Ji"
    "xYrU9+zgbLpAbroacvE3a3lcsb/zO7aXQC7+jeWoHPSV+6PpxgvkQcAFQJWcVAAy0daA64EHHgj/"
    "6l/9q1SC4cVwrBPIlcO5XmvIVRXIDTteaDsBFwBVclIByERbA67/8B/+Q7jssstSCego3/5aDueG"
    "vWW1PK5yzbtyOBf1EsiV27wrB33lW2FhOgIuAKrkpAKQiTYGXA8//HDYb7/9wpYtW9IQoI1iWNYx"
    "Xdt0vTwsohycDRLIlUO4cn/5O+I4ytNbHi/VEnABUCUnFYBMtDHguuCCC8LFF1+cSgDNFkOwTghX"
    "vv11ulthy8HZILesdkK4tWvXhn333bcY1iYCLgCq5KQCkIm2BVyPPfZY2GeffcLjjz+ehgDQzWtf"
    "+9rwqU99KpXaQ8AFQJVmp78AMFIf+9jHhFsAPfj85z8fHnnkkVQCALoRcAEwctu2bQsf+tCHUgmA"
    "mcRj5kc/+tFUAgC6EXABMHLXX399eOihh1IJgD2JtV5/8YtfpBIAMJX73gEy0aY2uF7+8peH2267"
    "LZUA6hUbau80Aj+1sffyUx3j+2LD7tHU98X+OCwqv2/q0xzLDcaXG5+f6X3l8ZUbpo/K31vubwNt"
    "cAFQJScVgEy0JeC65557wsEHH5xKQNPEp/t1lJ8IOFMwtHDhwqKLpr6vHOT0+r5eA6SZ3te2MKiN"
    "BFwAVMlJBSATbQm4Xv/612tLhlaYO3fus2ralIOXQWruxPEtXbq06I96DXLK74vvie+NBn0fjIqA"
    "C4AqOakAZKINAdfWrVvDr/7qr3p64gQqh0FTa+7EUCe+HpWDnKnvK9fIqTrwmWl85e8t10AChiPg"
    "AqBKTioAmWhDwPXZz342nHbaaanEdGIQ1Al8ZrpFrBwM9XrrVznwmel9M31v+X3l742/Vcu3t5Xf"
    "BzCVgAuAKjmpAGSiDQHXv/k3/ybccsstqVSN6QKVckATldsLmi7wmSmg6bUm0CDvmxogAUwCARcA"
    "VXJSAchE7gHX008/XdTgKgc+saZSfGJYNFPQVH5fucYQAPkScAFQJScVgEy0oQYXAHQIuACokmcn"
    "AwAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAA"
    "WRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcA"
    "AAAAWRNwAQAAAJC1WekvAA23Y6fU23h33XVX+MAHPhBmz54dZs2aFVasWFEMnzNnTli2bFnRv2jR"
    "orBgwYKif8mSJWHevHlFf3w9vi9+dvny5cWw+Fp8T7Rw4cKii8qfi98RvwuAPOw8ZjtoA1AZJxWA"
    "TOQUcL31rW8N73vf+1Jp9DqB2n777VeEbdM59dRTw0033ZRK4zN//vzw/e9/P6xcuTINyVNc5+9+"
    "97tTabxuuOGGcPjhh6dSCPvss0947LHHUqnZ4vYQt4XFixeHVatWhec+97lh7733DgcccEA48MAD"
    "w4te9KIwd+7c9O7qPf7448X38UsbN27cdVyhOgIuAACYQDHgysULX/jCGMaNvdv5Iz1NUXfHHXdc"
    "18+No/vQhz6UpipPTz/99I59992367yNo7vlllvSlD1jxYoVXd+XY7dkyZIdRx111I5LLrlkx49/"
    "/OM0h9XZtGlT1++d5C4uE6q3c9kCQGW0wQVApX74wx+G++67L5Xo1eWXX5768vTXf/3Xxbqnfps3"
    "bw5f//rXw1ve8pbwvOc9Lxx99NHhG9/4RnoVAGAyCbgAqJQf2oO55557ZrydsumuuOKK1MeoxbDr"
    "Va96Vfit3/qtrLchAIBhCLgAqNTf/u3fpj76lWtItGnTpnDNNdekEuNy++23hzVr1oQ3v/nNYevW"
    "rWkoAMBkEHABUCkB1+A+//nPF7ef5eaTn/xk2LZtWyoxTnE9vP/97w+//du/HR5++OE0FACg/QRc"
    "AFTm0UcfDffff38q0a/49Lqrr746lfLh9sTmufPOO4tbFh944IE0BACg3TyaFyATOTxx6qabbgrH"
    "HntsKo3f3nvvHX70ox+l0u6OP/74cP3116dSMxx22GHhb/7mb1Kp+e64447itrimueWWW8LLX/7y"
    "VAph5cqVRYA4aVavXh3+7u/+rvjbq8ceeyzstddeqUQUb8ON29BUP/3pT8Ov//qvF/2LFy8O8+fP"
    "L/qXLl0a5s6dW/QvW7YszJkzp+hfsWJFmDVrVtHF/ii+Ft8TzZ49OyxfvrzonzdvXliyZEnRH8cb"
    "xx8tWLAgLFq0qOiPf2M5iu+Nn4n+2T/7Z7vG2WQ7l4PfIgAAMGmKZ6o33J/92Z8969H64+723nvv"
    "NGXdHXfccV0/N+7uvvvuS1PYfGeddVbXeRh3d8stt6QpfMaKFSu6vm8SuoMOOmjHli1b0pLYs02b"
    "NnUdzyR3cZl009Rldfrpp6cpbLad0woAlfFfE4BM5PBj4Pd///fDF77whVQavxxrcEXnn39+uPTS"
    "S1OpuZ544onw3Oc+t5HthqnB9WznnHNOuOyyy1JpZmpw7W66Glx1LqtYIyvWAovf26nZ1anttXDh"
    "wqLr1ODq1Bjr1Bb7tV/7tXDiiSemMTWXGlwAVMlJBSATOQRcL3rRi8K9996bSuOXa8C1atWq8NBD"
    "D+263ampLr/88vDv//2/T6VmEXDt7q/+6q/CK1/5ylSanoBrd9MFXDHcPfXUU3cFUDFkimHT1ACq"
    "EzzF2xLj33JY1bkVsXPrYflWw7YTcAEAwASKAVeT/fznP9+x88fdrltkmtDleoti7L70pS+lqWyu"
    "Qw89tOu0N6Fzi+LuXbxVcfv27WmJTM8tirt3092iyHB2LlsAqIynKAJQiR/84Adh27ZtqcSwer2d"
    "bFy++93vFk/qIx/33HNP+PznP59KAADtIuACoBL3339/6qMKN998c9i4cWMqNU/TAzi6+9CHPpT6"
    "AADaxX3vAJlo+u0csVH0t7zlLanUDHtqg+uLX/xiuO+++1Kpu7jYr7766j2+rw4XXnhheOc735lK"
    "zbF169bwvOc9r2iXaNTOPffc8Cu/8iupNL0zzjgj7LPPPqkUwsUXXxy2bNmSSt3F+YrvG9ZBBx1U"
    "tO/Wq7iNxfbBfvaznxV/Y1gc27Krq0bk//gf/6NohHw6cTm8973vTaXpVbW8cjBdG1wMRxtcAFTJ"
    "SQUgE00PuN7znveEa665pmg0uazTuHJHp9Hljk6jzGVTf0hOHUenMeaOTmPNZbEcx/vSl740DRnc"
    "uBqjX716dRHQlee9CT772c+G0047LZVGa8OGDeHggw9OpWpV1bj6unXrwlVXXZVKg3nyySeLWnyx"
    "If8bb7wxDa3Gf/pP/ym84x3vSKXBVbW81q5d+6wHAlTp1ltvDbfddlsqDU7AVQ8BFwAATKAYcDEe"
    "42yM/sYbb0xT0Rxr167tOq2j6DZs2JCmonpVNa6+bt26NMZqfOtb39qx//77d/2uQbrDDjssjXk4"
    "VS2v9evXpzFWL46723f228V5pXo7ly0AVEYbXADQYB//+MdTXzM88MADldSIoXeHH3540aB/rOlU"
    "hTvuuKO4vRAAoE0EXADQYDfccEN4+OGHU2n84i1zjN6KFSuKWxUPPPDANGRwsW0vD4UAANpGwAUA"
    "DRbDiE996lOpNF5PPfVU+MQnPpFKjFpsUy62fxbbrRvW97///dQHANAOAi4AaLh4m2ITmqu56aab"
    "wqOPPppKjEN8QmMVDfw/+OCDqQ8AoB0EXADQcPF2sia0e9W09sAm1bnnnpv6Brd58+bUBwDQDgIu"
    "AMjAuNu+ijV+vva1r6US4/Sbv/mbYdWqVak0GI3MAwBtI+ACgBEYtt2k6667LvzkJz9JpdG78sor"
    "i/bABlVFu1E8Y9asWcWtisOYM2dO6gMAaAcBFwCMwJIlS8LatWtTqX9btmwpGhgfh6effnroxuWP"
    "Pvro1EcV/vk//+epbzCLFi1KfQAA7SDgAoAROeuss1LfYMbVBtY3v/nNsHHjxlTqX7yd7nd+53dS"
    "iSoMWyPuec97XuoDAGgHARcAjMiJJ54YVqxYkUr9u+eee8Idd9yRSqNz2WWXpb7BnHzyyWoMVeyJ"
    "J55IfYN5/vOfn/oAANphVvoLQMPt2Cn1MmLHH398uP7661NpMDHYeuyxx8Ib3/jG8OEPfzgN7d85"
    "55wzdODUj0cffbSo7TNM+1t33313Ec6deeaZacjgNmzYEA4++OBUqlZcP3vttVcqDW7dunXhqquu"
    "SqV6xGUQl+kgYu2vTZs2haVLl6Yhg6lqea1fvz5cdNFFqVSt733ve0U3rN/93d8N8+bNSyWqMis2"
    "KAcAAEyWGHAxHscdd1wMF4fqVqxYUYzrv/23/9b19V67JUuW7PjZz35WjGsULrnkkq7T0Wt30EEH"
    "FeO58soru77eb7dhw4ZifHXYtGlT1+/st1u3bl0aYz3idM6dO7frd/fSHXLIIWlMw6lqea1fvz6N"
    "kUmzc/0DQGXcoggAI/SSl7wkHHLIIanUv82bN4cvfOELqVS/Ydv9GrbdMXZ37bXXDlWj7tWvfnXq"
    "AwBoDwEXAIzY2WefnfoGc8UVV6S+en37298O9913Xyr1L94Kd8opp6QSVYjB1qWXXppKg7FOAIA2"
    "EnABwIj9/u///lCNrt9+++3h3nvvTaX6DFt7K7ZdFp+gSHViuDVM6HjYYYeFAw44IJUAANpDwAUA"
    "I7Zy5cpw0kknpdJgLr/88tRXj5/+9KfhS1/6UioN5owzzkh9VOGmm24KF154YSoN5q1vfWvqAwBo"
    "FwEXAIxBfBriMD73uc+FrVu3plL1PvvZz4YtW7akUv+e+9znhqOPPjqVGFasTRdrxA3T9lasvaX9"
    "LQCgrQRcADAGL3vZy8L++++fSv179NFHw/XXX59K1Ru2na/TTjstzJkzJ5UY1F133RVe+cpXFu22"
    "DRNuxfbQPvShD4VZs2alIQAA7eIqByATTX2k+o9+9KNw7rnnhoULFxbl2LbUggULiv5ly5YVIUf8"
    "Ub1ixYpi2Lx588KSJUuK/vi+TltUnfdG8b3xM7Ech0flz8W/sRyVP1eXWHNm2DApztNjjz2WSs+4"
    "+OKLwwUXXJBK/TviiCPCX//1X6dSdTZs2DDUkx6j2E5Uua2nq666Kpx55pmpNLg4bQcffHAqVSuu"
    "n7322iuVBrdu3bpifvv11FNPhUceeaRYdn/3d38XbrzxxnDnnXemV4fzrne9K7z97W9PpWpUtbzW"
    "r18fLrroolRikuw8zvstAgAAkyYGXE30t3/7tzF4a0S3YsWKolu1atWOvffee8dhhx2WpnI4xx13"
    "XNfv66eL0zXVj3/84x1z587t+v5eu3/4h39IY6vOG97whq7f1Wu3Zs2aNKZfuvLKK7u+t99uw4YN"
    "aYzV27RpU9fv7LeL67SzLfbaLVmypOu4quiOPfbYHdu3b09zWZ2qltf69evTGJk0O9c/AFTGLYoA"
    "DOUnP/lJ6hu/xx9/vOji7XsbN24MDz30UHqlmWI7VcO2iTRITaGZPPnkk+Ezn/lMKg2mippaOYu3"
    "Ena2xV67zZs3p09Xa82aNeGLX/ximD3bJR8A0G6udgAYSpMCrhydddZZqW8wsfHxYdpmmuraa68t"
    "ApdBxVtOTz755FRinI488shw8803h8WLF6chzfSOd7yjuCV5FN2Xv/zl9K0AQNsIuAAYyj/90z+l"
    "PgYRnzS4evXqVOrfww8/HL761a+m0vA+9rGPpb7BnHDCCbvaW2N83vSmN4WvfOUru9qtAwBoOwEX"
    "AEOZ2nA6/YkN5J9xxhmpNJjLLrss9Q3nf/2v/xW+853vpNJgXve616U+xmHvvfcugq0PfOADux7E"
    "AAAwCQRcAAxFwDW8YUOhr33ta0VNrmHF2x2HEcOV+GRHRi/Wmnvve98bvve974VjjjkmDQUAmBwC"
    "LgCG8vOf/zz1Mah99923aC9pULENrk984hOpNJhf/OIXQzdYv27dOo2Zj0lsN+2WW24J3/72t9MQ"
    "AIDJMiv9BaDhmvpI9Xh73Sc/+clUapZYo+hHP/pRKg3u+OOPD9dff30qDSbWsJmptlt80t0pp5yS"
    "Sv2LIdkDDzxQNKQ9iNi4/EknnZRKg/nBD35QTEc3MTyr4umKGzZsCAcffHAqVSuun7322iuV8rV2"
    "7dpw8cUXh5e+9KVpSD1yXF7XXXddsT9X5YknnnjWk1DjbaHlds+WL1++K/SdO3duWLp0adEfxfd1"
    "biON74nv7YgPBpg/f34qhbBy5crUF8KCBQuKhzl0xGNLZ7/v9v1xPOXvbZKd0+23CAAATJoYcDXR"
    "unXrYvDWyG7vvfdOUzmc4447ruv4++l2/ghNY+tuy5YtO1atWtX1s712N998cxpb/4466qiu4+y1"
    "O+KII9KYurvyyiu7fq7fbsOGDWmM1du0aVPX78y1O/fcc3c88cQTae6ql+Pyuu6669LUVyOHZbB2"
    "7do0tc2zc/oAoDLuIwCABli4cGE49dRTU2kwV1xxRerrz8aNG8PXv/71VBpMFbWzqNZHP/rRcMgh"
    "h4S77rorDWESxdpgADAJVAsGyERT/9v9mte8JlxzzTWp1Cw53aIY3XvvveFFL3pRKvUv3gL1yCOP"
    "hOc85zlpSG8uuuii8I53vCOV+hfn7X//7//9rFujpnKL4vjE29muvPLKcPLJJ6ch1XCL4jNtn8Xj"
    "TPmWws7thLHcCZeWLVtWPDG1fCtiXC/xdsN4l17ch8rDos5nyrcddm5rjMPj61H5M92+u8ncoggA"
    "ABOouJ+jgaq4fa+uLqdbFDvWrFnT9fO9dn/xF3+RxtSb7du371i9enXXcfXanXXWWWls03OL4vi7"
    "j3zkI2lOq+EWRYa1c50AQGXcoggADXL22WenvsH0e5tivDXxoYceSqXBuD0xD3/4h3/Y2AdCAAAM"
    "S8AFwFA6TwijGvE2splu9duTeJvjHXfckUp7dvnll6e+wey///7hsMMOSyWaLgaot9xySyoBALSH"
    "XyUADKX8aHuGFx/nf8opp6TSYHoNrWJ7XTfccEMqDeass85KfXSsW7euCJH21P3VX/1V0SZUp4vt"
    "lF1yySXhvPPOC0ceeeSudpmqtG3btiJEffjhh9MQAIB20LAjQCaa2l7JG9/4xvDhD384lcYr1nyK"
    "Da137LfffpU8QW5Ujcx3xBpYa9asSaX+xeUQA4xOI9TTufjii8MFF1yQSv2Ly/qHP/xhWL16dRoy"
    "vUlqZH79+vVFw/3D2r59e7j99tvDpz71qfDpT386bNmyJb0yvBNOOCH81//6X1NpMFUtr6OOOioc"
    "ffTRqVSvV7/61cVxgWbQyDwAVXJSAchEUwOuWBPlnnvuSaXiB8tuNU86TwPriE/8ik/+6ig/Eaxj"
    "T58pP42sbqMOuKL4NMV4u+GgLrvssnDOOeek0u7i5vSCF7wg3H///WlI/4455pjwla98JZVmJuAa"
    "Tgws3/KWt4TPfe5zacjwbrzxxvC7v/u7qdS/Ji8v8iDgAgCACRQDLsZjlE9R7IhPQ+w2nl67Qw89"
    "NI2pu1tvvbXr5/rpvvSlL6Wx7dkkPUVx/fr1aYzVi8tx7ty5Xb+33+7AAw8snqI5qByWF822c/0D"
    "QGW0wQUADXT66ac/63bLft15553hu9/9birt7uMf/3jqG0ysufN7v/d7qcSonHHGGeHaa68datvo"
    "iDUEY9tfAABtIOACgAZ6znOeE0466aRUGky8TbGbeGvZNddck0qDOfXUU4vbRhm9GCy+973vTaXh"
    "fPCDH0x9AAB5E3ABQEOdffbZqW8wn/3sZ8PWrVtT6ZeqaLA81iRifM4///xw2GGHpdLgbrvttrBx"
    "48ZUAgDIl4ALABrqiCOOCPvuu28q9W/Tpk3F7WxTDXt74kEHHRRe8pKXpBLjENvmfs973pNKwxn2"
    "aYoAAE0g4AKAhopPihy2ptTll1+e+p5x1113Peupl4NQe6sZDj/88HDggQem0uC+8Y1vpD4AgHwJ"
    "uACgwV73utcN1aB4vAXtgQceSKXp2+XqVZyW2P4WzXDCCSekvsF961vfik9pTSUAgDwJuACgwVav"
    "Xh2OPvroVBpMpxbXz372s/DFL36x6B/Uq1/96rBq1apUYtxe9rKXpb7Bbd68OfzgBz9IJQCAPAm4"
    "AKDhzjrrrNQ3mE984hNh+/btRbj1+OOPp6GDOfPMM1MfTXDAAQekvuF8//vfT30AAHkScAFAwx17"
    "7LFD1Zp69NFHww033BCuuOKKNGQwcRqOOeaYVKIJVq5cmfqG8+CDD6Y+AIA8CbgAoOFiu1fD1py6"
    "8MILw+23355Kg4mNyw/THhjVq2p9xBC0jS666KLiiZPDdo899lgaIwDQVAIuAMjA2WefnfoGc++9"
    "96a+wa1bty710RQ/+clPUt9wnnrqqdQHAJAnARcAZGD//fcPa9euTaXRW7NmTTjwwANTiab4n//z"
    "f6Y+AIDJJuACgEwM29j8ME4//fTUR5N861vfSn3DWbRoUeoDAMiTgAsAMvGa17wmrFixIpVGJ4Yf"
    "f/AHf5BKNMWOHTvC1VdfnUrDWb58eeoDAMiTgAsAMrFw4cJw2mmnpdLoHHfccWMJ1pjZV77ylXD/"
    "/fen0nD+5b/8l6kPACBPAi4AyMg4blN83etel/poim3btoU//uM/TqXhHXDAAakPACBPAi4AyMiL"
    "X/zicMghh6RS/VavXh1e8YpXpBJN8ba3va2SJ2NGS5YsCfvuu28qAQDkScAFQGW+/OUvh6uuuqro"
    "rr322qIcu1tuuSXceuutRbdhw4bw3//7fy+6H/3oR0W3cePG8NhjjxXd448/nsbGdM4+++zUV78z"
    "zjgjzJ7tcqFJ/st/+S/hfe97XyoN77d/+7etYwAAAEZjRwaOPPLIHXFSq+xWrFhRdHvttdeOvffe"
    "u+j23XffHQcddFDRHXLIITvWrl1bdPH7jzvuuKI76aSTdpx//vlpyoYTx9dt2vrp4jxU5bHHHtux"
    "aNGirt9TdfcP//AP6VuHc+WVV3Ydf7/dhg0b0hirt2nTpq7f2W+3fv36NMZqPfXUUzv++I//uOt3"
    "DtO9//3vT9/Qn6YvryiOu9t39tvFeaV6O5ctAFTGv+sAqEwd7fjEGl2x2/kDs6jpFbsf/vCH4Z57"
    "7im6u+++O9x2221Fd/PNN4frr7++6K655pqiFlkbxQbf4xMV67Z27dqw3377pRLj8vTTT4ebbrop"
    "HHzwweHiiy9OQ6vze7/3e6kPACBfAi4AKvP85z8/9VG3UdymeOaZZ6Y+RikGWg888EC47rrrwn/8"
    "j/8x7L333uHYY4+trM2tMiEmANAWs9JfABouh9s5vvGNb4RXvepVqTR+MRiIbXxN59FHHw2bN29O"
    "pemdc845Re2wYcRaV7H9sVmz9nzqXb58eXjOc56TSt3FzeEFL3hBuP/++9OQasWGxx955JHi70y2"
    "bt0a/s//+T+pNL1Ym+4tb3lLKg3uK1/5SnjhC1+YStP71V/91bBw4cJUCkXNvz3tQk888UT49V//"
    "9VQaXNzu9tlnn1TqTWx/7sknnwz/7//9v2I64lMSR+Ezn/lMOPXUU1PpGTFg+8d//MdUmt44l1ev"
    "Om38DSvWIF25cmUqUZWdx0O/RQAAYNLEgKvp/umf/mlXmzVN6GJ7XTOpom2tOrrzzjsvTeHM/vzP"
    "/7zr56vozjrrrPQtM7vlllu6fn7cXZyustgGWrf3TXK3//77F+16TVVV21pt6rTBVY+dyxYAKuMW"
    "RQAqE2sd7fzRnErU7fTTTw9z585NpWqtW7cu9dFWb3vb22rbfgAARk3ABUClDj300NRH3X7lV34l"
    "vPrVr06l6sSQ8mUve1kq0UZr1qwpAlIAgLZw3ztAJnK5neODH/xgeNOb3pRK47WnNriOP/744omL"
    "TXPeeeeF//yf/3MqzeyrX/1q+J3f+Z1Uqsa73vWu8Pa3vz2VZnbrrbeGI444IpWa45Zbbgkvf/nL"
    "UykU7SfFp3ESilpb3/3ud6d96mlsD2yvvfZKJaLp2uCK21Q8zkSzZ88u2s+b2j9nzpywbNmyon/e"
    "vHm72rUr9y9YsCAsWrSo6I9tx3Xaj4vD4mvR4sWLw/z584v++Ln4+an98Xv23XffPbbh1xTa4AIA"
    "gAlUNFiSgb//+79/Vts14+za3gZXtG3bth2rV6/uOp5Burlz5+548MEH09j3TBtc+XWXXHJJWird"
    "aYNr9266NriauKz+5m/+Jk1d8+2cXgCojFsUAajUgQceGJ773OemEnWLtUPOOOOMVBreK17xirB6"
    "9epUom1OOeWUcP7556cSOYo18OJTWWO3atWqogZZ7OLTReOtp7EDgEmkWjBAJnL6b/frX//68NGP"
    "fjSVxmcSblGM4jzG25Kq8IUvfCGcfPLJqbRnblHMR1xPX/va13bd5jYdtyjubrpbFLdu3RouuOCC"
    "oj8u13gbYVS+5bDXWwtjWB11vifevRdDrKh8m2ObuEURAAAmUHE/RyZuvvnmZ90yM65uEm5R7Djy"
    "yCO7jqufLt7Gt/MHexpjb9yimEd31FFH7di8eXNaGjNzi+Lu3XS3KDKcncsWACrjFkUAKhdrzrhN"
    "cbTOPvvs1De40047bVeD1rTHOeecE/7yL/9yV+0iAIA2EnABULl4O81rX/vaVGIU4u2WsT2eYZx5"
    "5pmpjzaIt8BdccUV4bLLLtvjbYkAALkTcAFQiypqFNG7WPNqmFDxoIMOCi95yUtSidwde+yx4e//"
    "/u/DWWedlYYAALSbgAuAWuy///7Fj+xRi7VWOk8Y69YodJu97nWvS339U+OuHY466qiigf0bbrgh"
    "7LPPPmkoAED7eXIJQCZybJD33nvvDbfffvuup4NF5aeFRf2Wly9fHmbP/uX/Z6aW+9GWpyiW/dZv"
    "/VaxzPsxd+7c8OMf/3igWxw9RXH8Ypgct+XY1lbsH5anKO5uuqcoMhxPUQQAgAlUPHKKSrXpKYod"
    "V1xxRddxztSdcMIJ6dP98xTF0XarVq3asWbNmh3nnnvujssuu2zH97///TSH1fEUxd07T1Gsx85l"
    "CwCV8V8TgEz4MVC9b37zm+HBBx9MpeZ44QtfGF760pemUn82b94crr766lTqzZo1a8IBBxyQSv15"
    "+OGHw9e//vVUao54q175SZ6f+9znwi9+8YtUaq5YoSXeXtsR+2Mtxdj9i3/xL8LChQvTK/WJyyku"
    "L37pD/7gDzTUXwM1uACokpMKQCYEXAC0iYALgCppZB4AAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4"
    "AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMjarPQXgIbbsVPqzd7TTz8d/vEf/zHM"
    "mjUrrFixIg0NRX8cBkD77TzeO+ADUBknFYBMtCngip7//OeH+++/P5V2t2TJkjB37tyif/HixWH+"
    "/Pm79ce/sRzNmzev+Ew0e/bssHz58qI/Kgdny5YtC3PmzCn64/vj56Jexhs/Fz/fsXLlytQXiu+L"
    "3xstXbp017QD0J2AC4AqOakAZKJtAdfFF18cLrjgglRqr0WLFnUNzsr9gwRn5XAufseCBQuK/vg3"
    "lqPyeAcJ/RYuXFh0AHUQcAFQJScVgEy0LeB65JFHwurVq8O2bdvSEJosBmudmmzlcG5qcFYO56YL"
    "zqYL5HqtLVceb7m/l/ECzSHgAqBKTioAmWhbwBWdfvrp4dOf/nQqQf3KQV1UDs7K/eXacuXbV8sh"
    "2nS18OJv9unalusl9CsHcnEa4rREU8NEyJ2AC4AqOakAZKKNAdedd94ZfvM3fzOVgH6UQ7RycDZd"
    "IFe+5bQcok0N5OqohTf1O8q33jK5BFwAVMlJBSATbQy4ole+8pXh5ptvTiVg0sRArFttuXJ/ubZc"
    "uSbbIG3LlQO58ninC+ei8njL4Vx5vPRPwAVAlZxUADLR1oDrG9/4RnjVq16VSgB5iqFZt0Cu3D81"
    "OCvXlisHZ51bZGP+8853vnPXbaptI+ACoEpOKgCZaGvAFcXbFOPtigD80nHHHRe+/OUvp1L7CLgA"
    "qJLGDwAYu7e97W2pD4CO9evXpz4AYE8EXACMXaylcNBBB6USAPG4+OIXvziVAIA9US0YIBNtvkUx"
    "irfhnHDCCakEkL/Yjlan4fqpDeKX29/qtLkVdZ52+Y53vKP1wb9bFAGokpMKQCbaHnDF2Yu1Fe65"
    "5540BJhE8YmFHeUnIfbyVMWo3Fh7bPg9PjUxKj89caawqfyd5c+Un7IYc5nydJa/szydzEzABUCV"
    "nFQAMtH2gCtSiwv6N93T+2YKccohTDkEirWIyk/smy746TVsKn8mhkMxJIrKwdHU72RyCLgAqJKT"
    "CkAmJiHgUouLUSjfNhZNd6tYOfgp197pNcQpfya+Ht/XEWv/dH7blz8z3XfG905XswlyJeACoEpO"
    "KgCZmISAK1KLq5li0NIJfqarvdNriFOuvTP1M+WwqZfgZ+rny8FP+fOddo2A5hBwAVAlJxWATExK"
    "wNWWWlzlWkIz3SrWrXHpqNfaO+XgZ6bgaLqwqfydcTo6t4pN/U6Aqgm4AKiSkwpAJiYl4Ir+8i//"
    "MrzhDW/YFeL02t5P7I/DonJwVA5+ZgqbpguOZgqbyt9Znk4AZibgAqBKTioAmZikgAuA9hNwAVCl"
    "Z/7lDAAAAACZEnABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3AB"
    "AAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZ"
    "E3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkLVZ6S8ADbdjp9TLDK677rrwyU9+"
    "MixZsiTMmzevGDZ37tywdOnSoj9atmxZmDNnTtEf3xPf27FixYowa9Yzp8dFixaFBQsWFP2zZ88O"
    "y5cvL/qj8mtxXHGcHeXvnvoaAM/Yeaz1WwSAyjipAGRCwNWbrVu3hjVr1oR77rknDWmWGH7FwC2a"
    "P39+WLx4cdEfrVy5MvXt/tp0wVscFl/rWLhwYdFFU0O5mYK38mtTQ79yIBi/N35/R3m6APoh4AKg"
    "Sk4qAJkQcPXu/vvvD4ceemh4/PHH0xBGIQZfMZjrKAd2MazrvDY1XCu/NjWUizXvpgsEew3eyq9N"
    "DQRnCgvLoR9QPQEXAFVyUgHIhICrP/FWxX/7b/9tKsFwYsg2Xa22qcFbOdibWhuuHLzNFK6VX5up"
    "tt3U18rTNfW7Y3AYA8Ro6jTDOAi4AKiSkwpAJgRc/Xv7298e/uzP/iyVgOmUw7WZArty8DbTLbD9"
    "tHs3U1g4SCA4dbpoLgEXAFVyUgHIhICrf3GRnXjiiUVtLmAyxfArBmfRoLfATlcTL7Z3d9FFF+1q"
    "947+CLgAqJKTCkAmBFyDefLJJ8Phhx8e7r777jQEoBqf+tSnwmtf+9pUol8CLgCq9ExDDADQUrFG"
    "xo033hhWr16dhgAM74/+6I+EWwDQIP5rApAJNbiGs2HDhqIm1+bNm9MQgMEcc8wx4YYbbth1CyOD"
    "UYMLgCqpwQXARHjxi18crr766l3t7QAM4oUvfGH43Oc+J9wCgIbxXxOATKjBVY2rrroqnHnmmakE"
    "8GydpzOWG52Pf2M5dp/+9KfDC17wgmI4w1GDC4AqOakAZELAVZ13v/vd4cILL0wloE4zBUZR+UmG"
    "K1euLP6Wn1oYn4K4YMGCor/8ZMM9vTcOi69F3d4b3xffH5WngdERcAFQJScVgEwIuKr1xje+MXz4"
    "wx9OJchHHYFRzBk6463yvTATARcAVXJSAciEgKta27dvD695zWvCddddl4YwqfYUGHVCotmzZ4fl"
    "y5cX/XsKdsrvLdcU6vbe2MX+qDze+DeWI4ERbSTgAqBKTioAmRBwVW/r1q3hyCOPDN/5znfSEIYV"
    "aw51App+AqNyCFSugVRXYBTHFccJjI+AC4AqOakAZELAVY9NmzaFww47LNx3331pSLNUHRjF2krx"
    "N+V07y2HQJ33lgOjhQsXFl3U7b0AvRJwAVAlJxWATAi46vPggw+Gf/fv/l3YsmVLUd5Tu0V7auOo"
    "HBiVby3rvLeXwKjzXoC2EnABUCUnFYBMCLgAaBMBFwBV0vgEAAAAAFkTcAEAAACQNQEXAAAAAFkT"
    "cAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAA"
    "AFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEX"
    "AAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFmblf4C0HA7dkq9tMT3vve9cO6554b/"
    "+3//b5g/f34aGsLKlStTXwgLFy4suo7ly5eH2bOf+f/U3Llzw9KlS4v+aNmyZWHOnDlF/7x588KS"
    "JUuK/ij2x2FRfE98b8dMry1atCgsWLCg6J81a1ZYsWJF0R9NnTaAfuw8pvgtAkBlnFQAMiHgaqfN"
    "mzeHt7/97eEDH/hAGpKvcvgVA7vFixen0rNDu6mvzRTMxQAvBnnRTOHbTK/FQDAGgx3l16I43Z3f"
    "2XF4fL2jHCgC1RJwAVAlJxWATAi42u2OO+4If/iHfxjuvvvuNISmiUFbOXwrB3NTQ7tyMDdTwBZ/"
    "309XK27qa1PDt5mCuZlCw5lqAcIoCbgAqJKTCkAmBFztt3379vCRj3wk/Mmf/El4/PHH01AYnRiE"
    "TRfMxQCvcyvt1Bpz5dcGrTEXx3nKKaeEVatWFWXaT8AFQJWcVAAyIeCaHI888khx2+LHP/7xNATa"
    "be3ateHSSy8NL3nJS9IQJoGAC4AqOakAZELANXnuuuuucN5554XvfOc7aQi0y7777hve9773hRNP"
    "PDENYZIIuACoklZTAaChYm2Wb3/72+ELX/hC2HvvvdNQyF+8TfHP//zPiyeJCrcAgCr4rwlAJtTg"
    "mmxbt24tAoH3vve9YcuWLWko5CW273X22WeHP/3TP9XWFmpwAVApJxWATAi4iB5++OGiEfpPfvKT"
    "Ydu2bWkoNN8xxxxT3I544IEHpiFMOgEXAFVyUgHIhICLsnvvvTdccMEF4cYbb0xDoHqxxlV8smLn"
    "KYkrV64shse/8+bNK17rPCUxPlUxvj/efth5rfOUxP322y8cfvjhxWehQ8AFQJWcVAAyIeCim1tv"
    "vTW8+c1vDnfffXcaQpvFsCgGTTFMiiHS8uXLi3IMoDohVBw2Z86cImjqvNYJoWLoFD/XLaDqvBY/"
    "F4Oq+B1QJwEXAFVyUgHIhICL6cRN4/Of/3y48MILww9/+MM0lFGIYdDs2bOfFTSVazR1QqgYGC1d"
    "ujQsXLiw6OJrMYSKQVP8G8ud2k6d18pBU+dz0CYCLgCq5KQCkAkBF3vyi1/8InzsYx8L73znO8Oj"
    "jz6ahk6Ozu10U2s0xRCpE0JNrbUUw6Opt9V1q+00NYTqvAYMTsAFQJWcVAAyIeCiVz/72c/CpZde"
    "WjTovXnz5jR09DqBUzloirrdHtctaJpao6kcNHVqNHWCpvg5v5UhLwIuAKrkpAKQCQEX/Yq1uN71"
    "rneFj3zkI8UTF2NYFIOmGBDFUEj7TcA4CbgAqJKTCkAmBFwMavv27UVoBdAkAi4AquSkApAJARcA"
    "bSLgAqBKs9NfAAAAAMiSgAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4"
    "AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACA"
    "rAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsA"
    "AACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMia"
    "gAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYsVnpLwAAAMBE2LFT6p3W"
    "rJ1SLxmwsgAAAICJ0EuwNZWgKw9WEgAAANBqMwVb5QCr1/fRPFYOAAAA0FrdQqtewqpBP8d4WDEA"
    "QCN0u4jsxoUlAAyu1/NtR+7n3anzO8j8VDEO6melAAAjNfUisSouNqF6g+6v9kdohkk/506d/850"
    "97Jcps5j+TNTX6MZrBQAoFa9XETWwcUn9K+O/dW+CKOT0zl32Gnt5TvL31F+fz/fPd3nysNpBiuk"
    "T/3sCHVrww7VpOUZjXqZNmH+c9uOmrbNlJWXZVOnM7f1vSdN3h4G0ab107R1U9eybfI22O885zIv"
    "TZ7Ocel3XXcziuU6dToH/c4q5nc6o1gOuem2vC2nPatzO+2mKeukn/kedpr39F1Tx19+f7/f3fls"
    "+XPl8dEMs9NfMtTvTtk0uU8/QNXicbEjDcpOmvzGTX+aLOcdmCLtGiPZN0b1PTApip03SYPGLk1O"
    "IQ1qhJnCqPhaN+nlZykPb9o8IuACgEYqrgyTNKjR0qQ2flrTZLogZeKlXWHk+8I4vhPapth5d0rF"
    "xsphGmcyXchFcwm4Mpf7QYPxsw1B88X9NErFRkmTlt1xJE224x8Tadzbvn0PBhP3nSgVs5AmeeTT"
    "XP7OQYOq6aZb8NVcAi7GYrqDBQDTa9KxM05LlIrZSrPhnMTEaMr2br+D/uS+z+Q2/bkv70kl4AIc"
    "wCEjTdhf23jMaOM8wVS2c8hP3G+jVMxampXGzUuarGdJL5EZAVcL2AEBJsu4jvvxe6NUbJ00e86p"
    "tJJtG/LT1v02p/mKtyNGqUjDCbgYubYeqHNnvUBeRr3PTtIxYpLmlclgm4b8tH2/bdL8pQxrtxCr"
    "2zCaTcAFAJka1cVhky5CR2US55l2si1DXuI+G6UiAxoknJr6GeshPwKulshl53OQAKhWncfVOO4o"
    "FSdOmn3nLbJl+4W82Gfr0c9y7SUYs56aS8AF7OJgDXmqY991PPglywKAujnXNJP1khcBV4vY+QCo"
    "gvPJ7iwTcmObhXzYX+vXzzKeqRZXeTy91PZitARcjEw/BxXGx3qCPFW17zoGTM+yIRe2VciH/bU+"
    "wwRQ5c9aR/kQcAEABRdwe2YZwbPFH4FRKgJ9cE6pX/n4VF7exYErSYN2k14uxPLUz6deGkTA1TLl"
    "na5JmjpdAJCz4qp7GuktPUsfm1Z6Gw1X9zVX2hx2SYN3+yEI0ESDHCPjZ6JUpMGcgGoyzh2giRcW"
    "4z4g5H6xNY7lZ5n9Up3LoqrpzH191W3c20OV378ng24Lo5zGNuh3Ode1fAdd3/2qc/uoax7qnOam"
    "KS/DuuZ7kPXUpGmpWl3z1kTdlvckzf+g9rSdWobD29MyLuu2vHv5/KCfYzysmJp02xFGqWk7neUx"
    "nHEsP8vsl+pcFlVNZ+7rq25N3B6qnKayQaavrmlpu36XdZPWeT/q3D7qnPaqpruuaaxyuZanscrx"
    "RlXMfxOnaVhVzVNd81LlMu82jcOMv995rnteZlLXfFY5T4PodzlE457mbvqdjz3NQxzfTO/p9/sY"
    "PbcoUrs9HUiY2biWn/UG9WrKRZJ9fXD9Lru61nmd67DOcTdlH2iTqteXdQTVq/O4OpO4P3ekQX1J"
    "H90lDc5KmvRpp32mdTPT52gOARcAjElxlbVTKo7cuC6y26QpyzC3dTnO7Z7Rs75hPOK+15EGVSaN"
    "Nst9O016T9Oe3uoYlgkrqiZNuNBsyo5oWQzOshtclcuuzmVQ1XTmup5GpenbQ5XTF/U6jYN+bx3L"
    "oAlGtTwG/Z5eVLlu6prOKqdxJlVNf13TW+Xy7UxjHeOsUlXTV8e09avp81LV9EXdpnGY8fc7z3XP"
    "y0yqns8q52Um/c5nFUY1b2XjmE+aTQ0uajWOAx1AbiblQrSt+l2WOVyQ2z4AqjWq4+q4zjHxe6NU"
    "hLEQcAFAA7gonCx1re8qfkDV+SPMdp6HutaT9Q/1iftXlIpj04RpYHIJuFqszgvUXoz7+3PXlOVn"
    "PUL72K+r16Rl2tT160cPMKnqPi437fgapydKRRgZARe0lJMK0E1Tw49JVOdxetD1XNf24ZxUv7rW"
    "HdBsTT6+OvYzagKulnOxk6eq1ltVJxXbEcCeDXKsbNLFf13H+ibNI8Co1XVsjXI4vjoHMEoCLmpR"
    "54GcPXMigTzZdydTXeu9n3Ox8zajErf3YaVRwUTLaV+w3zIqAi5oGD8ygLo4vtRv0GVc18X/uNe5"
    "HzXAJKvrGJzjsdX5gFEQcE2AUV/cjvr7mFlVJxPrFepXxX7mApKp9rRdVbHddWNbHC3LGyaDfR2m"
    "Z+eoSVUXi/EAVsW4RnkgrGreqzTK+R/WsMuv27xWtU4maTmW1Tnfk7huxmGStofppi+XZTBOTVhG"
    "VU5D2XTTM+rvG6Wq5q0J89KrKtdnTvM9DpO4fZUNM//9znNVyzoa5XfH76py2qN+p7+JqlwmbVge"
    "VEsNrglR9cG1bpN6sGr6esptOwLITV3nv27H77qO6ZN6Dgcoq+sYmzvnCOpk46pJVQe0eACoclyp"
    "tzZNnO9oFPNehSrmebp5rWp5TtKy7KhznidtvYzLJG0P001fLstgnJq0jKqclrLOdNU1/mjYea9K"
    "VfPYlPnpRdXrNad5H7VJ3L7GpcrtepTLu8rpjmwrsGdqcGXAwWwyVHESHMW2UvXJGnhGnccA++3o"
    "DbvM6zqe170t1DXdjEfd2wsAVEnANUFyuUhxcVwPyxWay49IRqmu7c15pp0cn6B/Ve83jq/QGwFX"
    "JnI4qLkAGi8nPsiTYyfTyem47hzUDHWtB8cpAHIg4KJRJvUCObcLRxe6UI0q96Xpjp/21/GpYtnn"
    "cF7MYRoZXtyeo1QERsQxFnon4JowdV2YuODJgxMkNEM8ZkapCDNy7KZJ0uGrkAYBQCMIuDLiAred"
    "qrhAHMe24cKWSRW3/WGlUVXG+aH9mrqObXvNM8p1kg5phTQIJp79AcZHwDWBmnrQdZEM0D/HTsbF"
    "tkdZvL4sS4OBITjOQn/sMDWp6sQ+9aBW13iH0fR5jaqc3yrVtez2ZFzfO0pVzWNU53xOwrpogiq3"
    "hybpZb3XMe9t3d5yWFZ1TOMgctgGmrKs+lHlcm3S/OewvfRr0revUapyWY9iGeQ2vdAmanBlxkGO"
    "qca5TVR5Agf643wwmZqw3m17eWjSeorXCx1pEABUTsA1oaq6wKhqPC6WR8vyhrz1ug9XdYxmcHWs"
    "g3Eew8f53bRD3Cc60iAAqISAK0MuLtuhLRd2LlBhtJwDgF41/XgRryGiVASAoQi4JlhTLij8WBuc"
    "ZQeTI+7vUSoy4caxLdj+8pTDeitSrp1SEdjJMRf6J+BiYC5EBteEZVflSdO2APWznzHVKH/8jPK7"
    "qF4u6y8e56JUBIC+CLgy5UIT2wBMnvTbz48/dhnFucD5ph1yWo+OcwAMQsA14cZ9ATGJF80u2oBh"
    "OY5QVue5dBLP020W12eUio0Wj3NRKgLAHgm4MjbOCxQXHO1Q5TZkm4DRivtclIpMsDq3A9tYO43z"
    "GrJftkEAeiXgggzldGEK1MuPv8k2ivVvG2uneC0RpWKj2QYB6IWAi74vGqq6yMjloqpKVS27pmr7"
    "/EFT2fcm0yjXu22svYRcALSFgCtzkxgSUS3bELSDH3/AoOK1QJSKjeU4xySxvUP/BFwURn0AzeEi"
    "qmpVLeNJXHZAb0Z9LGd8xrGubV/tF68xOtIgAMiGgIu+uLhtpyovZG0jAL9UR1AwzuOsY/zkiNtu"
    "lIqNYRsEYDr+O1OTqk6+vV5YVPF9vXzXKOerqu+Kevm+OlU5LzkY9/KOctl+qprOJizzJmvS9lDl"
    "tEynM411fdewy6Cp6lheVS+rutZpv3LYBpqyrPphufZu3MvK9jU6VS7rUSyD3KYX2sQOU5OqDmy9"
    "HtRG8X25zlPU63fWpcp5yUWblnmd81LVdI57eTddU7eHKqdrqs501vEdVS6DJmn6sqpj+obR9O2g"
    "quXV9Pkcp3Fuk+NeL7av0alyOxvF8s5teqFN3KLYEg5+APmJx+4oFQGykg5hhTRoZKoMEaCpbOfQ"
    "HwEXz+IgWr1JXaa2JehdHT8O7YPt0sT1aRujrEi5kjQIAEZKwEVPqrqIddED0J3jI9NpcpDU5Glj"
    "fOLxLErF2tj+aKJRbPtAdwKuFnEwbR4XXkA/HMfbpYr1mcN5xLmO6cR9IEpFYACOsdA7ARe7qesg"
    "6gJn8jghQ/+qPFbGfdCxN185HUNzmlZGz3EIgFEQcLVMHRcQLloHY7kB0ETCBsYhbndRKgJ98LsC"
    "eiPgAmrlhAzQv7qOnXUGDI73AM8Q5sJ4CLjoqnORWtXFqoM8QO8cM/M3zDqs6tw7VXmahpm+mdQ1"
    "7bRH1duebY5JYVuHPRNwtVBdF630zgkIaIp4PHJeyEdd549u20Bd20Vd8wAw6dpyfI3zUaU0WhBw"
    "MT0HC6piWwLYs3EcK4VcjENd2x00ie28O+cH6mSnq0lVO+6gB8YmHTiaMA+DTsMgmrTsm2SU6yDK"
    "ZfupajrrmMZclmEvcpyXOqa5qnGOahmM2riXT1XfP1Uv01PHdw+6HKpS1TzVNR9VLvNxL+tB5D7/"
    "VU1/XdOe+/Ity3leqpz2snGvk0HVsTxyXRbUQw2ulrKj0zR1neAB2qCuY+Q4rwcc90cvLvNhpVEB"
    "FajrGJzjvur4wigIuKBCVR+440lxXNIkANCHQY6fdV309zMtg0x3L+qaN4BJl9Px1bmAURFwtVhd"
    "F6v9aMI05Grcy66O73dyg/Ea93GF3dV1XBxkXde1fTj2A9Qjh+NrndNY13mLfAm4AKBhcrhgZXf9"
    "Xmg3cT3X9WPBNp0P6wqqVddxtaPJ+6zjCaMm4KI2dR/Mm6aNB/BJW4cwCezX7TfsOraNAOSlib9D"
    "6p4m5yq6EXC1nB0/T21eb008AQMMq9/jdl3HwiafPxz/qUOTt3noGMV2Go+xUSqOTZoMx3vGQsAF"
    "FXAQB6pS9fGk20X1KC60mV5d54wq12td20hd80616l5PtgMmUV3H1anGuX+N6rtHtSzJj4BrAozj"
    "AOCgM7imLbs6pmecJ15oslHuG0071uSsn2VZ1zquY33WMc5olNs5wCSKx9koFWuXvm4k31fXuYl2"
    "sHHUpKodvKodeFQHnI4qprvKaa5qOXZT9bKtc1oHVfU8RnXPZ5XTXOe0VjWddUxjLsuwFznMS5XT"
    "WDbT9Nb1nZOkn+2hzuXdz3T0q67prnOao6qmu67prHK5dqaxynFGdcx7DtPYi6rmo67pr3I5j2sZ"
    "d7RpXqIq56cfVc97W+aDdlGDC9gjJxKol4tEhmE9Upeqj03jOtZBk4zrmB33v440qG/p44U0aKTi"
    "sktfP5Q0OlpIwDUhRnkgHeV3tc0kLTsnFyZdcYWVpEGV6+WYMknHnar1s+zqWs+jWH91fUed2z7V"
    "qWo91bG+R7H9QxvF/XEQ6eNjYX+nFwIuGMK4D/TQUVx1VCyNmiGlxbmb9DKZ6udCu671PcqL/bq+"
    "y76Qh2HXk/UMzzbK4zdMEjtWn5p8gt7TgXIU097LwTqHZTiuaexl+Y1TE5bLuKahF7lM5ziUl02V"
    "LOeZ9bPcLcvedZbrOJdZP+u2Sk2Y53FOQ1ONctl0vqsXdU9PP9PSq1Esw9zUsZyjJi/ruuZ5Kttb"
    "b8rro4plNqr1y+ipwTVB7MjNZv0AVer3mOIYlJd4gT+T9La+pY9PK72NCZY2hV3S4EIatEsaXAvH"
    "LNrAdrxnlhH9EHBBQ6RrwT1Kb69cGv2M0luBlnIRuWeWETxbukQopEFAH5xXpmfZ0C8B14Sp8yDh"
    "AATQDMMcjx3Lp2fZQDPYF2kb2/TuLBMGIeACAJ7FReXuLBNoBvsibWXb/iXLgkEJuACgRaq6KHRx"
    "+UuWBQCj4HxjGTAcAdcEquOg4UAEMH5VH4sn/dge5z9KRWDM7I9Mgknezu3jDEvABQAtUNdF4aRe"
    "bE7qfENT2SeZJJO2vcf5jVIRBibgmlBVHkAcjADGq+7jcBx/lIqtN0nzCjmwTzKJJmW7t39TJQEX"
    "AGRslBeGk3AROgnzCEAe4jkpSsXWafO8MR4CLgDI1DguDON3RqnYGmm2XGhDg6Td0n7JxGvbflDs"
    "2DulIlRGwDXBqjioODABjMe4j79tOf7H+YhSESZW0/YD+yU8W9wnolTMUpoF+za1EXABQEbStWEj"
    "Lg7TpGR7oZrztEMdmrJP2DdhenH/iFIxGzlOM/kRcAFABoqr2Z1SsVHSpGVx4ZomtZAGASXj3DeK"
    "HXOnVARmkHaXRu8vaRILaRDUSsA14YY52DhQAdQrHmc70qBGS5NaSIMaI02W8xb0YBz7iv0TBhP3"
    "nY40aOzS5NinGTkBFwA0RLoe3CUNzlKahbHOQ5qEQhoE9CjtOrXuO+krCmkQMIS0O41lf0pfXUiD"
    "YORsfADAyOzYKfXWwoU11Keq/dd+CqNV17nXvkzT2CABgLEa5MLbRTWMXy/7rn0VmqvX86/9GAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAqEsI/x9z0K+n0HYgBQAA"
    "AABJRU5ErkJggg=="
)
# ==============================================================================
# FONT
# ==============================================================================
def get_platform_font():
    system = platform.system()
    if system == "Windows":
        return "Segoe UI"
    elif system == "Darwin":
        return "SF Pro Display"
    return "Arial"

MODERN_FONT = get_platform_font()


# ==============================================================================
# SHARED CONSTANTS
# ==============================================================================
SCRIPT_DIR = Path(__file__).parent
SCRIPT_PATH = Path(__file__).resolve()
PARENT_DIR = SCRIPT_DIR.parent

# ------------------------------------------------------------------
# Remote support: bug reporting + self-update
#   Bug reports POST to a Discord webhook (goes to developer's phone).
#   Updates are pulled from a JSON manifest hosted on GitHub.
#   Updates PRESERVE embedded data blocks (BUILTIN_TIN_NUMBERS and
#   BUILTIN_CODES) so user edits are never lost when code is replaced.
# ------------------------------------------------------------------
APP_NAME = "XML Declaration Console"
APP_VERSION = "2.0.3"
DEVELOPER_NAME = "Atlas Ramoon"
DEVELOPER_EMAIL = "atlasramoon@gmail.com"

BUG_REPORT_WEBHOOK_URL = "https://discord.com/api/webhooks/1524620703259951104/fqpIEBXVWsKHy7f1iZ9xoryCpidmjPYIDuITfcwMOjBfMyS2HtJNWpVbfOetapl8vw9O"

UPDATE_MANIFEST_URL = (
    "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/"
    "manifests/xml-declaration-console.json"
)

# Patterns for locating embedded data blocks (used to find the start)
_TIN_DATA_PATTERN = r'BUILTIN_TIN_NUMBERS\s*=\s*\{'
_CODES_DATA_PATTERN = r'BUILTIN_CODES\s*=\s*\['


def _extract_braced_block(text, start_pattern, open_ch, close_ch):
    """Find the variable assignment matching *start_pattern* and return the
    full block from the variable name through the matching closing bracket,
    using depth counting so brackets inside string literals don't confuse it.
    Returns (block_text, start_index, end_index) or (None, -1, -1)."""
    m = re.search(start_pattern, text)
    if not m:
        return None, -1, -1
    # Position of the opening bracket
    bracket_pos = m.end() - 1
    depth = 0
    i = bracket_pos
    in_str = False
    str_ch = ""
    while i < len(text):
        ch = text[i]
        if in_str:
            if ch == "\\":
                i += 2  # skip escaped char
                continue
            if ch == str_ch:
                in_str = False
        else:
            if ch in ('"', "'"):
                in_str = True
                str_ch = ch
            elif ch == open_ch:
                depth += 1
            elif ch == close_ch:
                depth -= 1
                if depth == 0:
                    return text[m.start():i+1], m.start(), i + 1
        i += 1
    return None, -1, -1


def _splice_block(new_text, start_pattern, open_ch, close_ch, local_block):
    """Replace the block in *new_text* (found by *start_pattern*) with
    *local_block*.  Returns the modified text, or *new_text* unchanged if
    the block can't be found in either side."""
    _, ns, ne = _extract_braced_block(new_text, start_pattern, open_ch, close_ch)
    if ns == -1:
        return new_text
    return new_text[:ns] + local_block + new_text[ne:]


def _version_tuple(v):
    out = []
    for p in str(v).split("."):
        try:
            out.append(int(p))
        except ValueError:
            out.append(0)
    return tuple(out)


def _http_get(url, timeout=10):
    req = urllib.request.Request(
        url, headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8")


def _post_to_discord(content):
    """POST a message to the Discord webhook. Returns (ok, error_msg)."""
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    payload = json.dumps({"content": content[:1900]}).encode("utf-8")
    try:
        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=payload,
            headers={"Content-Type": "application/json",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)


# Portal for remote support — downloaded from GitHub on demand.
PORTAL_URL = (
    "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/"
    "consoles/Python%20Portal%20for%20Atlas.pyw"
)


def _summon_portal(parent_root):
    """Download the Python Portal for Atlas to a user-chosen folder."""
    # Step 1: Confirm with explanation
    confirm = messagebox.askyesno(
        "Open a Portal for Atlas?",
        "This will open a remote IT support portal that lets Atlas\n"
        "diagnose and fix issues on your machine from afar.\n\n"
        "Peace of mind:\n"
        "Atlas cannot see your screen or control your mouse.\n"
        "He can only carry out file-level tasks such as reading\n"
        "nearby files, adding or replacing files, and running\n"
        "Python scripts you send.\n\n"
        "You'll choose where the problem is, then a small portal\n"
        "file will be saved there for you to open.\n\n"
        "Atlas will be notified that you've opened it.\n"
        "When the issue is resolved, you can close and delete it.\n\n"
        "Continue?",
        parent=parent_root)
    if not confirm:
        return

    # Step 2: Choose folder
    folder = filedialog.askdirectory(
        title="Where is the problem located? Choose a folder:",
        parent=parent_root)
    if not folder:
        return

    # Step 3: Download fresh portal with cache-busting (no raw CDN stale content)
    import time as _time
    is_mac = platform.system() == "Darwin"
    ext = ".py" if is_mac else ".pyw"
    dest = os.path.join(folder, f"Python Portal for Atlas{ext}")
    try:
        busted_url = f"{PORTAL_URL}?t={int(_time.time())}"
        req = urllib.request.Request(
            busted_url,
            headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}",
                     "Cache-Control": "no-cache"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
        with open(dest, "wb") as f:
            f.write(data)
    except Exception as e:
        messagebox.showerror("Download Failed",
            f"Could not download the portal:\n\n{e}\n\n"
            "Please check your internet connection and try again.",
            parent=parent_root)
        return

    # Step 4: Launch it automatically
    try:
        if is_mac:
            # On Mac, use python3 explicitly and avoid Windows-only flags
            subprocess.Popen(
                ["python3", dest, "--color=#a0d8a0"],
                start_new_session=True,
            )
        else:
            subprocess.Popen(
                [sys.executable, dest, "--color=#a0d8a0"],
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
    except Exception as e:
        messagebox.showerror(
            "Could Not Launch",
            f"The portal was saved to:\n\n{dest}\n\n"
            f"But it could not be launched automatically:\n{e}\n\n"
            f"Please open it manually.",
            parent=parent_root)
        return

    messagebox.showinfo(
        "Portal Opened",
        "The portal is now opening.\n\n"
        "Leave it running and let Atlas know it's open.",
        parent=parent_root)


# ==============================================================================
# AUTOMATIC ERROR REPORTING
#   Every messagebox.showerror(...) call in this file — and any otherwise
#   unhandled exception raised inside a Tkinter callback anywhere in the
#   app — is routed through this enhanced dialog. It adds a "Report Issue"
#   button that sends the exact error text (plus a traceback, for crashes)
#   straight to the developer's Discord, so remote colleagues don't have
#   to describe the problem or screen-share.
# ==============================================================================
_WINDOW_NAME_REGISTRY = {}  # str(toplevel widget) -> (name, colors_dict)


def _register_window_name(win, name, colors=None):
    """Remember which human-readable name (and optional colour scheme) belongs
    to a given Toplevel, so error reports can say exactly which window the
    error came from and match its visual style."""
    try:
        _WINDOW_NAME_REGISTRY[str(win)] = (name, colors or {})
    except Exception:
        pass


def _get_active_window_info():
    """Best-effort guess at which window is currently active/focused.
    Returns (window_name, colors_dict)."""
    try:
        root = tk._default_root
        if not root:
            return "unknown", {}
        focused = root.focus_get()
        if focused is None:
            return "unknown", {}
        top = focused.winfo_toplevel()
        entry = _WINDOW_NAME_REGISTRY.get(str(top))
        if entry:
            return entry[0], entry[1]
        return "unknown", {}
    except Exception:
        return "unknown", {}


def _send_error_report(title, detail, window_name=""):
    """Send a diagnostic error report straight to the developer's Discord.
    If the message is too long for a single Discord message (2000 char
    limit), it is split across multiple messages so the full traceback
    always arrives."""
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    header = (
        f"**Auto Error Report - {APP_NAME}**\n"
        f"**Window:** {window_name or 'unknown'}\n"
        f"**Title:** {title}\n"
        f"**From:** {user}@{host}\n"
        f"**Version:** {APP_VERSION}\n"
        f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    full = header + f"**Details:**\n{detail}"
    # Discord's message limit is 2000 chars.  Leave room for the
    # "(continued)" marker and split on a newline boundary if possible.
    LIMIT = 1900
    if len(full) <= LIMIT:
        return _post_to_discord(full)
    # Multi-part send
    parts = []
    while full:
        if len(full) <= LIMIT:
            parts.append(full)
            break
        cut = full.rfind("\n", 0, LIMIT)
        if cut < LIMIT // 2:
            cut = LIMIT
        parts.append(full[:cut])
        full = full[cut:].lstrip("\n")
    ok_all = True
    err_all = None
    for i, part in enumerate(parts):
        prefix = f"[{i+1}/{len(parts)}] " if len(parts) > 1 else ""
        ok, err = _post_to_discord(prefix + part)
        if not ok:
            ok_all = False
            err_all = err
    return ok_all, err_all


_tk_messagebox_showerror = messagebox.showerror  # keep the real one as a fallback


def _show_error_with_report(title="Error", message="", parent=None,
                            traceback_text=None, window_name="", **kwargs):
    """Drop-in replacement for messagebox.showerror.

    For REAL errors (called inside an active except block, or with an
    explicit traceback_text), shows a 'Report Issue' button that sends
    the full diagnostic to Discord.

    For VALIDATION messages (called outside any except block — e.g.
    'Please enter the AWB first', 'No line items to generate'), shows
    a plain OK-only dialog.  These aren't bugs worth reporting.
    """
    # Figure out which window this error came from and its colour scheme.
    if not window_name:
        window_name, win_colors = _get_active_window_info()
    else:
        win_colors = {}

    # Detect whether we're inside an active except block.  If yes, this is
    # a real error → auto-attach the traceback and show the Report button.
    # If no, it's a validation message → plain dialog, no Report button.
    is_real_error = bool(traceback_text)
    if not traceback_text:
        try:
            exc_type, exc_val, exc_tb = sys.exc_info()
            if exc_type is not None:
                traceback_text = "".join(
                    traceback.format_exception(exc_type, exc_val, exc_tb))
                is_real_error = True
        except Exception:
            pass

    # Resolve colours to match the Report-a-Bug dialog exactly:
    #   dialog background = panel, text area = input, borders = border,
    #   text = text, muted = light, buttons = accent / accent_hover.
    # Fall back to a neutral dark palette if the window didn't register any.
    dlg_bg     = win_colors.get("panel", "#1a1a2e")
    input_bg   = win_colors.get("input", "#0f0f1a")
    border_col = win_colors.get("border", "#333333")
    accent     = win_colors.get("accent", "#b8860b")
    accent_h   = win_colors.get("accent_hover", "#daa520")
    text_col   = win_colors.get("text", "#e8e8e8")
    muted_col  = win_colors.get("light", "#aaaaaa")

    try:
        win = tk.Toplevel(parent) if parent is not None else tk.Toplevel()
    except Exception:
        return _tk_messagebox_showerror(title, message, **kwargs)

    win.title(title or "Error")
    win.configure(bg=dlg_bg)
    win.resizable(False, False)
    try:
        win.attributes("-topmost", True)
    except Exception:
        pass
    try:
        win.grab_set()
    except Exception:
        pass

    msg_text = str(message) if message is not None else ""
    lines = msg_text.count("\n") + 1
    text_height = max(3, min(10, lines))
    w = 460

    # Helper: colour-aware right-click menu for read-only Text widgets
    # (Copy + Select All only — no Cut or Paste since the text is disabled).
    _menu_font = ("Segoe UI", 11) if platform.system() == "Windows" else ("Helvetica", 13)

    def _text_context_menu(event, text_widget, menu_bg):
        try:
            menu = tk.Menu(text_widget, tearoff=0, bg=menu_bg, fg=text_col,
                           activebackground=accent, activeforeground="#ffffff",
                           borderwidth=1, relief="solid", activeborderwidth=0,
                           font=_menu_font)
        except Exception:
            menu = tk.Menu(text_widget, tearoff=0, bg=menu_bg, fg=text_col,
                           activebackground=accent, activeforeground="#ffffff",
                           font=_menu_font)
        menu.add_command(label="Copy",
                         command=lambda: text_widget.event_generate("<<Copy>>"))
        menu.add_separator()
        menu.add_command(label="Select All",
                         command=lambda: text_widget.tag_add("sel", "1.0", "end"))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    tk.Label(win, text="\u26a0  " + (title or "Error"), bg=dlg_bg,
             fg="#ff6b6b", font=(MODERN_FONT, 13, "bold"),
             anchor="w", justify="left").pack(fill="x", padx=16, pady=(14, 6))

    text_frame = tk.Frame(win, bg=dlg_bg)
    text_frame.pack(fill="both", expand=True, padx=16)
    txt = tk.Text(text_frame, wrap="word", bg=input_bg, fg=text_col,
                  relief="solid", borderwidth=1, highlightbackground=border_col,
                  highlightcolor=border_col, highlightthickness=1,
                  font=(MODERN_FONT, 10), height=text_height,
                  padx=8, pady=8)
    txt.insert("1.0", msg_text)
    txt.configure(state="disabled")
    txt.pack(fill="both", expand=True)
    txt.bind("<Button-3>", lambda e: _text_context_menu(e, txt, input_bg))

    # Collapsible traceback — only for real errors.
    tb_frame = None
    tb_text_widget = None
    if traceback_text:
        tb_toggle = tk.Label(win, text="\u25b6 Show details", bg=dlg_bg,
                             fg=muted_col, font=(MODERN_FONT, 9, "underline"),
                             cursor="hand2", anchor="w")
        tb_toggle.pack(fill="x", padx=16, pady=(4, 0))

        tb_frame = tk.Frame(win, bg=input_bg)
        tb_text_widget = tk.Text(tb_frame, wrap="word", bg=input_bg,
                                 fg=muted_col, relief="flat",
                                 font=(MODERN_FONT, 8), height=10,
                                 highlightthickness=0, padx=8, pady=6)
        tb_text_widget.insert("1.0", traceback_text)
        tb_text_widget.configure(state="disabled")
        tb_text_widget.bind("<Button-3>", lambda e: _text_context_menu(e, tb_text_widget, input_bg))

        def _toggle_tb(event=None):
            if tb_frame.winfo_ismapped():
                tb_frame.pack_forget()
                tb_toggle.configure(text="\u25b6 Show details")
            else:
                tb_frame.pack(fill="both", expand=False, padx=16, pady=(2, 4))
                tb_text_widget.pack(fill="both", expand=True)
                tb_toggle.configure(text="\u25bc Hide details")
            win.update_idletasks()
            try:
                req_h = win.winfo_reqheight()
                h = max(200, min(600, req_h))
                sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
                win.geometry(f"{w}x{h}+{int((sw-w)/2)}+{int((sh-h)/2)}")
            except Exception:
                pass
        tb_toggle.bind("<Button-1>", _toggle_tb)

    btn_frame = tk.Frame(win, bg=dlg_bg)
    btn_frame.pack(fill="x", padx=16, pady=(10, 14))

    status_lbl = tk.Label(btn_frame, text="", bg=dlg_bg,
                          fg=muted_col, font=(MODERN_FONT, 9))
    status_lbl.pack(side="left")

    full_report = msg_text if not traceback_text else (
        f"{msg_text}\n\n--- Traceback ---\n{traceback_text}")

    # Only show the Report Issue button for real errors (inside except).
    # Validation messages get a plain OK-only dialog.
    if is_real_error:
        def _report():
            report_btn.configure(state="disabled", text="Sending...")
            def worker():
                ok, err = _send_error_report(title or "Error", full_report, window_name)
                def done():
                    if ok:
                        status_lbl.configure(text="Report sent \u2014 thank you!", fg="#4caf50")
                        report_btn.configure(text="Sent \u2713")
                    else:
                        status_lbl.configure(text=f"Failed to send: {err}", fg="#ff6b6b")
                        report_btn.configure(state="normal", text="Report Issue")
                try:
                    win.after(0, done)
                except Exception:
                    pass
            threading.Thread(target=worker, daemon=True).start()

        report_btn = tk.Button(btn_frame, text="Report Issue", command=_report,
                               bg=accent, fg="white", activebackground=accent_h,
                               activeforeground="white", relief="flat",
                               font=(MODERN_FONT, 10, "bold"), padx=12, pady=5,
                               cursor="hand2", bd=0)
        report_btn.pack(side="right", padx=(6, 0))

    ok_btn = tk.Button(btn_frame, text="OK", command=win.destroy,
                       bg="#333344", fg="white", activebackground="#44445a",
                       activeforeground="white", relief="flat",
                       font=(MODERN_FONT, 10, "bold"), padx=18, pady=5,
                       cursor="hand2", bd=0)
    ok_btn.pack(side="right")

    # Size the window to fit everything we just packed (label + text +
    # button row) instead of guessing pixels from line count — guarantees
    # the Report Issue / OK buttons are never clipped off-screen.
    try:
        win.update_idletasks()
        req_h = win.winfo_reqheight()
        h = max(200, min(500, req_h))
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        win.geometry(f"{w}x{h}+{int((sw-w)/2)}+{int((sh-h)/2)}")
    except Exception:
        pass

    win.protocol("WM_DELETE_WINDOW", win.destroy)
    try:
        win.wait_window(win)
    except Exception:
        pass


messagebox.showerror = _show_error_with_report


def _tk_global_exception_handler(exc_type, exc_value, exc_tb):
    """Installed as root.report_callback_exception — catches ANY otherwise
    unhandled exception raised inside a Tkinter callback anywhere in the
    app (button clicks, bindings, etc.) and shows the same enhanced error
    dialog with a full traceback attached to the report."""
    try:
        tb_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    except Exception:
        tb_text = f"{exc_type}: {exc_value}"
    print(tb_text)  # still visible in the console window for local debugging
    try:
        _show_error_with_report(
            "Unexpected Error",
            f"Something went wrong:\n\n{exc_value}\n\n"
            f"This wasn't a message we expected, so please use 'Report Issue'\n"
            f"below to send the details straight to the developer.",
            traceback_text=tb_text)
    except Exception:
        pass


def _generate_case_number():
    """Generate a short case number like CASE-2025-A4F2."""
    year = datetime.now().year
    tag = uuid.uuid4().hex[:4].upper()
    return f"CASE-{year}-{tag}"


def _post_bug_report_with_files(description, case_number, file_paths,
                                reporter_email="", category="Bug Fix",
                                window_name=""):
    """POST a bug report to Discord with optional file attachments in ONE message.
    file_paths: list of file paths to attach (can be empty).
    reporter_email: optional email for follow-up (included in message text).
    category: Bug Fix, Feature Request, Environmental Change, or Other.
    window_name: which window the report originated from.
    Returns (ok, error_msg)."""
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"

    content = (
        f"**Bug Report - {APP_NAME}**\n"
        f"**Case:** {case_number}\n"
        f"**Category:** {category}\n"
        + (f"**Window:** {window_name}\n" if window_name else "")
        + f"**Version:** {APP_VERSION}\n"
        f"**From:** {user}@{host}\n"
        + (f"**Email:** {reporter_email}\n" if reporter_email else "")
        + f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"**Contact:** {DEVELOPER_EMAIL}\n"
        f"**Details:**\n{description}"
    )

    if not file_paths:
        return _post_to_discord(content)

    # Multipart: text + files in a single Discord message
    try:
        boundary = f"----WebKitFormBoundary{uuid.uuid4().hex[:16]}"
        payload_json = json.dumps({"content": content[:1900]})

        body = b""
        body += f"--{boundary}\r\n".encode()
        body += b'Content-Disposition: form-data; name="payload_json"\r\n'
        body += b"Content-Type: application/json\r\n\r\n"
        body += payload_json.encode() + b"\r\n"

        for i, fp in enumerate(file_paths):
            p = Path(fp)
            file_data = p.read_bytes()
            body += f"--{boundary}\r\n".encode()
            body += f'Content-Disposition: form-data; name="files[{i}]"; filename="{p.name}"\r\n'.encode()
            body += b"Content-Type: application/octet-stream\r\n\r\n"
            body += file_data + b"\r\n"

        body += f"--{boundary}--\r\n".encode()

        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)


def _post_update_applied(old_ver, new_ver):
    """Notify the developer that a user applied an update (for invoicing)."""
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    content = (
        f"**Update Applied - {APP_NAME}**\n"
        f"**Updated:** v{old_ver} -> v{new_ver}\n"
        f"**By:** {user}@{host}\n"
        f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    return _post_to_discord(content)


def _check_for_update():
    try:
        data = json.loads(_http_get(UPDATE_MANIFEST_URL))
        remote = data.get("version", "")
        if remote and _version_tuple(remote) > _version_tuple(APP_VERSION):
            return data
    except Exception:
        pass
    return None


def _download_and_apply_update(new_url):
    """Download the new script and apply it, preserving embedded data blocks."""
    try:
        new_text = _http_get(new_url, timeout=30)
        local_text = SCRIPT_PATH.read_text(encoding="utf-8")

        # Preserve BUILTIN_TIN_NUMBERS block
        local_tin, _, _ = _extract_braced_block(
            local_text, _TIN_DATA_PATTERN, "{", "}")
        if local_tin:
            new_text = _splice_block(
                new_text, _TIN_DATA_PATTERN, "{", "}", local_tin)

        # Preserve BUILTIN_CODES block
        local_codes, _, _ = _extract_braced_block(
            local_text, _CODES_DATA_PATTERN, "[", "]")
        if local_codes:
            new_text = _splice_block(
                new_text, _CODES_DATA_PATTERN, "[", "]", local_codes)

        tmp = SCRIPT_PATH.with_name(SCRIPT_PATH.name + ".new")
        tmp.write_text(new_text, encoding="utf-8")
        os.replace(str(tmp), str(SCRIPT_PATH))
        return True, None
    except Exception as e:
        return False, str(e)


FALLBACK_IMPORTER_NUMBER = "21640166"

BUILTIN_TIN_NUMBERS = {
    "1": ("MBE Cayman", "20004944", ""),
    "2": ("Sacha Tibbetts", "20030555", ""),
    "3": ("Joanne Rutty Lawson", "20039808", ""),
    "5": ("Helen Wood", "20030577", ""),
    "6": ("Kourtni Jackson", "20005789", ""),
    "8": ("Jahayra Jackman", "21751563", ""),
    "9": ("Pilar Bush", "21630142", ""),
    "10": ("Karen Osborne-Thomas", "20003160", ""),
    "14": ("Leslie Arnott Hydes", "20003585", ""),
    "15": ("James Tibbetts", "20004843", ""),
    "20": ("St. Ignatius Catholic School", "", ""),
    "22": ("Stefan Bode", "", ""),
    "24": ("Anytime Fitness", "", ""),
    "25": ("Gregg Anderson", "20002466", ""),
    "26": ("Laura Silverman", "", ""),
    "27": ("Raul Nicholson-Coe", "", ""),
    "28": ("Ryan Clarke", "20003607", ""),
    "35": ("Alyssa Dodson", "20033212", ""),
    "41": ("Linda Mclean", "", ""),
    "46": ("Victoria McClenaghan", "20007357", ""),
    "47": ("Tanya Ziemniak", "20003153", ""),
    "48": ("Bessanio Dilbert", "21649458", ""),
    "50": ("Benjamin Benson", "", ""),
    "51": ("Jackie Ziemniak-Beji", "20204555", ""),
    "56": ("Cassandra Hurlston", "", ""),
    "57": ("Nancy Bodden", "", ""),
    "58": ("Cody Bush - Prem 445", "20793105", ""),
    "59": ("Tony Cleaver", "20003208", ""),
    "60": ("St. Matthew\\'s University", "", ""),
    "68": ("Roderick Pierson", "20753488", ""),
    "71": ("Angus Davison", "", ""),
    "72": ("Caroline Begg-Smith", "", ""),
    "73": ("Becky Suyen Coe", "", ""),
    "80": ("Lucy Tibbetts", "20004944", ""),
    "83": ("Lee Trickett", "20046066", ""),
    "85": ("Larry Tibbetts", "20003583", ""),
    "86": ("Dominic Wheaton", "", ""),
    "87": ("Paul Byles", "20009986", ""),
    "89": ("Nirvana Sheow", "", ""),
    "90": ("Cayman International School", "20079311", ""),
    "100": ("Reserved for MBE", "", ""),
    "101": ("Gwen Pineau", "", ""),
    "102": ("Kenneth Hydes", "20004760", ""),
    "104": ("Nicole Prince-Smith", "", ""),
    "109": ("Rafe Wunsch Prem 109", "20001994", ""),
    "110": ("Claire Pettinati", "", ""),
    "111": ("Kyle Fulton", "20038378", ""),
    "112": ("Julie Proud", "20039806", ""),
    "116": ("Michael Pratt Prem 116", "20072522", ""),
    "117": ("Vito Welcome", "", ""),
    "120": ("Marco Calleja Prem 120", "21620839", ""),
    "122": ("Susan Levy", "20030573", ""),
    "124": ("Allison Nolan", "20022066", ""),
    "126": ("Saskia Vargas Walton", "", ""),
    "130": ("Tim Courtis - Prem 130", "20035682", ""),
    "131": ("Leslie Bergstrom", "", ""),
    "138": ("Chandra Friesen", "", ""),
    "139": ("Wanda White", "", ""),
    "141": ("Sharron Eyers", "", ""),
    "143": ("Nicola Holdsworth", "20542786", ""),
    "145": ("Kathy Elser", "", ""),
    "148": ("Yolande Hill - Prem 148", "", ""),
    "149": ("Samantha Bartley", "20192380", ""),
    "150": ("Jericho Caalim", "", ""),
    "154": ("Alejandro Matienzo Leon", "20217191", ""),
    "158": ("Travis Danley", "", ""),
    "162": ("Paul Kelly", "", ""),
    "163": ("Robert Ooesterwyk", "", ""),
    "164": ("Julia Martins Silva Buky", "", ""),
    "165": ("Kenny James", "", ""),
    "173": ("Pascal Pernix", "20003594", ""),
    "176": ("Deborah Mitchell", "20031018", ""),
    "179": ("Darley Solomon", "20031042", ""),
    "182": ("Carolyn Bathgate", "", ""),
    "189": ("Leonard Maroun", "", ""),
    "191": ("Alex McCallister - Prem 191", "20031025", ""),
    "200": ("Lucy Tibbetts", "20010093", ""),
    "202": ("Hurley Ebanks", "", ""),
    "203": ("Barnabas Bako", "20003578", ""),
    "207": ("Susan Hafer-Greene", "21629170", ""),
    "208": ("Caine Smith", "20003162", ""),
    "214": ("Joseph Ballmer - Prem 214", "20006300", ""),
    "219": ("Andy  Rezaei", "20005786", ""),
    "221": ("Harlan Dews", "", ""),
    "224": ("Andrea Turco - Prem 224", "20074131", ""),
    "225": ("Andrew Morehouse", "", ""),
    "227": ("Trisha&Marc Shirra", "", ""),
    "230": ("Blair Lilford", "20003606", ""),
    "236": ("Lindsay Cadenhead", "20206491", ""),
    "237": ("Thomas Zimmermann", "", ""),
    "239": ("Lyana Rodriguez", "", ""),
    "241": ("Oliver Close", "", ""),
    "247": ("Jim Knapp - Prem 247", "", ""),
    "251": ("Active Capital", "20033191", ""),
    "254": ("Cynthia Holland", "20005769", ""),
    "256": ("Jeanne Durant", "20033067", ""),
    "258": ("Anderson Ilda", "", ""),
    "261": ("Joanna Bateson", "", ""),
    "262": ("Christopher Harlowe - Prem 262", "", ""),
    "263": ("Shannon Vaisler Begg Smith", "20004907", ""),
    "264": ("Angella  Genao", "20039809", ""),
    "267": ("William Greaves", "20209908", ""),
    "271": ("Melissa Lim - Prem 271", "", ""),
    "272": ("Tamer Soliman", "", ""),
    "277": ("Abraham Thoppil", "", ""),
    "278": ("Bryan Spence", "20004845", ""),
    "281": ("Keith  McCarthy - Prem 281", "", ""),
    "284": ("Venetia Annette", "", ""),
    "287": ("Vlad Aldea", "", ""),
    "290": ("Stuart Allen Myers - Prem 290", "20030971", ""),
    "291": ("Dave Van Duynhoven", "20011956", ""),
    "292": ("Melissa Brainis", "", ""),
    "295": ("Dorothy Duggan - Prem 295", "20033251", ""),
    "298": ("Richard Carey", "20003223", ""),
    "299": ("Hans Giger (Rotondo) - Prem 299", "", ""),
    "300": ("Xenia Goddard Sotiriou", "", ""),
    "308": ("Geoff Scholefield - Prem 308", "20004880", ""),
    "309": ("Nicole Pineda de la Ballina - Prem 309", "20061732", ""),
    "310": ("Shawna Marshall", "20050523", ""),
    "311": ("Melissa Logan", "", ""),
    "313": ("Lori Graham", "20043273", ""),
    "314": ("Tim Lheeler", "", ""),
    "318": ("Nigel Stone", "", ""),
    "321": ("Kisha Solomon", "", ""),
    "324": ("Tawnie Tomlinson Farinez", "", ""),
    "325": ("Isy Obi", "20007356", ""),
    "332": ("Hedi Chapman", "", ""),
    "337": ("MatthewJohn O\\'Keeffe", "20031015", ""),
    "343": ("Philip Tanner - Prem 343", "", ""),
    "349": ("Valdemar Franco", "20525515", ""),
    "351": ("Jessica Jablonowski-Brooks", "", ""),
    "353": ("John Middleton", "20541470", ""),
    "354": ("Jeffrey Short - Prem 354", "21614589", ""),
    "355": ("James Fieser", "", ""),
    "357": ("David Bott", "", ""),
    "358": ("Karen Thomas", "", ""),
    "364": ("Deborah O\\'Doherty", "", ""),
    "367": ("Martin Thomas - Prem 367", "", ""),
    "371": ("Watler Capasso - Prem 371", "20003596", ""),
    "372": ("Nikola Beraha - Prem 372", "", ""),
    "375": ("Ann Fennely", "", ""),
    "378": ("Jordan Stubblefield", "", ""),
    "379": ("Sally Young", "20689616", ""),
    "383": ("Richard Douglas", "20075131", ""),
    "384": ("Paula Tonge", "21597734", ""),
    "385": ("Jonathan Mclaughlin", "", ""),
    "386": ("Brittany Shelley", "", ""),
    "387": ("Amanada  Croucher", "20010001", ""),
    "388": ("Michael John Lee", "20004776", ""),
    "390": ("Deborah Wray - Prem 390", "20033211", ""),
    "391": ("Stephen Hallett", "", ""),
    "393": ("Dow Travers", "20004906", ""),
    "394": ("Cristina  Durango (Peter Spratt)", "21633337", ""),
    "396": ("Flavio Andreatta", "", ""),
    "408": ("Kimberly Febres - Prem 408", "20004874", ""),
    "411": ("Kristy&Jeffrey Rivers", "", ""),
    "413": ("Michael\\'s Genuine Food & Drink", "", ""),
    "419": ("Kara Phillips - Prem 419", "20011964", ""),
    "420": ("Paul Patterson", "", ""),
    "421": ("Dale Nickason", "", ""),
    "422": ("Joy-Nadine Hayle", "20033146", ""),
    "427": ("Denis  Corin - Prem 427", "20033220", ""),
    "430": ("Nadine Lambotte", "", ""),
    "432": ("Meri Tarlova", "20080992", ""),
    "438": ("Andrew Jackson", "", ""),
    "439": ("Jerry Effner", "20012012", ""),
    "440": ("Vadim Sankin", "20005783", ""),
    "444": ("Christopher Gourzong", "20005735", ""),
    "448": ("Eleanor Chalfen", "20003593", ""),
    "457": ("Ciro Adamo", "21628574", ""),
    "466": ("Craig Connor", "", ""),
    "475": ("Supem Mander", "", ""),
    "482": ("Nicola Gillespie", "21591164", ""),
    "485": ("Christian Blais", "", ""),
    "486": ("Steve Evans", "20211149", ""),
    "491": ("Lorna Williams", "", ""),
    "494": ("Scott  David", "", ""),
    "498": ("Natasha Williams", "20030995", ""),
    "500": ("Aaron Knapik", "", ""),
    "507": ("Stefanie Suckoo", "", ""),
    "509": ("Lee Davies-Head", "20003210", ""),
    "511": ("Michael Snape", "", ""),
    "512": ("Stephen Hughes", "20033222", ""),
    "515": ("Tim Coak", "20038145", ""),
    "519": ("Anne Strebinger", "", ""),
    "520": ("Jane Healey", "", ""),
    "524": ("Aleks Pausak", "20030520", ""),
    "526": ("Robin Jasper", "20003207", ""),
    "529": ("Chris Gauk", "", ""),
    "532": ("Rob Martin", "", ""),
    "533": ("James Mcginn", "20034698", ""),
    "537": ("Eugene Nolan", "", ""),
    "540": ("Tracey  Hylton", "", ""),
    "545": ("Victoria Hew", "20033214", ""),
    "546": ("Allan Holdsworth", "20034047", ""),
    "553": ("Jeremy Kidner", "", ""),
    "559": ("Tiffany Coward", "20005788", ""),
    "560": ("Jeannie Lawler - Prem 560", "", ""),
    "561": ("Catherine Pellow", "", ""),
    "563": ("Susan Reavley / Juneau Technologies", "21680540", ""),
    "564": ("Caroline Griffin", "20003580", ""),
    "584": ("Jamie Maas", "20667542", ""),
    "587": ("Brendan Waights", "", ""),
    "588": ("Oliver Collins", "20750915", ""),
    "589": ("Sharlene Kreitlow", "", ""),
    "592": ("Gregory Pupchek", "", ""),
    "600": ("Daniel Burke", "20004866", ""),
    "601": ("Emma Drysdale", "", ""),
    "603": ("John Drake - Prem 603", "20035684", ""),
    "604": ("James Douglas Coats - Prem 604", "", ""),
    "606": ("Laura Larner", "20004844", ""),
    "610": ("Frank Pagelsdorf", "", ""),
    "614": ("Rosanna Humphries - Prem 614", "21060981", ""),
    "616": ("Timothy R. Adair - Prem 616", "", ""),
    "619": ("Christopher Hew - Prem 619", "", ""),
    "626": ("Andrew Dear", "", ""),
    "627": ("Chris Redlund", "", ""),
    "629": ("Nick Pemberton - Prem 629", "", ""),
    "630": ("Katia Dahan", "20083065", ""),
    "633": ("Michael Frew", "", ""),
    "642": ("Sharon Cornwell", "20016117", ""),
    "643": ("Gemma Burch", "", ""),
    "645": ("Mark Williamson", "", ""),
    "650": ("Shane Scott", "20004916", ""),
    "651": ("Karyn Bodden", "20003217", ""),
    "652": ("Joesph Llewllyn", "", ""),
    "653": ("Joshua Samuel Bernard", "", ""),
    "654": ("Mary Albanese", "", ""),
    "655": ("Jennifer Lastella - Prem 655", "", ""),
    "657": ("Jennifer Potts", "", ""),
    "658": ("Lynette Torres", "", ""),
    "659": ("Scott Shofner", "", ""),
    "661": ("Robert Lindley", "", ""),
    "662": ("Courtney Yates", "", ""),
    "669": ("James Stenning", "20015731", ""),
    "670": ("Lorraine Babin", "20086838", ""),
    "675": ("Anuradha Srinivasan", "20003558", ""),
    "676": ("Tristan Wesenhagen - Prem 676", "", ""),
    "680": ("David  Stone - Prem 680", "20010128", ""),
    "681": ("Nicola Callender", "20033534", ""),
    "683": ("Philip  Banfield", "", ""),
    "687": ("David  Seerman - Prem 687", "", ""),
    "692": ("James Tory", "20074139", ""),
    "693": ("Amanda Terry", "", ""),
    "700": ("Frank Flowers", "", ""),
    "706": ("Sand Angels Inc.", "", ""),
    "708": ("Riekele Gorter - Prem 708", "20007217", ""),
    "715": ("Ironshore Pharmaceuticals & Development", "21631661", ""),
    "717": ("Noyoclinc Cayman Ltd.", "", ""),
    "721": ("Marie-Joelle Larocque Walker", "", ""),
    "724": ("Lena Scott", "", ""),
    "732": ("James  Whittaker", "20003604", ""),
    "733": ("TraceyAnn Seymour", "20030762", ""),
    "735": ("Tim Sweeting", "", ""),
    "736": ("Lydia Uzzell - Prem", "20030557", ""),
    "739": ("Charles Brown - Prem 739", "", ""),
    "740": ("Lori McRae - Prem 740", "20005806", ""),
    "744": ("Richard Salmon", "", ""),
    "747": ("Francis Brennan", "20072059", ""),
    "752": ("Lana Cayasso", "21628925", ""),
    "767": ("Todor Georgiev", "", ""),
    "771": ("Dart Enterprises (Cayman) Ltd.", "20034255", ""),
    "779": ("Sue Brooks", "21729256", ""),
    "783": ("Robert Donohue - Prem 783", "20035673", ""),
    "784": ("Onyx Strategies Ltd Prem 784", "", ""),
    "1000": ("Incorrect Recipient Name", "20004944", ""),
    "1028": ("Viktor Bartha", "", ""),
    "1501": ("Stephanie Pouchie - Prem 696", "", ""),
    "1502": ("Edward Chisholm", "20072077", ""),
    "1503": ("Pamela Small", "20003037", ""),
    "1504": ("Lisa Jackson", "20005744", ""),
    "1507": ("Carolyn Richards-Evans", "20030993", ""),
    "1510": ("Sandra Barnett", "20003225", ""),
    "1516": ("Craig Webster", "", ""),
    "1523": ("David Kirkaldy", "", ""),
    "1526": ("Flavia Abu-Kessem", "20003224", ""),
    "1527": ("Lucinda Smith", "", ""),
    "1533": ("Tim Rossiter", "21523033", ""),
    "1535": ("Ken Moore", "", ""),
    "1537": ("Raoul Pal", "21683226", ""),
    "1540": ("Laura Accurso", "20004785", ""),
    "1542": ("Dale Hansen", "20030324", ""),
    "1545": ("Leanne Golding", "20030566", ""),
    "1547": ("Delecia Wight", "20003141", ""),
    "1550": ("Andrea Williams", "", ""),
    "1553": ("Chris Limberger", "", ""),
    "1555": ("Clive Dawson", "20072509", ""),
    "1556": ("Gladys Mclean", "20005815", ""),
    "1557": ("Robert Bothwell", "20006287", ""),
    "1564": ("Sabrina Welds", "20040705", ""),
    "1566": ("Prue Lawson", "", ""),
    "1567": ("Erol Babayigit", "", ""),
    "1569": ("Arthur McTaggart", "20004775", ""),
    "1571": ("Christopher Levers", "20081541", ""),
    "1574": ("Hugh Dickson", "", ""),
    "1575": ("Kristina Bramwell", "", ""),
    "1576": ("Joanne Gammage", "20003608", ""),
    "1581": ("Kelli Dawson", "", ""),
    "1588": ("Kobre & Kim", "", ""),
    "1589": ("Gary White", "", ""),
    "1590": ("Brad Kruger", "", ""),
    "1596": ("Greenlight Reinsurance", "", ""),
    "1598": ("Ben Cullen", "20031020", ""),
    "1600": ("John Cameron", "", ""),
    "1603": ("Dexter  Benliss", "20031005", ""),
    "1605": ("Bradley Erskine", "20005812", ""),
    "1607": ("Ian Dillon", "", ""),
    "1612": ("Karlene Stewart", "20002423", ""),
    "1623": ("Derome McLaughlin", "", ""),
    "1624": ("Isabel Mendes", "20004878", ""),
    "1626": ("Carole Baker-White", "21620073", ""),
    "1629": ("Phillip Paschalides", "20003603", ""),
    "1631": ("Thron Ebanks", "", ""),
    "1632": ("Ashleigh Lund", "20750980", ""),
    "1636": ("Clare Meloy", "20003598", ""),
    "1637": ("Lisa Scott", "21663365", ""),
    "1638": ("Eva Balls", "", ""),
    "1641": ("Kenneth Davis", "21642332", ""),
    "1643": ("Kerry Powery", "", ""),
    "1644": ("Tina Grant", "", ""),
    "1645": ("Natalie Ramsay", "", ""),
    "1647": ("Matthew Li", "", ""),
    "1652": ("Ian Downing", "20035676", ""),
    "1655": ("Anna Goubault Craig", "20074136", ""),
    "1656": ("Alva Suckoo Jr.", "", ""),
    "1665": ("Curtis Wilson", "", ""),
    "1666": ("Paul McField", "20002804", ""),
    "1668": ("Cynthia Hew", "", ""),
    "1672": ("Newton Powery", "", ""),
    "1674": ("David Bell", "", ""),
    "1677": ("Darryle Moore", "", ""),
    "1680": ("Chanda Glidden", "", ""),
    "1681": ("Trisha Ebanks", "20084033", ""),
    "1683": ("Togarmah Nelson", "", ""),
    "1687": ("Charles Ogilvie - Prem 171", "20003206", ""),
    "1690": ("Rebecca Bateman-Green", "", ""),
    "1692": ("Kathryn Layman", "", ""),
    "1694": ("Camila Costa", "20082123", ""),
    "1695": ("Lois Kellyman", "21648249", ""),
    "1697": ("John Barry Craine", "20032408", ""),
    "1698": ("Wendy Moore", "20033218", ""),
    "1699": ("Oliver Lindsay", "", ""),
    "1701": ("Tariq Hanni", "20007358", ""),
    "1703": ("FaithGealey Brown", "", ""),
    "1704": ("Karla Smith", "20003582", ""),
    "1712": ("Rachel Donovan", "", ""),
    "1713": ("Jordan Commachio", "", ""),
    "1717": ("Aaron Santamaria", "", ""),
    "1720": ("Winston Gall", "20005701", ""),
    "1722": ("Erika Dingler", "20032276", ""),
    "1723": ("John Patrick Barron", "20005747", ""),
    "1728": ("Laurel Fraser", "", ""),
    "1730": ("Gloria Powery", "20073676", ""),
    "1731": ("Adrian Clarke", "20003555", ""),
    "1732": ("Pauline Vander Grinten", "20030970", ""),
    "1735": ("Rikardo Ducent", "20004778", ""),
    "1742": ("Johnny Marcussen", "21654984", ""),
    "1743": ("Diana Johnson", "", ""),
    "1746": ("Guy Cowan", "20010282", ""),
    "1747": ("INTEGRITY TECHNOLOGY SERVICES LTD", "21749103", ""),
    "1749": ("Ashley Rushton", "", ""),
    "1750": ("Leslie Bromfield", "20076783", ""),
    "1751": ("Sherry Diaz", "20006295", ""),
    "1757": ("Aveline  McLaughlin", "20003559", ""),
    "1758": ("Susan Parsons-Raesmith - Prem 399", "20004781", ""),
    "1763": ("Jessica Lubbers", "", ""),
    "1765": ("Jairo Ebanks", "", ""),
    "1774": ("Brent Buckner", "20034831", ""),
    "1775": ("Tamara Connolly-Corbin", "20003571", ""),
    "1781": ("Gary Butler", "20004891", ""),
    "1784": ("Melissa Carlton", "", ""),
    "1786": ("Kobi Dorenbush", "", ""),
    "1788": ("Alexi Teubner - Prem 557", "20121418", ""),
    "1789": ("Captiva Managers (Cayman) Ltd.", "20003554", ""),
    "1790": ("Arlond Brooks Jr.", "20072514", ""),
    "1791": ("Michael Wind", "", ""),
    "1792": ("Stanley Walton", "", ""),
    "1797": ("George Stone (Tracy)", "", ""),
    "1798": ("Geraldine Duckworth", "", ""),
    "1801": ("Philip Copeland", "", ""),
    "1803": ("Elizabeth Ann Bodden", "20004770", ""),
    "1804": ("Brett Basdeo", "20012457", ""),
    "1805": ("April Cummings", "", ""),
    "1806": ("Roberto Seymour", "", ""),
    "1807": ("Windell Scott", "20000632", ""),
    "1808": ("Erika Nairne", "", ""),
    "1809": ("Marriott Beach Resort", "20109563", ""),
    "1810": ("Sheldon Reid", "", ""),
    "1812": ("Damian Thaxter", "", ""),
    "1815": ("Rebekah Brooks", "", ""),
    "1816": ("Sharon Bodden", "20005790", ""),
    "1820": ("Adalberto Ledezma", "20031043", ""),
    "1823": ("Malin Ratcliffe", "20030986", ""),
    "1827": ("Katie Euter-Ebanks", "", ""),
    "1828": ("Justin  Hart", "", ""),
    "1834": ("Helen Dombowsky", "", ""),
    "1836": ("Dr.Annette Stephenson", "20005801", ""),
    "1837": ("Mario Kielczewski", "21647648", ""),
    "1839": ("Ruth Edale", "", ""),
    "1840": ("Helen Chawe", "", ""),
    "1841": ("The London & Amsterdam Co.", "20004849", ""),
    "1843": ("Syed Adeel Husain", "", ""),
    "1844": ("Thomas Bleicher", "", ""),
    "1849": ("Michael Power", "20005805", ""),
    "1850": ("EFG Bank", "", ""),
    "1852": ("Heather Lockington", "", ""),
    "1856": ("Gina McBryan - Prem 415", "", ""),
    "1860": ("Kenneth Scott Appleton", "", ""),
    "1863": ("Alexander Conolly", "20005748", ""),
    "1864": ("Janet Sairsingh", "20005804", ""),
    "1865": ("Lana Watler-Rowell", "", ""),
    "1869": ("Regina Brophy", "", ""),
    "1870": ("Lennette Scott", "", ""),
    "1871": ("Melanie Moore", "", ""),
    "1873": ("Laurel Schmid", "", ""),
    "1874": ("Julia Plumley", "20595937", ""),
    "1876": ("Lorna Whitelocke", "20010280", ""),
    "1885": ("Richard Nash", "", ""),
    "1888": ("Teresa Solomon", "20005750", ""),
    "1890": ("Michelle Prince-Haylock", "", ""),
    "1891": ("Theresa Broderick", "20085230", ""),
    "1898": ("Angelina Partridge", "", ""),
    "1900": ("BrittHay Electric Ltd.", "", ""),
    "1901": ("Loren Catt", "20033113", ""),
    "1907": ("Catriona McKinnon", "", ""),
    "1911": ("Dawn Pence", "20003600", ""),
    "1913": ("Darryn Forbes", "", ""),
    "1914": ("Aleksandra Morris", "20204552", ""),
    "1915": ("Michelle Richie", "20005717", ""),
    "1916": ("Dave Stephenson", "", ""),
    "1918": ("Tiffany Polloni", "", ""),
    "1920": ("Lincoln Charles", "21727843", ""),
    "1921": ("Joy Wallace-Grant", "", ""),
    "1922": ("Kristen Orr-Depner", "", ""),
    "1924": ("Karen Bomford", "20955671", ""),
    "1925": ("David Lillico", "20033033", ""),
    "1927": ("Christine Godfray", "", ""),
    "1936": ("Puja Kohli", "", ""),
    "1938": ("Michael  Pearson - Prem 769", "20003599", ""),
    "1942": ("Richard Harris", "", ""),
    "1949": ("Lendell Layman", "21614796", ""),
    "1957": ("Elissa Costello", "20034041", ""),
    "1959": ("Michael Padarin", "", ""),
    "1962": ("Josee Sevigny", "", ""),
    "1963": ("Eric Jackson", "20032283", ""),
    "1965": ("Joan Anderson", "", ""),
    "1966": ("Jennifer  Laforge", "", ""),
    "1969": ("Yaima Delgado Diaz-Ebanks", "", ""),
    "1970": ("Cleon Green", "", ""),
    "1972": ("Abbas Mokhtari", "", ""),
    "1973": ("Dale Williams", "20071513", ""),
    "1975": ("Jason Watler", "", ""),
    "1976": ("David Searle", "", ""),
    "1977": ("Peter Goddard", "20531836", ""),
    "1978": ("Rhian Campbell", "20005778", ""),
    "1979": ("Nealy Moxam", "", ""),
    "1982": ("James Cary", "", ""),
    "1985": ("Daniel Campbell", "", ""),
    "1987": ("Viviana Ebanks", "20033050", ""),
    "1988": ("Ivet Ferguson", "21304759", ""),
    "1989": ("Linda Hearn", "", ""),
    "1990": ("Jonnie Lochner", "", ""),
    "1994": ("Christine Dolbeare", "20003610", ""),
    "1997": ("David  Marshall", "", ""),
    "1999": ("David Allen", "20002414", ""),
    "2000": ("Caribbean Publishing", "", ""),
    "2001": ("Bryce Rivers", "", ""),
    "2002": ("Ulrich Ecke", "20078875", ""),
    "2003": ("Eve Vandenbol", "20005725", ""),
    "2006": ("Sean Foley", "20003219", ""),
    "2007": ("Rafael Toledo", "", ""),
    "2008": ("Claire Kluyver", "", ""),
    "2009": ("Ross Tibbetts", "20003588", ""),
    "2010": ("Lloyd Lenard B.", "", ""),
    "2015": ("Gary Montemayor", "20032281", ""),
    "2016": ("Raymond&Cathy Cherry", "", ""),
    "2017": ("Robert Taylor - Prem 729", "20005703", ""),
    "2018": ("Stephen Ashworth", "", ""),
    "2020": ("Thomas Ballaert", "", ""),
    "2021": ("Rocco Cecere", "", ""),
    "2022": ("David Lurie", "", ""),
    "2032": ("Troy Melanson", "", ""),
    "2033": ("Alan Patino", "", ""),
    "2034": ("Jason Lord", "21641187", ""),
    "2038": ("Kyle Klischuk", "", ""),
    "2039": ("Awujoola Oyeleke", "", ""),
    "2040": ("Derek Haines", "", ""),
    "2041": ("Andrew Ngyou", "", ""),
    "2042": ("Mary Robin Winters", "20003573", ""),
    "2043": ("Gabe Powell Jr", "20047755", ""),
    "2045": ("Justin Colgan", "", ""),
    "2046": ("Sophia  Graham", "", ""),
    "2048": ("Marybeth Montrichard", "", ""),
    "2050": ("Tamara Georgakopoulos", "21684991", ""),
    "2051": ("Alfred Williams", "", ""),
    "2052": ("Ajoni Ambersley", "", ""),
    "2053": ("Angel Yates", "", ""),
    "2054": ("Nadine Dumas", "20010220", ""),
    "2057": ("Jennifer  Ahearn", "20005698", ""),
    "2058": ("Teresita Ebanks", "", ""),
    "2060": ("Steven Mirabella", "", ""),
    "2062": ("Kirkland Nixon", "", ""),
    "2063": ("Joshua Aaron Mclaughlin", "20072523", ""),
    "2064": ("Jacquie Johnston", "", ""),
    "2065": ("Bryan  Hunter", "20003159", ""),
    "2066": ("Brittney Kellett", "20031007", ""),
    "2071": ("Sarah Ferguson", "21419823", ""),
    "2072": ("Desmond Campbell", "20007215", ""),
    "2073": ("Jimel McLean", "", ""),
    "2075": ("Paul Ong", "", ""),
    "2077": ("Edward Hayward", "20012055", ""),
    "2078": ("Larry Ebanks", "", ""),
    "2079": ("Marina Din", "", ""),
    "2086": ("Wayne Griffith", "", ""),
    "2087": ("Kelsey Matern", "", ""),
    "2089": ("Ahisha Bodden", "", ""),
    "2090": ("Sukesh Shah", "", ""),
    "2096": ("Devin Petrone", "", ""),
    "2100": ("Jessica Sylvester", "", ""),
    "2101": ("Rebecca Gallagher", "", ""),
    "2104": ("Matthew Grunewald", "20033108", ""),
    "2105": ("Kieran Walsh", "20191932", ""),
    "2106": ("Sharaline Jospeh", "20072510", ""),
    "2111": ("Dierdre Flynn", "", ""),
    "2114": ("Lee Parry", "20003203", ""),
    "2115": ("JohnLee Hudson", "", ""),
    "2116": ("Brian Fertitta", "", ""),
    "2117": ("Carlos Moreno", "", ""),
    "2118": ("Terese Parker", "", ""),
    "2119": ("Travis Ritch", "20505171", ""),
    "2121": ("Dennellia Stewart", "", ""),
    "2123": ("Claire Thomas", "", ""),
    "2124": ("Samuel Banks JR - Prem 539", "21667073", ""),
    "2125": ("William Millard", "", ""),
    "2126": ("Eva Donker", "", ""),
    "2127": ("Natalie Belgrave", "", ""),
    "2128": ("Alison Lebitschnig", "", ""),
    "2129": ("Rachael Windhaber", "", ""),
    "2130": ("Noel Reilly", "", ""),
    "2131": ("Lindsey Swift", "", ""),
    "2132": ("Ailbhe Kane", "", ""),
    "2134": ("Jaime Cormack", "", ""),
    "2135": ("Kerry Smithies", "", ""),
    "2136": ("Joseph Gladish", "20010132", ""),
    "2137": ("James Scutt", "", ""),
    "2138": ("Telicia Ebanks", "", ""),
    "2140": ("Bill Shewan", "", ""),
    "2141": ("Lyne O\\'Doherty", "", ""),
    "2142": ("Laura Goucke", "20217490", ""),
    "2143": ("Rebecca Kingsley-Bates", "", ""),
    "2144": ("Jamaal Anderson", "20009975", ""),
    "2146": ("Ishmael DaSIlva", "20030607", ""),
    "2147": ("Duncan Hancock", "20004772", ""),
    "2148": ("Darla Dilbert", "", ""),
    "2149": ("Shane Troughton", "", ""),
    "2150": ("Thora Forbes", "", ""),
    "2151": ("Camilla Anderson", "", ""),
    "2152": ("Jesse Basded", "", ""),
    "2153": ("Bonny Vincent", "", ""),
    "2154": ("Brian Ogilvie", "20004865", ""),
    "2155": ("Sarah Derebenstedt", "", ""),
    "2156": ("Erin Turpin", "", ""),
    "2157": ("Patricia Priestley", "", ""),
    "2158": ("Charles Quappe", "20033056", ""),
    "2159": ("Jonathan Allen", "20143365", ""),
    "2160": ("Daniel McKenna", "", ""),
    "2161": ("Craig Merren", "20012860", ""),
    "2162": ("Cesar Rico Aparil", "", ""),
    "2163": ("Mikhail  Bakalov", "20003132", ""),
    "2164": ("Simon Watson", "", ""),
    "2165": ("Francina Russell", "", ""),
    "2166": ("John McDow (Martinez)", "20005724", ""),
    "2168": ("Kayren Bodden", "", ""),
    "2169": ("Lydia Jackman", "", ""),
    "2170": ("Terry Stevenson", "", ""),
    "2171": ("Margaret Morash", "", ""),
    "2172": ("Carol O\\'neill", "", ""),
    "2174": ("Geraldine Ebanks", "20004842", ""),
    "2175": ("Jan Neveril", "", ""),
    "2176": ("Marie Shepheard", "", ""),
    "2177": ("Rosa Harris", "", ""),
    "2178": ("Yves Martel", "20005752", ""),
    "2179": ("Harry Lynch", "", ""),
    "2180": ("Allison Thorburn", "", ""),
    "2181": ("Kimberly Dawson", "", ""),
    "2182": ("Nicole Frolick", "", ""),
    "2183": ("Roxane Ebanks", "20022079", ""),
    "2184": ("Erik Fell", "20033070", ""),
    "2185": ("Angela Tanzillo-Swarts", "", ""),
    "2186": ("Anne Dolan", "20031032", ""),
    "2187": ("Rafael Fadipe", "", ""),
    "2188": ("Cheri Langston", "", ""),
    "2189": ("Yvette Cacho", "20035687", ""),
    "2190": ("Paul Osborne", "20061747", ""),
    "2191": ("JeanLyn Ebanks", "", ""),
    "2193": ("Bill Newton", "20086824", ""),
    "2194": ("John Douglas", "", ""),
    "2195": ("Vicki Calvert-Das", "", ""),
    "2196": ("Benjamin Leftwich", "20033284", ""),
    "2197": ("Liza Smith", "20061749", ""),
    "2199": ("Andres Bowman", "20011966", ""),
    "2200": ("Chelsea Green", "20003556", ""),
    "2201": ("Theodore Peyton", "", ""),
    "2202": ("Chela Ebanks", "", ""),
    "2203": ("National Gallery of the Cayman Islands", "", ""),
    "2204": ("Christina Bodden", "", ""),
    "2205": ("Donald Miller", "", ""),
    "2206": ("Alan Rohleder", "", ""),
    "2207": ("Lucy Johns", "21650026", ""),
    "2208": ("Antarjot Ahluwalia", "", ""),
    "2209": ("Tricia Whittaker-cain", "", ""),
    "2210": ("Rasheed Suazo", "20004887", ""),
    "2211": ("Damon Bilchuris", "20003601", ""),
    "2212": ("Alexander Randall", "", ""),
    "2213": ("XavierDalvin Ebanks", "", ""),
    "2214": ("Desiree Ward", "", ""),
    "2216": ("Shekeisha Connor", "20036231", ""),
    "2217": ("Kameron Mclean", "", ""),
    "2218": ("Paul Harris", "20006298", ""),
    "2219": ("Christina Travers", "", ""),
    "2220": ("Lilia Conolly", "20109619", ""),
    "2221": ("Carlene Alexander-Kay", "20004777", ""),
    "2222": ("Evelyn Tibbetts-Farrar", "", ""),
    "2223": ("Dale Hall", "", ""),
    "2224": ("Ron Walker", "", ""),
    "2225": ("Keiran Hutchison", "", ""),
    "2226": ("Nasaria Chollette", "", ""),
    "2227": ("Channing Mclean", "", ""),
    "2228": ("John Calhoun", "20076772", ""),
    "2229": ("Charley Ebanks", "20005704", ""),
    "2230": ("Marcia Kramer", "", ""),
    "2231": ("Sherry Bispath", "", ""),
    "2232": ("Adam Milburn", "", ""),
    "2233": ("Charles Lawrence", "", ""),
    "2234": ("Kimberly Watler", "", ""),
    "2235": ("Robert Campbell", "", ""),
    "2236": ("Colin Lumsden", "", ""),
    "2239": ("Marshall Jonathan", "", ""),
    "2240": ("Dolphin Discovery Cayman Ltd.", "", ""),
    "2241": ("Alison Armstrong", "21636870", ""),
    "2242": ("Denise Gower", "", ""),
    "2244": ("Lisa Hart-Boni", "", ""),
    "2245": ("Stephanie Krco", "", ""),
    "2246": ("Hana Trekell", "", ""),
    "2247": ("Krysten Schieltz - Prem 147", "", ""),
    "2248": ("Daniel Gower", "", ""),
    "2250": ("Jonathon Anglin", "20109634", ""),
    "2251": ("Aneal Sobie", "20033100", ""),
    "2252": ("Tisha Cooper", "21630705", ""),
    "2253": ("Logan-Caleb Minshew", "", ""),
    "2254": ("Kevin Shrestha", "", ""),
    "2255": ("SeanDanielJoseph Bodden", "21661168", ""),
    "2256": ("Ronald Hughes", "", ""),
    "2257": ("Elizabeth Kennedy", "", ""),
    "2258": ("Natasha Bodden", "20567317", ""),
    "2259": ("James Berry", "", ""),
    "2260": ("Kieron Rankin", "20004764", ""),
    "2261": ("Roshawn Robb", "", ""),
    "2262": ("Priscilla Puzo", "", ""),
    "2263": ("Hannah Carter", "", ""),
    "2264": ("Judy Yung", "20010213", ""),
    "2265": ("Adrian Rowe", "", ""),
    "2266": ("Karim Panju", "", ""),
    "2267": ("Richard Spencer", "", ""),
    "2268": ("Janet Mackey", "", ""),
    "2270": ("Kurt Klischuk", "20005764", ""),
    "2271": ("David Campbell", "20035672", ""),
    "2273": ("Heather Cassidy", "", ""),
    "2274": ("Sean Rankine", "", ""),
    "2275": ("Krista McLean", "", ""),
    "2277": ("Corey Cato", "", ""),
    "2278": ("Emily Panke", "", ""),
    "2279": ("Lori Scott", "20010002", ""),
    "2280": ("Trevor Murgio", "20340774", ""),
    "2282": ("Clifton Rose", "", ""),
    "2283": ("Jared Awe", "20010129", ""),
    "2284": ("Kendra Sepulveda", "", ""),
    "2285": ("Wayne Alexander", "20011952", ""),
    "2286": ("Melisa Makridakis", "", ""),
    "2287": ("Stephney Williams", "", ""),
    "2288": ("Bryan Mclaughlin", "", ""),
    "2289": ("Kendra Rankin", "", ""),
    "2290": ("Krister Mercado", "", ""),
    "2291": ("Tatiana Pattman", "", ""),
    "2292": ("Yasmine Mardelli", "", ""),
    "2293": ("Isaac Espinoza", "", ""),
    "2294": ("Gelia Frederick", "", ""),
    "2295": ("Tamara (Tara)  Nielsen", "", ""),
    "2296": ("Sean Knight", "", ""),
    "2297": ("KimberlyF. Charoo", "", ""),
    "2298": ("Tammea Ebanks", "20004877", ""),
    "2299": ("Satisha Brandon", "", ""),
    "2300": ("Eryka Simmons", "20003605", ""),
    "2301": ("Thomas Madden", "", ""),
    "2302": ("Cinthya Zerpa", "21636020", ""),
    "2303": ("Michelle Wight", "", ""),
    "2304": ("Jennifer Philpott", "20005753", ""),
    "2305": ("Richard Andrews", "20006299", ""),
    "2306": ("Ranjit Gill", "", ""),
    "2307": ("Justin  Tibbetts", "20035690", ""),
    "2308": ("Debora Gill", "", ""),
    "2309": ("Charolette Hoffman", "20005800", ""),
    "2310": ("Michelle Ebanks", "", ""),
    "2311": ("Paisley Taylor", "", ""),
    "2312": ("Esmond Wright", "20010126", ""),
    "2313": ("Daniel Marinescu", "", ""),
    "2314": ("Michaela Pultr", "20039278", ""),
    "2315": ("Allison Lawrence", "", ""),
    "2316": ("Roman Ipfling", "", ""),
    "2317": ("Kristin Powell", "", ""),
    "2318": ("Fernando Soto", "", ""),
    "2319": ("Lisa Royal", "20004773", ""),
    "2320": ("Guillermo Perez-Ramirez", "", ""),
    "2321": ("Ashley Martynec", "", ""),
    "2322": ("Karrah Glen", "", ""),
    "2323": ("Wanda Rice", "", ""),
    "2324": ("Andrew Small", "21633942", ""),
    "2325": ("Charlene Rivers", "21767406", ""),
    "2326": ("Damien Austin", "", ""),
    "2327": ("Frankie Pappas", "", ""),
    "2328": ("Joshua Ebanks", "", ""),
    "2329": ("Christine Maltman", "20002365", ""),
    "2330": ("Susan Jones", "21639746", ""),
    "2331": ("Michael Rivers", "", ""),
    "2332": ("Michael Shield", "20007229", ""),
    "2333": ("Alice Fallon", "", ""),
    "2334": ("Darcia Hamilton", "", ""),
    "2335": ("ChrishaunnaT Trowers", "", ""),
    "2336": ("Caroline Anderson", "20005811", ""),
    "2337": ("Ashley Higgs", "", ""),
    "2338": ("Elizabeth Ross", "20003579", ""),
    "2339": ("Helen Ebanks", "", ""),
    "2340": ("JamesDale Lavender", "", ""),
    "2341": ("Hannah Reid", "20030553", ""),
    "2342": ("Jimmy Bodden", "20005755", ""),
    "2343": ("Gary Smith", "20003226", ""),
    "2344": ("Celia DaCosta Gomez", "", ""),
    "2345": ("Renato Cabrera", "", ""),
    "2346": ("Evett Brown", "20030976", ""),
    "2347": ("Stephen Boyd", "20073032", ""),
    "2348": ("Patricia Mendez", "20076505", ""),
    "2349": ("Paul Palmer", "", ""),
    "2350": ("Benecia L. Jackson", "", ""),
    "2351": ("Nicole Crance", "", ""),
    "2352": ("Nan Erb", "", ""),
    "2353": ("Sheila Aronfeld", "20030589", ""),
    "2354": ("Cordon Harris", "", ""),
    "2355": ("Rachel Fisher", "", ""),
    "2356": ("Ingrid Pierce", "20754881", ""),
    "2359": ("Doug Moffatt", "21597803", ""),
    "2360": ("Sheldon Clarke", "", ""),
    "2361": ("April Panton", "20009993", ""),
    "2362": ("Carol Jackson", "21775481", ""),
    "2363": ("Jordan McLaughlin", "", ""),
    "2364": ("Christopher Keefe", "", ""),
    "2365": ("Fabio Castillio", "", ""),
    "2366": ("Dogulas Cameron", "21633320", ""),
    "2367": ("Julie hurlston", "", ""),
    "2368": ("Heidui Whittaker", "", ""),
    "2369": ("Souhilla Moore", "20031027", ""),
    "2370": ("James Padden", "20072511", ""),
    "2372": ("Leanne Wright", "", ""),
    "2373": ("Penelope Wunsch", "20001994", ""),
    "2374": ("Angela Smith", "", ""),
    "2375": ("Jennifer Cowdroy", "20031006", ""),
    "2376": ("Geetha Alagirisamy", "", ""),
    "2377": ("MohammodHabi Masood", "", ""),
    "2378": ("Yanique Mitchell", "", ""),
    "2379": ("Giselle Toth", "", ""),
    "2380": ("Avril Ward", "", ""),
    "2381": ("Amber Myhand", "", ""),
    "2382": ("Alasdair Foster", "20030582", ""),
    "2383": ("Shantie Ramoutar", "", ""),
    "2384": ("Todd Hazlewood", "21597697", ""),
    "2385": ("Jason Wong", "", ""),
    "2386": ("Heidi Whittaker", "", ""),
    "2387": ("Collen Martin", "", ""),
    "2388": ("Ricardo Radwanski", "", ""),
    "2389": ("Donald Gass - Prem 755", "20006582", ""),
    "2390": ("Gina  Pick", "", ""),
    "2391": ("Daniel Cimring", "20211152", ""),
    "2392": ("Lee Quenard", "", ""),
    "2393": ("Anoushka Seebaluk", "20005754", ""),
    "2396": ("Retno Purbasanti", "", ""),
    "2398": ("Dr.Sharon Chambers", "", ""),
    "2399": ("James Shapland", "", ""),
    "2400": ("Heather Barnett", "", ""),
    "2401": ("Rosemarie Lambert", "20005700", ""),
    "2402": ("Kirstie Miller", "", ""),
    "2403": ("Jennifer Dixon", "20080363", ""),
    "2404": ("Agnieszka Linkowski", "", ""),
    "2405": ("Anna-Marie Propper", "", ""),
    "2406": ("Bartlomiej Jeske", "", ""),
    "2407": ("Marcelo Suarez Castillio", "20030996", ""),
    "2408": ("Victoria Hunter", "", ""),
    "2409": ("Adam Kambeitz", "", ""),
    "2410": ("Kierstin Brie-Ashley", "", ""),
    "2411": ("Melissa Thomas", "20792910", ""),
    "2412": ("Darlene Young", "20033201", ""),
    "2414": ("Raymond Bennett", "", ""),
    "2415": ("Justin Appleyard", "20007453", ""),
    "2416": ("Roland Talanow", "20004946", ""),
    "2417": ("Al Ramtahal", "", ""),
    "2418": ("Sandria Gardner", "", ""),
    "2419": ("Candace Rankin", "20047747", ""),
    "2420": ("Joy Simpson", "", ""),
    "2421": ("Ebone Solomon", "", ""),
    "2423": ("Kirk  Douglas", "21597861", ""),
    "2424": ("M. Natasha Haye", "", ""),
    "2425": ("Camille Koo", "", ""),
    "2426": ("Karilee Campbell- Thomas", "", ""),
    "2427": ("Nancy  Whittaker", "20039274", ""),
    "2428": ("Andrea Goodall", "20033141", ""),
    "2429": ("Selborn Handsack", "", ""),
    "2430": ("Cheryl Ann Bramwell", "", ""),
    "2431": ("Aziza LaPierre", "20109618", ""),
    "2432": ("Carey  Burns", "20004892", ""),
    "2433": ("Dapinderjit Ghuman", "", ""),
    "2434": ("Lisa E Williams", "", ""),
    "2435": ("Claire Leadbeater", "", ""),
    "2436": ("Rhian  Minty", "", ""),
    "2437": ("Kelly Pettit", "", ""),
    "2439": ("Laura Powell", "20306562", ""),
    "2440": ("Ghislaine Egan", "20007488", ""),
    "2441": ("Lisa Boushy", "", ""),
    "2442": ("Karthryn  Dinspel- Powell", "", ""),
    "2443": ("Harry Boyce", "20004792", ""),
    "2444": ("Vikram Dookhy", "", ""),
    "2445": ("Cayman ABA Ltd", "21634507", ""),
    "2446": ("Sebastien Nehme", "", ""),
    "2447": ("Christine Groves", "20003570", ""),
    "2448": ("Junior Wong", "", ""),
    "2449": ("Sonia Persaud", "20003157", ""),
    "2450": ("Robert Coombes", "20005727", ""),
    "2451": ("Heidi  Blair", "", ""),
    "2452": ("Susan Grizzel", "", ""),
    "2453": ("Junier  Ferguson", "", ""),
    "2454": ("Nichola Le Prevost", "", ""),
    "2455": ("Chantol Hurlston", "20009977", ""),
    "2456": ("Danny Raymond", "", ""),
    "2457": ("Melissa Vail", "", ""),
    "2458": ("David Mars", "", ""),
    "2459": ("Ania Gutierrez", "", ""),
    "2460": ("Wilma Nelson", "", ""),
    "2461": ("Lucy Sleep-Powery", "20004484", ""),
    "2462": ("Padraic J. Linnane", "", ""),
    "2463": ("Angela  Lumley", "", ""),
    "2464": ("Tanya Dube", "21523050", ""),
    "2465": ("Connie Tang", "", ""),
    "2466": ("Alexandra Hughes", "", ""),
    "2467": ("Regina Ecclefield", "", ""),
    "2468": ("Erica Ffrench", "20664833", ""),
    "2469": ("Karin Kolbl Tibbetts", "20003202", ""),
    "2470": ("Robyn Pharr", "", ""),
    "2471": ("Carli Du Buisson", "20006713", ""),
    "2472": ("Yeisis  Anez", "21523024", ""),
    "2474": ("Marco  Miranda", "20003615", ""),
    "2475": ("Ashley Stafford", "20037833", ""),
    "2476": ("Lionel Durrant", "20004873", ""),
    "2477": ("Barbara MacDonald", "", ""),
    "2478": ("Chadd Bush", "", ""),
    "2479": ("Cayman Arts Festival", "20021071", ""),
    "2480": ("Nette Bulgin", "", ""),
    "2481": ("Michael Balkissoon", "20086487", ""),
    "2482": ("Michael Parton", "20030357", ""),
    "2483": ("Dion Minzett", "", ""),
    "2484": ("Sari Sumaryono", "20034832", ""),
    "2485": ("Sonita Malan", "20003218", ""),
    "2486": ("Joanna Powery", "20010277", ""),
    "2487": ("P Natalia Turnquest", "", ""),
    "2488": ("Tishel Mclean", "", ""),
    "2489": ("Roland Iton", "20039279", ""),
    "2490": ("Samantha  Ramirez", "20004769", ""),
    "2492": ("Chris-Lobain Wedderbrun", "", ""),
    "2493": ("Paul Murphy", "", ""),
    "2494": ("Denielle Pope", "", ""),
    "2495": ("Andrea Nixon", "", ""),
    "2496": ("Aleks Beckford", "", ""),
    "2497": ("Daniel McCarthy", "", ""),
    "2498": ("Carol  Dixon", "20003215", ""),
    "2499": ("Charles Johnson", "", ""),
    "2500": ("Osman Atiq", "", ""),
    "2501": ("Justin Robinson", "", ""),
    "2502": ("Yolande Lopez", "21523023", ""),
    "2503": ("Bridgette Powery", "20009981", ""),
    "2504": ("Paula Ross", "20004868", ""),
    "2505": ("Natalee Steele", "21723117", ""),
    "2506": ("Jeff George", "20032425", ""),
    "2507": ("John Davies", "", ""),
    "2508": ("Carol Edwards", "", ""),
    "2509": ("Deri Hill", "", ""),
    "2510": ("Carlo Lee", "", ""),
    "2511": ("Renee Rankin", "", ""),
    "2512": ("Stafford Millwood", "", ""),
    "2513": ("Alexandra Bartlett", "20005775", ""),
    "2514": ("Daniel  Tathum", "20003155", ""),
    "2520": ("Rena Strecker", "20024805", ""),
    "2521": ("Consie Nekita Williams", "20030989", ""),
    "2522": ("David  Nielsen", "20073684", ""),
    "2523": ("Paul Foura", "", ""),
    "2524": ("Samara  Ebanks", "20003595", ""),
    "2525": ("Wade  Ebanks", "", ""),
    "2526": ("Ricardo  Mclaughlin", "", ""),
    "2527": ("Gerrit J V Rensburg", "20004908", ""),
    "2528": ("Sharon  Campbell", "", ""),
    "2529": ("Michael  Johnson", "21630786", ""),
    "2530": ("Hannah  Robinson-Scholfield", "", ""),
    "2531": ("Este Rabey", "", ""),
    "2532": ("Asha  Charles", "", ""),
    "2533": ("Hevard Smith", "20035691", ""),
    "2534": ("Derren  Burlington", "20217192", ""),
    "2535": ("Orla O\\'Conner", "20031041", ""),
    "2536": ("Atlas Ramoon", "20664126", ""),
    "2538": ("Gary Todd", "", ""),
    "2539": ("Ceili  Fitzgerald", "", ""),
    "2540": ("Scott Suddaby", "", ""),
    "2541": ("Shanique Wright", "", ""),
    "2542": ("Bertha  Moore", "", ""),
    "2543": ("Deniel J. James", "20033195", ""),
    "2544": ("Jo Ann Smith", "21691924", ""),
    "2545": ("Ashleigh Bodden", "", ""),
    "2546": ("Robert  Robinson", "", ""),
    "2547": ("Laurel Ellis", "20005810", ""),
    "2548": ("Saulo Castro Martinez", "", ""),
    "2549": ("Jennifer Mckinney", "", ""),
    "2561": ("Dwight Scott", "", ""),
    "2562": ("Daniel Ebanks", "", ""),
    "2563": ("Venice Campbell", "20009997", ""),
    "2564": ("Candy Bodden", "20084031", ""),
    "2565": ("William Mclaughlin", "", ""),
    "2567": ("Robert Anthony", "", ""),
    "2568": ("Paul  Labrosse", "21740946", ""),
    "2569": ("Maria Kharitidi", "", ""),
    "2570": ("Heather  Nugent", "20030560", ""),
    "2571": ("Patrick Schmid", "20215797", ""),
    "2572": ("Mary Cardona", "", ""),
    "2573": ("Anisha Montemayor", "21633039", ""),
    "2574": ("Shenice  Mcfield", "20078878", ""),
    "2575": ("Helen Gagnon", "", ""),
    "2576": ("Michael Tapajna", "20004945", ""),
    "2578": ("Oscar E. Vaca", "20003227", ""),
    "2586": ("Pamela Thompson", "20003590", ""),
    "2587": ("Lynda Nutt", "20030369", ""),
    "2588": ("Karen  Galloway-Blake", "20034045", ""),
    "2589": ("Grace Myrie", "", ""),
    "2590": ("Bobeth O\\\\\\'Garro", "", ""),
    "2591": ("William Deckelman", "", ""),
    "2592": ("David  Coveney", "", ""),
    "2593": ("Jodi Smith-Jones", "20004858", ""),
    "2594": ("Sydney Libby", "", ""),
    "2595": ("Carlington Grant", "", ""),
    "2596": ("Amber  Ramsey", "20004853", ""),
    "2597": ("Kirksey William", "20010000", ""),
    "2628": ("Alison  Murphy", "20039807", ""),
    "2629": ("Brandon McLean", "20030550", ""),
    "2639": ("Maureen Watler", "20016094", ""),
    "2643": ("Hannah Myrie", "21752984", ""),
    "2644": ("Sue Merren", "20005734", ""),
    "2645": ("jose  vela", "", ""),
    "2646": ("Jeremy Bodden", "", ""),
    "2648": ("Christopher Wall", "", ""),
    "2649": ("Shaun Maloney", "20011792", ""),
    "2650": ("Darron Conolly", "20010218", ""),
    "2651": ("Maria Barnes-Campbell", "20546685", ""),
    "2653": ("Marius Deysel", "", ""),
    "2654": ("Cindy Tiofilo", "", ""),
    "2655": ("Marc Williams", "20037532", ""),
    "2656": ("Laura Bryson", "20220487", ""),
    "2657": ("Justine Jones", "", ""),
    "2658": ("Patrick Agemian", "20618684", ""),
    "2661": ("Harvey Stephenson", "", ""),
    "2662": ("Bianca Francis", "20032406", ""),
    "2664": ("Daniela Rico", "", ""),
    "2665": ("brittany  solomon", "", ""),
    "2666": ("Lakiva Reid", "", ""),
    "2667": ("Kathryn Wild", "", ""),
    "2669": ("Kristy Capewell", "", ""),
    "2670": ("M y J Nissen", "", ""),
    "2671": ("Richard Elder", "", ""),
    "2672": ("Dale Stoudt", "20005728", ""),
    "2673": ("Donald McLean", "20004909", ""),
    "2674": ("Martin Tedd", "", ""),
    "2675": ("Margo Richardson", "", ""),
    "2676": ("Jetena Bodden", "20212028", ""),
    "2677": ("Joshua Browne", "", ""),
    "2678": ("Ramona Tudorancea", "20072508", ""),
    "2683": ("Alexandra Stewart", "", ""),
    "2684": ("Linda Mason", "20009999", ""),
    "2686": ("Dimitrious Nasirpour", "", ""),
    "2687": ("kristy Blackburn", "", ""),
    "2688": ("Michael Carroll", "", ""),
    "2689": ("Camille Humphreys", "", ""),
    "2690": ("Candice Gliksman", "", ""),
    "2691": ("Androulla Hill", "", ""),
    "2692": ("Frances Fharina Nones", "", ""),
    "2693": ("Rhonda  Whitely", "", ""),
    "2694": ("Elmelinda Maldonado", "", ""),
    "2695": ("shawn solomon", "", ""),
    "2696": ("Innocenza La Civita", "", ""),
    "2697": ("Chelsea Mcfield", "20003618", ""),
    "2698": ("Taneisha  Welcome", "", ""),
    "2699": ("Jade Lyn", "", ""),
    "2700": ("Darren Goodall", "", ""),
    "2701": ("Courtney y Chris  Burke", "", ""),
    "2702": ("Jirarat Chaidee", "20541446", ""),
    "2703": ("Kirt Munroe", "20003553", ""),
    "2704": ("Andie Welsch", "20005813", ""),
    "2705": ("George scharnikow", "", ""),
    "2706": ("Karen Whyte", "", ""),
    "2707": ("Courtney Criswell", "20005729", ""),
    "2708": ("Wayne Forrester", "20061740", ""),
    "2709": ("Emmanuelle Cook", "21636295", ""),
    "2710": ("Christopher Jackson", "", ""),
    "2711": ("Janine Gregory", "20003213", ""),
    "2712": ("Wendy Jackson Didier", "", ""),
    "2713": ("Lincoln  Bodden", "20010130", ""),
    "2714": ("Fraser Hughes", "", ""),
    "2715": ("Judith Kimote", "", ""),
    "2716": ("Nadine Gray", "20080993", ""),
    "2717": ("Tiphanie Wilmot", "20009785", ""),
    "2718": ("Daniella Bodden", "20005770", ""),
    "2719": ("Karen Plummer", "", ""),
    "2720": ("Gyongyver Bartha", "", ""),
    "2721": ("Sean Inggs", "", ""),
    "2722": ("Shari Espeut", "", ""),
    "2723": ("Ashley Rivers", "", ""),
    "2724": ("Andrea Lacey", "", ""),
    "2725": ("Catherine Rennie", "", ""),
    "2727": ("Nadine Holness", "", ""),
    "2728": ("Kai  Bryant", "20003584", ""),
    "2729": ("Alastair Bird", "", ""),
    "2730": ("Kristan  La Tache", "20121433", ""),
    "2731": ("Allen Shankar Ebanks", "20004767", ""),
    "2732": ("Janiecia Connolly", "21688627", ""),
    "2733": ("Luis  Guillen", "20750928", ""),
    "2734": ("Belinda  Joseph", "20003602", ""),
    "2735": ("Brad Jones", "", ""),
    "2736": ("Lauren Herbert", "", ""),
    "2737": ("Jonathan Oriole", "", ""),
    "2738": ("Victoria Banks", "", ""),
    "2739": ("Tanya Wright", "20218572", ""),
    "2740": ("Janet  Young", "20072515", ""),
    "2741": ("Edwin Pellot-Rosa", "", ""),
    "2742": ("Kody Zander", "20002533", ""),
    "2743": ("John Ackerley", "21642029", ""),
    "2744": ("Natalie Wright", "20030587", ""),
    "2745": ("tricia purchas", "20011423", ""),
    "2746": ("David Irvine", "20035675", ""),
    "2747": ("Greg Bennett", "21620067", ""),
    "2748": ("meallisa Douglas", "20010567", ""),
    "2749": ("Freedi Montejo", "", ""),
    "2750": ("Johnelle Hylton", "", ""),
    "2751": ("Corriel Orrett", "", ""),
    "2752": ("Carlos Ebanks", "20675161", ""),
    "2753": ("Brian Shum Ying Nin", "20002264", ""),
    "2754": ("Prashant Sharma", "", ""),
    "2755": ("Joel  Bodden", "20061728", ""),
    "2756": ("Leslie Ford", "20003592", ""),
    "2757": ("Christine Eden", "", ""),
    "2758": ("Marsha-gay James Watson", "", ""),
    "2759": ("Lisaida Swaby Oliva", "20003576", ""),
    "2760": ("Ella Pineda", "", ""),
    "2761": ("Noah Seymour", "", ""),
    "2762": ("Robert Cameron  Leitch", "", ""),
    "2763": ("Jason Wood", "", ""),
    "2764": ("Nicolas Franco", "20221009", ""),
    "2765": ("Carrie Patraulea", "20004787", ""),
    "2766": ("Natalie Martin", "", ""),
    "2767": ("Joel  Clark", "", ""),
    "2768": ("Christopher Bouck", "20003613", ""),
    "2769": ("Yanique Mcfarlane", "", ""),
    "2770": ("Lois Myles-Davis", "20667518", ""),
    "2771": ("Donna Paterson", "", ""),
    "2772": ("Chelsea Blake", "", ""),
    "2774": ("Bernadette Carey", "20003557", ""),
    "2775": ("Rommel Coe", "", ""),
    "2776": ("Kathleen Corkey", "", ""),
    "2777": ("Melanie Stafford", "", ""),
    "2778": ("Lesley Walker", "", ""),
    "2779": ("Ronald Hatch", "20007463", ""),
    "2780": ("Mary Grace Melanio", "", ""),
    "2781": ("Christopher Sitarski", "", ""),
    "2782": ("Nickolas DaCosta", "", ""),
    "2783": ("Kial Whorms", "20002256", ""),
    "2784": ("Tiffany Lucas", "20006297", ""),
    "2785": ("Simon Cooper", "20206869", ""),
    "2786": ("Natasha Hydes", "20215137", ""),
    "2787": ("Brian Phelps", "20030568", ""),
    "2788": ("Sheneen Powell", "", ""),
    "2789": ("Miranda Camp", "20398416", ""),
    "2790": ("Jeremy Josephs", "20037518", ""),
    "2791": ("CHERYL YOUNG", "", ""),
    "2792": ("Jessie Huber", "", ""),
    "2793": ("Darryn Monaghan", "", ""),
    "2794": ("Samantha Green", "20033203", ""),
    "2795": ("Peter  Fiesel", "", ""),
    "2796": ("Cydric Desbarida", "", ""),
    "2797": ("Jillian Burley", "20084209", ""),
    "2798": ("James Parham", "20003158", ""),
    "2799": ("John Ern Tze Tan", "20037526", ""),
    "2800": ("Laurence Dawkes", "", ""),
    "2801": ("Zalisha Mohamed", "", ""),
    "2802": ("Michael Chester", "", ""),
    "2803": ("Addie Ade-Yusuf", "", ""),
    "2804": ("Ali Beckford", "20046067", ""),
    "2805": ("Corvin McLean", "20210883", ""),
    "2806": ("Matthew Laudon", "", ""),
    "2807": ("Amitabha Basu", "", ""),
    "2808": ("Hugh Lanigan", "", ""),
    "2809": ("Shonari Gow", "20032296", ""),
    "2810": ("Charlene Brady", "", ""),
    "2811": ("Steven Stewart", "", ""),
    "2812": ("Alison  Glasgow", "", ""),
    "2813": ("Andrew y Sally Peedom", "20218555", ""),
    "2814": ("Heather Carrigan", "20005702", ""),
    "2815": ("Staci Scott", "20030322", ""),
    "2816": ("Josh Caballero", "", ""),
    "2817": ("TETSUZAN RON", "", ""),
    "2818": ("Roger Campbell", "", ""),
    "2819": ("Morelda Gentles", "21631665", ""),
    "2820": ("Donna  Morris", "", ""),
    "2821": ("IMTIAZ ALI", "", ""),
    "2822": ("Richard Love", "", ""),
    "2823": ("Sophia Webb", "", ""),
    "2824": ("Jonelle  Campbell", "20003569", ""),
    "2825": ("Leon Young", "20191681", ""),
    "2826": ("Orla O\\'Regan", "", ""),
    "2827": ("Joanne Mimnagh", "20546020", ""),
    "2828": ("Ria Scott", "", ""),
    "2829": ("Dominic Orlando", "20036177", ""),
    "2830": ("Beth McCrae", "20121404", ""),
    "2831": ("Ann Serrant", "", ""),
    "2832": ("Carrie Emslie", "20009992", ""),
    "2833": ("Kennisha Dominguez", "20004861", ""),
    "2834": ("alicia hawthorne", "20004911", ""),
    "2835": ("Natalie Lavis", "", ""),
    "2836": ("Georgena Seymour", "", ""),
    "2837": ("Divad Aguirre-Britton", "20016124", ""),
    "2838": ("Rainer Dreyer", "", ""),
    "2839": ("Janell Taylor", "21597808", ""),
    "2840": ("Marco Pompei", "21628496", ""),
    "2841": ("Alasdair Foster", "20030582", ""),
    "2842": ("Sweet Roselle  Viado Labitad", "20010123", ""),
    "2843": ("Al-Siddique Allie", "20039811", ""),
    "2844": ("Laurie Jackson", "", ""),
    "2845": ("Aisha Zenith Smith", "20007212", ""),
    "2846": ("Paolo Cognolato", "", ""),
    "2847": ("Julie Adam", "", ""),
    "2849": ("Danielle  Connolly", "", ""),
    "2850": ("Diane Donovan", "20003161", ""),
    "2851": ("Cheslyn Van Aswegen", "20003611", ""),
    "2852": ("Kevin Forrest", "", ""),
    "2853": ("Alaa Hijazi", "", ""),
    "2854": ("Laura Young", "", ""),
    "2855": ("Sheryl Black", "", ""),
    "2856": ("skyler warren", "", ""),
    "2857": ("Yannick Whorms", "20031040", ""),
    "2858": ("Maxine Bird", "", ""),
    "2859": ("Tony De Quintal", "", ""),
    "2861": ("Karen Hall", "", ""),
    "2862": ("Mark Robinson", "", ""),
    "2863": ("Sandra Feduszczak", "", ""),
    "2864": ("Heidi Adair", "20010278", ""),
    "2865": ("Clayton McGowan", "20210926", ""),
    "2866": ("Jonathon Fraser", "", ""),
    "2867": ("Jorge Zaldivar", "20003199", ""),
    "2868": ("luciana chilmaza", "", ""),
    "2869": ("Natalie Bogle", "21635221", ""),
    "2870": ("Caroline Barnes", "", ""),
    "2871": ("Nicoela McCoy", "20004784", ""),
    "2872": ("Pablo Mejia", "", ""),
    "2873": ("Andrea Balajadia", "", ""),
    "2874": ("Schian Scott", "20751007", ""),
    "2875": ("Crystal Godet", "", ""),
    "2876": ("Sherine Stewart", "", ""),
    "2877": ("Simon Crompton", "21639628", ""),
    "2878": ("JAMES CHAPMAN", "20003138", ""),
    "2879": ("Allain Forget", "", ""),
    "2880": ("TROPICAL EYE  CARE LTD", "20032286", ""),
    "2881": ("Chris  Hadome", "20003574", ""),
    "2882": ("Nikisha Scott Powell", "20004465", ""),
    "2883": ("Alfredo Cardoza", "", ""),
    "2884": ("Freddy  Diaz", "", ""),
    "2885": ("Jouri Haylock", "", ""),
    "2886": ("Jouri Haylock", "", ""),
    "2887": ("Connie Godet", "20002062", ""),
    "2888": ("Jason Taylor", "", ""),
    "2889": ("ANDREA CHRISTIAN", "20005799", ""),
    "2890": ("Anthony Baker", "", ""),
    "2891": ("Peter Tyers-Smith", "", ""),
    "2892": ("Richard Gordon", "20072080", ""),
    "2893": ("linda thompson", "", ""),
    "2894": ("Ying Han", "", ""),
    "2895": ("Brooke Clark", "", ""),
    "2896": ("Katrina Jurn", "20005726", ""),
    "2897": ("Franklin Smith", "", ""),
    "2898": ("Brenda McGrath", "20074327", ""),
    "2899": ("David Griffin", "20553412", ""),
    "2900": ("Alistair Abbott", "", ""),
    "2901": ("Naddine Davis", "", ""),
    "2902": ("Stewart Siebens", "20004789", ""),
    "2903": ("Xavier  Ebanks", "", ""),
    "2904": ("Maja Ludford-Thomas", "21641186", ""),
    "2905": ("Anthony Akiwumi", "", ""),
    "2906": ("URSULA ZUBIZARRETA", "", ""),
    "2907": ("Peter Pasold", "", ""),
    "2908": ("Jeremy Walsh", "20754879", ""),
    "2909": ("Anju Whittaker", "", ""),
    "2910": ("Kevin McLean", "20002248", ""),
    "2911": ("Jose Vela", "", ""),
    "2912": ("javier zaglul", "", ""),
    "2913": ("Kathleen Chapman", "20004486", ""),
    "2914": ("Jenifer Royer Thompson", "", ""),
    "2915": ("Mr leslie Hydes Jr", "20003585", ""),
    "2916": ("Paula Stonoga Ferreira", "20080356", ""),
    "2917": ("Lori Halldorson", "20003586", ""),
    "2918": ("Michael Loewen", "20191713", ""),
    "2919": ("Amanda Vierra", "", ""),
    "2920": ("Martina Boyle", "20033216", ""),
    "2921": ("Mustapha Elouarghani", "20004485", ""),
    "2922": ("Dwayne McCalla", "20322607", ""),
    "2923": ("delmy smith", "", ""),
    "2924": ("Darren Kirchman", "20061719", ""),
    "2925": ("James Pridgen", "20014442", ""),
    "2926": ("Renzo Escobedo", "20005793", ""),
    "2927": ("Elizabeth Anne Digey", "", ""),
    "2928": ("Monte Thornton", "20005743", ""),
    "2929": ("Brandon  Caruana", "20072516", ""),
    "2930": ("ANITA LAPIERRE", "20032291", ""),
    "2931": ("Ernest Jehangir", "", ""),
    "2932": ("Richard Strommer", "20003597", ""),
    "2933": ("Bernadette Beckles", "20005809", ""),
    "2934": ("Jayde Johnson", "21620837", ""),
    "2935": ("Tommy Ebanks", "", ""),
    "2936": ("Elisa Giovannini", "20003591", ""),
    "2937": ("dean morales", "", ""),
    "2938": ("Joannah  Small", "", ""),
    "2939": ("Fiona Theaker", "20072525", ""),
    "2940": ("Kacia Harding", "", ""),
    "2941": ("Deborah Kerr-White", "21615178", ""),
    "2942": ("Kiralee Harnett", "", ""),
    "2943": ("Justin Connolly", "", ""),
    "2944": ("Brad Marquardt", "20033207", ""),
    "2945": ("LUDWICK  BERRY", "20074122", ""),
    "2946": ("Cheryl McAllister", "", ""),
    "2947": ("Carol Britton", "20003614", ""),
    "2948": ("Jody Powery-Gilbert", "20033536", ""),
    "2949": ("Jason Taylor", "", ""),
    "2950": ("Fiona De Serpa Pimentel", "", ""),
    "2951": ("Daniel  Pride", "", ""),
    "2953": ("Mayurie  Perera", "20683352", ""),
    "2954": ("Leo Comerton", "20004771", ""),
    "2955": ("jermaine shims", "", ""),
    "2956": ("Kenneth Thomas", "20018237", ""),
    "2957": ("Edward Harper", "20480633", ""),
    "2958": ("Ramone  Richardson", "", ""),
    "2959": ("Sophia Scott", "", ""),
    "2960": ("Danis Ariel Lemus", "20033058", ""),
    "2961": ("Jule Limoli", "20073671", ""),
    "2962": ("Alicia Dixon", "", ""),
    "2963": ("Robert Meyer", "20048791", ""),
    "2964": ("Victoria Merren", "", ""),
    "2965": ("alec cox", "", ""),
    "2966": ("Tishay  Heath", "", ""),
    "2967": ("James McGinn", "20034698", ""),
    "2968": ("EMIL KALINOWSKI", "20004864", ""),
    "2969": ("Adrienne Politowicz", "20030972", ""),
    "2970": ("Daphnie Frederick", "20005722", ""),
    "2971": ("Grace Boos", "", ""),
    "2972": ("mikol watler", "21635745", ""),
    "2973": ("Hannah Reid", "20030553", ""),
    "2974": ("Dawn Mclean-Bradly", "20004889", ""),
    "2975": ("Max Pairaudeau", "", ""),
    "2976": ("rachael fowler", "20010131", ""),
    "2977": ("Alexandra Rodgers", "20072064", ""),
    "2978": ("Scott Elliott", "", ""),
    "2979": ("Sharnelle Silburn", "", ""),
    "2980": ("Imilsy Carbo", "20032284", ""),
    "2981": ("Letecia Bodden", "20003156", ""),
    "2982": ("Alicia Dunbar", "", ""),
    "2983": ("John Lawless", "21164702", ""),
    "2984": ("Carien Harcombe", "", ""),
    "2985": ("Paul  Williams", "20080354", ""),
    "2986": ("Joyce Darlene Groskreutz-Forbes", "", ""),
    "2987": ("Robert Stuke", "20004876", ""),
    "2988": ("Lena Ebanks", "20004774", ""),
    "2989": ("Alicia Parry", "21616753", ""),
    "2990": ("Oliver Goodwin", "", ""),
    "2991": ("tammy nixon", "", ""),
    "2992": ("Louise Sargison", "", ""),
    "2993": ("Kalesha Edwards", "20005756", ""),
    "2994": ("Neriah Leblanc", "20007454", ""),
    "2995": ("Melissa Gordon", "20121449", ""),
    "2996": ("Devinda Nanayakkara", "", ""),
    "2997": ("PROFESSIONAL YACHT MANAGEMENT", "20073063", ""),
    "2998": ("Charles Bush", "20072062", ""),
    "2999": ("Antony McFarlane", "20006301", ""),
    "3000": ("Karyn Singfield", "", ""),
    "3001": ("Destinie Watler", "21597929", ""),
    "3002": ("Andrew  Rollins", "20032200", ""),
    "3003": ("MOHAMED MAFAS ABDUL RAHUMAN LEBBE", "", ""),
    "3004": ("Alicia Proud-Rabess", "20087390", ""),
    "3005": ("Belkis Soler", "20612694", ""),
    "3006": ("ALEKA BECKFORD", "", ""),
    "3007": ("Gladis Sandra  Green", "", ""),
    "3008": ("Marlon Lorde", "", ""),
    "3009": ("Paul  Henry", "", ""),
    "3010": ("Alexander Stoll", "20010127", ""),
    "3011": ("Ardil Salem", "", ""),
    "3012": ("CHRISTOHPER Piddock", "", ""),
    "3013": ("Jonathan Cohen", "20012056", ""),
    "3014": ("Claudia Coke", "20010221", ""),
    "3015": ("Dinah Ebanks", "", ""),
    "3016": ("Elsi Amor", "", ""),
    "3017": ("paulminto  miller", "20004983", ""),
    "3018": ("Simon Burton", "20034048", ""),
    "3019": ("Guglielmo Battaglia", "21628573", ""),
    "3020": ("Fergus Dignan", "", ""),
    "3021": ("Paul-Martin Seguin", "20080358", ""),
    "3022": ("dayana  brant bryan", "", ""),
    "3023": ("Alexander Henry", "", ""),
    "3024": ("Rodney Verma", "", ""),
    "3025": ("Joshua Williams", "", ""),
    "3026": ("Nicolette Steenbhom", "20009781", ""),
    "3027": ("Shanta Gooden", "20005745", ""),
    "3028": ("Daymon Pardue", "20008348", ""),
    "3029": ("anisha creary", "20082124", ""),
    "3030": ("Michele Yura", "20005803", ""),
    "3031": ("Stefan", "", ""),
    "3032": ("Stefan  Prior", "", ""),
    "3033": ("Joshua Preuss", "20011954", ""),
    "3034": ("Gina Conolly", "", ""),
    "3035": ("Alex  Richardson", "", ""),
    "3036": ("Malgorzata Gralak", "20215393", ""),
    "3037": ("Edlyn Ruiz", "20003212", ""),
    "3038": ("SAMANTHA BENNETT", "20003572", ""),
    "3039": ("Estrella Powery", "", ""),
    "3040": ("Remo Dalimonte", "", ""),
    "3041": ("George Williams", "", ""),
    "3042": ("Aoife Brophy", "", ""),
    "3043": ("Robert Lynch", "", ""),
    "3044": ("Elizabeth Buratti Clifton", "", ""),
    "3045": ("Andrew doussept", "20005721", ""),
    "3046": ("Lachlan Hewitt", "", ""),
    "3047": ("Brian Bodden", "20035695", ""),
    "3048": ("Christopher McLaughlin", "20011969", ""),
    "3049": ("Ameka McGowan", "20003617", ""),
    "3050": ("Bryan Lo", "", ""),
    "3051": ("Robert Sleutz", "20005761", ""),
    "3052": ("Roberto Ebanks", "20010223", ""),
    "3053": ("Michael Peck", "20074312", ""),
    "3054": ("Shanda Johnson", "21666648", ""),
    "3055": ("Edy Garcia", "20004846", ""),
    "3056": ("Judicial  Administration", "", ""),
    "3057": ("Robert Moseley", "", ""),
    "3058": ("Christopher Kelly", "", ""),
    "3059": ("Fawne Taylor", "20210931", ""),
    "3060": ("Jacqueline Bentley", "20005654", ""),
    "3061": ("Vasilena Lacheva", "", ""),
    "3062": ("Christina Cavazos", "20004847", ""),
    "3063": ("renell  benjamin", "", ""),
    "3064": ("Sandra Myles", "", ""),
    "3065": ("John Bruton", "20032997", ""),
    "3066": ("rachel lonergan", "20004762", ""),
    "3067": ("Felix Manzanares", "20073674", ""),
    "3068": ("Sary Dixon", "20217188", ""),
    "3069": ("Marilou Cariazo", "", ""),
    "3070": ("Prescilian Rivers", "20191903", ""),
    "3071": ("Ever Cuevas", "", ""),
    "3072": ("Terryann Pretlove", "", ""),
    "3073": ("Daniel Varszegi", "", ""),
    "3074": ("Eloy  Escalante", "20003209", ""),
    "3075": ("Tracey Elliott", "20031038", ""),
    "3076": ("Shalisa Barnett", "20003589", ""),
    "3077": ("Audiophile Group", "20033068", ""),
    "3078": ("HALO Industries, LLC", "", ""),
    "3079": ("Jasmine Brown", "20006296", ""),
    "3080": ("Paul Llewellyn", "", ""),
    "3081": ("Cherie Grad", "20005706", ""),
    "3082": ("Maria Pineda", "20212035", ""),
    "3083": ("Bal Watler", "21648281", ""),
    "3084": ("Nancy Oram", "", ""),
    "3085": ("Christopher Radford", "", ""),
    "3086": ("Simon Morgan-Dyer", "20009784", ""),
    "3087": ("Leilani Hammock", "", ""),
    "3088": ("Georgette Bakhit", "20030574", ""),
    "3089": ("Kristina Grant", "20951212", ""),
    "3090": ("Angelika Bodden", "", ""),
    "3091": ("Melissa  Rios nelson", "20425842", ""),
    "3092": ("Terez Rivers", "20003575", ""),
    "3093": ("Bertrand Stern", "20750976", ""),
    "3094": ("Jarvis Thompson - McLean", "", ""),
    "3095": ("Anthony Walton", "20032293", ""),
    "3096": ("Karl-Phillip Schlossstein", "", ""),
    "3097": ("Krishna Arch-Dilbert", "20084213", ""),
    "3098": ("Liam Hardie", "21657657", ""),
    "3099": ("Robert Tatum", "", ""),
    "3100": ("Catherine Ouriach", "20003139", ""),
    "3101": ("Andrea Laguna", "", ""),
    "3102": ("Troy Peliwan", "", ""),
    "3103": ("Daisy Lee", "20009995", ""),
    "3104": ("Lauren Kelly", "20003221", ""),
    "3105": ("Jasen Campbell", "20191698", ""),
    "3106": ("Adam Loynton", "20033142", ""),
    "3107": ("Pamela Thompson", "", ""),
    "3108": ("Nicole Makin", "", ""),
    "3109": ("Takiyah Smith", "", ""),
    "3110": ("Musa Machembe", "20505181", ""),
    "3111": ("Kelly Fisher", "", ""),
    "3112": ("Jeremy Scantlebury", "", ""),
    "3113": ("stephanie edwards", "20004890", ""),
    "3114": ("Kimberly Valliere", "", ""),
    "3115": ("Robert Tocchio", "", ""),
    "3116": ("Angie Yee", "20677814", ""),
    "3117": ("kadie morris", "", ""),
    "3118": ("Chloe Scott", "20003552", ""),
    "3119": ("Rudolf Hoehler", "20004875", ""),
    "3120": ("Taleke Moore", "", ""),
    "3121": ("Amadello Mena-Hebbert", "20219321", ""),
    "3122": ("Jesse Livingston", "20005772", ""),
    "3123": ("Jonathan Welds", "20032282", ""),
    "3124": ("Christopher  Winschel", "20005798", ""),
    "3125": ("Guy Healey", "", ""),
    "3126": ("Robert Anthony Mordeno", "", ""),
    "3127": ("J. Aliya Dunstan", "20009782", ""),
    "3128": ("Francesca Trees", "20005795", ""),
    "3129": ("Veko Ramgeet", "", ""),
    "3130": ("Yvonne Blanco", "", ""),
    "3131": ("Alhana Hurlston", "20005796", ""),
    "3132": ("Sophie Ginder", "20002970", ""),
    "3133": ("Julien Breton", "20004786", ""),
    "3134": ("Zulema Bush-Ramos", "20204545", ""),
    "3135": ("Shanda Johnson", "20033538", ""),
    "3136": ("Mikaiyah Thomas", "20005699", ""),
    "3137": ("Sharnelle Miller", "", ""),
    "3138": ("Anna Madden", "", ""),
    "3139": ("Andrew Anniford", "20030372", ""),
    "3140": ("Ilona Groark", "", ""),
    "3141": ("Juana  Amador", "20004788", ""),
    "3142": ("Rihards Brinkis", "20002076", ""),
    "3143": ("Richard Perdido", "20751018", ""),
    "3144": ("Giles Fletcher", "20003587", ""),
    "3145": ("Kirk Collins", "20003211", ""),
    "3146": ("Mikhail Campbell", "20030981", ""),
    "3147": ("Lisa Tatosian", "20001581", ""),
    "3148": ("Rebecca Biancardi", "21686965", ""),
    "3149": ("Colin WIlson", "", ""),
    "3150": ("Julie Barden", "", ""),
    "3151": ("Nari Ramdon", "20003163", ""),
    "3152": ("Cheryl Reid-Al Masri", "", ""),
    "3153": ("Jeanne Meaney-Uhl", "20003164", ""),
    "3154": ("Rachel Clark", "20208808", ""),
    "3155": ("Ray Quan", "", ""),
    "3156": ("Watch Impressions", "", ""),
    "3157": ("Anzita Mcfield", "20007216", ""),
    "3158": ("Reeva McLaughlin", "", ""),
    "3159": ("Kayla Wood", "20003609", ""),
    "3160": ("Peter Letko", "20220029", ""),
    "3161": ("Hannah Welcome", "20003612", ""),
    "3162": ("Keri Andrews", "20003214", ""),
    "3163": ("Carlos Sierra", "20003616", ""),
    "3164": ("Eric Vincent", "", ""),
    "3165": ("Karla Reyna", "20003216", ""),
    "3166": ("Kayla Bush", "", ""),
    "3167": ("Tricia McDoom", "20555293", ""),
    "3168": ("Deepu  Fernandez", "26104421", ""),
    "3169": ("Derrie Boggess", "20005794", ""),
    "3170": ("Natasha Marius", "20005792", ""),
    "3171": ("Marcello Piacentini", "20004848", ""),
    "3172": ("Sarah Pennacchini", "20076779", ""),
    "3173": ("Kristi Green", "20005763", ""),
    "3174": ("JASON EBANKS", "20003140", ""),
    "3175": ("davion cotterell", "20034830", ""),
    "3176": ("Louise Craig", "", ""),
    "3177": ("Rupesh Daya", "20009984", ""),
    "3178": ("Dwene Ebanks", "21777897", ""),
    "3179": ("Malcolm Hurlston", "20038427", ""),
    "3180": ("Mark Nicoll", "", ""),
    "3181": ("Trinda Blackmore", "21611432", ""),
    "3182": ("Caitlin Connor", "", ""),
    "3183": ("Cristina Derkach", "20007228", ""),
    "3184": ("Pradeep Kumar Betha", "20004763", ""),
    "3185": ("Kenroy  Sutherland", "20004766", ""),
    "3186": ("karen neal", "", ""),
    "3187": ("Mark Jones", "", ""),
    "3188": ("Daniel Cowan", "", ""),
    "3189": ("Laura McLaughlin", "20005760", ""),
    "3190": ("Jennifer Ebanks", "", ""),
    "3191": ("Priscilla Pouchie", "20010816", ""),
    "3192": ("Mario Iwazaki", "20004867", ""),
    "3193": ("Bessanio Dilbert", "", ""),
    "3194": ("Mac Imrie", "20009996", ""),
    "3195": ("Craig Ross", "20012054", ""),
    "3196": ("Jill D_x0019_ Aloisio", "20572304", ""),
    "3197": ("Yosha Alphonse", "20030997", ""),
    "3198": ("Marlies van Sloten", "20004859", ""),
    "3199": ("Robert Watler", "20220928", ""),
    "3200": ("Jordin Barnes-Tabora", "20109604", ""),
    "3201": ("Diana Virtue", "", ""),
    "3202": ("MICHAEL MILLER", "20016126", ""),
    "3203": ("Gary Hatswell", "20009990", ""),
    "3204": ("Devon Gow", "", ""),
    "3205": ("Jennifer  Dyer", "21657010", ""),
    "3206": ("Kerri-Lee Saunders", "", ""),
    "3207": ("Catherine Tick", "20031011", ""),
    "3208": ("Vincent Ebanks", "", ""),
    "3209": ("Jonathan Ebanks", "", ""),
    "3210": ("Mladen Kojic", "", ""),
    "3211": ("Craig Lamb", "20075130", ""),
    "3212": ("Dina  Bodden", "20008268", ""),
    "3213": ("Carli Ferreira", "20037522", ""),
    "3214": ("Samantha  McHayle", "20007370", ""),
    "3215": ("Ashley Smith-Phipps", "20073019", ""),
    "3216": ("Abayomi Sodimu", "20072499", ""),
    "3217": ("Winston Hunter", "", ""),
    "3218": ("Vincent Budesa", "20007359", ""),
    "3219": ("Dino Hydes", "20006581", ""),
    "3220": ("Juan Leiva - Vargas", "20055831", ""),
    "3221": ("Kathleen Spencer", "20482740", ""),
    "3222": ("Kayla Parsons", "20011958", ""),
    "3223": ("Natasha Playne", "20681185", ""),
    "3224": ("Godson Shaw", "", ""),
    "3225": ("Julia Allman", "20007155", ""),
    "3226": ("Ondine Bult", "", ""),
    "3227": ("Janeka  Ebanks", "20030362", ""),
    "3228": ("Tara Smith", "20009985", ""),
    "3229": ("Timisha Edwards", "20073034", ""),
    "3230": ("Grecia Iuculano", "", ""),
    "3231": ("Daegan McLaughlin", "20217189", ""),
    "3232": ("Peggy Campney", "20010219", ""),
    "3233": ("Tianah Cunningham", "", ""),
    "3234": ("Garfield Gordon", "", ""),
    "3235": ("Jodyann  Brown", "", ""),
    "3236": ("Stephen Leontsinis", "20754852", ""),
    "3237": ("Felicia Ramsay", "", ""),
    "3238": ("Damari Parker", "20084208", ""),
    "3239": ("Marina Caruntu", "20009983", ""),
    "3240": ("Selena Scott", "20031322", ""),
    "3241": ("salomon.h maria", "", ""),
    "3242": ("Angeline Murillo", "20429413", ""),
    "3243": ("Hamish Caithness", "20005161", ""),
    "3244": ("Sherri Fleming", "", ""),
    "3245": ("Test test", "", ""),
    "3246": ("CHRISTI EBANKS", "", ""),
    "3247": ("Jason Deamer", "", ""),
    "3248": ("Jerlane Ricketts", "20030575", ""),
    "3249": ("Niall Hanna", "20010125", ""),
    "3250": ("Kristina  Deckelman", "", ""),
    "3251": ("Michael  Joseph", "", ""),
    "3252": ("Jamel Johnson", "", ""),
    "3253": ("Jennifer Fox", "", ""),
    "3254": ("Tiffany Jeremiah", "20220906", ""),
    "3255": ("Jaslyne Bridges", "20036183", ""),
    "3256": ("Trust Munyuki", "20011955", ""),
    "3257": ("Leonardi Caraballo", "20009978", ""),
    "3258": ("Byron Clarke", "", ""),
    "3259": ("Kerry-Ann Stewart", "20076775", ""),
    "3260": ("Oliver Collins", "", ""),
    "3261": ("Daniel  Dixon", "20009980", ""),
    "3262": ("amanda hurlston", "21742843", ""),
    "3263": ("Shanee Mayorquin", "20708545", ""),
    "3264": ("Jason Hydes", "20011951", ""),
    "3265": ("Emma Carroll", "", ""),
    "3266": ("Stacey Ford", "20010003", ""),
    "3267": ("Alan Buchanan", "", ""),
    "3268": ("Friedrich Sajovitz", "20010281", ""),
    "3269": ("Adam Stang", "", ""),
    "3270": ("Jose Raul Cruz", "20010279", ""),
    "3271": ("Romayne Hylton", "20009989", ""),
    "3272": ("Brianna Wilkerson", "", ""),
    "3273": ("Greg Errol Fidelis Bisquera", "20010274", ""),
    "3274": ("Andrea  Rosewicz", "", ""),
    "3275": ("Anfernee Wright", "20034046", ""),
    "3276": ("Elizabeth  Kenny", "20206484", ""),
    "3277": ("Brittanni Seymour", "20001691", ""),
    "3278": ("James Lydeard", "", ""),
    "3279": ("Shana Chin", "", ""),
    "3280": ("Andrew Mclaughlin", "20667548", ""),
    "3281": ("Smarter Living Cayman Ltd", "20033145", ""),
    "3282": ("Spencer Vickers", "20011797", ""),
    "3283": ("Christine Sage", "20012010", ""),
    "3284": ("Schuyler Young", "20011795", ""),
    "3286": ("Simeon Dandie", "20015483", ""),
    "3287": ("trecika vernon", "", ""),
    "3288": ("Thomas Lowe", "", ""),
    "3289": ("Merlyn Brown", "20030999", ""),
    "3290": ("Vicki  Huckstep", "", ""),
    "3291": ("Shannon McDonald", "", ""),
    "3292": ("Chris Mills", "", ""),
    "3293": ("Deavour Rose", "21634286", ""),
    "3294": ("Colin Crumpton", "20033026", ""),
    "3295": ("Caroline Mills", "", ""),
    "3296": ("Michael Cao", "", ""),
    "3297": ("Byron  Clarke", "20011791", ""),
    "3298": ("Leo Fankhaenel", "", ""),
    "3299": ("Tania Parchment", "", ""),
    "3300": ("Daniel Bishop", "20030552", ""),
    "3301": ("Shari Smith", "20030559", ""),
    "3302": ("Joanne Meakin", "20011793", ""),
    "3303": ("Rachel Schneider", "20011794", ""),
    "3304": ("Robert Chilman", "", ""),
    "3305": ("Sarai Soto", "20012011", ""),
    "3306": ("laura thomas", "", ""),
    "3307": ("Nicolas Ramos Rego", "", ""),
    "3308": ("Alex Howard", "20033205", ""),
    "3309": ("Parker Ebanks", "", ""),
    "3310": ("Pamela Hazelwood", "20061726", ""),
    "3311": ("Trisha Johnson", "20011798", ""),
    "3312": ("Romelyn Lucero", "20034044", ""),
    "3313": ("Charnissa  Richardson", "20220028", ""),
    "3314": ("Bryan Cascante", "", ""),
    "3315": ("Shanen  O\\'Leary", "20220027", ""),
    "3316": ("Catherine  Bodden", "", ""),
    "3317": ("Dayana Arroliga", "", ""),
    "3318": ("Solenn Carriou", "20011997", ""),
    "3319": ("Daniel Lowe", "20016120", ""),
    "3320": ("Stephen Small", "20016118", ""),
    "3321": ("Mickal Solomon", "", ""),
    "3322": ("Robert Morrison", "21620012", ""),
    "3323": ("Dave Ehnes", "20012458", ""),
    "3324": ("Vanetta Barclay", "20017684", ""),
    "3325": ("George Palmer", "20029049", ""),
    "3326": ("Burmon Scott", "", ""),
    "3327": ("Alan Whyte", "20213047", ""),
    "3328": ("Carissa Liro-Hudson", "20032278", ""),
    "3329": ("TESSA EBANKS", "", ""),
    "3331": ("Aleksandr Kuznecov", "20009324", ""),
    "3332": ("Nicholas Cann", "", ""),
    "3333": ("Jenai Bellafonte", "", ""),
    "3334": ("Margaret Mendes", "", ""),
    "3335": ("Kimbert y Angelica  Solomon", "20086272", ""),
    "3336": ("Kieran  Donovan", "20072073", ""),
    "3337": ("Marta Wile", "", ""),
    "3338": ("Kayleigh  McLaughlin", "20017017", ""),
    "3339": ("SUZZETTE STEWART", "20034244", ""),
    "3340": ("Lionel Djondo", "20011761", ""),
    "3341": ("Troy Leacock", "20033192", ""),
    "3342": ("Bethsaida Lopez Herrera", "", ""),
    "3343": ("Davina Tresidder", "20016121", ""),
    "3344": ("Jacqueline  Morris", "20030580", ""),
    "3345": ("Caston Powery", "20016092", ""),
    "3346": ("CODY MAHAFFEY", "", ""),
    "3347": ("Kendra y Julian Morris", "20017686", ""),
    "3348": ("Joanna Mcniven", "20033072", ""),
    "3349": ("Willem Boshoff", "", ""),
    "3350": ("Leland Short", "20076767", ""),
    "3351": ("ANNA WATLER", "", ""),
    "3352": ("Estefany Andreina Panqueva Martinez", "20033132", ""),
    "3353": ("Ohare Brown", "20033197", ""),
    "3354": ("Loreta Ballart Acosta", "20017688", ""),
    "3355": ("Rene  DelaHaye", "20018238", ""),
    "3356": ("Nadine Cloutier", "20016514", ""),
    "3357": ("Eadaoin McArthur", "", ""),
    "3358": ("Amber Russell", "20027759", ""),
    "3359": ("Michael  Kellyman", "20033053", ""),
    "3360": ("Carl Hey", "", ""),
    "3361": ("KIRK STEELE", "20030554", ""),
    "3362": ("Anthony Smellie", "20440266", ""),
    "3363": ("larry Vaceannie", "", ""),
    "3364": ("Edilberto Beniquez", "20030551", ""),
    "3365": ("Jayson Atanacio", "21727498", ""),
    "3366": ("Leah Bootsma", "", ""),
    "3367": ("Priscilla Dickson", "", ""),
    "3368": ("Niall Simpson", "20061731", ""),
    "3369": ("Richard Ebanks", "", ""),
    "3370": ("Richard Ebanks", "20025758", ""),
    "3371": ("Rose Ritch", "20031033", ""),
    "3372": ("Cayman Dental  Services", "20035203", ""),
    "3373": ("David McGibbon", "20033147", ""),
    "3374": ("Caelan  McLaughlin", "20109612", ""),
    "3375": ("Mandy Singh", "21706575", ""),
    "3376": ("Zackary Thomas", "", ""),
    "3377": ("Tonishea  Heslop-Fraser", "", ""),
    "3378": ("Mansi Chitalia Patel", "", ""),
    "3379": ("Sinitta Mc Lean", "20030320", ""),
    "3380": ("Jasmine Williams", "20030363", ""),
    "3381": ("THOMAS MORAHAN", "20029281", ""),
    "3382": ("Pamela Booth", "20030974", ""),
    "3383": ("Heather Halsey", "", ""),
    "3384": ("Tanya Thompson", "20030619", ""),
    "3385": ("Lea Scott", "", ""),
    "3386": ("Danielle Connolly", "", ""),
    "3387": ("Kirstie  Johnson", "", ""),
    "3388": ("Emily O'Keeffe", "21634217", ""),
    "3389": ("Geovanna Dominguez", "", ""),
    "3390": ("Amalia Herrera", "20036178", ""),
    "3391": ("DAYAN CARDENAS", "20033027", ""),
    "3392": ("Richard Fear", "20033148", ""),
    "3393": ("betina farkas", "20030615", ""),
    "3394": ("Audrey Wellington", "20027963", ""),
    "3395": ("simon ashdown", "20033031", ""),
    "3396": ("Adam Baker", "21632642", ""),
    "3397": ("Peter Ranger", "20030593", ""),
    "3398": ("Ryan Walrond", "20002217", ""),
    "3399": ("Charles Jackson", "20034043", ""),
    "3400": ("Mark Whaley", "", ""),
    "3401": ("Nathaniel Luker", "20030381", ""),
    "3402": ("Catherine Day", "20030985", ""),
    "3403": ("Felicia Schvartz", "20072063", ""),
    "3404": ("Nazar Tsitsyala", "", ""),
    "3405": ("Jorge Lopez", "20016098", ""),
    "3406": ("Lotoya Smith", "20034042", ""),
    "3407": ("Brian Braggs", "20061748", ""),
    "3408": ("Daniel Herrmann", "20035685", ""),
    "3409": ("Wilmoth Shillingford", "20030373", ""),
    "3410": ("Ben Knudson", "", ""),
    "3411": ("Marika  Haynes", "20031029", ""),
    "3412": ("Janine Nyyssonen", "20035688", ""),
    "3413": ("Scott  MacLaren", "20030565", ""),
    "3414": ("evert brunekreef", "20030379", ""),
    "3415": ("Tejan Massally", "", ""),
    "3416": ("Annissa Sheow", "20033126", ""),
    "3417": ("Deborah Mitchell", "20031018", ""),
    "3418": ("AIja Ebanks", "20030510", ""),
    "3419": ("Christopher Olson", "20031031", ""),
    "3420": ("Holly Caird", "", ""),
    "3421": ("Donna Kay Stoddart", "20037066", ""),
    "3422": ("blake ducharme", "20109594", ""),
    "3423": ("HILDA MARCH", "20062883", ""),
    "3424": ("Justin Miller", "21646103", ""),
    "3425": ("Sarah Esplin- Jones", "20031044", ""),
    "3426": ("Stephanie Peterson", "20030990", ""),
    "3427": ("Felisiana Ebanks", "20001296", ""),
    "3428": ("Mark Tarsh", "20031035", ""),
    "3429": ("Lisa Kelly", "20075132", ""),
    "3430": ("James Grandage", "20032204", ""),
    "3431": ("Deborah  Wray", "20033211", ""),
    "3432": ("Bounito Levy", "20074137", ""),
    "3433": ("Duane Tibbetts", "", ""),
    "3434": ("KRISHNA MANI", "", ""),
    "3435": ("Jessica Redhead", "20033069", ""),
    "3436": ("Jennifer Butler", "20033209", ""),
    "3437": ("Andrew Tomlinson", "20083720", ""),
    "3438": ("John Jacob", "", ""),
    "3439": ("Kirk Douglas", "21651677", ""),
    "3440": ("Tyleisha Galbraith", "20074297", ""),
    "3441": ("Chantelle Day", "20217487", ""),
    "3442": ("Keith  Johnston", "", ""),
    "3443": ("Melanie Bokelman", "20033286", ""),
    "3444": ("Samuel Wiss", "", ""),
    "3445": ("Imelda Lindsay", "", ""),
    "3446": ("John Garret Jamito", "", ""),
    "3447": ("Michelle Hydes", "", ""),
    "3448": ("Camilla Testa", "", ""),
    "3449": ("Robyn  McCoy", "20035696", ""),
    "3450": ("Kerry Whittaker", "", ""),
    "3451": ("Michelle  Veldhoven", "", ""),
    "3452": ("Graeme Brown", "20037073", ""),
    "3453": ("Luis  Guillen", "20750928", ""),
    "3454": ("CHASE GREEN", "", ""),
    "3455": ("Chaturika Fonseka", "20632977", ""),
    "3456": ("Tanya Newlove", "", ""),
    "3457": ("Christina Trumbach", "21640110", ""),
    "3458": ("Richard Christian", "20037521", ""),
    "3459": ("Benjamin  Tonner", "", ""),
    "3460": ("Nan Erb", "", ""),
    "3461": ("ANGELEE BEERSINGH", "21666496", ""),
    "3462": ("Fana Harper Tyson", "20037069", ""),
    "3463": ("Marco Du Toit", "", ""),
    "3464": ("Ashley Wood", "", ""),
    "3465": ("Sabrina Foster", "20038144", ""),
    "3466": ("Etienne Jensen-Fontaine", "", ""),
    "3467": ("Bethany Ebanks", "20215704", ""),
    "3468": ("Kilian Werner", "20109588", ""),
    "3469": ("Miles Ruby", "20061915", ""),
    "3470": ("patricia  kohler", "20037520", ""),
    "3471": ("Madalyn Tavares", "20220494", ""),
    "3472": ("Emily Farren", "", ""),
    "3473": ("Jamaal Brown", "20041720", ""),
    "3474": ("Lauren Dombowsky", "20076781", ""),
    "3475": ("Julene Witter", "20073030", ""),
    "3476": ("Radhames Polanco", "21636015", ""),
    "3477": ("Kenroy  Henderson", "20425877", ""),
    "3478": ("Mark Drummond", "20047741", ""),
    "3479": ("Araceli Garcia", "20220486", ""),
    "3480": ("Alexia Walton", "20008496", ""),
    "3481": ("Kelsey Rae-Smith", "20037539", ""),
    "3482": ("Jonathan Williams", "20221177", ""),
    "3483": ("Camila Timoni", "20211332", ""),
    "3484": ("Bianca Tica", "20080362", ""),
    "3485": ("Brody Thomas", "", ""),
    "3486": ("Nigel May", "21677708", ""),
    "3487": ("Siddhesh Namugade", "", ""),
    "3488": ("Kenton Toews", "20038143", ""),
    "3489": ("Sara Fernandez", "", ""),
    "3490": ("Carolyn Tibbetts", "20061485", ""),
    "3491": ("LEIF EDWARDS-BEST", "", ""),
    "3492": ("Richard Bodden", "20075138", ""),
    "3493": ("Dawinskin Batista Javier", "20217492", ""),
    "3494": ("Colin Nestor", "20453555", ""),
    "3495": ("neal lomax", "20040703", ""),
    "3496": ("Benji Asquith", "20074301", ""),
    "3497": ("Thomas McPhee", "20081544", ""),
    "3498": ("Reynaldo Powery", "20047753", ""),
    "3499": ("Tashla Aimable", "20083051", ""),
    "3500": ("CATHY WALKER", "", ""),
    "3501": ("Victoria Ramos", "20109591", ""),
    "3502": ("DENSIE RANKIN", "20041719", ""),
    "3503": ("Danie Hiron", "20109626", ""),
    "3504": ("Brandon Caruana", "20072516", ""),
    "3505": ("PETAGAYE DILLION", "20044943", ""),
    "3506": ("Pauline Vayssiere", "20425818", ""),
    "3507": ("laurence DALLASERRA", "", ""),
    "3508": ("Tifiany Rose", "", ""),
    "3509": ("Kristina Gliguroska", "20567319", ""),
    "3510": ("Shantannia Bryan", "20061718", ""),
    "3511": ("Kioko Muasya", "20220485", ""),
    "3512": ("Shian O'Connor", "20062880", ""),
    "3513": ("Rosa Dacosta", "20045292", ""),
    "3514": ("Rosa Dacosta", "20045292", ""),
    "3515": ("Deshae Terry", "20073056", ""),
    "3516": ("Alexandra Maharaj", "20042660", ""),
    "3517": ("Felicia Connor", "20191731", ""),
    "3518": ("Adam Rossiter", "20044935", ""),
    "3519": ("Evelyn Lunn", "20550003", ""),
    "3520": ("MISHAEL MAY ANGLO", "20049121", ""),
    "3521": ("Aleisha Lalor", "", ""),
    "3522": ("Nadia Balleram", "20100089", ""),
    "3523": ("Vickie Clarke", "20047742", ""),
    "3524": ("Paula Ebanks", "20049125", ""),
    "3525": ("David McIntyre", "20072512", ""),
    "3526": ("Jennifer McConville", "20047744", ""),
    "3527": ("Sebastian Mottram", "20213044", ""),
    "3528": ("Lukas Schroeter", "", ""),
    "3529": ("Kimberly Glasgow", "20045291", ""),
    "3530": ("Andrew Woodcock", "20100083", ""),
    "3531": ("Titus Ebanks", "20109620", ""),
    "3532": ("Daniel Nguyen", "", ""),
    "3533": ("Cara Leeland", "", ""),
    "3534": ("Remi Tetot", "", ""),
    "3535": ("Adam Stang", "", ""),
    "3536": ("Maylyn Phillips", "21620020", ""),
    "3537": ("glenn anderson", "", ""),
    "3538": ("Kami Butcher", "20078626", ""),
    "3539": ("Anais Tatum", "20050509", ""),
    "3540": ("Robert Myers", "20046747", ""),
    "3541": ("Emily Murphy", "20047435", ""),
    "3542": ("Nancy McLeod", "", ""),
    "3543": ("Serge Beaudet", "20084210", ""),
    "3544": ("Rose-Marie Watt", "20085236", ""),
    "3545": ("Bradley de Schiffert", "20046745", ""),
    "3546": ("Sofia Samaniego", "20047379", ""),
    "3547": ("Kathleen  Lambert", "20047746", ""),
    "3548": ("Ashley Ebanks - Hoybia", "20061735", ""),
    "3549": ("Natasha  Poroosotum", "", ""),
    "3550": ("Rochelle Brooks", "20078619", ""),
    "3551": ("Nadine Jones", "", ""),
    "3552": ("SOPHIA HUNT", "20080994", ""),
    "3553": ("Tyler  Christian", "20072517", ""),
    "3554": ("Kerryann Simpson", "20086842", ""),
    "3555": ("Trifina Scott", "20047740", ""),
    "3556": ("Trifina Scott", "20047740", ""),
    "3557": ("Porshia Jackson", "", ""),
    "3558": ("William Treadwell", "", ""),
    "3559": ("Ian Morgan", "20049123", ""),
    "3560": ("Tania Slijper", "", ""),
    "3561": ("Denise C Reid", "20049127", ""),
    "3562": ("David Lloyd", "20204540", ""),
    "3563": ("Liam Kay", "20049128", ""),
    "3564": ("Yesica  Rodriguez federico", "", ""),
    "3565": ("Kristin Koopman", "", ""),
    "3566": ("INGRID JONES", "", ""),
    "3567": ("Vidya Ravella", "20617215", ""),
    "3568": ("AB Strategies", "21665589", ""),
    "3569": ("Bridget Ashworth", "20050512", ""),
    "3570": ("Lucinda Burgess-Shannon", "20071516", ""),
    "3571": ("Diane Eleanor Musson-McDonald", "20050521", ""),
    "3572": ("Rusty Gipson", "20055829", ""),
    "3573": ("Katrina Gomez", "20050529", ""),
    "3574": ("Julian Foster", "20052604", ""),
    "3575": ("Hazel Francis", "20051111", ""),
    "3576": ("Paola  Canales", "20061914", ""),
    "3577": ("Kirk Fletcher", "20075140", ""),
    "3578": ("NAUDE DREYER", "20052492", ""),
    "3579": ("Dennis Turner", "20052603", ""),
    "3580": ("Humberto  Campo", "", ""),
    "3581": ("Kristin Koopman", "", ""),
    "3582": ("Lacey Manz", "20109624", ""),
    "3583": ("Akranee Inyotha", "20055832", ""),
    "3584": ("Gigi Gaea", "", ""),
    "3585": ("Sue-Ann Bodden", "20073681", ""),
    "3586": ("Julian Coetzer", "20072513", ""),
    "3587": ("Jack McGregor", "", ""),
    "3588": ("Debbriana Linwood", "", ""),
    "3589": ("Kyle Saunders", "20078873", ""),
    "3590": ("Nuria Villacis", "", ""),
    "3591": ("Liesl Richter", "20109613", ""),
    "3592": ("Henry Nichols", "20062882", ""),
    "3593": ("Finley Josephs", "20055828", ""),
    "3594": ("Josche Wagner", "21636998", ""),
    "3595": ("Louise Spencer", "21706296", ""),
    "3596": ("Miranda Bourne", "20092393", ""),
    "3597": ("Javier Elias", "", ""),
    "3598": ("Lauren Downs", "20061473", ""),
    "3599": ("Aysha Jackson", "", ""),
    "3600": ("Ramjeet Johnson", "20554662", ""),
    "3601": ("Christy Robson", "20073683", ""),
    "3602": ("Alanis Linwood", "20061909", ""),
    "3603": ("Caitlyn  Richardson", "", ""),
    "3604": ("Lidka Scott", "20061711", ""),
    "3605": ("Jeffrey Fieler", "20061910", ""),
    "3606": ("Timothy Secord", "", ""),
    "3607": ("Derek Smith", "", ""),
    "3608": ("Disraeli Veliz", "20480642", ""),
    "3609": ("AbuBakar Nyanzi", "20061904", ""),
    "3610": ("Ara Alfonso Valdulla", "20061905", ""),
    "3611": ("Martha Coe", "20218459", ""),
    "3612": ("Robin Lowe", "20072061", ""),
    "3613": ("Romaine Barnes", "", ""),
    "3614": ("Helen Watts", "20061907", ""),
    "3615": ("Kacey Frederick", "20072519", ""),
    "3616": ("PATRICK BEERSINGH", "", ""),
    "3617": ("Antoine Powell", "21630682", ""),
    "3618": ("Jenna A Nixon", "20071431", ""),
    "3619": ("Jennifer  Moseley", "21641326", ""),
    "3620": ("Kimberly Scott", "20062864", ""),
    "3621": ("Mario Smith", "20062865", ""),
    "3622": ("Daniel Pallett", "20062869", ""),
    "3623": ("Thea Ebanks", "20061710", ""),
    "3624": ("DaleLa Solomon", "", ""),
    "3625": ("Ying Li", "20061501", ""),
    "3626": ("linda thompson", "", ""),
    "3627": ("Sabrina  Dennis-Elgueta", "", ""),
    "3628": ("David Self", "20084035", ""),
    "3629": ("Lina Parillon", "", ""),
    "3630": ("KELLI JACKSON", "20059980", ""),
    "3631": ("Jayvant Lakshman", "20061743", ""),
    "3632": ("Stephen Quinland", "20061738", ""),
    "3633": ("Test Test", "", ""),
    "3634": ("Kimberley  Rusyn", "", ""),
    "3635": ("Carleigh  Bell", "20073036", ""),
    "3636": ("Sushanna MccArthy", "", ""),
    "3637": ("Cherylann Hill", "", ""),
    "3638": ("Tomasz Ryk", "20073060", ""),
    "3639": ("Kasey Reid", "20425817", ""),
    "3640": ("Simone Somers Vine", "20080360", ""),
    "3641": ("Nick Freeland", "20109614", ""),
    "3642": ("Melinda Andrews", "20013159", ""),
    "3643": ("Mike McMahan", "20086284", ""),
    "3644": ("Viktoria Acsai", "20086833", ""),
    "3645": ("Wanda Viscount", "20073679", ""),
    "3646": ("Affiong Ate", "", ""),
    "3647": ("Lauren Merren", "20541455", ""),
    "3648": ("Colin Meronuk", "20013589", ""),
    "3649": ("Peter Sadler", "20078849", ""),
    "3650": ("Archibald Brathwaite", "20071426", ""),
    "3651": ("Ever Cuevas", "", ""),
    "3652": ("Bonnie Murugesu", "", ""),
    "3653": ("Rose Fagan", "20204554", ""),
    "3654": ("Marc Sparg", "21634696", ""),
    "3655": ("Gary Meek", "20061461", ""),
    "3656": ("Paul Lorraine", "", ""),
    "3657": ("Kimberly Riley-Cater", "", ""),
    "3658": ("Robert Rees", "", ""),
    "3659": ("Giannie McLaughlin", "20072060", ""),
    "3660": ("Lisa Jarvis", "20071514", ""),
    "3661": ("Shanice Dawes", "", ""),
    "3662": ("Dustin Springett", "20083062", ""),
    "3663": ("Catherine Tathum", "", ""),
    "3664": ("Marco Larrea", "20667511", ""),
    "3665": ("Donna Goodall", "20072066", ""),
    "3666": ("Maxine Bryson", "20071510", ""),
    "3667": ("Sandra Whitaker", "20071427", ""),
    "3668": ("Vikash Sobrun", "", ""),
    "3669": ("Deon Masters", "", ""),
    "3670": ("Paul Smith", "20074117", ""),
    "3671": ("Shanon  Bello de Monzon", "20071429", ""),
    "3672": ("Janell Dyer", "", ""),
    "3673": ("Keisha Stephenson-Reid", "20217486", ""),
    "3674": ("Amaury Cabrera", "", ""),
    "3675": ("Vanessa Ebanks", "20070905", ""),
    "3676": ("Vijayabalan Murugesu", "20073685", ""),
    "3677": ("Chelsea Bodden", "", ""),
    "3678": ("Pablo Mejia", "", ""),
    "3679": ("Cherylann Hill", "20087834", ""),
    "3680": ("Sandra Karlsson", "20058209", ""),
    "3681": ("Kyel West", "20440267", ""),
    "3682": ("Melissa Ferrer", "20072067", ""),
    "3683": ("PHILIP CATER", "", ""),
    "3684": ("Keith Mahan", "20072068", ""),
    "3685": ("Elena garcia", "", ""),
    "3686": ("Paolo Castello", "20070907", ""),
    "3687": ("Benjamin Bodden", "20210890", ""),
    "3688": ("Vanessa Williams", "20076503", ""),
    "3689": ("Grace Ann Mulgrave", "20073023", ""),
    "3690": ("Vanessa Gilman", "21629450", ""),
    "3691": ("Sherdon Kelly", "20541453", ""),
    "3692": ("Javier Hernandez", "20072526", ""),
    "3693": ("Marc Halley", "20006735", ""),
    "3694": ("Tracey Burns", "20072506", ""),
    "3695": ("Rhodian Bodden", "", ""),
    "3696": ("Tracey Jackman", "20072521", ""),
    "3697": ("Richard Pooley", "20204549", ""),
    "3698": ("Jamie Love", "", ""),
    "3699": ("Nilani Perera", "20072081", ""),
    "3700": ("Jia Jun Chua", "", ""),
    "3701": ("Veritas Trustees Limited", "", ""),
    "3702": ("Eric Paassen", "20085232", ""),
    "3703": ("Pieter Olthof", "21635337", ""),
    "3704": ("Wang Xu", "", ""),
    "3705": ("Debra Parsons", "20078872", ""),
    "3706": ("Heather Wendell", "20049800", ""),
    "3707": ("Alexandra Clynes", "20084207", ""),
    "3708": ("Andrew Stam", "20074315", ""),
    "3709": ("Evelyn Swaby", "", ""),
    "3710": ("Michael Pearson", "20074132", ""),
    "3711": ("Fleur O'Driscoll", "20074296", ""),
    "3712": ("Julie Butler", "20075141", ""),
    "3713": ("Veronica  Pizzio", "20075119", ""),
    "3714": ("Veronica  Pizzio", "20075119", ""),
    "3715": ("Candace Dulude", "20078635", ""),
    "3716": ("Lynden John", "20074331", ""),
    "3717": ("Joshua Rice", "20074120", ""),
    "3718": ("Maria Abegail  Monterde", "20213049", ""),
    "3719": ("Beth Johnson", "20078876", ""),
    "3720": ("Josie Frazer", "20060620", ""),
    "3721": ("Alisha Jackson-Forbes", "20191604", ""),
    "3722": ("Tish Welch", "20215804", ""),
    "3723": ("Calvin Myrie", "20087833", ""),
    "3724": ("Marcella Wright-Aguirre", "20075145", ""),
    "3725": ("KENDRA ASHTON", "20324131", ""),
    "3726": ("Quincy Scott", "21672162", ""),
    "3727": ("Jordan Scott", "21631671", ""),
    "3728": ("Sherianne Gajadhar", "20083049", ""),
    "3729": ("Jordan Tranel", "20081543", ""),
    "3730": ("Margie Barnes", "20078636", ""),
    "3731": ("J Michelle Couraud", "20078858", ""),
    "3732": ("Cassandra Van Rooyen", "20076768", ""),
    "3733": ("Rose Estefan Mameng", "20109590", ""),
    "3734": ("Haniya Soomro", "20085231", ""),
    "3735": ("Rustom Mameng", "20206486", ""),
    "3736": ("Keisha Ann  Ensong", "20100112", ""),
    "3737": ("Elizabeth Muschamp", "20078638", ""),
    "3738": ("Sarah Farquhar", "20109606", ""),
    "3739": ("Gail Graham", "20078642", ""),
    "3740": ("Anne Fleming", "20667530", ""),
    "3741": ("Aimee McKie", "20078643", ""),
    "3742": ("Vicdarwis  Suarez", "", ""),
    "3743": ("Sharis Ford", "20191664", ""),
    "3744": ("Emmary Ruiz", "20082117", ""),
    "3745": ("Samantha Paul", "20083063", ""),
    "3746": ("Ravi Mykoo", "20078874", ""),
    "3747": ("Andrew Vincent", "20080359", ""),
    "3748": ("Sara Galletly", "20081542", ""),
    "3749": ("Samantha Jackson", "", ""),
    "3750": ("Tiara Peverall", "20109609", ""),
    "3751": ("Craig Couch", "", ""),
    "3752": ("Jason  Giannitti", "", ""),
    "3753": ("Paul Connolly", "20073670", ""),
    "3754": ("matthew mcculty", "20204558", ""),
    "3755": ("Lars Elstrodt", "", ""),
    "3756": ("Renee Eccleston", "", ""),
    "3757": ("Gus Harsfai", "20121612", ""),
    "3758": ("kamar scott", "21622517", ""),
    "3759": ("Andy Parr", "20086840", ""),
    "3760": ("Scott Brady", "", ""),
    "3761": ("Christophe  DUFOUR", "20681234", ""),
    "3762": ("Doralee Wright", "20084211", ""),
    "3763": ("Alicia Proud", "", ""),
    "3764": ("Rochelle Terry", "20084200", ""),
    "3765": ("Scott Elphinstone", "20084203", ""),
    "3766": ("Jerray  Brown", "20109616", ""),
    "3767": ("Trecia Hew", "20109603", ""),
    "3768": ("Ainslee  Bodden", "", ""),
    "3769": ("Yoania  Ebanks", "20218568", ""),
    "3770": ("MORICE FABIEN", "20204537", ""),
    "3771": ("Erica Green", "20204543", ""),
    "3772": ("Danielle Miller", "20453525", ""),
    "3773": ("Fabiola Pritchard", "20086269", ""),
    "3774": ("Ella Bergstrom", "21646167", ""),
    "3775": ("Roukeim Pitterson", "", ""),
    "3776": ("Cindy Jefferson Bulgin", "21646782", ""),
    "3777": ("Juan Granados", "20210886", ""),
    "3778": ("Peter Fyfe", "20087395", ""),
    "3779": ("Hoaran Xin", "20086418", ""),
    "3780": ("Laura Robinson", "20210936", ""),
    "3781": ("Damien McGovern", "21600486", ""),
    "3782": ("Tara Roney", "", ""),
    "3783": ("Jenavieve van den Bergh", "20139674", ""),
    "3784": ("Cristiano Vincentini", "20191762", ""),
    "3785": ("Sarah Philp", "20218437", ""),
    "3786": ("Scott Fleurie", "20139631", ""),
    "3787": ("Angelo Giuzio", "20217498", ""),
    "3788": ("Grace Carroll", "20204536", ""),
    "3789": ("Izabella Almeida", "20480632", ""),
    "3790": ("Jack Sorrill", "", ""),
    "3791": ("Jade Scott", "20482730", ""),
    "3792": ("Sarah  Superfine", "20204533", ""),
    "3793": ("Danielle Robson", "20220032", ""),
    "3794": ("Insurance Managers Association of Cayman Ltd.", "20209904", ""),
    "3795": ("ARLETA GALKA", "20213050", ""),
    "3796": ("Laurence Mercier", "20525487", ""),
    "3797": ("Winston  Williams", "20216732", ""),
    "3798": ("Segal Forbes", "20204539", ""),
    "3799": ("Lisandra Clarke", "20429416", ""),
    "3800": ("Victoria  Foulds", "20210923", ""),
    "3801": ("Tracy Gibbs", "20206489", ""),
    "3802": ("Curtis Wyatt", "20217196", ""),
    "3803": ("Belinda Standen", "20210938", ""),
    "3804": ("MIchael  Hagen", "", ""),
    "3805": ("Alessandro  Lemos", "20453511", ""),
    "3806": ("Jessica Bell", "20212033", ""),
    "3807": ("Dinesh Khadka", "", ""),
    "3808": ("Diana Virtue", "20541468", ""),
    "3809": ("Kathryn Ebanks", "20209905", ""),
    "3810": ("Andrea Hughes", "20033221", ""),
    "3811": ("Derek Lloyd", "20210884", ""),
    "3812": ("Yvonne Reid", "20243816", ""),
    "3813": ("Nicki Robnett", "20210934", ""),
    "3814": ("Eileen Keens", "20553418", ""),
    "3815": ("Jonathan Owens", "20211333", ""),
    "3816": ("Senia Ebanks", "20212030", ""),
    "3817": ("Yanet Swaby Ebanks", "", ""),
    "3818": ("Tyleisha Galbraith", "", ""),
    "3819": ("Jacques Redelinghuys", "20211610", ""),
    "3820": ("Krysta Parchman", "", ""),
    "3821": ("Janice Grafton", "20217198", ""),
    "3822": ("Christina Ulett", "20580091", ""),
    "3823": ("Florence McGrath", "", ""),
    "3824": ("Elisa Brown", "", ""),
    "3825": ("Dayna  Scott", "", ""),
    "3826": ("Colleen Gaio", "", ""),
    "3827": ("Elizabeth Gaio", "20215702", ""),
    "3828": ("Chrisitian Cardo", "21730158", ""),
    "3829": ("Maria Perla Datingaling", "", ""),
    "3830": ("Cayman Finance", "20221186", ""),
    "3831": ("Julia Vrbancic", "20215803", ""),
    "3832": ("Krista Kerr", "20217491", ""),
    "3833": ("Susan Clifford-Ebanks", "20549984", ""),
    "3834": ("Sharesse McDonald", "20324129", ""),
    "3835": ("Derren Burlington", "20243147", ""),
    "3836": ("Janice Mc Lauchlan", "20220999", ""),
    "3837": ("Michael Philbrick", "20204422", ""),
    "3838": ("Bernarda Gomez", "20215395", ""),
    "3839": ("Owen Merodon", "20215800", ""),
    "3840": ("Jessica Swinamer", "20215802", ""),
    "3841": ("Stephen Pascoe", "20220912", ""),
    "3842": ("Jo-Anne Merodon", "20217494", ""),
    "3843": ("Peter Dinsdale", "20256889", ""),
    "3844": ("RHONA SHORTER", "20243618", ""),
    "3845": ("Illann  Power", "", ""),
    "3846": ("Shaneika  Brown", "", ""),
    "3847": ("Gelyn Acosta", "20425841", ""),
    "3848": ("Natasha Thompson", "", ""),
    "3849": ("Lenny Verhoeven", "20217485", ""),
    "3850": ("Dawn Lawtey", "20243731", ""),
    "3851": ("Patrice Stewart", "20218553", ""),
    "3852": ("Roberto Terc", "20220033", ""),
    "3853": ("James Wrigley", "20220492", ""),
    "3854": ("Swan Sandoval", "21635780", ""),
    "3855": ("Eveadine  Watson", "", ""),
    "3856": ("Ashleigh  Dixon", "20243646", ""),
    "3857": ("Evelyn  Swaby", "", ""),
    "3858": ("Scott Browning", "20221001", ""),
    "3859": ("Pierre Landry", "20220026", ""),
    "3860": ("Leslie-anne Bucalbos Elaydo", "20221328", ""),
    "3861": ("Margarita Kamerzanova", "20398437", ""),
    "3862": ("Darren Ratz", "20240358", ""),
    "3863": ("Guemas Matthieu", "20243702", ""),
    "3864": ("Lundie Richards", "20243730", ""),
    "3865": ("Susann Espinosa", "20243759", ""),
    "3866": ("Nicolas Ruiz", "20243674", ""),
    "3867": ("Patricia Healy", "20750917", ""),
    "3868": ("Jamila  McLean", "20966072", ""),
    "3869": ("Fanisha Nunez- Bush", "20243788", ""),
    "3870": ("Anigie Gonzalez", "", ""),
    "3871": ("Raulito Gonzalez", "20453547", ""),
    "3872": ("Kaisha Morrison", "", ""),
    "3873": ("Hope Conolly", "20300846", ""),
    "3874": ("Samuel Muir", "20440272", ""),
    "3875": ("ERICK LUCIO", "20354254", ""),
    "3876": ("Tara-Lee Schmarr", "20002264", ""),
    "3877": ("James Meehan", "20306715", ""),
    "3878": ("Rochelle Holness", "20324143", ""),
    "3879": ("Megan Ebanks", "20354275", ""),
    "3880": ("Stephen Broadbelt", "20518986", ""),
    "3881": ("Paula Rossano", "20525490", ""),
    "3882": ("Jody McFarland", "20354234", ""),
    "3883": ("Sandhya Shakya", "20425794", ""),
    "3884": ("Peter Goddard", "20531836", ""),
    "3885": ("Mallory Creed", "20453554", ""),
    "3886": ("Joy Andrade", "20667523", ""),
    "3887": ("D'Angela Andrade", "20518675", ""),
    "3888": ("MATTHEW FREDERICK", "20480708", ""),
    "3889": ("Jessica Reed", "20506458", ""),
    "3890": ("Alyson Medeiros", "20518847", ""),
    "3891": ("Rochell Foster", "20553919", ""),
    "3892": ("Simon Tyrrell", "20440258", ""),
    "3893": ("Sharn  Ricketts Lawrence", "20440260", ""),
    "3894": ("Elenie Falconer-Cloutier", "20541451", ""),
    "3895": ("Teresa Grimes", "20518676", ""),
    "3896": ("Niall Quinlan", "20453518", ""),
    "3897": ("Al-Siddique Allie", "20039811", ""),
    "3898": ("David Scott", "20548996", ""),
    "3899": ("simone pagnozzi", "20480606", ""),
    "3900": ("VANESSA EVANS", "20750991", ""),
    "3901": ("Betina Ridley", "20453478", ""),
    "3902": ("Chris Palmer", "20525501", ""),
    "3903": ("Reuben Meade", "21629172", ""),
    "3904": ("Ronald Challenger", "20539203", ""),
    "3905": ("CHRISTINA POWERY", "20525513", ""),
    "3906": ("Tomasz  Mialkos", "20480684", ""),
    "3907": ("David Mellinas Ucles", "20550936", ""),
    "3908": ("Heather Pierre", "20531827", ""),
    "3909": ("Marlene Cohen", "20546600", ""),
    "3910": ("Robert Betts", "20471670", ""),
    "3911": ("Shamus Smith", "20541454", ""),
    "3912": ("Nicholas Teasdale", "20480605", ""),
    "3913": ("Granger Haugh", "20480615", ""),
    "3914": ("Rizza Recto", "20546545", ""),
    "3915": ("Inbaraj Samuel Sundram", "20480674", ""),
    "3916": ("Andrew McCartney", "20480391", ""),
    "3917": ("Jayne Battersby", "", ""),
    "3918": ("Tatyana Salazar", "", ""),
    "3919": ("Amber Myrie", "20550006", ""),
    "3920": ("Mike Olckers", "20542967", ""),
    "3921": ("Compass Media", "20687570", ""),
    "3922": ("Bruno  Deluche", "20505170", ""),
    "3923": ("LORAN LEWIS", "", ""),
    "3924": ("Selena Frederick", "", ""),
    "3925": ("Nicole Pillay", "20518290", ""),
    "3926": ("Matthew Steemson", "20525488", ""),
    "3927": ("Stephanie van Deventer", "20542784", ""),
    "3928": ("Dewayne  Mclean", "20541475", ""),
    "3929": ("Paul  Inniss", "", ""),
    "3930": ("Haley D_x0019_ Hue", "", ""),
    "3931": ("Christopher Dyckman", "20551703", ""),
    "3932": ("Elljanna Dixon", "20564111", ""),
    "3933": ("Shaquiri Flores", "", ""),
    "3934": ("Bar Alezrah", "", ""),
    "3935": ("Cleto Bodden", "", ""),
    "3936": ("Imre Kozma", "20548997", ""),
    "3937": ("Laura Pageon", "20548998", ""),
    "3938": ("Joela Masayon", "", ""),
    "3939": ("Anna Marin Chua", "20550946", ""),
    "3940": ("Astrid Pilato", "20550000", ""),
    "3941": ("Jonasci  Carter", "20550944", ""),
    "3942": ("sean evans", "20632902", ""),
    "3943": ("PRAKASH RAMNANI", "20549990", ""),
    "3944": ("RUDY KUDRITZKI", "20681170", ""),
    "3945": ("meliki miller", "20667098", ""),
    "3946": ("Robyn Larkin", "20550956", ""),
    "3947": ("Treasha Bodden", "20052493", ""),
    "3948": ("Lijin kumar Pushpangadan", "20550958", ""),
    "3949": ("Kerri-Ann Gillies", "20553138", ""),
    "3950": ("Francois  Sevenster", "20553921", ""),
    "3951": ("ANdrew Twssdie", "20551707", ""),
    "3952": ("Jake Cris Bolanos", "20551701", ""),
    "3953": ("Dimitri Rohan", "21727878", ""),
    "3954": ("Shawn Bryan", "20750912", ""),
    "3955": ("Mirquella  De la rosa", "20580090", ""),
    "3956": ("Richard Douglas", "20553409", ""),
    "3957": ("Shane Martin", "", ""),
    "3958": ("ANDREA LAGUNA", "", ""),
    "3959": ("Paula Lombardo", "20564517", ""),
    "3960": ("Talia Davidson", "20574456", ""),
    "3961": ("Krista  Ebanks", "20564104", ""),
    "3962": ("Paul Koenig", "20572309", ""),
    "3963": ("Vana  Bennett", "20564109", ""),
    "3964": ("Patrick Mulrenan", "20555297", ""),
    "3965": ("Danielle Brown", "21523328", ""),
    "3966": ("Roberto Clemente", "20564109", ""),
    "3967": ("Prasanna Ketheeswaran", "20580093", ""),
    "3968": ("Jerome Bailey", "", ""),
    "3969": ("Derricka Neysmith", "", ""),
    "3970": ("Cassandra  Williams", "", ""),
    "3971": ("Leonard Lewis", "21633358", ""),
    "3972": ("Tsui Lam", "20572314", ""),
    "3973": ("Lisa Broadbridge", "", ""),
    "3974": ("David Hardy", "20010435", ""),
    "3975": ("Walter van der Merwe", "21636234", ""),
    "3976": ("Stuart Sentance", "20005786", ""),
    "3977": ("neil sherlock", "", ""),
    "3978": ("Herman Myrie", "", ""),
    "3979": ("Cherie-Anne Henderson-Dam", "", ""),
    "3980": ("Dominic Pouchie", "20006623", ""),
    "3981": ("James Ball", "", ""),
    "3982": ("Brielle Watler", "20754850", ""),
    "3983": ("Shane McCoon", "", ""),
    "3984": ("Kristen Thomson", "21642339", ""),
    "3985": ("Robert Quinn", "20691570", ""),
    "3986": ("Terry Ballard", "21631620", ""),
    "3987": ("Kate Theron", "", ""),
    "3988": ("Erin Parkinson", "", ""),
    "3989": ("Alexander Manfield", "20594299", ""),
    "3990": ("Cherrie  Graham", "", ""),
    "3991": ("Bernice Scott", "20667355", ""),
    "3992": ("Heather Ketterer", "", ""),
    "3993": ("Tim Baildam", "21628514", ""),
    "3994": ("Alexandra Franklin", "20691592", ""),
    "3995": ("ANITA BROSNAN", "20580101", ""),
    "3996": ("Lyea Rivers", "20531467", ""),
    "3997": ("Georgina Loxton", "", ""),
    "3998": ("Justin Moryto", "21591173", ""),
    "3999": ("renante areola", "", ""),
    "4000": ("Sharelle Escalante", "", ""),
    "4001": ("Bianca  Leacock", "", ""),
    "4002": ("Benjamin  Bartholomew", "", ""),
    "4003": ("Vidya Ravella", "20617215", ""),
    "4004": ("Isidora Eden", "21630672", ""),
    "4005": ("Gregory Kidd", "", ""),
    "4006": ("Renee Bernardo", "20012720", ""),
    "4007": ("Elizabeth Blake", "20595925", ""),
    "4008": ("Peter Sipos", "", ""),
    "4009": ("Dayan Cardenas Duquesne", "20033027", ""),
    "4010": ("Rashad Jervis", "20659968", ""),
    "4011": ("Luke Murray", "20667513", ""),
    "4012": ("Davania Hurlston", "", ""),
    "4013": ("Paula Aline Santos de Paula", "20595922", ""),
    "4014": ("Patricia Lobo", "", ""),
    "4015": ("Dawn Eaton", "20595879", ""),
    "4016": ("navaro cummings", "21655976", ""),
    "4017": ("James C. McCombs III", "", ""),
    "4018": ("Priscilla  Jensen", "", ""),
    "4019": ("Jill Taylor", "20750952", ""),
    "4020": ("Rachel Davis", "20659970", ""),
    "4021": ("Lindia Lewis", "", ""),
    "4022": ("Dexter  Beckford", "", ""),
    "4023": ("Kimberley Hecimovic", "21640894", ""),
    "4024": ("Helen Day", "20667526", ""),
    "4025": ("Helen Day", "20667526", ""),
    "4026": ("Richard Collett", "20680847", ""),
    "4027": ("Jada Smith", "21646737", ""),
    "4028": ("Tamoy Phillips", "", ""),
    "4029": ("Maria  Johnsson", "20722646", ""),
    "4030": ("Fiona O Dea", "20664843", ""),
    "4031": ("Sarah Hough", "", ""),
    "4032": ("Patrizia Bruzio", "20632927", ""),
    "4033": ("Britta Bush", "20667537", ""),
    "4034": ("Nickolas Hamilton-Jackson", "21642198", ""),
    "4035": ("Amanda Lankford", "21640703", ""),
    "4036": ("Christina Elliott-Ebanks", "21634175", ""),
    "4037": ("Carly Dignam", "", ""),
    "4038": ("Madelaine Sahit", "", ""),
    "4039": ("Andre  Savoury", "20750930", ""),
    "4040": ("Wanda Mellaneo", "20675156", ""),
    "4041": ("Vanessa Hansen Allott", "21068545", ""),
    "4042": ("Catherine Molitor", "20595929", ""),
    "4043": ("Ranald Henderson", "21633331", ""),
    "4044": ("Michael y Lexi Binckes", "21614595", ""),
    "4045": ("Mouya Williams", "20725841", ""),
    "4046": ("Tatum Acutt", "20667547", ""),
    "4047": ("Alexandra McAlpine-Culas", "", ""),
    "4048": ("Carol Bell", "", ""),
    "4049": ("Ann Jones", "20754846", ""),
    "4050": ("Jason Buckle", "20012774", ""),
    "4051": ("MORVEN BODDEN", "20750973", ""),
    "4052": ("Marcus Cubillo", "20709101", ""),
    "4053": ("Dionicia  Moya", "21633341", ""),
    "4054": ("Patrina Dixon", "", ""),
    "4055": ("Sogol Lohi", "20681161", ""),
    "4056": ("Genevieve Preston", "20595883", ""),
    "4057": ("Eleanor Shakespeare", "20595926", ""),
    "4058": ("MacDonald Machingura", "", ""),
    "4059": ("Tamsin West", "", ""),
    "4060": ("Karen Coles", "", ""),
    "4061": ("Zorina  McCoon", "21202426", ""),
    "4062": ("Gretchen Goodbody Gringley", "20595880", ""),
    "4063": ("Daniel J  Spiegel", "", ""),
    "4064": ("Kedeshia  Thomas", "", ""),
    "4065": ("Anu O_x0019_ Driscoll", "", ""),
    "4066": ("Camilo Herrera David", "", ""),
    "4067": ("Jon-Andrew Japal", "21642438", ""),
    "4068": ("Laura Durston", "", ""),
    "4069": ("Aoife Lynch", "20750963", ""),
    "4070": ("Trudi Higginbotham", "", ""),
    "4071": ("Shana Myles", "", ""),
    "4072": ("Madennys Ebanks", "", ""),
    "4073": ("Stephanie  Ditta", "20659976", ""),
    "4074": ("Shay Miller", "", ""),
    "4075": ("Kim Ferreira", "20632952", ""),
    "4076": ("Marcus Rowe", "20750913", ""),
    "4077": ("Anne Briggs", "", ""),
    "4078": ("Daniel Florek", "20675154", ""),
    "4079": ("Racquel Sutherland", "21686448", ""),
    "4080": ("Harry Ebanks", "", ""),
    "4081": ("Tanya Wigmore", "20681153", ""),
    "4082": ("SALLY Young", "20689616", ""),
    "4083": ("Michel Jacobs", "21616662", ""),
    "4084": ("Lynn Corkin", "20667099", ""),
    "4085": ("Janelle Tibbetts", "", ""),
    "4086": ("Lorna Murphy", "20817382", ""),
    "4087": ("Kellee Ann Holness", "21523112", ""),
    "4088": ("Dieshalee Bush-Terry", "20691590", ""),
    "4089": ("Javiera Aguayo", "20659790", ""),
    "4090": ("Chantelle MacKenzie", "21633351", ""),
    "4091": ("David Blake-Cownie", "", ""),
    "4092": ("Patrick Rosenfeld", "", ""),
    "4093": ("Andrew Croft", "21628123", ""),
    "4094": ("Keanna Douglas", "21625820", ""),
    "4095": ("Allison Gonsalves", "21634430", ""),
    "4096": ("David Cownie", "20084035", ""),
    "4097": ("Joseph  Jackson", "20632976", ""),
    "4098": ("Belinda Hart", "", ""),
    "4099": ("Elliot Power", "20691582", ""),
    "4100": ("Elvira Pearce", "20595900", ""),
    "4101": ("Rose Williams", "", ""),
    "4102": ("Robert Arsenault", "21719011", ""),
    "4103": ("Krystle Barrett", "20667520", ""),
    "4104": ("Wian Lategan", "20622501", ""),
    "4105": ("Wesley  Cullum", "21597722", ""),
    "4106": ("Triona Clarke", "20681140", ""),
    "4107": ("Fiona Graham", "", ""),
    "4108": ("Nicola Davies", "20793078", ""),
    "4109": ("Alisa James", "", ""),
    "4110": ("Edward Scott", "20667535", ""),
    "4111": ("Lisa Small", "", ""),
    "4112": ("Roger Small", "", ""),
    "4113": ("Shelley Leonard", "", ""),
    "4114": ("Justin Schmidt", "", ""),
    "4115": ("Mark Choice", "21611438", ""),
    "4116": ("Darryl Greer", "21597799", ""),
    "4117": ("Susan Hafer-Greene", "21629170", ""),
    "4118": ("Tami Maines", "", ""),
    "4119": ("Paula McCartney", "20632926", ""),
    "4120": ("kemaya samuels", "21620384", ""),
    "4121": ("Susan Ardill", "20666541", ""),
    "4122": ("Sabine Schommarz", "20751012", ""),
    "4123": ("Allison Anglin", "20667531", ""),
    "4124": ("Richard Murphy", "21637723", ""),
    "4125": ("Emer Timmons", "20667534", ""),
    "4126": ("Jaime-Lee Eccles", "20667543", ""),
    "4127": ("Nick Ferreira", "", ""),
    "4128": ("Iain Currie", "21637072", ""),
    "4129": ("Jheanelle Jones", "", ""),
    "4130": ("Todon Leshikar", "20632878", ""),
    "4131": ("Brittny Bustillo", "20838201", ""),
    "4132": ("Rebecca Reid", "21629667", ""),
    "4133": ("Breda Verling", "20681235", ""),
    "4134": ("Yiming Liu", "21634249", ""),
    "4135": ("Jatinder Kumar", "", ""),
    "4136": ("Allison Olarou", "20659963", ""),
    "4137": ("Paul Kelly", "21628495", ""),
    "4138": ("Thomas Akdeniz", "21066911", ""),
    "4139": ("Melissa  Llewellyn", "", ""),
    "4140": ("Denis Collins", "20594667", ""),
    "4141": ("Natasha  Macfadyen", "", ""),
    "4142": ("Tamika McField", "", ""),
    "4143": ("Elvis Jackson", "", ""),
    "4144": ("Robby Polk", "20750916", ""),
    "4145": ("Lisa Andre", "20667519", ""),
    "4146": ("SHARLENE BRENKUS", "20632901", ""),
    "4147": ("Nicholas Gaze", "20618678", ""),
    "4148": ("Heather Froude", "21523315", ""),
    "4149": ("Nicholas  Manning", "20632951", ""),
    "4150": ("Sofronia Tsagari", "", ""),
    "4151": ("Brandon Cadle", "", ""),
    "4152": ("ERIN WINCZURA", "20681208", ""),
    "4153": ("Margaret Wilks", "21730347", ""),
    "4154": ("GABRIEL CADENAS", "", ""),
    "4155": ("Gregery Barnes", "", ""),
    "4156": ("Craig  Pascoe", "20667528", ""),
    "4157": ("Jenna Erin Hydes", "21068307", ""),
    "4158": ("Ailie MacGeoch", "20633000", ""),
    "4159": ("James Wignall", "20667102", ""),
    "4160": ("Elizabeth  Sharples", "20667527", ""),
    "4161": ("Olivia Zimmer", "", ""),
    "4162": ("Barbie Bodden", "", ""),
    "4163": ("Lauren Hale", "", ""),
    "4164": ("Hayley Palmer", "", ""),
    "4165": ("Frank O'Leary", "20595888", ""),
    "4166": ("Camilo Herrera", "", ""),
    "4167": ("Travis Thompson", "20622514", ""),
    "4168": ("Sonja Salmon", "21629174", ""),
    "4169": ("Zoe Rogers", "20622500", ""),
    "4170": ("Taiba Khan", "21066729", ""),
    "4171": ("Priva VPN", "", ""),
    "4172": ("Kristina Ring", "20667100", ""),
    "4173": ("Taiba Khan", "21066729", ""),
    "4174": ("Kendra Ebanks", "20632975", ""),
    "4175": ("Rodney Dixon", "", ""),
    "4176": ("Tashia  James", "", ""),
    "4177": ("Darren  Ebanks", "21068212", ""),
    "4178": ("Tyra McField", "", ""),
    "4179": ("Monique Munden", "", ""),
    "4180": ("Joanne Woodcock", "", ""),
    "4181": ("monica walton", "21637510", ""),
    "4182": ("Sheynae Watson", "20751001", ""),
    "4183": ("Delano Eksteen", "20595962", ""),
    "4184": ("Francis Omar Campos", "20951530", ""),
    "4185": ("Louise Johnston", "", ""),
    "4186": ("Oxana Hendrikse", "", ""),
    "4187": ("Jill Mojica", "20667524", ""),
    "4188": ("Leianne Daykin", "20664844", ""),
    "4189": ("Sarah Harkness", "", ""),
    "4190": ("Stormi Woodhall", "", ""),
    "4191": ("Lisa O Donoghue", "", ""),
    "4192": ("Gabriel Alvero", "", ""),
    "4193": ("Kayla Robinson", "20667521", ""),
    "4194": ("Hannah Gethin", "", ""),
    "4195": ("Stevana Hanna", "20667086", ""),
    "4196": ("Krista  Wight", "20754871", ""),
    "4197": ("Hermes Cuello", "20618686", ""),
    "4198": ("Mark Botha", "20618683", ""),
    "4199": ("Susanna de Saram", "20750999", ""),
    "4200": ("Nickisha  Stephenson", "20632937", ""),
    "4201": ("Angella Watler", "20594523", ""),
    "4202": ("Agata Pawlak-Mayo", "", ""),
    "4203": ("Kesy Cook", "", ""),
    "4204": ("Steve Ali", "20966104", ""),
    "4205": ("Siska Putri", "20621483", ""),
    "4206": ("Elaine Anderson", "20595928", ""),
    "4207": ("Fredrich van der Merwe", "21639774", ""),
    "4208": ("Harry Rasmussen", "21641565", ""),
    "4209": ("Tesla Oviedo Thomas", "", ""),
    "4210": ("Samantha Payne", "21523060", ""),
    "4211": ("Amanda Hilditch", "20667539", ""),
    "4212": ("Daniel de Wet", "", ""),
    "4213": ("Dan Hall", "", ""),
    "4214": ("Ian MacLean", "", ""),
    "4215": ("Erin Reddington", "20595938", ""),
    "4216": ("Courtney MacQueen", "21628096", ""),
    "4217": ("John Paul Hennelly", "", ""),
    "4218": ("Natasha Casebolt", "", ""),
    "4219": ("Jonathan Walmsley", "20618681", ""),
    "4220": ("Ludwig Richter", "21523103", ""),
    "4221": ("Belinda Taylor", "20085887", ""),
    "4222": ("Natasha Last", "20750961", ""),
    "4223": ("Lori Monk", "20944841", ""),
    "4224": ("Heidi van Batenburg-Stafford", "20667545", ""),
    "4225": ("Elena Nelmes", "", ""),
    "4226": ("Collette y Darren Mclaughlin", "", ""),
    "4227": ("Gregory Davis", "", ""),
    "4228": ("Isabelle Baron", "", ""),
    "4229": ("Mark Taylor", "", ""),
    "4230": ("Francine  Wright", "", ""),
    "4231": ("Tarryn Humphries", "21634717", ""),
    "4232": ("Martin Edelenbos", "20667476", ""),
    "4233": ("JANET FRANCIS", "20618682", ""),
    "4234": ("Fiona Barrie", "21611597", ""),
    "4235": ("Rachael Zimmer", "20681179", ""),
    "4236": ("Jermaine Gregory", "20618761", ""),
    "4237": ("Aaron  Huber", "21633329", ""),
    "4238": ("Brian Kinsella", "21611673", ""),
    "4239": ("Scott McCarty", "", ""),
    "4240": ("Davian Jones", "21628575", ""),
    "4241": ("Timothy Swallow", "21066791", ""),
    "4242": ("Lym Liu", "", ""),
    "4243": ("April Laspinas", "20667529", ""),
    "4244": ("Sarah Curtis", "21597791", ""),
    "4245": ("Bradley Bregani", "", ""),
    "4246": ("Sean Ladley", "21628493", ""),
    "4247": ("Anthony Nelson", "21615230", ""),
    "4248": ("Chereel Daley", "", ""),
    "4249": ("Lucas Robbins", "20618685", ""),
    "4250": ("Hannah Foreman", "20951268", ""),
    "4251": ("Tarcie ann Banhan", "", ""),
    "4252": ("April  General", "20675155", ""),
    "4253": ("Richard Marian", "20659972", ""),
    "4254": ("Garth MacDonald", "20206419", ""),
    "4255": ("Eva Manning", "21629453", ""),
    "4256": ("Stephen Jordan", "20618762", ""),
    "4257": ("Joshua Gonzalez", "", ""),
    "4258": ("Janet Gardner", "21523310", ""),
    "4259": ("Plamka Evtimov", "20618766", ""),
    "4260": ("Kerry-Ann Jackson", "21597863", ""),
    "4261": ("Miguel Riverol", "20793037", ""),
    "4262": ("Shakir Johnson", "20955670", ""),
    "4263": ("Karina Maclean", "20667522", ""),
    "4264": ("Andrew Cousins", "21068116", ""),
    "4265": ("Christine Maltman", "20002365", ""),
    "4266": ("Mikhalia  Byndloss", "", ""),
    "4267": ("Michael Fernandez", "", ""),
    "4268": ("Shelley George", "21066760", ""),
    "4269": ("Dominic  Ross", "21722168", ""),
    "4270": ("Benjamin Pershick", "", ""),
    "4271": ("Safon Campbell", "20659975", ""),
    "4272": ("Tenisha Thompson", "20681145", ""),
    "4273": ("Michelle Reynolds Henry", "21523101", ""),
    "4274": ("Adrien Royston", "20659962", ""),
    "4275": ("Valerie Randrianasolo", "20675148", ""),
    "4276": ("Sarah Hope", "20667090", ""),
    "4277": ("Rob Malloy", "20618738", ""),
    "4278": ("Lisa Embleton", "21673621", ""),
    "4279": ("Craig Morgan", "20618733", ""),
    "4280": ("Jhaneen  Bodden", "", ""),
    "4281": ("Dajsha Samuels", "20667536", ""),
    "4282": ("Alison Canning", "20667541", ""),
    "4283": ("Leandro Heck", "21634244", ""),
    "4284": ("Sarah Ross", "20793050", ""),
    "4285": ("Christine Demsys", "20659964", ""),
    "4286": ("Adiaris Williams-Rodriguez", "21628143", ""),
    "4287": ("Cora Kerr", "20618737", ""),
    "4288": ("Idan Kfir Roberts", "", ""),
    "4289": ("Hunter Willkom", "20659966", ""),
    "4290": ("Ewan Jacques", "", ""),
    "4291": ("Manju Shakya", "20681188", ""),
    "4292": ("Alison Duncan", "20754888", ""),
    "4293": ("Hannah Fowler", "", ""),
    "4294": ("Bernice Greyling", "20667538", ""),
    "4295": ("Emma Byrne", "", ""),
    "4296": ("Nicko Henry", "21523068", ""),
    "4297": ("Douglas Dodds", "", ""),
    "4298": ("Michael Diamond", "", ""),
    "4299": ("Eden Jordan", "20667546", ""),
    "4300": ("Raul Phillips", "20966093", ""),
    "4301": ("Olivaire  Watler", "21638213", ""),
    "4302": ("Melissa Hamilton", "20667532", ""),
    "4303": ("Angela Barkhouse", "21630734", ""),
    "4304": ("Mary  Austin", "21630066", ""),
    "4305": ("Akinyele Logan", "20667509", ""),
    "4306": ("Luisa O_x0019_ Neil", "20677813", ""),
    "4307": ("Chen kaylee", "", ""),
    "4308": ("Ma Letecia Emmeline Pet", "", ""),
    "4309": ("Helen O'Sullivan", "20659980", ""),
    "4310": ("Chantal Byrd", "20659961", ""),
    "4311": ("Stellerie Noyons", "20681158", ""),
    "4312": ("Amy Powell", "", ""),
    "4313": ("Gabriela Gibson", "", ""),
    "4314": ("Lisa Wood", "21630061", ""),
    "4315": ("Kalia Elliott", "21639418", ""),
    "4316": ("Lissett  Higgins", "", ""),
    "4317": ("Susannah Sweetman", "", ""),
    "4318": ("Sommer McField", "", ""),
    "4319": ("Juliet Bell", "20754878", ""),
    "4320": ("Rizlaine D'Ouissi", "", ""),
    "4321": ("Chariza Jane Are", "", ""),
    "4322": ("Elizabeth Rae", "20659965", ""),
    "4323": ("Tyler Lawson", "20675159", ""),
    "4324": ("Bernadette Myles", "", ""),
    "4325": ("Antonea Ebanks", "", ""),
    "4326": ("Meghan Ricker", "", ""),
    "4327": ("Sarah Wheeler", "20667087", ""),
    "4328": ("Majin  Paul", "20753500", ""),
    "4329": ("Marcela Rondon narvaez", "", ""),
    "4330": ("Conor Given", "20793000", ""),
    "4331": ("Bilal Ahmad", "", ""),
    "4332": ("Irish Bartolata", "", ""),
    "4333": ("Sai Nidval", "", ""),
    "4334": ("Iva Rozinkova", "20667544", ""),
    "4335": ("Anne-Marie Leadbetter", "20689367", ""),
    "4336": ("Marie Brennan", "21597809", ""),
    "4337": ("Sabine Schommarz", "20751012", ""),
    "4338": ("Tara Robinson", "20681151", ""),
    "4339": ("SHARI PATTERSON", "20681163", ""),
    "4340": ("Sharon Steer", "21202390", ""),
    "4341": ("LLOYD CAMPBELL", "21631641", ""),
    "4342": ("Simone Brady", "", ""),
    "4343": ("KIERAN MEHIGAN", "", ""),
    "4344": ("Samantha Sinclair", "2006436", ""),
    "4345": ("Rodrigo Gordillo", "20681176", ""),
    "4346": ("Lassanger  Parkinson", "20778478", ""),
    "4347": ("Elizabeth  Owens", "", ""),
    "4348": ("JOHN PAUL RAMOS", "", ""),
    "4349": ("Racquel Barnes", "20667091", ""),
    "4350": ("DELIA SLATER", "20793092", ""),
    "4351": ("Sheryl Campbell", "21641600", ""),
    "4352": ("Jody-Ann Moore", "21068339", ""),
    "4353": ("Natalie Hart", "20667095", ""),
    "4354": ("Katrina Thompson", "21523113", ""),
    "4355": ("William Korkie", "20681133", ""),
    "4356": ("yasmin Wheatle-Smith", "21673439", ""),
    "4357": ("Nadine  Bryan", "20753493", ""),
    "4358": ("Julianne Yacyshyn", "20750909", ""),
    "4359": ("Daniel Artusi", "20681219", ""),
    "4360": ("Setia Carlson", "21630341", ""),
    "4361": ("Hilary Cuff", "21667141", ""),
    "4362": ("Kimberly  McPherson", "20696865", ""),
    "4363": ("Camile Gooden", "21640129", ""),
    "4364": ("Sanjay Siddappa", "", ""),
    "4365": ("Joshaun Barclay", "", ""),
    "4366": ("Lecia McLaughlin", "20750036", ""),
    "4367": ("Nigel Smith", "21067714", ""),
    "4368": ("Karen Hydes", "20675150", ""),
    "4369": ("Michelle Erwin", "20675149", ""),
    "4370": ("Marivic Montecalbo", "21068147", ""),
    "4371": ("Cardel McBean", "21533389", ""),
    "4372": ("Mayzie Kelly", "", ""),
    "4373": ("Douglas Harrell", "20675151", ""),
    "4374": ("Kierstin Stewart", "21758560", ""),
    "4375": ("Lucy Diggle", "", ""),
    "4376": ("Daphne Ewing-Chow", "", ""),
    "4377": ("Tennille Broderick", "20750997", ""),
    "4378": ("Khieandre Grant", "", ""),
    "4379": ("Kayla Banks", "", ""),
    "4380": ("Phyllis Miller", "21634216", ""),
    "4381": ("Marcia Duggon", "20753606", ""),
    "4382": ("Melisa Bent-Hamilton", "", ""),
    "4383": ("Kirsty Farrell", "21614591", ""),
    "4384": ("Miriam Berry", "21678851", ""),
    "4385": ("Whyman Ebanks", "21066639", ""),
    "4386": ("Deviekin  Tibbetts", "", ""),
    "4387": ("Tiara Myles", "", ""),
    "4388": ("Nicholas Kedney", "21067777", ""),
    "4389": ("Scimone Campbell", "21066912", ""),
    "4390": ("EMENIKE MCLEAN-MYLES", "", ""),
    "4391": ("Ruwan Jayasekera", "20751016", ""),
    "4392": ("Selvin Whyte", "", ""),
    "4393": ("David Olson", "21597864", ""),
    "4394": ("Malika Chow", "21636212", ""),
    "4395": ("Laura Ebanks", "21620026", ""),
    "4396": ("Benedict Hicks", "21068177", ""),
    "4397": ("GLEN WIGNEY", "21620160", ""),
    "4398": ("Kayon Whyte", "20951179", ""),
    "4399": ("Lauren Tims", "20951163", ""),
    "4400": ("Britt Viljoen", "21631656", ""),
    "4401": ("Lorenzo Brown", "20000392", ""),
    "4402": ("Karen Stephen-Dalton", "", ""),
    "4403": ("David Bogran", "", ""),
    "4404": ("Gemma Henry", "21630056", ""),
    "4405": ("Eric Hertha", "", ""),
    "4406": ("Antonette Baptist", "", ""),
    "4407": ("Carl Brenton", "20663653", ""),
    "4408": ("Pearlyn Henry-Burrell", "", ""),
    "4409": ("Anita Zagorski", "20023066", ""),
    "4410": ("Robin Bond", "20689358", ""),
    "4411": ("Barbara / Robert Oosterwyk", "", ""),
    "4412": ("Ryan Macleod", "20792936", ""),
    "4413": ("Jeremy Anderson", "21630130", ""),
    "4414": ("Petrina Moore", "", ""),
    "4415": ("YJhDEcbtWk kCBDKzUgNXm", "", ""),
    "4416": ("Martin Kelly", "20751028", ""),
    "4417": ("William Foster", "21726664", ""),
    "4418": ("Keysha Sailsman", "20792949", ""),
    "4420": ("Tetiana Lotts", "20750905", ""),
    "4421": ("ANGANEE LEWIS", "20595827", ""),
    "4422": ("LESSLEY CHRISTUDOSS", "", ""),
    "4423": ("Robert Tyler", "21617738", ""),
    "4424": ("Sherene  Morgan", "", ""),
    "4425": ("Cora Miller", "21630709", ""),
    "4426": ("Judith Mauer", "20754865", ""),
    "4427": ("Wendy Hernandez", "21619177", ""),
    "4428": ("Hopeton Lindo", "20011697", ""),
    "4429": ("Jasmine  Powery", "20779004", ""),
    "4430": ("Adam Lambert", "21686049", ""),
    "4431": ("Dennis Pascal", "20951521", ""),
    "4432": ("Constance Robertson", "20750919", ""),
    "4433": ("Nicole Gagliano", "20754863", ""),
    "4434": ("Theo Louw", "", ""),
    "4435": ("SHAN XU", "", ""),
    "4436": ("Cory Allenger", "", ""),
    "4437": ("Kindu Ebanks", "21666452", ""),
    "4438": ("Abby Guilmette", "21001224", ""),
    "4439": ("Rachel Funk", "20751024", ""),
    "4440": ("Juliet Fenn", "", ""),
    "4441": ("Delecia Ebanks", "21689035", ""),
    "4442": ("Simon Miller", "21597736", ""),
    "4443": ("Hannah Daval-Bowden", "21597887", ""),
    "4444": ("deckers restaurant", "21068369", ""),
    "4445": ("Alaina Danley", "20086212", ""),
    "4446": ("Anita Zagorski", "20023066", ""),
    "4447": ("Leanne Thorne-Jeanson", "", ""),
    "4448": ("Bradley  Connolly", "", ""),
    "4449": ("Gary Franklin", "21597920", ""),
    "4450": ("Tianah Mason", "", ""),
    "4451": ("Franz  Manderson", "21595049", ""),
    "4452": ("Heidi Kiss", "21628567", ""),
    "4453": ("Tanya Campbell", "21597735", ""),
    "4454": ("Cheryl Thompson", "21520203", ""),
    "4455": ("Farieza Hussain", "21634253", ""),
    "4456": ("Danielle Klischuk", "20006714", ""),
    "4457": ("Petra Steinlechner", "20751025", ""),
    "4458": ("Alisha Bailey", "21620163", ""),
    "4459": ("Paula Mcfarlane", "", ""),
    "4460": ("shaundelle rodrigues", "20753602", ""),
    "4461": ("Denise Warren", "", ""),
    "4462": ("Arnelio Triana", "", ""),
    "4463": ("Rhojen mason", "20751021", ""),
    "4464": ("leyah neil", "21629456", ""),
    "4465": ("Ivan Webb", "21628545", ""),
    "4466": ("Lanette van der merwe", "21611514", ""),
    "4467": ("Karina Lama", "20753501", ""),
    "4468": ("Jevon Gregory", "20001387", ""),
    "4469": ("Abigail Tajah", "", ""),
    "4470": ("O_x0019_ Brian  Clarke", "", ""),
    "4471": ("Edward Ford", "21637171", ""),
    "4472": ("Cashema Rankine", "", ""),
    "4473": ("Charlotte  Nicholson", "20750929", ""),
    "4474": ("Godofredo Jr Ariem", "", ""),
    "4475": ("Jason Powers", "20951260", ""),
    "4476": ("Martin Steyn", "20792975", ""),
    "4477": ("Ariana Bodden", "20951155", ""),
    "4478": ("Michele Aubert", "20944862", ""),
    "4479": ("Tiffany Dilbert", "20750994", ""),
    "4480": ("Simone Barrett", "", ""),
    "4481": ("Angelique Collins", "20817398", ""),
    "4482": ("Renzo Escalante", "21628505", ""),
    "4483": ("Tamar White", "", ""),
    "4484": ("Emma Evans", "21687563", ""),
    "4485": ("Sherene Hutchinson", "20750911", ""),
    "4486": ("Mark Hawkins", "20817398", ""),
    "4487": ("John Bodden", "", ""),
    "4488": ("Shakira Cox", "21634496", ""),
    "4489": ("Anna Ghandilyan", "20750920", ""),
    "4490": ("DIANA  JOSEPH", "20026662", ""),
    "4491": ("Amy Wallis", "21082456", ""),
    "4492": ("Ross Cowan", "", ""),
    "4493": ("Mark Woollard", "21597840", ""),
    "4494": ("Kerrina Cecere", "21068457", ""),
    "4495": ("Richard Harwood", "20951277", ""),
    "4496": ("Brianny Molina", "20005267", ""),
    "4497": ("Crystal Rayworth", "21644981", ""),
    "4498": ("Jessica Murrie", "21638217", ""),
    "4499": ("Christopher Samuels", "21628551", ""),
    "4500": ("Theonie Samuels", "20966115", ""),
    "4501": ("Timothy Bradley", "20754861", ""),
    "4502": ("Jason Ta", "", ""),
    "4503": ("Sheryl Campbell", "21641600", ""),
    "4504": ("Charles Parkerson", "21620071", ""),
    "4505": ("Kimberly Wood", "20220082", ""),
    "4506": ("Alex Mclaughlin", "20792923", ""),
    "4507": ("Craig Arch", "", ""),
    "4508": ("Aliyha Nelson", "", ""),
    "4509": ("Brooke Sippel", "", ""),
    "4510": ("travis dilbert", "", ""),
    "4511": ("Caroline Lewis", "", ""),
    "4512": ("lodovico Testori", "20753492", ""),
    "4513": ("Karen Hunter", "21068398", ""),
    "4514": ("Brendan Smith", "20040527", ""),
    "4515": ("Taylor Ebanks Romero", "21665431", ""),
    "4516": ("Alice Bayles", "", ""),
    "4517": ("jeny scott", "", ""),
    "4518": ("Claudia Rodriguez", "21617738", ""),
    "4519": ("Jevon  Pearson", "", ""),
    "4520": ("Shelly Roderick", "21523051", ""),
    "4521": ("Laura Clemens", "21636662", ""),
    "4522": ("Laura Favella", "21636592", ""),
    "4523": ("Lorette Powelll", "21068338", ""),
    "4524": ("Warren Keens", "21595050", ""),
    "4525": ("Timothy Griezitis", "21643618", ""),
    "4526": ("Massiel altagracia  Garrido ortiz", "", ""),
    "4527": ("Rosalie  Sherwood", "21597817", ""),
    "4528": ("Tani Coronado", "", ""),
    "4529": ("Jessica Lumbre", "", ""),
    "4530": ("Eric Lacasse", "", ""),
    "4531": ("Melissa Bridgemohan", "21630686", ""),
    "4532": ("Nordeth  shippy", "", ""),
    "4533": ("Genevieve White", "", ""),
    "4534": ("Charity Putman", "21644472", ""),
    "4535": ("Chase Leacock", "", ""),
    "4536": ("Errol Black", "", ""),
    "4537": ("TRISHANN VERNON", "21082307", ""),
    "4538": ("Laura Friedman", "21068276", ""),
    "4539": ("CAROLINE  MCLEAN", "21069284", ""),
    "4540": ("Jiyu Zhou", "21523294", ""),
    "4541": ("Rachel Masterton", "21067685", ""),
    "4542": ("Rachel Bell", "", ""),
    "4543": ("Rachel Warner", "21067683", ""),
    "4544": ("Karla Jackson", "20009594", ""),
    "4545": ("S Jane Williams", "21591159", ""),
    "4546": ("Macario Jr Gallo", "20793077", ""),
    "4547": ("Megan English", "", ""),
    "4548": ("Ibereayo  Shell", "21591144", ""),
    "4549": ("Michelle  Foxon", "20955647", ""),
    "4550": ("Philippa Walsh", "", ""),
    "4551": ("Martyn Bould", "21630657", ""),
    "4552": ("Martyn Bould", "21630657", ""),
    "4553": ("Shannon Russell", "", ""),
    "4554": ("Thais Rodriguez", "21066699", ""),
    "4555": ("Jewel Robinson", "", ""),
    "4556": ("Shantol Wilson", "20793091", ""),
    "4557": ("John White", "21595048", ""),
    "4558": ("Michael Murphy", "21614555", ""),
    "4559": ("Shawn Knight", "21631660", ""),
    "4560": ("Bradley Rose", "21113903", ""),
    "4561": ("JC TEST", "", ""),
    "4562": ("Kelly Sage", "21620842", ""),
    "4563": ("Rochelle  Morris", "", ""),
    "4564": ("Letitia Solomon", "21597751", ""),
    "4565": ("Marjorie Pitter", "", ""),
    "4566": ("Siobhan Keane", "21591164", ""),
    "4567": ("Lasma Purs", "21597866", ""),
    "4568": ("Matthew Phillips", "21655367", ""),
    "4569": ("Jodie Whittaker", "", ""),
    "4570": ("Kaneesa Ebanks-Wilson", "", ""),
    "4571": ("Amit Sajnani", "", ""),
    "4572": ("Carlos Fredrick", "21641662", ""),
    "4573": ("ROB ELKERTON", "21620148", ""),
    "4574": ("Taneisha  Gordon", "", ""),
    "4575": ("Tawshea  Taylor", "21641662", ""),
    "4576": ("Jose Rivas", "20966082", ""),
    "4577": ("Alvina Barnes", "21639074", ""),
    "4578": ("Garrett Verling", "21630298", ""),
    "4579": ("Jonathan Nunez", "", ""),
    "4580": ("FgGmOqeAhUZwQ lWjFLNPUmYXMCnq", "", ""),
    "4581": ("Andrea Singh", "20955669", ""),
    "4582": ("Ronan Guilfoyle", "", ""),
    "4583": ("Shauna Smith", "20955648", ""),
    "4584": ("Brianna Rodriguez", "21520220", ""),
    "4585": ("Brian Robinson", "21596392", ""),
    "4586": ("Jay Ehrhart", "20955616", ""),
    "4587": ("Patricia Parchment", "", ""),
    "4588": ("Luca Del Ciotto", "", ""),
    "4589": ("Jonas Bush", "20966092", ""),
    "4590": ("David Guilmette", "21206781", ""),
    "4591": ("Linda Laidlaw", "21597843", ""),
    "4592": ("Rebecca Stoner", "21629096", ""),
    "4593": ("Shaniah  Silburn", "21645117", ""),
    "4594": ("Erika Bodden", "20951287", ""),
    "4595": ("Benjamin Austin", "21628553", ""),
    "4596": ("Dora Ebanks", "21637432", ""),
    "4597": ("Alice  Aherne", "21633326", ""),
    "4598": ("Esther Taylor", "21597725", ""),
    "4599": ("Francis Butterworth", "21628569", ""),
    "4600": ("Emma  Santiago", "21639105", ""),
    "4601": ("Katy Bayles", "21634646", ""),
    "4602": ("Camille Josephs", "20955657", ""),
    "4603": ("Stacy McAfee", "", ""),
    "4604": ("Alanna Trundle", "", ""),
    "4605": ("Van Zyl Fourie", "", ""),
    "4606": ("Julie Koutroubis", "", ""),
    "4607": ("chris armistead", "", ""),
    "4608": ("Lotoya Stewart", "20034042", ""),
    "4609": ("Djon Brown", "21597730", ""),
    "4610": ("Kadejah Bodden", "21202404", ""),
    "4611": ("Stephen Aleria", "20721374", ""),
    "4612": ("Fabio  Tancredi", "", ""),
    "4613": ("Furius Whelan", "21591880", ""),
    "4614": ("Casey Santamaria", "21523334", ""),
    "4615": ("Gadiel Piercy", "21641841", ""),
    "4616": ("Kishan Morgan", "", ""),
    "4617": ("SERENE BROWN", "", ""),
    "4618": ("Serene BROWN", "", ""),
    "4619": ("Celeste Charnley", "20951153", ""),
    "4620": ("Eugenie Myrie", "21164571", ""),
    "4621": ("Lauren Butler", "21595047", ""),
    "4622": ("Michelle Joven", "", ""),
    "4623": ("Gerlyn Mae Dela Rosa", "", ""),
    "4624": ("Morgan Albo", "21633349", ""),
    "4625": ("Joanna Marie Altares", "21591239", ""),
    "4626": ("Louisa Mentz", "", ""),
    "4627": ("Laura Johnston", "21596771", ""),
    "4628": ("John Reid", "21591381", ""),
    "4629": ("Pekko Kuusela", "21646771", ""),
    "4630": ("Rodica Pirnau", "21633944", ""),
    "4631": ("Hollie Fenton", "21597886", ""),
    "4632": ("Marcia Mitchell", "21597729", ""),
    "4633": ("Claire Crawford", "", ""),
    "4634": ("Julian Schoefer", "20955606", ""),
    "4635": ("Tara Robinson", "20681151", ""),
    "4636": ("Wolraad Euvrard", "21597727", ""),
    "4637": ("Erik Solomon", "20443361", ""),
    "4638": ("Asaph Scott", "", ""),
    "4639": ("Rebecca Steller", "21591228", ""),
    "4640": ("Wanda Parchment", "", ""),
    "4641": ("Belinda Musonza", "21533395", ""),
    "4642": ("Craig Sulak", "21611671", ""),
    "4643": ("Maxine Gardener", "21523296", ""),
    "4644": ("Michael Yacyshyn", "21597813", ""),
    "4645": ("Lucy McLaughlin", "21630062", ""),
    "4646": ("Matthew Southgate", "", ""),
    "4647": ("Derrick Elliott", "21628566", ""),
    "4648": ("Samantha Scott", "21211173", ""),
    "4649": ("Martin Thomas", "", ""),
    "4650": ("Donna-Kay Smith", "21612493", ""),
    "4651": ("Stefan Charette", "21611428", ""),
    "4652": ("Nicola Logan-Jackson", "21597795", ""),
    "4653": ("Sherry-Suli Hernandez", "21597737", ""),
    "4654": ("Abiann Gayle", "21533407", ""),
    "4655": ("LaShawntae Robinson", "21740257", ""),
    "4656": ("Laura Ebanks", "21630656", ""),
    "4657": ("Kyle Farrington", "21596764", ""),
    "4658": ("Mariah McIntyre", "", ""),
    "4659": ("Ozair Siddiqui", "21647824", ""),
    "4660": ("Ajinkya Shidhore", "21304751", ""),
    "4661": ("Movine Huggan", "", ""),
    "4662": ("oshane christian", "21622027", ""),
    "4663": ("Lageorgia  Miller", "21591241", ""),
    "4664": ("Jonathan Smellie", "21588414", ""),
    "4665": ("Ethan Ebanks", "21588411", ""),
    "4666": ("Alim Harji", "21103228", ""),
    "4667": ("Yvonne Kinglocke", "21655204", ""),
    "4668": ("Jody Ebanks", "21588415", ""),
    "4669": ("Wanda Thompson", "", ""),
    "4670": ("Heidi Wallace", "21103447", ""),
    "4671": ("Margaret Young", "", ""),
    "4672": ("Joan illiams", "", ""),
    "4673": ("Claudette Bell-White", "", ""),
    "4674": ("Joan Williams", "", ""),
    "4675": ("Krista  Drobac", "21596766", ""),
    "4676": ("Meredith Hew", "21131760", ""),
    "4677": ("Thomas Marks", "", ""),
    "4678": ("Vixton  Bennett", "", ""),
    "4679": ("Emily Warden", "21164578", ""),
    "4680": ("Jaryd Moore", "21129109", ""),
    "4681": ("Athena Smith", "20707667", ""),
    "4682": ("Channelle Thomas", "21616754", ""),
    "4683": ("Jayne Lawless", "", ""),
    "4684": ("Jesus Estevez Lescaille", "", ""),
    "4685": ("Lorny  Becker", "", ""),
    "4686": ("Jodeesa Hamilton", "21164583", ""),
    "4687": ("Erin  Bodden", "21636996", ""),
    "4688": ("NAMITHA ABRAHAM", "21597819", ""),
    "4689": ("Marcia Muttoo", "21620838", ""),
    "4690": ("anna kay king", "21533373", ""),
    "4691": ("Sandy Nohemy  Ávila Mendoza", "21591188", ""),
    "4692": ("Gema Brett", "21164584", ""),
    "4693": ("Kameron George", "21164591", ""),
    "4694": ("Judith Furer", "21635775", ""),
    "4695": ("Chandan  Thawrani", "21591394", ""),
    "4696": ("Rachel Boraston", "21633314", ""),
    "4697": ("Eleanor Cook", "", ""),
    "4698": ("Steve Tippetts", "21637474", ""),
    "4699": ("Carla Watler", "21533368", ""),
    "4700": ("Tracey Kirby", "21156212", ""),
    "4701": ("Olivia Shanks", "21164570", ""),
    "4702": ("Sara Mair Doak", "", ""),
    "4703": ("Deand  Quest", "21634156", ""),
    "4704": ("Joana Mclean", "", ""),
    "4705": ("Jason Mears", "21620034", ""),
    "4706": ("Andrea Edwards", "21636076", ""),
    "4707": ("Joy Oremule", "21591201", ""),
    "4708": ("Trisha Dilbert", "21631645", ""),
    "4709": ("Kelsey Still", "21164464", ""),
    "4710": ("Christian Taylor", "21202425", ""),
    "4711": ("David Manouchehri", "21367007", ""),
    "4712": ("Yuri Eden", "21356715", ""),
    "4713": ("Sharon Lamb", "21628491", ""),
    "4714": ("Crystal Heffernan", "21168613", ""),
    "4715": ("Sonja Santor", "21622488", ""),
    "4716": ("Alaaeddine Sahibi", "", ""),
    "4717": ("Jessica Powell", "21620031", ""),
    "4718": ("Anya Park", "21596394", ""),
    "4719": ("Kelley  Ebanks", "", ""),
    "4720": ("Jozef Vogel", "21597870", ""),
    "4721": ("Aileen Hunn", "21533718", ""),
    "4722": ("Daniel Tyndale", "21523322", ""),
    "4723": ("Sigrid Menschaart", "", ""),
    "4724": ("Melinda Mclean", "21211249", ""),
    "4725": ("Sheneak Neil", "", ""),
    "4726": ("Jordon Forbes", "", ""),
    "4727": ("Daisha Coleman", "21591253", ""),
    "4728": ("jody rhodes", "21523308", ""),
    "4729": ("Scott Lewis", "21176108", ""),
    "4730": ("Andiara  Reis Prates", "21533404", ""),
    "4731": ("Gilmer G Bradshaw", "21630055", ""),
    "4732": ("Richard  Maitland", "21631613", ""),
    "4733": ("Liliana Forbes", "21628542", ""),
    "4734": ("Todd Dillabough", "", ""),
    "4735": ("Donna-Dene  Ramirez Rosales", "", ""),
    "4736": ("Jose  Venner", "21591177", ""),
    "4737": ("Jennifer Sangaroonthong", "", ""),
    "4738": ("Jennifer McLemore", "21180621", ""),
    "4739": ("Mark Robson", "21591200", ""),
    "4740": ("Ernest Henry", "21523321", ""),
    "4741": ("Colin Martin", "", ""),
    "4742": ("Manisha Gupta", "21591254", ""),
    "4743": ("Shashank Mathur", "21638538", ""),
    "4744": ("Ian Smith", "21628564", ""),
    "4745": ("Keira Christian", "21235267", ""),
    "4746": ("Claire Murphy", "21235283", ""),
    "4747": ("Ricardo Bodington", "", ""),
    "4748": ("Randa mcpherson", "", ""),
    "4749": ("Savitr  Sastri BV", "21211258", ""),
    "4750": ("Tammi Barnes", "", ""),
    "4751": ("Falisha  Munroe", "21597935", ""),
    "4752": ("Jaundre Jansen Van Rensburg", "21235275", ""),
    "4753": ("Regan Cairns", "21628490", ""),
    "4754": ("Tim O'Sullivan", "21523041", ""),
    "4755": ("Benedict Havey", "", ""),
    "4756": ("Henry Hill", "21628560", ""),
    "4757": ("Yvonne Plamondon", "21235276", ""),
    "4758": ("Maria Bragina", "", ""),
    "4759": ("Sara Jan", "21520248", ""),
    "4760": ("Praajakta Pargaonkar", "", ""),
    "4761": ("Maria Parra", "", ""),
    "4762": ("Eva Gosciniak", "21666049", ""),
    "4763": ("erika aufiero", "21639835", ""),
    "4764": ("Apollinia Lavia", "21356784", ""),
    "4765": ("Ethania Gayle", "21631871", ""),
    "4766": ("Britaney Thompson", "21356768", ""),
    "4767": ("Tashanta Solomon", "21597745", ""),
    "4768": ("Paul Davies", "21591189", ""),
    "4769": ("NiiAkwei General", "", ""),
    "4770": ("Claud Lewis", "21591391", ""),
    "4771": ("JINGHE HAN", "", ""),
    "4772": ("Alexandra Anitoaie", "21641053", ""),
    "4773": ("by Kadejah B", "", ""),
    "4774": ("Azuka Obi", "", ""),
    "4775": ("Yichen Yue", "21235258", ""),
    "4776": ("Thakurdai  Mohamed", "", ""),
    "4777": ("LAURA SILVAN", "21356745", ""),
    "4778": ("Jorge Manso Oliva", "21635894", ""),
    "4779": ("Brian Daly", "", ""),
    "4780": ("Kaitlyn McGee", "", ""),
    "4781": ("Tamanna Keir", "21235238", ""),
    "4782": ("ANNA MATTHEWS", "", ""),
    "4783": ("Joshua Hecht", "21591378", ""),
    "4784": ("Rachel Pan", "", ""),
    "4785": ("Julie Hughes", "21356756", ""),
    "4786": ("Rick Schlesinger", "", ""),
    "4787": ("Abigail Stoddart", "21596396", ""),
    "4788": ("Emma Linney", "21638675", ""),
    "4789": ("christine Connor", "", ""),
    "4790": ("Priscilla Shire(Richard McAree)", "21725108", ""),
    "4791": ("James Cooper", "21700212", ""),
    "4792": ("Bridgette von Gerhardt", "21640007", ""),
    "4793": ("Michaela Walker", "", ""),
    "4794": ("Hussain Alhasani", "", ""),
    "4795": ("Charmaigne De las Alas", "21597738", ""),
    "4796": ("Linh Vu", "21356739", ""),
    "4797": ("Andrea Haughton", "21638229", ""),
    "4798": ("KENNETH MITCHELL", "", ""),
    "4799": ("Courtney Duval", "20217191", ""),
    "4800": ("Sheryl  Walton-Williams", "", ""),
    "4801": ("Eric Bergstrom", "21588416", ""),
    "4802": ("Ronalee Dobroslavic", "21629071", ""),
    "4803": ("Karla Montes", "21588410", ""),
    "4804": ("michelle Biase", "", ""),
    "4805": ("Brandon Ramsay", "21356782", ""),
    "4806": ("Aoife Molloy", "21639788", ""),
    "4807": ("Yasmin Mitchell", "21356724", ""),
    "4808": ("Debbie Facey", "", ""),
    "4809": ("Marissa Mejia", "21621590", ""),
    "4810": ("Anthony Fraser", "21326276", ""),
    "4811": ("Kim McLaughlin", "21356750", ""),
    "4812": ("Aniruddha Kayal", "", ""),
    "4813": ("alec cox", "21745619", ""),
    "4814": ("Wendy Poloney", "21588418", ""),
    "4815": ("Kate Ryley", "21588407", ""),
    "4816": ("Ashley  Smith", "20073019", ""),
    "4817": ("Alejandro  Gaona", "21643379", ""),
    "4818": ("James Macallum", "21651525", ""),
    "4819": ("Jose Rochez", "21597867", ""),
    "4820": ("Ronan O'Shea", "", ""),
    "4821": ("Catriona  Lefkos", "21751812", ""),
    "4822": ("lakiva reid", "", ""),
    "4823": ("Wayne Lindsey", "", ""),
    "4824": ("Samantha Jackson", "21690250", ""),
    "4825": ("Wendy  Foreman", "21597723", ""),
    "4826": ("Aravind Nair", "21356783", ""),
    "4827": ("Clare-Louise McGrath", "", ""),
    "4828": ("Ricardo Smith", "", ""),
    "4829": ("Ricardo Smith", "", ""),
    "4830": ("Mariam Yusuf", "21591229", ""),
    "4831": ("Kathryn Cozzens", "21356755", ""),
    "4832": ("Phillip Ebanks", "21620177", ""),
    "4833": ("Andrew Linford", "21356791", ""),
    "4834": ("Quincy Cusack", "21614550", ""),
    "4835": ("Harris Caribbean", "", ""),
    "4836": ("Delphinia  Ebanks", "", ""),
    "4837": ("Kevin Moran", "", ""),
    "4838": ("Shelley Wilkinson", "21591219", ""),
    "4839": ("Adam Vanicek", "21626818", ""),
    "4840": ("Janice Fraser", "21484046", ""),
    "4841": ("Darren Burke", "21591182", ""),
    "4842": ("Linda Diaz", "21523365", ""),
    "4843": ("Fausto Bogran", "21616659", ""),
    "4844": ("JOHAN OTTO", "21591380", ""),
    "4845": ("Christopher Rivers", "", ""),
    "4846": ("WANDA ADONZA HARRISON", "21597724", ""),
    "4847": ("shannon williams", "21520259", ""),
    "4848": ("Bharat Arora", "", ""),
    "4849": ("Kanika Banerjee", "21523377", ""),
    "4850": ("Divya Wahi", "21484043", ""),
    "4851": ("Rexford Tutor", "21659000", ""),
    "4852": ("john friesen", "21591259", ""),
    "4853": ("Pauline Small", "", ""),
    "4854": ("Juliet Carty", "", ""),
    "4855": ("Maryke Nieuwoudt", "21643276", ""),
    "4856": ("Jinghe Han", "", ""),
    "4857": ("Davene Watson", "21621553", ""),
    "4858": ("Earl Peterson", "21484041", ""),
    "4859": ("Sarah Ryan", "21520258", ""),
    "4860": ("Tomi Owolabi", "", ""),
    "4861": ("Hill Smith", "", ""),
    "4862": ("George Bashforth", "21620068", ""),
    "4863": ("Ramon  Rojas Paredes", "21683055", ""),
    "4864": ("Brian Nounev", "", ""),
    "4865": ("Carlos Aybar", "21596390", ""),
    "4866": ("Danielle Greaves", "21484056", ""),
    "4867": ("Lydia Watling", "21523364", ""),
    "4868": ("Lucille Shelley", "21523352", ""),
    "4869": ("Casandra Bush", "", ""),
    "4870": ("Bo Hansson", "", ""),
    "4871": ("Lynne Wester", "21523371", ""),
    "4872": ("Matthew Evans", "21523359", ""),
    "4873": ("Andre Stephano", "21484057", ""),
    "4874": ("Zoe Du Bois", "21520246", ""),
    "4875": ("Sherifa Lampart", "", ""),
    "4876": ("Yamilet Clarke", "21520253", ""),
    "4877": ("Wrendon Timothy", "21523032", ""),
    "4878": ("Sandy  Hermiston", "", ""),
    "4879": ("Zayde Garcia", "", ""),
    "4880": ("Christopher Ali", "21484063", ""),
    "4881": ("Gary Hill", "", ""),
    "4882": ("Terry y Gordon Scherer", "21629460", ""),
    "4883": ("Randall Fisher", "21591165", ""),
    "4884": ("Gary Allen", "21620080", ""),
    "4885": ("Anita Welds", "21533390", ""),
    "4886": ("Michael Clark", "", ""),
    "4887": ("Stephanie Hampson", "", ""),
    "4888": ("Courtenay Wolfe", "", ""),
    "4889": ("Amanda Connor", "21484059", ""),
    "4890": ("Karen  Gardener", "21523302", ""),
    "4891": ("Norjehan Hextall", "21520283", ""),
    "4892": ("Stewart Southey", "", ""),
    "4893": ("Soren Kolbe", "21520238", ""),
    "4894": ("Jonetta Mosley-Matchett", "21597874", ""),
    "4895": ("Jodian  McLeod", "", ""),
    "4896": ("Felicia Connor", "20191731", ""),
    "4897": ("Winsome Martinez", "", ""),
    "4898": ("Russell Goldenberg", "21611435", ""),
    "4899": ("Jen Strangeway", "21523345", ""),
    "4900": ("Derek Byrne", "21591225", ""),
    "4901": ("Lydia Datzreiter", "", ""),
    "4902": ("Tiffani Straker", "21520237", ""),
    "4903": ("Michael Drobac", "", ""),
    "4904": ("Amy Slee", "", ""),
    "4905": ("Roberta Anderson-Southam", "21520267", ""),
    "4906": ("Michael Crothers", "21520284", ""),
    "4907": ("Thomas Haynes", "21591170", ""),
    "4908": ("Louise Cowley", "21611447", ""),
    "4909": ("Leidy Borjas", "21636126", ""),
    "4910": ("Yislianys  Silva", "", ""),
    "4911": ("De-anna Trowers", "21484048", ""),
    "4912": ("Nicholas Ricketts", "21723978", ""),
    "4913": ("Gillian Lynch", "21591157", ""),
    "4914": ("Admiral Anderson", "21533399", ""),
    "4915": ("Monique  Brown", "21597812", ""),
    "4916": ("Angela Pretorius", "21523340", ""),
    "4917": ("Laetitia  Bush", "21620157", ""),
    "4918": ("Nick Reyes", "", ""),
    "4919": ("Dusty Norman", "21484042", ""),
    "4920": ("Richard Lawtey", "21520275", ""),
    "4921": ("Brett Carrington", "", ""),
    "4922": ("Paul Greb", "21591153", ""),
    "4923": ("Natasha Chan", "20550004", ""),
    "4924": ("Anthony Esposito", "20505169", ""),
    "4925": ("Moesha Ramsay-Howell", "21725138", ""),
    "4926": ("Andrea y Chris Duty", "21533487", ""),
    "4927": ("William Inniss", "21520222", ""),
    "4928": ("Michelle Majid", "21634519", ""),
    "4929": ("Will Kerr", "20003220", ""),
    "4930": ("Robert Helina", "21633465", ""),
    "4931": ("Meila Johnson", "20595217", ""),
    "4932": ("Karen Chatburn", "21523339", ""),
    "4933": ("Yacely Parchment", "21638291", ""),
    "4934": ("Charlotte  Wright", "20004886", ""),
    "4935": ("Marlie Du toit", "20003581", ""),
    "4936": ("Samuel Antonio  Martin Allen", "21523053", ""),
    "4937": ("Lisa Rankin", "", ""),
    "4938": ("Maureen Robinson", "", ""),
    "4939": ("Don Ebanks", "21659461", ""),
    "4940": ("Deandra  Ford", "20073670", ""),
    "4941": ("Dewald Cloete", "21484044", ""),
    "4942": ("Valerie Anderson", "21520229", ""),
    "4943": ("Sai Gajjala", "", ""),
    "4944": ("Kirsten  Cellier", "21523108", ""),
    "4945": ("Glenn Kennedy", "21597895", ""),
    "4946": ("Andrew Bowie", "21533400", ""),
    "4947": ("Natasha Miglecz", "21591154", ""),
    "4948": ("Joel Edwards", "21523351", ""),
    "4949": ("Sonny Powell", "21591158", ""),
    "4950": ("Grisel  B Ebanks", "", ""),
    "4951": ("Deborah  Roberts", "21597944", ""),
    "4952": ("Nechelle  Calimlim", "21597796", ""),
    "4953": ("Frances Robinson", "", ""),
    "4954": ("Shelley Addleson", "21640962", ""),
    "4955": ("Cheyenna Calderon-Hoaglund", "21532640", ""),
    "4956": ("Emisha Powell", "21669462", ""),
    "4957": ("Catherine Valiquette", "21533366", ""),
    "4958": ("Shurnee Dawkins", "21523069", ""),
    "4959": ("Sarah Dinyer", "21597752", ""),
    "4960": ("Dave Tibbetts", "21645356", ""),
    "4961": ("Gregory Vasic", "21523316", ""),
    "4962": ("Adelle Myers", "21596402", ""),
    "4963": ("Dorothy Rau", "21597937", ""),
    "4964": ("Nadine Abou Nohra", "21523070", ""),
    "4965": ("Sherine Thomas", "21597740", ""),
    "4966": ("Sheenah Hislop", "21523093", ""),
    "4967": ("Nordra Walcott", "21523061", ""),
    "4968": ("RAMEL CARLOS", "21632636", ""),
    "4969": ("Damian Kielczewski", "21591387", ""),
    "4970": ("Samantha Fulton", "21597801", ""),
    "4971": ("Cassius Ebanks", "", ""),
    "4972": ("Lisa Hindle", "21614561", ""),
    "4973": ("shannon williams", "21520259", ""),
    "4974": ("Kidan  Brooks", "21597862", ""),
    "4975": ("Kavita Maharaj-Alexander", "21597798", ""),
    "4976": ("fakry jaffar", "21629168", ""),
    "4977": ("Caitlin Dunne", "21639877", ""),
    "4978": ("Christopher Conolly", "", ""),
    "4979": ("Christopher  McTaggart", "20216667", ""),
    "4980": ("Khaya Dube", "21741358", ""),
    "4981": ("Cheyenne Dixon", "21611612", ""),
    "4982": ("Zoe Day", "", ""),
    "4983": ("Brian Robinson", "21596392", ""),
    "4984": ("Myrna  Van der zee", "", ""),
    "4985": ("Romaine  Little", "21633791", ""),
    "4986": ("Aurore Grasset", "21595846", ""),
    "4987": ("Jayda Powery", "21597741", ""),
    "4988": ("Angela Choi", "21596395", ""),
    "4989": ("Rodney Alison", "21614554", ""),
    "4990": ("Sharon Looney", "21611549", ""),
    "4991": ("Nour Khaleq", "", ""),
    "4992": ("Simon Smith", "21611430", ""),
    "4993": ("beverly edgington", "21596393", ""),
    "4994": ("Triniti Dixon", "21591218", ""),
    "4995": ("Steven Sokohl", "21643267", ""),
    "4996": ("Memory Sagud", "", ""),
    "4997": ("Robert  DaCosta", "", ""),
    "4998": ("JASMINE  EBANKS", "", ""),
    "4999": ("Sylvester Dube", "", ""),
    "5000": ("Mail Boxes Etc.", "20004944", ""),
    "5001": ("Neville Boon", "21597860", ""),
    "5002": ("Dail Davis", "21597739", ""),
    "5003": ("Sharon n Ray Brown", "21597754", ""),
    "5004": ("Cherry Gonzales", "", ""),
    "5005": ("Cherry Gonzales", "", ""),
    "5006": ("Javier Mckenzie", "21328826", ""),
    "5007": ("Andrene Levy", "", ""),
    "5008": ("Chelsea Green", "20003556", ""),
    "5009": ("Zane Kuttner", "21597719", ""),
    "5010": ("Kieran Honey", "", ""),
    "5011": ("shane dwyer", "21611479", ""),
    "5012": ("Ines Maria Mendez", "21628502", ""),
    "5013": ("phil cassingham", "", ""),
    "5014": ("Beverley Speirs", "", ""),
    "5015": ("Timothy Derksen", "21597732", ""),
    "5016": ("Lehia  Bryan", "", ""),
    "5017": ("Zainab Fatima", "21597720", ""),
    "5018": ("Daria Zamkova", "21615185", ""),
    "5019": ("Paulinda MendozaWilliams", "21614570", ""),
    "5020": ("Gina Cox", "21611565", ""),
    "5021": ("Nylah Rampersad", "", ""),
    "5022": ("Sebastian Goerlich", "21611670", ""),
    "5023": ("Neil Dempsey", "21620019", ""),
    "5024": ("Bransens Bransens", "", ""),
    "5025": ("Lana Kelly", "21611574", ""),
    "5026": ("Carlene Newell", "21611722", ""),
    "5027": ("EMILY MOORE", "", ""),
    "5028": ("Alex Shuker", "", ""),
    "5029": ("Tayla Niemand", "21620008", ""),
    "5030": ("Dahlia Evans", "21680900", ""),
    "5031": ("Katie Alpers", "21611563", ""),
    "5032": ("Brian Patterson", "21611451", ""),
    "5033": ("Bernadette Watler", "21612499", ""),
    "5034": ("Nuvia  Garcia", "21611433", ""),
    "5035": ("Michael Baulk", "21611436", ""),
    "5036": ("Tamara Worboys", "", ""),
    "5037": ("Mina Whorms", "21622494", ""),
    "5038": ("Mary Burke-Cameron", "21614560", ""),
    "5039": ("Anna MacRae", "21611685", ""),
    "5040": ("Antonio Hafner", "21614613", ""),
    "5041": ("Victoria Iacoviello", "21611431", ""),
    "5042": ("Shannon Hydes", "", ""),
    "5043": ("Michelle Lewis", "21634890", ""),
    "5044": ("Antonia Aguayo", "21611491", ""),
    "5045": ("Aniya Anderson", "", ""),
    "5046": ("Venessa Ebanks", "21639678", ""),
    "5047": ("Jenni Huys", "21611560", ""),
    "5048": ("Ashley Bowers", "21615186", ""),
    "5049": ("Beverly  Shuford", "21614597", ""),
    "5050": ("Hosea Harris", "", ""),
    "5051": ("Brian Tang", "", ""),
    "5052": ("Juan Pablo (JP) Fukushi", "21630059", ""),
    "5053": ("Jocelyn  Fuller", "21614578", ""),
    "5054": ("Leroy Jordan", "21614605", ""),
    "5055": ("Graeme Love", "21628565", ""),
    "5056": ("Sunshine Carag", "21614549", ""),
    "5057": ("Liana Etsebeth", "", ""),
    "5058": ("Kaitlyn Thompson", "", ""),
    "5059": ("Michael Ebanks", "21638525", ""),
    "5060": ("Rebekah Clark", "21620265", ""),
    "5061": ("Jonalie Pepito", "21620027", ""),
    "5062": ("Ryan Harding", "20063839", ""),
    "5063": ("Francine Bloomfield", "21614590", ""),
    "5064": ("Stacey Clark", "21614563", ""),
    "5065": ("Trisha Anthony", "", ""),
    "5066": ("Jonathan Moffatt", "21614576", ""),
    "5067": ("Christopher Arseneau", "", ""),
    "5068": ("David  Watler", "21614610", ""),
    "5069": ("Tanesha Hall", "21614547", ""),
    "5070": ("Abena Agard", "", ""),
    "5071": ("Patrick Keenan", "21616729", ""),
    "5072": ("Charl Grobler", "21616739", ""),
    "5073": ("Layman Scott III", "21614551", ""),
    "5074": ("Carson Yates", "", ""),
    "5075": ("Philippe  Deslandes", "21620016", ""),
    "5076": ("LINAH MHOSHIWA", "21616669", ""),
    "5077": ("Kristy Childers", "21616671", ""),
    "5078": ("Guiesha Smith", "21669356", ""),
    "5079": ("Deniss Suhanovs", "21616718", ""),
    "5080": ("Elda Fuentes", "21631667", ""),
    "5081": ("Jordana Ebanks-Hurlston", "", ""),
    "5082": ("Tashalee Christie", "21727844", ""),
    "5083": ("Deandre Simpson", "21616720", ""),
    "5084": ("Rahshan Halstead", "21628489", ""),
    "5085": ("Maria Antonia Borget", "21617754", ""),
    "5086": ("Amanda Kong", "21616724", ""),
    "5087": ("Michael Scott", "21616663", ""),
    "5088": ("Andre Johnson", "21620077", ""),
    "5089": ("Maria Tom-Pack", "21617750", ""),
    "5090": ("Benson  Brooks", "", ""),
    "5091": ("Marlon Joseph", "21618679", ""),
    "5092": ("Keylie Ebanks", "", ""),
    "5093": ("mayfran corbalan", "21617733", ""),
    "5094": ("neydi campbell", "", ""),
    "5095": ("Bianca Gwekwete", "21618676", ""),
    "5096": ("shaneil smith", "21620009", ""),
    "5097": ("Wayne  Bodden", "21620169", ""),
    "5098": ("Linda Key", "21620079", ""),
    "5099": ("Niki O_x0019_ Hara", "", ""),
    "5100": ("Anika von Gerhardt", "21618672", ""),
    "5101": ("Khylar Miller", "21654094", ""),
    "5102": ("Dallas Neatham", "21618677", ""),
    "5103": ("Avatar Mathura", "21618675", ""),
    "5104": ("Lethea Welcome", "21620025", ""),
    "5105": ("Laura McGeever", "21622548", ""),
    "5106": ("Anneleise Richards", "21629167", ""),
    "5107": ("David Brown", "21619152", ""),
    "5108": ("CAVEN  BOOTHE", "", ""),
    "5109": ("Forde  Pierson", "", ""),
    "5110": ("Tyler Eaton", "", ""),
    "5111": ("Alcot Gould", "", ""),
    "5112": ("Rodolfo Membreno", "", ""),
    "5113": ("Olisiea Blake", "21621577", ""),
    "5114": ("Shunell  Hoyte", "21620840", ""),
    "5115": ("Nadia Rivers", "21622028", ""),
    "5116": ("Ana Chisholm", "21622038", ""),
    "5117": ("Richard Caruso", "21622024", ""),
    "5118": ("Susan Redding", "21621545", ""),
    "5119": ("DAEQUAN ISAACS", "", ""),
    "5120": ("James McIntyre", "21622030", ""),
    "5121": ("Louise Baulk", "21622512", ""),
    "5122": ("Brad Huntington", "21622036", ""),
    "5123": ("Verla Joy Basdeo", "21622022", ""),
    "5124": ("denise claux", "21622496", ""),
    "5125": ("alan cartolano", "", ""),
    "5126": ("Paulo Cirulli", "21621548", ""),
    "5127": ("Joseph  Watler Jr", "21638215", ""),
    "5128": ("Susan  Watson", "21591159", ""),
    "5129": ("Sue watson", "", ""),
    "5130": ("Jessica Wright", "", ""),
    "5131": ("Lorna Bent", "21628541", ""),
    "5132": ("Fernando Mendes", "21624342", ""),
    "5133": ("Jordan  Lamoral", "21624036", ""),
    "5134": ("Hugh Lockwood", "21624350", ""),
    "5135": ("Beverley Hodkin", "21624336", ""),
    "5136": ("Anetia  DaCosta", "21629463", ""),
    "5137": ("otis myles", "21630357", ""),
    "5138": ("Diane Vasic", "21625339", ""),
    "5139": ("Adrian  Murenzi", "21624321", ""),
    "5140": ("Sarah Hobbs", "21642425", ""),
    "5141": ("Jesse Prince", "21626804", ""),
    "5142": ("Ian MacFarlaine", "", ""),
    "5143": ("Emily Saunders", "21625338", ""),
    "5144": ("Dorith Brenda Ebanks", "21627722", ""),
    "5145": ("Brooke Powell", "21638544", ""),
    "5146": ("jemar vernon", "", ""),
    "5147": ("Roger  Watler", "21625003", ""),
    "5148": ("Patricia Stoll", "", ""),
    "5149": ("MARIA PADUA", "", ""),
    "5150": ("Katrina  Watler", "21625004", ""),
    "5151": ("Victorino Dela Cruz", "21626800", ""),
    "5152": ("Lucy  Gardner", "", ""),
    "5153": ("Alan Dodds", "21626820", ""),
    "5154": ("Sidiam Marquez", "21626798", ""),
    "5155": ("Jameal Welcome", "21628559", ""),
    "5156": ("Luis Daniel  Ebanks", "", ""),
    "5157": ("Donna Molina", "21633347", ""),
    "5158": ("REMYA  K REGHUNADHAN", "", ""),
    "5159": ("Andrade McLaughlin", "21637024", ""),
    "5160": ("Khylar Miller", "21626801", ""),
    "5161": ("Suntera Limited", "21714001", ""),
    "5162": ("Junie  McField", "21626799", ""),
    "5163": ("Annette  Francis", "21633346", ""),
    "5164": ("Jonathan Turnham", "21628581", ""),
    "5165": ("Denis Landry", "21629464", ""),
    "5166": ("Justin Ladzinski", "21630662", ""),
    "5167": ("holly Keating", "", ""),
    "5168": ("Jenny Crawshay", "21633325", ""),
    "5169": ("Jane Scaletta", "21630347", ""),
    "5170": ("BETTY BROWN", "21630679", ""),
    "5171": ("Ian De Villiers", "21628554", ""),
    "5172": ("Abigail  Reid", "21628101", ""),
    "5173": ("Olessya Moretti", "21630674", ""),
    "5174": ("cody culbert", "", ""),
    "5175": ("matthew davidson", "", ""),
    "5176": ("Christopher Paul", "21630676", ""),
    "5177": ("Kenola Ricot-Williams", "21633360", ""),
    "5178": ("David Burtton", "21629063", ""),
    "5179": ("Lauren Nelson", "21630655", ""),
    "5180": ("William Burns", "21630414", ""),
    "5181": ("Danette Mclaughlin", "20012776", ""),
    "5182": ("Ana Casildo", "21633359", ""),
    "5183": ("Scott Towler", "21633339", ""),
    "5184": ("Lisa Beauchamp", "21633344", ""),
    "5185": ("janardhanan ajaikumar", "", ""),
    "5186": ("Victoria Haynes", "", ""),
    "5187": ("Rob Barton", "21633343", ""),
    "5188": ("jianqiao xu", "", ""),
    "5189": ("Christine Wheatley", "21633334", ""),
    "5190": ("ZUNILDA  GORDON PENA", "", ""),
    "5191": ("MARK  WILSON", "", ""),
    "5192": ("John Dabbs", "21633332", ""),
    "5193": ("Nicky Sinclair", "21637365", ""),
    "5194": ("Marian Foster", "", ""),
    "5195": ("Jason Eastman", "21635054", ""),
    "5196": ("Mitchelle Muhlanga", "21636225", ""),
    "5197": ("ELAINE MANOR", "", ""),
    "5198": ("Kisha Ebanks", "21653471", ""),
    "5199": ("Romaine Little", "21633791", ""),
    "5200": ("Hardeep Riyat", "", ""),
    "5201": ("Terryann A Robinson", "21634250", ""),
    "5202": ("Kristan Cooper", "", ""),
    "5203": ("Jacky Rowland", "21638747", ""),
    "5204": ("Vanessas Rankine", "21639370", ""),
    "5205": ("Quynh  Pham", "", ""),
    "5206": ("Paige Waller", "", ""),
    "5207": ("RICHARD SIOSON", "21636265", ""),
    "5208": ("Adis Kevorkian", "21636260", ""),
    "5209": ("Jerry Dee", "", ""),
    "5210": ("Trecia Russell-Jones", "", ""),
    "5211": ("Niall Gallagher", "21634983", ""),
    "5212": ("Mark Magliocco", "", ""),
    "5213": ("Carrie Lyn-Bodden", "21655303", ""),
    "5214": ("Supipi Gamage", "21637309", ""),
    "5215": ("Gregg Arnold", "20617277", ""),
    "5216": ("Jonathan Richardson", "21686573", ""),
    "5217": ("Chad Phillipps", "", ""),
    "5218": ("Xavi Rivera", "21637908", ""),
    "5219": ("Raheem Webb", "21637861", ""),
    "5220": ("Aravindhan Jothisekaran", "", ""),
    "5221": ("Takudzwa Kathleen Zinyakatira", "21641947", ""),
    "5222": ("Courtney Lindsay", "", ""),
    "5223": ("William  Powery", "21636253", ""),
    "5224": ("Linbern Eden", "21640263", ""),
    "5225": ("Amber Primrose", "20607761", ""),
    "5226": ("Natasha Luchies", "21643600", ""),
    "5227": ("Anouska  Reid", "21646072", ""),
    "5228": ("Alessandra Petrizzelli", "21638527", ""),
    "5229": ("Johan Mannerheim", "21635700", ""),
    "5230": ("David Pattaway", "21639490", ""),
    "5231": ("FITZGERALD DAVIS", "21677726", ""),
    "5232": ("Jada Elliott", "21639494", ""),
    "5233": ("Taimoon Stewart", "21639487", ""),
    "5234": ("Lianet Hydes", "", ""),
    "5235": ("James Macfee", "21642408", ""),
    "5236": ("Tayla Richmond", "20117192", ""),
    "5237": ("Keisha Johnson", "21640612", ""),
    "5238": ("Iglesias  Shepherd", "", ""),
    "5239": ("Sharon Alger", "21640182", ""),
    "5240": ("Carlo Nikko Calarion", "21637086", ""),
    "5241": ("Kent Green", "21730552", ""),
    "5242": ("Rachel Costello", "21639739", ""),
    "5243": ("Northview Services Ltd.", "21641399", ""),
    "5244": ("Wendy Johnston", "21651048", ""),
    "5245": ("Leisa Welcome", "21643058", ""),
    "5246": ("Evelin Ritch", "21642338", ""),
    "5247": ("Ashlyn Goubault", "", ""),
    "5248": ("ZOVNITH AMÉRICA WATCHMAN", "", ""),
    "5249": ("Carol Jurchison", "21641032", ""),
    "5250": ("Jada Brown", "", ""),
    "5251": ("Jenna Munruddin", "21642080", ""),
    "5252": ("Amanda Gray", "21642327", ""),
    "5253": ("Mitchell Demeter", "", ""),
    "5254": ("Dearbhaile O'Kelly", "21642655", ""),
    "5255": ("Michael Evans", "21640902", ""),
    "5256": ("Hazel O Brien", "21639745", ""),
    "5257": ("Vamelyn Clarke", "21642667", ""),
    "5258": ("Sangeeta Awasthi", "", ""),
    "5259": ("Dannelle  Gordon", "21644606", ""),
    "5260": ("Lisa Bortolotto", "", ""),
    "5261": ("Anna Cummings", "", ""),
    "5262": ("Elad Belaish", "21643083", ""),
    "5263": ("Patrick  Comrie", "21643355", ""),
    "5264": ("Andrew Campbell", "21649116", ""),
    "5265": ("Carina Sebastian", "21644661", ""),
    "5266": ("David  Oakley", "21646118", ""),
    "5267": ("Greg Melehov", "21646105", ""),
    "5268": ("Peter Williams", "21643501", ""),
    "5269": ("Nadine McBean", "21647060", ""),
    "5270": ("Alma Smith", "21605265", ""),
    "5271": ("Erika Wilson", "", ""),
    "5272": ("Paul Peat", "21644937", ""),
    "5273": ("Patrick Ramirez", "21645400", ""),
    "5274": ("Samuel Bachet", "21675951", ""),
    "5275": ("Lindsay Thompson", "", ""),
    "5276": ("Jordan Chisholm", "21649506", ""),
    "5277": ("Willem Pienaar", "21650285", ""),
    "5278": ("Health City", "21691634", ""),
    "5279": ("sistemas mbe", "", ""),
    "5280": ("Roberta Hawthorne", "21634971", ""),
    "5281": ("Charles Lewinson Jr.", "21645688", ""),
    "5282": ("Isobel Tomkinson", "", ""),
    "5283": ("Shanae Lawrence", "", ""),
    "5284": ("Trisha Hennings-Jackson", "21647083", ""),
    "5285": ("Carl Hawkes", "21645710", ""),
    "5286": ("Stuart Miller", "", ""),
    "5287": ("Jodie Foster", "", ""),
    "5288": ("Hendrik Lombard", "", ""),
    "5289": ("Marco Angelo Perez", "", ""),
    "5290": ("Kemar  Scarlett", "", ""),
    "5291": ("Jacqueline  Bryan", "21647793", ""),
    "5292": ("Ashley Roque", "21645883", ""),
    "5293": ("Jonathan Evans", "21648542", ""),
    "5294": ("Kelly Anne Wright", "21648954", ""),
    "5295": ("Maxwell Maxwell Ritch", "20070940", ""),
    "5296": ("DEAN HOWELL", "", ""),
    "5297": ("Anel Marais", "21650972", ""),
    "5298": ("Sarah Martin", "21649909", ""),
    "5299": ("Sefu Bernard", "20051586", ""),
    "5300": ("Evelyn Banks", "", ""),
    "5301": ("Neesa Wilson", "", ""),
    "5302": ("Dhalia  Francis", "", ""),
    "5303": ("Michael Ryan", "21639873", ""),
    "5304": ("Thalia Rego Ramos", "21650712", ""),
    "5305": ("The Legal Godfairy", "", ""),
    "5306": ("Simone  Wright", "21679900", ""),
    "5307": ("Jesper Kristensen", "20084037", ""),
    "5308": ("Melissa or Mike McWatt", "", ""),
    "5309": ("Richard Shugrue", "21649398", ""),
    "5310": ("Nataline Ramoon", "", ""),
    "5311": ("Arthur Dzaghgouni", "", ""),
    "5312": ("Michael Hydes", "21653032", ""),
    "5313": ("PETURA PETERKIN", "", ""),
    "5314": ("Paul James", "21654465", ""),
    "5315": ("Kayla Stewart", "", ""),
    "5316": ("Kolton Miller", "21654475", ""),
    "5317": ("Derrick Harper", "21649186", ""),
    "5318": ("Jessica Crawford", "21655109", ""),
    "5319": ("Anneka Greenway", "21655531", ""),
    "5320": ("Barbara Wolf", "21645056", ""),
    "5321": ("Daniel Eaton", "", ""),
    "5322": ("Wesley Robinson", "20828038", ""),
    "5323": ("Vanessa Bodden", "21656486", ""),
    "5324": ("Tansi Churchill", "20084738", ""),
    "5325": ("Thian-Hok Liem", "", ""),
    "5326": ("Yana Yagupolsky", "21650059", ""),
    "5327": ("Dimitri Stocker", "", ""),
    "5328": ("Alexandra Potts", "20840913", ""),
    "5329": ("Daryn Schwulst", "21655882", ""),
    "5330": ("Angelo Solimando", "20550360", ""),
    "5331": ("Corey Randolph", "21657970", ""),
    "5332": ("Inga Mayorquin", "21430718", ""),
    "5333": ("MARVIN HIRSH", "20693239", ""),
    "5334": ("Nicholas Gargaro", "21686014", ""),
    "5335": ("Kevin Copes", "", ""),
    "5336": ("Farrah Sbaiti", "21640893", ""),
    "5337": ("Stacie Ann  Chambers", "", ""),
    "5338": ("Hayley Reid", "21658049", ""),
    "5339": ("Marco Olivier", "21659139", ""),
    "5340": ("Joseph Barker-Willis", "21658465", ""),
    "5341": ("kerstin nickchen", "", ""),
    "5342": ("Laura OByrne", "21647533", ""),
    "5343": ("Martin And Ellamay Oakley", "21660185", ""),
    "5344": ("susan sheridan", "", ""),
    "5345": ("Jack  Fleming", "21661177", ""),
    "5346": ("Stephanie  Wright", "", ""),
    "5347": ("Sergio Rodiles", "", ""),
    "5348": ("Ryan Berckmans", "21661691", ""),
    "5349": ("Jennifer Steinberg", "21661517", ""),
    "5350": ("Brad LiLaing Singh", "21662358", ""),
    "5351": ("Jessica Sukar", "", ""),
    "5352": ("Balbina Briles", "21662299", ""),
    "5353": ("GEAN GRAHAM", "", ""),
    "5354": ("Roberth Davis-Cordero", "21645974", ""),
    "5355": ("Gregory McTaggart", "21652263", ""),
    "5356": ("Marica  Martin", "", ""),
    "5357": ("Alieah McFarlane", "21692238", ""),
    "5358": ("Michael Wilson", "21664330", ""),
    "5359": ("Sivaganapathy  Thulasingam", "", ""),
    "5360": ("Leslie Hydes Jr", "20003585", ""),
    "5361": ("Jordan McLaughlin", "21664163", ""),
    "5362": ("Faith Maina", "21733826", ""),
    "5363": ("Alessandro Di Felice", "20032771", ""),
    "5364": ("Endre Koszec", "21637902", ""),
    "5365": ("Linford Pierson", "21633619", ""),
    "5366": ("Richome  Scarlett", "", ""),
    "5367": ("Melanie  Gillians", "21664412", ""),
    "5368": ("Vasudevan Pitchandi", "21665133", ""),
    "5369": ("Anthony Abenoja", "", ""),
    "5370": ("Taylor More", "21669212", ""),
    "5371": ("Jonathan Chacon", "", ""),
    "5372": ("Nicholas Wright", "", ""),
    "5373": ("Thomas Bryan Ebanks", "21666630", ""),
    "5374": ("Nathalie Trim-Johnson", "21665957", ""),
    "5375": ("Charlie Reaney", "21667079", ""),
    "5376": ("Amonea Barnes", "21675910", ""),
    "5377": ("Jennifer Easton", "", ""),
    "5378": ("Jacqueline Ebanks", "20018557", ""),
    "5379": ("Evelyn Swaby Swaby", "", ""),
    "5380": ("Ryan King", "21647840", ""),
    "5381": ("Ben Calderhead", "21667221", ""),
    "5382": ("Taneisha  Peters", "", ""),
    "5383": ("Kimberly Ebanks", "21698839", ""),
    "5384": ("Lisa Yun", "21664051", ""),
    "5385": ("Jeremy Hadley", "21669242", ""),
    "5386": ("Cindy  conway", "21669697", ""),
    "5387": ("Amandai Brown", "21669350", ""),
    "5388": ("Andrei Baicoianu", "", ""),
    "5389": ("Shawn Birkett", "20056961", ""),
    "5390": ("Yolanda  Phillips", "21635231", ""),
    "5391": ("Aleta Botha", "21669615", ""),
    "5392": ("Gordon Miller", "", ""),
    "5393": ("Kristen  Rankin", "21651471", ""),
    "5394": ("Randall Sudlow", "21619200", ""),
    "5395": ("James Hutchings", "21670535", ""),
    "5396": ("Theresa Terry", "21678930", ""),
    "5397": ("Diana  Webb", "21677638", ""),
    "5398": ("Michelle Manuel", "", ""),
    "5399": ("Catriona Steele", "", ""),
    "5400": ("Christopher Heusler", "21672720", ""),
    "5401": ("Jasmine Myles", "21673812", ""),
    "5402": ("Carlos Javier  Liriano de los Santos", "", ""),
    "5403": ("Dina Kahiri", "21636235", ""),
    "5404": ("Jerhon Johnson", "21658632", ""),
    "5405": ("Taneisha  Peters", "", ""),
    "5406": ("Josephine Gerolao", "20015958", ""),
    "5407": ("Everold  Manning", "21709367", ""),
    "5408": ("Sherlene McDougall", "21688989", ""),
    "5409": ("Sam Hall", "", ""),
    "5410": ("Marlon Riera", "21673370", ""),
    "5411": ("Kathleen Helvester", "21676617", ""),
    "5412": ("Valerie Blume", "21674507", ""),
    "5413": ("Diana Dyer", "21673917", ""),
    "5414": ("Tashoi Blair", "21674202", ""),
    "5415": ("Lesley Tafadzwa Deglis", "21676023", ""),
    "5416": ("Shanna-lee  shim", "21688838", ""),
    "5417": ("Edward Swaby", "21675059", ""),
    "5418": ("Jerica Hurlston", "21675076", ""),
    "5419": ("Brittany Balcewich", "21676708", ""),
    "5420": ("Jacqueline Duffus", "", ""),
    "5421": ("Andrew Jones", "21645922", ""),
    "5422": ("Yvette AustinElbaz", "21643296", ""),
    "5423": ("Louise Davis", "21674935", ""),
    "5424": ("Nelma Antonio", "21674597", ""),
    "5425": ("Johnoy  Stone", "", ""),
    "5426": ("Beth Waterfall", "21675816", ""),
    "5427": ("Martine  Ferguson", "21676623", ""),
    "5428": ("Alice Solomon", "21713764", ""),
    "5429": ("Ruby Rose Julian", "21655548", ""),
    "5430": ("Beatriz Benitez", "21640152", ""),
    "5431": ("KImberly Stetz", "21677733", ""),
    "5432": ("kenrick whitaker", "", ""),
    "5433": ("BEAL WILSON", "21685478", ""),
    "5434": ("Chantelle Erasmus", "21676318", ""),
    "5435": ("Alissa Moberg", "", ""),
    "5436": ("Hannah Smith", "21668090", ""),
    "5437": ("Craig-Anthony Jervis", "21688383", ""),
    "5438": ("Latesha Ritch", "21660461", ""),
    "5439": ("Olisiea Blake", "21621577", ""),
    "5440": ("Neika Seymour", "21635740", ""),
    "5441": ("Bonnie Gerow", "21655640", ""),
    "5442": ("Sabrina Nunez", "21677787", ""),
    "5443": ("Claire La-Roda Thomas", "21678478", ""),
    "5444": ("Mirza Kaufmann", "21678322", ""),
    "5445": ("Brenda Chin", "21679553", ""),
    "5446": ("Martin  Lancaster", "21686789", ""),
    "5447": ("Mylene Marques", "", ""),
    "5448": ("Martina Allen", "21673837", ""),
    "5449": ("Martina Allen", "21673837", ""),
    "5450": ("Ritchel Manlincon", "", ""),
    "5451": ("Yuri Sanchez", "20770317", ""),
    "5452": ("Gerardo  Soler", "", ""),
    "5453": ("Halle Manderson", "21639258", ""),
    "5454": ("Ruben Cisternino", "21658992", ""),
    "5455": ("Julia Milwood", "21679871", ""),
    "5456": ("Sandra  Millwood", "21679872", ""),
    "5457": ("Karla  Jaquez-garabitos", "", ""),
    "5458": ("Christopher Connolty", "21649768", ""),
    "5459": ("Amir Palmer", "", ""),
    "5460": ("Vladimir Ubiparip", "", ""),
    "5461": ("Pamela Duffy", "21681231", ""),
    "5462": ("Boris Brady", "", ""),
    "5463": ("Cory Thackeray", "21680692", ""),
    "5464": ("Aleksei Seren", "21683905", ""),
    "5465": ("Kadie Wright", "21681221", ""),
    "5466": ("Coretta  Edwards", "", ""),
    "5467": ("Test Tester", "", ""),
    "5468": ("Joselene Vegiz", "", ""),
    "5469": ("Jamie Decker", "", ""),
    "5470": ("Rosita Parkes", "21682743", ""),
    "5471": ("Bronwyn  Orford", "", ""),
    "5472": ("Juan Pablo Urrutia", "21683896", ""),
    "5473": ("Julie Corsetti", "21683106", ""),
    "5474": ("Lucia Ebanks", "21682931", ""),
    "5475": ("Yamil Test Isa Test", "", ""),
    "5476": ("Norma Jefferson - El", "", ""),
    "5477": ("Gale Zappacosta", "21683674", ""),
    "5478": ("Camille Da Silva", "21683516", ""),
    "5479": ("Dwayne Chip McLellan", "21656477", ""),
    "5480": ("SANDRA HAMILTON", "21683728", ""),
    "5481": ("Nicholas  Vasic", "21684128", ""),
    "5482": ("Bethalia Bailey", "21667169", ""),
    "5483": ("Laura Le Marechal", "21684215", ""),
    "5484": ("Ryan Charles", "", ""),
    "5485": ("Dorlisa Piercy", "21207067", ""),
    "5486": ("Nicola  White", "21750980", ""),
    "5487": ("mercy Troendle", "21688550", ""),
    "5488": ("James Tulloch", "21685060", ""),
    "5489": ("Soul-Ann Weekes", "", ""),
    "5490": ("Timothy Dilbert", "20011086", ""),
    "5491": ("paul ulett", "21685049", ""),
    "5492": ("KATIE FILIPO DA CRUZ", "21685232", ""),
    "5493": ("Jessica  Eden", "21634878", ""),
    "5494": ("Erin logan", "", ""),
    "5495": ("Diedre-Ana Bodden", "21673055", ""),
    "5496": ("Luis Monzon", "", ""),
    "5497": ("Rachel Williams", "21634220", ""),
    "5498": ("Chantae Stimpson", "", ""),
    "5500": ("Isabel Scott", "", ""),
    "5501": ("Graeme McIntyre", "21685653", ""),
    "5502": ("Anna McIntyre", "21685721", ""),
    "5503": ("Lisa Czudnochowsky", "21686481", ""),
    "5504": ("Tanja Ebanks", "21685000", ""),
    "5505": ("Jahmaulle Watson", "", ""),
    "5506": ("Jahmaulle Watson", "", ""),
    "5507": ("kyla Ebanks", "", ""),
    "5508": ("Demetri Chambers", "", ""),
    "5509": ("Todd Aird", "21687049", ""),
    "5510": ("robert o sullivan", "", ""),
    "5511": ("Leyda Ritch", "21687414", ""),
    "5512": ("Alonzo De Freitas", "21678097", ""),
    "5513": ("Alexandra Douglas", "", ""),
    "5514": ("Will Dean", "21687020", ""),
    "5515": ("Harris Scher", "", ""),
    "5516": ("Marcelo Spinardi", "21691077", ""),
    "5517": ("Tisa  Pickersgill", "", ""),
    "5518": ("Regina Davis", "21690784", ""),
    "5519": ("Simone  Johnson", "", ""),
    "5520": ("Leigh-Ann Edwards", "21652209", ""),
    "5521": ("Plamena Kostova", "21688571", ""),
    "5522": ("Patricia  Paez", "21681951", ""),
    "5523": ("SHANNILLE MCLEAN", "21641679", ""),
    "5524": ("Tanasa Gallimore", "21725968", ""),
    "5525": ("Jaime Fernandez", "", ""),
    "5526": ("Jaime Fernandez", "", ""),
    "5527": ("Jaime Fernandez", "", ""),
    "5528": ("Ben Lindsey", "21688998", ""),
    "5529": ("Kathryn Turley", "21657654", ""),
    "5530": ("Ionut RUSU BOCA", "21684107", ""),
    "5531": ("Brett Prior", "21690513", ""),
    "5532": ("Barrington Campbell", "21733999", ""),
    "5533": ("Teka Lee", "", ""),
    "5534": ("Roberto Hernandez", "", ""),
    "5535": ("Alexandra Cutus", "21688574", ""),
    "5536": ("Marco Antonio Mendoza", "21691922", ""),
    "5537": ("Lyndon Waite", "", ""),
    "5538": ("William Steen", "21691946", ""),
    "5539": ("Tracy Holst", "", ""),
    "5540": ("Tim Howard", "21683284", ""),
    "5541": ("Shashel Wright", "", ""),
    "5542": ("Mark Fraser", "", ""),
    "5543": ("Mohamed Balti", "21696160", ""),
    "5544": ("Peter-Gay Watson", "21693230", ""),
    "5545": ("Cynthia Mortimer", "21721740", ""),
    "5546": ("Paula Avelli", "21688698", ""),
    "5547": ("Robert  Anniford", "", ""),
    "5548": ("John Ebanks", "20084467", ""),
    "5549": ("Jonathan MacDonald", "21694124", ""),
    "5550": ("Emily Bynoe", "21729323", ""),
    "5551": ("Berta Zambrano", "", ""),
    "5552": ("Jonathan MacDonald", "21694124", ""),
    "5553": ("William G Stewart", "", ""),
    "5554": ("Karen Bentham", "21694302", ""),
    "5555": ("Brittani McGregor", "21640187", ""),
    "5556": ("Julie Test", "", ""),
    "5557": ("Wong Sew Sen Ah Kang", "21694694", ""),
    "5558": ("Sarah Sussman", "21694735", ""),
    "5559": ("Janesher Palmer", "", ""),
    "5560": ("Iam Tester", "", ""),
    "5561": ("MBE TESTING", "", ""),
    "5562": ("Winston Rennis", "21718866", ""),
    "5563": ("Lomay Russell", "21695044", ""),
    "5564": ("Shanna-kay Holness", "21676105", ""),
    "5565": ("Barleigh McCarthy", "20069051", ""),
    "5566": ("Marina Harris", "", ""),
    "5567": ("Test Testings", "", ""),
    "5568": ("Diana Martinez", "", ""),
    "5569": ("Test Testagain", "", ""),
    "5570": ("Test Testing Again", "", ""),
    "5571": ("Diana Karen Martinez", "", ""),
    "5572": ("JAIME FERNANDEZ", "", ""),
    "5573": ("Diana Martinez", "", ""),
    "5574": ("Duchess Natasha Gayle", "", ""),
    "5575": ("Christopher Marable", "21696006", ""),
    "5576": ("Natalie Ebanks", "", ""),
    "5577": ("Vanessa Rubach", "21695904", ""),
    "5578": ("Soleil Watts", "21723786", ""),
    "5579": ("Lateishea Cooper", "21694211", ""),
    "5580": ("Steven Woollett", "21696295", ""),
    "5581": ("Island Primary", "", ""),
    "5582": ("Island Primary", "", ""),
    "5583": ("Jennifer Cowdroy", "20031006", ""),
    "5584": ("Tresian Thomas", "", ""),
    "5585": ("Charlin Jakeline", "21699589", ""),
    "5586": ("Shanagaye Walker", "", ""),
    "5587": ("Douglas Weick", "21696363", ""),
    "5588": ("Lotoya Reid", "21635616", ""),
    "5589": ("Isobel Stott", "21679688", ""),
    "5590": ("Kristen Reid", "", ""),
    "5591": ("Mihail Cristian Mircioiu", "21697522", ""),
    "5592": ("Giezis Carbajal", "21697618", ""),
    "5593": ("Robert Wharton", "", ""),
    "5594": ("Andrea Johnson", "", ""),
    "5595": ("Brooke Parchment", "21638952", ""),
    "5596": ("Sofia Tobutt", "21793189", ""),
    "5597": ("MBE Harbour Walk", "", ""),
    "5598": ("Marilyn Manuel", "", ""),
    "5599": ("Shawnica Robinson", "21637625", ""),
    "5600": ("Test Tester", "", ""),
    "5601": ("Aretha Giscombe", "21701159", ""),
    "5602": ("Anyarin Wirotthanaphat", "21688586", ""),
    "5603": ("Mihai Lupu", "21699923", ""),
    "5604": ("Fedy Sungap", "21733916", ""),
    "5605": ("Michael Christiansen", "21700104", ""),
    "5606": ("Matthew Fisher", "21687395", ""),
    "5607": ("Marilyn Campbell-Davis", "20012935", ""),
    "5608": ("Abiah Cole", "21714746", ""),
    "5609": ("Fay Aranguren", "", ""),
    "5610": ("Lucas Tibbetts", "21735633", ""),
    "5611": ("Pretty Thulisine Moloi", "21700466", ""),
    "5612": ("Suzette Darby", "21700354", ""),
    "5613": ("Mya Bennett", "21702834", ""),
    "5614": ("Madeleine Smith", "", ""),
    "5615": ("Chyna Powery", "21706570", ""),
    "5616": ("Jonah Ebanks", "", ""),
    "5617": ("Jessica Leggatt", "", ""),
    "5618": ("Nicole Chisholm", "21701911", ""),
    "5619": ("Alehandro Hudgson", "21641357", ""),
    "5620": ("Shamara Lewis", "", ""),
    "5621": ("Ian Kawaley", "21686248", ""),
    "5622": ("Chez'arie McGill", "21675804", ""),
    "5623": ("Patricia Toro-Sigmon", "21702722", ""),
    "5624": ("Kerry-Ann Lewis", "", ""),
    "5625": ("Kera Brown", "21719560", ""),
    "5626": ("Vienie Vernon", "20070629", ""),
    "5627": ("Chester Flake", "", ""),
    "5628": ("James Kattan", "21703522", ""),
    "5629": ("David Jordine", "21703891", ""),
    "5630": ("Song Wang", "21703533", ""),
    "5631": ("Dominic Rivers", "21703639", ""),
    "5632": ("Darryl Hather", "21704539", ""),
    "5633": ("Reineer Dominguez", "21699769", ""),
    "5634": ("Shaneice Williams", "21700616", ""),
    "5635": ("Anthony Wallace", "21703882", ""),
    "5636": ("Adam Sax", "21684125", ""),
    "5637": ("Witney Nathan", "21711131", ""),
    "5638": ("Cristian Bezi", "", ""),
    "5639": ("Diane Langford", "21683034", ""),
    "5640": ("Roy Yamm", "", ""),
    "5641": ("Annabelle Bush", "21728063", ""),
    "5642": ("Corey Cato", "21634897", ""),
    "5643": ("Carmen Carcamo", "21705641", ""),
    "5644": ("Paul Helvestor", "", ""),
    "5645": ("Joel Webster", "21705548", ""),
    "5646": ("Summer Du Preez", "21704922", ""),
    "5647": ("Peter Heiss", "20607763", ""),
    "5648": ("Andrea Gorocica", "", ""),
    "5649": ("Shari Walton-Rankin", "20215928", ""),
    "5650": ("Charlotte Crocker", "21704922", ""),
    "5651": ("Edward Howell", "21731016", ""),
    "5652": ("Sherona Palmer", "", ""),
    "5653": ("Neeshan Peters", "20039797", ""),
    "5654": ("Mckaine Gallimore", "21708247", ""),
    "5655": ("Jodian Taylor", "21703940", ""),
    "5656": ("Peta-gaye Leslie", "", ""),
    "5657": ("Shalini Amarasinghe", "", ""),
    "5658": ("Landie Ebanks", "21707873", ""),
    "5659": ("Marc Daryl Obillo", "", ""),
    "5660": ("Hussayne Wallace", "", ""),
    "5661": ("Loretta Gillispie", "21650087", ""),
    "5662": ("Karina Tibbetts", "20001237", ""),
    "5663": ("Primrose Clarke", "20724497", ""),
    "5664": ("Jermaine Salmon", "", ""),
    "5665": ("Jabari Powery", "21709288", ""),
    "5666": ("Venessa Paragh", "21709947", ""),
    "5667": ("Queenie Alegonza", "21709941", ""),
    "5668": ("Miranda Ramoon", "21090899", ""),
    "5669": ("Demar Williams", "", ""),
    "5670": ("Norma Ebanks", "", ""),
    "5671": ("Daniel Szymanski", "21710872", ""),
    "5672": ("Oleksandr Topalov", "21702455", ""),
    "5673": ("Natalie Oliver", "21710356", ""),
    "5674": ("Itay Lewinski", "", ""),
    "5675": ("Jose Hernandez", "21710333", ""),
    "5676": ("Cadian Drummond", "21706296", ""),
    "5677": ("Thiago Chavier Da Silva", "21710767", ""),
    "5678": ("Chris Thibedeau", "", ""),
    "5679": ("Michael Mullings", "21710672", ""),
    "5680": ("Nadisha Bennett", "", ""),
    "5681": ("Leonardo Brown", "", ""),
    "5682": ("Veronica Jones-Mathias", "21712236", ""),
    "5683": ("Andrea Platt", "", ""),
    "5684": ("Tim Warren", "21711486", ""),
    "5685": ("Zarria Morgan", "21719405", ""),
    "5686": ("Zakisha Young", "21685718", ""),
    "5687": ("Shennell Peters", "21711953", ""),
    "5688": ("Andreea Rivard", "21636210", ""),
    "5689": ("Ashley -Ann Clarke", "21637874", ""),
    "5690": ("Tanasha McRae", "21675300", ""),
    "5691": ("Amir James Rasteh", "21695815", ""),
    "5692": ("Jodian Taylor", "21675120", ""),
    "5693": ("Bazz Morgan", "", ""),
    "5694": ("Jenno Lantaya", "21713047", ""),
    "5695": ("Tanja Ebanks", "21685000", ""),
    "5696": ("Margaret Hungwe", "", ""),
    "5697": ("Natalie Lazenby", "21729587", ""),
    "5698": ("Shanley Mclean", "21713749", ""),
    "5699": ("Natalie Lazenby", "21729587", ""),
    "5700": ("Dante DaCosta", "21713191", ""),
    "5701": ("Shawn Saunders", "21713398", ""),
    "5702": ("Keyda Barrett-King", "", ""),
    "5703": ("Gregory Peacock", "21713558", ""),
    "5704": ("Sherica Stewart", "", ""),
    "5705": ("Tara Bush", "", ""),
    "5706": ("Jacklyn Walton", "21683936", ""),
    "5707": ("Indirea Binns", "20749951", ""),
    "5708": ("Lucy Renault", "21681655", ""),
    "5709": ("Aristea Underwood", "", ""),
    "5710": ("Sandria Walker", "21714252", ""),
    "5711": ("Nathalie Fakhry", "21651642", ""),
    "5712": ("Jonathan Anglin", "20109634", ""),
    "5713": ("Sammantha Clarke", "", ""),
    "5714": ("Treshorna Randall", "21715123", ""),
    "5715": ("Omar Evans", "", ""),
    "5716": ("Gabriela Illa", "", ""),
    "5717": ("Cayman Leasing", "21653093", ""),
    "5718": ("Mickhale Green", "", ""),
    "5719": ("Gillian Hernandez", "", ""),
    "5720": ("Matthew Kitching", "21714975", ""),
    "5721": ("Annabelle Hurlston-Bush", "", ""),
    "5722": ("Stephen Jay", "", ""),
    "5723": ("Donna Fosters", "", ""),
    "5724": ("Alpacino Andrews", "", ""),
    "5725": ("Christina Rankine", "", ""),
    "5726": ("Cindy Gouws", "21704191", ""),
    "5727": ("Ashley Levine", "21716916", ""),
    "5728": ("Gina Lawrence", "", ""),
    "5729": ("Dariel Pouza", "21729733", ""),
    "5730": ("Abiah Cole", "21714746", ""),
    "5731": ("Kaneil Barrett", "20009537", ""),
    "5732": ("Natasia Ebanks", "21717044", ""),
    "5733": ("Cerwyn Lithgow", "21718685", ""),
    "5734": ("Roe Laesea Francis", "", ""),
    "5735": ("Carolina Caraballo Garcia", "21723911", ""),
    "5736": ("Erica  Christie", "21680677", ""),
    "5737": ("Jaida Alexander", "21718274", ""),
    "5738": ("Sean Pitcher", "21721221", ""),
    "5739": ("Michael Edwards II", "21710313", ""),
    "5740": ("Stacey Dorush", "21718431", ""),
    "5741": ("Daniel Canencia", "21650286", ""),
    "5742": ("Roxenette Gil Valerio", "21719204", ""),
    "5743": ("Glenn Kennedy", "21597895", ""),
    "5744": ("mario altamirano quispe", "21719231", ""),
    "5745": ("Michelle Allen", "", ""),
    "5746": ("Madonna Hiromi Sato", "21696206", ""),
    "5747": ("Toussaint Burke", "", ""),
    "5748": ("Billie Bryan", "21720534", ""),
    "5749": ("Jasmine McLaughlin", "", ""),
    "5750": ("Shaquille Brown", "21724886", ""),
    "5751": ("Kelano Bynes", "21730157", ""),
    "5752": ("Nicole Moore", "21650426", ""),
    "5753": ("Brittany Bennett", "21635704", ""),
    "5754": ("Bettylee Johnson", "21723476", ""),
    "5755": ("Ranald Henderson", "21634188", ""),
    "5756": ("Tom Marks", "", ""),
    "5757": ("Zhiheng Rao", "21720977", ""),
    "5758": ("Jason Ebanks", "20003140", ""),
    "5759": ("Vina Bianca San Juan", "21721739", ""),
    "5760": ("Joseph Revitte", "21721392", ""),
    "5761": ("Janeth Mani", "21723947", ""),
    "5762": ("Elizabeth Mahadeo", "21721736", ""),
    "5763": ("Elda Fuentes", "21631667", ""),
    "5764": ("Alice Mclaughlin", "21675306", ""),
    "5765": ("Nahkiesha  Ebanks", "21636259", ""),
    "5766": ("Jelena Vujovic", "21274327", ""),
    "5767": ("David Gordon", "", ""),
    "5768": ("Matthew Clubbe", "", ""),
    "5769": ("Daniel Baumslag", "", ""),
    "5770": ("Lalika Mcfield", "21728732", ""),
    "5771": ("Anna Milanowska", "21665742", ""),
    "5772": ("Mr Diaz/Happy Days Paddle & Coffee Ltd", "21725289", ""),
    "5773": ("Ashley Jones", "", ""),
    "5774": ("Ralna Simmonds", "", ""),
    "5775": ("Che Andrews", "", ""),
    "5776": ("Jane Moseley", "21715922", ""),
    "5777": ("JanWanda Dewar", "20579781", ""),
    "5778": ("Dimitri Ritch", "21679558", ""),
    "5779": ("Roe Francis", "", ""),
    "5780": ("Ashley Robinson", "", ""),
    "5781": ("Marlon Campbell", "", ""),
    "5782": ("Alex Karanja", "", ""),
    "5783": ("Daisy Lines", "", ""),
    "5784": ("George E Hydes Jr", "", ""),
    "5785": ("daniel murphy", "20003072", ""),
    "5786": ("Rudy Inoa", "", ""),
    "5787": ("Ahmed Marrero", "21620239", ""),
    "5788": ("Jonathan Stott", "21725857", ""),
    "5789": ("Gregory Morgan", "20039721", ""),
    "5790": ("KnK Leds Accessories", "", ""),
    "5791": ("Ben Lindsey", "21688998", ""),
    "5792": ("Pascal Gajraj", "21667899", ""),
    "5793": ("Aylem Fernandez", "", ""),
    "5794": ("David WALKER", "21635377", ""),
    "5795": ("Alison JURKOWSKI", "21726930", ""),
    "5796": ("Yorgena Polanco", "21717476", ""),
    "5797": ("Krystal Morales-Ebanks", "", ""),
    "5798": ("Omeila Crosebourne", "21727857", ""),
    "5799": ("Rosemary Becerra", "21699600", ""),
    "5800": ("Travis Bodden", "", ""),
    "5801": ("Shanda Gallego", "20076505", ""),
    "5802": ("Kirby Taylor", "21729118", ""),
    "5803": ("Rochelle Morgan", "21737225", ""),
    "5804": ("Carla Gordon", "", ""),
    "5805": ("Christian Palikuca", "21728747", ""),
    "5806": ("Kristoff powell", "", ""),
    "5807": ("Karen Ford", "21728739", ""),
    "5808": ("Anthony Hall", "21728483", ""),
    "5809": ("Julian collins", "", ""),
    "5810": ("Kerrian Ann Marie", "", ""),
    "5811": ("Cynthia Boyett", "21732690", ""),
    "5812": ("Russell Linford", "", ""),
    "5813": ("Malique Ebanks", "21715448", ""),
    "5814": ("Holly Mackenzie", "21729967", ""),
    "5815": ("Michael Lewis", "20753886", ""),
    "5816": ("Tavanna Harrison", "", ""),
    "5817": ("Gizelle Watler", "21717793", ""),
    "5818": ("Jesse Arch", "", ""),
    "5819": ("Jade Spencer", "21714732", ""),
    "5820": ("Justin Montiel", "21759057", ""),
    "5821": ("WOLFF DIGITAL MARKETING SEZC", "21716571", "SPCL ECO ZONE"),
    "5822": ("David Bennett", "21695648", ""),
    "5823": ("Patricia Mandish", "", ""),
    "5824": ("Rande rivers", "21730933", ""),
    "5825": ("Nathan Higgins", "21731693", ""),
    "5826": ("Massiel Martinez-Roman", "", ""),
    "5827": ("Jessica Bradbury", "21731443", ""),
    "5828": ("Bailey Welds", "21731535", ""),
    "5829": ("Kira Webster", "21731794", ""),
    "5830": ("Aaron Froelich", "21717596", ""),
    "5831": ("Ditte Henderson", "21706811", ""),
    "5832": ("Nicola Reynolds", "", ""),
    "5833": ("Andre Jackson", "21719079", ""),
    "5834": ("Madeline Harrop", "21732509", ""),
    "5835": ("Rah'Shawn Gardner", "21722128", ""),
    "5836": ("Stephen Ebanks", "", ""),
    "5837": ("Brian Walpole", "21704367", ""),
    "5838": ("Jennifer Norris", "21733377", ""),
    "5839": ("Tashauna Mckenzie", "", ""),
    "5840": ("Shekirah Ebanks", "", ""),
    "5841": ("Meghan Fitzpatrick", "21703349", ""),
    "5842": ("Mark Russell", "", ""),
    "5843": ("Craig Uwins", "21699404", ""),
    "5844": ("Chloe Allotey", "21733255", ""),
    "5845": ("Gary Phillips", "", ""),
    "5847": ("Shantel Miller", "", ""),
    "5848": ("Nekeisha Henry", "", ""),
    "5849": ("Bobisher Harrison", "21706697", ""),
    "5850": ("Mari Chaccour", "", ""),
    "5851": ("Adam Bobrowski", "21712250", ""),
    "5852": ("Emma Joy", "21739348", ""),
    "5853": ("Colleen and Dana Armitage and Ball", "21676553", ""),
    "5854": ("Paula Chang", "21736659", ""),
    "5855": ("Monique Wilson", "", ""),
    "5856": ("Hillegonda Spits", "21736304", ""),
    "5857": ("Ahmed Daoudi Mjida", "", ""),
    "5858": ("Sidney Terry", "21736425", ""),
    "5859": ("Aminata Shepherd", "", ""),
    "5860": ("Shanay Hanson", "21745981", ""),
    "5861": ("STEPHEN BAILEY", "", ""),
    "5862": ("Paulo Simoes", "21737568", ""),
    "5863": ("Dayana Powery", "", ""),
    "5864": ("Kamil Szumanski", "21737760", ""),
    "5865": ("Kritza Kirkconnell", "21728417", ""),
    "5866": ("Sadek Palmer", "", ""),
    "5867": ("Marissa Visayas", "", ""),
    "5868": ("Zenia Blanco", "", ""),
    "5869": ("Alexandros Kousoulas", "21738496", ""),
    "5870": ("Rachel Adams", "21753324", ""),
    "5871": ("Adam Fox", "", ""),
    "5872": ("Lindsay Mulligan", "21739077", ""),
    "5873": ("Cathy Scott", "21740504", ""),
    "5874": ("Belfry Lisbeth Jimenez", "21713030", ""),
    "5875": ("Matthew Greenberg", "21739429", ""),
    "5876": ("Flor Navarro", "", ""),
    "5877": ("Radagi Batista javier", "", ""),
    "5878": ("Stephen Crowther", "", ""),
    "5879": ("Elizabeth Stephenson", "21740547", ""),
    "5880": ("Nicola Ebanks", "", ""),
    "5881": ("Meleta Townsend -LAWRENCE", "21711847", ""),
    "5882": ("Grace Christian-Welcome", "21686981", ""),
    "5883": ("Isabella Hubbell", "", ""),
    "5884": ("Brittany Chang Upstone", "", ""),
    "5885": ("Roxanne Smith", "21742351", ""),
    "5886": ("Drew Nicholls", "", ""),
    "5887": ("Molly Snowberger", "21704134", ""),
    "5888": ("Arianna McFarlane", "", ""),
    "5889": ("Noel Cayasso-Smith", "20008077", ""),
    "5890": ("Chastine Rankine", "", ""),
    "5891": ("Jardelyn Viernes", "", ""),
    "5892": ("Arihana Lisboa", "", ""),
    "5893": ("Nada El Khashab", "21741918", ""),
    "5894": ("Regina Ebanks", "", ""),
    "5895": ("Janielle Buttrum", "", ""),
    "5896": ("Chena Bodden", "", ""),
    "5897": ("Ruthney Shaw", "21757453", ""),
    "5898": ("Shaniek Beaumont", "21734126", ""),
    "5899": ("Stephen James Pittman", "20000068", ""),
    "5900": ("Cristhian Prado", "", ""),
    "5901": ("Ace Dampil", "21745616", ""),
    "5902": ("Lewis Bebora", "21743125", ""),
    "5903": ("Aleah Copeland", "21693308", ""),
    "5904": ("AART VAN DIJK", "21675066", ""),
    "5905": ("Sarah Torino", "21744317", ""),
    "5906": ("Dylan Evans", "", ""),
    "5907": ("Stuart Barton", "21734931", ""),
    "5908": ("Leah Beckles", "21697248", ""),
    "5909": ("Tajeme Mckenzie", "21744233", ""),
    "5910": ("Rita Pinhal Smith", "21712225", ""),
    "5911": ("Olaniyi David", "", ""),
    "5912": ("Zane Powles", "", ""),
    "5913": ("LEEANN CHRISTIAN", "", ""),
    "5914": ("Ala Hasan Yousef Abu Eish", "", ""),
    "5915": ("Elan Groves", "20021340", ""),
    "5916": ("Delroy Taylor", "21744642", ""),
    "5917": ("Sara Simpson", "", ""),
    "5918": ("Luiyi Urbina", "", ""),
    "5919": ("Dinesh D silva", "", ""),
    "5920": ("Luiyi Urbina", "", ""),
    "5921": ("Tahirah Henriques", "21746310", ""),
    "5922": ("Jeremy Hurst", "21631310", ""),
    "5923": ("MARILYN SWING", "", ""),
    "5924": ("Analin Brinson", "21742506", ""),
    "5925": ("Rossmery Massiel De la Rosa", "", ""),
    "5926": ("Stephen Gilsenan", "21747030", ""),
    "5927": ("Summer Ross", "", ""),
    "5928": ("Tishawana Gauntlett", "21747259", ""),
    "5929": ("rossmery massiel de la rosa perez", "", ""),
    "5930": ("Rosario Ciantar", "", ""),
    "5931": ("Valeska Alexander", "21610838", ""),
    "5932": ("Elizabeth Ordd Suazo", "21749151", ""),
    "5933": ("JOHNOY STONE", "21737994", ""),
    "5934": ("Jeffrey debins Ebanks herrera", "21748294", ""),
    "5935": ("Lauren Rose James", "", ""),
    "5936": ("Yanique Thompson", "", ""),
    "5937": ("Sabrina Bonthorne", "", ""),
    "5938": ("Lynette Manderson", "21748645", ""),
    "5939": ("Angela Davis", "", ""),
    "5940": ("Adiel Gordon", "", ""),
    "5941": ("Vivia Banks", "", ""),
    "5942": ("Halfranm Hydes", "21762172", ""),
    "5943": ("Diandra McCoy", "", ""),
    "5944": ("Shaunagh Marie Dempsey", "", ""),
    "5945": ("Dara Scott", "21727429", ""),
    "5946": ("Gary Bahre", "", ""),
    "5947": ("Hugh Black", "21750122", ""),
    "5948": ("Zalisha McLaughlin", "20739654", ""),
    "5949": ("Dorothy Manzanares", "21696161", ""),
    "5950": ("Casey Walker", "21751469", ""),
    "5951": ("PIPPA CASSIDY", "21659188", ""),
    "5952": ("Warwin Cruz", "", ""),
    "5953": ("Tyrese Richards", "", ""),
    "5954": ("Corey Bonthorne", "21653662", ""),
    "5955": ("Justin Rankin", "21625976", ""),
    "5956": ("Valerie Parsons", "20593442", ""),
    "5957": ("Reuben Smith", "", ""),
    "5958": ("Jessica Tesconi", "", ""),
    "5959": ("Rhisto Reyes", "21752742", ""),
    "5960": ("Indigo Insurance Ltd", "21762471", ""),
    "5961": ("Emma Milburn", "21719954", ""),
    "5962": ("Gemma Cowan", "21682831", ""),
    "5963": ("Leydis Hechavarria Cardona", "", ""),
    "5964": ("Jose Madrigal", "21752726", ""),
    "5965": ("Kevin Ramirez", "", ""),
    "5966": ("Ryan Punambolam", "21742784", ""),
    "5967": ("Chez Tschetter", "21752045", ""),
    "5968": ("Laura Odone", "21756274", ""),
    "5969": ("GEORGE BROOKS", "21756181", ""),
    "5970": ("Alizee Laurent", "21689528", ""),
    "5971": ("Latina Young", "20013038", ""),
    "5972": ("CIG Department of Counselling", "21742655", ""),
    "5973": ("Daniel Baigent", "", ""),
    "5974": ("Yamelis Bonilla Diaz", "", ""),
    "5975": ("Ben Brown", "21757361", ""),
    "5976": ("Ashleigh Dorrington-Yi", "", ""),
    "5977": ("Daver Hair", "", ""),
    "5978": ("Robert Newburger", "21759729", ""),
    "5979": ("Carrie Lamon", "", ""),
    "5980": ("Vlastimil Hynek", "", ""),
    "5981": ("Tiffany Mcfarlane", "", ""),
    "5982": ("Gaylene Layden", "21758325", ""),
    "5983": ("Akime Palmer", "", ""),
    "5984": ("Jose Vasquez", "", ""),
    "5985": ("David Colebatch", "21755344", ""),
    "5986": ("Christopher Jarvis", "21683684", ""),
    "5987": ("Julien Mendes", "", ""),
    "5988": ("Joel Landell", "", ""),
    "5989": ("Madlin Diaz", "", ""),
    "5990": ("Hrachya Sargsyan", "21760593", ""),
    "5991": ("Yasheika Walters", "21681297", ""),
    "5992": ("Shannon Passley", "21667799", ""),
    "5993": ("Djemy Victor", "", ""),
    "5994": ("Laurent Pelissier", "21760048", ""),
    "5995": ("Atlas-TEST Ramoon", "", ""),
    "5996": ("Aneal Sobie", "20033100", ""),
    "5997": ("Eve Wong", "", ""),
    "5998": ("Shernett Brown", "21762017", ""),
    "5999": ("Shaniya Seymour", "", ""),
    "6000": ("Jason Smith", "21705380", ""),
    "6001": ("Emlyn Jones", "", ""),
    "6002": ("David Samuel", "", ""),
    "6003": ("Kevon Wong", "21738529", ""),
    "6004": ("Israel Garcia De La Rosa", "21764276", ""),
    "6005": ("Betty Ann Duty", "21698242", ""),
    "6006": ("Michelle Melvill", "", ""),
    "6007": ("Saday Chiu Swaby", "21615765", ""),
    "6008": ("Judy Ann Minott", "", ""),
    "6009": ("Ian Morrison", "", ""),
    "6010": ("Alyssa Watler", "21091218", ""),
    "6011": ("Daniella Le Beurrier", "21769350", ""),
    "6012": ("Do bert ebanks", "", ""),
    "6013": ("Kaliq Dixon-Rickfield", "21768625", ""),
    "6014": ("Nicolas Masso", "21767824", ""),
    "6015": ("Zoe Masters", "21747695", ""),
    "6017": ("Kemesha Richards", "", ""),
    "6018": ("Sharon Chambers", "21710170", ""),
    "6019": ("Theodore Owens", "21769164", ""),
    "6020": ("Ruby Kimberley Boncales", "21669234", ""),
    "6021": ("Victoria Jurkowski", "", ""),
    "6022": ("Thabani Dzowa", "", ""),
    "6023": ("Anescka Lewis", "", ""),
    "6024": ("Aaliyah Levy", "", ""),
    "6025": ("Tiffany Lu", "21770223", ""),
    "6026": ("Nicholas van Wyk", "21756194", ""),
    "6027": ("Melissa Rivas", "21766456", ""),
    "6028": ("Chandler Veldey", "", ""),
    "6029": ("Junior Sukhu", "", ""),
    "6030": ("Raymond Hainey", "21747697", ""),
    "6031": ("Gaynell Gordon", "21747602", ""),
    "6032": ("Paul Palmer", "", ""),
    "6033": ("Sabrina Cane", "20000663", ""),
    "6034": ("KENESHA DINNEALL-BANTON", "21773759", ""),
    "6035": ("Morrill Scott Jr.", "21762016", ""),
    "6036": ("Samantha Clarke Duncan", "21675284", ""),
    "6037": ("Doshille Osman", "21774176", ""),
    "6038": ("Tammy Patino", "21754224", ""),
    "6039": ("VERUSCHA VANESSA BAKER", "21639615", ""),
    "6040": ("Brittany Dixon", "", ""),
    "6041": ("Sebastian Scholz", "21774544", ""),
    "6042": ("Paul Parker", "", ""),
    "6043": ("Serena Litteral", "21748111", ""),
    "6044": ("Tian Peralto Levers", "", ""),
    "6045": ("Yuanyuan Li", "", ""),
    "6046": ("Amy Donnelly", "21743990", ""),
    "6047": ("Vasil Blaze", "", ""),
    "6048": ("Makayla Blackman", "21764557", ""),
    "6049": ("Lewis Cohen", "21752602", ""),
    "6050": ("Judith Kay", "21766132", ""),
    "6051": ("Junior Sukhu", "", ""),
    "6052": ("Jeremy Easton", "21765857", ""),
    "6053": ("Caroline Sully Forsberg", "21776447", ""),
    "6054": ("Emily Palmer", "21778194", ""),
    "6055": ("Chanel Bodden", "21778546", ""),
    "6056": ("Caroline Sully Forsberg", "21776447", ""),
    "6057": ("Segal N Forbes", "21717288", ""),
    "6058": ("Rhesa Long", "21767823", ""),
    "6059": ("Saskia Stevenson", "21686787", ""),
    "6060": ("Estrella Deleon", "", ""),
    "6061": ("Florian Ziegler", "21774938", ""),
    "6062": ("Nicola Ellis", "21760016", ""),
    "6063": ("Lianne Watler", "21644232", ""),
    "6064": ("Luigi Comboni", "21726931", ""),
    "6065": ("Claudia Barcelo-Patulot", "", ""),
    "6066": ("David Shaw", "21781258", ""),
    "6067": ("Yvonne Powekk", "", ""),
    "6068": ("Allison Ramkissoon Jebodh", "21774925", ""),
    "6069": ("Arien Hazell", "21782640", ""),
    "6070": ("Sarah Little", "21783047", ""),
    "6071": ("Colton Schultz", "21783280", ""),
    "6072": ("Debra Stewart", "21783290", ""),
    "6073": ("Kevin Valdespino", "", ""),
    "6074": ("James Hutchings", "21670535", ""),
    "6075": ("Carlos Abarca", "21786662", ""),
    "6076": ("Kenneth Chaplin", "21785715", ""),
    "6077": ("Stephen Roberts", "", ""),
    "6078": ("Stephen Chapman", "", ""),
    "6079": ("Tamara Bugembe", "", ""),
    "6080": ("Emily Jackson", "", ""),
    "6081": ("Lucinda Hislop", "21701458", ""),
    "6082": ("Cody McCoy", "", ""),
    "6083": ("Danielle Beriault", "", ""),
    "6084": ("Allan Lyall", "21789759", ""),
    "6085": ("Vanessa Ann Alvarez", "", ""),
    "6086": ("Shari Seymour", "", ""),
    "6087": ("Cesar Gallego", "21789483", ""),
    "6088": ("Armani Mclaughlin", "21749671", ""),
    "6089": ("Dalton Nelson", "21766765", ""),
    "6090": ("Hayley O\\'Connell", "21791022", ""),
    "6091": ("Dorbin Scott", "", ""),
    "6092": ("Maryse Harvey", "", ""),
    "6093": ("Alexander Peintner", "21792296", ""),
    "6094": ("Johnny Ferrari", "21672787", ""),
    "6095": ("Tiffinii Fischer", "", ""),
    "6096": ("Althea Swaby", "", ""),
    "6097": ("Natalie Bolland", "", ""),
    "6098": ("Terri-Lynn Brown", "21658587", ""),
    "6099": ("Simone D'Amico", "", ""),
    "6100": ("Patrick Rogers", "", ""),
    "6101": ("Caroline Diviney", "", ""),
    "6102": ("Kim Dennison", "", ""),
    "6103": ("ANNMARIE SIMPSON", "", ""),
    "6104": ("ELSPETH CRAMB", "21785514", ""),
    "6105": ("Jahida Dixon", "", ""),
    "6106": ("Heather Halsey", "21767589", ""),
    "6107": ("Sophia Ebanks", "", ""),
    "6108": ("Lucia Gallardo", "", ""),
    "6109": ("Rodrigo Huerta Amoros", "21759344", ""),
    "6110": ("Joel Small", "21630300", ""),
    "6111": ("Paolo Castello", "20070907", ""),
    "6112": ("David Jameson", "", ""),
    "6113": ("Rajhni  Murphy", "21796481", ""),
    "6114": ("Tyra Cuthbert", "21798427", ""),
    "6115": ("Agentic Technologies SEZC / Derek Manuge", "21797360", "SPCL ECO ZONE"),
    "6116": ("Dharshani Naidoo", "", ""),
    "6117": ("Lewin Parsons", "", ""),
    "6118": ("Ken Carlew Labarda", "", ""),
    "6119": ("Zachary Blacher", "21798190", ""),
    "6120": ("MARJAN BEIG", "", ""),
    "6121": ("Dennis Neylan", "21708519", ""),
    "6122": ("Cedric Gidarisingh", "20004605", ""),
    "6123": ("Michelle Plunkett", "", ""),
    "6124": ("Rosario Ciantar", "21799098", ""),
    "6125": ("D'Angela Andrade", "20518675", ""),
    "6126": ("Jean Gordon", "", ""),
    "6127": ("Mary Macayanan", "", ""),
    "6128": ("Deobrah Sylvester Murray", "", ""),
    "6129": ("Maya Smith", "21702895", ""),
    "6130": ("Kayah Clarke", "", ""),
    "6131": ("Kerri Medera", "21801026", ""),
    "6132": ("Adrien Lafeuille", "", ""),
    "6133": ("Jenalyn Sagao", "", ""),
    "6134": ("Abraham Thomas", "21798282", ""),
    "6135": ("DEVS TEST", "", ""),
    "6136": ("Timothy  Heath", "21641954", ""),
    "6137": ("Alejandro Angel", "21647767", ""),
    "6138": ("Gordon Jackson", "21802216", ""),
    "6139": ("Gregory Richardson", "21760157", ""),
    "6140": ("Shontadria Kemp", "", ""),
    "6141": ("Lisa Caplan", "", ""),
    "6142": ("Whitney Dykeman", "21687965", ""),
    "6143": ("James Beall", "21801803", ""),
    "6144": ("Matthew Wood", "21716342", ""),
}

MASTER_IMPORTER_DEFAULT = "20000561"
DEFAULT_COMMODITY_CODE = "98010029"

# Common customs procedure codes for the dropdown
PROCEDURE_OPTIONS = ["HOME", "BLD MAT", "SCHOOL", "RETAILER", "SPCL ECO ZONE"]
UNIT_OPTIONS = ["NO", "KG", "G", "L", "LB", "M2", "M3", "PCS", "PR", "DOZ"]

BUILTIN_CODES = [
    ("Ackee (Canned)", "98020047", "KG", "HOME"),
    ("Ackee (Fresh)", "98020046", "KG", "HOME"),
    ("Agarbatti and other odoriferous preparations which operate by burning", "33074100", "KG", "HOME"),
    ("Ammunition (Farmers/sport shooting association)", "98020084", "NO", "HOME"),
    ("Ammunition (Non-Farmers)", "98020083", "NO", "HOME"),
    ("Antiques of an age exceeding one hundred years", "98010057", "NO", "HOME"),
    ("Ashtrays", "71141911", "NO", "HOME"),
    ("Auto Parts and accessories", "87081000", "NO", "HOME"),
    ("Baba Roots", "98020058", "KG", "HOME"),
    ("Baggage and household effects", "98020015", "KG", "HOME"),
    ("Bammy", "98020048", "KG", "HOME"),
    ("Bed linen, towels, kitchen linen", "98020063", "NO", "HOME"),
    ("Bed Spreads, bed linen, towels", "98010033", "NO", "HOME"),
    ("Beef (fresh, frozen)", "98020038", "KG", "HOME"),
    ("Beer", "98020023", "L", "HOME"),
    ("Bicycles and cycles", "87120000", "NO", "HOME"),
    ("Bike parts and acc", "87149900", "KG", "HOME"),
    ("binoculars", "90051000", "NO", "HOME"),
    ("Boat or dock fenders", "40169400", "NO", "HOME"),
    ("Bullion (gold, silver)", "98010028", "G", "HOME"),
    ("Buns", "98020049", "KG", "HOME"),
    ("Calendars of any kinds, printed, invl. Calendar blocks", "49100000", "KG", "HOME"),
    ("Cameras; lens, filters, flashes and tripods for cameras", "98010022", "NO", "HOME"),
    ("candles, Tapers and the like", "34060000", "KG", "HOME"),
    ("Car/motor vehicle engine parts (seals, gaskets, filters)", "98010037", "NO", "HOME"),
    ("Car/motor vehicle parts and accessories", "98010035", "NO", "HOME"),
    ("Cases for cell phones", "98010021", "NO", "HOME"),
    ("Cases for jewellery", "98010032", "NO", "HOME"),
    ("CD's", "98020074", "NO", "HOME"),
    ("Cell phones", "98010020", "NO", "HOME"),
    ("Cell phones (Iphones, androids)", "98020073", "NO", "HOME"),
    ("Champagne", "98020024", "L", "HOME"),
    ("Chandeliers and other wall lighting", "98020085", "NO", "HOME"),
    ("Chewing tobacco not exceeding 250 g per importer aged 18 years or more", "98020019", "G", "HOME"),
    ("Children's picture, drawing or colouring books", "49030000", "NO", "HOME"),
    ("Cider, Wine coolers", "98020028", "L", "HOME"),
    ("Cigarettes (Over the allowable limit)", "98020036", "NO", "HOME"),
    ("Cigarettes containing tobacco", "98010059", "NO", "HOME"),
    ("Cigarettes not exceeding 200 in number per importer aged 18 years or more", "98020018", "NO", "HOME"),
    ("Cigarillos, not exceeding 200 in number per importer aged 18 years or more", "98020031", "NO", "HOME"),
    ("Cigars (export)", "98010024", "NO", "HOME"),
    ("Cigars, cheroots and cigarillos, containing tobacco", "98020034", "NO", "HOME"),
    ("Clothing", "98010017", "NO", "HOME"),
    ("Coffee (exc Nepresso pods)", "09011100", "KG", "HOME"),
    ("Coins", "71181020", "G", "HOME"),
    ("Computer equipment ((laptops, desk tops, ipads, tablets, usb, flash drives)", "98010048", "NO", "HOME"),
    ("Computers, Laptops, Ipads, Tablets", "98020071", "NO", "HOME"),
    ("Contact lenses, non-contact lenses, eye glasses, sun glasses,  and spectacles;", "98010040", "NO", "HOME"),
    ("Cosmetics (make up)", "98010045", "KG", "HOME"),
    ("Debit cards, credit cards, RSA tokens", "98010049", "NO", "HOME"),
    ("Dessert Wines", "98020027", "L", "HOME"),
    ("Discs, DVD's, CD's, software", "98010039", "NO", "HOME"),
    ("Documents", "98010010", "NO", "HOME"),
    ("Doors, windows and their frames", "98020087", "NO", "HOME"),
    ("Drinking Glasses", "70139100", "NO", "HOME"),
    ("Drones", "98010051", "NO", "HOME"),
    ("Dutiable personal goods or household articles purchased by residents returning from an overseas visit of an aggregate value not exceeding CI$350", "98020010", "KG", "HOME"),
    ("DVD's", "98020075", "NO", "HOME"),
    ("Electrical apparatus", "98020086", "NO", "HOME"),
    ("Electrical transformers,static converters (incl.battery chargers)", "85044000", "NO", "HOME"),
    ("Electronic cigarettes", "98010023", "NO", "HOME"),
    ("Electronic cigarettes (e cigarettes)", "98020037", "NO", "HOME"),
    ("Envelopes, office supplies", "98010047", "NO", "HOME"),
    ("Essential Oils", "33011900", "KG", "HOME"),
    ("Eye glasses, lenses", "90014000", "PR", "HOME"),
    ("Fabric on roll", "98010050", "M2", "HOME"),
    ("Fabrics (coton, etc.)", "98020064", "M3", "HOME"),
    ("Fertilizer", "31010000", "KG", "HOME"),
    ("Firearms (Farmers/sport shooting association)", "98020082", "NO", "HOME"),
    ("Firearms (Non-farmers)", "98020081", "NO", "HOME"),
    ("Fish (fresh)", "98010025", "KG", "HOME"),
    ("Flags of all nations", "63079010", "KG", "HOME"),
    ("Footwear (shoes)", "98010018", "PR", "HOME"),
    ("Frames and mountings for spectacles, goggles or the like", "90031100", "NO", "HOME"),
    ("Fruit (fresh, frozen)", "98020043", "KG", "HOME"),
    ("Furniture", "94035000", "NO", "HOME"),
    ("Glass mirrors", "70099200", "NO", "HOME"),
    ("Guitars, string instruments", "92029000", "NO", "HOME"),
    ("Hair (human)", "98010043", "KG", "HOME"),
    ("Hair (synthetic)", "98010044", "KG", "HOME"),
    ("Hair Conditioner", "98020051", "KG", "HOME"),
    ("Hand tools (hammers, screwdrivers, etc.)", "98020069", "NO", "HOME"),
    ("Handbags (not of leather)", "98020065", "NO", "HOME"),
    ("Headphones", "85183000", "NO", "HOME"),
    ("Hearing aids", "90214000", "NO", "HOME"),
    ("House hold items", "39249000", "KG", "HOME"),
    ("Household effects proved to have been in bona fide use and ownership abroad by the passenger for a period of not less than six months and not imported for sale or commercial exchange", "98020014", "KG", "HOME"),
    ("Hoverboards, scooters and electric Bicycles", "87119020", "NO", "HOME"),
    ("Imitation jewellery", "98010030", "NO", "HOME"),
    ("insecticides", "38086100", "KG", "HOME"),
    ("Instruments and tools of trade, to be used for the purpose of the passenger's trade or profession, proved to have been in bona fide use and ownership abroad by the passenger for a period of not less than six months and not imported for sale or commercial exchange", "98020013", "KG", "HOME"),
    ("Jewellery of Precious or semi-precious stones;  and articles of goldsmiths? or silversmiths? wares and parts thereof, of precious metal or of metal clad with precious metal", "98010029", "NO", "HOME"),
    ("Jewelry (costume)", "98020067", "NO", "HOME"),
    ("Jewelry (gold, silver, platinum)", "98020066", "NO", "HOME"),
    ("Lamps and lighting fittings", "94054000", "NO", "HOME"),
    ("Leatherbag", "42022100", "NO", "HOME"),
    ("Line telephone sets with cordless handsets", "98020072", "NO", "HOME"),
    ("Liqueurs and cordials", "98020091", "L", "HOME"),
    ("Locks of base metal (incl. door locks)", "83014000", "DOZ", "BLD MAT"),
    ("Medication", "30049000", "KG", "HOME"),
    ("Miscellaneous food products including dietary supplements", "98010027", "KG", "HOME"),
    ("Miscellaneous foods", "98020050", "KG", "HOME"),
    ("Miscellaneous goods outside the personal allowance", "98020022", "NO", "HOME"),
    ("Miscellaneous products not elsewhere specified (n.e.s.)", "98010061", "KG", "HOME"),
    ("Motor Vehicles Parts and Accessories (incl. rims)", "98020079", "NO", "HOME"),
    ("Motorcycles parts and accessories", "87141000", "KG", "HOME"),
    ("Mutton/Goat (fresh, frozen)", "98020040", "KG", "HOME"),
    ("Notebooks/composition books", "98010038", "NO", "HOME"),
    ("Office or school supplies, of plastics, n.e.s.", "39261000", "KG", "SCHOOL"),
    ("Office supplies", "98010047", "NO", "HOME"),
    ("Original artwork of a kind otherwise classifiable", "98010056", "NO", "HOME"),
    ("Orthopaedic appliances, including crutches, surgical belts and trusses; splints and other fracture appliances; artificial parts of the body;  hearing aids and other appliances which are worn or carried,", "98010053", "NO", "HOME"),
    ("Other Electronics", "98020078", "NO", "HOME"),
    ("Paints and varnishes", "32081000", "L", "HOME"),
    ("Parts of non-electric razors of base metal", "82129000", "NO", "HOME"),
    ("Party supplies", "95059000", "NO", "HOME"),
    ("Pearls", "98010031", "NO", "HOME"),
    ("Perfumes (only)", "33030010", "KG", "HOME"),
    ("Personal Effects", "9802--", "NO", "HOME"),
    ("Pharmaceutical products under Chapter 30", "98010012", "KG", "HOME"),
    ("Pharmaceutical products under Chapter 30 for Health Services Authority (H.S.A)", "98010013", "KG", "HOME"),
    ("Photosensitive semiconductor devices,(solar)", "85414010", "NO", "HOME"),
    ("Pianos", "92012000", "NO", "HOME"),
    ("Planting seeds", "12099100", "KG", "HOME"),
    ("Pneumatic Tires", "40111000", "NO", "HOME"),
    ("Pork (fresh, frozen)", "98020039", "KG", "HOME"),
    ("Postal packages not specified by commodity", "98010060", "NO", "HOME"),
    ("Power tools (with battery or motor)", "98020070", "NO", "HOME"),
    ("Preparations for oral or dental hygiene", "98020052", "KG", "HOME"),
    ("Prepared glues and other prepared adhesives put up for retail", "35052000", "KG", "HOME"),
    ("Printed books", "49019900", "NO", "HOME"),
    ("Printers, copying machines and facsimile machines", "844339000", "NO", "HOME"),
    ("Professional books (medical, accounting, law)", "98010034", "NO", "HOME"),
    ("Raw and semi-processed materials, nes", "98030000", "NO", "HOME"),
    ("Rizzla (Cigarette paper)", "98020056", "NO", "HOME"),
    ("Saddlery and harness for any animal", "42010000", "NO", "HOME"),
    ("Salon Items (Hair conditioner, etc.)", "98020051", "KG", "HOME"),
    ("Samples (carpet, tiles)", "98010041", "NO", "HOME"),
    ("Samples (liquour for promotional purposes))", "98010042", "NO", "HOME"),
    ("Sanitary towels (pads)", "96190000", "NO", "HOME"),
    ("School Supplies", "39261000", "KG", "SCHOOL"),
    ("Seasoning/condiments (ketchup, mustard)", "98020044", "KG", "HOME"),
    ("Sheath contraceptives (Condoms)", "98020055", "DOZ", "HOME"),
    ("Shellfish/Crustaceans (Shrimp, Lobster, Conch, etc.)(fresh, frozen)", "98020041", "KG", "HOME"),
    ("Shutters, blinds, inlc. Venetian blinds, and similar articles", "39253000", "NO", "HOME"),
    ("Smoking tobacco", "98020035", "LB", "HOME"),
    ("SOAPS, Shampoo, soap cleansers", "33051000", "KG", "HOME"),
    ("Sodas", "98020057", "L", "HOME"),
    ("Sparkling wine", "98020025", "L", "HOME"),
    ("Spirits - Unsweetened greater than  50 % vol.", "98020029", "L", "HOME"),
    ("Spirits - Unsweetened less than 50 % vol.", "98020030", "L", "HOME"),
    ("Spirits, not exceeding 1 litre per importer aged 18 years or more", "98020016", "L", "HOME"),
    ("Sports equipment", "95069100", "NO", "HOME"),
    ("Statuettes and other ornamental articles, of plastic", "39264000", "KG", "HOME"),
    ("Sunglasses", "90041000", "DOZ", "HOME"),
    ("Supplements", "98010027", "KG", "HOME"),
    ("Table Wines", "98020026", "L", "HOME"),
    ("Tableware and Kitchenware", "39241000", "KG", "HOME"),
    ("Tableware and kitchenware of porcelain and china", "98010055", "PCS", "HOME"),
    ("Tea", "09021000", "KG", "HOME"),
    ("Tobacco, not stemmed/stripped", "98020032", "LB", "HOME"),
    ("Tobacco, partly or wholly stemmed/stripped", "98020033", "LB", "HOME"),
    ("Toiletries", "98020053", "KG", "HOME"),
    ("Toilette waters, colognes, eau de parfum", "33030020", "KG", "HOME"),
    ("Toys", "98010052", "NO", "HOME"),
    ("TV's", "98020076", "NO", "HOME"),
    ("Umbrella", "66011000", "DOZ", "HOME"),
    ("Unspecified materials (plumbing fixtures, faucets, switches)", "98020088", "NO", "BLD MAT"),
    ("Used personal baggage", "98020089", "KG", "HOME"),
    ("Used personal effects", "98020090", "KG", "HOME"),
    ("Used personal effects and baggage of Passengers previously permanently resident abroad who are entering the Islands on a bona fide change of residence, that is, for a period exceeding twelve months", "98020011", "KG", "HOME"),
    ("Vaccines for human medicine", "98010014", "KG", "HOME"),
    ("Vaccines for veterinary medicine", "98010015", "KG", "HOME"),
    ("Vanilla", "98020045", "KG", "HOME"),
    ("Vegetables", "98010026", "KG", "HOME"),
    ("Vegetables (fresh and frozen only)", "98020042", "KG", "HOME"),
    ("Vehicle tyres (tires)", "98020080", "NO", "HOME"),
    ("Video Games and consoles", "98020077", "NO", "HOME"),
    ("Vitamins", "98010016", "KG", "HOME"),
    ("Watch bands", "91131000", "NO", "HOME"),
    ("Watches", "98010019", "NO", "HOME"),
    ("Watches/travel clocks (not including parts, 22%)", "98020068", "NO", "HOME"),
    ("Weave (Human or Synthetic)", "98020059", "KG", "HOME"),
    ("Whey", "04041000", "KG", "HOME"),
    ("Wigs (Of human hair)", "98020061", "KG", "HOME"),
    ("Wigs (Of synthetic textile materials)", "98020060", "NO", "HOME"),
    ("Wines, not exceeding 4 litres of  per importer aged 18 years or more", "98020017", "L", "HOME"),
    ("Wool knitting yarn", "51081000", "KG", "HOME"),
]


def truncate_desc(text, limit=50):
    """Truncate a description to `limit` chars at the nearest word boundary.
    COLS rejects overly long descriptions; the commodity code already
    carries the full legal text on the officer's side."""
    text = text.strip()
    if len(text) <= limit:
        return text
    trunc = text[:limit]
    last_space = trunc.rfind(" ")
    if last_space > 20:
        trunc = trunc[:last_space]
    return trunc


class _TreeValue:
    """Wrapper so _generate_xml can call row['cby'].get() on tree values."""
    __slots__ = ("_v",)
    def __init__(self, value=""):
        self._v = str(value) if value is not None else ""
    def get(self):
        return self._v
    def set(self, value):
        self._v = str(value) if value is not None else ""

# ==============================================================================
# SHARED HELPERS
# ==============================================================================
def money(value):
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0.00")

def money_str(value):
    return format(money(value), ".2f")

def clean_text(text):
    text = str(text).lower().strip()
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return " ".join(text.split())

def clean_dock_receipt(val):
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        return val_str[:-2]
    return val_str

# Build lookup dicts from BUILTIN_CODES (description -> (code, unit, proc))
_BUILTIN_DESC_MAP = {}   # cleaned desc -> (code, unit, proc)
_BUILTIN_CODE_MAP = {}   # code -> (desc, unit, proc)
for _desc, _code, _unit, _proc in BUILTIN_CODES:
    _clean = clean_text(_desc)
    _BUILTIN_DESC_MAP[_clean] = (_code, _unit, _proc)
    _BUILTIN_CODE_MAP[_code] = (_desc, _unit, _proc)

# ==============================================================================
# MODE CONFIGURATIONS
# ==============================================================================
OCEAN_CONFIG = {
    "mode": "ocean",
    "title": "Ocean XML Declaration Console",
    "subtitle": "Build Customs SAD XML files from manifest data",
    "excel_file": PARENT_DIR / "Create Declaration Files for Ocean" / "Input Ocean Manifest Data.xlsx",
    "output_folder": PARENT_DIR / "Create Declaration Files for Ocean" / "Output XML Files",
    "transport_mode": "SEA",
    "discharge_port": "KYGEC",
    "exporter_address": "8001 NW 79th Ave",
    "vessel_fixed": None,
    "vessels": [
        ("Caribe Navigator", "SBCN"),
        ("Caribe Voyager", "CBVG"),
        ("Caribe Legend", "CL"),
        ("Tropic Mariner", "CV"),
        ("Seaboard Sun", "SS"),
        ("Bf Fortaleza", "FORTA"),
    ],
    "header_fields": [
        ("awb", "AWB / Master BL #"),
        ("vessel", "Vessel Code"),
        ("weight_lb", "Total Weight (LB)"),
        ("departure", "Departure Date (YYYY-MM-DD)"),
        ("arrival", "Arrival Date (YYYY-MM-DD)"),
        ("weight_cf", "Total Weight (CF)"),
        ("voyage", "Voyage #"),
        ("pkg_count", "Master Pkg Count"),
        ("master_importer", "Master Importer #"),
    ],
    "col_defs": [
        ("cby", "CBY", 70),
        ("dock", "Dock Receipt#", 110),
        ("desc", "Description", 210),
        ("code", "Comm. Code", 90),
        ("unit", "Unit", 55),
        ("proc", "Procedure", 85),
        ("qty", "Qty", 45),
        ("value", "Value USD", 85),
        ("freight", "Freight USD", 90),
        ("insurance", "Insurance USD", 100),
        ("importer", "Importer #", 95),
        ("origin", "Origin", 55),
    ],
    "colors": {
        "bg":         "#0a2417",
        "panel":      "#0f3320",
        "input":      "#10381f",
        "border":     "#3cb371",
        "text":       "#e8f5e9",
        "light":      "#90ee90",
        "accent":     "#2e8b57",
        "accent_hover": "#3cb371",
        "danger":     "#7a1f1f",
        "danger_hover": "#a52a2a",
        "btn_text":   "#ffffff",
    },
    "house_weight_zero": False,
    "excel_sheet": 0,
    "excel_header_labels": {
        "manifest_date": "Manifest Date",
        "arrival_date":  "Date of Arrival",
        "master_bl":     "BL#",
        "pkg_count":     "No. Pcs. BG",
        "weight_cf":     "Total Weight CF",
        "weight_lb":     "Total Weight LB",
        "vessel_code":   "Vessel Code",
        "voyage_no":     "Voyage #",
    },
    "excel_dock_col": "Dock Receipt#",
    "cols_sheet": "COLS",
    "cols_header_row": 9,
    "cols_data_start_row": 10,
    "cols_header_label_col": 2,
    "cols_header_value_col": 3,
    "cols_column_map": {
        "cby": 1,
        "dock": 0,
        "desc": 4,
        "value": 7,
        "freight": 8,
        "insurance": 9,
        "qty": 5,
        "importer": None,
    },
    "cols_header_labels": {
        "manifest_date": "Manifest Date",
        "arrival_date":  "Date of Arrival",
        "master_bl":     "BL#",
        "pkg_count":     "No. Pcs.",
        "weight_cf":     "Total Weight CF",
        "weight_lb":     None,
        "vessel_code":   None,
        "voyage_no":     None,
    },
    "commodity_sheet": "Commodity Codes",
    "commodity_desc_col": 1,
    "commodity_code_col": 2,
    "commodity_unit_col": None,
    "commodity_proc_col": None,
}

AIR_CONFIG = {
    "mode": "air",
    "title": "Air XML Declaration Console",
    "subtitle": "Build Customs SAD XML files from air manifest data",
    "excel_file": PARENT_DIR / "Create Declaration Files for Air" / "Input Air Manifest Data.xlsx",
    "output_folder": PARENT_DIR / "Create Declaration Files for Air" / "Output XML Files",
    "transport_mode": "AIR",
    "discharge_port": "KYGCM",
    "exporter_address": "2250 NW 114th Ave",
    "vessel_fixed": "KX",
    "vessels": [],
    "header_fields": [
        ("awb", "AWB #"),
        ("voyage", "KX Flight Number"),
        ("weight_lb", "Gross Weight (LB)"),
        ("departure", "Departure Date (YYYY-MM-DD)"),
        ("arrival", "Arrival Date (YYYY-MM-DD)"),
        ("weight_cf", "Gross Volume (CF)"),
        ("pkg_count", "# Bags/Pcs"),
        ("master_importer", "Master Importer #"),
    ],
    "col_defs": [
        ("cby", "CBY", 70),
        ("dock", "Package #", 110),
        ("desc", "Description", 210),
        ("code", "Comm. Code", 90),
        ("unit", "Unit", 55),
        ("proc", "Procedure", 85),
        ("qty", "Qty", 45),
        ("value", "Value USD", 85),
        ("freight", "Freight USD", 90),
        ("insurance", "Insurance USD", 100),
        ("importer", "Importer #", 95),
        ("origin", "Origin", 55),
    ],
    "colors": {
        "bg":         "#1a0a1e",
        "panel":      "#2a1030",
        "input":      "#3a1840",
        "border":     "#d946b5",
        "text":       "#fce4ec",
        "light":      "#f48fb1",
        "accent":     "#c2185b",
        "accent_hover": "#e91e63",
        "danger":     "#7a1f1f",
        "danger_hover": "#a52a2a",
        "btn_text":   "#ffffff",
    },
    "house_weight_zero": True,
    "excel_sheet": "Create",
    "excel_header_labels": {
        "manifest_date": "Manifest Date",
        "arrival_date":  "Date of Arrival",
        "master_bl":     "AWB#",
        "pkg_count":     "# Bags/Pcs AWB",
        "weight_cf":     "GrossVol CF",
        "weight_lb":     "GrossWt LB",
        "vessel_code":   None,
        "voyage_no":     "KX Flight Number",
    },
    "excel_dock_col": "Package#",
    "cols_sheet": " COLS",
    "cols_header_row": 8,
    "cols_data_start_row": 9,
    "cols_header_label_col": 3,
    "cols_header_value_col": 4,
    "cols_column_map": {
        "cby": 2,
        "dock": 16,
        "desc": 6,
        "value": 8,
        "freight": 9,
        "insurance": 10,
        "qty": None,
        "importer": 4,
    },
    "cols_header_labels": {
        "manifest_date": "Manifest Date",
        "arrival_date":  "Date of Arrival",
        "master_bl":     "AWB#",
        "pkg_count":     "# Bags/Pcs AWB",
        "weight_cf":     None,
        "weight_lb":     None,
        "vessel_code":   None,
        "voyage_no":     None,
    },
    "commodity_sheet": "Commodity Codes",
    "commodity_desc_col": 1,
    "commodity_code_col": 2,
    "commodity_unit_col": 3,
    "commodity_proc_col": None,
}

CONFIGS = {"ocean": OCEAN_CONFIG, "air": AIR_CONFIG}


# ==============================================================================
# XML STRUCTURE BUILDER  (parameterised by mode config)
# ==============================================================================
def build_sad_structure(cfg, hdr, bill_type, bill_no, content_text,
                        pkg_count, pkg_type="BG",
                        gross_wt=None, gross_vol=None):
    root = ET.Element("SADEntry")
    ET.SubElement(root, "Date").text = hdr["arrival"]
    ET.SubElement(root, "Regime").text = "IM1"

    exporter = ET.SubElement(root, "Exporter")
    ET.SubElement(exporter, "Name").text = "E-box Logistics"
    ET.SubElement(exporter, "Address").text = cfg["exporter_address"]
    ET.SubElement(exporter, "City").text = "Miami"
    ET.SubElement(exporter, "State").text = "FL"
    ET.SubElement(exporter, "Country").text = "USA"
    ET.SubElement(exporter, "Phone").text = "3457451400"

    consignment = ET.SubElement(root, "Consignment")
    ET.SubElement(consignment, "DepartureDate").text = hdr["departure"]
    ET.SubElement(consignment, "ArrivalDate").text = hdr["arrival"]
    ET.SubElement(consignment, "ExportCountry").text = "USA"
    ET.SubElement(consignment, "ImportCountry").text = "CYM"
    ET.SubElement(consignment, "ShippingPort").text = "USMIA"
    ET.SubElement(consignment, "DischargePort").text = cfg["discharge_port"]
    ET.SubElement(consignment, "TransportMode").text = cfg["transport_mode"]

    shipment = ET.SubElement(root, "Shipment")
    vessel_code = cfg["vessel_fixed"] or hdr["vessel"]
    ET.SubElement(shipment, "VesselCode").text = vessel_code
    ET.SubElement(shipment, "VoyageNo").text = str(hdr["voyage"])
    ET.SubElement(shipment, "ShippingAgent").text = "KX"
    ET.SubElement(shipment, "ShippingAgentName").text = "KX"
    ET.SubElement(shipment, "BillNumber").text = bill_no
    ET.SubElement(shipment, "BillType").text = bill_type
    if bill_type == "HOUSE":
        ET.SubElement(shipment, "MasterBillNumber").text = hdr["awb"]

    packages = ET.SubElement(root, "Packages")
    ET.SubElement(packages, "PkgCount").text = str(pkg_count)
    ET.SubElement(packages, "PkgType").text = pkg_type
    wt = gross_wt if gross_wt is not None else hdr["weight_lb"]
    vol = gross_vol if gross_vol is not None else hdr["weight_cf"]
    ET.SubElement(packages, "GrossWt").text = money_str(wt)
    ET.SubElement(packages, "GrossWtUnit").text = "LB"
    ET.SubElement(packages, "GrossVol").text = money_str(vol)
    ET.SubElement(packages, "GrossVolUnit").text = "CF"
    ET.SubElement(packages, "Contents").text = content_text

    return root


# ==============================================================================
# LAUNCHER WINDOW  (small dark picker)
# ==============================================================================
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

LAUNCHER_BG = "#1a1a2e"
LAUNCHER_PANEL = "#16213e"
LAUNCHER_ACCENT = "#0f3460"
LAUNCHER_TEXT = "#e0e0e0"


# ==============================================================================
# SUPPORT MIXIN — bug report + self-update functionality shared by all windows
# ==============================================================================
class SupportMixin:
    """Provides bug-reporting and self-update features to any window class
    that has a ``self.win`` (CTkToplevel), ``self._support_bg`` (bg color
    string), and ``self._window_name`` (identifying label for bug reports).

    Call ``_set_support_palette(colors)`` in ``__init__`` to theme the
    support dialogs to match the window's own colour scheme."""

    # ---- Palette setup --------------------------------------------------
    def _set_support_palette(self, colors):
        """Populate instance-level dialog colours from a *colors* dict
        (keys: bg, input, border, text, light, accent, accent_hover).
        Falls back to sensible defaults for missing keys."""
        bg       = colors.get("bg", "#1a1a2e")
        panel    = colors.get("panel", bg)
        inp      = colors.get("input", "#0f0f1a")
        border   = colors.get("border", "#333333")
        text     = colors.get("text", "#e8e8e8")
        light    = colors.get("light", "#aaaaaa")
        accent   = colors.get("accent", "#2e8b57")
        accent_h = colors.get("accent_hover", accent)

        self._SUP_DLG_BG       = panel
        self._SUP_DLG_LIGHT    = text
        self._SUP_DLG_MUTED    = light
        self._SUP_DLG_INPUT    = inp
        self._SUP_DLG_BORDER   = border
        self._SUP_DLG_ACCENT   = accent
        self._SUP_DLG_ACCENT_H = accent_h
        self._SUP_DLG_GREEN    = accent
        self._SUP_DLG_GREEN_H  = accent_h
        # Info-box: dark background (input colour) so light text is readable,
        # with the accent colour used only for the heading
        self._SUP_DLG_INFO_BG  = inp
        self._SUP_DLG_INFO_HDR = accent
        self._SUP_DLG_INFO_TXT = text

    # ---- Tooltip helpers ------------------------------------------------
    def _show_tooltip(self, widget, text):
        self._hide_tooltip()
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + 18
        line_count = text.count("\n") + 1
        est_height = line_count * 16 + 16
        screen_h = widget.winfo_screenheight()
        if y + est_height > screen_h - 20:
            y = widget.winfo_rooty() - est_height - 8
        tw = tk.Toplevel(self.win)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=text, justify="left",
                         bg=self._SUP_DLG_BG, fg=self._SUP_DLG_LIGHT,
                         relief="solid", bd=1, padx=10, pady=8,
                         font=(MODERN_FONT, 10), wraplength=320)
        label.pack()
        self._tooltip_win = tw

    def _hide_tooltip(self):
        if self._tooltip_win is not None:
            try:
                self._tooltip_win.destroy()
            except Exception:
                pass
            self._tooltip_win = None

    # ---- Support icon ---------------------------------------------------
    def _attach_support_icon(self, parent_frame):
        """Create the bug / apply-fixes icon button in *parent_frame*."""
        self._support_btn = ctk.CTkButton(
            parent_frame, text="\U0001f41e", width=34, height=28,
            fg_color=self._support_bg, hover_color="#24507a",
            corner_radius=6,
            font=("Segoe UI Emoji", 15),
            command=lambda: self._on_support_click(self._window_name))
        self._support_btn.pack(side="right")
        self._support_btn.bind("<Enter>",
            lambda e: self._show_tooltip(self._support_btn, self._support_tooltip))
        self._support_btn.bind("<Leave>",
            lambda e: self._hide_tooltip())

    # ---- Background update check ---------------------------------------
    def _check_update_bg(self):
        result = _check_for_update()
        if result:
            self._pending_update = result
            try:
                self.win.after(0, self._refresh_support_icon)
            except Exception:
                pass

    def _refresh_support_icon(self):
        try:
            if self._pending_update:
                ver = self._pending_update.get("version", "?")
                self._support_btn.configure(
                    text="Apply Fixes", fg_color="#b8860b",
                    hover_color="#daa520",
                    font=(MODERN_FONT, 11, "bold"), width=90)
                self._support_tooltip = f"Apply Fixes \u2014 v{ver} available"
            else:
                self._support_btn.configure(
                    text="\U0001f41e", fg_color=self._support_bg,
                    hover_color="#24507a",
                    font=("Segoe UI Emoji", 15), width=34)
                self._support_tooltip = "Report a Bug"
        except Exception:
            pass

    # ---- Click handler --------------------------------------------------
    def _on_support_click(self, window_name):
        self._hide_tooltip()
        if self._pending_update:
            self._apply_fixes_dialog()
        else:
            self._report_bug_dialog(window_name)

    # ---- Bug report dialog ---------------------------------------------
    def _report_bug_dialog(self, window_name):
        dlg = ctk.CTkToplevel(self.win)
        dlg.title(f"Report a Bug \u2014 {window_name}")
        dlg.configure(fg_color=self._SUP_DLG_BG)
        dlg.transient(self.win)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 640
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(dlg, text=f"Report a Bug \u2014 {window_name}",
                     font=(MODERN_FONT, 15, "bold"),
                     text_color=self._SUP_DLG_LIGHT).pack(
            anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Describe the problem in as much detail as you can \u2014 "
                 f"what you\ndid, what happened, and what you expected. This "
                 f"goes directly\nto the developer ({DEVELOPER_NAME}).",
            font=(MODERN_FONT, 11), text_color=self._SUP_DLG_MUTED,
            anchor="w", justify="left").pack(
            anchor="w", padx=16, pady=(0, 6))

        # ---- Check for Updates bar ----
        update_bar = ctk.CTkFrame(dlg, fg_color="transparent")
        update_bar.pack(fill="x", padx=16, pady=(0, 6))

        check_btn = ctk.CTkButton(
            update_bar, text="Check for Updates",
            command=lambda: None,  # set below
            fg_color=self._SUP_DLG_ACCENT, hover_color=self._SUP_DLG_ACCENT_H,
            width=140, height=26, corner_radius=4,
            font=(MODERN_FONT, 10, "bold"))
        check_btn.pack(side="left")

        update_status = ctk.CTkLabel(
            update_bar, text="",
            font=(MODERN_FONT, 10), text_color=self._SUP_DLG_MUTED,
            anchor="w")
        update_status.pack(side="left", padx=(10, 0))

        # Spinner animation state
        _spin = {"frame": 0, "running": False}

        def _spin_step():
            if not _spin["running"]:
                return
            _spin["frame"] = (_spin["frame"] + 1) % 8
            dots = "." * ((_spin["frame"] % 4) + 1)
            update_status.configure(text=f"Checking{dots}")
            dlg.after(200, _spin_step)

        def _do_check():
            _spin["running"] = True
            check_btn.configure(state="disabled", text="Checking...")
            _spin_step()

            def _worker():
                result = _check_for_update()
                dlg.after(0, lambda: _check_done(result))

            threading.Thread(target=_worker, daemon=True).start()

        def _check_done(result):
            _spin["running"] = False
            check_btn.configure(state="normal", text="Check for Updates")
            if result:
                ver = result.get("version", "?")
                self._pending_update = result
                update_status.configure(
                    text=f"v{ver} available!",
                    text_color=self._SUP_DLG_LIGHT)
                # Replace the check button with an "Update Now" button
                check_btn.configure(
                    text=f"Update Now \u2014 v{ver}",
                    command=self._apply_fixes_dialog,
                    fg_color="#b8860b", hover_color="#daa520",
                    width=180)
            else:
                update_status.configure(
                    text=f"You're up to date (v{APP_VERSION})",
                    text_color=self._SUP_DLG_MUTED)

        check_btn.configure(command=_do_check)

        # Category selector
        cat_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        cat_frame.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(cat_frame, text="Category:",
                     font=(MODERN_FONT, 11, "bold"),
                     text_color=self._SUP_DLG_LIGHT,
                     anchor="w").pack(anchor="w", pady=(0, 2))
        cat_var = ctk.StringVar(value="Bug Fix")
        cat_row = ctk.CTkFrame(cat_frame, fg_color="transparent")
        cat_row.pack(fill="x")
        for cat in ("Bug Fix", "Feature Request",
                     "Environmental Change", "Other"):
            ctk.CTkRadioButton(
                cat_row, text=cat, variable=cat_var, value=cat,
                font=(MODERN_FONT, 10), text_color=self._SUP_DLG_LIGHT,
                fg_color=self._SUP_DLG_ACCENT,
                hover_color=self._SUP_DLG_ACCENT_H,
                border_color=self._SUP_DLG_BORDER).pack(
                side="left", padx=(0, 12))

        box = ctk.CTkTextbox(
            dlg, height=140, fg_color=self._SUP_DLG_INPUT,
            border_color=self._SUP_DLG_BORDER, border_width=1,
            corner_radius=4, text_color=self._SUP_DLG_LIGHT,
            font=(MODERN_FONT, 11))
        box.pack(fill="both", expand=True, padx=16)
        box.focus_set()

        # Optional email field
        email_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        email_frame.pack(fill="x", padx=16, pady=(4, 0))
        ctk.CTkLabel(
            email_frame,
            text="Your email (optional \u2014 for updates on this report)",
            font=(MODERN_FONT, 10), text_color=self._SUP_DLG_MUTED,
            anchor="w").pack(anchor="w")
        email_var = ctk.StringVar(value="")
        email_entry = ctk.CTkEntry(
            email_frame, textvariable=email_var, height=28,
            fg_color=self._SUP_DLG_INPUT, border_color=self._SUP_DLG_BORDER,
            border_width=1, corner_radius=4, text_color=self._SUP_DLG_LIGHT,
            font=(MODERN_FONT, 11))
        email_entry.pack(fill="x", pady=(1, 0))

        # "What to expect" info box
        info = ctk.CTkFrame(dlg, fg_color=self._SUP_DLG_INFO_BG, corner_radius=6)
        info.pack(fill="x", padx=16, pady=(6, 8))
        ctk.CTkLabel(info, text="What to expect",
                     font=(MODERN_FONT, 11, "bold"),
                     text_color=self._SUP_DLG_INFO_HDR,
                     anchor="w").pack(anchor="w", padx=10, pady=(8, 2))
        ctk.CTkLabel(info,
            text=f"\u2022 {DEVELOPER_NAME} will review your report within "
                 f"24-48 hours\n"
                 f"\u2022 When a fix is ready, click 'Check for Updates' above\n"
                 f"   or relaunch the console \u2014 the bug icon will show "
                 f"'Apply Fixes'\n"
                 f"\u2022 {DEVELOPER_NAME} will coordinate with management if "
                 f"the fix\n"
                 f"   requires substantial work\n"
                 f"\u2022 Write down your case number for follow-up: contact "
                 f"{DEVELOPER_EMAIL}",
            font=(MODERN_FONT, 10), text_color=self._SUP_DLG_INFO_TXT,
            anchor="w", justify="left").pack(
            anchor="w", padx=10, pady=(0, 8))

        # Portal summon icon (remote support) in the button row
        _portal_canvas = tk.Canvas(btns, width=24, height=24,
                                    bg=self._SUP_DLG_BG, highlightthickness=0)
        _portal_canvas.pack(side="right")
        _portal_canvas.create_oval(2, 2, 22, 22, outline="#a0d8a0", width=2)
        _portal_canvas.create_oval(7, 7, 17, 17, fill="#a0d8a0", outline="")
        _portal_canvas.configure(cursor="hand2")
        _portal_canvas.bind("<Button-1>", lambda e: (dlg.destroy(), _summon_portal(self.win)))

        def _next():
            desc = box.get("0.0", "end").strip()
            if not desc:
                messagebox.showwarning("Empty",
                                       "Please describe the bug first.")
                return
            email = email_var.get().strip()
            category = cat_var.get()
            dlg.destroy()
            self._show_attach_files_dialog(
                desc, email, category, window_name)

        ctk.CTkButton(btns, text="Next", command=_next,
                      fg_color=self._SUP_DLG_GREEN,
                      hover_color=self._SUP_DLG_GREEN_H, width=100,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        ctk.CTkButton(btns, text="Cancel", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    # ---- Attach files dialog -------------------------------------------
    def _show_attach_files_dialog(self, description, reporter_email="",
                                  category="Bug Fix", window_name=""):
        """After the bug description, offer to attach sample files.
        The actual Discord send happens here \u2014 text + files in ONE message."""
        case_num = _generate_case_number()
        dlg = ctk.CTkToplevel(self.win)
        dlg.title(f"Attach Files? (Case {case_num})")
        dlg.configure(fg_color=self._SUP_DLG_BG)
        dlg.transient(self.win)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 420
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(dlg, text="Attach Sample Files?",
                     font=(MODERN_FONT, 15, "bold"),
                     text_color=self._SUP_DLG_LIGHT).pack(
            anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Based on your description, {DEVELOPER_NAME} may need "
                 f"sample\nfiles to reproduce the issue in a test "
                 f"environment.\n\n"
                 f"The console script itself is always attached "
                 f"automatically\nso {DEVELOPER_NAME} gets your exact version "
                 f"for testing.\n\n"
                 f"If you'd like, attach additional Excel or CSV files "
                 f"below.\nThis is completely optional and files are sent "
                 f"privately\nalongside your bug report.",
            font=(MODERN_FONT, 11), text_color=self._SUP_DLG_MUTED,
            anchor="w", justify="left").pack(
            anchor="w", padx=16, pady=(0, 8))

        file_list_frame = ctk.CTkFrame(
            dlg, fg_color=self._SUP_DLG_INPUT, corner_radius=4,
            border_width=1, border_color=self._SUP_DLG_BORDER)
        file_list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        file_list_label = ctk.CTkLabel(
            file_list_frame, text="No files selected",
            font=(MODERN_FONT, 11), text_color=self._SUP_DLG_MUTED,
            anchor="w", justify="left")
        file_list_label.pack(anchor="w", padx=10, pady=10)

        selected_files = []

        def _pick_files():
            paths = filedialog.askopenfilenames(
                title="Select files to attach",
                filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv"),
                           ("All files", "*.*")])
            if paths:
                selected_files.clear()
                selected_files.extend(paths)
                names = [Path(p).name for p in paths]
                display = "\n".join(names[:6])
                if len(names) > 6:
                    display += f"\n...and {len(names)-6} more"
                file_list_label.configure(
                    text=display, text_color=self._SUP_DLG_LIGHT)

        def _submit():
            submit_btn.configure(state="disabled", text="Sending...")
            skip_btn.configure(state="disabled")

            def _worker():
                all_files = [str(SCRIPT_PATH)] + list(selected_files)
                ok, err = _post_bug_report_with_files(
                    description, case_num, all_files,
                    reporter_email, category, window_name)
                self.win.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                dlg.destroy()
                n_files = len(selected_files)
                if n_files:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted with {n_files} "
                        f"file(s).\n\n"
                        f"Your report and files have been sent to "
                        f"{DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' "
                        "here.")
                else:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted.\n\n"
                        f"Your report has been sent to "
                        f"{DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' "
                        "here.")
            else:
                submit_btn.configure(state="normal", text="Submit Report")
                skip_btn.configure(state="normal")
                messagebox.showerror("Could Not Send",
                    f"The report could not be sent:\n{err}\n\n"
                    "Check your internet connection and try again.")

        ctk.CTkButton(btns, text="Browse Files", command=_pick_files,
                      fg_color=self._SUP_DLG_ACCENT,
                      hover_color=self._SUP_DLG_ACCENT_H, width=120,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        submit_btn = ctk.CTkButton(
            btns, text="Submit Report", command=_submit,
            fg_color=self._SUP_DLG_GREEN, hover_color=self._SUP_DLG_GREEN_H,
            width=130, height=30, corner_radius=5,
            font=(MODERN_FONT, 11, "bold"))
        submit_btn.pack(side="left", padx=(8, 0))
        skip_btn = ctk.CTkButton(
            btns, text="Skip & Send", command=_submit,
            fg_color="#667788", hover_color="#556677", width=110,
            height=30, corner_radius=5, font=(MODERN_FONT, 11))
        skip_btn.pack(side="left", padx=(8, 0))

    # ---- Apply fixes dialog --------------------------------------------
    def _apply_fixes_dialog(self):
        upd = self._pending_update
        if not upd:
            return

        dlg = ctk.CTkToplevel(self.win)
        dlg.title("Apply Fixes")
        dlg.configure(fg_color=self._SUP_DLG_BG)
        dlg.transient(self.win)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 380
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ver = upd.get("version", "?")
        ctk.CTkLabel(dlg, text=f"Fixes Available  (v{ver})",
                     font=(MODERN_FONT, 15, "bold"),
                     text_color=self._SUP_DLG_LIGHT).pack(
            anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(
            dlg,
            text=f"From {DEVELOPER_NAME}. Click Apply Fixes to download,\n"
                 f"then choose how to restart.",
            font=(MODERN_FONT, 11), text_color=self._SUP_DLG_MUTED,
            anchor="w").pack(anchor="w", padx=16, pady=(0, 8))

        box = ctk.CTkTextbox(
            dlg, height=170, fg_color=self._SUP_DLG_INPUT,
            border_color=self._SUP_DLG_BORDER, border_width=1,
            corner_radius=4, text_color=self._SUP_DLG_LIGHT,
            font=(MODERN_FONT, 11), wrap="word")
        box.pack(fill="both", expand=True, padx=16)
        box.insert("0.0", upd.get("changelog", "(no description provided)"))
        box.configure(state="disabled")

        def _apply():
            apply_btn.configure(state="disabled", text="Applying...")

            def _worker():
                ok, err = _download_and_apply_update(upd.get("url", ""))
                self.win.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                old_ver = APP_VERSION
                new_ver = upd.get("version", "?")
                threading.Thread(
                    target=lambda: _post_update_applied(old_ver, new_ver),
                    daemon=True).start()
                # Clear the pending update so the bug icon goes back
                # to 🐞 — the update is on disk, they just need to
                # restart.  This way they can still report a new bug
                # if they click "Later" instead of restarting.
                self._pending_update = None
                try:
                    self.win.after(0, self._refresh_support_icon)
                except Exception:
                    pass
                dlg.destroy()
                # Show a restart-choice dialog with two options:
                # "Restart Only" and "Restart + Restore Progress".
                self._show_restart_choice(new_ver)
            else:
                apply_btn.configure(state="normal", text="Apply Fixes")
                messagebox.showerror("Update Failed",
                    f"Could not apply the update:\n{err}\n\n"
                    "Check your internet connection and try again.")

        apply_btn = ctk.CTkButton(
            btns, text="Apply Fixes", command=_apply,
            fg_color="#b8860b", hover_color="#daa520",
            width=130, height=30, corner_radius=5,
            font=(MODERN_FONT, 11, "bold"))
        apply_btn.pack(side="left")
        ctk.CTkButton(btns, text="Later", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    def _show_restart_choice(self, new_ver):
        """After an update is downloaded, ask the user how to restart.
        Offers two options: plain restart, or restart with session
        preservation (saves UploadWindow treeview + decl numbers)."""
        dlg = ctk.CTkToplevel(self.win)
        dlg.title("Restart to Apply")
        dlg.configure(fg_color=self._SUP_DLG_BG)
        dlg.transient(self.win)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 480, 220
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        ctk.CTkLabel(dlg, text=f"Updated to v{new_ver}",
                     font=(MODERN_FONT, 15, "bold"),
                     text_color=self._SUP_DLG_LIGHT).pack(
            anchor="w", padx=20, pady=(18, 4))
        ctk.CTkLabel(dlg,
            text="The console needs to restart to apply the changes.\n"
                 "Choose how you'd like to restart:",
            font=(MODERN_FONT, 11), text_color=self._SUP_DLG_MUTED,
            anchor="w", justify="left").pack(anchor="w", padx=20, pady=(0, 12))

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=20, pady=(0, 16))

        def _restart_only():
            dlg.destroy()
            self._restart_app(preserve_session=False)

        def _restart_preserve():
            dlg.destroy()
            self._restart_app(preserve_session=True)

        ctk.CTkButton(btns, text="Restart Only",
                      command=_restart_only,
                      fg_color="#2471a3", hover_color="#3498db",
                      width=120, height=32, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        ctk.CTkButton(btns, text="Restart + Restore Progress",
                      command=_restart_preserve,
                      fg_color="#b8860b", hover_color="#daa520",
                      width=200, height=32, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=(8, 0))
        ctk.CTkButton(btns, text="Later",
                      command=dlg.destroy,
                      fg_color="#444444", hover_color="#555555",
                      width=70, height=32, corner_radius=6,
                      font=(MODERN_FONT, 11)).pack(side="right")

    # ---- Session save / restore (for live bug-fix restarts) ------------
    # When the user clicks "Restart + Restore Progress", we serialize the
    # UploadWindow's treeview rows + declaration numbers to a temp JSON
    # file with "used": false.  On the next launch, UploadWindow.__init__
    # calls _restore_session() which:
    #   1. Reads the JSON
    #   2. Checks "used" is false and the file is < 24 hours old
    #   3. Restores the data into the treeview + dicts
    #   4. Marks "used": true and writes it back
    #   5. Tries to delete the file
    #
    # The "used" flag is the primary one-time-use guarantee: even if the
    # file can't be deleted (locked temp dir, antivirus, strict perms),
    # it will never be restored again because "used" is true.
    #
    # The 24-hour timestamp is a safety net for edge cases like a crash
    # during restart — the file survives with "used": false, and the
    # next successful launch picks it up.  But after 24 hours it's
    # considered stale and ignored, so it can't haunt someone days later.

    _SESSION_FILE = "mbe_session_restore.json"
    _SESSION_EXPIRY_SECONDS = 86400  # 24 hours (safety net only)

    def _save_session_for_restore(self):
        """Serialize the UploadWindow's treeview + decl data to a temp
        JSON file so it can be restored after a restart-restart update.
        Cross-platform: uses tempfile.gettempdir()."""
        import tempfile, os, json, datetime

        # Find the UploadWindow via the launcher — it holds the
        # declaration numbers and treeview that we want to preserve.
        upload_win = None
        try:
            upload_win = self.launcher._upload_win
        except Exception:
            pass
        if not upload_win:
            return False

        try:
            rows = []
            for item in upload_win._tree.get_children():
                vals = upload_win._tree.set(item)
                rows.append({
                    "file":   vals.get("file", ""),
                    "cby":    vals.get("cby", ""),
                    "status": vals.get("status", ""),
                    "docs":   vals.get("docs", ""),
                    "decl":   vals.get("decl", ""),
                })

            session = {
                "written_at": datetime.datetime.now().isoformat(),
                "used": False,
                "active_window": self._window_name,
                "rows": rows,
                "decl_numbers": dict(upload_win._decl_numbers),
                "master_decl": upload_win._master_decl,
                "xml_folder": str(upload_win._xml_folder) if upload_win._xml_folder else None,
            }

            path = os.path.join(tempfile.gettempdir(), self._SESSION_FILE)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(session, f)
            return True
        except Exception:
            return False

    def _restore_session(self):
        """Check for a session-restore JSON in the temp folder.
        If it exists, is unused ("used": false), and is < 24 hours old,
        restore treeview rows and declaration data.  Then mark it as
        used and try to delete the file.
        Returns True if data was restored, False otherwise."""
        import tempfile, os, json, datetime

        path = os.path.join(tempfile.gettempdir(), self._SESSION_FILE)
        if not os.path.exists(path):
            return False

        try:
            with open(path, "r", encoding="utf-8") as f:
                session = json.load(f)
        except Exception:
            # Corrupted file — try to delete, move on.
            try:
                os.remove(path)
            except Exception:
                pass
            return False

        # Check "used" flag — primary one-time-use guarantee.
        # If already used, this file can never be restored again,
        # even if it couldn't be deleted last time.
        if session.get("used", False):
            try:
                os.remove(path)
            except Exception:
                pass
            return False

        # Safety net: reject anything older than 24 hours.  This handles
        # the case where a file persists for days (locked temp dir) —
        # we don't want stale data from last week reappearing.
        try:
            written = datetime.datetime.fromisoformat(session["written_at"])
            age = (datetime.datetime.now() - written).total_seconds()
            if age > self._SESSION_EXPIRY_SECONDS:
                try:
                    os.remove(path)
                except Exception:
                    pass
                return False
        except Exception:
            try:
                os.remove(path)
            except Exception:
                pass
            return False

        # Restore treeview rows
        restored_count = 0
        try:
            for row in session.get("rows", []):
                self._tree.insert("", "end", values=(
                    row.get("file", ""),
                    row.get("cby", ""),
                    row.get("status", ""),
                    row.get("docs", ""),
                    row.get("decl", ""),
                ))
                restored_count += 1
        except Exception:
            pass

        # Restore declaration data
        try:
            self._decl_numbers = dict(session.get("decl_numbers", {}))
        except Exception:
            self._decl_numbers = {}
        try:
            self._master_decl = session.get("master_decl", "")
        except Exception:
            self._master_decl = ""
        try:
            folder = session.get("xml_folder")
            if folder:
                from pathlib import Path
                self._xml_folder = Path(folder)
        except Exception:
            pass

        # Mark as used and write back — this is the one-time-use lock.
        # Even if the delete fails, the "used": true flag prevents any
        # future session from restoring the same data again.
        try:
            session["used"] = True
            with open(path, "w", encoding="utf-8") as f:
                json.dump(session, f)
        except Exception:
            pass

        # Now try to delete the file.  If it fails, the "used" flag
        # above guarantees it won't be used again.
        try:
            os.remove(path)
        except Exception:
            pass

        return restored_count > 0

    # ---- Restart --------------------------------------------------------
    def _restart_app(self, preserve_session=False):
        if preserve_session:
            self._save_session_for_restore()
        try:
            subprocess.Popen([sys.executable, str(SCRIPT_PATH)])
        except Exception:
            pass
        try:
            self.win.destroy()
        except Exception:
            pass
        # Also destroy the launcher root if accessible
        try:
            if hasattr(self, 'launcher') and self.launcher:
                self.launcher.root.destroy()
        except Exception:
            pass
        sys.exit(0)


class Launcher:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.report_callback_exception = _tk_global_exception_handler
        self.root.title(f"XML Declaration Builder v{APP_VERSION}")
        self.root.configure(fg_color=LAUNCHER_BG)
        _register_window_name(self.root, "Launcher", {"bg": LAUNCHER_BG, "panel": LAUNCHER_PANEL, "accent": LAUNCHER_ACCENT, "accent_hover": "#1a4a7a", "text": "#e8e8e8"})
        w, h = 420, 370
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{int((sw-w)/2)}+{int((sh-h)/2)}")
        self.root.resizable(False, False)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Logo
        try:
            _img = Image.open(io.BytesIO(base64.b64decode(MBE_LOGO_B64)))
            _lw, _lh = _img.size
            _tw = 90
            _th = max(1, int(_lh * _tw / _lw))
            _logo = ctk.CTkImage(light_image=_img, dark_image=_img, size=(_tw, _th))
            ctk.CTkLabel(self.root, image=_logo, text="").pack(pady=(20, 4))
        except Exception:
            pass

        ctk.CTkLabel(self.root, text="XML Declaration Builder",
                     font=(MODERN_FONT, 16, "bold"), text_color=LAUNCHER_TEXT).pack(pady=(4, 2))
        ctk.CTkLabel(self.root, text="Choose a shipment type to begin",
                     font=(MODERN_FONT, 11), text_color="#888888").pack(pady=(0, 16))

        btn_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        btn_frame.pack(pady=(0, 20))

        ctk.CTkButton(btn_frame, text="Build XML files for Ocean",
                      command=lambda: self._open("ocean"),
                      fg_color="#2e8b57", hover_color="#3cb371",
                      width=170, height=42, corner_radius=8,
                      font=(MODERN_FONT, 13, "bold")).pack(side="left", padx=10)

        ctk.CTkButton(btn_frame, text="Build XML files for Air",
                      command=lambda: self._open("air"),
                      fg_color="#c2185b", hover_color="#e91e63",
                      width=170, height=42, corner_radius=8,
                      font=(MODERN_FONT, 13, "bold")).pack(side="left", padx=10)

        # TIN Numbers + Codes buttons
        btn_row = ctk.CTkFrame(self.root, fg_color="transparent")
        btn_row.pack(pady=(0, 20))
        ctk.CTkButton(btn_row, text="TIN Numbers",
                      command=self._open_tin,
                      fg_color="#0f3460", hover_color="#1a4a7a",
                      width=120, height=28, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=5)
        ctk.CTkButton(btn_row, text="Item Codes",
                      command=self._open_codes,
                      fg_color="#0f3460", hover_color="#1a4a7a",
                      width=120, height=28, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=5)

        # Divider — separates the build/lookup tools from the standalone
        # upload action. corner_radius=0 so the bar renders as a crisp line
        # (rounded corners collapse a short bar and make it invisible), and a
        # taller/brighter bar so it stays visible under display scaling.
        ctk.CTkFrame(self.root, fg_color="#9aa0d4", height=3, width=300,
                     corner_radius=0).pack(pady=(18, 16))

        # Upload Declarations button (standalone)
        ctk.CTkButton(self.root, text="Upload Declarations to COLS",
                      command=self._open_upload,
                      fg_color="#6c3483", hover_color="#8e44ad",
                      width=280, height=36, corner_radius=8,
                      font=(MODERN_FONT, 12, "bold")).pack(pady=(0, 22))

        self._console = None
        self._codes_win = None
        self._tin_win = None
        self._upload_win = None

        # Check if we're restarting from a session restore — if so,
        # auto-open the window the user was in so they don't just see
        # the launcher and think they lost everything.
        self._check_session_restore()

    def _check_session_restore(self):
        """Check if a session-restore file exists and is unused.
        If so, auto-open the window the user was in so they don't
        just see the launcher and think they lost everything.
        The actual data restore happens in UploadWindow.__init__."""
        import tempfile, os, json, datetime
        path = os.path.join(tempfile.gettempdir(),
                            SupportMixin._SESSION_FILE)
        if not os.path.exists(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                session = json.load(f)
        except Exception:
            return
        if session.get("used", False):
            return
        # Check it's not stale (> 24 hours)
        try:
            written = datetime.datetime.fromisoformat(session["written_at"])
            age = (datetime.datetime.now() - written).total_seconds()
            if age > SupportMixin._SESSION_EXPIRY_SECONDS:
                return
        except Exception:
            return
        # Auto-open the window the user was in
        active = session.get("active_window", "")
        self.root.after(500, lambda: self._auto_open_window(active))

    def _auto_open_window(self, window_name):
        """Open the window that was active before the restart."""
        if "Upload" in window_name:
            self._open_upload()
        elif "Ocean" in window_name or "Air" in window_name:
            # We don't know which mode — default to ocean.  The console
            # window doesn't have session restore, so this is just a
            # convenience to get them back to roughly where they were.
            self._open("ocean")
        elif "TIN" in window_name:
            self._open_tin()
        elif "Codes" in window_name or "Item" in window_name:
            self._open_codes()

    def _open_tin(self):
        if self._tin_win is not None:
            try:
                self._tin_win.win.deiconify()
                self._tin_win.win.lift()
                return
            except Exception:
                pass
        self._tin_win = TINWindow(self)

    def _open_codes(self):
        if self._codes_win is not None:
            try:
                self._codes_win.win.deiconify()
                self._codes_win.win.lift()
                return
            except Exception:
                pass
        self._codes_win = CodesWindow(self)

    def _open_upload(self):
        if self._upload_win is not None:
            try:
                self._upload_win.win.deiconify()
                self._upload_win.win.lift()
                return
            except Exception:
                pass
        self.root.withdraw()
        self._upload_win = UploadWindow(self)

    def _open(self, mode):
        self.root.withdraw()
        cfg = CONFIGS[mode]
        self._console = ConsoleWindow(self, cfg)

    def show(self):
        self.root.deiconify()

    def _on_close(self):
        self.root.destroy()

    def run(self):
        self.root.mainloop()


# ==============================================================================
# CONSOLE WINDOW  (full editor – parameterised by mode config)
# ==============================================================================
class ConsoleWindow(SupportMixin):
    def __init__(self, launcher, cfg):
        self.launcher = launcher
        self.cfg = cfg
        c = cfg["colors"]

        # Support-mixin state
        self._pending_update = None
        self._support_tooltip = "Report a Bug"
        self._tooltip_win = None
        self._support_bg = c["bg"]
        self._window_name = f"XML Builder - {cfg['mode'].title()}"
        self._set_support_palette(c)

        self.win = ctk.CTkToplevel(launcher.root)
        self.win.title(f"{cfg['title']} v{APP_VERSION}")
        self.win.configure(fg_color=c["bg"])
        _register_window_name(self.win, self._window_name, c)
        WIN_W, WIN_H = 1290, 760
        sw, sh = self.win.winfo_screenwidth(), self.win.winfo_screenheight()
        self.win.geometry(f"{WIN_W}x{WIN_H}+{int((sw-WIN_W)/2)}+{max(0, int((sh-WIN_H)/2))}")
        self.win.minsize(1100, 560)
        self.win.protocol("WM_DELETE_WINDOW", self._close_all)

        self.header_entries = {}
        self._desc_to_code = {}
        self._code_to_unit = {}
        self._code_to_proc = {}
        self._manifest_dir = None  # set when Build from Manifest is used

        # Vessel display helpers
        self._vessel_display = [f"{name} ({code})" for name, code in cfg["vessels"]]
        self._code_to_display = {code.upper(): f"{name} ({code})" for name, code in cfg["vessels"]}

        self._build_ui()
        self._show_placeholder()
        self.win.after(100, self.win.focus_force)

        # Check for updates in the background
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    # -- UI construction --------------------------------------------------
    def _build_ui(self):
        c = self.cfg["colors"]
        win = self.win

        # ---- Title bar ----
        title_frame = ctk.CTkFrame(win, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=(16, 6))

        # Back button (left)
        ctk.CTkButton(title_frame, text="\u2190 Back",
                      command=self._go_back,
                      fg_color=c["accent"], hover_color=c["accent_hover"],
                      width=80, height=28, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold"),
                      text_color=c["btn_text"]).pack(side="left")

        # Logo (right)
        try:
            _img = Image.open(io.BytesIO(base64.b64decode(MBE_LOGO_B64)))
            _lw, _lh = _img.size
            _tw = 130
            _th = max(1, int(_lh * _tw / _lw))
            _logo = ctk.CTkImage(light_image=_img, dark_image=_img, size=(_tw, _th))
            ctk.CTkLabel(title_frame, image=_logo, text="").pack(side="right")
        except Exception:
            pass

        ctk.CTkLabel(title_frame, text=self.cfg["title"],
                     font=(MODERN_FONT, 20, "bold"), text_color=c["light"]).pack(side="left", padx=(16, 0))
        ctk.CTkLabel(title_frame, text=self.cfg["subtitle"],
                     font=(MODERN_FONT, 12), text_color=c["text"]).pack(side="left", padx=(14, 0))

        # ---- Header fields ----
        header_panel = ctk.CTkFrame(win, fg_color=c["panel"], corner_radius=8)
        header_panel.pack(fill="x", padx=20, pady=(0, 10))
        ctk.CTkLabel(header_panel, text="Shipment Header",
                     font=(MODERN_FONT, 14, "bold"), text_color=c["light"]).pack(anchor="w", padx=14, pady=(10, 4))

        header_grid = ctk.CTkFrame(header_panel, fg_color="transparent")
        header_grid.pack(fill="x", padx=10, pady=(0, 12))

        for idx, (key, label) in enumerate(self.cfg["header_fields"]):
            r, col = divmod(idx, 3)
            block = ctk.CTkFrame(header_grid, fg_color="transparent")
            block.grid(row=r, column=col, padx=8, pady=6, sticky="w")
            ctk.CTkLabel(block, text=label, width=170, anchor="w",
                         font=(MODERN_FONT, 11), text_color=c["text"]).pack(side="left")
            if key == "vessel" and self._vessel_display:
                widget = ctk.CTkComboBox(block, width=190, height=28, values=self._vessel_display,
                                         fg_color=c["input"], border_color=c["border"], border_width=1,
                                         corner_radius=5, text_color=c["text"],
                                         button_color=c["accent"], button_hover_color=c["accent_hover"],
                                         dropdown_fg_color=c["panel"], dropdown_text_color=c["text"],
                                         dropdown_hover_color=c["accent"])
                widget.set("")
            else:
                widget = ctk.CTkEntry(block, width=190, height=28, fg_color=c["input"],
                                      border_color=c["border"], border_width=1, corner_radius=5,
                                      text_color=c["text"])
            widget.pack(side="left")
            self.header_entries[key] = widget

        # Sensible defaults
        self._set_header("pkg_count", "1")
        self._set_header("weight_lb", "0.00")
        self._set_header("weight_cf", "0.00")
        self._set_header("master_importer", MASTER_IMPORTER_DEFAULT)
        if self.cfg.get("vessel_fixed") and "vessel" in self.header_entries:
            self._set_header("vessel", self.cfg["vessel_fixed"])
        if self.cfg["mode"] == "air" and "voyage" in self.header_entries:
            self._set_header("voyage", "909")

        # ---- Action bar ----
        action_frame = ctk.CTkFrame(win, fg_color="transparent")
        action_frame.pack(fill="x", padx=20, pady=(0, 8))

        self._make_btn(action_frame, "Build from Manifest", self._build_from_manifest, width=170).pack(side="left", padx=(0, 8))
        self._make_btn(action_frame, "+ Add Row", lambda: self._add_row(),
                       fg="#1f5e3a" if self.cfg["mode"] == "ocean" else "#a8630a",
                       hover="#2e8b57" if self.cfg["mode"] == "ocean" else "#cc7a14",
                       width=110).pack(side="left", padx=8)
        self._make_btn(action_frame, "Clear All",
                       lambda: (self._clear_all_rows(), self._show_placeholder()),
                       fg=c["danger"], hover=c["danger_hover"], width=100).pack(side="left", padx=8)
        self._make_btn(action_frame, "Delete Row",
                       self._remove_selected_row,
                       fg=c["danger"], hover=c["danger_hover"], width=100).pack(side="left", padx=8)
        self._make_btn(action_frame, "Generate XML Files", self._generate_xml, width=190).pack(side="right", padx=(8, 0))

        # ---- Items table header ----
        ctk.CTkLabel(win, text="Line Items  (grouped by CBY on output)  -  double-click a cell to edit",
                     font=(MODERN_FONT, 14, "bold"), text_color=c["light"]).pack(anchor="w", padx=22, pady=(4, 2))

        # ---- Treeview table ----
        tree_frame = ctk.CTkFrame(win, fg_color=c["panel"], corner_radius=6)
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(4, 8))

        col_keys = [k for k, _, _ in self.cfg["col_defs"]]

        self._tree = ttk.Treeview(tree_frame, columns=col_keys, show="headings",
                                  selectmode="browse")
        for key, label, w in self.cfg["col_defs"]:
            self._tree.heading(key, text=label, anchor="w")
            # Description column stretches to fill remaining width
            self._tree.column(key, width=w, minwidth=w, anchor="w",
                              stretch=(key == "desc"))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=c["input"], foreground=c["text"],
                        fieldbackground=c["input"], bordercolor=c["border"],
                        rowheight=28, font=(MODERN_FONT, 11))
        style.configure("Treeview.Heading",
                        background=c["panel"], foreground=c["light"],
                        font=(MODERN_FONT, 11, "bold"), relief="flat")
        style.map("Treeview",
                  background=[("selected", c["accent"])],
                  foreground=[("selected", c["btn_text"])])
        style.map("Treeview.Heading", background=[("active", c["accent"])])
        # Tag for rows with blank descriptions (red highlight)
        self._tree.tag_configure("blank_desc", background="#5c1a1a", foreground="#ff6b6b")

        tree_scroll = ctk.CTkScrollbar(tree_frame, command=self._tree.yview)
        self._tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side="right", fill="y", padx=(2, 4), pady=4)
        self._tree.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)

        self._tree.bind("<Double-1>", self._on_tree_double_click)
        self._tree.bind("<Delete>", lambda e: self._remove_selected_row())
        # Tab navigation between cells
        self._tree.bind("<Tab>", self._on_tab_next)
        self._tree.bind("<Shift-Tab>", self._on_tab_prev)
        self._editing_row_id = None
        self._editing_key = None

        self._placeholder = None
        self._editing_window = None

        # ---- Status bar (with support icon in the corner) ----
        status_frame = ctk.CTkFrame(win, fg_color="transparent")
        status_frame.pack(fill="x", padx=22, pady=(0, 12))

        self.status_label = ctk.CTkLabel(
            status_frame, text="Standalone mode. Use 'Build from Manifest' to load from a manifest Excel, "
                      "or enter line items manually.",
            font=(MODERN_FONT, 11), text_color=c["text"], anchor="w")
        self.status_label.pack(side="left", fill="x", expand=True)

        self._attach_support_icon(status_frame)

    # -- Small helpers ----------------------------------------------------
    def _make_btn(self, parent, text, command, fg=None, hover=None, width=150):
        c = self.cfg["colors"]
        return ctk.CTkButton(parent, text=text, command=command, width=width, height=34,
                             fg_color=fg or c["accent"], hover_color=hover or c["accent_hover"],
                             text_color=c["btn_text"],
                             font=(MODERN_FONT, 12, "bold"), corner_radius=6)

    def _set_header(self, key, value):
        w = self.header_entries[key]
        w.delete(0, "end")
        w.insert(0, value)

    def _set_vessel(self, code):
        code = str(code).strip()
        if self.cfg.get("vessel_fixed"):
            if "vessel" in self.header_entries:
                self._set_header("vessel", self.cfg["vessel_fixed"])
        elif "vessel" in self.header_entries:
            self.header_entries["vessel"].set(self._code_to_display.get(code.upper(), code))

    def _set_status(self, text):
        self.status_label.configure(text=text)

    # -- Row management (Treeview-based) ---------------------------------
    def _show_placeholder(self):
        """Placeholder disabled — empty space is cleaner."""
        pass

    def _hide_placeholder(self):
        """Placeholder disabled."""
        pass

    def _on_tree_double_click(self, event):
        if self._editing_window is not None:
            try:
                self._editing_window.destroy()
            except Exception:
                pass
            self._editing_window = None
        self._autocomplete_list = None
        region = self._tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self._tree.identify_column(event.x)
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            return
        col_idx = int(col.replace("#", "")) - 1
        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        if col_idx < 0 or col_idx >= len(col_keys):
            return
        key = col_keys[col_idx]
        self._start_edit(row_id, key)

    def _on_tab_next(self, event):
        """Tab moves to the next cell to the right, or first cell of next row."""
        if self._editing_window is not None:
            return  # let the editing widget handle Tab
        sel = self._tree.selection()
        if not sel:
            return
        row_id = sel[0]
        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        # Find current column
        cur_col = self._tree.identify_column(event.x)
        cur_idx = int(cur_col.replace("#", "")) - 1 if cur_col.startswith("#") else 0
        # Try next column
        next_idx = cur_idx + 1
        if next_idx >= len(col_keys):
            # Wrap to next row, first column
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx + 1 < len(children):
                    row_id = children[cur_row_idx + 1]
                    next_idx = 0
                else:
                    return  # last row, last col - nowhere to go
            except ValueError:
                return
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)
        self._tree.see(row_id)
        self._start_edit(row_id, col_keys[next_idx])
        return "break"

    def _on_tab_prev(self, event):
        """Shift+Tab moves to the previous cell to the left, or last cell of prev row."""
        if self._editing_window is not None:
            return  # let the editing widget handle Tab
        sel = self._tree.selection()
        if not sel:
            return
        row_id = sel[0]
        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        cur_col = self._tree.identify_column(event.x)
        cur_idx = int(cur_col.replace("#", "")) - 1 if cur_col.startswith("#") else 0
        prev_idx = cur_idx - 1
        if prev_idx < 0:
            # Wrap to previous row, last column
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx > 0:
                    row_id = children[cur_row_idx - 1]
                    prev_idx = len(col_keys) - 1
                else:
                    return
            except ValueError:
                return
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)
        self._tree.see(row_id)
        self._start_edit(row_id, col_keys[prev_idx])
        return "break"

    def _start_edit(self, row_id, key):
        """Open an edit popup on the given cell. Shared by double-click and Tab nav."""
        if self._editing_window is not None:
            try:
                self._editing_window.destroy()
            except Exception:
                pass
            self._editing_window = None
        if self._autocomplete_list is not None:
            self._autocomplete_list.place_forget()
            self._autocomplete_list.destroy()
            self._autocomplete_list = None

        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        col_idx = col_keys.index(key)
        col = f"#{col_idx + 1}"
        bbox = self._tree.bbox(row_id, col)
        if not bbox:
            return
        x, y, w, h = bbox
        cx = self._tree.winfo_rootx() + x
        cy = self._tree.winfo_rooty() + y
        c = self.cfg["colors"]
        current_val = self._tree.set(row_id, key)
        self._editing_window = tk.Toplevel(self.win)
        self._editing_window.overrideredirect(True)
        self._editing_window.geometry(f"{w}x{h}+{cx}+{cy}")
        self._editing_window.attributes("-topmost", True)
        self._editing_row_id = row_id
        self._editing_key = key
        # Convert cell position to self.win-relative coordinates.
        # bbox gives coords relative to the tree widget. Use winfo_rootx/y
        # (reliable at click time since the window is already rendered) to
        # walk the full widget hierarchy offset back to self.win.
        tree_rel_x = self._tree.winfo_rootx() - self.win.winfo_rootx()
        tree_rel_y = self._tree.winfo_rooty() - self.win.winfo_rooty()
        self._edit_cell_x = x + tree_rel_x
        self._edit_cell_y = y + tree_rel_y
        self._edit_cell_h = h
        # Select the row so Tab navigation knows where we are
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)

        # Helper: save current and move to next cell (Tab)
        def tab_to_next():
            self._save_current_edit()
            self.win.after(10, self._tab_to_next_cell)

        # Helper: save current and move to prev cell (Shift+Tab)
        def tab_to_prev():
            self._save_current_edit()
            self.win.after(10, self._tab_to_prev_cell)

        if key == "proc":
            widget = ctk.CTkComboBox(self._editing_window, values=PROCEDURE_OPTIONS,
                                     width=w, height=h, fg_color=c["input"],
                                     border_color=c["border"], border_width=1,
                                     corner_radius=2, text_color=c["text"],
                                     button_color=c["accent"],
                                     button_hover_color=c["accent_hover"],
                                     dropdown_fg_color=c["panel"],
                                     dropdown_text_color=c["text"],
                                     dropdown_hover_color=c["accent"])
            widget.set(current_val)
            widget.pack(fill="both", expand=True)
            widget.bind("<<ComboboxSelected>>",
                        lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<FocusOut>",
                        lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<Tab>", lambda e: (tab_to_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_to_prev(), "break"))
            widget.focus_set()
        elif key == "desc":
            # Description column gets autocomplete from BUILTIN_CODES
            widget = ctk.CTkEntry(self._editing_window, width=w, height=h,
                                  fg_color=c["input"], border_color=c["accent"],
                                  border_width=2, corner_radius=2, text_color=c["text"],
                                  font=(MODERN_FONT, 11))
            widget.pack(fill="both", expand=True)
            widget.insert(0, current_val)
            widget.focus_set()
            widget.select_range(0, "end")
            self._editing_row_id = row_id
            self._editing_key = key
            self._desc_entry = widget
            # Build autocomplete list popup
            self._autocomplete_list = tk.Listbox(self.win, height=8,
                                                 bg=c["input"], fg=c["text"],
                                                 selectbackground=c["accent"],
                                                 selectforeground=c["btn_text"],
                                                 borderwidth=1, relief="solid",
                                                 font=(MODERN_FONT, 11),
                                                 activestyle="none")
            self._autocomplete_list.place_forget()
            self._autocomplete_list.bind("<<ListboxSelect>>",
                                         lambda e: self._on_autocomplete_select(row_id))
            widget.bind("<KeyRelease>", lambda e: self._update_autocomplete(widget, row_id))
            widget.bind("<Return>", lambda e: (self._finish_desc_edit(row_id, widget.get()), self._tab_to_next_cell()))
            widget.bind("<Escape>", lambda e: self._cancel_edit())
            widget.bind("<FocusOut>",
                        lambda e: self.win.after(150, lambda: self._finish_desc_edit(row_id, widget.get())))
            widget.bind("<Down>", lambda e: self._autocomplete_focus(widget))
            widget.bind("<Tab>", lambda e: (tab_to_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_to_prev(), "break"))
            # No initial dropdown — suggestions appear as the user types
        elif key == "cby":
            # CBY column gets autocomplete from BUILTIN_TIN_NUMBERS
            widget = ctk.CTkEntry(self._editing_window, width=w, height=h,
                                  fg_color=c["input"], border_color=c["accent"],
                                  border_width=2, corner_radius=2, text_color=c["text"],
                                  font=(MODERN_FONT, 11))
            widget.pack(fill="both", expand=True)
            widget.insert(0, current_val)
            widget.focus_set()
            widget.select_range(0, "end")
            self._editing_row_id = row_id
            self._editing_key = key
            # Build autocomplete list popup showing "CBY# - Name (TIN)"
            self._autocomplete_list = tk.Listbox(self.win, height=8,
                                                 bg=c["input"], fg=c["text"],
                                                 selectbackground=c["accent"],
                                                 selectforeground=c["btn_text"],
                                                 borderwidth=1, relief="solid",
                                                 font=(MODERN_FONT, 11),
                                                 activestyle="none")
            self._autocomplete_list.place_forget()
            self._autocomplete_list.bind("<<ListboxSelect>>",
                                         lambda e: self._on_cby_autocomplete_select(row_id))
            widget.bind("<KeyRelease>", lambda e: self._update_cby_autocomplete(widget))
            widget.bind("<Return>", lambda e: (self._finish_cby_edit(row_id, widget.get()), self._tab_to_next_cell()))
            widget.bind("<Escape>", lambda e: self._cancel_edit())
            widget.bind("<FocusOut>",
                        lambda e: self.win.after(150, lambda: self._finish_cby_edit(row_id, widget.get())))
            widget.bind("<Down>", lambda e: self._autocomplete_focus(widget))
            widget.bind("<Tab>", lambda e: (tab_to_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_to_prev(), "break"))
            # No initial dropdown — suggestions appear as the user types
        elif key == "importer":
            # Importer column - suggest TIN based on the CBY in this row
            widget = ctk.CTkEntry(self._editing_window, width=w, height=h,
                                  fg_color=c["input"], border_color=c["accent"],
                                  border_width=2, corner_radius=2, text_color=c["text"],
                                  font=(MODERN_FONT, 11))
            widget.pack(fill="both", expand=True)
            widget.insert(0, current_val)
            widget.focus_set()
            widget.select_range(0, "end")
            # Look up the CBY for this row and show a hint
            cby_val = self._tree.set(row_id, "cby").strip()
            hint = ""
            if cby_val:
                tin_entry = _get_tin_entry(cby_val)
                if tin_entry:
                    name, tin = tin_entry[0], tin_entry[1]
                    if tin:
                        hint = f"  ({name}: {tin})"
                    else:
                        hint = f"  ({name}: no TIN on file)"
                else:
                    hint = f"  (CBY {cby_val} not in database)"
            widget.bind("<Return>", lambda e: (self._finish_importer_edit(row_id, widget.get(), cby_val), self._tab_to_next_cell()))
            widget.bind("<Escape>", lambda e: self._cancel_edit())
            widget.bind("<FocusOut>",
                        lambda e: self.win.after(150, lambda: self._finish_importer_edit(row_id, widget.get(), cby_val)))
            widget.bind("<Tab>", lambda e: (tab_to_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_to_prev(), "break"))
        else:
            widget = ctk.CTkEntry(self._editing_window, width=w, height=h,
                                  fg_color=c["input"], border_color=c["accent"],
                                  border_width=2, corner_radius=2, text_color=c["text"],
                                  font=(MODERN_FONT, 11))
            widget.pack(fill="both", expand=True)
            widget.insert(0, current_val)
            widget.focus_set()
            widget.select_range(0, "end")
            widget.bind("<Return>", lambda e: (self._finish_edit(row_id, key, widget.get()), self._tab_to_next_cell()))
            widget.bind("<Escape>", lambda e: self._cancel_edit())
            widget.bind("<FocusOut>", lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<Tab>", lambda e: (tab_to_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_to_prev(), "break"))
        self._editing_row_id = row_id
        self._editing_key = key

    def _update_cby_autocomplete(self, entry):
        """Update the CBY autocomplete listbox based on what's typed."""
        if self._autocomplete_list is None:
            return
        typed = entry.get().strip().lower()
        matches = []
        if typed:
            for cby, entry in BUILTIN_TIN_NUMBERS.items():
                name = entry[0] if entry else ""
                tin = entry[1] if len(entry) > 1 else ""
                if typed in cby or typed in name.lower() or typed in f"{cby} {name.lower()}":
                    matches.append((cby, name, tin))
                    if len(matches) >= 20:
                        break
        # Sort by CBY number
        matches.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 0)

        self._autocomplete_list.delete(0, "end")
        for cby, name, tin in matches:
            display = f"#{cby} - {name}"
            if tin:
                display += f" ({tin})"
            self._autocomplete_list.insert("end", display)

        if matches:
            entry_x = self._edit_cell_x
            entry_y = self._edit_cell_y
            entry_h = self._edit_cell_h
            list_h = min(len(matches), 8) * 22
            win_h = self.win.winfo_height()
            if entry_y + entry_h + list_h > win_h:
                place_y = max(0, entry_y - list_h)
            else:
                place_y = entry_y + entry_h
            self._autocomplete_list.configure(height=min(len(matches), 8))
            self._autocomplete_list.place(x=entry_x, y=place_y,
                                          width=400, height=list_h)
            self._autocomplete_list.lift()
        else:
            self._autocomplete_list.place_forget()

    def _on_cby_autocomplete_select(self, row_id):
        """When a CBY suggestion is clicked, fill in CBY + importer number."""
        if self._autocomplete_list is None:
            return
        sel = self._autocomplete_list.curselection()
        if not sel:
            return
        selected = self._autocomplete_list.get(sel[0])
        # Parse "#cby - name (tin)" format
        # Extract CBY number
        cby = selected.split(" - ")[0].lstrip("#").strip()
        self._finish_cby_edit(row_id, cby)

    def _finish_cby_edit(self, row_id, value):
        """Finish editing CBY and auto-fill importer number from TIN database."""
        if self._autocomplete_list is not None:
            self._autocomplete_list.place_forget()
            self._autocomplete_list.destroy()
            self._autocomplete_list = None
        if self._editing_window is not None:
            try:
                self._tree.set(row_id, "cby", value)
            except Exception:
                pass
            # Auto-fill importer number from TIN database
            tin_entry = _get_tin_entry(value)
            if tin_entry:
                name, tin = tin_entry[0], tin_entry[1]
                if tin:
                    try:
                        self._tree.set(row_id, "importer", tin)
                    except Exception:
                        pass
            try:
                self._editing_window.destroy()
            except Exception:
                pass
            self._editing_window = None

    def _finish_importer_edit(self, row_id, value, cby_val):
        """Finish editing importer number. If empty and CBY has a TIN, auto-fill it."""
        if self._editing_window is not None:
            value = value.strip()
            if not value and cby_val:
                # Field was left empty - auto-fill from TIN database
                tin_entry = _get_tin_entry(cby_val)
                if tin_entry and tin_entry[1]:
                    value = tin_entry[1]
            try:
                self._tree.set(row_id, "importer", value)
            except Exception:
                pass
            try:
                self._editing_window.destroy()
            except Exception:
                pass
            self._editing_window = None

    def _update_autocomplete(self, entry, row_id):
        """Update the autocomplete listbox based on what's typed in the entry."""
        if self._autocomplete_list is None:
            return
        typed = entry.get().strip().lower()
        matches = []
        if typed:
            for desc, code, unit, proc in BUILTIN_CODES:
                if typed in desc.lower():
                    matches.append((desc, code, unit, proc))
                    if len(matches) >= 20:
                        break
        else:
            # Show first 20 if nothing typed
            for desc, code, unit, proc in BUILTIN_CODES[:20]:
                matches.append((desc, code, unit, proc))

        self._autocomplete_list.delete(0, "end")
        for desc, code, unit, proc in matches:
            self._autocomplete_list.insert("end", desc)

        if matches:
            # Position the listbox below the editing entry, or above if
            # there isn't enough room below (e.g. first row near top)
            entry_x = self._edit_cell_x
            entry_y = self._edit_cell_y
            entry_h = self._edit_cell_h
            list_h = min(len(matches), 8) * 22
            win_h = self.win.winfo_height()
            if entry_y + entry_h + list_h > win_h:
                # Not enough room below — place above the entry
                place_y = max(0, entry_y - list_h)
            else:
                place_y = entry_y + entry_h
            self._autocomplete_list.configure(height=min(len(matches), 8))
            self._autocomplete_list.place(x=entry_x, y=place_y,
                                          width=350, height=list_h)
            self._autocomplete_list.lift()
        else:
            self._autocomplete_list.place_forget()

    def _autocomplete_focus(self, entry):
        """Move focus to the autocomplete listbox."""
        if self._autocomplete_list is not None and self._autocomplete_list.size() > 0:
            self._autocomplete_list.focus_set()
            self._autocomplete_list.selection_set(0)
            self._autocomplete_list.activate(0)

    def _on_autocomplete_select(self, row_id):
        """When a suggestion is clicked, fill in desc + code + unit + proc."""
        if self._autocomplete_list is None:
            return
        sel = self._autocomplete_list.curselection()
        if not sel:
            return
        selected_desc = self._autocomplete_list.get(sel[0])
        self._finish_desc_edit(row_id, selected_desc)

    def _finish_desc_edit(self, row_id, value):
        """Finish editing the description and auto-fill code/unit/proc if matched."""
        if self._autocomplete_list is not None:
            self._autocomplete_list.place_forget()
            self._autocomplete_list.destroy()
            self._autocomplete_list = None
        if self._editing_window is not None:
            try:
                self._tree.set(row_id, "desc", value)
            except Exception:
                pass
            # Look up code/unit/proc from BUILTIN_CODES
            match = _BUILTIN_DESC_MAP.get(clean_text(value))
            if match:
                code, unit, proc = match
                try:
                    self._tree.set(row_id, "code", code)
                    self._tree.set(row_id, "unit", unit)
                    self._tree.set(row_id, "proc", proc)
                except Exception:
                    pass
            try:
                self._editing_window.destroy()
            except Exception:
                pass
            self._editing_window = None

    def _finish_edit(self, row_id, key, value):
        if self._editing_window is not None:
            try:
                self._tree.set(row_id, key, value)
            except Exception:
                pass
            try:
                self._editing_window.destroy()
            except Exception:
                pass
            self._editing_window = None
        # Update red tag when description is edited
        if key == "desc":
            if value.strip():
                self._tree.item(row_id, tags=())
            else:
                self._tree.item(row_id, tags=("blank_desc",))

    def _cancel_edit(self):
        if self._autocomplete_list is not None:
            self._autocomplete_list.place_forget()
            self._autocomplete_list.destroy()
            self._autocomplete_list = None
        if self._editing_window is not None:
            try:
                self._editing_window.destroy()
            except Exception:
                pass
            self._editing_window = None
        self._editing_row_id = None
        self._editing_key = None

    def _save_current_edit(self):
        """Save whatever is in the current editing widget without closing the nav flow."""
        if self._editing_window is None or self._editing_row_id is None:
            return
        row_id = self._editing_row_id
        key = self._editing_key
        # Find the widget inside the editing window
        for child in self._editing_window.winfo_children():
            try:
                value = child.get()
                break
            except Exception:
                continue
        else:
            return
        # Save using the appropriate finish method
        if key == "desc":
            self._finish_desc_edit(row_id, value)
        elif key == "cby":
            self._finish_cby_edit(row_id, value)
        elif key == "importer":
            cby_val = self._tree.set(row_id, "cby").strip()
            self._finish_importer_edit(row_id, value, cby_val)
        else:
            self._finish_edit(row_id, key, value)

    def _tab_to_next_cell(self):
        """After saving current cell, move to the next cell to the right."""
        if self._editing_row_id is None:
            return
        row_id = self._editing_row_id
        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        key = self._editing_key
        if key not in col_keys:
            return
        cur_idx = col_keys.index(key)
        next_idx = cur_idx + 1
        if next_idx >= len(col_keys):
            # Wrap to next row
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx + 1 < len(children):
                    row_id = children[cur_row_idx + 1]
                    next_idx = 0
                else:
                    return  # last cell
            except ValueError:
                return
        self._start_edit(row_id, col_keys[next_idx])

    def _tab_to_prev_cell(self):
        """After saving current cell, move to the previous cell to the left."""
        if self._editing_row_id is None:
            return
        row_id = self._editing_row_id
        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        key = self._editing_key
        if key not in col_keys:
            return
        cur_idx = col_keys.index(key)
        prev_idx = cur_idx - 1
        if prev_idx < 0:
            # Wrap to previous row, last column
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx > 0:
                    row_id = children[cur_row_idx - 1]
                    prev_idx = len(col_keys) - 1
                else:
                    return
            except ValueError:
                return
        self._start_edit(row_id, col_keys[prev_idx])

    def _remove_selected_row(self):
        selected = self._tree.selection()
        if not selected:
            return
        for item in selected:
            self._tree.delete(item)
        if not self._tree.get_children():
            self._show_placeholder()

    def _remove_row(self, item_id):
        try:
            self._tree.delete(item_id)
        except Exception:
            pass
        if not self._tree.get_children():
            self._show_placeholder()

    def _add_row(self, data=None):
        self._hide_placeholder()
        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        values = []
        for key in col_keys:
            val = ""
            if data and data.get(key) not in (None, ""):
                val = str(data[key])
            elif key == "qty":
                val = "1"
            elif key == "origin":
                val = "USA"
            elif key == "proc":
                val = "HOME"
            values.append(val)
        # Check if description is blank — highlight the row red
        tags = ()
        if data and not str(data.get("desc", "")).strip():
            tags = ("blank_desc",)
        item_id = self._tree.insert("", "end", values=values, tags=tags)
        return item_id

    def _clear_all_rows(self):
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._hide_placeholder()

    def _get_all_rows(self):
        col_keys = [k for k, _, _ in self.cfg["col_defs"]]
        rows = []
        for item in self._tree.get_children():
            vals = self._tree.set(item)
            row = {}
            for key in col_keys:
                row[key] = _TreeValue(vals.get(key, ""))
            rows.append(row)
        return rows

    # -- Build from Manifest (reads COLS sheet) ---------------------------
    def _build_from_manifest(self):
        """Open a file picker, then parse the COLS sheet in a background thread
        with a loading popup so the UI doesn't freeze."""
        initial_dir = str(self.cfg["excel_file"].parent) if self.cfg["excel_file"].parent.exists() else str(Path.home())
        file_path = filedialog.askopenfilename(
            title="Select " + self.cfg["mode"].title() + " Manifest Excel File",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not file_path:
            return

        # --- Loading popup (show BEFORE importing pandas to avoid freeze) ---
        loader = ctk.CTkToplevel(self.win)
        loader.title("Building from Manifest")
        loader.configure(fg_color=self.cfg["colors"]["bg"])
        loader.geometry("340x140")
        loader.transient(self.win)
        loader.grab_set()
        sw, sh = loader.winfo_screenwidth(), loader.winfo_screenheight()
        loader.geometry(f"+{int((sw-340)/2)}+{int((sh-140)/2)}")
        loader.resizable(False, False)
        loader.overrideredirect(False)

        c = self.cfg["colors"]
        ctk.CTkLabel(loader, text="Building from Manifest...",
                     font=(MODERN_FONT, 15, "bold"), text_color=c["light"]).pack(pady=(24, 8))
        ctk.CTkLabel(loader, text="Loading, please wait...",
                     font=(MODERN_FONT, 11), text_color=c["text"]).pack(pady=(0, 12))
        progress = ctk.CTkProgressBar(loader, width=280, height=14,
                                      progress_color=c["accent"],
                                      fg_color=c["input"])
        progress.pack(pady=(0, 20))
        progress.set(0.05)
        loader.update()

        # Now import pandas (this is the slow part — popup is already showing)
        try:
            import pandas as pd
        except ImportError:
            loader.destroy()
            messagebox.showerror("Missing Dependency",
                                 "pandas is required.\n\n"
                                 "Install with:\n    pip install pandas openpyxl")
            return

        # --- Parse in background thread ---
        result = {"data": None, "error": None}

        def parse_worker():
            try:
                cfg = self.cfg
                xlsx = Path(file_path)
                sheet = cfg["cols_sheet"]
                cm = cfg["cols_column_map"]
                hl = cfg["cols_header_labels"]
                hdr_row = cfg["cols_header_row"]
                data_start = cfg["cols_data_start_row"]

                # Read COLS sheet - try the configured name, then variations
                # (some manifests have " COLS" with a leading space, others "COLS")
                xl = pd.ExcelFile(xlsx)
                sheet_names = [s.strip() for s in xl.sheet_names]
                sheet_found = None
                for sn in xl.sheet_names:
                    if sn.strip().lower() == sheet.strip().lower():
                        sheet_found = sn
                        break
                if sheet_found is None:
                    # Try case-insensitive match on "cols"
                    for sn in xl.sheet_names:
                        if sn.strip().lower() == "cols":
                            sheet_found = sn
                            break
                if sheet_found is None:
                    raise ValueError(f"Could not find a COLS sheet. Available sheets: {xl.sheet_names}")
                raw_df = pd.read_excel(xlsx, sheet_name=sheet_found, header=None)

                # Parse header overview cells
                hv = {"manifest_date": "", "arrival_date": "", "master_bl": "",
                      "pkg_count": "1", "weight_cf": "0.00", "weight_lb": "0.00",
                      "vessel_code": "", "voyage_no": ""}
                for i in range(min(hdr_row, len(raw_df))):
                    row_vals = [str(v).strip() if not pd.isna(v) else "" for v in raw_df.iloc[i].tolist()]
                    for key, label in hl.items():
                        if label and label in row_vals:
                            col_idx = row_vals.index(label)
                            val_idx = col_idx + 1
                            if val_idx < len(raw_df.columns):
                                val = raw_df.iloc[i].tolist()[val_idx]
                                if pd.isna(val) or str(val).strip().lower() in ["", "nan"]:
                                    continue
                                if key in ("manifest_date", "arrival_date"):
                                    try:
                                        hv[key] = pd.to_datetime(val).strftime("%Y-%m-%d")
                                    except Exception:
                                        hv[key] = str(val).strip()
                                elif key == "pkg_count":
                                    try:
                                        hv[key] = str(int(float(val)))
                                    except Exception:
                                        hv[key] = str(val).strip()
                                elif key in ("weight_cf", "weight_lb"):
                                    hv[key] = money_str(val)
                                else:
                                    hv[key] = str(val).strip()

                # Use built-in codes database (not from manifest)
                desc_to_code = dict(_BUILTIN_DESC_MAP)
                code_to_unit = {code: unit for code, (_, unit, _) in _BUILTIN_DESC_MAP.values() for code in [code]} if False else {}
                code_to_proc = {}
                for _clean_desc, (_code, _unit, _proc) in _BUILTIN_DESC_MAP.items():
                    code_to_unit[_code] = _unit
                    code_to_proc[_code] = _proc
                # Also map by original desc for exact matching
                for _desc, _code, _unit, _proc in BUILTIN_CODES:
                    desc_to_code[clean_text(_desc)] = _code

                def find_best_match(description):
                    cleaned = clean_text(description)
                    if cleaned in desc_to_code:
                        return desc_to_code[cleaned]
                    for k, v in desc_to_code.items():
                        if cleaned in k or k in cleaned:
                            return v
                    words = cleaned.split()
                    if len(words) >= 2:
                        first_two = " ".join(words[:2])
                        for k, v in desc_to_code.items():
                            if first_two in k:
                                return v
                    matches = get_close_matches(cleaned, desc_to_code.keys(), n=1, cutoff=0.60)
                    if matches:
                        return desc_to_code[matches[0]]
                    return DEFAULT_COMMODITY_CODE

                # Build item rows data
                items = []
                for i in range(data_start, len(raw_df)):
                    row_data = raw_df.iloc[i].tolist()

                    cby_val = ""
                    if cm["cby"] is not None and cm["cby"] < len(row_data):
                        cby_val = clean_dock_receipt(row_data[cm["cby"]])
                    if not cby_val or str(cby_val).lower() in ("", "nan", "0", "0.0", "none", "store"):
                        continue

                    desc_str = ""
                    if cm["desc"] is not None and cm["desc"] < len(row_data):
                        desc_str = str(row_data[cm["desc"]]).strip()
                    # Treat 0/0.0/NaN as blank descriptions (empty Excel cells)
                    if desc_str.lower() in ("nan", "0", "0.0", "shipment total"):
                        desc_str = ""
                    # Don't skip rows with blank descriptions — include them
                    # so the user sees the gap and can fix it.

                    dock_val = ""
                    if cm["dock"] is not None and cm["dock"] < len(row_data):
                        dock_val = clean_dock_receipt(row_data[cm["dock"]])

                    importer_number = ""
                    if cm["importer"] is not None and cm["importer"] < len(row_data):
                        raw_imp = row_data[cm["importer"]]
                        try:
                            importer_number = str(int(float(raw_imp)))
                        except Exception:
                            importer_number = str(raw_imp).strip()
                        if importer_number.lower() in ("", "nan", "0", "none"):
                            importer_number = ""
                    # If no importer from manifest, look up by CBY in built-in TIN list
                    if not importer_number and cby_val:
                        tin_entry = _get_tin_entry(str(cby_val))
                        if tin_entry:
                            importer_number = tin_entry[1]

                    qty_val = "1"
                    if cm["qty"] is not None and cm["qty"] < len(row_data):
                        try:
                            qty_val = str(int(float(row_data[cm["qty"]])))
                        except Exception:
                            qty_val = "1"

                    value_val = money_str(row_data[cm["value"]]) if cm["value"] is not None and cm["value"] < len(row_data) else "0.00"
                    freight_val = money_str(row_data[cm["freight"]]) if cm["freight"] is not None and cm["freight"] < len(row_data) else "0.00"
                    insurance_val = money_str(row_data[cm["insurance"]]) if cm["insurance"] is not None and cm["insurance"] < len(row_data) else "0.00"

                    if desc_str and desc_to_code:
                        code = find_best_match(desc_str)
                    else:
                        code = ""  # No code for blank descriptions

                    items.append({
                        "cby": cby_val,
                        "dock": dock_val,
                        "desc": desc_str,
                        "code": code,
                        "unit": code_to_unit.get(code, "NO"),
                        "proc": _apply_special_proc(cby_val, code_to_proc.get(code, "HOME")),
                        "qty": qty_val,
                        "value": value_val,
                        "freight": freight_val,
                        "insurance": insurance_val,
                        "importer": importer_number,
                        "origin": "USA",
                    })

                result["data"] = {"hv": hv, "items": items,
                                   "desc_to_code": desc_to_code,
                                   "code_to_unit": code_to_unit,
                                   "code_to_proc": code_to_proc,
                                   "filename": xlsx.name,
                                   "manifest_dir": xlsx.parent}

            except Exception as e:
                result["error"] = str(e)

        def on_progress(value):
            try:
                progress.set(value)
                loader.update_idletasks()
            except Exception:
                pass

        def run_thread():
            # Animate the progress bar gently while parsing
            import time
            pv = [0.1]
            stop = [False]

            def animate():
                while not stop[0]:
                    pv[0] = min(0.9, pv[0] + 0.03)
                    on_progress(pv[0])
                    time.sleep(0.08)

            anim_thread = threading.Thread(target=animate, daemon=True)
            anim_thread.start()

            parse_worker()
            stop[0] = True
            on_progress(1.0)

            # Schedule UI update on main thread
            self.win.after(100, lambda: finish_ui())

        def finish_ui():
            try:
                loader.destroy()
            except Exception:
                pass

            if result["error"]:
                messagebox.showerror("Load Error", f"Failed to read the manifest:\n\n{result['error']}")
                return

            data = result["data"]
            hv = data["hv"]
            items = data["items"]

            # Store commodity lookup tables
            self._desc_to_code = data["desc_to_code"]
            self._code_to_unit = data["code_to_unit"]
            self._code_to_proc = data["code_to_proc"]
            self._manifest_dir = data.get("manifest_dir")

            # Populate header widgets
            cfg = self.cfg
            self._set_header("awb", hv["master_bl"])
            if hv.get("vessel_code") and "vessel" in self.header_entries:
                self._set_vessel(hv["vessel_code"])
            elif cfg.get("vessel_fixed") and "vessel" in self.header_entries:
                self._set_header("vessel", cfg["vessel_fixed"])
            if hv.get("voyage_no"):
                self._set_header("voyage", hv["voyage_no"])
            if hv.get("manifest_date"):
                self._set_header("departure", hv["manifest_date"])
            if hv.get("arrival_date"):
                self._set_header("arrival", hv["arrival_date"])
            if hv.get("weight_lb"):
                self._set_header("weight_lb", hv["weight_lb"])
            if hv.get("weight_cf"):
                self._set_header("weight_cf", hv["weight_cf"])
            if hv.get("pkg_count"):
                self._set_header("pkg_count", hv["pkg_count"])

            # Populate treeview - instant with native widget
            self._clear_all_rows()
            for item in items:
                self._add_row(item)

            self._set_status(f"Loaded {len(items)} line item(s) from {data['filename']} (COLS sheet).")
            messagebox.showinfo("Loaded", f"Loaded {len(items)} line item(s) from the COLS sheet.\n\n"
                                           "Some header fields may need manual entry.")

        # Start the background thread
        work_thread = threading.Thread(target=run_thread, daemon=True)
        work_thread.start()



    # -- Collect header ---------------------------------------------------
    def _collect_header(self):
        vessel_raw = self.header_entries["vessel"].get().strip() if "vessel" in self.header_entries else ""
        m = re.search(r"\((.*?)\)", vessel_raw)
        vessel_code = m.group(1).strip() if m else vessel_raw
        return {
            "awb": self.header_entries["awb"].get().strip(),
            "vessel": vessel_code,
            "voyage": self.header_entries["voyage"].get().strip(),
            "departure": self.header_entries["departure"].get().strip(),
            "arrival": self.header_entries["arrival"].get().strip(),
            "weight_lb": self.header_entries["weight_lb"].get().strip() or "0.00",
            "weight_cf": self.header_entries["weight_cf"].get().strip() or "0.00",
            "pkg_count": self.header_entries["pkg_count"].get().strip() or "1",
            "master_importer": self.header_entries["master_importer"].get().strip() or MASTER_IMPORTER_DEFAULT,
        }

    # -- Generate XML -----------------------------------------------------
    def _generate_xml(self):
        cfg = self.cfg
        hdr = self._collect_header()
        if not hdr["awb"]:
            messagebox.showerror("Missing AWB", "Please enter the AWB / Master BL # before generating.")
            return

        # Read all rows from the treeview
        all_rows = self._get_all_rows()
        active_rows = []
        for row in all_rows:
            cby = row["cby"].get().strip()
            desc = row["desc"].get().strip()
            if not cby and not desc:
                continue
            if not cby:
                continue
            active_rows.append(row)

        if not active_rows:
            messagebox.showerror("No Line Items", "There are no line items with a CBY to generate.")
            return

        # Check for blank descriptions
        blank_desc_cby = []
        for row in active_rows:
            desc = row["desc"].get().strip()
            cby = row["cby"].get().strip()
            if not desc or desc.lower() in ("0", "0.0", "nan"):
                blank_desc_cby.append(cby)

        if blank_desc_cby:
            # Count occurrences per CBY so we show each once with a count
            counts = Counter(blank_desc_cby)
            cby_list = ", ".join(
                f"{cby} ({n})" if n > 1 else cby
                for cby, n in counts.items())
            # Custom dialog with proper button labels
            dialog = ctk.CTkToplevel(self.win)
            dialog.title("Description Box Blank")
            dialog.configure(fg_color="#1a1a2e")
            dialog.geometry("460x240")
            dialog.resizable(False, False)
            dialog.transient(self.win)
            dialog.grab_set()
            dialog.update_idletasks()
            px = self.win.winfo_rootx() + (self.win.winfo_width() - 460) // 2
            py = self.win.winfo_rooty() + (self.win.winfo_height() - 240) // 2
            dialog.geometry(f"+{px}+{py}")
            ctk.CTkLabel(dialog, text="Description Box Blank",
                         font=(MODERN_FONT, 14, "bold"), text_color="#e8e8e8").pack(pady=(20, 4))
            ctk.CTkLabel(dialog,
                         text=f"The following CBY(s) have blank descriptions:\n{cby_list}",
                         font=(MODERN_FONT, 12), text_color="#ff6b6b",
                         justify="center").pack(pady=(0, 16))
            choice = [None]
            def on_cancel():
                choice[0] = False
                dialog.destroy()
            def on_proceed():
                choice[0] = True
                dialog.destroy()
            btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
            btn_frame.pack(pady=(0, 20))
            ctk.CTkButton(btn_frame, text="Cancel Export", command=on_cancel,
                         fg_color="#c0392b", hover_color="#e74c3c",
                         width=140, height=32, corner_radius=6,
                         text_color="#ffffff", font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=(30, 10))
            ctk.CTkButton(btn_frame, text="Proceed - with Issues", command=on_proceed,
                         fg_color="#e8820c", hover_color="#ffa726",
                         width=160, height=32, corner_radius=6,
                         text_color="#ffffff", font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=(10, 30))
            dialog.bind("<Escape>", lambda e: on_cancel())
            dialog.bind("<Return>", lambda e: on_proceed())
            dialog.wait_window()
            if not choice[0]:
                return  # User chose to cancel and fix

        try:
            # If we loaded from a manifest, save XML files next to it.
            # Otherwise fall back to the default output folder.
            if self._manifest_dir:
                # Use the departure/manifest date for the subfolder name
                date_str = hdr.get("departure") or hdr.get("arrival") or "Unknown Date"
                # Clean the date for use as a folder name (replace any invalid chars)
                safe_date = re.sub(r'[<>:"/\|?*]', '-', date_str)
                out = self._manifest_dir / f"XML files for {safe_date}"
            else:
                out = cfg["output_folder"]
            out.mkdir(parents=True, exist_ok=True)
            created = []

            # ---- MASTER ----
            master_wt = hdr["weight_lb"]
            master_vol = hdr["weight_cf"]
            master_root = build_sad_structure(cfg, hdr, "MASTER", hdr["awb"],
                                              "Consolidated Shipment", hdr["pkg_count"],
                                              pkg_type="BG",
                                              gross_wt=master_wt, gross_vol=master_vol)
            mi = ET.Element("Importer")
            ET.SubElement(mi, "Number").text = hdr["master_importer"]
            master_root.insert(2, mi)
            ET.SubElement(master_root, "MoneyDeclaredFlag").text = "N"
            master_fn = f"HBL-Master {hdr['awb']}.xml"
            ET.ElementTree(master_root).write(out / master_fn, encoding="utf-8", xml_declaration=True)
            created.append(master_fn)

            # ---- GROUP BY CBY ----
            groups = {}
            order = []
            for row in active_rows:
                cby = row["cby"].get().strip()
                if cby not in groups:
                    groups[cby] = []
                    order.append(cby)
                groups[cby].append(row)

            skipped = 0
            for cby in order:
                rows = groups[cby]
                try:
                    first_row = rows[0]
                    first_dock = clean_dock_receipt(first_row["dock"].get().strip())
                    importer_number = first_row["importer"].get().strip() or FALLBACK_IMPORTER_NUMBER

                    total_value = Decimal("0.00")
                    total_freight = Decimal("0.00")
                    total_insurance = Decimal("0.00")
                    unique_descs = set()
                    for row in rows:
                        total_value += money(row["value"].get())
                        total_freight += money(row["freight"].get())
                        total_insurance += money(row["insurance"].get())
                        d = row["desc"].get().strip()
                        if d:
                            unique_descs.add(d)

                    if len(unique_descs) == 1:
                        contents = truncate_desc(list(unique_descs)[0]).upper()
                    elif len(unique_descs) > 1:
                        contents = "MIXED GOODS"
                    else:
                        contents = "GOODS"

                    # Air house XMLs use 0.00 for weight/volume; ocean uses header values
                    if cfg["house_weight_zero"]:
                        h_wt, h_vol = "0.00", "0.00"
                    else:
                        h_wt, h_vol = hdr["weight_lb"], hdr["weight_cf"]

                    house = build_sad_structure(cfg, hdr, "HOUSE", first_dock, contents,
                                                len(rows), pkg_type="PK",
                                                gross_wt=h_wt, gross_vol=h_vol)

                    imp_node = ET.Element("Importer")
                    ET.SubElement(imp_node, "Number").text = importer_number
                    house.insert(2, imp_node)

                    valuation = ET.SubElement(house, "Valuation")
                    ET.SubElement(valuation, "Currency").text = "USD"
                    ET.SubElement(valuation, "NetCost").text = money_str(total_value)
                    ET.SubElement(valuation, "NetInsurance").text = money_str(total_insurance)
                    ET.SubElement(valuation, "NetFreight").text = money_str(total_freight)
                    ET.SubElement(valuation, "TermsOfDelivery").text = "FOB"

                    row_idx = 1
                    for row in rows:
                        comm_code = row["code"].get().strip() or DEFAULT_COMMODITY_CODE
                        item_node = ET.SubElement(house, "Items", row=str(row_idx))
                        ET.SubElement(item_node, "Code").text = comm_code
                        ET.SubElement(item_node, "Desc").text = truncate_desc(row["desc"].get())
                        ET.SubElement(item_node, "Origin").text = row["origin"].get().strip() or "USA"
                        ET.SubElement(item_node, "Qty").text = row["qty"].get().strip() or "1"
                        ET.SubElement(item_node, "QtyUnit").text = row["unit"].get().strip() or "NO"
                        ET.SubElement(item_node, "Cost").text = money_str(row["value"].get())
                        ET.SubElement(item_node, "Insurance").text = money_str(row["insurance"].get())
                        ET.SubElement(item_node, "Freight").text = money_str(row["freight"].get())
                        ET.SubElement(item_node, "InvNumber").text = clean_dock_receipt(row["dock"].get().strip())

                        proc_node = ET.SubElement(item_node, "Procedure")
                        _proc_code = row["proc"].get().strip() or "HOME"
                        _proc_code = _apply_special_proc(cby, _proc_code)
                        ET.SubElement(proc_node, "Code").text = _proc_code
                        ET.SubElement(proc_node, "ImporterNumber").text = importer_number
                        row_idx += 1

                    ET.SubElement(house, "MoneyDeclaredFlag").text = "N"
                    h_fn = f"HBL-CBY {cby} {first_dock}.xml"
                    ET.ElementTree(house).write(out / h_fn, encoding="utf-8", xml_declaration=True)
                    created.append(h_fn)
                except Exception as e:
                    skipped += 1
                    print(f"Skipping CBY {cby}: {e}")

            self._set_status(f"Generated {len(created)} XML file(s) into '{out.name}'.")
            summary = (f"Master entries:  1\n"
                       f"House entries:   {len(created) - 1}\n"
                       f"Groups skipped:  {skipped}\n\n"
                       f"XML files saved in:\n{out}")
            messagebox.showinfo("XML Generation Complete", summary)

        except Exception as e:
            messagebox.showerror("Generation Error", f"An error occurred while generating XML:\n\n{e}")

    # -- Navigation -------------------------------------------------------
    def _close_all(self):
        """X button: close everything (console + launcher)."""
        self.win.destroy()
        self.launcher.root.destroy()

    def _go_back(self):
        """Back button: return to the launcher."""
        self.win.destroy()
        self.launcher.show()


# ==============================================================================
# CODES MANAGEMENT WINDOW
# ==============================================================================
class CodesWindow(SupportMixin):
    """Searchable, editable table of commodity codes, units, and procedures.
    Changes are saved back to the script file so they persist across machines
    (via Dropbox sync)."""

    def __init__(self, launcher):
        self.launcher = launcher

        # Support-mixin state
        self._pending_update = None
        self._support_tooltip = "Report a Bug"
        self._tooltip_win = None
        self._support_bg = "#0f1117"
        self._window_name = "Item Codes"
        self._set_support_palette({
            "bg": "#0f1117", "panel": "#141a2a", "input": "#0a0e16",
            "border": "#1a2a4a", "text": "#c8d6e5", "light": "#5a7a9a",
            "accent": "#0f3460", "accent_hover": "#1a4a7a",
        })

        self.win = ctk.CTkToplevel()
        self.win.title(f"Commodity Codes Management v{APP_VERSION}")
        self.win.configure(fg_color="#0f1117")
        _register_window_name(self.win, self._window_name, {"bg": "#0f1117", "panel": "#141a2a", "accent": "#0f3460", "accent_hover": "#1a4a7a", "text": "#c8d6e5"})
        self.win.geometry("900x600")
        sw, sh = self.win.winfo_screenwidth(), self.win.winfo_screenheight()
        self.win.geometry(f"+{int((sw-900)/2)}+{int((sh-600)/2)}")
        self.win.protocol("WM_DELETE_WINDOW", self._on_close)
        self.win.transient(launcher.root)

        # Title bar with Back button
        title_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        title_frame.pack(fill="x", padx=16, pady=(12, 0))
        ctk.CTkButton(title_frame, text="\u2190 Back",
                      command=self._on_close,
                      fg_color="#333", hover_color="#444",
                      width=80, height=28, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold"),
                      text_color="#e8e8e8").pack(side="left")
        ctk.CTkLabel(title_frame, text="Commodity Codes Database",
                     font=(MODERN_FONT, 18, "bold"), text_color="#e8e8e8").pack(side="left", padx=(16, 0))

        # Subtitle
        ctk.CTkLabel(self.win, text=f"{len(BUILTIN_CODES)} codes loaded  -  double-click a cell to edit  -  changes save to the script file",
                     font=(MODERN_FONT, 11), text_color="#888").pack(pady=(8, 12))

        # Search bar
        search_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        search_frame.pack(fill="x", padx=20, pady=(0, 8))
        ctk.CTkLabel(search_frame, text="Search:", font=(MODERN_FONT, 12),
                     text_color="#ccc").pack(side="left", padx=(0, 8))
        self._search_var = ctk.StringVar()
        self._search_var.trace("w", lambda *a: self._filter_tree())
        search_entry = ctk.CTkEntry(search_frame, textvariable=self._search_var,
                                    width=300, height=30, fg_color="#1a1a2e",
                                    border_color="#333", border_width=1, corner_radius=5,
                                    text_color="#e8e8e8", font=(MODERN_FONT, 12))
        search_entry.pack(side="left", padx=(0, 12))

        # Add button
        ctk.CTkButton(search_frame, text="+ Add Code", command=self._add_code,
                      fg_color="#2e8b57", hover_color="#3cb371",
                      width=110, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        # Delete button
        ctk.CTkButton(search_frame, text="Delete Selected", command=self._delete_selected,
                      fg_color="#c0392b", hover_color="#e74c3c",
                      width=130, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        # Save button
        ctk.CTkButton(search_frame, text="Save Changes", command=self._save_to_script,
                      fg_color="#e8820c", hover_color="#ffa726",
                      width=130, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="right", padx=4)

        # Treeview
        tree_frame = ctk.CTkFrame(self.win, fg_color="#1a1a2e", corner_radius=6)
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        cols = ["desc", "code", "unit", "proc"]
        labels = ["Description", "Code", "Unit", "Procedure"]
        widths = [400, 120, 80, 120]
        self._tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        for col, label, w in zip(cols, labels, widths):
            self._tree.heading(col, text=label, anchor="w")
            self._tree.column(col, width=w, minwidth=w, anchor="w",
                              stretch=(col == "desc"))

        style = ttk.Style()
        style.configure("Treeview", background="#1a1a2e", foreground="#e8e8e8",
                        fieldbackground="#1a1a2e", bordercolor="#333",
                        rowheight=26, font=(MODERN_FONT, 11))
        style.configure("Treeview.Heading", background="#0f1117", foreground="#ffa726",
                        font=(MODERN_FONT, 11, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", "#2e8b57")],
                  foreground=[("selected", "#fff")])

        tree_scroll = ctk.CTkScrollbar(tree_frame, command=self._tree.yview)
        self._tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side="right", fill="y", padx=(2, 4), pady=4)
        self._tree.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)

        self._tree.bind("<Double-1>", self._on_double_click)
        self._tree.bind("<Tab>", self._codes_tab_next)
        self._tree.bind("<Shift-Tab>", self._codes_tab_prev)
        self._editing_win = None
        self._editing_row_id = None
        self._editing_key = None
        self._deleted_desc_set = set()
        self._dirty = False

        # Populate
        self._populate_tree()

        # Status (with support icon in the corner)
        status_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        status_frame.pack(fill="x", padx=22, pady=(0, 8))
        self._status = ctk.CTkLabel(status_frame, text="Ready", font=(MODERN_FONT, 11),
                                    text_color="#888", anchor="w")
        self._status.pack(side="left", fill="x", expand=True)
        self._attach_support_icon(status_frame)

        # Check for updates in the background
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    def _populate_tree(self, filter_text=""):
        for item in self._tree.get_children():
            self._tree.delete(item)
        for desc, code, unit, proc in BUILTIN_CODES:
            if filter_text:
                combined = f"{desc} {code} {unit} {proc}".lower()
                if filter_text.lower() not in combined:
                    continue
            self._tree.insert("", "end", values=(desc, code, unit, proc))
        # Clear deletion tracking when full list is reloaded
        if not filter_text:
            self._deleted_desc_set = set()

    def _filter_tree(self):
        self._populate_tree(self._search_var.get().strip())

    def _on_double_click(self, event):
        if self._editing_win is not None:
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
        region = self._tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            return
        self._add_code(existing_item=row_id)

    def _codes_tab_next(self, event):
        if self._editing_win is not None:
            return
        sel = self._tree.selection()
        if not sel:
            return
        row_id = sel[0]
        cols = ["desc", "code", "unit", "proc"]
        cur_col = self._tree.identify_column(event.x)
        cur_idx = int(cur_col.replace("#", "")) - 1 if cur_col.startswith("#") else 0
        next_idx = cur_idx + 1
        if next_idx >= len(cols):
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx + 1 < len(children):
                    row_id = children[cur_row_idx + 1]
                    next_idx = 0
                else:
                    return
            except ValueError:
                return
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)
        self._tree.see(row_id)
        self._codes_start_edit(row_id, cols[next_idx])
        return "break"

    def _codes_tab_prev(self, event):
        if self._editing_win is not None:
            return
        sel = self._tree.selection()
        if not sel:
            return
        row_id = sel[0]
        cols = ["desc", "code", "unit", "proc"]
        cur_col = self._tree.identify_column(event.x)
        cur_idx = int(cur_col.replace("#", "")) - 1 if cur_col.startswith("#") else 0
        prev_idx = cur_idx - 1
        if prev_idx < 0:
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx > 0:
                    row_id = children[cur_row_idx - 1]
                    prev_idx = len(cols) - 1
                else:
                    return
            except ValueError:
                return
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)
        self._tree.see(row_id)
        self._codes_start_edit(row_id, cols[prev_idx])
        return "break"

    def _codes_start_edit(self, row_id, key):
        """Open edit popup on a cell. Shared by double-click and Tab nav."""
        if self._editing_win is not None:
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
        cols = ["desc", "code", "unit", "proc"]
        col_idx = cols.index(key)
        col = f"#{col_idx + 1}"
        bbox = self._tree.bbox(row_id, col)
        if not bbox:
            return
        x, y, w, h = bbox
        cx = self._tree.winfo_rootx() + x
        cy = self._tree.winfo_rooty() + y
        current_val = self._tree.set(row_id, key)
        self._editing_win = tk.Toplevel(self.win)
        self._editing_win.overrideredirect(True)
        self._editing_win.geometry(f"{w}x{h}+{cx}+{cy}")
        self._editing_win.attributes("-topmost", True)
        self._editing_row_id = row_id
        self._editing_key = key
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)

        def tab_next():
            self._codes_save_current_and_tab(1)
        def tab_prev():
            self._codes_save_current_and_tab(-1)

        if key == "unit":
            widget = ctk.CTkComboBox(self._editing_win, values=UNIT_OPTIONS,
                                     width=w, height=h, fg_color="#1a1a2e",
                                     border_color="#333", border_width=1, corner_radius=2,
                                     text_color="#e8e8e8", button_color="#e8820c",
                                     button_hover_color="#ffa726",
                                     dropdown_fg_color="#1a1a2e",
                                     dropdown_text_color="#e8e8e8",
                                     dropdown_hover_color="#2e8b57")
            widget.set(current_val)
            widget.pack(fill="both", expand=True)
            widget.bind("<<ComboboxSelected>>", lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<FocusOut>", lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<Tab>", lambda e: (tab_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_prev(), "break"))
            widget.focus_set()
        elif key == "proc":
            widget = ctk.CTkComboBox(self._editing_win, values=PROCEDURE_OPTIONS,
                                     width=w, height=h, fg_color="#1a1a2e",
                                     border_color="#333", border_width=1, corner_radius=2,
                                     text_color="#e8e8e8", button_color="#e8820c",
                                     button_hover_color="#ffa726",
                                     dropdown_fg_color="#1a1a2e",
                                     dropdown_text_color="#e8e8e8",
                                     dropdown_hover_color="#2e8b57")
            widget.set(current_val)
            widget.pack(fill="both", expand=True)
            widget.bind("<<ComboboxSelected>>", lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<FocusOut>", lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<Tab>", lambda e: (tab_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_prev(), "break"))
            widget.focus_set()
        else:
            widget = ctk.CTkEntry(self._editing_win, width=w, height=h,
                                  fg_color="#1a1a2e", border_color="#e8820c",
                                  border_width=2, corner_radius=2, text_color="#e8e8e8",
                                  font=(MODERN_FONT, 11))
            widget.pack(fill="both", expand=True)
            widget.insert(0, current_val)
            widget.focus_set()
            widget.select_range(0, "end")
            widget.bind("<Return>", lambda e: (self._finish_edit(row_id, key, widget.get()), self._codes_tab_to_cell(1)))
            widget.bind("<Escape>", lambda e: self._cancel_edit())
            widget.bind("<FocusOut>", lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<Tab>", lambda e: (tab_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_prev(), "break"))

    def _codes_save_current_and_tab(self, direction):
        """Save current edit and move to next/prev cell."""
        if self._editing_row_id is None:
            return
        row_id = self._editing_row_id
        key = self._editing_key
        # Get value from the editing widget
        for child in self._editing_win.winfo_children():
            try:
                value = child.get()
                break
            except:
                continue
        else:
            return
        self._finish_edit(row_id, key, value)
        self.win.after(10, lambda: self._codes_tab_to_cell(direction))

    def _codes_tab_to_cell(self, direction):
        """Move to next (1) or prev (-1) cell after saving."""
        if self._editing_row_id is None:
            return
        row_id = self._editing_row_id
        cols = ["desc", "code", "unit", "proc"]
        key = self._editing_key
        if key not in cols:
            return
        cur_idx = cols.index(key)
        next_idx = cur_idx + direction
        if next_idx >= len(cols):
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx + 1 < len(children):
                    row_id = children[cur_row_idx + 1]
                    next_idx = 0
                else:
                    return
            except ValueError:
                return
        elif next_idx < 0:
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx > 0:
                    row_id = children[cur_row_idx - 1]
                    next_idx = len(cols) - 1
                else:
                    return
            except ValueError:
                return
        self._codes_start_edit(row_id, cols[next_idx])

    def _finish_edit(self, row_id, key, value):
        if self._editing_win is not None:
            try:
                self._tree.set(row_id, key, value)
            except Exception:
                pass
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
            self._dirty = True
            self._status.configure(text="Edited - click 'Save Changes' to persist")

    def _cancel_edit(self):
        if self._editing_win is not None:
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
        self._editing_row_id = None
        self._editing_key = None

    def _add_code(self, existing_item=None):
        """Open a popup dialog to enter (or edit) a commodity code.
        If existing_item is provided, the dialog is pre-filled with that
        row's values and saves update the row instead of inserting."""
        is_edit = existing_item is not None
        dlg = ctk.CTkToplevel(self.win)
        dlg.title("Edit Commodity Code" if is_edit else "Add Commodity Code")
        dlg.configure(fg_color="#0f1117")
        dlg.geometry("460x520")
        dlg.transient(self.win)
        dlg.attributes("-topmost", True)
        # Center over the parent window
        self.win.update_idletasks()
        px = self.win.winfo_rootx() + (self.win.winfo_width() - 460) // 2
        py = self.win.winfo_rooty() + (self.win.winfo_height() - 520) // 2
        dlg.geometry(f"+{px}+{py}")
        dlg.grab_set()

        ctk.CTkLabel(dlg, text="Edit Commodity Code" if is_edit else "Add Commodity Code",
                     font=(MODERN_FONT, 16, "bold"),
                     text_color="#e8e8e8").pack(pady=(16, 4))
        ctk.CTkLabel(dlg, text="Fill in the details and click Save." if is_edit
                     else "Fill in the details and click Add.",
                     font=(MODERN_FONT, 11), text_color="#888").pack(pady=(0, 12))

        form = ctk.CTkFrame(dlg, fg_color="transparent")
        form.pack(fill="x", padx=24, pady=(0, 8))

        def add_row(label_text, widget):
            ctk.CTkLabel(form, text=label_text, anchor="w",
                         font=(MODERN_FONT, 12), text_color="#ccc").pack(fill="x", pady=(8, 2))
            widget.pack(fill="x")

        # Pre-fill values if editing
        cur_desc, cur_code, cur_unit, cur_proc = "", "", "NO", "HOME"
        if is_edit:
            vals = self._tree.set(existing_item)
            cur_desc = vals.get("desc", "")
            cur_code = vals.get("code", "")
            cur_unit = vals.get("unit", "NO") or "NO"
            cur_proc = vals.get("proc", "HOME") or "HOME"

        desc_entry = ctk.CTkEntry(form, height=30, fg_color="#1a1a2e",
                                  border_color="#333", border_width=1, corner_radius=5,
                                  text_color="#e8e8e8", font=(MODERN_FONT, 12))
        desc_entry.insert(0, cur_desc)
        add_row("Description:", desc_entry)

        code_entry = ctk.CTkEntry(form, height=30, fg_color="#1a1a2e",
                                  border_color="#333", border_width=1, corner_radius=5,
                                  text_color="#e8e8e8", font=(MODERN_FONT, 12))
        code_entry.insert(0, cur_code)
        add_row("Code:", code_entry)
        # Numeric-only warning label (hidden by default)
        code_warn = ctk.CTkLabel(form, text="Numbers only please.",
                                 font=(MODERN_FONT, 10), text_color="#e74c3c", anchor="w")
        code_warn.pack_forget()

        def _validate_code(char):
            if char.isdigit() or char == "":
                code_warn.pack_forget()
                return True
            code_warn.pack(fill="x", pady=(2, 0))
            self.win.after(2000, lambda: code_warn.pack_forget())
            return False
        code_entry.configure(validate="key",
                             validatecommand=(self.win.register(_validate_code), "%S"))

        unit_combo = ctk.CTkComboBox(form, values=UNIT_OPTIONS, height=30,
                                     fg_color="#1a1a2e", border_color="#333",
                                     border_width=1, corner_radius=5,
                                     text_color="#e8e8e8", button_color="#e8820c",
                                     button_hover_color="#ffa726",
                                     dropdown_fg_color="#1a1a2e",
                                     dropdown_text_color="#e8e8e8",
                                     dropdown_hover_color="#2e8b57")
        unit_combo.set(cur_unit)
        add_row("Unit:", unit_combo)

        # Item codes only use HOME/BLD MAT/SCHOOL procedures; RETAILER and
        # SPCL ECO ZONE are customer-level (TIN) procedures, not item-level.
        _ITEM_PROC_OPTIONS = [p for p in PROCEDURE_OPTIONS
                              if p not in ("RETAILER", "SPCL ECO ZONE")]
        proc_combo = ctk.CTkComboBox(form, values=_ITEM_PROC_OPTIONS, height=30,
                                     fg_color="#1a1a2e", border_color="#333",
                                     border_width=1, corner_radius=5,
                                     text_color="#e8e8e8", button_color="#e8820c",
                                     button_hover_color="#ffa726",
                                     dropdown_fg_color="#1a1a2e",
                                     dropdown_text_color="#e8e8e8",
                                     dropdown_hover_color="#2e8b57")
        # If the existing proc is RETAILER/SPCL ECO ZONE (legacy data),
        # default to HOME since those aren't valid item-level procedures.
        if cur_proc in ("RETAILER", "SPCL ECO ZONE"):
            cur_proc = "HOME"
        proc_combo.set(cur_proc)
        add_row("Procedure:", proc_combo)

        status_lbl = ctk.CTkLabel(dlg, text="", font=(MODERN_FONT, 11),
                                  text_color="#e74c3c", anchor="w")
        status_lbl.pack(fill="x", padx=28, pady=(4, 0))

        def confirm():
            desc = desc_entry.get().strip()
            code = code_entry.get().strip()
            unit = unit_combo.get().strip() or "NO"
            proc = proc_combo.get().strip() or "HOME"
            if not desc:
                status_lbl.configure(text="Description is required.")
                desc_entry.focus_set()
                return
            if not code:
                status_lbl.configure(text="Code is required.")
                code_entry.focus_set()
                return
            # When editing, skip duplicate check if the code hasn't changed
            if not (is_edit and code.lower() == cur_code.strip().lower()):
                # Check for duplicate code in the tree AND the full in-memory
                # database (the tree may be filtered, hiding the duplicate).
                dup_item = None
                dup_desc_from_mem = None
                for item in self._tree.get_children():
                    if item == existing_item:
                        continue
                    if self._tree.set(item, "code").strip().lower() == code.lower():
                        dup_item = item
                        break
                if dup_item is None:
                    for mem_desc, mem_code, _mem_unit, _mem_proc in BUILTIN_CODES:
                        if mem_code.strip().lower() == code.lower():
                            dup_desc_from_mem = mem_desc
                            break
                if dup_item is not None or dup_desc_from_mem is not None:
                    if dup_item is not None:
                        dup_desc = self._tree.set(dup_item, "desc").strip()
                    else:
                        dup_desc = dup_desc_from_mem
                    if not messagebox.askyesno(
                            "Duplicate Code",
                            f"Code '{code}' already exists for:\n\n"
                            f"  {dup_desc}\n\n"
                            f"Do you want to override it with the new details?",
                            parent=dlg):
                        return  # user chose No - stay on the dialog
                    # Override the duplicate row (in tree if visible, else memory)
                    if dup_item is not None:
                        self._tree.set(dup_item, "desc", desc)
                        self._tree.set(dup_item, "unit", unit)
                        self._tree.set(dup_item, "proc", proc)
                        self._tree.see(dup_item)
                        self._tree.selection_set(dup_item)
                    else:
                        for i, (mem_desc, mem_code, _u, _p) in enumerate(BUILTIN_CODES):
                            if mem_code.strip().lower() == code.lower():
                                BUILTIN_CODES[i] = (desc, code, unit, proc)
                                break
                    self._dirty = True
                    self._status.configure(
                        text=f"Updated '{desc}' - click 'Save Changes' to persist")
                    dlg.grab_release()
                    dlg.destroy()
                    return
            # Save: update existing row or insert new
            if is_edit:
                self._tree.set(existing_item, "desc", desc)
                self._tree.set(existing_item, "code", code)
                self._tree.set(existing_item, "unit", unit)
                self._tree.set(existing_item, "proc", proc)
                self._tree.see(existing_item)
                self._tree.selection_set(existing_item)
            else:
                item_id = self._tree.insert("", "end", values=(desc, code, unit, proc))
                self._tree.see(item_id)
                self._tree.selection_set(item_id)
            self._dirty = True
            self._status.configure(
                text=f"{'Updated' if is_edit else 'Added'} '{desc}' - click 'Save Changes' to persist")
            dlg.grab_release()
            dlg.destroy()

        def cancel():
            dlg.grab_release()
            dlg.destroy()

        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=24, pady=(8, 16))
        ctk.CTkButton(btn_frame, text="Cancel", command=cancel,
                      fg_color="#333", hover_color="#444",
                      width=100, height=32, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold"),
                      text_color="#e8e8e8").pack(side="right", padx=(8, 0))
        ctk.CTkButton(btn_frame, text="Save" if is_edit else "Add", command=confirm,
                      fg_color="#2e8b57", hover_color="#3cb371",
                      width=100, height=32, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold")).pack(side="right")

        # Enter key confirms, Escape cancels
        dlg.bind("<Return>", lambda e: confirm())
        dlg.bind("<Escape>", lambda e: cancel())
        desc_entry.focus_set()

    def _delete_selected(self):
        selected = self._tree.selection()
        if not selected:
            return
        if not messagebox.askyesno("Confirm", "Delete the selected code(s)?"):
            return
        for item in selected:
            desc = self._tree.set(item, "desc").strip().lower()
            if desc:
                self._deleted_desc_set.add(desc)
            self._tree.delete(item)
        self._dirty = True
        self._status.configure(text="Deleted - click 'Save Changes' to persist")

    def _save_to_script(self):
        """Save the current tree contents back to the BUILTIN_CODES list in the script file.
        Merges with any changes saved by other users since this window was opened."""
        # Start with the full in-memory list (has ALL codes, not just visible ones)
        my_codes = list(BUILTIN_CODES)
        # Override with any visible tree edits
        visible_codes = {}
        for item in self._tree.get_children():
            vals = self._tree.set(item)
            desc = vals.get("desc", "").strip()
            code = vals.get("code", "").strip()
            unit = vals.get("unit", "NO").strip() or "NO"
            proc = vals.get("proc", "HOME").strip() or "HOME"
            if desc and code:
                visible_codes[desc.lower()] = (desc, code, unit, proc)
        # Merge visible edits into the full list
        if visible_codes:
            my_codes_dict = {desc.lower(): (desc, code, unit, proc) for desc, code, unit, proc in my_codes}
            my_codes_dict.update(visible_codes)
            my_codes = list(my_codes_dict.values())

        # Apply explicit deletions
        deleted_set = getattr(self, '_deleted_desc_set', set())
        my_codes = [(desc, code, unit, proc) for desc, code, unit, proc in my_codes
                    if desc.lower() not in deleted_set]

        if not my_codes and not deleted_set:
            messagebox.showerror("Error", "No valid codes to save.")
            return

        # Re-read the script file to get the latest version from disk
        # (someone else may have saved changes via Dropbox sync)
        script_path = Path(__file__).resolve()
        content = script_path.read_text(encoding="utf-8")

        # Parse the current BUILTIN_CODES block from the file on disk
        import re as _re
        disk_codes = []
        pattern = r'BUILTIN_CODES = \[.*?\n\]'
        m = _re.search(pattern, content, flags=_re.DOTALL)
        if m:
            block_text = m.group(0)
            # Extract tuples from the block
            for line in block_text.split("\n"):
                line = line.strip()
                if line.startswith("(") and line.endswith("),"):
                    try:
                        # Parse: ("desc", "code", "unit", "proc"),
                        inner = line[1:-2]  # strip ( and ),
                        parts = inner.split('", "')
                        if len(parts) == 4:
                            desc = parts[0].lstrip('"')
                            code = parts[1]
                            unit = parts[2]
                            proc = parts[3].rstrip('"')
                            disk_codes.append((desc, code, unit, proc))
                    except Exception:
                        continue

        # Merge: disk entries + my entries, apply explicit deletions
        merged = {}
        for desc, code, unit, proc in disk_codes:
            merged[desc.lower()] = (desc, code, unit, proc)
        for desc, code, unit, proc in my_codes:
            merged[desc.lower()] = (desc, code, unit, proc)
        # Remove explicitly deleted entries
        deleted_set = getattr(self, '_deleted_desc_set', set())
        for key in list(merged.keys()):
            if key in deleted_set:
                del merged[key]

        final_codes = list(merged.values())
        # Sort alphabetically by description
        final_codes.sort(key=lambda x: x[0].lower())

        # Build the new BUILTIN_CODES block
        lines = ["BUILTIN_CODES = ["]
        for desc, code, unit, proc in final_codes:
            safe_desc = desc.replace('\\', '\\\\').replace('"', '\\"')
            lines.append(f'    ("{safe_desc}", "{code}", "{unit}", "{proc}"),')
        lines.append("]")
        new_block = "\n".join(lines)

        new_content = _re.sub(pattern, lambda m: new_block, content, count=1, flags=_re.DOTALL)

        if new_content == content:
            messagebox.showwarning("No Change", "No changes to save.")
            return

        try:
            script_path.write_text(new_content, encoding="utf-8")
            # Update the in-memory list
            BUILTIN_CODES.clear()
            BUILTIN_CODES.extend(final_codes)
            # Rebuild lookup maps
            _BUILTIN_DESC_MAP.clear()
            _BUILTIN_CODE_MAP.clear()
            for _desc, _code, _unit, _proc in final_codes:
                _clean = clean_text(_desc)
                _BUILTIN_DESC_MAP[_clean] = (_code, _unit, _proc)
                _BUILTIN_CODE_MAP[_code] = (_desc, _unit, _proc)
            # Clear deletions and refresh the tree
            self._deleted_desc_set = set()
            self._dirty = False
            self._search_var.set("")  # clear filter to show full list
            self._populate_tree()
            self._status.configure(text=f"Saved {len(final_codes)} codes (merged with disk)")
            messagebox.showinfo("Saved", f"Saved {len(final_codes)} codes to the script file.\n\n"
                                          "Changes merged with any updates from other users\n"
                                          "and will sync via Dropbox.")
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save:\n\n{e}")

    def _on_close(self):
        if self._dirty:
            answer = messagebox.askyesnocancel("Unsaved Changes",
                                                "You have unsaved changes.\n\n"
                                                "Click Yes to save and close.\n"
                                                "Click No to discard and close.\n"
                                                "Click Cancel to go back.")
            if answer is None:
                return  # cancel - don't close
            if answer:
                self._save_to_script()
        self.launcher._codes_win = None
        self.win.destroy()



def _get_tin_entry(cby):
    """Get TIN entry as a 3-tuple (name, tin, special_proc).
    Handles both old 2-tuple and new 3-tuple formats."""
    entry = BUILTIN_TIN_NUMBERS.get(str(cby))
    if entry is None:
        return None
    if len(entry) >= 3:
        return entry
    elif len(entry) == 2:
        return (entry[0], entry[1], "")
    else:
        return (str(entry[0]) if entry else "", "", "")

def _apply_special_proc(cby, current_proc):
    """Apply special procedure override from TIN database.
    RETAILER overrides HOME only (not SCHOOL or BLD MAT).
    SPCL ECO ZONE overrides ALL procedures.
    Returns the (possibly overridden) procedure code."""
    entry = _get_tin_entry(cby)
    if not entry:
        return current_proc
    special = entry[2] if len(entry) > 2 else ""
    if special == "SPCL ECO ZONE":
        return "SPCL ECO ZONE"
    elif special == "RETAILER" and current_proc == "HOME":
        return "RETAILER"
    return current_proc

# ==============================================================================
# TIN NUMBERS MANAGEMENT WINDOW
# ==============================================================================
class TINWindow(SupportMixin):
    """Searchable, editable table of customer TIN/importer numbers.
    Supports Quick Sync from the AoA Log Master Excel file.
    Changes save back to the script file (syncs via Dropbox)."""

    def __init__(self, launcher):
        self.launcher = launcher

        # Support-mixin state
        self._pending_update = None
        self._support_tooltip = "Report a Bug"
        self._tooltip_win = None
        self._support_bg = "#0f1117"
        self._window_name = "TIN Numbers"
        self._set_support_palette({
            "bg": "#0f1117", "panel": "#141a2a", "input": "#0a0e16",
            "border": "#1a2a4a", "text": "#c8d6e5", "light": "#5a7a9a",
            "accent": "#0f3460", "accent_hover": "#1a4a7a",
        })

        self.win = ctk.CTkToplevel()
        self.win.title(f"TIN Numbers Management v{APP_VERSION}")
        self.win.configure(fg_color="#0f1117")
        _register_window_name(self.win, self._window_name, {"bg": "#0f1117", "panel": "#141a2a", "accent": "#0f3460", "accent_hover": "#1a4a7a", "text": "#c8d6e5"})
        self.win.geometry("960x600")
        sw, sh = self.win.winfo_screenwidth(), self.win.winfo_screenheight()
        self.win.geometry(f"+{int((sw-960)/2)}+{int((sh-600)/2)}")
        self.win.protocol("WM_DELETE_WINDOW", self._on_close)
        self.win.transient(launcher.root)

        # Title bar with Back button
        title_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        title_frame.pack(fill="x", padx=16, pady=(12, 0))
        ctk.CTkButton(title_frame, text="\u2190 Back",
                      command=self._on_close,
                      fg_color="#333", hover_color="#444",
                      width=80, height=28, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold"),
                      text_color="#e8e8e8").pack(side="left")
        ctk.CTkLabel(title_frame, text="TIN Numbers Database",
                     font=(MODERN_FONT, 18, "bold"), text_color="#e8e8e8").pack(side="left", padx=(16, 0))

        # Subtitle
        ctk.CTkLabel(self.win,
                     text=f"{len(BUILTIN_TIN_NUMBERS)} customers loaded  -  double-click a cell to edit  -  Special Procedure requires CBC approval  -  fallback TIN: {FALLBACK_IMPORTER_NUMBER}",
                     font=(MODERN_FONT, 11), text_color="#888").pack(pady=(8, 12))

        # Search + action bar
        search_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        search_frame.pack(fill="x", padx=20, pady=(0, 8))
        ctk.CTkLabel(search_frame, text="Search:", font=(MODERN_FONT, 12),
                     text_color="#ccc").pack(side="left", padx=(0, 8))
        self._search_var = ctk.StringVar()
        self._search_var.trace("w", lambda *a: self._filter_tree())
        search_entry = ctk.CTkEntry(search_frame, textvariable=self._search_var,
                                    width=250, height=30, fg_color="#1a1a2e",
                                    border_color="#333", border_width=1, corner_radius=5,
                                    text_color="#e8e8e8", font=(MODERN_FONT, 12))
        search_entry.pack(side="left", padx=(0, 8))

        ctk.CTkButton(search_frame, text="Quick Sync with AoA List",
                      command=self._quick_sync,
                      fg_color="#2e8b57", hover_color="#3cb371",
                      width=180, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        ctk.CTkButton(search_frame, text="Save Changes",
                      command=self._save_to_script,
                      fg_color="#e8820c", hover_color="#ffa726",
                      width=120, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="right", padx=4)

        # Treeview
        tree_frame = ctk.CTkFrame(self.win, fg_color="#1a1a2e", corner_radius=6)
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 8))

        cols = ["cby", "name", "tin", "sproc"]
        labels = ["CBY #", "Customer Name", "TIN #", "Special Procedure"]
        widths = [80, 350, 120, 160]
        self._tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        for col, label, w in zip(cols, labels, widths):
            self._tree.heading(col, text=label, anchor="w")
            self._tree.column(col, width=w, minwidth=w, anchor="w",
                              stretch=(col == "name"))

        style = ttk.Style()
        style.configure("Treeview", background="#1a1a2e", foreground="#e8e8e8",
                        fieldbackground="#1a1a2e", bordercolor="#333",
                        rowheight=26, font=(MODERN_FONT, 11))
        style.configure("Treeview.Heading", background="#0f1117", foreground="#ffa726",
                        font=(MODERN_FONT, 11, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", "#2e8b57")],
                  foreground=[("selected", "#fff")])

        tree_scroll = ctk.CTkScrollbar(tree_frame, command=self._tree.yview)
        self._tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side="right", fill="y", padx=(2, 4), pady=4)
        self._tree.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)

        self._tree.bind("<Double-1>", self._on_double_click)
        self._tree.bind("<Tab>", self._tin_tab_next)
        self._tree.bind("<Shift-Tab>", self._tin_tab_prev)
        self._editing_win = None
        self._editing_row_id = None
        self._editing_key = None
        self._deleted_cby_set = set()
        self._dirty = False

        # Delete button
        del_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        del_frame.pack(fill="x", padx=20, pady=(0, 8))
        ctk.CTkButton(del_frame, text="Delete Selected",
                      command=self._delete_selected,
                      fg_color="#c0392b", hover_color="#e74c3c",
                      width=130, height=28, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        ctk.CTkButton(del_frame, text="+ Add Customer",
                      command=self._add_entry,
                      fg_color="#2e8b57", hover_color="#3cb371",
                      width=130, height=28, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)

        # Status (with support icon in the corner)
        status_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        status_frame.pack(fill="x", padx=22, pady=(0, 8))
        self._status = ctk.CTkLabel(status_frame, text="Ready", font=(MODERN_FONT, 11),
                                    text_color="#888", anchor="w")
        self._status.pack(side="left", fill="x", expand=True)
        self._attach_support_icon(status_frame)

        # Populate
        self._populate_tree()

        # Check for updates in the background
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    def _sync_tree_to_memory(self):
        """Sync current treeview values back to BUILTIN_TIN_NUMBERS in memory.
        This preserves edits made in filtered mode when the tree is repopulated."""
        _DISPLAY_TO_CODE = {"": "", "Retailer": "RETAILER", "Special Economic Zone": "SPCL ECO ZONE"}
        for item in self._tree.get_children():
            vals = self._tree.set(item)
            cby = vals.get("cby", "").strip()
            if not cby:
                continue
            name = vals.get("name", "").strip()
            tin = vals.get("tin", "").strip()
            sproc_display = vals.get("sproc", "").strip()
            sproc = _DISPLAY_TO_CODE.get(sproc_display, sproc_display)
            BUILTIN_TIN_NUMBERS[cby] = (name, tin, sproc)

    def _populate_tree(self, filter_text="", skip_memory_sync=False):
        # Sync any tree edits back to memory before repopulating.
        # skip_memory_sync=True is used after Quick Sync, which already
        # updated BUILTIN_TIN_NUMBERS directly — syncing the stale tree
        # values back would clobber those updates.
        if not skip_memory_sync:
            self._sync_tree_to_memory()
        for item in self._tree.get_children():
            self._tree.delete(item)
        _SPROC_DISPLAY = {"": "", "RETAILER": "Retailer", "SPCL ECO ZONE": "Special Economic Zone"}
        for cby, entry in sorted(BUILTIN_TIN_NUMBERS.items(),
                                        key=lambda x: int(x[0]) if x[0].isdigit() else 0):
            name = entry[0] if entry else ""
            tin = entry[1] if len(entry) > 1 else ""
            sproc = entry[2] if len(entry) > 2 else ""
            sproc_display = _SPROC_DISPLAY.get(sproc, sproc)
            if filter_text:
                combined = f"{cby} {name} {tin} {sproc_display}".lower()
                if filter_text.lower() not in combined:
                    continue
            self._tree.insert("", "end", values=(cby, name, tin, sproc_display))
        # Clear deletion tracking when full list is reloaded
        if not filter_text:
            self._deleted_cby_set = set()

    def _filter_tree(self, skip_memory_sync=False):
        self._populate_tree(self._search_var.get().strip(),
                            skip_memory_sync=skip_memory_sync)

    def _on_double_click(self, event):
        if self._editing_win is not None:
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
        region = self._tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            return
        self._add_entry(existing_item=row_id)

    def _tin_tab_next(self, event):
        if self._editing_win is not None:
            return
        sel = self._tree.selection()
        if not sel:
            return
        row_id = sel[0]
        cols = ["cby", "name", "tin", "sproc"]
        cur_col = self._tree.identify_column(event.x)
        cur_idx = int(cur_col.replace("#", "")) - 1 if cur_col.startswith("#") else 0
        next_idx = cur_idx + 1
        if next_idx >= len(cols):
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx + 1 < len(children):
                    row_id = children[cur_row_idx + 1]
                    next_idx = 0
                else:
                    return
            except ValueError:
                return
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)
        self._tree.see(row_id)
        self._tin_start_edit(row_id, cols[next_idx])
        return "break"

    def _tin_tab_prev(self, event):
        if self._editing_win is not None:
            return
        sel = self._tree.selection()
        if not sel:
            return
        row_id = sel[0]
        cols = ["cby", "name", "tin", "sproc"]
        cur_col = self._tree.identify_column(event.x)
        cur_idx = int(cur_col.replace("#", "")) - 1 if cur_col.startswith("#") else 0
        prev_idx = cur_idx - 1
        if prev_idx < 0:
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx > 0:
                    row_id = children[cur_row_idx - 1]
                    prev_idx = len(cols) - 1
                else:
                    return
            except ValueError:
                return
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)
        self._tree.see(row_id)
        self._tin_start_edit(row_id, cols[prev_idx])
        return "break"

    def _tin_start_edit(self, row_id, key):
        """Open edit popup on a cell. Shared by double-click and Tab nav."""
        if self._editing_win is not None:
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
        cols = ["cby", "name", "tin", "sproc"]
        col_idx = cols.index(key)
        col = f"#{col_idx + 1}"
        bbox = self._tree.bbox(row_id, col)
        if not bbox:
            return
        x, y, w, h = bbox
        cx = self._tree.winfo_rootx() + x
        cy = self._tree.winfo_rooty() + y
        current_val = self._tree.set(row_id, key)
        self._editing_win = tk.Toplevel(self.win)
        self._editing_win.overrideredirect(True)
        self._editing_win.geometry(f"{w}x{h}+{cx}+{cy}")
        self._editing_win.attributes("-topmost", True)
        self._editing_row_id = row_id
        self._editing_key = key
        self._tree.selection_set(row_id)
        self._tree.focus(row_id)

        def tab_next():
            self._tin_save_current_and_tab(1)
        def tab_prev():
            self._tin_save_current_and_tab(-1)

        if key == "sproc":
            # Dropdown for Special Procedure
            _SPROC_OPTIONS = ["", "Retailer", "Special Economic Zone"]
            widget = ctk.CTkComboBox(self._editing_win, values=_SPROC_OPTIONS,
                                     width=w, height=h, fg_color="#1a1a2e",
                                     border_color="#e8820c", border_width=2, corner_radius=2,
                                     text_color="#e8e8e8", font=(MODERN_FONT, 11),
                                     button_color="#e8820c", button_hover_color="#ffa726",
                                     dropdown_fg_color="#1a1a2e",
                                     dropdown_text_color="#e8e8e8",
                                     dropdown_hover_color="#2a2a4e")
            widget.set(current_val)
            widget.pack(fill="both", expand=True)
            widget.focus_set()

            _sproc_resolved = [False]
            def _on_sproc_select(value=None):
                if _sproc_resolved[0]:
                    return
                val = value if value is not None else widget.get()
                # Check if user selected a special procedure that needs confirmation
                if val and val != current_val:
                    _sproc_resolved[0] = True
                    if not self._confirm_special_proc(val):
                        # User cancelled - revert
                        widget.set(current_val)
                        self._cancel_edit()
                        return
                    _sproc_resolved[0] = False  # allow further edits
                self._finish_edit(row_id, key, val)
                _sproc_resolved[0] = True

            widget.bind("<<ComboboxSelected>>", lambda e: _on_sproc_select(widget.get()))
            widget.bind("<Return>", lambda e: _on_sproc_select(widget.get()))
            widget.bind("<Escape>", lambda e: self._cancel_edit())
            widget.bind("<FocusOut>", lambda e: _on_sproc_select(widget.get()))
            widget.bind("<Tab>", lambda e: (tab_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_prev(), "break"))
        else:
            widget = ctk.CTkEntry(self._editing_win, width=w, height=h,
                                  fg_color="#1a1a2e", border_color="#e8820c",
                                  border_width=2, corner_radius=2, text_color="#e8e8e8",
                                  font=(MODERN_FONT, 11))
            widget.pack(fill="both", expand=True)
            widget.insert(0, current_val)
            widget.focus_set()
            widget.select_range(0, "end")
            widget.bind("<Return>", lambda e: (self._finish_edit(row_id, key, widget.get()), self._tin_tab_to_cell(1)))
            widget.bind("<Escape>", lambda e: self._cancel_edit())
            widget.bind("<FocusOut>", lambda e: self._finish_edit(row_id, key, widget.get()))
            widget.bind("<Tab>", lambda e: (tab_next(), "break"))
            widget.bind("<Shift-Tab>", lambda e: (tab_prev(), "break"))

    def _tin_save_current_and_tab(self, direction):
        """Save current edit and move to next/prev cell."""
        if self._editing_row_id is None:
            return
        row_id = self._editing_row_id
        key = self._editing_key
        for child in self._editing_win.winfo_children():
            try:
                value = child.get()
                break
            except:
                continue
        else:
            return
        self._finish_edit(row_id, key, value)
        self.win.after(10, lambda: self._tin_tab_to_cell(direction))

    def _tin_tab_to_cell(self, direction):
        """Move to next (1) or prev (-1) cell after saving."""
        if self._editing_row_id is None:
            return
        row_id = self._editing_row_id
        cols = ["cby", "name", "tin", "sproc"]
        key = self._editing_key
        if key not in cols:
            return
        cur_idx = cols.index(key)
        next_idx = cur_idx + direction
        if next_idx >= len(cols):
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx + 1 < len(children):
                    row_id = children[cur_row_idx + 1]
                    next_idx = 0
                else:
                    return
            except ValueError:
                return
        elif next_idx < 0:
            children = self._tree.get_children()
            try:
                cur_row_idx = children.index(row_id)
                if cur_row_idx > 0:
                    row_id = children[cur_row_idx - 1]
                    next_idx = len(cols) - 1
                else:
                    return
            except ValueError:
                return
        self._tin_start_edit(row_id, cols[next_idx])

    def _finish_edit(self, row_id, key, value):
        if self._editing_win is not None:
            try:
                self._tree.set(row_id, key, value)
            except Exception:
                pass
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
            self._dirty = True
            self._status.configure(text="Edited - click 'Save Changes' to persist")

    def _cancel_edit(self):
        if self._editing_win is not None:
            try:
                self._editing_win.destroy()
            except Exception:
                pass
            self._editing_win = None
        self._editing_row_id = None
        self._editing_key = None

    def _confirm_special_proc(self, display_name, parent=None):
        """Show a confirmation popup when assigning a special procedure.
        Returns True if user confirms, False otherwise.
        If parent is provided, the dialog is transient to that widget
        (so it appears on top of an already-grabbed dialog)."""
        _DISPLAY_TO_CODE = {"Retailer": "RETAILER", "Special Economic Zone": "SPCL ECO ZONE"}
        code = _DISPLAY_TO_CODE.get(display_name, display_name)
        owner = parent if parent is not None else self.win
        dialog = ctk.CTkToplevel(owner)
        dialog.title("Special Procedure Confirmation")
        dialog.configure(fg_color="#1a1a2e")
        if display_name == "Retailer":
            dialog.geometry("520x320")
            dw, dh = 520, 320
        else:
            dialog.geometry("480x260")
            dw, dh = 480, 260
        dialog.resizable(False, False)
        dialog.transient(owner)
        dialog.attributes("-topmost", True)
        dialog.grab_set()
        dialog.update_idletasks()
        px = owner.winfo_rootx() + (owner.winfo_width() - dw) // 2
        py = owner.winfo_rooty() + (owner.winfo_height() - dh) // 2
        dialog.geometry(f"+{px}+{py}")
        ctk.CTkLabel(dialog, text="Special Procedure Assignment",
                     font=(MODERN_FONT, 14, "bold"), text_color="#e8e8e8").pack(pady=(20, 4))
        if display_name == "Retailer":
            ctk.CTkLabel(dialog,
                         text=f"You are assigning: {display_name}\n\n"
                              f"Retailer will only work if ONE of the following is true:\n\n"
                              f"1. The customer has RETAILER granted by CBC\n"
                              f"   on their own account (permanent).\n\n"
                              f"2. We have delegated Retailer to the customer,\n"
                              f"   which must be accepted first and expires after 60 days.\n\n"
                              f"Otherwise the declaration will be rejected by COLS.\n"
                              f"Has this been confirmed?",
                         font=(MODERN_FONT, 12), text_color="#ff6b6b",
                         justify="center").pack(pady=(0, 16))
        else:
            ctk.CTkLabel(dialog,
                         text=f"You are assigning: {display_name}\n\n"
                              f"This procedure requires prior approval from CBC.\n"
                              f"The customer must already be granted this permission\n"
                              f"in COLS, otherwise the declaration will be rejected.\n\n"
                              f"Has CBC approved this special procedure for this customer?",
                         font=(MODERN_FONT, 12), text_color="#ff6b6b",
                         justify="center").pack(pady=(0, 16))
        choice = [None]
        def on_confirm():
            choice[0] = True
            dialog.destroy()
        def on_cancel():
            choice[0] = False
            dialog.destroy()
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=(0, 20))
        ctk.CTkButton(btn_frame, text="Cancel - Not Yet Assigned", command=on_cancel,
                     fg_color="#c0392b", hover_color="#e74c3c",
                     width=180, height=32, corner_radius=6,
                     text_color="#ffffff", font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=(20, 10))
        ctk.CTkButton(btn_frame, text="Yes - Already Approved by CBC", command=on_confirm,
                     fg_color="#2e8b57", hover_color="#3cb371",
                     width=210, height=32, corner_radius=6,
                     text_color="#ffffff", font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=(10, 20))
        dialog.bind("<Escape>", lambda e: on_cancel())
        dialog.wait_window()
        return choice[0] is True

    def _add_entry(self, existing_item=None):
        """Open a popup dialog to enter (or edit) a customer.
        If existing_item is provided, the dialog is pre-filled with that
        row's values and saves update the row instead of inserting.
        If the name AND TIN match an existing entry (when adding new),
        the user is asked whether to override it."""
        _SPROC_OPTIONS = ["", "Retailer", "Special Economic Zone"]
        is_edit = existing_item is not None
        dlg = ctk.CTkToplevel(self.win)
        dlg.title("Edit Customer" if is_edit else "Add Customer")
        dlg.configure(fg_color="#0f1117")
        dlg.geometry("460x560")
        dlg.transient(self.win)
        dlg.attributes("-topmost", True)
        # Center over the parent window
        self.win.update_idletasks()
        px = self.win.winfo_rootx() + (self.win.winfo_width() - 460) // 2
        py = self.win.winfo_rooty() + (self.win.winfo_height() - 560) // 2
        dlg.geometry(f"+{px}+{py}")
        dlg.grab_set()

        ctk.CTkLabel(dlg, text="Edit Customer" if is_edit else "Add Customer",
                     font=(MODERN_FONT, 16, "bold"),
                     text_color="#e8e8e8").pack(pady=(16, 4))
        ctk.CTkLabel(dlg, text="Fill in the details and click Save." if is_edit
                     else "Fill in the details and click Add.",
                     font=(MODERN_FONT, 11), text_color="#888").pack(pady=(0, 12))

        form = ctk.CTkFrame(dlg, fg_color="transparent")
        form.pack(fill="x", padx=24, pady=(0, 8))

        def add_row(label_text, widget):
            ctk.CTkLabel(form, text=label_text, anchor="w",
                         font=(MODERN_FONT, 12), text_color="#ccc").pack(fill="x", pady=(8, 2))
            widget.pack(fill="x")

        # Pre-fill values if editing
        cur_cby, cur_name, cur_tin, cur_sproc = "", "", "", ""
        if is_edit:
            vals = self._tree.set(existing_item)
            cur_cby = vals.get("cby", "")
            cur_name = vals.get("name", "")
            cur_tin = vals.get("tin", "")
            cur_sproc = vals.get("sproc", "")

        cby_entry = ctk.CTkEntry(form, height=30, fg_color="#1a1a2e",
                                 border_color="#333", border_width=1, corner_radius=5,
                                 text_color="#e8e8e8", font=(MODERN_FONT, 12))
        cby_entry.insert(0, cur_cby)
        add_row("CBY #:", cby_entry)

        name_entry = ctk.CTkEntry(form, height=30, fg_color="#1a1a2e",
                                  border_color="#333", border_width=1, corner_radius=5,
                                  text_color="#e8e8e8", font=(MODERN_FONT, 12))
        name_entry.insert(0, cur_name)
        add_row("Customer Name:", name_entry)

        tin_entry = ctk.CTkEntry(form, height=30, fg_color="#1a1a2e",
                                 border_color="#333", border_width=1, corner_radius=5,
                                 text_color="#e8e8e8", font=(MODERN_FONT, 12))
        tin_entry.insert(0, cur_tin)
        add_row("TIN #:", tin_entry)
        # Numeric-only warning label (hidden by default)
        tin_warn = ctk.CTkLabel(form, text="Numbers only please.",
                                font=(MODERN_FONT, 10), text_color="#e74c3c", anchor="w")
        tin_warn.pack_forget()

        def _validate_tin(char):
            if char.isdigit() or char == "":
                tin_warn.pack_forget()
                return True
            tin_warn.pack(fill="x", pady=(2, 0))
            self.win.after(2000, lambda: tin_warn.pack_forget())
            return False
        tin_entry.configure(validate="key",
                            validatecommand=(self.win.register(_validate_tin), "%S"))

        sproc_combo = ctk.CTkComboBox(form, values=_SPROC_OPTIONS, height=30,
                                      fg_color="#1a1a2e", border_color="#333",
                                      border_width=1, corner_radius=5,
                                      text_color="#e8e8e8", button_color="#e8820c",
                                      button_hover_color="#ffa726",
                                      dropdown_fg_color="#1a1a2e",
                                      dropdown_text_color="#e8e8e8",
                                      dropdown_hover_color="#2e8b57")
        sproc_combo.set(cur_sproc)
        add_row("Special Procedure:", sproc_combo)

        status_lbl = ctk.CTkLabel(dlg, text="", font=(MODERN_FONT, 11),
                                  text_color="#e74c3c", anchor="w")
        status_lbl.pack(fill="x", padx=28, pady=(4, 0))

        def confirm():
            cby = cby_entry.get().strip()
            name = name_entry.get().strip()
            tin = tin_entry.get().strip()
            sproc_display = sproc_combo.get().strip()
            if not cby:
                status_lbl.configure(text="CBY # is required.")
                cby_entry.focus_set()
                return
            if not name:
                status_lbl.configure(text="Customer Name is required.")
                name_entry.focus_set()
                return
            # Confirm special procedure assignment if it's Retailer or
            # Special Economic Zone (and has changed from the current value)
            if sproc_display in ("Retailer", "Special Economic Zone"):
                if sproc_display != cur_sproc.strip():
                    if not self._confirm_special_proc(sproc_display, parent=dlg):
                        return  # user cancelled - stay on the dialog
            # When editing, skip duplicate check if name+TIN haven't changed
            unchanged = (is_edit and name.lower() == cur_name.strip().lower()
                         and tin.lower() == cur_tin.strip().lower())
            if not unchanged:
                # Check for duplicate: name AND TIN match an existing entry.
                # Search both the visible tree AND the full in-memory database
                # (the tree may be filtered, hiding the duplicate).
                self._sync_tree_to_memory()
                dup_item = None
                # First check the visible tree
                for item in self._tree.get_children():
                    if item == existing_item:
                        continue
                    ex_name = self._tree.set(item, "name").strip().lower()
                    ex_tin = self._tree.set(item, "tin").strip().lower()
                    if name and name.lower() == ex_name and tin.lower() == ex_tin:
                        dup_item = item
                        break
                # If not found in tree, check the full in-memory database
                dup_cby_from_mem = None
                if dup_item is None:
                    for mem_cby, (mem_name, mem_tin, _mem_sproc) in BUILTIN_TIN_NUMBERS.items():
                        if is_edit and mem_cby == cur_cby.strip():
                            continue
                        if (name and name.lower() == mem_name.strip().lower()
                                and tin.lower() == mem_tin.strip().lower()):
                            dup_cby_from_mem = mem_cby
                            break
                if dup_item is not None or dup_cby_from_mem is not None:
                    if dup_item is not None:
                        ex_cby = self._tree.set(dup_item, "cby").strip()
                    else:
                        ex_cby = dup_cby_from_mem
                    if not messagebox.askyesno(
                            "Duplicate Customer",
                            f"A customer '{name}' with TIN '{tin or '(none)'}' already exists"
                            f" (CBY #{ex_cby}).\n\n"
                            f"Do you want to override it with the new details?",
                            parent=dlg):
                        return  # user chose No - stay on the dialog
                    # Override the duplicate row (in memory if not visible in tree)
                    if dup_item is not None:
                        self._tree.set(dup_item, "cby", cby)
                        self._tree.set(dup_item, "name", name)
                        self._tree.set(dup_item, "tin", tin)
                        self._tree.set(dup_item, "sproc", sproc_display)
                        self._tree.see(dup_item)
                        self._tree.selection_set(dup_item)
                    else:
                        # Entry exists in memory but not in filtered tree -
                        # update memory directly and mark old CBY for deletion
                        self._deleted_cby_set.add(dup_cby_from_mem)
                        BUILTIN_TIN_NUMBERS[cby] = (name, tin,
                                                    {"": "", "Retailer": "RETAILER",
                                                     "Special Economic Zone": "SPCL ECO ZONE"}.get(sproc_display, sproc_display))
                    self._dirty = True
                    self._status.configure(
                        text=f"Updated '{name}' - click 'Save Changes' to persist")
                    dlg.grab_release()
                    dlg.destroy()
                    return
            # Save: update existing row or insert new
            if is_edit:
                self._tree.set(existing_item, "cby", cby)
                self._tree.set(existing_item, "name", name)
                self._tree.set(existing_item, "tin", tin)
                self._tree.set(existing_item, "sproc", sproc_display)
                self._tree.see(existing_item)
                self._tree.selection_set(existing_item)
            else:
                item_id = self._tree.insert("", "end", values=(cby, name, tin, sproc_display))
                self._tree.see(item_id)
                self._tree.selection_set(item_id)
            self._dirty = True
            self._status.configure(
                text=f"{'Updated' if is_edit else 'Added'} '{name}' - click 'Save Changes' to persist")
            dlg.grab_release()
            dlg.destroy()

        def cancel():
            dlg.grab_release()
            dlg.destroy()

        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=24, pady=(8, 16))
        ctk.CTkButton(btn_frame, text="Cancel", command=cancel,
                      fg_color="#333", hover_color="#444",
                      width=100, height=32, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold"),
                      text_color="#e8e8e8").pack(side="right", padx=(8, 0))
        ctk.CTkButton(btn_frame, text="Save" if is_edit else "Add", command=confirm,
                      fg_color="#2e8b57", hover_color="#3cb371",
                      width=100, height=32, corner_radius=6,
                      font=(MODERN_FONT, 11, "bold")).pack(side="right")

        # Enter key confirms, Escape cancels
        dlg.bind("<Return>", lambda e: confirm())
        dlg.bind("<Escape>", lambda e: cancel())
        cby_entry.focus_set()

    def _delete_selected(self):
        selected = self._tree.selection()
        if not selected:
            return
        if not messagebox.askyesno("Confirm", "Delete the selected entry(s)?"):
            return
        for item in selected:
            cby = self._tree.set(item, "cby").strip()
            if cby:
                self._deleted_cby_set.add(cby)
            self._tree.delete(item)
        self._dirty = True
        self._status.configure(text="Deleted - click 'Save Changes' to persist")

    def _quick_sync(self):
        """Open a file picker to select an AoA Log Master Excel file,
        then update TIN numbers from it with a progress bar."""
        file_path = filedialog.askopenfilename(
            title="Select AoA Log Master Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not file_path:
            return

        # Show progress bar window IMMEDIATELY (before file read)
        prog_win = ctk.CTkToplevel(self.win)
        prog_win.title("Syncing")
        prog_win.configure(fg_color="#0f1117")
        prog_win.geometry("400x120")
        sw, sh = prog_win.winfo_screenwidth(), prog_win.winfo_screenheight()
        prog_win.geometry(f"+{int((sw-400)/2)}+{int((sh-120)/2)}")
        prog_win.transient(self.win)
        prog_win.grab_set()
        prog_win.protocol("WM_DELETE_WINDOW", lambda: None)  # prevent closing

        ctk.CTkLabel(prog_win, text="Syncing with AoA List...",
                     font=(MODERN_FONT, 14, "bold"), text_color="#e8e8e8").pack(pady=(16, 8))
        self._sync_progress = ctk.CTkProgressBar(prog_win, width=340, height=16,
                                                 progress_color="#2e8b57",
                                                 fg_color="#1a1a2e")
        self._sync_progress.set(0)
        self._sync_progress.pack(pady=(0, 8))
        self._sync_status_label = ctk.CTkLabel(prog_win, text="Opening file...",
                                               font=(MODERN_FONT, 11), text_color="#888")
        self._sync_status_label.pack()
        prog_win.update_idletasks()

        # Run everything in background thread (openpyxl file read + sync)
        def do_sync():
            # Use openpyxl directly — it's already imported elsewhere in the
            # app, so there's no heavy import delay (unlike pandas).
            try:
                import openpyxl
            except Exception as e:
                self.win.after(0, lambda: self._sync_error(prog_win, f"Could not load Excel engine:\n\n{e}"))
                return

            self.win.after(0, lambda: self._update_sync_progress(0, "Reading file..."))

            # Open the workbook (read-only, data_only for cached values)
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            except Exception as e:
                self.win.after(0, lambda: self._sync_error(prog_win, f"Could not open the file:\n\n{e}"))
                return

            # Check for the AOA Log sheet (case-insensitive, space-insensitive)
            sheet_found = None
            for sn in wb.sheetnames:
                if sn.strip().lower() == "aoa log":
                    sheet_found = sn
                    break
            if sheet_found is None:
                self.win.after(0, lambda: self._sync_error(prog_win,
                    f"This does not appear to be an AoA Log Master file.\n\n"
                    f"It must contain a sheet named 'AOA Log' with columns:\n"
                    f"  CBY#, Full Name, Tin #\n\n"
                    f"Sheets found in this file: {wb.sheetnames}"))
                wb.close()
                return

            ws = wb[sheet_found]

            # Read all rows into a list (read_only mode requires iteration)
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(all_rows) < 1:
                self.win.after(0, lambda: self._sync_error(prog_win, "The AOA Log sheet is empty."))
                return

            # Validate the header row looks right
            header = [str(v).strip().lower() if v is not None else '' for v in all_rows[0]]
            if 'cby' not in header[0].lower() or 'name' not in header[1].lower():
                self.win.after(0, lambda: self._sync_error(prog_win,
                    f"The AOA Log sheet doesn't have the expected columns.\n\n"
                    f"Expected: CBY#, Full Name, Tin #\n"
                    f"Found: {header[:4]}"))
                return

            # Process rows
            added = 0
            updated = 0
            unchanged = 0
            total = len(all_rows) - 1
            for i in range(1, len(all_rows)):
                row = all_rows[i]
                cby = str(row[0]).strip() if row[0] is not None else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                tin = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                if not cby or cby.lower() == 'nan':
                    continue
                if not name or name.lower() in ('nan', '', 'none', '#n/a', '#n/a!', '#ref!', '#value!'):
                    continue
                try:
                    cby_int = str(int(float(cby)))
                except:
                    cby_int = cby
                if not tin or tin.lower() in ('nan', '', '0', 'none'):
                    tin_int = ''
                else:
                    try:
                        tin_int = str(int(float(tin)))
                    except:
                        # TIN contains letters (e.g. "rejected", "pending approval")
                        # — ignore it so the fallback TIN is used instead
                        tin_int = ''

                existing = _get_tin_entry(cby_int)
                if existing is None:
                    BUILTIN_TIN_NUMBERS[cby_int] = (name, tin_int, "")
                    added += 1
                elif tin_int and existing[1] != tin_int:
                    new_name = name if name and name.lower() != 'nan' else existing[0]
                    sproc = existing[2] if len(existing) > 2 else ""
                    BUILTIN_TIN_NUMBERS[cby_int] = (new_name, tin_int, sproc)
                    updated += 1
                elif not existing[0] and name and name.lower() != 'nan':
                    sproc = existing[2] if len(existing) > 2 else ""
                    BUILTIN_TIN_NUMBERS[cby_int] = (name, existing[1], sproc)
                    updated += 1
                else:
                    unchanged += 1

                # Update progress every 50 rows
                if i % 50 == 0 or i == total:
                    progress = i / total
                    self.win.after(0, lambda p=progress, r=i, t=total:
                                   self._update_sync_progress(p, f"Processing row {r} of {t}..."))

            # Done - update UI from main thread
            self.win.after(0, lambda: self._finish_sync(prog_win, added, updated, unchanged))

        threading.Thread(target=do_sync, daemon=True).start()

    def _sync_error(self, prog_win, message):
        """Close progress window and show error (called from main thread)."""
        prog_win.destroy()
        self._sync_progress = None
        self._sync_status_label = None
        messagebox.showerror("Sync Error", message)

    def _update_sync_progress(self, progress, status_text):
        """Update the sync progress bar (called from main thread via .after)."""
        if hasattr(self, '_sync_progress') and self._sync_progress:
            self._sync_progress.set(progress)
        if hasattr(self, '_sync_status_label') and self._sync_status_label:
            self._sync_status_label.configure(text=status_text)

    def _finish_sync(self, prog_win, added, updated, unchanged):
        """Called when sync is complete - close progress bar and show results."""
        prog_win.destroy()
        self._sync_progress = None
        self._sync_status_label = None
        if added > 0 or updated > 0:
            self._dirty = True
        self._filter_tree(skip_memory_sync=True)
        self._status.configure(text=f"Sync complete: {added} new, {updated} updated, {unchanged} unchanged")
        messagebox.showinfo("Sync Complete",
                            f"Added: {added} new entries\n"
                            f"Updated: {updated} entries\n"
                            f"Unchanged: {unchanged} entries\n\n"
                            f"Click 'Save Changes' to persist to the script file.")

    def _save_to_script(self):
        """Save the current tree contents back to BUILTIN_TIN_NUMBERS in the script file.
        Merges with any changes saved by other users since this window was opened.
        Works correctly even when the tree is filtered."""
        # Start with the full in-memory dict (has ALL entries, not just visible ones)
        my_data = dict(BUILTIN_TIN_NUMBERS)
        # Override with any visible tree edits (the tree may be filtered, so only
        # override what's actually shown - those are the ones the user may have edited)
        for item in self._tree.get_children():
            vals = self._tree.set(item)
            cby = vals.get("cby", "").strip()
            name = vals.get("name", "").strip()
            tin = vals.get("tin", "").strip()
            sproc_display = vals.get("sproc", "").strip()
            _DISPLAY_TO_CODE = {"": "", "Retailer": "RETAILER", "Special Economic Zone": "SPCL ECO ZONE"}
            sproc = _DISPLAY_TO_CODE.get(sproc_display, sproc_display)
            if cby:
                my_data[cby] = (name, tin, sproc)

        # Apply explicit deletions
        deleted_set = getattr(self, '_deleted_cby_set', set())
        for cby in deleted_set:
            my_data.pop(cby, None)

        if not my_data and not deleted_set:
            messagebox.showerror("Error", "No valid entries to save.")
            return

        # Re-read the script file to get the latest version from disk
        script_path = Path(__file__).resolve()
        content = script_path.read_text(encoding="utf-8")

        # Parse the current BUILTIN_TIN_NUMBERS block from the file on disk
        import re as _re
        disk_data = {}
        pattern = r'BUILTIN_TIN_NUMBERS = \{.*?\n\}'
        m = _re.search(pattern, content, flags=_re.DOTALL)
        if m:
            block_text = m.group(0)
            for line in block_text.split("\n"):
                line = line.strip()
                if line.startswith('"') and ": (" in line:
                    try:
                        cby_part, rest = line.split('": (', 1)
                        cby = cby_part.lstrip('"')
                        rest = rest.rstrip(',').rstrip(')').rstrip(')')
                        inner = rest.strip()
                        if inner.startswith('('):
                            inner = inner[1:]
                        parts = inner.split('", "')
                        if len(parts) == 2:
                            name = parts[0].lstrip('"')
                            tin = parts[1].rstrip('"')
                            disk_data[cby] = (name, tin, "")
                        elif len(parts) == 3:
                            name = parts[0].lstrip('"')
                            tin = parts[1]
                            sproc = parts[2].rstrip('"')
                            disk_data[cby] = (name, tin, sproc)
                    except Exception:
                        continue

        # Merge: disk entries that aren't in my_data and weren't deleted = added by someone else
        merged = dict(disk_data)
        merged.update(my_data)  # my edits override disk
        # Remove explicitly deleted entries
        for cby in deleted_set:
            merged.pop(cby, None)

        # Build the new BUILTIN_TIN_NUMBERS block
        lines = ["BUILTIN_TIN_NUMBERS = {"]
        for cby in sorted(merged.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            entry = merged[cby]
            name = entry[0] if entry else ""
            tin = entry[1] if len(entry) > 1 else ""
            sproc = entry[2] if len(entry) > 2 else ""
            safe_name = name.replace('\\', '\\\\').replace('"', '\\"')
            lines.append(f'    "{cby}": ("{safe_name}", "{tin}", "{sproc}"),')
        lines.append("}")
        new_block = "\n".join(lines)

        new_content = _re.sub(pattern, lambda m: new_block, content, count=1, flags=_re.DOTALL)

        if new_content == content:
            messagebox.showwarning("No Change", "No changes to save.")
            return

        try:
            script_path.write_text(new_content, encoding="utf-8")
            # Update the in-memory dict
            BUILTIN_TIN_NUMBERS.clear()
            BUILTIN_TIN_NUMBERS.update(merged)
            # Clear deletions and refresh the tree
            self._deleted_cby_set = set()
            self._dirty = False
            self._search_var.set("")  # clear filter to show full list
            self._populate_tree()
            self._status.configure(text=f"Saved {len(merged)} entries (merged with disk)")
            messagebox.showinfo("Saved", f"Saved {len(merged)} TIN entries to the script file.\n\n"
                                          "Changes merged with any updates from other users\n"
                                          "and will sync via Dropbox.")
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save:\n\n{e}")

    def _on_close(self):
        if self._dirty:
            answer = messagebox.askyesnocancel("Unsaved Changes",
                                                "You have unsaved changes.\n\n"
                                                "Click Yes to save and close.\n"
                                                "Click No to discard and close.\n"
                                                "Click Cancel to go back.")
            if answer is None:
                return  # cancel - don't close
            if answer:
                self._save_to_script()
        self.launcher._tin_win = None
        self.win.destroy()


# ==============================================================================
# UPLOAD DECLARATIONS TO COLS WINDOW
# ==============================================================================
class AutoPilotProgressBar:
    """A small always-on-top floating progress window for auto-pilot mode.
    Has a title bar with close/minimize, is resizable, and matches the
    console UI. If closed or minimized, it pops back up when the session
    completes. Persists on screen until manually closed after completion.
    Cross-platform (Win/Mac)."""
    def __init__(self, parent_win=None):
        self._parent = parent_win
        self._win = ctk.CTkToplevel()
        self._win.title("Auto-Pilot Progress")
        self._win.configure(fg_color="#0f0f1a")
        # Keep on top but allow normal window controls
        try:
            self._win.attributes("-topmost", True)
        except Exception:
            pass

        self._error_flash = False
        self._issue_count = 0
        self._total = 0
        self._current = 0
        self._phase = ""
        self._completed = False  # track if session is done
        self._user_closed = False  # track if user manually closed

        # Size: matches console proportions, resizable
        self._w = 340
        self._h = 120
        self._min_w = 280
        self._min_h = 100
        sw = self._win.winfo_screenwidth()
        sh = self._win.winfo_screenheight()
        x = sw - self._w - 30
        y = sh - self._h - 80
        self._win.geometry(f"{self._w}x{self._h}+{x}+{y}")
        self._win.minsize(self._min_w, self._min_h)
        self._win.resizable(True, True)

        # Handle X button — mark as user-closed, don't destroy yet
        # (we'll destroy in close() or reopen on complete)
        self._win.protocol("WM_DELETE_WINDOW", self._on_user_close)

        # --- Phase label ---
        self._phase_label = ctk.CTkLabel(
            self._win, text="Uploading House Declarations...",  # default phase
            font=(MODERN_FONT, 13, "bold"), text_color="#e8e8e8",
            anchor="w")
        self._phase_label.pack(fill="x", padx=12, pady=(8, 2))

        # --- Progress bar (tkinter Canvas for custom colors) ---
        self._bar_canvas = tk.Canvas(
            self._win, height=16,
            bg="#0f0f1a", highlightthickness=1,
            highlightbackground="#333333")
        self._bar_canvas.pack(fill="x", padx=12, pady=(0, 2))

        # --- Detail label ---
        self._detail_label = ctk.CTkLabel(
            self._win, text="",
            font=(MODERN_FONT, 10), text_color="#888888",
            anchor="w")
        self._detail_label.pack(fill="x", padx=12, pady=(0, 2))

        # --- Activity log (real-time status from console) ---
        self._activity_label = ctk.CTkLabel(
            self._win, text="",
            font=(MODERN_FONT, 9), text_color="#666666",
            anchor="w", wraplength=300)
        self._activity_label.pack(fill="x", padx=12, pady=(0, 6))

        self._win.update_idletasks()

        # Flash animation state
        self._flash_on = False
        self._flash_after_id = None

    def _on_user_close(self):
        """User clicked X. If session not complete, just withdraw (hide).
        If complete, actually destroy."""
        if self._completed:
            self._user_closed = True
            self._win.destroy()
        else:
            # Session still running — hide but keep alive
            # It will pop back up when complete() is called
            self._win.withdraw()

    def set_phase(self, phase, detail=""):
        """Update the phase label and detail text."""
        self._phase = phase
        self._phase_label.configure(text=phase)
        if detail:
            self._detail_label.configure(text=detail)

    def set_activity(self, text):
        """Update the real-time activity line (mirrors console status bar)."""
        # Truncate if too long to fit
        if len(text) > 80:
            text = text[:77] + "..."
        self._activity_label.configure(text=text)

    def set_progress(self, current, total):
        """Update the progress bar. current/total determines fill."""
        self._current = current
        self._total = total
        self._draw_bar()
        if total > 0:
            pct = int(current / total * 100)
            self._detail_label.configure(
                text=f"{current} / {total}  ({pct}%)  |  Issues: {self._issue_count}")

    def add_issue(self):
        """Increment the issue count and flash red."""
        self._issue_count += 1
        self._error_flash = True
        self._start_flash()
        self._detail_label.configure(
            text=f"{self._current} / {self._total}  |  Issues: {self._issue_count}")

    def _start_flash(self):
        """Flash the bar red briefly on error."""
        if self._flash_after_id:
            try:
                self._win.after_cancel(self._flash_after_id)
            except Exception:
                pass
        self._flash_on = True
        self._draw_bar()
        self._flash_after_id = self._win.after(600, self._stop_flash)

    def _stop_flash(self):
        self._flash_on = False
        self._flash_after_id = None
        self._draw_bar()

    def _draw_bar(self):
        """Draw the progress bar on the canvas."""
        self._bar_canvas.delete("all")
        w = self._bar_canvas.winfo_width()
        h = self._bar_canvas.winfo_height()
        if w <= 1:
            w = self._w - 32
        if h <= 1:
            h = 16

        # Background (empty bar)
        self._bar_canvas.create_rectangle(0, 0, w, h, fill="#1a1a2e", outline="")

        # Fill
        if self._total > 0:
            fill_w = int(w * (self._current / self._total))
        else:
            fill_w = 0

        if self._flash_on:
            color = "#b71c1c"  # red flash
        elif "declaration session complete" in self._phase.lower():
            color = "#1b5e20"  # green - complete
        elif "attach" in self._phase.lower():
            color = "#e8820c"  # orange - attaching docs
        elif "upload" in self._phase.lower() or "retry" in self._phase.lower():
            color = "#1a237e"  # blue - uploading
        else:
            color = "#e8820c"  # default orange

        if fill_w > 0:
            self._bar_canvas.create_rectangle(0, 0, fill_w, h, fill=color, outline="")

    def _play_done_sound(self):
        """Play a notification sound when session is complete.
        Cross-platform: uses system bell + platform-specific sound."""
        try:
            self._win.bell()
        except Exception:
            pass
        try:
            if sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["afplay", "/System/Library/Sounds/Glass.aiff"],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            elif sys.platform == "win32":
                import winsound
                winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        except Exception:
            pass

    def complete(self, uploaded, errors, missed, docs_attached, docs_issues):
        """Show the final complete state with green bar and summary.
        If the window was minimized or closed by the user, pop it back up."""
        self._completed = True
        total_issues = errors + missed + docs_issues
        self._phase = "Declaration Session Complete"
        self._phase_label.configure(
            text="Declaration Session Complete",
            text_color="#4caf50")
        self._current = self._total if self._total > 0 else 1
        self._error_flash = False
        self._flash_on = False
        self._draw_bar()
        detail = (f"Uploaded: {uploaded}  |  Docs: {docs_attached}  |  "
                  f"Issues: {total_issues}")
        if total_issues > 0:
            detail += f"  ({errors} errors, {missed} missed, {docs_issues} docs)"
            self._detail_label.configure(text=detail, text_color="#ff6b6b")
        else:
            self._detail_label.configure(text=detail, text_color="#4caf50")
        self._activity_label.configure(text="Declarations complete. Review COLS and submit when ready.")

        # If user minimized or closed the window, pop it back up
        if not self._user_closed:
            try:
                self._win.deiconify()  # un-minimize / un-hide
                self._win.lift()
                try:
                    self._win.attributes("-topmost", True)
                except Exception:
                    pass
                # Flash the title bar to get attention (Windows)
                if sys.platform == "win32":
                    try:
                        self._win.attributes("-topmost", False)
                        self._win.after(100, lambda: self._win.attributes("-topmost", True))
                    except Exception:
                        pass
            except Exception:
                pass

        # Play notification sound
        self._play_done_sound()

    def close(self):
        """Destroy the floating bar. Called by the main script on shutdown."""
        if self._flash_after_id:
            try:
                self._win.after_cancel(self._flash_after_id)
            except Exception:
                pass
        try:
            self._win.destroy()
        except Exception:
            pass

    def is_alive(self):
        """Check if the window still exists and hasn't been user-closed."""
        if self._user_closed:
            return False
        try:
            self._win.winfo_exists()
            return True
        except Exception:
            return False

class UploadWindow(SupportMixin):
    """Window for uploading XML declaration files to the COLS website via Selenium."""

    UPLOAD_URL = "https://online.gov.ky/cols/faces/uploaddeclaration"
    SHORT_DELAY = 1.0
    UPLOAD_DELAY = 2.0
    WAIT_TIMEOUT = 15

    # Status colors for treeview rows
    STATUS_COLORS = {
        "pending":   ("#1a1a2e", "#e8e8e8"),    # dark bg, light text
        "uploading": ("#1a237e", "#ffffff"),    # blue
        "uploaded":  ("#1b5e20", "#ffffff"),    # green
        "error":     ("#b71c1c", "#ffffff"),    # red - failed validation
        "missed":    ("#6a1b9a", "#ffffff"),    # purple - skipped/missed by worker
        "skipped":   ("#424242", "#999999"),    # grey
        "docs_attach": ("#1a237e", "#ffffff"),   # blue - docs being attached
        "docs_ok":     ("#1b5e20", "#ffffff"),   # green - docs attached successfully
        "docs_manual": ("#b71c1c", "#ffffff"),   # red - docs need manual attention
    }

    def __init__(self, launcher):
        self.launcher = launcher

        # Support-mixin state
        self._pending_update = None
        self._support_tooltip = "Report a Bug"
        self._tooltip_win = None
        self._support_bg = "#0f0f1a"
        self._window_name = "Upload Declarations"
        self._set_support_palette({
            "bg": "#0f0f1a", "panel": "#1a1520", "input": "#0a0a12",
            "border": "#e8820c", "text": "#e8e8e8", "light": "#a0a0a0",
            "accent": "#e8820c", "accent_hover": "#f0a030",
        })

        self.win = ctk.CTkToplevel()
        self.win.title(f"Upload Declarations to COLS v{APP_VERSION}")
        self.win.configure(fg_color="#0f0f1a")
        _register_window_name(self.win, self._window_name, {"bg": "#0f0f1a", "panel": "#1a1520", "accent": "#6c3483", "accent_hover": "#8e44ad", "text": "#e8e8e8"})
        w, h = 820, 610
        sw, sh = self.win.winfo_screenwidth(), self.win.winfo_screenheight()
        self.win.geometry(f"{w}x{h}+{int((sw-w)/2)}+{int((sh-h)/2)}")

        self._xml_folder = None
        self._xml_files = []          # list of Path objects
        self._master_files = []       # HBL-Master*.xml files
        self._house_files = []        # HBL-CBY*.xml files
        self._driver = None           # Selenium driver
        self._upload_thread = None
        self._paused = False
        self._stop_requested = False
        self._uploading = False
        self._decl_numbers = {}       # filename -> declaration number
        self._master_decl = ""        # declaration number assigned to master
        self._edge_launched = False   # True after Edge is open and user is logged in
        self._edge_ready = False      # True after Edge browser is open

        # --- Title bar ---
        title_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        title_frame.pack(fill="x", padx=16, pady=(12, 8))
        ctk.CTkButton(title_frame, text="\u2190 Back to Menu",
                      command=self._on_close,
                      fg_color="#0f3460", hover_color="#1a4a7a",
                      text_color="#e8e8e8",
                      font=(MODERN_FONT, 12, "bold"), width=120, height=30,
                      corner_radius=6).pack(side="left", padx=(0, 10))
        ctk.CTkLabel(title_frame, text="Upload Declarations to COLS",
                     font=(MODERN_FONT, 16, "bold"), text_color="#e8e8e8").pack(side="left")

        # Logo (right)
        try:
            _img = Image.open(io.BytesIO(base64.b64decode(MBE_LOGO_B64)))
            _lw, _lh = _img.size
            _tw = 100
            _th = max(1, int(_lh * _tw / _lw))
            _logo = ctk.CTkImage(light_image=_img, dark_image=_img, size=(_tw, _th))
            ctk.CTkLabel(title_frame, image=_logo, text="").pack(side="right")
        except Exception:
            pass

        # --- Folder picker row ---
        folder_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        folder_frame.pack(fill="x", padx=16, pady=(0, 8))
        self._folder_btn = ctk.CTkButton(folder_frame, text="Choose XML Folder",
                                         command=self._pick_folder,
                                         fg_color="#e8820c", hover_color="#ffa726",
                                         text_color="#e8e8e8", font=(MODERN_FONT, 12, "bold"),
                                         width=150, height=30, corner_radius=6)
        self._folder_btn.pack(side="left", padx=(0, 8))
        self._folder_label = ctk.CTkLabel(folder_frame, text="No folder selected",
                                          font=(MODERN_FONT, 11), text_color="#888888")
        self._folder_label.pack(side="left")

        # --- Automation checkboxes ---
        auto_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        auto_frame.pack(fill="x", padx=16, pady=(0, 6))

        self._auto_retry_var = ctk.IntVar(value=0)
        self._auto_retry_cb = ctk.CTkCheckBox(auto_frame, text="Auto-retry missed",
                                              variable=self._auto_retry_var,
                                              font=(MODERN_FONT, 11), text_color="#e8e8e8",
                                              fg_color="#e8820c", hover_color="#ffa726",
                                              height=20)
        self._auto_retry_cb.pack(side="left", padx=(0, 20))

        self._auto_attach_var = ctk.IntVar(value=0)
        self._auto_attach_cb = ctk.CTkCheckBox(auto_frame, text="Auto-attach supporting docs",
                                               variable=self._auto_attach_var,
                                               font=(MODERN_FONT, 11), text_color="#e8e8e8",
                                               fg_color="#e8820c", hover_color="#ffa726",
                                               height=20)
        self._auto_attach_cb.pack(side="left")
        self._autopilot_bar = None  # floating progress bar for auto-pilot mode

        # --- Treeview ---
        tree_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        tree_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        tree_style = ttk.Style()
        tree_style.theme_use("clam")
        # Configure both the named style AND the base style to ensure dark theme
        tree_style.configure("Upload.Treeview", background="#1a1a2e", foreground="#e8e8e8",
                             fieldbackground="#1a1a2e", borderwidth=0, rowheight=26,
                             font=(MODERN_FONT, 10))
        tree_style.configure("Upload.Treeview.Heading", background="#0f3460",
                             foreground="#e8e8e8", font=(MODERN_FONT, 10, "bold"), relief="flat")
        tree_style.map("Upload.Treeview",
                       background=[("selected", "#e8820c")],
                       foreground=[("selected", "#ffffff")])
        # Also set the base Treeview style in case it hasn't been set yet
        tree_style.configure("Treeview", background="#1a1a2e", foreground="#e8e8e8",
                             fieldbackground="#1a1a2e", bordercolor="#333",
                             rowheight=26, font=(MODERN_FONT, 10))
        tree_style.configure("Treeview.Heading", background="#0f1117", foreground="#ffa726",
                             font=(MODERN_FONT, 10, "bold"), relief="flat")
        tree_style.map("Treeview",
                       background=[("selected", "#e8820c")],
                       foreground=[("selected", "#ffffff")])

        tree_cols = ("file", "cby", "status", "docs", "decl")
        self._tree = ttk.Treeview(tree_frame, columns=tree_cols, show="headings",
                                  style="Upload.Treeview", selectmode="extended")
        self._tree.heading("file", text="XML File", anchor="w")
        self._tree.heading("cby", text="CBY", anchor="w")
        self._tree.heading("status", text="Status", anchor="w")
        self._tree.heading("docs", text="Docs", anchor="w")
        self._tree.heading("decl", text="Declaration #", anchor="w")
        self._tree.column("file", width=250, minwidth=180, anchor="w")
        self._tree.column("cby", width=70, minwidth=50, anchor="w")
        self._tree.column("status", width=100, minwidth=70, anchor="w")
        self._tree.column("docs", width=110, minwidth=80, anchor="w")
        self._tree.column("decl", width=120, minwidth=90, anchor="w")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Allow double-click to manually edit declaration number
        self._tree.bind("<Double-1>", self._on_tree_double_click)
        self._tree.bind("<Delete>", lambda e: self._remove_selected())
        self._decl_edit_win = None

        # Tag configurations for row colors
        for status, (bg, fg) in self.STATUS_COLORS.items():
            self._tree.tag_configure(status, background=bg, foreground=fg)

        # --- Control buttons ---
        ctrl_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        ctrl_frame.pack(fill="x", padx=16, pady=(0, 12))

        # Start button lives in a fixed-width slot so other buttons
        # never shift. The button itself is packed/unpacked instantly
        # when XML files are detected or the tree becomes empty.
        self._start_slot = ctk.CTkFrame(ctrl_frame, fg_color="transparent",
                                         width=150, height=30)
        self._start_slot.pack(side="left", padx=(0, 6))
        self._start_slot.pack_propagate(False)
        self._start_btn = ctk.CTkButton(self._start_slot, text="Click here to Begin",
                                        command=self._start_upload,
                                        fg_color="#229954", hover_color="#1e8449",
                                        text_color="#ffffff", font=(MODERN_FONT, 12, "bold"),
                                        width=150, height=30, corner_radius=6)
        # NOT packed initially — appears instantly when XML files are loaded
        self._start_btn_visible = False

        self._pause_btn = ctk.CTkButton(ctrl_frame, text="Pause",
                                        command=self._pause_upload,
                                        fg_color="#c0392b", hover_color="#e74c3c",
                                        text_color="#e8e8e8", font=(MODERN_FONT, 12, "bold"),
                                        width=80, height=30, corner_radius=6,
                                        state="disabled")
        self._pause_btn.pack(side="left", padx=(0, 6))

        self._attach_btn = ctk.CTkButton(ctrl_frame, text="Attach Supporting Docs",
                                         command=self._attach_supporting_docs,
                                         fg_color="#117a65", hover_color="#138d75",
                                         text_color="#e8e8e8", font=(MODERN_FONT, 12, "bold"),
                                         width=160, height=30, corner_radius=6,
                                         state="disabled")
        self._attach_btn.pack(side="left", padx=(0, 6))

        self._copy_decl_btn = ctk.CTkButton(ctrl_frame, text="Copy Decl #s from COLS",
                                            command=self._copy_decl_from_cols,
                                            fg_color="#1a5276", hover_color="#2e86c1",
                                            text_color="#e8e8e8", font=(MODERN_FONT, 12, "bold"),
                                            width=150, height=30, corner_radius=6,
                                            state="disabled")
        self._copy_decl_btn.pack(side="right", padx=(6, 0))

        self._paste_btn = ctk.CTkButton(ctrl_frame, text="Paste Declarations to Manifest",
                                        command=self._quick_paste,
                                        fg_color="#6c3483", hover_color="#8e44ad",
                                        text_color="#e8e8e8", font=(MODERN_FONT, 12, "bold"),
                                        width=175, height=30, corner_radius=6,
                                        state="disabled")
        self._paste_btn.pack(side="right", padx=(6, 0))

        # --- Advanced toggle (collapsible) ---
        self._advanced_visible = False
        self._advanced_btn = ctk.CTkButton(self.win, text="Advanced ▶",
                                           command=self._toggle_advanced,
                                           fg_color="transparent", hover_color="#1a1a2e",
                                           text_color="#888888", font=(MODERN_FONT, 10),
                                           width=80, height=20, anchor="w")
        self._advanced_btn.pack(fill="x", padx=16, pady=(2, 0))

        self._advanced_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        # Not packed initially - hidden until toggled

        self._retry_btn = ctk.CTkButton(self._advanced_frame, text="Retry Missed",
                                        command=self._retry_missed,
                                        fg_color="#6a1b9a", hover_color="#8e24aa",
                                        text_color="#e8e8e8", font=(MODERN_FONT, 11, "bold"),
                                        width=110, height=28, corner_radius=6,
                                        state="disabled")
        self._retry_btn.pack(side="left", padx=(0, 6))

        self._skip_btn = ctk.CTkButton(self._advanced_frame, text="Skip Selected",
                                       command=self._skip_selected,
                                       fg_color="#555555", hover_color="#666666",
                                       text_color="#e8e8e8", font=(MODERN_FONT, 11, "bold"),
                                       width=100, height=28, corner_radius=6,
                                       state="disabled")
        self._skip_btn.pack(side="left", padx=(0, 6))

        self._remove_btn = ctk.CTkButton(self._advanced_frame, text="Remove Selected",
                                         command=self._remove_selected,
                                         fg_color="#7b241c", hover_color="#922b21",
                                         text_color="#e8e8e8", font=(MODERN_FONT, 11, "bold"),
                                         width=120, height=28, corner_radius=6,
                                         state="disabled")
        self._remove_btn.pack(side="left", padx=(0, 6))

        self._unskip_btn = ctk.CTkButton(self._advanced_frame, text="Unskip Selected",
                                         command=self._unskip_selected,
                                         fg_color="#2c3e50", hover_color="#34495e",
                                         text_color="#e8e8e8", font=(MODERN_FONT, 11, "bold"),
                                         width=110, height=28, corner_radius=6,
                                         state="disabled")
        self._unskip_btn.pack(side="left", padx=(0, 6))

        # Logging checkbox (ticked by default for now — will be unchecked
        # by default in a future version once everything is proven stable)
        self._logging_var = ctk.IntVar(value=1)
        self._logging_cb = ctk.CTkCheckBox(self._advanced_frame, text="Enable logging",
                                            variable=self._logging_var,
                                            font=(MODERN_FONT, 10), text_color="#888888",
                                            fg_color="#e8820c", hover_color="#ffa726",
                                            height=18)
        self._logging_cb.pack(side="left", padx=(16, 0))

        # "Master Already Uploaded" button — for when you did the master
        # in a previous session and just need to upload house files now.
        self._master_done_btn = ctk.CTkButton(
            self._advanced_frame, text="Master Already Uploaded",
            command=self._mark_master_done,
            fg_color="#1a5276", hover_color="#2471a3",
            text_color="#e8e8e8", font=(MODERN_FONT, 11, "bold"),
            width=160, height=28, corner_radius=6)
        self._master_done_btn.pack(side="left", padx=(16, 0))

        # --- Status bar (with support icon in the corner) ---
        status_frame = ctk.CTkFrame(self.win, fg_color="transparent")
        status_frame.pack(side="bottom", fill="x", padx=16, pady=(4, 10))
        self._status = ctk.CTkLabel(status_frame, text="Ready. Choose a folder to begin.",
                                    font=(MODERN_FONT, 11), text_color="#888888")
        self._status.pack(side="left", fill="x", expand=True)
        self._attach_support_icon(status_frame)

        self.win.protocol("WM_DELETE_WINDOW", self._on_close)

        # Restore session if we're restarting after a live update
        restored = self._restore_session()
        if restored:
            try:
                self._status.configure(
                    text="Restored from previous session. Ready to paste or upload.")
            except Exception:
                pass

        # Check for updates in the background
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    def _mark_master_done(self):
        """Mark master files as uploaded and ask for the declaration number.
        Used when the master was uploaded in a previous session and you
        only need to upload house files now."""
        # Mark all master files as uploaded in the tree
        marked = 0
        for item in self._tree.get_children():
            filename = self._tree.set(item, "file")
            status = self._tree.set(item, "status").lower()
            if ("HBL-Master" in filename or "MBL" in filename) and status not in ("uploaded", "error"):
                self._tree.set(item, "status", "Uploaded")
                self._tree.item(item, tags=("uploaded",))
                marked += 1

        if marked == 0 and not self._master_files:
            messagebox.showinfo("No Master Files",
                                "No master files found in the list.")
            return

        # Ask for the master declaration number
        entered_decl = [None]
        def ask_for_decl():
            dialog = ctk.CTkToplevel(self.win)
            dialog.title("Master Declaration Number")
            dialog.configure(fg_color="#1a1a2e")
            dialog.geometry("400x220")
            dialog.resizable(False, False)
            dialog.transient(self.win)
            dialog.grab_set()

            dialog.update_idletasks()
            px = self.win.winfo_rootx() + (self.win.winfo_width() - 400) // 2
            py = self.win.winfo_rooty() + (self.win.winfo_height() - 220) // 2
            dialog.geometry(f"+{px}+{py}")

            ctk.CTkLabel(dialog, text="Enter Master Declaration Number",
                         font=(MODERN_FONT, 14, "bold"), text_color="#e8e8e8").pack(pady=(24, 4))
            ctk.CTkLabel(dialog,
                         text="Enter the declaration number that was assigned\n"
                              "to the master when you uploaded it previously.",
                         font=(MODERN_FONT, 11), text_color="#888888",
                         justify="center").pack(pady=(0, 12))

            entry = ctk.CTkEntry(dialog, width=240, height=34,
                                 fg_color="#0f0f1a", border_color="#e8820c",
                                 border_width=2, corner_radius=6, text_color="#e8e8e8",
                                 font=(MODERN_FONT, 14))
            entry.pack(pady=(0, 12))
            entry.focus_set()

            def on_ok():
                entered_decl[0] = entry.get().strip()
                dialog.destroy()

            def on_cancel():
                dialog.destroy()

            ctk.CTkButton(dialog, text="OK", command=on_ok,
                         fg_color="#e8820c", hover_color="#ffa726",
                         width=100, height=30, corner_radius=6,
                         font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=(80, 8), pady=(0, 16))
            ctk.CTkButton(dialog, text="Cancel", command=on_cancel,
                         fg_color="#555555", hover_color="#666666",
                         width=100, height=30, corner_radius=6,
                         font=(MODERN_FONT, 12, "bold")).pack(side="left", pady=(0, 16))
            dialog.bind("<Return>", lambda e: on_ok())
            dialog.bind("<Escape>", lambda e: on_cancel())

        self.win.after(0, ask_for_decl)
        # Wait for the dialog to close (in the main thread)
        # We use after to check periodically
        def check_result():
            if entered_decl[0] is not None:
                if entered_decl[0]:
                    self._master_decl = entered_decl[0]
                    self._status.configure(
                        text=f"Master marked as uploaded. Declaration #{entered_decl[0]}. "
                             f"Ready to upload house files.")
                else:
                    self._status.configure(
                        text="Master marked as uploaded. No declaration number entered.")
            else:
                self.win.after(100, check_result)
        self.win.after(100, check_result)

    def _toggle_advanced(self):
        """Show/hide the advanced controls."""
        if self._advanced_visible:
            self._advanced_frame.pack_forget()
            self._advanced_btn.configure(text="Advanced ▶")
            self._advanced_visible = False
        else:
            self._advanced_frame.pack(fill="x", padx=16, pady=(2, 4),
                                      after=self._advanced_btn)
            self._advanced_btn.configure(text="Advanced ▼")
            self._advanced_visible = True

    def _log(self, message):
        """Write to the attach docs log if logging is enabled."""
        if not (getattr(self, '_logging_var', None) and self._logging_var.get()):
            return
        try:
            log_path = Path.home() / "Documents" / "attach_docs_log.txt"
            with open(log_path, "a", encoding="utf-8") as logf:
                logf.write(message)
        except Exception:
            pass

    def _upload_log(self, message):
        """Write to the XML upload log if logging is enabled. Kept in a
        separate file (upload_log.txt) so it can be diffed against the
        attach docs log to confirm every uploaded CBY also got docs."""
        if not (getattr(self, '_logging_var', None) and self._logging_var.get()):
            return
        try:
            log_path = Path.home() / "Documents" / "upload_log.txt"
            with open(log_path, "a", encoding="utf-8") as logf:
                logf.write(message)
                logf.flush()
        except Exception:
            pass

    def _skip_selected(self):
        """Mark selected rows as 'skipped' so they won't be uploaded."""
        selected = self._tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows first by clicking on them.")
            return
        for item in selected:
            filename = self._tree.set(item, "file")
            self._tree.set(item, "status", "Skipped")
            self._tree.item(item, tags=("skipped",))
            self._xml_files = [f for f in self._xml_files if f.name != filename]
            self._master_files = [f for f in self._master_files if f.name != filename]
            self._house_files = [f for f in self._house_files if f.name != filename]
        self._status.configure(text=f"Skipped {len(selected)} file(s). They will not be uploaded.")

    def _remove_selected(self):
        """Remove selected rows from the treeview entirely."""
        selected = self._tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows first by clicking on them.")
            return
        count = len(selected)
        answer = messagebox.askyesno("Remove Files",
                                     f"Remove {count} file(s) from the list?\n\n"
                                     f"This only removes them from the uploader window.\n"
                                     f"The actual XML files on disk are not deleted.")
        if not answer:
            return
        for item in selected:
            filename = self._tree.set(item, "file")
            self._tree.delete(item)
            self._xml_files = [f for f in self._xml_files if f.name != filename]
            self._master_files = [f for f in self._master_files if f.name != filename]
            self._house_files = [f for f in self._house_files if f.name != filename]
            if filename in self._decl_numbers:
                del self._decl_numbers[filename]
        self._status.configure(text=f"Removed {count} file(s) from the list.")

    def _unskip_selected(self):
        """Reset selected rows back to 'pending' so they will be uploaded."""
        selected = self._tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows first by clicking on them.")
            return
        restored = 0
        for item in selected:
            status = self._tree.set(item, "status").lower()
            if status == "skipped":
                filename = self._tree.set(item, "file")
                self._tree.set(item, "status", "Pending")
                self._tree.item(item, tags=("pending",))
                if self._xml_folder:
                    file_path = self._xml_folder / filename
                    if file_path.exists():
                        self._xml_files.append(file_path)
                        if "HBL-Master" in filename or "MBL" in filename:
                            self._master_files.append(file_path)
                        else:
                            self._house_files.append(file_path)
                        self._xml_files = natsorted(self._xml_files)
                        self._master_files = natsorted(self._master_files)
                        self._house_files = natsorted(self._house_files)
                        restored += 1
        self._status.configure(text=f"Unskipped {restored} file(s). They will be uploaded.")

    def _show_start_btn(self):
        """Show the Start button instantly."""
        if not self._start_btn_visible:
            self._start_btn.pack(fill="both", expand=True)
            self._start_btn_visible = True
        self._start_btn.configure(state="normal")

    def _hide_start_btn(self):
        """Hide the Start button instantly."""
        if self._start_btn_visible:
            self._start_btn.pack_forget()
            self._start_btn_visible = False

    def _pick_folder(self):
        """Open folder picker and load XML files."""
        folder = filedialog.askdirectory(title="Choose folder containing XML files")
        if not folder:
            return
        self._xml_folder = Path(folder)
        self._folder_label.configure(text=str(self._xml_folder))
        self._load_xml_files()

    def _load_xml_files(self):
        """Scan folder for XML files and populate treeview.
        Master files (HBL-Master*) are listed first, then house files (HBL-CBY*)."""
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._xml_files = []
        self._decl_numbers = {}
        self._master_files = []
        self._house_files = []

        if not self._xml_folder or not self._xml_folder.exists():
            return

        xml_files = natsorted(self._xml_folder.glob("*.xml"))
        for xf in xml_files:
            is_master = "HBL-Master" in xf.name or "MBL" in xf.name
            cby = ""
            if is_master:
                cby = "MASTER"
            else:
                m = re.search(r'CBY\s*(\d+)', xf.name)
                if m:
                    cby = m.group(1)
            self._xml_files.append(xf)
            if is_master:
                self._master_files.append(xf)
            else:
                self._house_files.append(xf)
            self._tree.insert("", "end", values=(xf.name, cby, "Pending", "", ""),
                              tags=("pending",))

        master_count = len(self._master_files)
        house_count = len(self._house_files)
        self._status.configure(
            text=f"Loaded {master_count} master + {house_count} house XML file(s)")
        if self._xml_files:
            self._show_start_btn()
            self._paste_btn.configure(state="normal")
            self._skip_btn.configure(state="normal")
            self._remove_btn.configure(state="normal")
            self._unskip_btn.configure(state="normal")
        else:
            self._hide_start_btn()

    def _set_row_docs_status(self, cby, docs_status, docs_text=""):
        """Update the Docs column text for a row matching the given CBY.
        Does NOT change row color — the row keeps its upload status color.
        The docs text itself communicates the state."""
        for item in self._tree.get_children():
            row_cby = self._tree.set(item, "cby")
            if row_cby == cby:
                self._tree.set(item, "docs", docs_text)
                break

    def _update_status(self, text):
        """Update the status bar and mirror to autopilot bar if active."""
        self._status.configure(text=text)
        if self._autopilot_bar and self._autopilot_bar.is_alive():
            self._autopilot_bar.set_activity(text)

    def _set_row_status(self, filename, status, decl_num=""):
        """Update a row's status and color in the treeview."""
        for item in self._tree.get_children():
            vals = self._tree.set(item, "file")
            if vals == filename:
                self._tree.set(item, "status", status.capitalize())
                if decl_num:
                    self._tree.set(item, "decl", decl_num)
                    self._decl_numbers[filename] = decl_num
                self._tree.item(item, tags=(status,))
                break

    def _start_upload(self):
        """Two-phase: first click launches Edge, second click starts uploading."""
        if not _SELENIUM_AVAILABLE:
            messagebox.showerror("Missing Dependency",
                                 "Selenium is not installed.\n\n"
                                 "Install it with:\n  pip install selenium natsort")
            return
        if not self._xml_files:
            messagebox.showwarning("No Files", "Please choose a folder with XML files first.")
            return

        if not self._edge_launched:
            # Phase 1: Launch Edge
            self._launch_edge()
        else:
            # Phase 2: Start uploading
            self._begin_upload()

    def _launch_edge(self):
        """Launch Edge browser and navigate to COLS. Does NOT start uploading."""
        self._edge_ready = False
        self._start_btn.configure(state="disabled", text="Launching...")
        self._status.configure(text="Starting Edge browser... please wait")
        threading.Thread(target=self._edge_launch_worker, daemon=True).start()

    def _edge_launch_worker(self):
        """Background thread: just launches Edge and navigates to COLS."""
        # Animate status while Edge is starting
        dots = ["", ".", "..", "..."]
        dot_idx = [0]
        def update_dots():
            if self._edge_ready:
                return
            self._status.configure(text=f"Starting Edge browser{dots[dot_idx[0] % 4]}... please wait")
            dot_idx[0] += 1
            self.win.after(400, update_dots)
        self.win.after(100, update_dots)

        try:
            options = Options()
            options.add_argument("--inprivate")
            options.add_experimental_option("detach", True)
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-popup-blocking")
            options.add_argument("--no-first-run")
            options.add_argument("--no-default-browser-check")
            # Performance: skip unnecessary loading for faster startup
            options.add_argument("--disable-sync")
            options.add_argument("--disable-background-networking")
            options.add_argument("--disable-default-apps")
            options.add_argument("--disable-notifications")
            options.add_argument("--disable-translate")
            options.add_argument("--disable-background-timer-throttling")
            options.add_argument("--disable-renderer-backgrounding")
            options.add_argument("--disable-backgrounding-occluded-windows")
            options.page_load_strategy = "eager"
            self._driver = webdriver.Edge(options=options)
            self._driver.get(self.UPLOAD_URL)
        except Exception as e:
            self.win.after(0, lambda: self._status.configure(
                text=f"Failed to start Edge: {e}"))
            self.win.after(0, lambda: self._start_btn.configure(
                state="normal", text="Click here to Begin"))
            return

        # Edge is ready - switch button to "Start Upload"
        self._edge_ready = True
        self._edge_launched = True
        self.win.after(0, lambda: self._status.configure(
            text="Edge is open. Log in to COLS, then click 'Start Upload'"))
        self.win.after(0, lambda: self._start_btn.configure(
            state="normal", text="Start Upload"))
        self.win.after(0, lambda: self._copy_decl_btn.configure(state="normal"))
        self.win.after(0, lambda: self._attach_btn.configure(state="normal"))

    def _begin_upload(self):
        """Phase 2: Start the actual file upload process."""
        if self._uploading:
            return
        self._uploading = True
        self._paused = False
        self._stop_requested = False
        self._start_btn.configure(state="disabled", text="Uploading...")
        self._pause_btn.configure(state="normal", text="Pause")
        self._attach_btn.configure(state="disabled")
        self._retry_btn.configure(state="disabled")
        self._status.configure(text="Starting upload...")
        threading.Thread(target=self._upload_worker, daemon=True).start()

    def _upload_worker(self):
        """Background thread: uploads files to COLS.
        Master files are uploaded first. After each master upload, we capture
        the declaration number and prompt the user to continue before uploading
        house files.
        Edge must already be launched and user logged in before calling this."""
        if not self._driver:
            self.win.after(0, lambda: self._status.configure(text="Error: Edge not launched"))
            self.win.after(0, self._reset_buttons)
            self._uploading = False
            return

        # Initialize the upload log (fresh file each run) with the full planned
        # list of CBYs. This is the authoritative list to diff against the
        # attach docs log later.
        if getattr(self, '_logging_var', None) and self._logging_var.get():
            try:
                import datetime
                _ulp = Path.home() / "Documents" / "upload_log.txt"
                with open(_ulp, "w", encoding="utf-8") as _uf:
                    _uf.write(f"upload_log — run started {datetime.datetime.now():%Y-%m-%d %H:%M:%S}\n")
                    _uf.write(f"Master files: {len(self._master_files)}  House files: {len(self._house_files)}\n")
                    _uf.write("\nPlanned house CBYs:\n")
                    for hf in self._house_files:
                        m = re.search(r'CBY\s*(\d+)', hf.name)
                        cby = m.group(1) if m else "?"
                        _uf.write(f"  CBY {cby}  <- {hf.name}\n")
                    _uf.write("\n" + "=" * 60 + "\n")
            except Exception as e:
                print(f"[Upload] Log init error: {e}")

        # ---- PHASE 1: Upload master files (if any) ----
        for master_file in self._master_files:
            if self._stop_requested:
                break
            while self._paused:
                time.sleep(0.5)
                if self._stop_requested:
                    break
            if self._stop_requested:
                break

            # Check if already uploaded
            current_status = ""
            for item in self._tree.get_children():
                if self._tree.set(item, "file") == master_file.name:
                    current_status = self._tree.set(item, "status").lower()
                    break
            if current_status in ("uploaded", "skipped", "error"):
                continue

            self.win.after(0, lambda f=master_file: self._set_row_status(f.name, "uploading"))
            self.win.after(0, lambda f=master_file: self._status.configure(
                text=f"Uploading MASTER: {f.name}..."))

            try:
                wait = WebDriverWait(self._driver, self.WAIT_TIMEOUT)
                file_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
                file_input.send_keys(str(master_file.resolve()))

                # Try multiple button XPath patterns (short timeout per attempt)
                upload_clicked = False
                button_xpaths = [
                    "//button[contains(text(), 'Upload Declaration')]",
                    "//button[contains(text(), 'Upload')]",
                    "//input[@type='submit' and contains(@value, 'Upload')]",
                    "//button[@type='submit']",
                    "//a[contains(text(), 'Upload')]",
                ]
                short_wait = WebDriverWait(self._driver, 3)  # 3s per XPath, not 15s
                for xpath in button_xpaths:
                    try:
                        upload_button = short_wait.until(
                            EC.element_to_be_clickable((By.XPATH, xpath)))
                        upload_button.click()
                        upload_clicked = True
                        break
                    except Exception:
                        continue

                if not upload_clicked:
                    raise Exception("Could not find upload button on page")

                # Smart wait: poll for page to settle instead of fixed sleep.
                # For master uploads, just wait for the page to stop loading.
                # Check if the file input reappears (page ready for next upload)
                # or an error popup appears.
                # Max wait is 15s — generous for slow COLS days.
                for _mw in range(30):  # 30 * 0.5s = 15s max
                    time.sleep(0.5)
                    try:
                        body_lower = self._driver.find_element(
                            By.TAG_NAME, "body").text.lower()
                        if any(err in body_lower for err in [
                            "failed to upload", "an error occurred",
                            "invalid file", "not a valid",
                            "rejected", "please make sure"]):
                            break  # Error appeared, stop waiting
                        # Check if file input reappeared (page ready)
                        inputs = self._driver.find_elements(
                            By.XPATH, "//input[@type='file']")
                        if inputs and inputs[0].is_displayed():
                            break  # Page is ready for next upload
                    except Exception:
                        pass

                # Mark master as uploaded - declaration number entered manually in the gate
                self.win.after(0, lambda f=master_file: self._set_row_status(
                    f.name, "uploaded"))
                self.win.after(0, lambda f=master_file: self._status.configure(
                    text=f"MASTER uploaded: {f.name}. Waiting for declaration number..."))

            except Exception as e:
                self.win.after(0, lambda f=master_file, e=e: self._set_row_status(f.name, "error"))
                self.win.after(0, lambda f=master_file, e=e: self._status.configure(
                    text=f"MASTER MISSED: {f.name} - {type(e).__name__}: {e}"))

            time.sleep(self.SHORT_DELAY)

        # ---- GATE: Ask user for master declaration number ----
        # Skip if already known (e.g., auto-retry of house files only)
        if self._stop_requested:
            self.win.after(0, self._upload_finished)
            return

        if not self._master_decl:
            # Show popup to get master declaration number
            self._master_decl = ""  # clear any stale value
            entered_decl = [None]
            def ask_for_decl():
                dialog = ctk.CTkToplevel(self.win)
                dialog.title("Master Declaration Number Required")
                dialog.configure(fg_color="#1a1a2e")
                dialog.geometry("400x240")
                dialog.resizable(False, False)
                dialog.transient(self.win)
                dialog.grab_set()

                # Center on parent
                dialog.update_idletasks()
                px = self.win.winfo_rootx() + (self.win.winfo_width() - 400) // 2
                py = self.win.winfo_rooty() + (self.win.winfo_height() - 240) // 2
                dialog.geometry(f"+{px}+{py}")

                ctk.CTkLabel(dialog, text="Master Declaration Number Required",
                             font=(MODERN_FONT, 14, "bold"), text_color="#e8e8e8").pack(pady=(24, 4))
                ctk.CTkLabel(dialog,
                             text="Enter the declaration number assigned to the master.\n"
                                  "You can find it on the COLS page after submitting.\n"
                                  "House files will NOT upload until this is entered.",
                             font=(MODERN_FONT, 11), text_color="#888888",
                             justify="center").pack(pady=(0, 12))

                entry = ctk.CTkEntry(dialog, width=240, height=34,
                                     fg_color="#0f0f1a", border_color="#e8820c",
                                     border_width=2, corner_radius=6, text_color="#e8e8e8",
                                     font=(MODERN_FONT, 14))
                entry.pack(pady=(0, 12))
                entry.focus_set()

                def on_ok():
                    entered_decl[0] = entry.get().strip()
                    dialog.destroy()

                def on_cancel():
                    dialog.destroy()

                btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
                btn_frame.pack(pady=(0, 16))
                ctk.CTkButton(btn_frame, text="OK - Start House Uploads", command=on_ok,
                              fg_color="#2e8b57", hover_color="#3cb371", width=170, height=32,
                              corner_radius=6, font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=6)
                ctk.CTkButton(btn_frame, text="Cancel", command=on_cancel,
                              fg_color="#555555", hover_color="#666666", width=80, height=32,
                              corner_radius=6, font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=6)

                entry.bind("<Return>", lambda e: on_ok())
                entry.bind("<Escape>", lambda e: on_cancel())
                dialog.protocol("WM_DELETE_WINDOW", on_cancel)

                self.win.wait_window(dialog)

            # Run the dialog on the main thread, wait for result in background thread
            import threading as _th
            evt = _th.Event()
            def run_dialog():
                ask_for_decl()
                evt.set()
            self.win.after(0, run_dialog)
            evt.wait()

            if entered_decl[0]:
                self._master_decl = entered_decl[0]
                # Update master rows in tree with the entered declaration number
                for master_file in self._master_files:
                    self.win.after(0, lambda f=master_file, d=entered_decl[0]:
                        self._set_row_status(f.name, "uploaded", d))
                self.win.after(0, lambda d=entered_decl[0]: self._status.configure(
                    text=f"Master declaration #: {d}. Starting house uploads..."))
            else:
                # User cancelled - don't proceed
                self.win.after(0, lambda: self._status.configure(
                    text="STOPPED: Master declaration number not entered. House files will not upload."))
                self.win.after(0, self._upload_finished)
                return


        # ---- PHASE 2: Upload house files ----
        # Show floating progress bar if any auto mode is enabled
        if (self._auto_retry_var.get() or self._auto_attach_var.get()) and not self._autopilot_bar:
            self._autopilot_bar = AutoPilotProgressBar(self.win)
        if self._autopilot_bar and self._autopilot_bar.is_alive():
            self._autopilot_bar.set_phase("Uploading House Declarations...")
            self._autopilot_bar.set_progress(0, len(self._house_files))
        for xml_file in self._house_files:
            if self._stop_requested:
                break
            while self._paused:
                time.sleep(0.5)
                if self._stop_requested:
                    break
            if self._stop_requested:
                break

            # Check if already uploaded
            current_status = ""
            for item in self._tree.get_children():
                if self._tree.set(item, "file") == xml_file.name:
                    current_status = self._tree.set(item, "status").lower()
                    break
            if current_status in ("uploaded", "skipped", "error"):
                continue

            # Pre-check: if this file is already on the COLS page (e.g., from
            # a previous session), mark as uploaded and skip. Only trust a FULL
            # filename match — the number-only match can hit the file-picker
            # label of a different file and cause false "uploaded" marks.
            try:
                body_check = self._driver.find_element(By.TAG_NAME, "body").text
                if xml_file.name in body_check:
                    _cby_m = re.search(r'CBY\s*(\d+)', xml_file.name)
                    self._upload_log(
                        f"CBY {_cby_m.group(1) if _cby_m else '?'}: PRE-CHECK skip "
                        f"(matched by filename)  ({xml_file.name})\n")
                    self.win.after(0, lambda fn=xml_file.name: self._set_row_status(fn, "uploaded"))
                    self.win.after(0, lambda fn=xml_file.name: self._status.configure(
                        text=f"Already on page: {fn} (skipped re-upload)"))
                    time.sleep(self.SHORT_DELAY)
                    continue
            except Exception:
                pass

            self.win.after(0, lambda f=xml_file: self._set_row_status(f.name, "uploading"))
            self.win.after(0, lambda f=xml_file: self._update_status(
                f"Uploading: {f.name}..."))

            try:
                wait = WebDriverWait(self._driver, self.WAIT_TIMEOUT)
                file_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))

                # Count "Passed" occurrences BEFORE uploading. Each successfully
                # uploaded declaration row shows "Passed" on the page. After we
                # click upload, we wait for this count to go up by 1 — proving a
                # new row genuinely appeared. This is fast (body.text includes the
                # new row's text quickly) and reliable ("Passed" only appears on
                # genuine uploads, not in the file picker label).
                # We also count "error" — if a row lands with an error status
                # instead of "Passed", the error count goes up. That also means
                # the upload settled (just with an error), so we stop waiting
                # and mark it as an error for human review.
                def _count_word(word):
                    try:
                        body = self._driver.find_element(By.TAG_NAME, "body").text
                        return body.lower().count(word)
                    except Exception:
                        return -1
                passed_before = _count_word("passed")
                error_before = _count_word("error")

                file_input.send_keys(str(xml_file.resolve()))

                # Click the Upload button. The button is always the same on COLS
                # — don't overthink it. Find it and click it immediately, no
                # visibility checks, no multiple XPath guessing. If the first
                # attempt fails, try a real Selenium .click() as fallback.
                # The "Passed" count wait after this handles verification AND
                # tells us when COLS is ready for the next file.
                upload_clicked = False
                time.sleep(0.3)  # brief moment for COLS to register the file
                try:
                    btn = self._driver.find_element(
                        By.XPATH, "//button[contains(text(), 'Upload Declaration')]")
                    self._driver.execute_script("arguments[0].click();", btn)
                    upload_clicked = True
                except Exception:
                    pass
                # Fallback: real Selenium click (fires ADF event handlers)
                if not upload_clicked:
                    try:
                        btn = self._driver.find_element(
                            By.XPATH, "//button[contains(text(), 'Upload Declaration')]")
                        btn.click()
                        upload_clicked = True
                    except Exception:
                        pass
                # Last fallback: try broader XPath patterns
                if not upload_clicked:
                    for xpath in [
                        "//button[contains(text(), 'Upload')]",
                        "//input[@type='submit' and contains(@value, 'Upload')]",
                        "//button[@type='submit']",
                        "//a[contains(text(), 'Upload')]",
                    ]:
                        try:
                            btn = self._driver.find_element(By.XPATH, xpath)
                            self._driver.execute_script("arguments[0].click();", btn)
                            upload_clicked = True
                            break
                        except Exception:
                            continue

                if not upload_clicked:
                    # Last resort: refresh and try once more
                    self.win.after(0, lambda f=xml_file: self._update_status(
                        f"Upload button not found, refreshing page for {f.name}..."))
                    try:
                        self._driver.navigate().refresh()
                        time.sleep(3)
                        retry_wait = WebDriverWait(self._driver, 10)
                        file_input2 = retry_wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
                        file_input2.send_keys(str(xml_file.resolve()))
                        time.sleep(0.3)
                        btn = retry_wait.until(
                            EC.presence_of_element_located(
                                (By.XPATH, "//button[contains(text(), 'Upload Declaration')]")))
                        self._driver.execute_script("arguments[0].click();", btn)
                        upload_clicked = True
                    except Exception:
                        pass
                    except Exception:
                        pass
                    if not upload_clicked:
                        raise Exception("Could not find upload button (even after refresh)")

                # Smart wait: poll for the "Passed" count OR "error" count to
                # increase — either means a new declaration row landed on the
                # page and the upload settled. "Passed" = success, "error" =
                # the row appeared with an error status (human review needed).
                # Either way we stop waiting immediately. Also check for error
                # popups (ADF dialogs with specific rejection messages).
                page_settled = False
                _row_had_error = False
                for _pw in range(60):  # 60 * 0.3s = 18s max
                    time.sleep(0.3)
                    try:
                        body_text = self._driver.find_element(
                            By.TAG_NAME, "body").text
                        body_lower = body_text.lower()
                        # Check for error popup (ADF rejection dialog)
                        if any(err in body_lower for err in [
                            "failed to upload", "an error occurred",
                            "invalid file", "not a valid",
                            "rejected", "please make sure"]):
                            page_settled = True
                            break
                        # Success: "Passed" count increased (new row appeared)
                        if passed_before >= 0 and _count_word("passed") > passed_before:
                            page_settled = True
                            break
                        # Error status: "error" count increased (row landed with
                        # error instead of "Passed" — upload settled, needs review)
                        if error_before >= 0 and _count_word("error") > error_before:
                            page_settled = True
                            _row_had_error = True
                            break
                    except Exception:
                        pass

                # Verify the file actually appeared on the COLS page.
                # Also detect error popups (ADF message dialogs) that appear
                # when COLS rejects the XML — these must be dismissed and
                # marked as errors (not missed) to avoid retry loops.
                upload_verified = False
                had_error_popup = False
                error_message = ""
                self._verify_match_kind = "none"
                try:
                    body_text = self._driver.find_element(
                        By.TAG_NAME, "body").text
                    body_lower = body_text.lower()

                    # Check for ADF error popup — the dialog has id d1_msgDlg_cancel
                    # and shows error text like "Failed to upload XML..."
                    # Also check for validation errors that appear inline.
                    error_phrases = [
                        "failed to upload xml",
                        "failed to upload",
                        "please make sure every start-tag",
                        "an error occurred",
                        "error processing",
                        "invalid file",
                        "only xml",
                        "not a valid",
                        "upload failed",
                        "rejected",
                        "validation error",
                        "must be numeric",
                        "is required",
                        "cannot be null",
                        "duplicate declaration",
                        "already exists",
                    ]
                    for phrase in error_phrases:
                        if phrase in body_lower:
                            had_error_popup = True
                            # Try to extract the actual error message
                            try:
                                # ADF error dialog text is in a TD with class x1e4
                                error_cells = self._driver.find_elements(
                                    By.XPATH, "//td[@class='x1e4']")
                                for cell in error_cells:
                                    cell_text = cell.text.strip()
                                    if cell_text and len(cell_text) > 5:
                                        error_message = cell_text[:200]
                                        break
                            except Exception:
                                pass
                            if not error_message:
                                # Fallback: find the error text in body
                                for line in body_text.split("\n"):
                                    for phrase in error_phrases:
                                        if phrase in line.lower():
                                            error_message = line.strip()[:200]
                                            break
                                    if error_message:
                                        break
                            break

                    # Verify upload: "Passed" count went up (a new declaration
                    # row genuinely appeared). This is the reliable signal —
                    # "Passed" only shows on uploaded rows, never in the picker.
                    if passed_before >= 0 and _count_word("passed") > passed_before:
                        upload_verified = True
                        self._verify_match_kind = "passed-count"

                    # If the row landed with "error" status instead of "Passed",
                    # the upload settled but the declaration has an error. Mark
                    # it as an error (not missed) so it doesn't get retried —
                    # humans need to review and fix it on COLS.
                    if _row_had_error and not upload_verified and not had_error_popup:
                        had_error_popup = True  # reuse the error-handling path
                        error_message = "Declaration uploaded with error status on COLS"
                        self._verify_match_kind = "error-row"

                    # NOTE: Inline error indicator check removed — it was scanning
                    # the ENTIRE page for any "Error" text, which matched errors
                    # from OTHER rows and caused false positives. Now, uncertain
                    # uploads (no explicit popup, no verification) fall through
                    # to "missed" status and get retried automatically.
                except Exception:
                    upload_verified = True  # Give benefit of the doubt on error

                # Dismiss error popup if present (click OK button)
                if had_error_popup:
                    try:
                        ok_btn = self._driver.find_element(
                            By.ID, "d1_msgDlg_cancel")
                        ok_btn.click()
                        time.sleep(0.5)
                    except Exception:
                        try:
                            # Fallback: find any link/button with text "OK"
                            ok_links = self._driver.find_elements(
                                By.XPATH, "//a[contains(text(), 'OK')]")
                            if ok_links:
                                ok_links[0].click()
                                time.sleep(0.5)
                        except Exception:
                            pass

                _cby_m = re.search(r'CBY\s*(\d+)', xml_file.name)
                _cby_log = _cby_m.group(1) if _cby_m else "?"

                if upload_verified and not had_error_popup:
                    self.win.after(0, lambda f=xml_file: self._set_row_status(
                        f.name, "uploaded"))
                    self.win.after(0, lambda f=xml_file: self._update_status(
                        f"Uploaded: {f.name}"))
                    self._upload_log(f"CBY {_cby_log}: UPLOADED "
                                     f"(verified by {getattr(self, '_verify_match_kind', '?')})  "
                                     f"({xml_file.name})\n")
                elif had_error_popup:
                    # Error popup on COLS — mark as error (not retried automatically)
                    self.win.after(0, lambda f=xml_file: self._set_row_status(
                        f.name, "error"))
                    if self._autopilot_bar and self._autopilot_bar.is_alive():
                        self._autopilot_bar.add_issue()
                    self.win.after(0, lambda f=xml_file, msg=error_message: self._update_status(
                        f"ERROR: {f.name} — {msg}" if msg else f"ERROR: {f.name} rejected by COLS"))
                    self._upload_log(f"CBY {_cby_log}: ERROR — {error_message or 'rejected by COLS'}  ({xml_file.name})\n")
                else:
                    self.win.after(0, lambda f=xml_file: self._set_row_status(
                        f.name, "missed"))
                    self.win.after(0, lambda f=xml_file: self._update_status(
                        f"WARNING: {f.name} not confirmed on page — will retry"))
                    if self._autopilot_bar and self._autopilot_bar.is_alive():
                        self._autopilot_bar.add_issue()
                    self._upload_log(f"CBY {_cby_log}: MISSED (not confirmed)  ({xml_file.name})\n")

            except Exception as e:
                self.win.after(0, lambda f=xml_file, e=e: self._set_row_status(f.name, "error"))
                self.win.after(0, lambda f=xml_file, e=e: self._status.configure(
                    text=f"Error: {f.name} - {type(e).__name__}: {e}"))
                _cby_m = re.search(r'CBY\s*(\d+)', xml_file.name)
                self._upload_log(f"CBY {_cby_m.group(1) if _cby_m else '?'}: EXCEPTION — {type(e).__name__}: {e}  ({xml_file.name})\n")

            # Update autopilot progress bar
            if self._autopilot_bar and self._autopilot_bar.is_alive():
                done = 0
                for item in self._tree.get_children():
                    s = self._tree.set(item, "status").lower()
                    if s in ("uploaded", "error", "skipped"):
                        done += 1
                total_files = len(self._house_files)
                self._autopilot_bar.set_progress(done, total_files)
            time.sleep(self.SHORT_DELAY)

        self.win.after(0, self._upload_finished)

    def _on_tree_double_click(self, event):
        """Allow manual editing of the Declaration # column."""
        if self._decl_edit_win is not None:
            try:
                self._decl_edit_win.destroy()
            except Exception:
                pass
            self._decl_edit_win = None

        region = self._tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self._tree.identify_column(event.x)
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            return
        col_idx = int(col.replace("#", "")) - 1
        cols = ["file", "cby", "status", "docs", "decl"]
        if col_idx < 0 or col_idx >= len(cols):
            return
        key = cols[col_idx]
        if key != "decl":
            return  # Only allow editing the Declaration # column

        bbox = self._tree.bbox(row_id, col)
        if not bbox:
            return
        x, y, w, h = bbox
        cx = self._tree.winfo_rootx() + x
        cy = self._tree.winfo_rooty() + y
        current_val = self._tree.set(row_id, key)

        self._decl_edit_win = tk.Toplevel(self.win)
        self._decl_edit_win.overrideredirect(True)
        self._decl_edit_win.geometry(f"{w}x{h}+{cx}+{cy}")
        self._decl_edit_win.attributes("-topmost", True)

        widget = ctk.CTkEntry(self._decl_edit_win, width=w, height=h,
                              fg_color="#0f0f1a", border_color="#e8820c",
                              border_width=2, corner_radius=2, text_color="#e8e8e8",
                              font=(MODERN_FONT, 10))
        widget.pack(fill="both", expand=True)
        widget.insert(0, current_val)
        widget.focus_set()
        widget.select_range(0, "end")

        def finish(e=None):
            value = widget.get().strip()
            self._tree.set(row_id, key, value)
            filename = self._tree.set(row_id, "file")
            if value:
                self._decl_numbers[filename] = value
            elif filename in self._decl_numbers:
                del self._decl_numbers[filename]
            try:
                self._decl_edit_win.destroy()
            except Exception:
                pass
            self._decl_edit_win = None

        def cancel(e=None):
            try:
                self._decl_edit_win.destroy()
            except Exception:
                pass
            self._decl_edit_win = None

        widget.bind("<Return>", finish)
        widget.bind("<Escape>", cancel)
        widget.bind("<FocusOut>", finish)

    def _extract_decl_number(self):
        """Try to extract the declaration number assigned by COLS after upload.
        Searches both visible text and page source. May need adjustment based on
        the actual COLS page structure."""
        if not self._driver:
            return ""

        # Patterns to try - ordered from most specific to most generic
        patterns = [
            r'Declaration\s*(?:Number|#|No\.?)\s*[:]?\s*(\d{5,})',
            r'Declaration\s*(?:Number|#|No\.?)\s*[:]?\s*([A-Z0-9-]{5,})',
            r'(?:Created|Submitted|Assigned|Generated).*?(\d{5,})',
            r'Reference\s*(?:Number|#)?\s*[:]?\s*(\d{5,})',
            r'(?:Dec|Decl)\.?\s*[:#]?\s*(\d{5,})',
            r'(?:Number|No\.?)\s*[:]?\s*(\d{7,})',
            r'(\d{7,})',  # Any 7+ digit number as last resort
        ]

        # Strategy 1: Search visible body text
        try:
            body_text = self._driver.find_element(By.TAG_NAME, "body").text
            for pattern in patterns:
                m = re.search(pattern, body_text, re.IGNORECASE)
                if m:
                    return m.group(1)
        except Exception:
            pass

        # Strategy 2: Search page source (catches hidden elements, JavaScript-rendered text)
        try:
            page_source = self._driver.page_source
            # In page source, look for patterns in text nodes
            for pattern in patterns:
                m = re.search(pattern, page_source, re.IGNORECASE)
                if m:
                    return m.group(1)
        except Exception:
            pass

        # Strategy 3: Look for any element containing "declaration" text
        try:
            elements = self._driver.find_elements(By.XPATH,
                "//*[contains(text(), 'declaration') or contains(text(), 'Declaration')]")
            for el in elements:
                text = el.text
                for pattern in patterns:
                    m = re.search(pattern, text, re.IGNORECASE)
                    if m:
                        return m.group(1)
        except Exception:
            pass

        return ""

    def _pause_upload(self):
        """Toggle pause/resume."""
        if self._paused:
            self._paused = False
            self._pause_btn.configure(text="Pause", fg_color="#c0392b", hover_color="#e74c3c")
            self._attach_btn.configure(state="disabled")
            self._status.configure(text="Resumed uploading...")
        else:
            self._paused = True
            self._pause_btn.configure(text="Resume", fg_color="#2e8b57", hover_color="#3cb371")
            self._attach_btn.configure(state="normal")
            self._status.configure(text="Paused. Click Resume to continue.")

    def _retry_missed(self):
        """Reset failed/skipped rows to pending and restart upload."""
        if not self._driver or not self._uploading:
            # Need to restart the whole process
            self._start_upload()
            return

        # Reset missed rows to pending (NOT errors - those need manual fixing)
        for item in self._tree.get_children():
            status = self._tree.set(item, "status").lower()
            if status in ("missed",):
                filename = self._tree.set(item, "file")
                self._tree.set(item, "status", "Pending")
                self._tree.item(item, tags=("pending",))

        # Re-run upload for pending files (worker handles master-first automatically)
        self._paused = False
        self._pause_btn.configure(text="Pause", state="normal", fg_color="#c0392b", hover_color="#e74c3c")
        self._stop_requested = False
        if self._autopilot_bar and self._autopilot_bar.is_alive():
            self._autopilot_bar.set_phase("Retrying Missed Declarations...")
        threading.Thread(target=self._upload_worker, daemon=True).start()

    def _upload_finished(self):
        """Called when upload loop completes."""
        self._uploading = False
        if self._edge_launched:
            self._start_btn.configure(state="normal", text="Start Upload")
        else:
            self._start_btn.configure(state="normal", text="Click here to Begin")
        self._pause_btn.configure(state="disabled", text="Pause", fg_color="#c0392b", hover_color="#e74c3c")
        self._attach_btn.configure(state="normal")
        self._retry_btn.configure(state="normal")

        # Count results
        uploaded = errors = missed = pending = 0
        for item in self._tree.get_children():
            s = self._tree.set(item, "status").lower()
            if s == "uploaded":
                uploaded += 1
            elif s == "error":
                errors += 1
            elif s == "missed":
                missed += 1
            elif s == "pending":
                pending += 1

        self._status.configure(
            text=f"Done: {uploaded} uploaded, {errors} errors, {missed} missed, {pending} pending")

        # Auto-retry missed files if checkbox is checked
        # (errors are NOT auto-retried - they need manual fixing)
        if missed > 0 and self._auto_retry_var.get() and not self._stop_requested:
            self._status.configure(
                text=f"Auto-retrying {missed} missed file(s)...")
            if self._autopilot_bar and self._autopilot_bar.is_alive():
                self._autopilot_bar.set_phase("Retrying Missed Declarations...")
                self._autopilot_bar.set_progress(0, missed)
            self.win.after(2000, self._retry_missed)  # Wait 2s then retry
            return

        # Auto-attach supporting docs if checkbox is checked
        # (proceed even if some files marked "error" — they may have uploaded
        #  successfully but been misdetected; genuinely errored files won't
        #  have "Upload Supporting Documents" links on COLS so they're skipped)
        if self._auto_attach_var.get() and uploaded > 0 and not self._stop_requested:
            self._status.configure(
                text="Auto-attaching supporting documents...")
            if self._autopilot_bar and self._autopilot_bar.is_alive():
                self._autopilot_bar.set_phase("Attaching Declaration Supporting Docs...")
                self._autopilot_bar.set_progress(0, uploaded)
            self.win.after(2000, self._auto_attach_supporting_docs)
            return

        # If no auto-attach will run, complete the autopilot bar
        if not (self._auto_attach_var.get() and uploaded > 0 and not self._stop_requested):
            if self._autopilot_bar and self._autopilot_bar.is_alive():
                self._autopilot_bar.complete(uploaded, errors, missed, 0, 0)
        if errors > 0:
            messagebox.showinfo("Upload Complete",
                                f"Uploaded: {uploaded}\nErrors: {errors}\nMissed: {missed}\n\n"
                                f"Errors need to be fixed manually:\n"
                                f"  1. Inspect the file on COLS to see the issue\n"
                                f"  2. Fix the XML file\n"
                                f"  3. Re-upload it manually on COLS\n"
                                f"  4. Double-click the Declaration # column to enter the number\n\n"
                                f"Missed files can be retried via 'Retry Missed' under Advanced.\n"
                                f"Or check 'Auto-retry missed' at the top for automatic retry.")
        else:
            messagebox.showinfo("Upload Complete",
                                f"All {uploaded} file(s) uploaded successfully!\n\n"
                                f"You can now attach supporting docs and submit on COLS.")

    def _reset_buttons(self):
        if self._edge_launched:
            self._start_btn.configure(state="normal", text="Start Upload")
        else:
            self._start_btn.configure(state="normal", text="Click here to Begin")
        self._pause_btn.configure(state="disabled", text="Pause", fg_color="#c0392b", hover_color="#e74c3c")
        self._retry_btn.configure(state="disabled")
        self._attach_btn.configure(state="disabled")

    def _auto_attach_supporting_docs(self):
        """Auto-attach: find PDFs and attach without showing confirmation dialog."""
        self._attach_supporting_docs(skip_dialog=True)

    def _attach_supporting_docs(self, skip_dialog=False):
        """Find PDFs in the manifest folder and attach them to declarations on COLS."""
        if not self._driver:
            messagebox.showwarning("No Browser", "Launch Edge first, then use this button.")
            return
        if not self._xml_folder:
            messagebox.showwarning("No Folder", "Choose an XML folder first.")
            return

        # Find PDFs
        pdfs = self._find_supporting_pdfs()
        if not pdfs:
            messagebox.showwarning("No PDFs Found",
                                   "Could not find any PDF files.\n\n"
                                   "Searched in:\n"
                                   f"  {self._xml_folder}\n"
                                   f"  {self._xml_folder.parent}\n"
                                   "and their subfolders.\n\n"
                                   "Make sure your PDF files are in the manifest\n"
                                   "folder or a subfolder next to the XML files.")
            return

        # Clear the log at the start of each run (only if logging is enabled)
        if getattr(self, '_logging_var', None) and self._logging_var.get():
            _log_path_init = Path.home() / "Documents" / "attach_docs_log.txt"
            with open(_log_path_init, "w", encoding="utf-8") as _lf:
                import datetime
                _lf.write(f"attach_docs_log — run started {datetime.datetime.now():%Y-%m-%d %H:%M:%S}\n")

        # Categorize PDFs: AWB/BOL vs invoices vs placeholder
        awb_bol_pdfs = []
        invoice_pdfs = []
        placeholder_pdf = None
        for pdf in pdfs:
            name = pdf.name.upper()
            # Remove leading # for detection
            clean_name = name.lstrip("#")
            if clean_name.startswith("378") or name.startswith("SMLU") or name.startswith("#SMLU"):
                awb_bol_pdfs.append(pdf)
            elif "MANIFEST" in name and "CUSTOM" in name:
                placeholder_pdf = pdf
            else:
                invoice_pdfs.append(pdf)

        # Match invoice PDFs to CBY numbers from the treeview
        cby_pdfs_map = {}  # cby_number -> list of PDF paths
        all_tree_cbys = []
        for item in self._tree.get_children():
            cby = self._tree.set(item, "cby")
            if not cby or cby == "MASTER":
                continue
            all_tree_cbys.append(cby)
            matched = []
            for pdf in invoice_pdfs:
                stem = pdf.stem  # filename without extension
                if stem.upper().startswith(cby.upper()):
                    # Check next char after CBY is not a digit
                    rest = stem[len(cby):]
                    if not rest or rest[0] in "_-. " or rest[0].isalpha():
                        matched.append(pdf)
            if matched:
                cby_pdfs_map[cby] = matched

        # For CBYs with no invoice, use the Manifest and Customs PDF as placeholder
        placeholder_cbys = []
        if placeholder_pdf:
            for cby in all_tree_cbys:
                if cby not in cby_pdfs_map:
                    cby_pdfs_map[cby] = [placeholder_pdf]
                    placeholder_cbys.append(cby)

        # Show confirmation dialog
        if not cby_pdfs_map and not awb_bol_pdfs:
            messagebox.showwarning("No Matches",
                                   "Found PDFs but could not match them to any CBY numbers.\n\n"
                                   f"PDFs found: {len(pdfs)}\n"
                                   f"CBYs in treeview: {len(all_tree_cbys)}\n\n"
                                   "Make sure PDF filenames start with the CBY number\n"
                                   "(e.g., '200_1.pdf' for CBY 200).")
            return

        if skip_dialog:
            # Auto mode: skip confirmation, go straight to attaching
            self._attach_btn.configure(state="disabled")
            threading.Thread(target=self._attach_docs_worker,
                             args=(cby_pdfs_map, awb_bol_pdfs, placeholder_pdf),
                             daemon=True).start()
        else:
            self._show_attach_confirm_dialog(cby_pdfs_map, awb_bol_pdfs, pdfs,
                                             placeholder_pdf, placeholder_cbys)

    def _show_attach_confirm_dialog(self, cby_pdfs_map, awb_bol_pdfs, all_pdfs,
                                     placeholder_pdf=None, placeholder_cbys=None):
        """Show confirmation dialog with PDFs grouped by CBY before attaching."""
        if placeholder_cbys is None:
            placeholder_cbys = []

        dialog = ctk.CTkToplevel(self.win)
        dialog.title("Attach Supporting Documents")
        dialog.configure(fg_color="#1a1a2e")
        dialog.geometry("550x500")
        dialog.resizable(False, False)
        dialog.transient(self.win)
        dialog.grab_set()

        dialog.update_idletasks()
        px = self.win.winfo_rootx() + (self.win.winfo_width() - 550) // 2
        py = self.win.winfo_rooty() + (self.win.winfo_height() - 500) // 2
        dialog.geometry(f"+{px}+{py}")

        ctk.CTkLabel(dialog, text="Supporting Documents Found",
                     font=(MODERN_FONT, 14, "bold"), text_color="#e8e8e8").pack(pady=(16, 4))
        ctk.CTkLabel(dialog,
                     text=f"Found {len(all_pdfs)} PDF file(s).\n"
                          "Review the matching below, then click Start.",
                     font=(MODERN_FONT, 11), text_color="#888888").pack(pady=(0, 12))

        scroll = ctk.CTkScrollableFrame(dialog, fg_color="transparent", height=320)
        scroll.pack(fill="both", expand=True, padx=20, pady=(0, 12))

        # Show AWB/BOL section
        if awb_bol_pdfs:
            ctk.CTkLabel(scroll, text="AWB / Bill of Lading (attached to all CBYs):",
                         font=(MODERN_FONT, 11, "bold"), text_color="#117a65").pack(anchor="w", pady=(4, 2))
            for pdf in awb_bol_pdfs:
                ctk.CTkLabel(scroll, text=f"  {pdf.name}",
                             font=(MODERN_FONT, 10), text_color="#e8e8e8").pack(anchor="w")
            ctk.CTkLabel(scroll, text="",
                         font=(MODERN_FONT, 8)).pack()

        # Show placeholder info
        if placeholder_pdf and placeholder_cbys:
            ctk.CTkLabel(scroll, text=f"Placeholder for CBYs with no invoice:",
                         font=(MODERN_FONT, 11, "bold"), text_color="#e8820c").pack(anchor="w", pady=(4, 2))
            ctk.CTkLabel(scroll, text=f"  {placeholder_pdf.name}",
                         font=(MODERN_FONT, 10), text_color="#e8e8e8").pack(anchor="w")
            ctk.CTkLabel(scroll, text=f"  Used for: {', '.join(sorted(placeholder_cbys))}",
                         font=(MODERN_FONT, 10), text_color="#888888").pack(anchor="w")
            ctk.CTkLabel(scroll, text="",
                         font=(MODERN_FONT, 8)).pack()

        # Show CBY sections
        if cby_pdfs_map:
            ctk.CTkLabel(scroll, text="Documents by CBY:",
                         font=(MODERN_FONT, 11, "bold"), text_color="#e8820c").pack(anchor="w", pady=(4, 2))
            for cby in sorted(cby_pdfs_map.keys()):
                is_placeholder = cby in placeholder_cbys
                label = f"  CBY {cby}:" + ("  [placeholder]" if is_placeholder else "")
                ctk.CTkLabel(scroll, text=label,
                             font=(MODERN_FONT, 10, "bold"), text_color="#e8e8e8").pack(anchor="w")
                for pdf in cby_pdfs_map[cby]:
                    ctk.CTkLabel(scroll, text=f"    {pdf.name}",
                                 font=(MODERN_FONT, 10), text_color="#888888").pack(anchor="w")

        # Show unmatched CBYs (no invoice AND no placeholder)
        tree_cbys = set()
        for item in self._tree.get_children():
            cby = self._tree.set(item, "cby")
            if cby and cby != "MASTER":
                tree_cbys.add(cby)
        unmatched = tree_cbys - set(cby_pdfs_map.keys())
        if unmatched:
            ctk.CTkLabel(scroll, text="",
                         font=(MODERN_FONT, 8)).pack()
            ctk.CTkLabel(scroll, text=f"CBYs with no PDFs at all: {', '.join(sorted(unmatched))}",
                         font=(MODERN_FONT, 10), text_color="#c0392b").pack(anchor="w")

        result = [None]
        def on_start():
            result[0] = True
            dialog.destroy()
        def on_cancel():
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=(0, 16))
        ctk.CTkButton(btn_frame, text="Start Attaching", command=on_start,
                      fg_color="#117a65", hover_color="#138d75", width=140, height=32,
                      corner_radius=6, font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Cancel", command=on_cancel,
                      fg_color="#555555", hover_color="#666666", width=80, height=32,
                      corner_radius=6, font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=6)

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        self.win.wait_window(dialog)

        if result[0]:
            self._attach_btn.configure(state="disabled")
            threading.Thread(target=self._attach_docs_worker,
                             args=(cby_pdfs_map, awb_bol_pdfs, placeholder_pdf),
                             daemon=True).start()

    def _find_supporting_pdfs(self):
        """Find PDFs in the manifest folder and subfolders.
        The manifest folder is either the XML folder itself or its parent
        (if the XML folder is a subfolder like 'XML Files')."""
        search_dirs = set()

        # Always search the XML folder itself
        search_dirs.add(self._xml_folder)

        # Search immediate subfolders of the XML folder
        try:
            for item in self._xml_folder.iterdir():
                if item.is_dir():
                    search_dirs.add(item)
        except Exception:
            pass

        # Always search the parent folder and its subfolders too.
        # The manifest PDF (Manifest and Customs*.pdf) typically lives
        # in the parent, and the XML folder may be named 'invoices',
        # 'XML files', or anything else.
        parent = self._xml_folder.parent
        if parent != self._xml_folder and parent.exists():
            search_dirs.add(parent)
            try:
                for item in parent.iterdir():
                    if item.is_dir():
                        search_dirs.add(item)
            except Exception:
                pass

        # Find all PDFs in the search dirs
        pdfs = set()
        for d in search_dirs:
            if d.exists():
                for pdf in d.glob("*.pdf"):
                    pdfs.add(pdf)
                for pdf in d.glob("*.PDF"):
                    pdfs.add(pdf)

        return sorted(pdfs, key=lambda p: p.name.lower())

    def _attach_docs_worker(self, cby_pdfs_map, awb_bol_pdfs, placeholder_pdf=None):
        """Background thread: attach supporting docs to each declaration on COLS."""
        self._uploading = True
        self._stop_requested = False
        self.win.after(0, lambda: self._pause_btn.configure(state="disabled"))

        self.win.after(0, lambda: self._update_status(
            "Looking for 'Upload Supporting Documents' links on COLS..."))

        try:
            links = self._driver.find_elements(
                By.XPATH, "//a[contains(text(), 'Upload Supporting Documents')]")
        except Exception as e:
            self.win.after(0, lambda: self._status.configure(
                text=f"Error finding links: {e}"))
            self.win.after(0, lambda: self._attach_btn.configure(state="normal"))
            self._uploading = False
            return

        if not links:
            self.win.after(0, lambda: messagebox.showwarning(
                "No Links Found",
                "Could not find any 'Upload Supporting Documents' links on the COLS page.\n\n"
                "Make sure you have uploaded XML files and they appear\n"
                "in the table on the COLS page."))
            self.win.after(0, lambda: self._status.configure(
                text="No 'Upload Supporting Documents' links found"))
            self.win.after(0, lambda: self._attach_btn.configure(state="normal"))
            self._uploading = False
            return

        attached_count = 0
        failed_count = 0
        skipped_count = 0
        # Only count visible links — hidden/stale links can cause the script
        # to keep trying past the real declarations.
        try:
            visible_links = [l for l in links if l.is_displayed()]
        except Exception:
            # Links went stale already — re-find them
            links = self._driver.find_elements(
                By.XPATH, "//a[contains(text(), 'Upload Supporting Documents')]")
            try:
                visible_links = [l for l in links if l.is_displayed()]
            except Exception:
                visible_links = links
        total_links = len(visible_links)
        consecutive_failures = 0
        MAX_CONSECUTIVE_FAILURES = 3

        _logging_enabled = getattr(self, '_logging_var', None) and self._logging_var.get()
        if _logging_enabled:
            log_path = Path.home() / "Documents" / "attach_docs_log.txt"
            try:
                with open(log_path, "a", encoding="utf-8") as logf:
                    logf.write(f"\n--- Attach session started ---\n")
                    logf.write(f"Total links found: {len(links)}  Visible: {total_links}\n")
                    logf.flush()
            except Exception as e:
                print(f"[Attach Docs] Logging setup error: {e}")
        else:
            import os
            log_path = os.devnull

        # Track processed links by their stable DOM id instead of position.
        # Oracle ADF re-stamps/recycles table rows as you scroll, so indexing
        # by position causes random skips. By id, we always pick the next
        # link we haven't done yet, regardless of how the list reshuffles.
        processed_link_ids = set()
        i = -1  # iteration counter (kept for status/error messages)

        # CBYs we've confirmed docs were attached to (persists across the
        # double-check pass). Used to skip re-doing ones already handled.
        attached_cbys = set()
        # Track CBYs we attempted but failed/skipped during the main pass.
        # If this is empty at the end, the double-check pass is skipped entirely
        # (no point re-walking everything if nothing was missed).
        missed_cbys = set()
        # After the main pass finishes, we do ONE final "double-check" pass to
        # catch any declaration that got skipped. It runs only once — some
        # genuinely can't attach (e.g. oversized PDF) and we don't want an
        # endless loop chasing them.
        doing_double_check = False

        while not self._stop_requested:
            # If too many consecutive failures, the page is likely in a bad state.
            # Stop rather than spamming clicks that open dialogs we can't read.
            if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                self.win.after(0, lambda: self._update_status(
                    f"Stopped after {consecutive_failures} consecutive failures — page may need manual review"))
                self.win.after(0, lambda: messagebox.showwarning(
                    "Attachment Halted",
                    f"The script stopped after {consecutive_failures} consecutive dialogs could not be read.\n\n"
                    f"This usually means the COLS page is in an unexpected state.\n"
                    f"Attached: {attached_count}  Failed: {failed_count}  Skipped: {skipped_count}\n\n"
                    f"Please check the COLS page, close any open dialogs, and try again."))
                break

            # ---- Find the next link we haven't processed yet ----
            def _find_next_link():
                """Return (link, link_id) for the first unprocessed visible link,
                or (None, None) if there are none currently in the DOM."""
                found = self._driver.find_elements(
                    By.XPATH, "//a[contains(text(), 'Upload Supporting Documents')]")
                for l in found:
                    try:
                        if not l.is_displayed():
                            continue
                        lid = l.get_attribute("id")
                    except Exception:
                        continue
                    if lid and lid not in processed_link_ids:
                        return l, lid
                return None, None

            link, link_id = _find_next_link()

            # If nothing unprocessed is currently rendered, scroll down to force
            # ADF to render more rows, then look again. If still nothing after
            # reaching the bottom, we're genuinely done.
            if link is None:
                prev_scroll = self._driver.execute_script("return window.pageYOffset;")
                self._driver.execute_script("window.scrollBy(0, 800);")
                time.sleep(1.2)
                link, link_id = _find_next_link()
                if link is None:
                    new_scroll = self._driver.execute_script("return window.pageYOffset;")
                    if new_scroll == prev_scroll:
                        # Bottom reached and no unprocessed links remain.
                        if not doing_double_check:
                            # Only run the double-check pass if the main pass
                            # actually missed something. If everything attached
                            # cleanly, skip the re-walk entirely — saves several
                            # minutes on good runs.
                            if not missed_cbys:
                                if _logging_enabled:
                                    with open(log_path, "a", encoding="utf-8") as logf:
                                        logf.write(f"\nAll links processed cleanly "
                                                   f"({len(attached_cbys)} attached) — "
                                                   f"skipping double-check (no misses).\n")
                                break
                            # Start the ONE-TIME double-check pass: scroll back to
                            # the top, forget which link ids we've seen (so we
                            # re-scan everything), and re-walk the list. The CBY
                            # skip-check below means already-attached declarations
                            # are closed immediately, so this only re-attaches misses.
                            doing_double_check = True
                            processed_link_ids = set()
                            consecutive_failures = 0
                            # Remember how many declarations the double-check pass
                            # will walk, so its own progress bar can fill 0 -> total.
                            double_check_total = max(total_links, 1)
                            self._driver.execute_script("window.scrollTo(0, 0);")
                            time.sleep(1.2)
                            self.win.after(0, lambda: self._update_status(
                                "Double-checking for any missed documents..."))
                            if self._autopilot_bar and self._autopilot_bar.is_alive():
                                self._autopilot_bar.set_phase("Double-checking Documents...")
                                self._autopilot_bar.set_progress(0, double_check_total)
                            if _logging_enabled:
                                with open(log_path, "a", encoding="utf-8") as logf:
                                    logf.write(f"\n--- Double-check pass started "
                                               f"({len(attached_cbys)} attached so far, "
                                               f"{len(missed_cbys)} missed) ---\n")
                            continue
                        # Double-check already done — we're truly finished.
                        if _logging_enabled:
                            with open(log_path, "a", encoding="utf-8") as logf:
                                logf.write(f"\nAll links processed "
                                           f"({len(attached_cbys)} attached total).\n")
                        break
                    # Scrolled but nothing new yet — loop again to keep scrolling.
                    continue

            # Mark as processed up front so a failure on this link doesn't cause
            # us to retry it forever.
            processed_link_ids.add(link_id)
            i += 1
            # The initial scan may have under-counted (ADF renders rows lazily),
            # so keep the displayed total at least as large as what we've done.
            if len(processed_link_ids) > total_links:
                total_links = len(processed_link_ids)

            if doing_double_check:
                # During the double-check, drive the bar 0 -> double_check_total
                # so the always-on-top window visibly moves instead of sitting
                # frozen at a full yellow fill.
                dc_total = max(double_check_total, len(processed_link_ids))
                self.win.after(0, lambda n=len(processed_link_ids), t=dc_total: self._update_status(
                    f"Double-checking {n} of {t}..."))
                if self._autopilot_bar and self._autopilot_bar.is_alive():
                    self._autopilot_bar.set_progress(len(processed_link_ids), dc_total)
            else:
                self.win.after(0, lambda n=len(processed_link_ids), t=total_links: self._update_status(
                    f"Attaching docs to declaration {n} of {t}..."))
                if self._autopilot_bar and self._autopilot_bar.is_alive():
                    self._autopilot_bar.set_progress(len(processed_link_ids), total_links)

            try:
                # Before clicking, make sure no dialog is already open from a
                # previous iteration. If one is lingering, close it first —
                # this prevents multiple dialogs stacking up.
                try:
                    existing = self._driver.find_elements(
                        By.XPATH, "//div[contains(text(), 'Supporting Documents - HBL')]")
                    if any(d.is_displayed() for d in existing):
                        self._close_supporting_docs_dialog()
                        time.sleep(1)
                except Exception:
                    pass

                # Scroll the link into view
                self._driver.execute_script("arguments[0].scrollIntoView(true);", link)
                time.sleep(0.5)

                # Click the link
                link.click()
                time.sleep(2)

                # Wait for dialog to appear and read title
                title_text = ""
                for attempt in range(10):  # Try for up to 5 seconds
                    try:
                        title_el = self._driver.find_element(
                            By.XPATH, "//div[contains(text(), 'Supporting Documents - HBL')]")
                        title_text = title_el.text
                        if title_text:
                            break
                    except Exception:
                        pass
                    time.sleep(0.5)

                if not title_text:
                    try:
                        title_el = self._driver.find_element(
                            By.XPATH, "//div[contains(@class, 'x1h9') and contains(text(), 'HBL')]")
                        title_text = title_el.text
                    except Exception:
                        pass

                with open(log_path, "a", encoding="utf-8") as logf:
                    logf.write(f"\nDialog title: '{title_text}'\n")
                    if not title_text:
                        logf.write(f"  WARNING: Could not read dialog title!\n")
                        # Dump page state for debugging
                        logf.write(f"  Current URL: {self._driver.current_url}\n")
                        divs = self._driver.find_elements(
                            By.XPATH, "//div[contains(@class, 'x1h9')]")
                        logf.write(f"  Dialog-like divs found: {len(divs)}\n")
                        for di, d in enumerate(divs[:5]):
                            try:
                                logf.write(f"    [{di}] text='{d.text[:100]}'\n")
                            except:
                                pass

                # Extract CBY from title (e.g., "Supporting Documents - HBL-CBY 2006 1091363994.xml")
                cby_match = re.search(r'CBY\s*(\d+)', title_text)
                if not cby_match:
                    self.win.after(0, lambda t=title_text: self._status.configure(
                        text=f"Could not find CBY in dialog title: {t}"))
                    # Aggressively close whatever might be open — press Escape
                    # multiple times and try close buttons, then wait longer.
                    self._close_supporting_docs_dialog()
                    time.sleep(0.5)
                    self._close_supporting_docs_dialog()
                    time.sleep(2)
                    consecutive_failures += 1
                    failed_count += 1
                    if not doing_double_check:
                        missed_cbys.add("?")  # unknown CBY — forces double-check
                    continue

                # Reset consecutive failure counter on success
                consecutive_failures = 0
                cby = cby_match.group(1)

                # If this CBY already got its docs (earlier in this pass, or in
                # the main pass before the double-check), just close and move on.
                if cby in attached_cbys:
                    if _logging_enabled:
                        with open(log_path, "a", encoding="utf-8") as logf:
                            logf.write(f"\nCBY {cby}: already attached, skipping.\n")
                    self._close_supporting_docs_dialog()
                    time.sleep(0.5)
                    continue

                # Update treeview: mark this CBY as Attaching...
                self.win.after(0, lambda c=cby: self._set_row_docs_status(c, "docs_attach", "Attaching..."))

                # Find PDFs for this CBY
                cby_pdfs = cby_pdfs_map.get(cby, [])
                all_pdfs_for_cby = cby_pdfs + awb_bol_pdfs

                # If no invoices matched but we have a placeholder, use it
                if not cby_pdfs and placeholder_pdf:
                    all_pdfs_for_cby = [placeholder_pdf] + awb_bol_pdfs
                    self.win.after(0, lambda c=cby: self._status.configure(
                        text=f"Using placeholder for CBY {c} (no invoice found)..."))

                if not all_pdfs_for_cby:
                    self.win.after(0, lambda c=cby: self._status.configure(
                        text=f"No PDFs found for CBY {c}, skipping..."))
                    self.win.after(0, lambda c=cby: self._set_row_docs_status(c, "docs_manual", "No PDFs - Check Manually"))
                    if self._autopilot_bar and self._autopilot_bar.is_alive():
                        self._autopilot_bar.add_issue()
                    self._close_supporting_docs_dialog()
                    skipped_count += 1
                    if not doing_double_check:
                        missed_cbys.add(cby)
                    time.sleep(1)
                    continue

                self.win.after(0, lambda c=cby, n=len(all_pdfs_for_cby): self._status.configure(
                    text=f"Attaching {n} document(s) to CBY {c}..."))

                if _logging_enabled:
                    with open(log_path, "a", encoding="utf-8") as logf:
                        logf.write(f"\n{'='*60}\n")
                        logf.write(f"CBY: {cby} | PDFs: {len(all_pdfs_for_cby)}\n")
                        for pdf in all_pdfs_for_cby:
                            logf.write(f"  - {pdf.name}\n")

                # Separate AWB/BOL from invoices
                cby_awb = [pdf for pdf in all_pdfs_for_cby
                           if pdf.name.upper().lstrip("#").startswith("378")
                           or pdf.name.upper().startswith("SMLU")
                           or pdf.name.upper().startswith("#SMLU")]
                cby_invoices = [pdf for pdf in all_pdfs_for_cby if pdf not in cby_awb]
                ordered_pdfs = cby_awb + cby_invoices

                with open(log_path, "a", encoding="utf-8") as logf:
                    logf.write(f"Ordered: {[p.name for p in ordered_pdfs]}\n")

                attach_failed = False
                for j, pdf in enumerate(ordered_pdfs):
                    if self._stop_requested:
                        break

                    with open(log_path, "a", encoding="utf-8") as logf:
                        logf.write(f"\n--- Field {j}: {pdf.name} ---\n")

                    # Step 1: If beyond the 2 default fields, click "Add Document"
                    if j >= 2:
                        with open(log_path, "a", encoding="utf-8") as logf:
                            logf.write(f"  Clicking 'Add Document' for extra field...\n")
                        try:
                            # Select 'Invoice' doc type BEFORE clicking Add Document
                            # (Diagnostic: user selects type first, then clicks Add)
                            from selenium.webdriver.support.ui import Select
                            dropdowns = self._driver.find_elements(
                                By.XPATH, "//select[contains(@id, 'soc1')]")
                            with open(log_path, "a", encoding="utf-8") as logf:
                                logf.write(f"  Found {len(dropdowns)} doc-type dropdowns\n")
                            if dropdowns:
                                # Select by visible text 'Invoice' first, fall back to value
                                try:
                                    Select(dropdowns[-1]).select_by_visible_text("Invoice")
                                except Exception:
                                    Select(dropdowns[-1]).select_by_value("1")
                                time.sleep(0.5)
                                with open(log_path, "a", encoding="utf-8") as logf:
                                    logf.write("  Selected Invoice doc type\n")
                            add_btn = self._driver.find_element(
                                By.XPATH, "//button[contains(text(), 'Add Document')]")
                            # Count existing file inputs before clicking
                            before = len(self._driver.find_elements(
                                By.XPATH, "//input[@type='file']"))
                            add_btn.click()
                            # Wait up to 5 s for a new file input to appear
                            for _aw in range(10):
                                time.sleep(0.5)
                                after = len(self._driver.find_elements(
                                    By.XPATH, "//input[@type='file']"))
                                if after > before:
                                    break
                            with open(log_path, "a", encoding="utf-8") as logf:
                                logf.write(f"  Add Document clicked: inputs {before}->{after}\n")
                        except Exception as e:
                            with open(log_path, "a", encoding="utf-8") as logf:
                                logf.write(f"  ERROR adding field: {e}\n")

                    # Step 2: RE-FIND dialog inputs fresh each time.
                    # ADF removes a filled input from the DOM after each attachment,
                    # so the next empty input is ALWAYS at index 0.
                    attached_ok = False
                    for retry in range(3):
                        try:
                            all_inputs = self._driver.find_elements(
                                By.XPATH, "//input[@type='file']")
                            dialog_inputs = [fi for fi in all_inputs
                                           if "if3" in (fi.get_attribute("id") or "")]
                            # Fallback: 'Add Document' new row may use different ID pattern
                            if not dialog_inputs:
                                dialog_inputs = all_inputs
                            with open(log_path, "a", encoding="utf-8") as logf:
                                logf.write(f"  Attempt {retry+1}: All={len(all_inputs)}, "
                                          f"Dialog={len(dialog_inputs)}\n")
                                for di, fi in enumerate(dialog_inputs):
                                    logf.write(f"    dialog_input[{di}]: "
                                              f"id={fi.get_attribute('id')}\n")

                            # Always use index 0 — ADF removes filled inputs from DOM
                            if len(dialog_inputs) > 0:
                                dialog_inputs[0].send_keys(str(pdf.resolve()))
                                time.sleep(3)  # Wait for ADF to re-render after attach
                                # Verify: check if filename appeared in dialog text.
                                # ADF shows the filename after successful attach.
                                # If the browser froze or attach failed silently,
                                # the filename won't be there — retry.
                                try:
                                    dialog_text = self._driver.find_element(
                                        By.XPATH, "//div[contains(@id,':d1')]").text
                                    pdf_stem = pdf.stem
                                    if pdf_stem in dialog_text or pdf.name in dialog_text:
                                        with open(log_path, "a", encoding="utf-8") as logf:
                                            logf.write(f"  Attached {pdf.name} to dialog_input[0] (verified)\n")
                                        attached_ok = True
                                        break
                                    else:
                                        with open(log_path, "a", encoding="utf-8") as logf:
                                            logf.write(f"  Retry {retry+1}: {pdf.name} not in dialog text yet, retrying...\n")
                                        time.sleep(1)
                                except Exception:
                                    # Can't read dialog text — assume success
                                    with open(log_path, "a", encoding="utf-8") as logf:
                                        logf.write(f"  Attached {pdf.name} to dialog_input[0] (no verify)\n")
                                    attached_ok = True
                                    break
                            else:
                                with open(log_path, "a", encoding="utf-8") as logf:
                                    logf.write(f"  Retry {retry+1}: no dialog inputs available\n")
                                time.sleep(1)
                        except Exception as e:
                            with open(log_path, "a", encoding="utf-8") as logf:
                                logf.write(f"  Retry {retry+1} error: {type(e).__name__}: {e}\n")
                            time.sleep(1)

                    if not attached_ok:
                        with open(log_path, "a", encoding="utf-8") as logf:
                            logf.write(f"  FAILED to attach {pdf.name} after 3 retries\n")
                        attach_failed = True
                        self.win.after(0, lambda n=pdf.name: self._status.configure(
                            text=f"Warning: could not attach {n}"))
                        self.win.after(0, lambda c=cby: self._set_row_docs_status(c, "docs_manual", "Check Manually"))
                        if self._autopilot_bar and self._autopilot_bar.is_alive():
                            self._autopilot_bar.add_issue()
                        if not doing_double_check:
                            missed_cbys.add(cby)

                # Step 3: Click the real 'Upload' <A> link in the dialog.
                # Diagnostic confirmed: Upload is an <A> inside a div whose id
                # ends in ':d1_ok'. Must use real Selenium .click() so ADF's
                # event handlers fire — JS .click() on upDecl bypasses them.
                upload_clicked = False
                if not attach_failed:
                    # Primary: find the A link inside the *:d1_ok footer div
                    for _w in range(20):
                        try:
                            upload_link = self._driver.find_element(
                                By.XPATH,
                                "//div[contains(@id,':d1_ok')]//a")
                            upload_link.click()
                            upload_clicked = True
                            time.sleep(3)
                            break
                        except Exception:
                            time.sleep(0.5)
                    # Fallback: find any <A> wrapping a SPAN with text 'Upload'
                    if not upload_clicked:
                        with open(log_path, "a", encoding="utf-8") as logf:
                            logf.write("  d1_ok link not found, trying span fallback...\n")
                        try:
                            upload_link = self._driver.find_element(
                                By.XPATH,
                                "//a[.//span[normalize-space(text())='Upload']]")
                            upload_link.click()
                            upload_clicked = True
                            time.sleep(3)
                        except Exception as e:
                            with open(log_path, "a", encoding="utf-8") as logf:
                                logf.write(f"  Span fallback failed: {e}\n")

                if not upload_clicked:
                    with open(log_path, "a", encoding="utf-8") as logf:
                        logf.write(f"\nWARNING: Could not find/click Upload link!\n")
                    self.win.after(0, lambda: self._status.configure(
                        text="Warning: could not find Upload link, skipping..."))
                    self.win.after(0, lambda c=cby: self._set_row_docs_status(c, "docs_manual", "Check Manually"))
                    self._close_supporting_docs_dialog()
                    if not doing_double_check:
                        missed_cbys.add(cby)
                    time.sleep(1)
                else:
                    with open(log_path, "a", encoding="utf-8") as logf:
                        logf.write(f"Upload clicked successfully for CBY {cby}\n")
                    # This CBY is now genuinely done — remember it so the
                    # double-check pass closes it immediately instead of redoing.
                    attached_cbys.add(cby)
                    # Wait for COLS to close the dialog automatically after upload.
                    # Poll for the dialog to disappear (max 10s), then proceed.
                    for _dw in range(20):
                        time.sleep(0.5)
                        try:
                            dialogs = self._driver.find_elements(
                                By.XPATH, "//div[contains(text(), 'Supporting Documents - HBL')]")
                            if not dialogs or not any(d.is_displayed() for d in dialogs):
                                break  # dialog is gone
                        except Exception:
                            break  # page changed, dialog likely gone

                attached_count += 1
                # Reclassify false "error" rows as "uploaded" — if docs were
                # attached, the file IS on COLS, so the error was a false positive.
                self.win.after(0, lambda c=cby: self._reclassify_error_if_docs_attached(c))
                self.win.after(0, lambda c=cby, n=len(all_pdfs_for_cby): self._update_status(
                    f"Attached {n} document(s) to CBY {c} ({attached_count} done)"))
                self.win.after(0, lambda c=cby: self._set_row_docs_status(c, "docs_ok", "Attached"))

            except Exception as e:
                failed_count += 1
                err_msg = f"{type(e).__name__}: {e}"
                self.win.after(0, lambda msg=err_msg: self._status.configure(
                    text=f"Failed (#{i+1}): {msg}"))
                print(f"[Attach Docs] Failed on link {i+1}: {err_msg}")
                if not doing_double_check:
                    missed_cbys.add("?")  # unknown — forces double-check
                try:
                    self._close_supporting_docs_dialog()
                except Exception:
                    pass
                time.sleep(2)

        # Show summary
        self._uploading = False
        self.win.after(0, lambda: self._attach_btn.configure(state="normal"))
        # Re-enable pause button if upload is paused
        if self._paused:
            self.win.after(0, lambda: self._pause_btn.configure(state="normal"))
        # Update autopilot bar to complete state
        if self._autopilot_bar and self._autopilot_bar.is_alive():
            docs_issues = failed_count + skipped_count
            # Count upload results from treeview
            up = err = miss = 0
            for item in self._tree.get_children():
                s = self._tree.set(item, "status").lower()
                if s == "uploaded": up += 1
                elif s == "error": err += 1
                elif s == "missed": miss += 1
            self._autopilot_bar.complete(up, err, miss, attached_count, docs_issues)
        self.win.after(0, lambda: self._update_status(
            f"Done: {attached_count} declarations had docs attached, "
                 f"{skipped_count} skipped, {failed_count} failed"))
        self.win.after(0, lambda: messagebox.showinfo(
            "Attach Supporting Docs Complete",
            f"Attached supporting documents to {attached_count} declaration(s).\n"
            f"Skipped: {skipped_count} (no matching PDFs)\n"
            f"Failed: {failed_count}\n\n"
            f"Review the COLS page to verify the documents are attached,\n"
            f"then check the declaration checkbox and click Submit when ready."))

    def _close_supporting_docs_dialog(self):
        """Close the Supporting Documents dialog on COLS."""
        # Strategy 1: Find close button by aria-label
        for xpath in [
            "//a[@aria-label='Close' and contains(@id, 'd1::close')]",
            "//a[@title='Close' and contains(@id, 'd1::close')]",
            "//a[@aria-label='Close']",
            "//a[contains(@class, 'x1h3')]",
            "//a[contains(@id, '::close')]",
        ]:
            try:
                close_btn = self._driver.find_element(By.XPATH, xpath)
                if close_btn.is_displayed():
                    close_btn.click()
                    time.sleep(1.5)
                    return
            except Exception:
                continue
        # Strategy 2: Press Escape (try multiple times)
        try:
            from selenium.webdriver.common.keys import Keys
            body = self._driver.find_element(By.TAG_NAME, "body")
            for _ in range(3):
                body.send_keys(Keys.ESCAPE)
                time.sleep(0.5)
        except Exception:
            pass

    def _reclassify_error_if_docs_attached(self, cby):
        """If a row is marked 'error' but docs were just attached to it on COLS,
        the file was actually uploaded successfully (false error). Reclassify
        as 'uploaded' so the copy feature can capture its declaration number."""
        for item in self._tree.get_children():
            row_cby = self._tree.set(item, "cby")
            row_status = self._tree.set(item, "status").lower()
            if row_cby == cby and row_status == "error":
                self._tree.set(item, "status", "Uploaded")
                self._tree.item(item, tags=("uploaded",))
                self._update_status(f"CBY {cby} reclassified: error → uploaded (docs attached)")
                break

    def _copy_decl_from_cols(self):
        """Read the COLS page for declaration numbers and filenames.
        COLS uses Oracle ADF with absolutely-positioned divs, not HTML tables.
        The _afrc attribute encodes column/row positions:
          _afrc="4 1 X 1..." = Declaration No column
          _afrc="6 1 X 1..." = Filename column
        We pair them by row number (X)."""
        if not self._driver:
            messagebox.showwarning("No Browser", "Launch Edge first, then use this button.")
            return

        self._status.configure(text="Reading declaration numbers from COLS page...")

        pairs = []  # list of (declaration_number, filename) tuples

        try:
            # Use JavaScript to extract data from the ADF grid
            # The _afrc attribute format is "col 1 row 1 start top"
            # Column 4 = Declaration No, Column 6 = Filename
            js_script = """
            var results = [];
            // Find all divs with _afrc attribute
            var allDivs = document.querySelectorAll('div[_afrc]');
            var declCells = {};
            var fileCells = {};

            for (var i = 0; i < allDivs.length; i++) {
                var div = allDivs[i];
                var afrc = div.getAttribute('_afrc');
                if (!afrc) continue;
                var parts = afrc.split(' ');
                if (parts.length < 2) continue;
                var col = parts[0];
                var row = parts[2];

                var text = div.textContent.trim();

                if (col === '4' && parseInt(row) > 2) {
                    // Declaration No column, skip header row (row 2)
                    declCells[row] = text;
                } else if (col === '6' && parseInt(row) > 2) {
                    // Filename column, skip header row
                    fileCells[row] = text;
                }
            }

            // Pair declaration numbers with filenames by row
            for (var row in fileCells) {
                var filename = fileCells[row];
                var declNum = declCells[row] || '-';
                if (filename && filename.indexOf('HBL-') >= 0) {
                    results.push([declNum, filename]);
                }
            }
            return results;
            """
            js_results = self._driver.execute_script(js_script)

            for item in js_results:
                decl_num = str(item[0]).strip()
                filename = str(item[1]).strip()
                # Skip rows where declaration number is "-" (not yet submitted)
                if decl_num and decl_num != "-" and filename:
                    pairs.append((decl_num, filename))
                elif filename:
                    # Include with empty decl number so user can see it and fill in
                    pairs.append(("", filename))

        except Exception as e:
            # Fallback: try reading by position (left:160px = decl, left:345px = filename)
            try:
                js_fallback = """
                var results = [];
                var spans = document.querySelectorAll('span');
                var byTop = {};
                for (var i = 0; i < spans.length; i++) {
                    var span = spans[i];
                    var text = span.textContent.trim();
                    if (!text) continue;
                    var parent = span.parentElement;
                    if (!parent) continue;
                    var left = parent.style.left;
                    var top = parent.style.top;
                    if (!left || !top) continue;
                    if (!byTop[top]) byTop[top] = {};
                    if (left === '160px') byTop[top]['decl'] = text;
                    if (left === '345px') byTop[top]['file'] = text;
                }
                for (var top in byTop) {
                    var entry = byTop[top];
                    if (entry['file'] && entry['file'].indexOf('HBL-') >= 0) {
                        results.push([entry['decl'] || '', entry['file']]);
                    }
                }
                return results;
                """
                js_results = self._driver.execute_script(js_fallback)
                for item in js_results:
                    decl_num = str(item[0]).strip()
                    filename = str(item[1]).strip()
                    if decl_num and decl_num != "-":
                        pairs.append((decl_num, filename))
                    elif filename:
                        pairs.append(("", filename))
            except Exception as e2:
                messagebox.showerror("Error", f"Could not read COLS page:\n{e2}")
                self._status.configure(text="Failed to read COLS page")
                return

        if not pairs:
            messagebox.showwarning("No Declaration Numbers Found",
                                   "Could not find any declaration data on the COLS page.\n\n"
                                   "Make sure you have uploaded files and they appear\n"
                                   "in the table on the COLS page.\n\n"
                                   "You can also enter them manually by double-clicking\n"
                                   "the Declaration # column in the uploader window.")
            self._status.configure(text="No declaration data found on COLS page")
            return

        # Show what we found and let user confirm
        result = self._show_decl_capture_dialog(pairs)
        if result is None:
            self._status.configure(text="Declaration number capture cancelled")
            return

        # Match declaration numbers to treeview rows by filename
        matched = 0
        for decl_num, col_filename in result:
            if not decl_num or not col_filename:
                continue
            for item in self._tree.get_children():
                tree_file = self._tree.set(item, "file")
                tree_status = self._tree.set(item, "status").lower()
                # Match by exact filename or by CBY number
                if tree_file == col_filename or col_filename in tree_file or tree_file in col_filename:
                    if tree_status in ("uploaded", "error"):
                        self._tree.set(item, "decl", decl_num)
                        self._decl_numbers[tree_file] = decl_num
                        matched += 1
                    break
                tree_cby = self._tree.set(item, "cby")
                if tree_cby and tree_cby != "MASTER":
                    col_cby_match = re.search(r'CBY\s*(\d+)', col_filename, re.IGNORECASE)
                    if col_cby_match and col_cby_match.group(1) == tree_cby:
                        if tree_status in ("uploaded", "error"):
                            self._tree.set(item, "decl", decl_num)
                            self._decl_numbers[tree_file] = decl_num
                            matched += 1
                        break

        self._status.configure(
            text=f"Copied {len(result)} declaration number(s) from COLS. Matched {matched} to files.")
        if matched > 0:
            messagebox.showinfo("Success",
                                f"Matched {matched} declaration number(s) to uploaded files.\n\n"
                                f"Review the Declaration # column to verify.\n"
                                f"Double-click any cell to edit manually.")
        else:
            messagebox.showwarning("No Matches",
                                f"Found {len(result)} declaration number(s) on COLS,\n"
                                f"but could not match them to files by filename.\n\n"
                                f"The filenames on COLS may differ from the XML filenames.\n"
                                f"Double-click the Declaration # column to enter them manually.")

    def _show_decl_capture_dialog(self, pairs):
        """Show a dialog with declaration numbers and filenames for user to confirm/edit.
        pairs is a list of (declaration_number, filename) tuples."""
        dialog = ctk.CTkToplevel(self.win)
        dialog.title("Declaration Numbers from COLS")
        dialog.configure(fg_color="#1a1a2e")
        dialog.geometry("550x450")
        dialog.resizable(False, False)
        dialog.transient(self.win)
        dialog.grab_set()

        dialog.update_idletasks()
        px = self.win.winfo_rootx() + (self.win.winfo_width() - 550) // 2
        py = self.win.winfo_rooty() + (self.win.winfo_height() - 450) // 2
        dialog.geometry(f"+{px}+{py}")

        ctk.CTkLabel(dialog, text="Declaration Numbers Found on COLS",
                     font=(MODERN_FONT, 14, "bold"), text_color="#e8e8e8").pack(pady=(16, 4))
        ctk.CTkLabel(dialog,
                     text="Review the declaration numbers and filenames below.\n"
                          "Edit if needed. They will be matched to uploaded files.",
                     font=(MODERN_FONT, 11), text_color="#888888").pack(pady=(0, 12))

        # Column headers
        header_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(0, 4))
        ctk.CTkLabel(header_frame, text="Declaration Number",
                     font=(MODERN_FONT, 11, "bold"), text_color="#e8820c",
                     width=140, anchor="w").pack(side="left", padx=(0, 8))
        ctk.CTkLabel(header_frame, text="Filename",
                     font=(MODERN_FONT, 11, "bold"), text_color="#888888",
                     width=340, anchor="w").pack(side="left")

        # Scrollable frame for entries
        scroll = ctk.CTkScrollableFrame(dialog, fg_color="transparent", height=260)
        scroll.pack(fill="both", expand=True, padx=20, pady=(0, 12))

        decl_entries = []
        file_entries = []
        for decl_num, filename in pairs:
            row_frame = ctk.CTkFrame(scroll, fg_color="transparent")
            row_frame.pack(fill="x", pady=2)

            de = ctk.CTkEntry(row_frame, width=140, height=28,
                              fg_color="#0f0f1a", border_color="#e8820c", border_width=1,
                              corner_radius=4, text_color="#e8e8e8",
                              font=(MODERN_FONT, 12))
            de.insert(0, decl_num)
            de.pack(side="left", padx=(0, 8))
            decl_entries.append(de)

            fe = ctk.CTkEntry(row_frame, width=340, height=28,
                              fg_color="#0f0f1a", border_color="#333", border_width=1,
                              corner_radius=4, text_color="#e8e8e8",
                              font=(MODERN_FONT, 11))
            fe.insert(0, filename)
            fe.pack(side="left")
            file_entries.append(fe)

        result = [None]
        def on_ok():
            result[0] = [(de.get().strip(), fe.get().strip())
                         for de, fe in zip(decl_entries, file_entries)
                         if de.get().strip()]
            dialog.destroy()
        def on_cancel():
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=(0, 16))
        ctk.CTkButton(btn_frame, text="Apply to Uploader Window", command=on_ok,
                      fg_color="#2e8b57", hover_color="#3cb371", width=180, height=32,
                      corner_radius=6, font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Cancel", command=on_cancel,
                      fg_color="#555555", hover_color="#666666", width=80, height=32,
                      corner_radius=6, font=(MODERN_FONT, 12, "bold")).pack(side="left", padx=6)

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        self.win.wait_window(dialog)
        return result[0]

    def _quick_paste(self):
        """Open a manifest file, find the COLS sheet, match CBY numbers,
        and paste declaration numbers into the declaration column.

        Reads with openpyxl (data_only) to find columns and CBY values,
        then edits the sheet XML directly inside the xlsx zip to write
        declaration values. This preserves all external links, drawings,
        and formulas — no Excel process needed, no corruption.
        """
        if not self._decl_numbers:
            # They haven't copied/uploaded declaration numbers yet.
            # Replicate the "Copy Decl #s from COLS" flow automatically
            # so they see the same review dialog before pasting.
            proceed = messagebox.askyesno(
                "Copy Declaration Numbers First",
                "You haven't copied any declaration numbers yet.\n\n"
                "Click Yes to copy them from the COLS page now\n"
                "(same as clicking 'Copy Decl #s from COLS').\n\n"
                "You'll be able to review and confirm the list before\n"
                "anything is pasted to the manifest.")
            if not proceed:
                return
            self._copy_decl_from_cols()
            if not self._decl_numbers:
                # Copy step already showed its own message (no browser,
                # nothing found, or the user cancelled the review dialog)
                return

        file_path = filedialog.askopenfilename(
            title="Choose manifest file",
            filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return

        # Build CBY -> decl map
        cby_to_decl = {}
        master_decl = None
        for filename, decl_num in self._decl_numbers.items():
            m = re.search(r'CBY\s*(\d+)', filename, re.IGNORECASE)
            if m and decl_num:
                cby_to_decl[m.group(1)] = decl_num
            elif ("HBL-Master" in filename or "MBL" in filename) and decl_num:
                master_decl = decl_num

        if not cby_to_decl and not master_decl:
            # Build a diagnostic dump of exactly what WAS captured so the
            # user can hit 'Report Issue' and the developer sees the guts
            # of why nothing matched — no guessing over the phone.
            diag = "Captured declaration numbers (filename -> decl#):\n"
            if self._decl_numbers:
                for fn, dn in self._decl_numbers.items():
                    diag += f"  {fn!r} -> {dn!r}\n"
            else:
                diag += "  (none)\n"
            diag += f"\nCBY regex used: CBY\\s*(\\d+)  (case-insensitive)\n"
            diag += f"Master patterns: 'HBL-Master' or 'MBL' in filename\n"
            diag += f"cby_to_decl result: {cby_to_decl}\n"
            diag += f"master_decl result: {master_decl!r}"
            _show_error_with_report("No Matches",
                "No declaration numbers with matching CBY numbers found.\n\n"
                "The uploaded files don't have CBY numbers in their names,\n"
                "or no house declarations were captured.\n\n"
                "Make sure the house XML files (HBL-CBY*.xml) were uploaded\n"
                "and their declaration numbers appear in the Declaration # column.",
                traceback_text=diag,
                window_name=self._window_name)
            return
        if not cby_to_decl and master_decl:
            diag = "Captured declaration numbers (filename -> decl#):\n"
            for fn, dn in self._decl_numbers.items():
                diag += f"  {fn!r} -> {dn!r}\n"
            diag += f"\nCBY regex used: CBY\\s*(\\d+)  (case-insensitive)\n"
            diag += f"cby_to_decl result: {cby_to_decl}\n"
            diag += f"master_decl result: {master_decl!r}"
            _show_error_with_report("No House Declarations",
                "Only a Master declaration was found — no House/CBY declarations.\n\n"
                "The Paste to Manifest feature needs house declaration numbers\n"
                "(from HBL-CBY*.xml files) to match against CBY rows in the manifest.\n\n"
                "Make sure the house XML files were uploaded and their declaration\n"
                "numbers appear in the Declaration # column.\n\n"
                f"Master declaration found: {master_decl}",
                traceback_text=diag,
                window_name=self._window_name)
            return

        try:
            import openpyxl
            import zipfile
            import shutil
            import xml.etree.ElementTree as ET
            from io import BytesIO
        except ImportError as e:
            messagebox.showerror("Error", f"Missing required library: {e}")
            return

        file_path = Path(file_path)

        # Check if the file is open in Excel (Windows locks open files)
        try:
            # Try opening for read+write AND renaming a temp file over it
            # (Excel may allow read-write but block file replacement)
            with open(str(file_path), 'r+b'):
                pass
        except PermissionError:
            self._show_manifest_open_error(file_path)
            return

        try:
            # ── Step 1: Read with openpyxl to find columns and CBY values ──────
            wb = openpyxl.load_workbook(str(file_path), data_only=True)

            cols_sheet_name = None
            for name in wb.sheetnames:
                if name.strip().upper().replace(" ", "") == "COLS":
                    cols_sheet_name = name
                    break
            if not cols_sheet_name:
                for name in wb.sheetnames:
                    if "cols" in name.strip().lower():
                        cols_sheet_name = name
                        break
            if not cols_sheet_name:
                messagebox.showerror("Sheet Not Found",
                                     f"Could not find a 'COLS' sheet.\n"
                                     f"Available: {', '.join(wb.sheetnames)}")
                return

            ws = wb[cols_sheet_name]

            # Find CBY and Declaration columns.
            # CBY header is unique to the data header row (not in the top section).
            # Once we find the CBY row, search that SAME row for the Decl column.
            # This avoids picking up 'COLS Declaration #' from the header section.
            cby_col = None
            decl_col = None
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=15):
                for cell in row:
                    val = str(cell.value).strip().lower() if cell.value else ""
                    if "cby" in val and cby_col is None:
                        cby_col = cell.column
                        header_row = cell.row
                        break
                if cby_col:
                    break

            if cby_col and header_row:
                # Search the SAME row for the Decl column
                for cell in ws[header_row]:
                    val = str(cell.value).strip().lower() if cell.value else ""
                    if "decl" in val and decl_col is None:
                        decl_col = cell.column

            if not cby_col:
                messagebox.showerror("Column Not Found",
                                     "Could not find a CBY column in the COLS sheet.")
                return
            if not decl_col:
                messagebox.showerror("Column Not Found",
                                     "Could not find a Declaration column in the COLS sheet.")
                return

            # Collect all writes: (row, col, value)
            writes = []
            pasted = 0

            # Master declaration into header section (rows 1-8)
            if self._master_decl:
                for row in ws.iter_rows(min_row=1, max_row=8):
                    for cell in row:
                        val = str(cell.value).strip().lower() if cell.value else ""
                        if ("master decl" in val or "cols declaration" in val
                                or ("declaration" in val and "#" in val
                                    and cell.row <= 5)):
                            target = ws.cell(row=cell.row, column=cell.column + 1)
                            if not target.value:
                                try:
                                    mval = int(self._master_decl)
                                except (ValueError, TypeError):
                                    break  # skip master, don't write non-numeric
                                writes.append((cell.row, cell.column + 1, mval))
                                pasted += 1
                            break
                    else:
                        continue
                    break

            # House declarations
            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
                cby_cell = row[cby_col - 1] if cby_col <= len(row) else None
                if cby_cell and cby_cell.value:
                    cby_raw = cby_cell.value
                    if isinstance(cby_raw, float) and cby_raw == int(cby_raw):
                        cby_val = str(int(cby_raw))
                    else:
                        cby_val = str(cby_raw).strip()
                    decl = cby_to_decl.get(cby_val)
                    if not decl:
                        decl = cby_to_decl.get(cby_val.lstrip("0"))
                    if not decl:
                        for k, v in cby_to_decl.items():
                            if k.lstrip("0") == cby_val.lstrip("0"):
                                decl = v
                                break
                    if decl:
                        decl_cell = row[decl_col - 1] if decl_col <= len(row) else None
                        if decl_cell and not decl_cell.value:
                            # Only write numeric declaration values — the
                            # zip-editing method writes <v>...</v> which is
                            # only valid for numbers.  Non-numeric values
                            # would produce invalid Excel XML and corrupt
                            # the file.  Skip them instead.
                            try:
                                write_val = int(decl)
                            except (ValueError, TypeError):
                                continue  # skip this cell, don't write
                            writes.append((decl_cell.row, decl_cell.column, write_val))
                            pasted += 1

            wb.close()

            if not writes:
                messagebox.showinfo("Quick Paste",
                                    "No empty declaration cells found to paste into.\n"
                                    "All matching cells may already have values.")
                return

            # ── Step 2: Edit the sheet XML directly inside the xlsx zip ────────
            # This preserves everything — external links, drawings, formulas.
            try:
                success = self._write_cells_to_xlsx(str(file_path), cols_sheet_name,
                                                    writes, cby_col, decl_col, header_row)
            except PermissionError:
                self._show_manifest_open_error(file_path)
                return

            if success:
                messagebox.showinfo("Quick Paste Complete",
                                    f"Pasted {pasted} declaration number(s) into:\n"
                                    f"{file_path.name}\n"
                                    f"Sheet: {cols_sheet_name}\n\n"
                                    f"CBYs matched: {', '.join(sorted(cby_to_decl.keys()))}")
                self._status.configure(
                    text=f"Quick Paste: {pasted} declarations pasted to {file_path.name}")
            else:
                # _write_cells_to_xlsx returned False without raising —
                # this means it couldn't find the sheet inside the xlsx zip
                # structure.  Don't attempt an openpyxl save fallback — that
                # can corrupt the manifest by stripping external links and
                # drawings.  Instead, report the issue so it can be fixed.
                raise RuntimeError(
                    f"Could not locate the sheet '{cols_sheet_name}' inside the\n"
                    f"xlsx zip structure. The manifest was not modified.")

        except Exception as e:
            if "permission" in str(e).lower() or isinstance(e, PermissionError):
                self._show_manifest_open_error(file_path)
            else:
                # Real error — show with Report button so the developer
                # gets the full traceback in Discord.
                messagebox.showerror("Error", f"Failed to paste declarations:\n{e}")

    def _show_manifest_open_error(self, file_path):
        """Show a prominent, unmissable error that the manifest file is open
        or otherwise locked / not writable."""
        file_name = file_path.name if hasattr(file_path, 'name') else str(file_path)
        # Use a custom Toplevel dialog so it's large, bold, and stays on top
        dlg = ctk.CTkToplevel()
        dlg.title("CANNOT WRITE TO MANIFEST FILE")
        dlg.configure(fg_color="#0f1117")
        dlg.geometry("520x340")
        dlg.attributes("-topmost", True)
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"+{int((sw-520)/2)}+{int((sh-340)/2)}")
        dlg.transient()
        dlg.grab_set()
        # Red warning header
        ctk.CTkLabel(dlg, text="CANNOT WRITE TO MANIFEST",
                     font=(MODERN_FONT, 20, "bold"),
                     text_color="#e74c3c").pack(pady=(24, 8))
        ctk.CTkLabel(dlg,
                     text=f"The manifest file could not be written to:\n\n"
                          f"  {file_name}\n\n"
                          f"Possible causes:\n"
                          f"  1. The file is OPEN in Excel — close it and try again\n"
                          f"  2. The file is on a read-only or network folder\n"
                          f"  3. Your user account lacks write permission\n"
                          f"  4. Antivirus is blocking the write\n\n"
                          f"If closing Excel doesn't help, try saving a copy\n"
                          f"to your Desktop and pasting to that instead.",
                     font=(MODERN_FONT, 12),
                     text_color="#e8e8e8", justify="left").pack(pady=(0, 16))
        ctk.CTkButton(dlg, text="OK", command=dlg.destroy,
                      fg_color="#e74c3c", hover_color="#c0392b",
                      width=120, height=36, corner_radius=6,
                      font=(MODERN_FONT, 13, "bold")).pack(pady=(0, 24))
        dlg.bind("<Return>", lambda e: dlg.destroy())
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        self._status.configure(
            text=f"Quick Paste FAILED - close {file_name} in Excel and try again")

    def _write_cells_to_xlsx(self, file_path, sheet_name, writes, cby_col, decl_col, header_row):
        """Write cell values directly into the xlsx zip using string replacement.
        Preserves everything — external links, drawings, formulas, namespaces.
        No Excel process, no ElementTree (which corrupts namespace prefixes).
        Returns True on success, False on failure."""
        import zipfile
        import re
        import shutil
        import os

        def col_letter(n):
            result = ""
            while n > 0:
                n, rem = divmod(n - 1, 26)
                result = chr(65 + rem) + result
            return result

        # Build cell_ref -> value string map
        cell_writes = {}
        for row, col, val in writes:
            ref = f"{col_letter(col)}{row}"
            cell_writes[ref] = str(val)

        temp_path = file_path + ".tmp"

        try:
            # Step 1: Find which sheet file inside the zip is our COLS sheet
            with zipfile.ZipFile(file_path, 'r') as zin:
                wb_xml = zin.read('xl/workbook.xml').decode('utf-8')

                # Find sheet rId by name (string search, no XML parsing)
                # Pattern: <sheet name=" COLS" ... r:id="rId2"/>
                sheet_rid = None
                # Try exact name match first, then flexible
                for pattern in [
                    rf'<sheet name="{re.escape(sheet_name)}"[^>]*r:id="(rId\d+)"',
                    rf'<sheet name="\s*{re.escape(sheet_name.strip())}"[^>]*r:id="(rId\d+)"',
                ]:
                    m = re.search(pattern, wb_xml)
                    if m:
                        sheet_rid = m.group(1)
                        break

                if not sheet_rid:
                    # Fallback: find any sheet with "cols" in the name
                    m = re.search(r'<sheet name="[^"]*[Cc][Oo][Ll][Ss][^"]*"[^>]*r:id="(rId\d+)"', wb_xml)
                    if m:
                        sheet_rid = m.group(1)

                if not sheet_rid:
                    return False

                # Find sheet file path from rels
                rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                m = re.search(rf'<Relationship Id="{re.escape(sheet_rid)}"[^>]*Target="([^"]+)"', rels_xml)
                if not m:
                    return False
                sheet_file = 'xl/' + m.group(1)

                # Read the sheet XML as a string
                sheet_xml = zin.read(sheet_file).decode('utf-8')

            # Step 2: Modify the sheet XML using string replacement
            modified = sheet_xml
            written = 0

            for cell_ref, value in cell_writes.items():
                # Pattern 1: Cell exists with a value — replace the value
                # <c r="F10" s="153"><v>old</v></c>  or  <c r="F10" t="s"><v>old</v></c>
                # We need to handle various attribute orders and self-closing cells

                # Pattern A: Self-closing empty cell: <c r="F10" s="153"/>
                pattern_a = rf'<c r="{cell_ref}"([^>]*)/>'
                m_a = re.search(pattern_a, modified)
                if m_a:
                    attrs = m_a.group(1)
                    # Remove t="s" or t="str" if present (we're writing a number)
                    attrs = re.sub(r'\s+t="[^"]*"', '', attrs)
                    replacement = f'<c r="{cell_ref}"{attrs}><v>{value}</v></c>'
                    modified = modified[:m_a.start()] + replacement + modified[m_a.end():]
                    written += 1
                    continue

                # Pattern B: Cell with existing value: <c r="F10" ...>...</c>
                # Could contain <f>formula</f><v>value</v> or just <v>value</v>
                pattern_b = rf'(<c r="{cell_ref}"[^>]*>)(.*?)(</c>)'
                m_b = re.search(pattern_b, modified, re.DOTALL)
                if m_b:
                    attrs = m_b.group(1)
                    inner = m_b.group(2)
                    # Remove t="s" or t="str" if present
                    attrs = re.sub(r'\s+t="[^"]*"', '', attrs)
                    # Remove any existing <v>...</v> and <f>...</f>
                    inner = re.sub(r'<v>.*?</v>', '', inner, flags=re.DOTALL)
                    inner = re.sub(r'<f>.*?</f>', '', inner, flags=re.DOTALL)
                    replacement = f'{attrs}<v>{value}</v></c>'
                    modified = modified[:m_b.start()] + replacement + modified[m_b.end():]
                    written += 1
                    continue

                # Pattern C: Cell doesn't exist in the XML at all
                # We need to insert it into the right row
                row_num = int(''.join(c for c in cell_ref if c.isdigit()))
                col_part = ''.join(c for c in cell_ref if c.isalpha())

                # Find the row element for this row number
                row_pattern = rf'(<row r="{row_num}"[^>]*>)'
                m_row = re.search(row_pattern, modified)
                if m_row:
                    # Insert the cell at the end of the row (before </row>)
                    row_start = m_row.start()
                    row_end_tag = modified.find('</row>', row_start)
                    if row_end_tag >= 0:
                        new_cell = f'<c r="{cell_ref}"><v>{value}</v></c>'
                        modified = modified[:row_end_tag] + new_cell + modified[row_end_tag:]
                        written += 1
                        continue

                # If we get here, the row doesn't exist either.
                # Find the sheetData closing tag and insert a new row before it.
                # Find the last row before </sheetData> to insert after it.
                sheet_data_end = modified.find('</sheetData>')
                if sheet_data_end >= 0:
                    new_row = f'<row r="{row_num}"><c r="{cell_ref}"><v>{value}</v></c></row>'
                    modified = modified[:sheet_data_end] + new_row + modified[sheet_data_end:]
                    written += 1

            if written == 0:
                return False

            # Step 3: Write the new zip — copy everything except the sheet file,
            # then write the modified sheet XML
            with zipfile.ZipFile(file_path, 'r') as zin:
                with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        if item.filename == sheet_file:
                            zout.writestr(item, modified)
                        else:
                            zout.writestr(item, zin.read(item.filename))

            # Step 4: Replace original with temp
            shutil.move(temp_path, file_path)
            return True

        except PermissionError:
            # Clean up temp file on error
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception:
                pass
            raise  # Let the caller handle the "file is open" message
        except Exception as e:
            # Clean up temp file on error
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception:
                pass
            # Re-raise so the caller's outer except can show a proper
            # error dialog with Report button — don't swallow it as
            # "manifest is open" which is misleading.
            raise

    def _has_progress(self):
        """Check if there's any upload progress that would be lost."""
        if self._uploading:
            return True
        for item in self._tree.get_children():
            status = self._tree.set(item, "status").lower()
            if status in ("uploaded", "uploading", "error", "missed"):
                return True
        if self._decl_numbers:
            return True
        return False

    def _on_close(self):
        """Close window with confirmation if there's active progress."""
        if self._has_progress() or self._uploading:
            answer = messagebox.askyesno(
                "Close Upload Window?",
                "You have upload progress in this session.\n\n"
                "Closing this window will:\n"
                "- Stop any active uploads\n"
                "- Lose captured declaration numbers\n"
                "- The Edge browser will stay open\n\n"
                "Are you sure you want to close?",
                icon="warning")
            if not answer:
                return
        self._stop_requested = True
        self._paused = False
        # Don't kill the browser - let user close it manually
        self.launcher._upload_win = None
        self.win.destroy()
        self.launcher.show()


# ==============================================================================
# MAIN
# ==============================================================================
if __name__ == "__main__":
    app = Launcher()
    app.run()
