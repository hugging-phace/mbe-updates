# -*- coding: utf-8 -*-
"""
Packages at Customs Console
Tracks packages held at Customs (ADD status) — either pending inspection
("Inspection") or held for further documentation ("Held").
Loads from / saves to the ADD Status Log Excel file.
Export PDF produces a legible printout for the driver.
"""
import os
import sys
import re
import platform
import warnings
import subprocess
import importlib
import json
import threading
import traceback
import getpass
import urllib.request
import uuid
from pathlib import Path
from datetime import datetime, date
import io
import base64

import tkinter as tk
from tkinter import messagebox, filedialog
import tkinter.ttk as ttk

# ------------------------------------------------------------------
# Dependency check — verify required packages, offer to install missing ones
# ------------------------------------------------------------------
REQUIRED_PACKAGES = {
    "customtkinter": "customtkinter",
    "openpyxl":      "openpyxl",
    "fpdf":          "fpdf2",   # module name : pip package name
    "lxml":          "lxml",    # required by fpdf2 for image embedding (PDF export)
}

def _check_and_install_dependencies():
    """Check which required packages are missing. If any are missing,
    offer to install them via pip (checking what's already installed first
    to avoid conflicts). Returns True if all deps are available."""
    missing = []
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing.append((module_name, pip_name))

    if not missing:
        return True

    # Build a readable list of what's missing
    missing_names = [f"  - {pip_name}" for _, pip_name in missing]
    msg = ("The following packages are required but not installed:\n\n"
           + "\n".join(missing_names)
           + "\n\nWould you like to install them now via pip?")

    root = tk.Tk()
    root.withdraw()
    result = messagebox.askyesno("Missing Dependencies", msg)
    if not result:
        root.destroy()
        return False

    # Build pip command
    pip_args = [sys.executable, "-m", "pip", "install"]
    for _, pip_name in missing:
        pip_args.append(pip_name)

    # Show a progress dialog while pip runs in a background thread
    prog = tk.Toplevel(root)
    prog.title("Installing Dependencies")
    prog.configure(bg="#1a3a5c")
    prog.resizable(False, False)
    prog.transient(root)
    try:
        prog.grab_set()
    except Exception:
        pass
    pw, ph = 380, 120
    sw, sh = prog.winfo_screenwidth(), prog.winfo_screenheight()
    prog.geometry(f"{pw}x{ph}+{(sw-pw)//2}+{(sh-ph)//2}")
    try:
        prog.attributes("-topmost", True)
    except Exception:
        pass

    tk.Label(prog, text="Installing packages via pip...",
             bg="#1a3a5c", fg="#ffffff",
             font=("Segoe UI", 12, "bold")).pack(pady=(18, 8))

    pkg_list = ", ".join(p for _, p in missing)
    tk.Label(prog, text=pkg_list,
             bg="#1a3a5c", fg="#aabbcc",
             font=("Segoe UI", 9)).pack(pady=(0, 8))

    # Indeterminate progress bar — pulses while pip runs
    bar = ttk.Progressbar(prog, mode="indeterminate", length=320)
    bar.pack(pady=(0, 14))
    bar.start(15)

    install_result = {"ok": False, "stderr": ""}

    def _worker():
        try:
            r = subprocess.run(pip_args, capture_output=True, text=True)
            install_result["ok"] = (r.returncode == 0)
            install_result["stderr"] = r.stderr
        except Exception as e:
            install_result["ok"] = False
            install_result["stderr"] = str(e)
        try:
            prog.after(0, prog.destroy)
        except Exception:
            pass

    threading.Thread(target=_worker, daemon=True).start()

    # Modal wait — blocks until the worker destroys the dialog
    try:
        prog.wait_window(prog)
    except Exception:
        pass

    bar.stop()

    if not install_result["ok"]:
        messagebox.showerror(
            "Installation Failed",
            f"Could not install packages:\n\n{install_result['stderr'][:500]}")
        root.destroy()
        return False

    root.destroy()

    # Verify all packages are now available
    for module_name, _ in missing:
        try:
            importlib.import_module(module_name)
        except ImportError:
            messagebox.showerror(
                "Still Missing",
                f"Package '{module_name}' still could not be imported\n"
                f"after installation. Please install it manually.")
            return False

    return True

# Run the check before importing the packages
if not _check_and_install_dependencies():
    sys.exit(1)

# Now safe to import
import customtkinter as ctk
from PIL import Image, ImageOps

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import openpyxl

# fpdf2 for PDF export
try:
    from fpdf import FPDF
    _FPDF_AVAILABLE = True
except ImportError:
    _FPDF_AVAILABLE = False

# ------------------------------------------------------------------
# Shared resources
# ------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent


def get_platform_font():
    system = platform.system()
    if system == "Windows":
        return "Segoe UI"
    elif system == "Darwin":
        return "SF Pro Display"
    return "Arial"


MODERN_FONT = get_platform_font()

# ==============================================================================
# EMBEDDED MBE LOGO (base64 PNG - no external file needed)
# ==============================================================================
MBE_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAABLgAAALaCAYAAAAySeo9AAAAAXNSR0IArs4c6QAAAARnQU1BAACx"
    "jwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAFOQSURBVHhe7d0L0CVVYSfwM+/3A7emEt3ZBYpF"
    "sTABMcTJIhlZUSDBAAsuJCADArsYLdlCYyglOxg1SgSzlo9VQMH3A1gkgI+IATQmUCzMsobFlaBO"
    "YGVZNg6Iw4wyw+yc5tyx+eZ+39xH9719+v5+VV1fn7739u139/1/p08HAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADIxK/0FoOF27JR6ASB7s3ZKvQAwNCcVgEwIuABokzoD"
    "runOmUI1gPZygAfIhIALgDapOmzq9zwp7AJoFwd1gEwIuABok6oCpmHPj4IugHZwMAfIhIALgDap"
    "Ilia7tw43bj7fT8A+XAgB8jEdBflAJCjYUOlqefFfsbX7Zw67PQAMF6z018AAIAsDBNuRfH9USoW"
    "uoVeAOTDfykAMuHCG4A2mRow9Wrq+bAznn7Ok+Xvnm58AORFDS4AACBLg4ZR5VBLoAXQDgIuAAAg"
    "C70GU/G1sjT4WcrjKptuOADNJuACAABaLeVchTToWaYbDkA+BFwAAEBWBFIATCXgAgAAGq/uWweF"
    "ZgB5E3ABAAATodeQrO4wDYDqCbgAAIBWiQFVN+nlghpbAO0i4AIAACaKcAugfQRcAABAq8QAqyMN"
    "KqRBwi2AFhJwAQAArVUOtKbepghAewi4AACAiTFdyFUerpYXQH4EXAAAQOMNUxNLYAXQfgIuAABg"
    "ovQbkAHQfAIuAAAgC1XW4ip/vtw/9X0A5EHABQAAZKkcTPWiW3jV7zgAaCb/nQDIhAtwANpkmJpS"
    "U8+Jg46rqvEAMH5qcAEAAFmZGkRNDap6IdwCaBcHcYBMDHLxDgBNVUWg1O3cuKfxDvIZAJrPgRwg"
    "E90uyAEgV1WFSsOeH4VbAO3gYA6QCQEXAG1SdbDU73lSsAXQLg7qAJkQcAHQJnUFTDOdL4VaAO3l"
    "AA+QCQEXAG0ibAKgSp6iCAAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAA"
    "ZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwA"
    "AAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDW"
    "BFwAAAAAZG1W+gtAw+3YKfVCz+68887wJ3/yJ+Ghhx4Kc+fODStXrgyzZ88Oy5cvDwsWLAiLFi0K"
    "S5cuLV5btmxZmDdv3rNei13sj6/F96xYsaJ4z5IlS3b7PEA/Zu2UegFgaE4qAJkQcNGPe++9twi2"
    "rrvuujSkfjH86oRn8+fPD4sXL+4annVeW7hwYdGVw7M5c+YU5U54FvvjsBjMdV7rfA7Im4ALgCo5"
    "qQBkQsBFLzZu3Bguuuii8JnPfCZs27YtDW2nGIDFsKxbeBa72B+HxWCsHJ51aqXFWmjxc+XwrBOs"
    "dV6Ln4/hW/wLVEvABUCVnFQAMiHgYiaPPvpoePe73x0+/OEPtz7YGpcYdMXgqxOedWqnxYCsc9tm"
    "Jzwrvzb1ltBOeBYDtXJAV/58HBZDN2gzARcAVXJSAciEgItufvrTn4ZLL7206DZv3pyG0hax5tnU"
    "Wmn9hmedz5fDs/j5OO74N47PbZ+Mg4ALgCo5qQBkQsBF2datW4vaWhdffHFRewuqEAOyqeHZ1IAt"
    "Duvc9jlTeFaulRZDtAMPPDB9CzxDwAVAlZxUADIh4CKKtx9eeeWV4U//9E+LJyNCLo444ojwwQ9+"
    "UNDFLgIuAKrkpAKQCQEXX/3qV8Mf/dEfFU9IhBzF2l1vetObwvr16zXcj4ALgEo5qQBkQsA1ub77"
    "3e+G888/P9x8881pCORt1apVRbtxp512Wgw50lAmjYALgCrNTn8BgIZ5+OGHw9lnnx0OOeQQ4Rat"
    "EtuNO/3008Phhx8eNmzYkIYCAAzOf00AMqEG1+T4+c9/Hi655JLwnve8x5MRGUq8JTA29B51GoSP"
    "Og3GR52nMkaxcfmo08B8FF+L74k6T2uMOg3LR52nN0axgflYMac8jtgofWyAPio/sTGOI05L5zUm"
    "ixpcAFTJSQUgEwKuyRDb2TrvvPPC/fffn4bQFJ0nDEadICjq9JdDnM4TBKNyEFQOk+J742fib/wY"
    "CkXlcZQDqXKwVA6TOuOIuk0HNJmAC4AqOakAZELA1W7xdsQ3vOEN4brrrktD2qvfWkXdQpxyraLy"
    "OMphUnl85XF0ahiVg6ByIFUOk8rjAKol4AKgSk4qAJkQcLVTXK1XXHFF8XTExx9/PA3tXQxiuoU4"
    "e6pVVA5xyqFQtyAodt1qGE03jnINo3iLWrxVrTwOgEjABUCVnFQAMiHgaqdYc+uOO+7YYzhVrrFU"
    "bsMIIFcCLgCq5KQCkAkBFwBtIuACoEqz018AAAAAyJKACwAAAICsCbgAAAAAyJqACwAAAICsCbgA"
    "AAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICs"
    "CbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAA"
    "AICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICszUp/AWi4HTulXir05JNPhle96lXhBz/4"
    "QZg/f36YPXt2WL58efHaokWLwoIFC541bMmSJWHevHlh7ty5YenSpcWwZcuWhTlz5hTD4+vlYXGc"
    "ixcvLoatWLEizJo1a9d4o5UrVxZ/O8Pi6/F9AG2383jntwgAlXFSAciEgKt627dvD8cff3y48cYb"
    "05BmiaHX1NAthmVxWAzPYogW9RuwxSAtjjuaKWDrhHndhgEMS8AFQJWcVAAyIeCq3utf//rw0Y9+"
    "NJXoR6zBNlOYFmu3xfeUQ7d+A7bOsG416MoBX2dY+buA5hNwAVAlJxWATAi4qnXxxReHCy64IJVo"
    "mxh0xYCtHLp1wrRuw7oFbAsXLiy6XgO28rBO6FcO8+I44riAZwi4AKiSkwpAJgRc1fniF78YTjnl"
    "lFSC0SrXfusEbJ1bT6Opw8oBWyeI6xamlWuwzTSsE9xFnVp15WEwKgIuAKrkpAKQCQFXNe64445w"
    "xBFHhC1btqQhQFmnBttMAVu3Wm3DPHihXNNtplp1tIuAC4AqOakAZELANbyNGzeGQw89NDz66KNp"
    "CJCTGHpNDd06QVwMz2KIFlURsHWGlQO2OK74+qpVq4oywxFwAVAlJxWATAi4hvPTn/40/Ot//a/D"
    "vffem4YA9C+GXjfccEN4xStekYYwKAEXAFXS0ikArbd9+/aizS3hFjCseHvziSeeWNzuDAA0h4AL"
    "gNZ761vfGr761a+mEsBwHn/88XDUUUcJuQCgQVQLBsiEWxQH8+lPfzqcfvrpqQRQnb322ivcdttt"
    "4dd+7dfSEPrhFkUAquSkApAJAVf/7rrrrnD44Yd7YiJQm9jg/De/+U0h1wAEXABUyUkFIBMCrv48"
    "8sgj4Td+4zfCQw89lIYAues8RTHqPD1xan98T3wCYrRw4cKii8pPSiz3l5+qWH7qYrm//ITGbv3x"
    "yYr77LNPMYzeCbgAqJKTCkAmBFy9e+qpp8IRRxwRvvOd76QhwNy5c3cFNuVQp9xfDm9mz54dli9f"
    "vlt/zCRWrFixW38Ug56OOLyTX5T74/jj98zUH6czBkxT+2mXnduE3yIAVMZJBSATAq7evfnNbw7v"
    "f//7UwmqMV2QM13/dIHN0qVLi7ApKg8v1zyarkZSueZRefh0NZXK4RU0jYALgCo5qQBkQsDVm2uv"
    "vTacdNJJqURTxMClE9iUw5vpgpxeAptyf/l2snJ/uUZSL/3x9/Z0QRZQLQEXAFVyUgHIhIBrzzZu"
    "3BgOOuig4hH+bTLdrWXl28bKIU0vt5ZF09U86vd2snJ/uUZSuR9gKgEXAFVyUgHIhIBrZtu2bQsv"
    "f/nLu7a71UuoU75tbLrApqpby8r909VUKtdCAmgjARcAVXJSAciEgGtmsWH5zZs3p5JbywCaTsAF"
    "QJWcVAAyIeACoE0EXABUaXb6CwAAAABZEnABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3AB"
    "AAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZ"
    "E3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAA"
    "AABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkLVZ6S8ADbdjp9TbSieffHK44447wvz588PixYuLYXPm"
    "zAnLli0r+qOVK1emvhCWL18eZs9+5v80S5cuDXPnzi36Fy5cWHTRvHnzwpIlS3brn2m8cXh8PYrv"
    "j5+LFi1aFBYsWFD0l6cRgMHM2in1AsDQnFQAMtHmgOuuu+4Kv/Ebv5FKeYnBWic4i4FbDN46ysFZ"
    "OYSL4VgMyaLpgrNyCDfTeHsJ5MrhXpyGOC3R1PECjJKAC4AqOakAZKLNAddrXvOacM0116QS47Ji"
    "xYrU9+zgbLpAbroacvE3a3lcsb/zO7aXQC7+jeWoHPSV+6PpxgvkQcAFQJWcVAAy0daA64EHHgj/"
    "6l/9q1SC4cVwrBPIlcO5XmvIVRXIDTteaDsBFwBVclIByERbA67/8B/+Q7jssstSCego3/5aDueG"
    "vWW1PK5yzbtyOBf1EsiV27wrB33lW2FhOgIuAKrkpAKQiTYGXA8//HDYb7/9wpYtW9IQoI1iWNYx"
    "Xdt0vTwsohycDRLIlUO4cn/5O+I4ytNbHi/VEnABUCUnFYBMtDHguuCCC8LFF1+cSgDNFkOwTghX"
    "vv11ulthy8HZILesdkK4tWvXhn333bcY1iYCLgCq5KQCkIm2BVyPPfZY2GeffcLjjz+ehgDQzWtf"
    "+9rwqU99KpXaQ8AFQJVmp78AMFIf+9jHhFsAPfj85z8fHnnkkVQCALoRcAEwctu2bQsf+tCHUgmA"
    "mcRj5kc/+tFUAgC6EXABMHLXX399eOihh1IJgD2JtV5/8YtfpBIAMJX73gEy0aY2uF7+8peH2267"
    "LZUA6hUbau80Aj+1sffyUx3j+2LD7tHU98X+OCwqv2/q0xzLDcaXG5+f6X3l8ZUbpo/K31vubwNt"
    "cAFQJScVgEy0JeC65557wsEHH5xKQNPEp/t1lJ8IOFMwtHDhwqKLpr6vHOT0+r5eA6SZ3te2MKiN"
    "BFwAVMlJBSATbQm4Xv/612tLhlaYO3fus2ralIOXQWruxPEtXbq06I96DXLK74vvie+NBn0fjIqA"
    "C4AqOakAZKINAdfWrVvDr/7qr3p64gQqh0FTa+7EUCe+HpWDnKnvK9fIqTrwmWl85e8t10AChiPg"
    "AqBKTioAmWhDwPXZz342nHbaaanEdGIQ1Al8ZrpFrBwM9XrrVznwmel9M31v+X3l742/Vcu3t5Xf"
    "BzCVgAuAKjmpAGSiDQHXv/k3/ybccsstqVSN6QKVckATldsLmi7wmSmg6bUm0CDvmxogAUwCARcA"
    "VXJSAchE7gHX008/XdTgKgc+saZSfGJYNFPQVH5fucYQAPkScAFQJScVgEy0oQYXAHQIuACokmcn"
    "AwAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAA"
    "WRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcA"
    "AAAAWRNwAQAAAJC1WekvAA23Y6fU23h33XVX+MAHPhBmz54dZs2aFVasWFEMnzNnTli2bFnRv2jR"
    "orBgwYKif8mSJWHevHlFf3w9vi9+dvny5cWw+Fp8T7Rw4cKii8qfi98RvwuAPOw8ZjtoA1AZJxWA"
    "TOQUcL31rW8N73vf+1Jp9DqB2n777VeEbdM59dRTw0033ZRK4zN//vzw/e9/P6xcuTINyVNc5+9+"
    "97tTabxuuOGGcPjhh6dSCPvss0947LHHUqnZ4vYQt4XFixeHVatWhec+97lh7733DgcccEA48MAD"
    "w4te9KIwd+7c9O7qPf7448X38UsbN27cdVyhOgIuAACYQDHgysULX/jCGMaNvdv5Iz1NUXfHHXdc"
    "18+No/vQhz6UpipPTz/99I59992367yNo7vlllvSlD1jxYoVXd+XY7dkyZIdRx111I5LLrlkx49/"
    "/OM0h9XZtGlT1++d5C4uE6q3c9kCQGW0wQVApX74wx+G++67L5Xo1eWXX5768vTXf/3Xxbqnfps3"
    "bw5f//rXw1ve8pbwvOc9Lxx99NHhG9/4RnoVAGAyCbgAqJQf2oO55557ZrydsumuuOKK1MeoxbDr"
    "Va96Vfit3/qtrLchAIBhCLgAqNTf/u3fpj76lWtItGnTpnDNNdekEuNy++23hzVr1oQ3v/nNYevW"
    "rWkoAMBkEHABUCkB1+A+//nPF7ef5eaTn/xk2LZtWyoxTnE9vP/97w+//du/HR5++OE0FACg/QRc"
    "AFTm0UcfDffff38q0a/49Lqrr746lfLh9sTmufPOO4tbFh944IE0BACg3TyaFyATOTxx6qabbgrH"
    "HntsKo3f3nvvHX70ox+l0u6OP/74cP3116dSMxx22GHhb/7mb1Kp+e64447itrimueWWW8LLX/7y"
    "VAph5cqVRYA4aVavXh3+7u/+rvjbq8ceeyzstddeqUQUb8ON29BUP/3pT8Ov//qvF/2LFy8O8+fP"
    "L/qXLl0a5s6dW/QvW7YszJkzp+hfsWJFmDVrVtHF/ii+Ft8TzZ49OyxfvrzonzdvXliyZEnRH8cb"
    "xx8tWLAgLFq0qOiPf2M5iu+Nn4n+2T/7Z7vG2WQ7l4PfIgAAMGmKZ6o33J/92Z8969H64+723nvv"
    "NGXdHXfccV0/N+7uvvvuS1PYfGeddVbXeRh3d8stt6QpfMaKFSu6vm8SuoMOOmjHli1b0pLYs02b"
    "NnUdzyR3cZl009Rldfrpp6cpbLad0woAlfFfE4BM5PBj4Pd///fDF77whVQavxxrcEXnn39+uPTS"
    "S1OpuZ544onw3Oc+t5HthqnB9WznnHNOuOyyy1JpZmpw7W66Glx1LqtYIyvWAovf26nZ1anttXDh"
    "wqLr1ODq1Bjr1Bb7tV/7tXDiiSemMTWXGlwAVMlJBSATOQRcL3rRi8K9996bSuOXa8C1atWq8NBD"
    "D+263ampLr/88vDv//2/T6VmEXDt7q/+6q/CK1/5ylSanoBrd9MFXDHcPfXUU3cFUDFkimHT1ACq"
    "EzzF2xLj33JY1bkVsXPrYflWw7YTcAEAwASKAVeT/fznP9+x88fdrltkmtDleoti7L70pS+lqWyu"
    "Qw89tOu0N6Fzi+LuXbxVcfv27WmJTM8tirt3092iyHB2LlsAqIynKAJQiR/84Adh27ZtqcSwer2d"
    "bFy++93vFk/qIx/33HNP+PznP59KAADtIuACoBL3339/6qMKN998c9i4cWMqNU/TAzi6+9CHPpT6"
    "AADaxX3vAJlo+u0csVH0t7zlLanUDHtqg+uLX/xiuO+++1Kpu7jYr7766j2+rw4XXnhheOc735lK"
    "zbF169bwvOc9r2iXaNTOPffc8Cu/8iupNL0zzjgj7LPPPqkUwsUXXxy2bNmSSt3F+YrvG9ZBBx1U"
    "tO/Wq7iNxfbBfvaznxV/Y1gc27Krq0bk//gf/6NohHw6cTm8973vTaXpVbW8cjBdG1wMRxtcAFTJ"
    "SQUgE00PuN7znveEa665pmg0uazTuHJHp9Hljk6jzGVTf0hOHUenMeaOTmPNZbEcx/vSl740DRnc"
    "uBqjX716dRHQlee9CT772c+G0047LZVGa8OGDeHggw9OpWpV1bj6unXrwlVXXZVKg3nyySeLWnyx"
    "If8bb7wxDa3Gf/pP/ym84x3vSKXBVbW81q5d+6wHAlTp1ltvDbfddlsqDU7AVQ8BFwAATKAYcDEe"
    "42yM/sYbb0xT0Rxr167tOq2j6DZs2JCmonpVNa6+bt26NMZqfOtb39qx//77d/2uQbrDDjssjXk4"
    "VS2v9evXpzFWL46723f228V5pXo7ly0AVEYbXADQYB//+MdTXzM88MADldSIoXeHH3540aB/rOlU"
    "hTvuuKO4vRAAoE0EXADQYDfccEN4+OGHU2n84i1zjN6KFSuKWxUPPPDANGRwsW0vD4UAANpGwAUA"
    "DRbDiE996lOpNF5PPfVU+MQnPpFKjFpsUy62fxbbrRvW97///dQHANAOAi4AaLh4m2ITmqu56aab"
    "wqOPPppKjEN8QmMVDfw/+OCDqQ8AoB0EXADQcPF2sia0e9W09sAm1bnnnpv6Brd58+bUBwDQDgIu"
    "AMjAuNu+ijV+vva1r6US4/Sbv/mbYdWqVak0GI3MAwBtI+ACgBEYtt2k6667LvzkJz9JpdG78sor"
    "i/bABlVFu1E8Y9asWcWtisOYM2dO6gMAaAcBFwCMwJIlS8LatWtTqX9btmwpGhgfh6effnroxuWP"
    "Pvro1EcV/vk//+epbzCLFi1KfQAA7SDgAoAROeuss1LfYMbVBtY3v/nNsHHjxlTqX7yd7nd+53dS"
    "iSoMWyPuec97XuoDAGgHARcAjMiJJ54YVqxYkUr9u+eee8Idd9yRSqNz2WWXpb7BnHzyyWoMVeyJ"
    "J55IfYN5/vOfn/oAANphVvoLQMPt2Cn1MmLHH398uP7661NpMDHYeuyxx8Ib3/jG8OEPfzgN7d85"
    "55wzdODUj0cffbSo7TNM+1t33313Ec6deeaZacjgNmzYEA4++OBUqlZcP3vttVcqDW7dunXhqquu"
    "SqV6xGUQl+kgYu2vTZs2haVLl6Yhg6lqea1fvz5cdNFFqVSt733ve0U3rN/93d8N8+bNSyWqMis2"
    "KAcAAEyWGHAxHscdd1wMF4fqVqxYUYzrv/23/9b19V67JUuW7PjZz35WjGsULrnkkq7T0Wt30EEH"
    "FeO58soru77eb7dhw4ZifHXYtGlT1+/st1u3bl0aYz3idM6dO7frd/fSHXLIIWlMw6lqea1fvz6N"
    "kUmzc/0DQGXcoggAI/SSl7wkHHLIIanUv82bN4cvfOELqVS/Ydv9GrbdMXZ37bXXDlWj7tWvfnXq"
    "AwBoDwEXAIzY2WefnfoGc8UVV6S+en37298O9913Xyr1L94Kd8opp6QSVYjB1qWXXppKg7FOAIA2"
    "EnABwIj9/u///lCNrt9+++3h3nvvTaX6DFt7K7ZdFp+gSHViuDVM6HjYYYeFAw44IJUAANpDwAUA"
    "I7Zy5cpw0kknpdJgLr/88tRXj5/+9KfhS1/6UioN5owzzkh9VOGmm24KF154YSoN5q1vfWvqAwBo"
    "FwEXAIxBfBriMD73uc+FrVu3plL1PvvZz4YtW7akUv+e+9znhqOPPjqVGFasTRdrxA3T9lasvaX9"
    "LQCgrQRcADAGL3vZy8L++++fSv179NFHw/XXX59K1Ru2na/TTjstzJkzJ5UY1F133RVe+cpXFu22"
    "DRNuxfbQPvShD4VZs2alIQAA7eIqByATTX2k+o9+9KNw7rnnhoULFxbl2LbUggULiv5ly5YVIUf8"
    "Ub1ixYpi2Lx588KSJUuK/vi+TltUnfdG8b3xM7Ech0flz8W/sRyVP1eXWHNm2DApztNjjz2WSs+4"
    "+OKLwwUXXJBK/TviiCPCX//1X6dSdTZs2DDUkx6j2E5Uua2nq666Kpx55pmpNLg4bQcffHAqVSuu"
    "n7322iuVBrdu3bpifvv11FNPhUceeaRYdn/3d38XbrzxxnDnnXemV4fzrne9K7z97W9PpWpUtbzW"
    "r18fLrroolRikuw8zvstAgAAkyYGXE30t3/7tzF4a0S3YsWKolu1atWOvffee8dhhx2WpnI4xx13"
    "XNfv66eL0zXVj3/84x1z587t+v5eu3/4h39IY6vOG97whq7f1Wu3Zs2aNKZfuvLKK7u+t99uw4YN"
    "aYzV27RpU9fv7LeL67SzLfbaLVmypOu4quiOPfbYHdu3b09zWZ2qltf69evTGJk0O9c/AFTGLYoA"
    "DOUnP/lJ6hu/xx9/vOji7XsbN24MDz30UHqlmWI7VcO2iTRITaGZPPnkk+Ezn/lMKg2mippaOYu3"
    "Ena2xV67zZs3p09Xa82aNeGLX/ximD3bJR8A0G6udgAYSpMCrhydddZZqW8wsfHxYdpmmuraa68t"
    "ApdBxVtOTz755FRinI488shw8803h8WLF6chzfSOd7yjuCV5FN2Xv/zl9K0AQNsIuAAYyj/90z+l"
    "PgYRnzS4evXqVOrfww8/HL761a+m0vA+9rGPpb7BnHDCCbvaW2N83vSmN4WvfOUru9qtAwBoOwEX"
    "AEOZ2nA6/YkN5J9xxhmpNJjLLrss9Q3nf/2v/xW+853vpNJgXve616U+xmHvvfcugq0PfOADux7E"
    "AAAwCQRcAAxFwDW8YUOhr33ta0VNrmHF2x2HEcOV+GRHRi/Wmnvve98bvve974VjjjkmDQUAmBwC"
    "LgCG8vOf/zz1Mah99923aC9pULENrk984hOpNJhf/OIXQzdYv27dOo2Zj0lsN+2WW24J3/72t9MQ"
    "AIDJMiv9BaDhmvpI9Xh73Sc/+clUapZYo+hHP/pRKg3u+OOPD9dff30qDSbWsJmptlt80t0pp5yS"
    "Sv2LIdkDDzxQNKQ9iNi4/EknnZRKg/nBD35QTEc3MTyr4umKGzZsCAcffHAqVSuun7322iuV8rV2"
    "7dpw8cUXh5e+9KVpSD1yXF7XXXddsT9X5YknnnjWk1DjbaHlds+WL1++K/SdO3duWLp0adEfxfd1"
    "biON74nv7YgPBpg/f34qhbBy5crUF8KCBQuKhzl0xGNLZ7/v9v1xPOXvbZKd0+23CAAATJoYcDXR"
    "unXrYvDWyG7vvfdOUzmc4447ruv4++l2/ghNY+tuy5YtO1atWtX1s712N998cxpb/4466qiu4+y1"
    "O+KII9KYurvyyiu7fq7fbsOGDWmM1du0aVPX78y1O/fcc3c88cQTae6ql+Pyuu6669LUVyOHZbB2"
    "7do0tc2zc/oAoDLuIwCABli4cGE49dRTU2kwV1xxRerrz8aNG8PXv/71VBpMFbWzqNZHP/rRcMgh"
    "h4S77rorDWESxdpgADAJVAsGyERT/9v9mte8JlxzzTWp1Cw53aIY3XvvveFFL3pRKvUv3gL1yCOP"
    "hOc85zlpSG8uuuii8I53vCOV+hfn7X//7//9rFujpnKL4vjE29muvPLKcPLJJ6ch1XCL4jNtn8Xj"
    "TPmWws7thLHcCZeWLVtWPDG1fCtiXC/xdsN4l17ch8rDos5nyrcddm5rjMPj61H5M92+u8ncoggA"
    "ABOouJ+jgaq4fa+uLqdbFDvWrFnT9fO9dn/xF3+RxtSb7du371i9enXXcfXanXXWWWls03OL4vi7"
    "j3zkI2lOq+EWRYa1c50AQGXcoggADXL22WenvsH0e5tivDXxoYceSqXBuD0xD3/4h3/Y2AdCAAAM"
    "S8AFwFA6TwijGvE2splu9duTeJvjHXfckUp7dvnll6e+wey///7hsMMOSyWaLgaot9xySyoBALSH"
    "XyUADKX8aHuGFx/nf8opp6TSYHoNrWJ7XTfccEMqDeass85KfXSsW7euCJH21P3VX/1V0SZUp4vt"
    "lF1yySXhvPPOC0ceeeSudpmqtG3btiJEffjhh9MQAIB20LAjQCaa2l7JG9/4xvDhD384lcYr1nyK"
    "Da137LfffpU8QW5Ujcx3xBpYa9asSaX+xeUQA4xOI9TTufjii8MFF1yQSv2Ly/qHP/xhWL16dRoy"
    "vUlqZH79+vVFw/3D2r59e7j99tvDpz71qfDpT386bNmyJb0yvBNOOCH81//6X1NpMFUtr6OOOioc"
    "ffTRqVSvV7/61cVxgWbQyDwAVXJSAchEUwOuWBPlnnvuSaXiB8tuNU86TwPriE/8ik/+6ig/Eaxj"
    "T58pP42sbqMOuKL4NMV4u+GgLrvssnDOOeek0u7i5vSCF7wg3H///WlI/4455pjwla98JZVmJuAa"
    "Tgws3/KWt4TPfe5zacjwbrzxxvC7v/u7qdS/Ji8v8iDgAgCACRQDLsZjlE9R7IhPQ+w2nl67Qw89"
    "NI2pu1tvvbXr5/rpvvSlL6Wx7dkkPUVx/fr1aYzVi8tx7ty5Xb+33+7AAw8snqI5qByWF822c/0D"
    "QGW0wQUADXT66ac/63bLft15553hu9/9birt7uMf/3jqG0ysufN7v/d7qcSonHHGGeHaa68datvo"
    "iDUEY9tfAABtIOACgAZ6znOeE0466aRUGky8TbGbeGvZNddck0qDOfXUU4vbRhm9GCy+973vTaXh"
    "fPCDH0x9AAB5E3ABQEOdffbZqW8wn/3sZ8PWrVtT6ZeqaLA81iRifM4///xw2GGHpdLgbrvttrBx"
    "48ZUAgDIl4ALABrqiCOOCPvuu28q9W/Tpk3F7WxTDXt74kEHHRRe8pKXpBLjENvmfs973pNKwxn2"
    "aYoAAE0g4AKAhopPihy2ptTll1+e+p5x1113Peupl4NQe6sZDj/88HDggQem0uC+8Y1vpD4AgHwJ"
    "uACgwV73utcN1aB4vAXtgQceSKXp2+XqVZyW2P4WzXDCCSekvsF961vfik9pTSUAgDwJuACgwVav"
    "Xh2OPvroVBpMpxbXz372s/DFL36x6B/Uq1/96rBq1apUYtxe9rKXpb7Bbd68OfzgBz9IJQCAPAm4"
    "AKDhzjrrrNQ3mE984hNh+/btRbj1+OOPp6GDOfPMM1MfTXDAAQekvuF8//vfT30AAHkScAFAwx17"
    "7LFD1Zp69NFHww033BCuuOKKNGQwcRqOOeaYVKIJVq5cmfqG8+CDD6Y+AIA8CbgAoOFiu1fD1py6"
    "8MILw+23355Kg4mNyw/THhjVq2p9xBC0jS666KLiiZPDdo899lgaIwDQVAIuAMjA2WefnfoGc++9"
    "96a+wa1bty710RQ/+clPUt9wnnrqqdQHAJAnARcAZGD//fcPa9euTaXRW7NmTTjwwANTiab4n//z"
    "f6Y+AIDJJuACgEwM29j8ME4//fTUR5N861vfSn3DWbRoUeoDAMiTgAsAMvGa17wmrFixIpVGJ4Yf"
    "f/AHf5BKNMWOHTvC1VdfnUrDWb58eeoDAMiTgAsAMrFw4cJw2mmnpdLoHHfccWMJ1pjZV77ylXD/"
    "/fen0nD+5b/8l6kPACBPAi4AyMg4blN83etel/poim3btoU//uM/TqXhHXDAAakPACBPAi4AyMiL"
    "X/zicMghh6RS/VavXh1e8YpXpBJN8ba3va2SJ2NGS5YsCfvuu28qAQDkScAFQGW+/OUvh6uuuqro"
    "rr322qIcu1tuuSXceuutRbdhw4bw3//7fy+6H/3oR0W3cePG8NhjjxXd448/nsbGdM4+++zUV78z"
    "zjgjzJ7tcqFJ/st/+S/hfe97XyoN77d/+7etYwAAAEZjRwaOPPLIHXFSq+xWrFhRdHvttdeOvffe"
    "u+j23XffHQcddFDRHXLIITvWrl1bdPH7jzvuuKI76aSTdpx//vlpyoYTx9dt2vrp4jxU5bHHHtux"
    "aNGirt9TdfcP//AP6VuHc+WVV3Ydf7/dhg0b0hirt2nTpq7f2W+3fv36NMZqPfXUUzv++I//uOt3"
    "DtO9//3vT9/Qn6YvryiOu9t39tvFeaV6O5ctAFTGv+sAqEwd7fjEGl2x2/kDs6jpFbsf/vCH4Z57"
    "7im6u+++O9x2221Fd/PNN4frr7++6K655pqiFlkbxQbf4xMV67Z27dqw3377pRLj8vTTT4ebbrop"
    "HHzwweHiiy9OQ6vze7/3e6kPACBfAi4AKvP85z8/9VG3UdymeOaZZ6Y+RikGWg888EC47rrrwn/8"
    "j/8x7L333uHYY4+trM2tMiEmANAWs9JfABouh9s5vvGNb4RXvepVqTR+MRiIbXxN59FHHw2bN29O"
    "pemdc845Re2wYcRaV7H9sVmz9nzqXb58eXjOc56TSt3FzeEFL3hBuP/++9OQasWGxx955JHi70y2"
    "bt0a/s//+T+pNL1Ym+4tb3lLKg3uK1/5SnjhC1+YStP71V/91bBw4cJUCkXNvz3tQk888UT49V//"
    "9VQaXNzu9tlnn1TqTWx/7sknnwz/7//9v2I64lMSR+Ezn/lMOPXUU1PpGTFg+8d//MdUmt44l1ev"
    "Om38DSvWIF25cmUqUZWdx0O/RQAAYNLEgKvp/umf/mlXmzVN6GJ7XTOpom2tOrrzzjsvTeHM/vzP"
    "/7zr56vozjrrrPQtM7vlllu6fn7cXZyustgGWrf3TXK3//77F+16TVVV21pt6rTBVY+dyxYAKuMW"
    "RQAqE2sd7fzRnErU7fTTTw9z585NpWqtW7cu9dFWb3vb22rbfgAARk3ABUClDj300NRH3X7lV34l"
    "vPrVr06l6sSQ8mUve1kq0UZr1qwpAlIAgLZw3ztAJnK5neODH/xgeNOb3pRK47WnNriOP/744omL"
    "TXPeeeeF//yf/3MqzeyrX/1q+J3f+Z1Uqsa73vWu8Pa3vz2VZnbrrbeGI444IpWa45Zbbgkvf/nL"
    "UykU7SfFp3ESilpb3/3ud6d96mlsD2yvvfZKJaLp2uCK21Q8zkSzZ88u2s+b2j9nzpywbNmyon/e"
    "vHm72rUr9y9YsCAsWrSo6I9tx3Xaj4vD4mvR4sWLw/z584v++Ln4+an98Xv23XffPbbh1xTa4AIA"
    "gAlUNFiSgb//+79/Vts14+za3gZXtG3bth2rV6/uOp5Burlz5+548MEH09j3TBtc+XWXXHJJWird"
    "aYNr9266NriauKz+5m/+Jk1d8+2cXgCojFsUAajUgQceGJ773OemEnWLtUPOOOOMVBreK17xirB6"
    "9epUom1OOeWUcP7556cSOYo18OJTWWO3atWqogZZ7OLTReOtp7EDgEmkWjBAJnL6b/frX//68NGP"
    "fjSVxmcSblGM4jzG25Kq8IUvfCGcfPLJqbRnblHMR1xPX/va13bd5jYdtyjubrpbFLdu3RouuOCC"
    "oj8u13gbYVS+5bDXWwtjWB11vifevRdDrKh8m2ObuEURAAAmUHE/RyZuvvnmZ90yM65uEm5R7Djy"
    "yCO7jqufLt7Gt/MHexpjb9yimEd31FFH7di8eXNaGjNzi+Lu3XS3KDKcncsWACrjFkUAKhdrzrhN"
    "cbTOPvvs1De40047bVeD1rTHOeecE/7yL/9yV+0iAIA2EnABULl4O81rX/vaVGIU4u2WsT2eYZx5"
    "5pmpjzaIt8BdccUV4bLLLtvjbYkAALkTcAFQiypqFNG7WPNqmFDxoIMOCi95yUtSidwde+yx4e//"
    "/u/DWWedlYYAALSbgAuAWuy///7Fj+xRi7VWOk8Y69YodJu97nWvS339U+OuHY466qiigf0bbrgh"
    "7LPPPmkoAED7eXIJQCZybJD33nvvDbfffvuup4NF5aeFRf2Wly9fHmbP/uX/Z6aW+9GWpyiW/dZv"
    "/VaxzPsxd+7c8OMf/3igWxw9RXH8Ypgct+XY1lbsH5anKO5uuqcoMhxPUQQAgAlUPHKKSrXpKYod"
    "V1xxRddxztSdcMIJ6dP98xTF0XarVq3asWbNmh3nnnvujssuu2zH97///TSH1fEUxd07T1Gsx85l"
    "CwCV8V8TgEz4MVC9b37zm+HBBx9MpeZ44QtfGF760pemUn82b94crr766lTqzZo1a8IBBxyQSv15"
    "+OGHw9e//vVUao54q175SZ6f+9znwi9+8YtUaq5YoSXeXtsR+2Mtxdj9i3/xL8LChQvTK/WJyyku"
    "L37pD/7gDzTUXwM1uACokpMKQCYEXAC0iYALgCppZB4AAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4"
    "AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMjarPQXgIbbsVPqzd7TTz8d/vEf/zHM"
    "mjUrrFixIg0NRX8cBkD77TzeO+ADUBknFYBMtCngip7//OeH+++/P5V2t2TJkjB37tyif/HixWH+"
    "/Pm79ce/sRzNmzev+Ew0e/bssHz58qI/Kgdny5YtC3PmzCn64/vj56Jexhs/Fz/fsXLlytQXiu+L"
    "3xstXbp017QD0J2AC4AqOakAZKJtAdfFF18cLrjgglRqr0WLFnUNzsr9gwRn5XAufseCBQuK/vg3"
    "lqPyeAcJ/RYuXFh0AHUQcAFQJScVgEy0LeB65JFHwurVq8O2bdvSEJosBmudmmzlcG5qcFYO56YL"
    "zqYL5HqtLVceb7m/l/ECzSHgAqBKTioAmWhbwBWdfvrp4dOf/nQqQf3KQV1UDs7K/eXacuXbV8sh"
    "2nS18OJv9unalusl9CsHcnEa4rREU8NEyJ2AC4AqOakAZKKNAdedd94ZfvM3fzOVgH6UQ7RycDZd"
    "IFe+5bQcok0N5OqohTf1O8q33jK5BFwAVMlJBSATbQy4ole+8pXh5ptvTiVg0sRArFttuXJ/ubZc"
    "uSbbIG3LlQO58ninC+ei8njL4Vx5vPRPwAVAlZxUADLR1oDrG9/4RnjVq16VSgB5iqFZt0Cu3D81"
    "OCvXlisHZ51bZGP+8853vnPXbaptI+ACoEpOKgCZaGvAFcXbFOPtigD80nHHHRe+/OUvp1L7CLgA"
    "qJLGDwAYu7e97W2pD4CO9evXpz4AYE8EXACMXaylcNBBB6USAPG4+OIXvziVAIA9US0YIBNtvkUx"
    "irfhnHDCCakEkL/Yjlan4fqpDeKX29/qtLkVdZ52+Y53vKP1wb9bFAGokpMKQCbaHnDF2Yu1Fe65"
    "5540BJhE8YmFHeUnIfbyVMWo3Fh7bPg9PjUxKj89caawqfyd5c+Un7IYc5nydJa/szydzEzABUCV"
    "nFQAMtH2gCtSiwv6N93T+2YKccohTDkEirWIyk/smy746TVsKn8mhkMxJIrKwdHU72RyCLgAqJKT"
    "CkAmJiHgUouLUSjfNhZNd6tYOfgp197pNcQpfya+Ht/XEWv/dH7blz8z3XfG905XswlyJeACoEpO"
    "KgCZmISAK1KLq5li0NIJfqarvdNriFOuvTP1M+WwqZfgZ+rny8FP+fOddo2A5hBwAVAlJxWATExK"
    "wNWWWlzlWkIz3SrWrXHpqNfaO+XgZ6bgaLqwqfydcTo6t4pN/U6Aqgm4AKiSkwpAJiYl4Ir+8i//"
    "MrzhDW/YFeL02t5P7I/DonJwVA5+ZgqbpguOZgqbyt9Znk4AZibgAqBKTioAmZikgAuA9hNwAVCl"
    "Z/7lDAAAAACZEnABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3AB"
    "AAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZ"
    "E3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkLVZ6S8ADbdjp9TLDK677rrwyU9+"
    "MixZsiTMmzevGDZ37tywdOnSoj9atmxZmDNnTtEf3xPf27FixYowa9Yzp8dFixaFBQsWFP2zZ88O"
    "y5cvL/qj8mtxXHGcHeXvnvoaAM/Yeaz1WwSAyjipAGRCwNWbrVu3hjVr1oR77rknDWmWGH7FwC2a"
    "P39+WLx4cdEfrVy5MvXt/tp0wVscFl/rWLhwYdFFU0O5mYK38mtTQ79yIBi/N35/R3m6APoh4AKg"
    "Sk4qAJkQcPXu/vvvD4ceemh4/PHH0xBGIQZfMZjrKAd2MazrvDY1XCu/NjWUizXvpgsEew3eyq9N"
    "DQRnCgvLoR9QPQEXAFVyUgHIhICrP/FWxX/7b/9tKsFwYsg2Xa22qcFbOdibWhuuHLzNFK6VX5up"
    "tt3U18rTNfW7Y3AYA8Ro6jTDOAi4AKiSkwpAJgRc/Xv7298e/uzP/iyVgOmUw7WZArty8DbTLbD9"
    "tHs3U1g4SCA4dbpoLgEXAFVyUgHIhICrf3GRnXjiiUVtLmAyxfArBmfRoLfATlcTL7Z3d9FFF+1q"
    "947+CLgAqJKTCkAmBFyDefLJJ8Phhx8e7r777jQEoBqf+tSnwmtf+9pUol8CLgCq9ExDDADQUrFG"
    "xo033hhWr16dhgAM74/+6I+EWwDQIP5rApAJNbiGs2HDhqIm1+bNm9MQgMEcc8wx4YYbbth1CyOD"
    "UYMLgCqpwQXARHjxi18crr766l3t7QAM4oUvfGH43Oc+J9wCgIbxXxOATKjBVY2rrroqnHnmmakE"
    "8GydpzOWG52Pf2M5dp/+9KfDC17wgmI4w1GDC4AqOakAZELAVZ13v/vd4cILL0wloE4zBUZR+UmG"
    "K1euLP6Wn1oYn4K4YMGCor/8ZMM9vTcOi69F3d4b3xffH5WngdERcAFQJScVgEwIuKr1xje+MXz4"
    "wx9OJchHHYFRzBk6463yvTATARcAVXJSAciEgKta27dvD695zWvCddddl4YwqfYUGHVCotmzZ4fl"
    "y5cX/XsKdsrvLdcU6vbe2MX+qDze+DeWI4ERbSTgAqBKTioAmRBwVW/r1q3hyCOPDN/5znfSEIYV"
    "aw51App+AqNyCFSugVRXYBTHFccJjI+AC4AqOakAZELAVY9NmzaFww47LNx3331pSLNUHRjF2krx"
    "N+V07y2HQJ33lgOjhQsXFl3U7b0AvRJwAVAlJxWATAi46vPggw+Gf/fv/l3YsmVLUd5Tu0V7auOo"
    "HBiVby3rvLeXwKjzXoC2EnABUCUnFYBMCLgAaBMBFwBV0vgEAAAAAFkTcAEAAACQNQEXAAAAAFkT"
    "cAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAA"
    "AFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEX"
    "AAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFmblf4C0HA7dkq9tMT3vve9cO6554b/"
    "+3//b5g/f34aGsLKlStTXwgLFy4suo7ly5eH2bOf+f/U3Llzw9KlS4v+aNmyZWHOnDlF/7x588KS"
    "JUuK/ij2x2FRfE98b8dMry1atCgsWLCg6J81a1ZYsWJF0R9NnTaAfuw8pvgtAkBlnFQAMiHgaqfN"
    "mzeHt7/97eEDH/hAGpKvcvgVA7vFixen0rNDu6mvzRTMxQAvBnnRTOHbTK/FQDAGgx3l16I43Z3f"
    "2XF4fL2jHCgC1RJwAVAlJxWATAi42u2OO+4If/iHfxjuvvvuNISmiUFbOXwrB3NTQ7tyMDdTwBZ/"
    "309XK27qa1PDt5mCuZlCw5lqAcIoCbgAqJKTCkAmBFztt3379vCRj3wk/Mmf/El4/PHH01AYnRiE"
    "TRfMxQCvcyvt1Bpz5dcGrTEXx3nKKaeEVatWFWXaT8AFQJWcVAAyIeCaHI888khx2+LHP/7xNATa"
    "be3ateHSSy8NL3nJS9IQJoGAC4AqOakAZELANXnuuuuucN5554XvfOc7aQi0y7777hve9773hRNP"
    "PDENYZIIuACoklZTAaChYm2Wb3/72+ELX/hC2HvvvdNQyF+8TfHP//zPiyeJCrcAgCr4rwlAJtTg"
    "mmxbt24tAoH3vve9YcuWLWko5CW273X22WeHP/3TP9XWFmpwAVApJxWATAi4iB5++OGiEfpPfvKT"
    "Ydu2bWkoNN8xxxxT3I544IEHpiFMOgEXAFVyUgHIhICLsnvvvTdccMEF4cYbb0xDoHqxxlV8smLn"
    "KYkrV64shse/8+bNK17rPCUxPlUxvj/efth5rfOUxP322y8cfvjhxWehQ8AFQJWcVAAyIeCim1tv"
    "vTW8+c1vDnfffXcaQpvFsCgGTTFMiiHS8uXLi3IMoDohVBw2Z86cImjqvNYJoWLoFD/XLaDqvBY/"
    "F4Oq+B1QJwEXAFVyUgHIhICL6cRN4/Of/3y48MILww9/+MM0lFGIYdDs2bOfFTSVazR1QqgYGC1d"
    "ujQsXLiw6OJrMYSKQVP8G8ud2k6d18pBU+dz0CYCLgCq5KQCkAkBF3vyi1/8InzsYx8L73znO8Oj"
    "jz6ahk6Ozu10U2s0xRCpE0JNrbUUw6Opt9V1q+00NYTqvAYMTsAFQJWcVAAyIeCiVz/72c/CpZde"
    "WjTovXnz5jR09DqBUzloirrdHtctaJpao6kcNHVqNHWCpvg5v5UhLwIuAKrkpAKQCQEX/Yq1uN71"
    "rneFj3zkI8UTF2NYFIOmGBDFUEj7TcA4CbgAqJKTCkAmBFwMavv27UVoBdAkAi4AquSkApAJARcA"
    "bSLgAqBKs9NfAAAAAMiSgAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4"
    "AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACA"
    "rAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsA"
    "AACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMia"
    "gAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYsVnpLwAAAMBE2LFT6p3W"
    "rJ1SLxmwsgAAAICJ0EuwNZWgKw9WEgAAANBqMwVb5QCr1/fRPFYOAAAA0FrdQqtewqpBP8d4WDEA"
    "QCN0u4jsxoUlAAyu1/NtR+7n3anzO8j8VDEO6melAAAjNfUisSouNqF6g+6v9kdohkk/506d/850"
    "97Jcps5j+TNTX6MZrBQAoFa9XETWwcUn9K+O/dW+CKOT0zl32Gnt5TvL31F+fz/fPd3nysNpBiuk"
    "T/3sCHVrww7VpOUZjXqZNmH+c9uOmrbNlJWXZVOnM7f1vSdN3h4G0ab107R1U9eybfI22O885zIv"
    "TZ7Ocel3XXcziuU6dToH/c4q5nc6o1gOuem2vC2nPatzO+2mKeukn/kedpr39F1Tx19+f7/f3fls"
    "+XPl8dEMs9NfMtTvTtk0uU8/QNXicbEjDcpOmvzGTX+aLOcdmCLtGiPZN0b1PTApip03SYPGLk1O"
    "IQ1qhJnCqPhaN+nlZykPb9o8IuACgEYqrgyTNKjR0qQ2flrTZLogZeKlXWHk+8I4vhPapth5d0rF"
    "xsphGmcyXchFcwm4Mpf7QYPxsw1B88X9NErFRkmTlt1xJE224x8Tadzbvn0PBhP3nSgVs5AmeeTT"
    "XP7OQYOq6aZb8NVcAi7GYrqDBQDTa9KxM05LlIrZSrPhnMTEaMr2br+D/uS+z+Q2/bkv70kl4AIc"
    "wCEjTdhf23jMaOM8wVS2c8hP3G+jVMxampXGzUuarGdJL5EZAVcL2AEBJsu4jvvxe6NUbJ00e86p"
    "tJJtG/LT1v02p/mKtyNGqUjDCbgYubYeqHNnvUBeRr3PTtIxYpLmlclgm4b8tH2/bdL8pQxrtxCr"
    "2zCaTcAFAJka1cVhky5CR2US55l2si1DXuI+G6UiAxoknJr6GeshPwKulshl53OQAKhWncfVOO4o"
    "FSdOmn3nLbJl+4W82Gfr0c9y7SUYs56aS8AF7OJgDXmqY991PPglywKAujnXNJP1khcBV4vY+QCo"
    "gvPJ7iwTcmObhXzYX+vXzzKeqRZXeTy91PZitARcjEw/BxXGx3qCPFW17zoGTM+yIRe2VciH/bU+"
    "wwRQ5c9aR/kQcAEABRdwe2YZwbPFH4FRKgJ9cE6pX/n4VF7exYErSYN2k14uxPLUz6deGkTA1TLl"
    "na5JmjpdAJCz4qp7GuktPUsfm1Z6Gw1X9zVX2hx2SYN3+yEI0ESDHCPjZ6JUpMGcgGoyzh2giRcW"
    "4z4g5H6xNY7lZ5n9Up3LoqrpzH191W3c20OV378ng24Lo5zGNuh3Ode1fAdd3/2qc/uoax7qnOam"
    "KS/DuuZ7kPXUpGmpWl3z1kTdlvckzf+g9rSdWobD29MyLuu2vHv5/KCfYzysmJp02xFGqWk7neUx"
    "nHEsP8vsl+pcFlVNZ+7rq25N3B6qnKayQaavrmlpu36XdZPWeT/q3D7qnPaqpruuaaxyuZanscrx"
    "RlXMfxOnaVhVzVNd81LlMu82jcOMv995rnteZlLXfFY5T4PodzlE457mbvqdjz3NQxzfTO/p9/sY"
    "PbcoUrs9HUiY2biWn/UG9WrKRZJ9fXD9Lru61nmd67DOcTdlH2iTqteXdQTVq/O4OpO4P3ekQX1J"
    "H90lDc5KmvRpp32mdTPT52gOARcAjElxlbVTKo7cuC6y26QpyzC3dTnO7Z7Rs75hPOK+15EGVSaN"
    "Nst9O016T9Oe3uoYlgkrqiZNuNBsyo5oWQzOshtclcuuzmVQ1XTmup5GpenbQ5XTF/U6jYN+bx3L"
    "oAlGtTwG/Z5eVLlu6prOKqdxJlVNf13TW+Xy7UxjHeOsUlXTV8e09avp81LV9EXdpnGY8fc7z3XP"
    "y0yqns8q52Um/c5nFUY1b2XjmE+aTQ0uajWOAx1AbiblQrSt+l2WOVyQ2z4AqjWq4+q4zjHxe6NU"
    "hLEQcAFAA7gonCx1re8qfkDV+SPMdp6HutaT9Q/1iftXlIpj04RpYHIJuFqszgvUXoz7+3PXlOVn"
    "PUL72K+r16Rl2tT160cPMKnqPi437fgapydKRRgZARe0lJMK0E1Tw49JVOdxetD1XNf24ZxUv7rW"
    "HdBsTT6+OvYzagKulnOxk6eq1ltVJxXbEcCeDXKsbNLFf13H+ibNI8Co1XVsjXI4vjoHMEoCLmpR"
    "54GcPXMigTzZdydTXeu9n3Ox8zajErf3YaVRwUTLaV+w3zIqAi5oGD8ygLo4vtRv0GVc18X/uNe5"
    "HzXAJKvrGJzjsdX5gFEQcE2AUV/cjvr7mFlVJxPrFepXxX7mApKp9rRdVbHddWNbHC3LGyaDfR2m"
    "Z+eoSVUXi/EAVsW4RnkgrGreqzTK+R/WsMuv27xWtU4maTmW1Tnfk7huxmGStofppi+XZTBOTVhG"
    "VU5D2XTTM+rvG6Wq5q0J89KrKtdnTvM9DpO4fZUNM//9znNVyzoa5XfH76py2qN+p7+JqlwmbVge"
    "VEsNrglR9cG1bpN6sGr6esptOwLITV3nv27H77qO6ZN6Dgcoq+sYmzvnCOpk46pJVQe0eACoclyp"
    "tzZNnO9oFPNehSrmebp5rWp5TtKy7KhznidtvYzLJG0P001fLstgnJq0jKqclrLOdNU1/mjYea9K"
    "VfPYlPnpRdXrNad5H7VJ3L7GpcrtepTLu8rpjmwrsGdqcGXAwWwyVHESHMW2UvXJGnhGnccA++3o"
    "DbvM6zqe170t1DXdjEfd2wsAVEnANUFyuUhxcVwPyxWay49IRqmu7c15pp0cn6B/Ve83jq/QGwFX"
    "JnI4qLkAGi8nPsiTYyfTyem47hzUDHWtB8cpAHIg4KJRJvUCObcLRxe6UI0q96Xpjp/21/GpYtnn"
    "cF7MYRoZXtyeo1QERsQxFnon4JowdV2YuODJgxMkNEM8ZkapCDNy7KZJ0uGrkAYBQCMIuDLiAred"
    "qrhAHMe24cKWSRW3/WGlUVXG+aH9mrqObXvNM8p1kg5phTQIJp79AcZHwDWBmnrQdZEM0D/HTsbF"
    "tkdZvL4sS4OBITjOQn/sMDWp6sQ+9aBW13iH0fR5jaqc3yrVtez2ZFzfO0pVzWNU53xOwrpogiq3"
    "hybpZb3XMe9t3d5yWFZ1TOMgctgGmrKs+lHlcm3S/OewvfRr0revUapyWY9iGeQ2vdAmanBlxkGO"
    "qca5TVR5Agf643wwmZqw3m17eWjSeorXCx1pEABUTsA1oaq6wKhqPC6WR8vyhrz1ug9XdYxmcHWs"
    "g3Eew8f53bRD3Cc60iAAqISAK0MuLtuhLRd2LlBhtJwDgF41/XgRryGiVASAoQi4JlhTLij8WBuc"
    "ZQeTI+7vUSoy4caxLdj+8pTDeitSrp1SEdjJMRf6J+BiYC5EBteEZVflSdO2APWznzHVKH/8jPK7"
    "qF4u6y8e56JUBIC+CLgy5UIT2wBMnvTbz48/dhnFucD5ph1yWo+OcwAMQsA14cZ9ATGJF80u2oBh"
    "OY5QVue5dBLP020W12eUio0Wj3NRKgLAHgm4MjbOCxQXHO1Q5TZkm4DRivtclIpMsDq3A9tYO43z"
    "GrJftkEAeiXgggzldGEK1MuPv8k2ivVvG2uneC0RpWKj2QYB6IWAi74vGqq6yMjloqpKVS27pmr7"
    "/EFT2fcm0yjXu22svYRcALSFgCtzkxgSUS3bELSDH3/AoOK1QJSKjeU4xySxvUP/BFwURn0AzeEi"
    "qmpVLeNJXHZAb0Z9LGd8xrGubV/tF68xOtIgAMiGgIu+uLhtpyovZG0jAL9UR1AwzuOsY/zkiNtu"
    "lIqNYRsEYDr+O1OTqk6+vV5YVPF9vXzXKOerqu+Kevm+OlU5LzkY9/KOctl+qprOJizzJmvS9lDl"
    "tEynM411fdewy6Cp6lheVS+rutZpv3LYBpqyrPphufZu3MvK9jU6VS7rUSyD3KYX2sQOU5OqDmy9"
    "HtRG8X25zlPU63fWpcp5yUWblnmd81LVdI57eTddU7eHKqdrqs501vEdVS6DJmn6sqpj+obR9O2g"
    "quXV9Pkcp3Fuk+NeL7av0alyOxvF8s5teqFN3KLYEg5+APmJx+4oFQGykg5hhTRoZKoMEaCpbOfQ"
    "HwEXz+IgWr1JXaa2JehdHT8O7YPt0sT1aRujrEi5kjQIAEZKwEVPqrqIddED0J3jI9NpcpDU5Glj"
    "fOLxLErF2tj+aKJRbPtAdwKuFnEwbR4XXkA/HMfbpYr1mcN5xLmO6cR9IEpFYACOsdA7ARe7qesg"
    "6gJn8jghQ/+qPFbGfdCxN185HUNzmlZGz3EIgFEQcLVMHRcQLloHY7kB0ETCBsYhbndRKgJ98LsC"
    "eiPgAmrlhAzQv7qOnXUGDI73AM8Q5sJ4CLjoqnORWtXFqoM8QO8cM/M3zDqs6tw7VXmahpm+mdQ1"
    "7bRH1duebY5JYVuHPRNwtVBdF630zgkIaIp4PHJeyEdd549u20Bd20Vd8wAw6dpyfI3zUaU0WhBw"
    "MT0HC6piWwLYs3EcK4VcjENd2x00ie28O+cH6mSnq0lVO+6gB8YmHTiaMA+DTsMgmrTsm2SU6yDK"
    "ZfupajrrmMZclmEvcpyXOqa5qnGOahmM2riXT1XfP1Uv01PHdw+6HKpS1TzVNR9VLvNxL+tB5D7/"
    "VU1/XdOe+/Ity3leqpz2snGvk0HVsTxyXRbUQw2ulrKj0zR1neAB2qCuY+Q4rwcc90cvLvNhpVEB"
    "FajrGJzjvur4wigIuKBCVR+440lxXNIkANCHQY6fdV309zMtg0x3L+qaN4BJl9Px1bmAURFwtVhd"
    "F6v9aMI05Grcy66O73dyg/Ea93GF3dV1XBxkXde1fTj2A9Qjh+NrndNY13mLfAm4AKBhcrhgZXf9"
    "Xmg3cT3X9WPBNp0P6wqqVddxtaPJ+6zjCaMm4KI2dR/Mm6aNB/BJW4cwCezX7TfsOraNAOSlib9D"
    "6p4m5yq6EXC1nB0/T21eb008AQMMq9/jdl3HwiafPxz/qUOTt3noGMV2Go+xUSqOTZoMx3vGQsAF"
    "FXAQB6pS9fGk20X1KC60mV5d54wq12td20hd80616l5PtgMmUV3H1anGuX+N6rtHtSzJj4BrAozj"
    "AOCgM7imLbs6pmecJ15oslHuG0071uSsn2VZ1zquY33WMc5olNs5wCSKx9koFWuXvm4k31fXuYl2"
    "sHHUpKodvKodeFQHnI4qprvKaa5qOXZT9bKtc1oHVfU8RnXPZ5XTXOe0VjWddUxjLsuwFznMS5XT"
    "WDbT9Nb1nZOkn+2hzuXdz3T0q67prnOao6qmu67prHK5dqaxynFGdcx7DtPYi6rmo67pr3I5j2sZ"
    "d7RpXqIq56cfVc97W+aDdlGDC9gjJxKol4tEhmE9Upeqj03jOtZBk4zrmB33v440qG/p44U0aKTi"
    "sktfP5Q0OlpIwDUhRnkgHeV3tc0kLTsnFyZdcYWVpEGV6+WYMknHnar1s+zqWs+jWH91fUed2z7V"
    "qWo91bG+R7H9QxvF/XEQ6eNjYX+nFwIuGMK4D/TQUVx1VCyNmiGlxbmb9DKZ6udCu671PcqL/bq+"
    "y76Qh2HXk/UMzzbK4zdMEjtWn5p8gt7TgXIU097LwTqHZTiuaexl+Y1TE5bLuKahF7lM5ziUl02V"
    "LOeZ9bPcLcvedZbrOJdZP+u2Sk2Y53FOQ1ONctl0vqsXdU9PP9PSq1Esw9zUsZyjJi/ruuZ5Kttb"
    "b8rro4plNqr1y+ipwTVB7MjNZv0AVer3mOIYlJd4gT+T9La+pY9PK72NCZY2hV3S4EIatEsaXAvH"
    "LNrAdrxnlhH9EHBBQ6RrwT1Kb69cGv2M0luBlnIRuWeWETxbukQopEFAH5xXpmfZ0C8B14Sp8yDh"
    "AATQDMMcjx3Lp2fZQDPYF2kb2/TuLBMGIeACAJ7FReXuLBNoBvsibWXb/iXLgkEJuACgRaq6KHRx"
    "+UuWBQCj4HxjGTAcAdcEquOg4UAEMH5VH4sn/dge5z9KRWDM7I9Mgknezu3jDEvABQAtUNdF4aRe"
    "bE7qfENT2SeZJJO2vcf5jVIRBibgmlBVHkAcjADGq+7jcBx/lIqtN0nzCjmwTzKJJmW7t39TJQEX"
    "AGRslBeGk3AROgnzCEAe4jkpSsXWafO8MR4CLgDI1DguDON3RqnYGmm2XGhDg6Td0n7JxGvbflDs"
    "2DulIlRGwDXBqjioODABjMe4j79tOf7H+YhSESZW0/YD+yU8W9wnolTMUpoF+za1EXABQEbStWEj"
    "Lg7TpGR7oZrztEMdmrJP2DdhenH/iFIxGzlOM/kRcAFABoqr2Z1SsVHSpGVx4ZomtZAGASXj3DeK"
    "HXOnVARmkHaXRu8vaRILaRDUSsA14YY52DhQAdQrHmc70qBGS5NaSIMaI02W8xb0YBz7iv0TBhP3"
    "nY40aOzS5NinGTkBFwA0RLoe3CUNzlKahbHOQ5qEQhoE9CjtOrXuO+krCmkQMIS0O41lf0pfXUiD"
    "YORsfADAyOzYKfXWwoU11Keq/dd+CqNV17nXvkzT2CABgLEa5MLbRTWMXy/7rn0VmqvX86/9GAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAqEsI/x9z0K+n0HYgBQAA"
    "AABJRU5ErkJggg=="
)


# ------------------------------------------------------------------
# Paths
# ------------------------------------------------------------------
INSPECTION_DIR = Path(__file__).parent
SCRIPT_PATH = Path(__file__).resolve()

# ------------------------------------------------------------------
# Remote support: bug reporting + self-update
#   Bug reports POST to a Discord webhook (goes to developer's phone).
#   Updates are pulled from a JSON manifest hosted on GitHub.
#   Updates PRESERVE the local CUSTOMS_DATA block so user data is
#   never lost when the code is replaced.
# ------------------------------------------------------------------
APP_NAME = "Packages at Customs Console"
APP_VERSION = "1.0.4"
DEVELOPER_NAME = "Atlas Ramoon"

BUG_REPORT_WEBHOOK_URL = "https://discord.com/api/webhooks/1524620703259951104/fqpIEBXVWsKHy7f1iZ9xoryCpidmjPYIDuITfcwMOjBfMyS2HtJNWpVbfOetapl8vw9O"

UPDATE_MANIFEST_URL = (
    "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/"
    "manifests/customs-console.json"
)

_DATA_BLOCK_PATTERN = r'CUSTOMS_DATA\s*=\s*\['


def _extract_braced_block(text, start_pattern, open_ch, close_ch):
    """Find the variable assignment matching *start_pattern* and return the
    full block from the variable name through the matching closing bracket,
    using depth counting so brackets inside string literals don't confuse it."""
    m = re.search(start_pattern, text)
    if not m:
        return None, -1, -1
    bracket_pos = m.end() - 1
    depth = 0
    i = bracket_pos
    in_str = False
    str_ch = ""
    while i < len(text):
        ch = text[i]
        if in_str:
            if ch == "\\":
                i += 2
                continue
            if ch == str_ch:
                in_str = False
        else:
            if ch in ('"', "'"):
                in_str = True
                str_ch = ch
            elif ch == open_ch:
                depth += 1
            elif ch == close_ch:
                depth -= 1
                if depth == 0:
                    return text[m.start():i+1], m.start(), i + 1
        i += 1
    return None, -1, -1


def _splice_block(new_text, start_pattern, open_ch, close_ch, local_block):
    _, ns, ne = _extract_braced_block(new_text, start_pattern, open_ch, close_ch)
    if ns == -1:
        return new_text
    return new_text[:ns] + local_block + new_text[ne:]


def _version_tuple(v):
    out = []
    for p in str(v).split("."):
        try:
            out.append(int(p))
        except ValueError:
            out.append(0)
    return tuple(out)


def _http_get(url, timeout=10):
    req = urllib.request.Request(
        url, headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8")


def _post_to_discord(content):
    """POST a message to the Discord webhook. Returns (ok, error_msg)."""
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    payload = json.dumps({"content": content[:1900]}).encode("utf-8")
    try:
        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=payload,
            headers={"Content-Type": "application/json",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)


def _generate_case_number():
    """Generate a short case number like CASE-2025-A4F2."""
    year = datetime.now().year
    tag = uuid.uuid4().hex[:4].upper()
    return f"CASE-{year}-{tag}"


def _post_bug_report_with_files(description, case_number, file_paths,
                                reporter_email="", category="Bug Fix"):
    """POST a bug report to Discord with optional file attachments in ONE message.
    file_paths: list of file paths to attach (can be empty).
    reporter_email: optional email for follow-up (included in message text).
    category: Bug Fix, Feature Request, Environmental Change, or Other.
    Returns (ok, error_msg)."""
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"

    content = (
        f"**Bug Report - {APP_NAME}**\n"
        f"**Case:** {case_number}\n"
        f"**Category:** {category}\n"
        f"**Version:** {APP_VERSION}\n"
        f"**From:** {user}@{host}\n"
        + (f"**Email:** {reporter_email}\n" if reporter_email else "")
        + f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"**Details:**\n{description}"
    )

    if not file_paths:
        # Simple text-only message
        return _post_to_discord(content)

    # Multipart: text + files in a single Discord message
    try:
        boundary = f"----WebKitFormBoundary{uuid.uuid4().hex[:16]}"
        payload_json = json.dumps({"content": content[:1900]})

        body = b""
        body += f"--{boundary}\r\n".encode()
        body += b'Content-Disposition: form-data; name="payload_json"\r\n'
        body += b"Content-Type: application/json\r\n\r\n"
        body += payload_json.encode() + b"\r\n"

        for i, fp in enumerate(file_paths):
            p = Path(fp)
            file_data = p.read_bytes()
            body += f"--{boundary}\r\n".encode()
            body += f'Content-Disposition: form-data; name="files[{i}]"; filename="{p.name}"\r\n'.encode()
            body += b"Content-Type: application/octet-stream\r\n\r\n"
            body += file_data + b"\r\n"

        body += f"--{boundary}--\r\n".encode()

        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)


def _post_update_applied(old_ver, new_ver):
    """Notify the developer that a user applied an update (for invoicing)."""
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    content = (
        f"**Update Applied - {APP_NAME}**\n"
        f"**Updated:** v{old_ver} -> v{new_ver}\n"
        f"**By:** {user}@{host}\n"
        f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    return _post_to_discord(content)


def _check_for_update():
    try:
        data = json.loads(_http_get(UPDATE_MANIFEST_URL))
        remote = data.get("version", "")
        if remote and _version_tuple(remote) > _version_tuple(APP_VERSION):
            return data
    except Exception:
        pass
    return None


# ------------------------------------------------------------------
# Self-update with local-data preservation.
#
# Update contract: the public GitHub copy of this script MUST ship with
# an EMPTY data block (CUSTOMS_DATA = []) so no customer information is
# ever published. Each user's live data lives ONLY in their local .pyw.
#
# On update we download the fresh GitHub version, re-read THIS running
# file to extract the local CUSTOMS_DATA block, and splice that local
# block back into the downloaded text via _splice_block() before the
# atomic os.replace(). That is what makes updates "splice local data in,
# never overwrite user entries."
#
# DO NOT remove the _splice_block() call or the local-block extraction,
# and DO NOT rename the CUSTOMS_DATA marker (see _DATA_BLOCK_PATTERN).
# Doing either would wipe every user's entries on the next update.
# ------------------------------------------------------------------
def _download_and_apply_update(new_url):
    try:
        new_text = _http_get(new_url, timeout=30)
        local_text = SCRIPT_PATH.read_text(encoding="utf-8")
        local_block, _, _ = _extract_braced_block(
            local_text, _DATA_BLOCK_PATTERN, "[", "]")
        if local_block:
            new_text = _splice_block(
                new_text, _DATA_BLOCK_PATTERN, "[", "]", local_block)
        tmp = SCRIPT_PATH.with_name(SCRIPT_PATH.name + ".new")
        tmp.write_text(new_text, encoding="utf-8")
        os.replace(str(tmp), str(SCRIPT_PATH))
        return True, None
    except Exception as e:
        return False, str(e)

# ------------------------------------------------------------------
# Embedded data — saved inside this script file (no external Excel).
#
# CUSTOMS_DATA holds the LIVE user data written by the running app
# (see _save_to_excel, which rewrites this block in place). It is the
# block that _download_and_apply_update preserves via _splice_block.
#
# The public GitHub copy of this file MUST keep this exactly empty:
#     CUSTOMS_DATA = []
# so customer package data is never published. Local machines populate
# it at runtime; those local entries survive updates because the marker
# name below is matched by _DATA_BLOCK_PATTERN and spliced back in.
#
# Do NOT rename this "CUSTOMS_DATA" marker: it is part of the update
# contract. Renaming it here without also updating _DATA_BLOCK_PATTERN
# (and _save_to_excel) will silently wipe every user's entries on the
# next update.
# ------------------------------------------------------------------
CUSTOMS_DATA = [
    {"store": "CB", "cby": "3092", "package": "1073614561", "still_add": "Yes", "manifest": "2024-05-30", "dec": "3378538", "reason": "Held", "paid": "Check COLS", "notes": "To  be Assessed #378-07236935 KX909-ABANDON/LOST"},
    {"store": "CB", "cby": "5059", "package": "1079778125", "still_add": "Yes", "manifest": "2025-08-13", "dec": "3766059", "reason": "Held", "paid": "Check COLS", "notes": "Get seizure notice written up"},
    {"store": "HW", "cby": "5002", "package": "1089045411", "still_add": "Yes", "manifest": "2026-03-02", "dec": "3976040", "reason": "Inspection", "paid": "Duties Paid", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1088931156", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1088948331", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1088957851", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1088959495", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1088963862", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1088972423", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1089175606", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1089398557", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1089403986", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1089422322", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1089449992", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "5801", "package": "1089474741", "still_add": "Yes", "manifest": "2026-04-06", "dec": "4012646", "reason": "Held", "paid": "Check COLS", "notes": ""},
    {"store": "CB", "cby": "2482", "package": "1090242832", "still_add": "Yes", "manifest": "2026-05-11", "dec": "4050573", "reason": "Held", "paid": "Pay and collect", "notes": ""},
    {"store": "CB", "cby": "5034", "package": "1090297710", "still_add": "Yes", "manifest": "2026-05-11", "dec": "4050611", "reason": "Held+Inspection", "paid": "Pay and collect", "notes": ""},
    {"store": "CB", "cby": "3344", "package": "1090559388", "still_add": "Yes", "manifest": "2026-05-26", "dec": "4066121", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "3344", "package": "1090686935", "still_add": "Yes", "manifest": "2026-05-26", "dec": "4066121", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "3344", "package": "1090720834", "still_add": "Yes", "manifest": "2026-05-26", "dec": "4066121", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "3344", "package": "1090734433", "still_add": "Yes", "manifest": "2026-05-26", "dec": "4066121", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "3344", "package": "1090817895", "still_add": "Yes", "manifest": "2026-05-26", "dec": "4066121", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "3344", "package": "1090927262", "still_add": "Yes", "manifest": "2026-05-26", "dec": "4066121", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "HW", "cby": "6003", "package": "1090608672", "still_add": "Yes", "manifest": "2026-06-01", "dec": "4071261", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "HW", "cby": "6003", "package": "1090630936", "still_add": "Yes", "manifest": "2026-06-01", "dec": "4071261", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "4908", "package": "1091200207", "still_add": "Yes", "manifest": "2026-06-10", "dec": "4083167", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "5679", "package": "1091196444", "still_add": "Yes", "manifest": "2026-06-10", "dec": "4083168", "reason": "Held", "paid": "Duties Paid", "notes": ""},
    {"store": "CB", "cby": "3013", "package": "1091062693", "still_add": "Yes", "manifest": "2026-06-22", "dec": "4095268", "reason": "Held", "paid": "Pay and collect", "notes": ""},
    {"store": "CB", "cby": "3013", "package": "1091148679", "still_add": "Yes", "manifest": "2026-06-22", "dec": "4095268", "reason": "Held", "paid": "Pay and collect", "notes": ""},
    {"store": "CB", "cby": "3013", "package": "1091160419", "still_add": "Yes", "manifest": "2026-06-22", "dec": "4095268", "reason": "Held", "paid": "Pay and collect", "notes": ""},
    {"store": "CB", "cby": "3013", "package": "1091431741", "still_add": "Yes", "manifest": "2026-06-22", "dec": "4095268", "reason": "Held", "paid": "Pay and collect", "notes": ""},
    {"store": "CB", "cby": "143", "package": "1091263455", "still_add": "Yes", "manifest": "2026-06-29", "dec": "4102210", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "2703", "package": "1091653727", "still_add": "Yes", "manifest": "2026-07-01", "dec": "4105866", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "2703", "package": "1091631092", "still_add": "Yes", "manifest": "2026-07-01", "dec": "4105866", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "2703", "package": "1091601463", "still_add": "Yes", "manifest": "2026-07-01", "dec": "4105866", "reason": "Held", "paid": "Not yet paid", "notes": ""},
    {"store": "CB", "cby": "2703", "package": "1091593757", "still_add": "Yes", "manifest": "2026-07-01", "dec": "4105866", "reason": "Held", "paid": "Not yet paid", "notes": ""},
]

# ------------------------------------------------------------------
# Theme colours  (Light mode)
# ------------------------------------------------------------------
BG       = "#1a3a5c"   # window background (blue)
PANEL    = "#ffffff"   # panels / cards
INPUT    = "#ffffff"   # input / treeview background
BORDER   = "#d0d5dd"   # subtle borders
DARK     = "#1a1a2e"   # primary text
LIGHT    = "#ffffff"   # light text (on blue background)
TEXT     = "#e8eef5"   # secondary text (on blue background)
MUTED    = "#aabbcc"   # muted / status text (on blue background)
ACCENT   = "#2e6fdb"   # primary accent (blue)
ACCENT_H = "#1a56c4"
GREEN    = "#16a34a"   # add / success
GREEN_H  = "#15803d"
ORANGE   = "#ea7c1a"   # save
ORANGE_H = "#c66812"
RED      = "#dc2626"   # delete
RED_H    = "#b91c1c"
PURPLE   = "#7c3aed"   # reference external files
PURPLE_H = "#6d28d9"
ROW_ALT  = "#c5d8eb"   # alternating row stripe (light blue)
HEAD_BG  = "#b0cce4"   # treeview heading background (light blue)
HEAD_FG  = "#1a1a2e"   # treeview heading text
SELECTED = "#94b8d8"   # selected row background (medium light blue)
SELECTED_FG = "#1a1a2e"
TREE_BG  = "#d6e4f0"   # treeview background (light blue)
# PDF font — fpdf2 built-in (no external font files needed)
PDF_FONT = "Helvetica"

# Reason dropdown values
REASON_VALUES = ["Held", "Inspection", "Held+Inspection"]

# COLS Status dropdown values
PAID_VALUES = ["Not yet paid", "Duties Paid", "Check COLS", "Pay and collect", "Inspect and collect"]

# Column definitions: (key, label, width, stretch)
COLUMNS = [
    ("store",       "Store",             60,  False),
    ("cby",         "CBY",               60,  False),
    ("package",     "Package #",        130,  False),
    ("still_add",   "Still ADD?",        90,  False),
    ("manifest",    "Manifest Date",    110,  False),
    ("dec",         "DEC #",             90,  False),
    ("reason",      "Reason",           130,  False),
    ("paid",        "COLS Payment Status", 150,  False),
    ("notes",       "Inspection Notes", 320,  True),
]


# ==============================================================================
# ERROR REPORTING INFRASTRUCTURE
# Drop-in replacement for messagebox.showerror that adds a "Report Issue"
# button for real errors (inside except blocks), sending the full traceback
# to the developer's Discord.  Validation messages get a plain OK dialog.
# ==============================================================================

_WINDOW_NAME_REGISTRY = {}


def _register_window_name(win, name, colors=None):
    try:
        _WINDOW_NAME_REGISTRY[str(win)] = (name, colors or {})
    except Exception:
        pass


def _get_active_window_info():
    try:
        root = tk._default_root
        if not root:
            return "unknown", {}
        focused = root.focus_get()
        if focused is None:
            return "unknown", {}
        top = focused.winfo_toplevel()
        entry = _WINDOW_NAME_REGISTRY.get(str(top))
        if entry:
            return entry[0], entry[1]
        return "unknown", {}
    except Exception:
        return "unknown", {}


def _send_error_report(title, detail, window_name=""):
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    header = (
        f"**Auto Error Report - {APP_NAME}**\n"
        f"**Window:** {window_name or 'unknown'}\n"
        f"**Title:** {title}\n"
        f"**From:** {user}@{host}\n"
        f"**Version:** {APP_VERSION}\n"
        f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    full = header + f"**Details:**\n{detail}"
    LIMIT = 1900
    if len(full) <= LIMIT:
        return _post_to_discord(full)
    parts = []
    while full:
        if len(full) <= LIMIT:
            parts.append(full)
            break
        cut = full.rfind("\n", 0, LIMIT)
        if cut < LIMIT // 2:
            cut = LIMIT
        parts.append(full[:cut])
        full = full[cut:].lstrip("\n")
    ok_all = True
    err_all = None
    for i, part in enumerate(parts):
        prefix = f"[{i+1}/{len(parts)}] " if len(parts) > 1 else ""
        ok, err = _post_to_discord(prefix + part)
        if not ok:
            ok_all = False
            err_all = err
    return ok_all, err_all


_tk_messagebox_showerror = messagebox.showerror


def _show_error_with_report(title="Error", message="", parent=None,
                            traceback_text=None, window_name="", **kwargs):
    """Drop-in replacement for messagebox.showerror.

    For REAL errors (called inside an active except block, or with an
    explicit traceback_text), shows a 'Report Issue' button that sends
    the full diagnostic to Discord.

    For VALIDATION messages (called outside any except block), shows
    a plain OK-only dialog.
    """
    if not window_name:
        window_name, win_colors = _get_active_window_info()
    else:
        win_colors = {}

    is_real_error = bool(traceback_text)
    if not traceback_text:
        try:
            exc_type, exc_val, exc_tb = sys.exc_info()
            if exc_type is not None:
                traceback_text = "".join(
                    traceback.format_exception(exc_type, exc_val, exc_tb))
                is_real_error = True
        except Exception:
            pass

    # Use the app's light colour scheme for the error dialog
    dlg_bg     = win_colors.get("panel", BG)
    input_bg   = win_colors.get("input", INPUT)
    border_col = win_colors.get("border", BORDER)
    accent     = win_colors.get("accent", ACCENT)
    accent_h   = win_colors.get("accent_hover", ACCENT_H)
    text_col   = win_colors.get("text", DARK)
    muted_col  = win_colors.get("muted", MUTED)

    try:
        win = tk.Toplevel(parent) if parent is not None else tk.Toplevel()
    except Exception:
        return _tk_messagebox_showerror(title, message, **kwargs)

    win.title(title or "Error")
    win.configure(bg=dlg_bg)
    win.resizable(False, False)
    try:
        win.attributes("-topmost", True)
    except Exception:
        pass
    try:
        win.grab_set()
    except Exception:
        pass

    msg_text = str(message) if message is not None else ""
    lines = msg_text.count("\n") + 1
    text_height = max(3, min(10, lines))
    w = 460

    _menu_font = ("Segoe UI", 11) if platform.system() == "Windows" else ("Helvetica", 13)

    def _text_context_menu(event, text_widget, menu_bg):
        try:
            menu = tk.Menu(text_widget, tearoff=0, bg=menu_bg, fg=text_col,
                           activebackground=accent, activeforeground="#ffffff",
                           borderwidth=1, relief="solid", activeborderwidth=0,
                           font=_menu_font)
        except Exception:
            menu = tk.Menu(text_widget, tearoff=0, bg=menu_bg, fg=text_col,
                           activebackground=accent, activeforeground="#ffffff",
                           font=_menu_font)
        menu.add_command(label="Copy",
                         command=lambda: text_widget.event_generate("<<Copy>>"))
        menu.add_separator()
        menu.add_command(label="Select All",
                         command=lambda: text_widget.tag_add("sel", "1.0", "end"))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    tk.Label(win, text="\u26a0  " + (title or "Error"), bg=dlg_bg,
             fg="#cc0000", font=(MODERN_FONT, 13, "bold"),
             anchor="w", justify="left").pack(fill="x", padx=16, pady=(14, 6))

    text_frame = tk.Frame(win, bg=dlg_bg)
    text_frame.pack(fill="both", expand=True, padx=16)
    txt = tk.Text(text_frame, wrap="word", bg=input_bg, fg=text_col,
                  relief="solid", borderwidth=1, highlightbackground=border_col,
                  highlightcolor=border_col, highlightthickness=1,
                  font=(MODERN_FONT, 10), height=text_height,
                  padx=8, pady=8)
    txt.insert("1.0", msg_text)
    txt.configure(state="disabled")
    txt.pack(fill="both", expand=True)
    txt.bind("<Button-3>", lambda e: _text_context_menu(e, txt, input_bg))

    tb_frame = None
    tb_text_widget = None
    if traceback_text:
        tb_toggle = tk.Label(win, text="\u25b6 Show details", bg=dlg_bg,
                             fg=muted_col, font=(MODERN_FONT, 9, "underline"),
                             cursor="hand2", anchor="w")
        tb_toggle.pack(fill="x", padx=16, pady=(4, 0))

        tb_frame = tk.Frame(win, bg=input_bg)
        tb_text_widget = tk.Text(tb_frame, wrap="word", bg=input_bg,
                                 fg=muted_col, relief="flat",
                                 font=(MODERN_FONT, 8), height=10,
                                 highlightthickness=0, padx=8, pady=6)
        tb_text_widget.insert("1.0", traceback_text)
        tb_text_widget.configure(state="disabled")
        tb_text_widget.bind("<Button-3>", lambda e: _text_context_menu(e, tb_text_widget, input_bg))

        def _toggle_tb(event=None):
            if tb_frame.winfo_ismapped():
                tb_frame.pack_forget()
                tb_toggle.configure(text="\u25b6 Show details")
            else:
                tb_frame.pack(fill="both", expand=False, padx=16, pady=(2, 4))
                tb_text_widget.pack(fill="both", expand=True)
                tb_toggle.configure(text="\u25bc Hide details")
            win.update_idletasks()
            try:
                req_h = win.winfo_reqheight()
                h = max(200, min(600, req_h))
                sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
                win.geometry(f"{w}x{h}+{int((sw-w)/2)}+{int((sh-h)/2)}")
            except Exception:
                pass
        tb_toggle.bind("<Button-1>", _toggle_tb)

    btn_frame = tk.Frame(win, bg=dlg_bg)
    btn_frame.pack(fill="x", padx=16, pady=(10, 14))

    status_lbl = tk.Label(btn_frame, text="", bg=dlg_bg,
                          fg=muted_col, font=(MODERN_FONT, 9))
    status_lbl.pack(side="left")

    full_report = msg_text if not traceback_text else (
        f"{msg_text}\n\n--- Traceback ---\n{traceback_text}")

    if is_real_error:
        def _report():
            report_btn.configure(state="disabled", text="Sending...")
            def worker():
                ok, err = _send_error_report(title or "Error", full_report, window_name)
                def done():
                    if ok:
                        status_lbl.configure(text="Report sent \u2014 thank you!", fg=GREEN)
                        report_btn.configure(text="Sent \u2713")
                    else:
                        status_lbl.configure(text=f"Failed to send: {err}", fg="#cc0000")
                        report_btn.configure(state="normal", text="Report Issue")
                try:
                    win.after(0, done)
                except Exception:
                    pass
            threading.Thread(target=worker, daemon=True).start()

        report_btn = tk.Button(btn_frame, text="Report Issue", command=_report,
                               bg=accent, fg="white", activebackground=accent_h,
                               activeforeground="white", relief="flat",
                               font=(MODERN_FONT, 10, "bold"), padx=14, pady=5,
                               cursor="hand2", bd=0)
        report_btn.pack(side="right", padx=(6, 0))

    ok_btn = tk.Button(btn_frame, text="OK", command=win.destroy,
                       bg="#555566", fg="white", activebackground="#666677",
                       activeforeground="white", relief="flat",
                       font=(MODERN_FONT, 10, "bold"), padx=18, pady=5,
                       cursor="hand2", bd=0)
    ok_btn.pack(side="right")

    try:
        win.update_idletasks()
        req_h = win.winfo_reqheight()
        h = max(200, min(500, req_h))
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        win.geometry(f"{w}x{h}+{int((sw-w)/2)}+{int((sh-h)/2)}")
    except Exception:
        pass

    win.protocol("WM_DELETE_WINDOW", win.destroy)
    try:
        win.wait_window(win)
    except Exception:
        pass


messagebox.showerror = _show_error_with_report


def _tk_global_exception_handler(exc_type, exc_value, exc_tb):
    try:
        tb_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    except Exception:
        tb_text = f"{exc_type}: {exc_value}"
    print(tb_text)
    try:
        _show_error_with_report(
            "Unexpected Error",
            f"Something went wrong:\n\n{exc_value}\n\n"
            f"This wasn't a message we expected, so please use 'Report Issue'\n"
            f"below to send the details straight to the developer.",
            traceback_text=tb_text)
    except Exception:
        pass


# ==============================================================================
# PACKAGES AT CUSTOMS CONSOLE
# ==============================================================================
class CustomsConsole:
    """Treeview-based console for managing packages held at Customs."""

    def __init__(self):
        ctk.set_appearance_mode("Light")
        ctk.set_default_color_theme("blue")
        self.root = ctk.CTk()
        self.root.title(f"Packages at Customs — ADD Status Log  v{APP_VERSION}")
        self.root.configure(fg_color=BG)
        self.root.report_callback_exception = _tk_global_exception_handler
        _register_window_name(self.root, "Customs Console", {
            "bg": BG, "panel": BG, "input": INPUT, "border": BORDER,
            "accent": ACCENT, "accent_hover": ACCENT_H,
            "text": DARK, "muted": MUTED, "light": LIGHT,
        })

        WIN_W, WIN_H = 1200, 740
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{WIN_W}x{WIN_H}+{int((sw-WIN_W)/2)}+{int((sh-WIN_H)/2)}")
        self.root.minsize(1000, 560)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        self._rows = []          # list of dicts — each row's data
        self._dirty = False      # unsaved changes?
        self._editing_win = None
        self._search_var = ctk.StringVar()
        self._tooltip_win = None
        self._pending_update = None
        self._support_tooltip = "Report a Bug"

        self._build_ui()
        self._load_from_excel()
        self._populate_tree()

        # Check for updates in the background so startup isn't blocked
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    # ------------------------------------------------------------------
    # Help icon + tooltip helper
    # ------------------------------------------------------------------
    def _add_help_icon(self, parent, text, side="left", padx=(2, 8)):
        """Add a small '?' icon next to a button. Hovering shows a tooltip."""
        icon = ctk.CTkLabel(parent, text="?", width=18, height=18,
                            font=(MODERN_FONT, 11, "bold"),
                            text_color=LIGHT, fg_color=ACCENT,
                            corner_radius=9, cursor="question_arrow")
        icon.pack(side=side, padx=padx)
        icon.bind("<Enter>", lambda e: self._show_tooltip(icon, text))
        icon.bind("<Leave>", lambda e: self._hide_tooltip())
        return icon

    def _show_tooltip(self, widget, text):
        self._hide_tooltip()
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + 18
        # Estimate tooltip height (rough: count lines * ~16px + padding)
        line_count = text.count("\n") + 1
        est_height = line_count * 16 + 16
        # If not enough room below, show above the icon
        screen_h = widget.winfo_screenheight()
        if y + est_height > screen_h - 20:
            y = widget.winfo_rooty() - est_height - 8
        tw = tk.Toplevel(self.root)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=text, justify="left",
                         bg="#1a1a2e", fg="#e8e8e8", relief="solid",
                         bd=1, padx=10, pady=8,
                         font=(MODERN_FONT, 10), wraplength=320)
        label.pack()
        self._tooltip_win = tw

    def _hide_tooltip(self):
        if self._tooltip_win is not None:
            try:
                self._tooltip_win.destroy()
            except Exception:
                pass
            self._tooltip_win = None

    # ------------------------------------------------------------------
    # Entry right-click context menu
    # ------------------------------------------------------------------
    def _attach_entry_context_menu(self, widget):
        """Attach a right-click Cut/Copy/Paste/Select All menu to a CTkEntry."""
        inner = getattr(widget, "_entry", widget)

        def _show(event):
            menu = tk.Menu(inner, tearoff=0,
                           bg="#1a1a2e", fg="#e8e8e8",
                           activebackground=ACCENT, activeforeground="#ffffff",
                           font=(MODERN_FONT, 11), borderwidth=1, relief="solid")
            menu.add_command(label="Cut",        command=lambda: inner.event_generate("<<Cut>>"))
            menu.add_command(label="Copy",       command=lambda: inner.event_generate("<<Copy>>"))
            menu.add_command(label="Paste",      command=lambda: inner.event_generate("<<Paste>>"))
            menu.add_separator()
            menu.add_command(label="Select All", command=lambda: inner.event_generate("<<SelectAll>>"))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        inner.bind("<Button-3>", _show)

    def _attach_textbox_context_menu(self, widget):
        """Attach a right-click Cut/Copy/Paste/Select All menu to a CTkTextbox."""
        inner = getattr(widget, "_textbox", widget)

        def _show(event):
            menu = tk.Menu(inner, tearoff=0,
                           bg="#1a1a2e", fg="#e8e8e8",
                           activebackground=ACCENT, activeforeground="#ffffff",
                           font=(MODERN_FONT, 11), borderwidth=1, relief="solid")
            menu.add_command(label="Cut",        command=lambda: inner.event_generate("<<Cut>>"))
            menu.add_command(label="Copy",       command=lambda: inner.event_generate("<<Copy>>"))
            menu.add_command(label="Paste",      command=lambda: inner.event_generate("<<Paste>>"))
            menu.add_separator()
            menu.add_command(label="Select All", command=lambda: inner.event_generate("<<SelectAll>>"))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        inner.bind("<Button-3>", _show)

    # ------------------------------------------------------------------
    # Progress bar helper (themed, non-blocking)
    # ------------------------------------------------------------------
    def _show_progress(self, title="Working...", indeterminate=True):
        """Show a themed progress bar overlay. Returns the toplevel + bar."""
        prog_win = ctk.CTkToplevel(self.root)
        prog_win.title("")
        prog_win.configure(fg_color=BG)
        prog_win.transient(self.root)
        prog_win.overrideredirect(True)  # no window chrome
        pw, ph = 340, 90
        sw, sh = prog_win.winfo_screenwidth(), prog_win.winfo_screenheight()
        prog_win.geometry(f"{pw}x{ph}+{int((sw-pw)/2)}+{int((sh-ph)/2)}")
        prog_win.grab_set()

        ctk.CTkLabel(prog_win, text=title,
                     font=(MODERN_FONT, 13, "bold"), text_color=LIGHT).pack(pady=(18, 8))

        if indeterminate:
            bar = ctk.CTkProgressBar(prog_win, width=280, height=14,
                                     progress_color=ACCENT,
                                     fg_color=PANEL)
            bar.pack(pady=(0, 10))
            bar.set(0)  # start at 0, we'll animate via after()
        else:
            bar = ctk.CTkProgressBar(prog_win, width=280, height=14,
                                     progress_color=GREEN,
                                     fg_color=PANEL)
            bar.pack(pady=(0, 10))
            bar.set(0)

        self.root.update_idletasks()
        return prog_win, bar

    def _close_progress(self, prog_win):
        """Close the progress bar overlay."""
        try:
            prog_win.grab_release()
            prog_win.destroy()
        except Exception:
            pass

    def _pulse_progress(self, bar, prog_win):
        """Animate an indeterminate progress bar."""
        try:
            current = bar.get()
            new = current + 0.05
            if new > 1.0:
                new = 0.0
            bar.set(new)
            prog_win.after(50, lambda: self._pulse_progress(bar, prog_win))
        except Exception:
            pass

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------
    def _build_ui(self):
        root = self.root

        # ---- Title bar ----
        title_frame = ctk.CTkFrame(root, fg_color="transparent")
        title_frame.pack(fill="x", padx=16, pady=(12, 0))

        ctk.CTkLabel(title_frame, text="Packages at Customs",
                     font=(MODERN_FONT, 20, "bold"), text_color=LIGHT).pack(side="left")
        ctk.CTkLabel(title_frame, text="Customs Tracker",
                     font=(MODERN_FONT, 12), text_color=TEXT).pack(side="left", padx=(14, 0))

        # Logo (top-right)
        try:
            _img = Image.open(io.BytesIO(base64.b64decode(MBE_LOGO_B64)))
            _lw, _lh = _img.size
            _tw = 130
            _th = max(1, int(_lh * _tw / _lw))
            _logo = ctk.CTkImage(light_image=_img, dark_image=_img, size=(_tw, _th))
            ctk.CTkLabel(title_frame, image=_logo, text="").pack(side="right")
        except Exception:
            pass

        # ---- Subtitle ----
        self._subtitle = ctk.CTkLabel(
            root, text="Double-click a row to edit  -  + Add Package opens a dialog  -  Ctrl/Shift+click to multi-select",
            font=(MODERN_FONT, 11), text_color=MUTED)
        self._subtitle.pack(pady=(8, 10))

        # ---- Search + action bar ----
        bar = ctk.CTkFrame(root, fg_color="transparent")
        bar.pack(fill="x", padx=20, pady=(0, 8))

        ctk.CTkLabel(bar, text="Search:", font=(MODERN_FONT, 12),
                     text_color=LIGHT).pack(side="left", padx=(0, 8))
        self._search_var.trace("w", lambda *a: self._filter_tree())
        search_entry = ctk.CTkEntry(bar, textvariable=self._search_var,
                                    width=200, height=30, fg_color=INPUT,
                                    border_color=BORDER, border_width=1, corner_radius=5,
                                    text_color=DARK, font=(MODERN_FONT, 12))
        search_entry.pack(side="left", padx=(0, 12))
        self._attach_entry_context_menu(search_entry)

        # Update ADD Status button (imports export-porestatus CSV)
        ctk.CTkButton(bar, text="Update ADD Status",
                      command=self._update_add_status,
                      fg_color=PURPLE, hover_color=PURPLE_H,
                      width=150, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        self._add_help_icon(bar,
            "Update ADD Status\n\n"
            "Imports the export-porestatus CSV from e-box to check which\n"
            "packages are still at ADD status:\n"
            "  - Flags packages no longer at ADD status as 'No'\n"
            "  - Adds new packages from the export\n"
            "  - Fills in missing manifest dates and store/CBY info\n\n"
            "HOW TO: Go to the e-box dashboard, click on status\n"
            "'Customs Documentation (ADD)', then hit Export.")

        # Fill DEC# from Manifest button (works on selected rows)
        ctk.CTkButton(bar, text="Fill DEC# from Manifest",
                      command=self._fill_from_manifest,
                      fg_color=GREEN, hover_color=GREEN_H,
                      width=180, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        self._add_help_icon(bar,
            "Fill DEC# from Manifest\n\n"
            "Select one or more rows in the list first, then click this button\n"
            "to choose a manifest file. It reads the Bags sheet and fills in\n"
            "the DEC# and Reason columns by matching package numbers.\n\n"
            "Read-only - does not modify the manifest file.")

        # ---- Visual partition before the standalone COLS button ----
        sep = ctk.CTkFrame(bar, width=3, height=34, fg_color="#5a8aba",
                           corner_radius=0)
        sep.pack(side="left", padx=(18, 18))

        # Update COLS Payment button (imports Declarations export)
        ctk.CTkButton(bar, text="Update COLS Payment",
                      command=self._update_cols_payment,
                      fg_color="#b8860b", hover_color="#daa520",
                      width=160, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        self._add_help_icon(bar,
            "Update COLS Payment\n\n"
            "Imports the Declarations export from COLS to check duty\n"
            "payment status:\n"
            "  - Marks packages as 'Duties Paid' if Status=Paid\n"
            "  - Marks packages older than 2 months as 'Check COLS'\n"
            "    (too old to appear in the export)\n\n"
            "TIP: For best results, export a 2-month paid list from COLS.\n"
            "Filter by last 2 months, load 'paid' status entries, then\n"
            "export as Excel from the site. A shorter range may miss\n"
            "packages that are actually paid.")

        # ---- Treeview ----
        tree_frame = ctk.CTkFrame(root, fg_color=TREE_BG, corner_radius=6)
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 8))

        col_keys   = [c[0] for c in COLUMNS]
        col_labels = [c[1] for c in COLUMNS]
        col_widths = [c[2] for c in COLUMNS]

        self._tree = ttk.Treeview(tree_frame, columns=col_keys,
                                  show="headings", selectmode="extended")
        for key, label, w, stretch in COLUMNS:
            self._tree.heading(key, text=label, anchor="w")
            self._tree.column(key, width=w, minwidth=w, anchor="w", stretch=stretch)

        style = ttk.Style()
        style.configure("Treeview", background=TREE_BG, foreground=DARK,
                        fieldbackground=TREE_BG, bordercolor=BORDER,
                        rowheight=26, font=(MODERN_FONT, 11))
        style.configure("Treeview.Heading", background=HEAD_BG, foreground=HEAD_FG,
                        font=(MODERN_FONT, 11, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", SELECTED)],
                  foreground=[("selected", SELECTED_FG)])

        # Tag for rows flagged "No" (no longer ADD) — light red background
        self._tree.tag_configure("not_add", background="#ffcccc", foreground="#990000")

        tree_scroll = ctk.CTkScrollbar(tree_frame, command=self._tree.yview)
        self._tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side="right", fill="y", padx=(2, 4), pady=4)
        self._tree.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)

        self._tree.bind("<Double-1>", self._on_double_click)
        self._tree.bind("<Tab>", self._tab_next)
        self._tree.bind("<Shift-Tab>", self._tab_prev)

        # ---- Bottom button row ----
        bottom = ctk.CTkFrame(root, fg_color="transparent")
        bottom.pack(fill="x", padx=20, pady=(0, 8))

        ctk.CTkButton(bottom, text="+ Add Package",
                      command=self._add_row,
                      fg_color=GREEN, hover_color=GREEN_H,
                      width=130, height=28, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        self._add_help_icon(bottom,
            "Add Package\n\n"
            "Opens a dialog to enter all package details.\n\n"
            "Package # field has a '+' button — click it to add\n"
            "more package numbers for the same customer/shipment.\n"
            "Each number becomes its own independent row with\n"
            "all other fields (Store, CBY, DEC#, etc.) identical.\n\n"
            "COLS Payment Status automatically updates all other\n"
            "packages that share the same DEC #.")

        ctk.CTkButton(bottom, text="Delete Selected",
                      command=self._delete_selected,
                      fg_color=RED, hover_color=RED_H,
                      width=130, height=28, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left", padx=4)
        self._add_help_icon(bottom,
            "Delete Selected\n\n"
            "Deletes all selected rows (Ctrl/Shift+click to multi-select).\n"
            "Asks for confirmation before deleting.")

        # Dynamic "delete all no longer ADD" button (hidden until needed)
        self._delete_not_add_btn = ctk.CTkButton(
            bottom, text="Delete All No Longer ADD",
            command=self._delete_all_not_add,
            fg_color=RED, hover_color=RED_H,
            width=200, height=28, corner_radius=5,
            font=(MODERN_FONT, 11, "bold"))
        # Packed dynamically when rows are flagged "No"

        # Right-aligned: Export PDF and Save
        ctk.CTkButton(bottom, text="Export PDF",
                      command=self._export_pdf,
                      fg_color=ACCENT, hover_color=ACCENT_H,
                      width=110, height=28, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="right", padx=4)
        self._add_help_icon(bottom,
            "Export PDF\n\n"
            "Generates a printable PDF for the driver.\n"
            "Three options:\n"
            "  - Full List: all packages\n"
            "  - GrabnGo: Held packages with duties paid/ready to collect\n"
            "  - Inspections: Inspection-only packages with duties paid",
            side="right", padx=(8, 2))
        ctk.CTkButton(bottom, text="Save",
                      command=self._save_to_excel,
                      fg_color=ORANGE, hover_color=ORANGE_H,
                      width=100, height=28, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="right", padx=4)
        self._add_help_icon(bottom,
            "Save\n\n"
            "Saves all changes directly into the script file itself.\n"
            "No external Excel file is created. Data syncs via Dropbox.",
            side="right", padx=(8, 2))

        # ---- Status bar (with support icon in the corner) ----
        status_frame = ctk.CTkFrame(root, fg_color="transparent")
        status_frame.pack(fill="x", padx=22, pady=(0, 8))

        self._status = ctk.CTkLabel(status_frame, text="Ready", font=(MODERN_FONT, 11),
                                    text_color=LIGHT, anchor="w")
        self._status.pack(side="left", fill="x", expand=True)

        # Bug / Apply-Fixes icon — bottom-right corner
        self._support_btn = ctk.CTkButton(
            status_frame, text="🐞", width=34, height=28,
            fg_color=BG, hover_color="#24507a", corner_radius=6,
            font=("Segoe UI Emoji", 15), command=self._on_support_click)
        self._support_btn.pack(side="right")
        self._support_btn.bind("<Enter>", lambda e: self._show_tooltip(self._support_btn, self._support_tooltip))
        self._support_btn.bind("<Leave>", lambda e: self._hide_tooltip())

    # ------------------------------------------------------------------
    # Data loading / saving
    # ------------------------------------------------------------------
    def _convert_paid_status(self, raw, manifest_str):
        """Convert old/legacy paid values to the new dropdown values.
        'Yes' -> 'Duties Paid'
        'N/A' or '' -> 'Not yet paid' (or 'Check COLS' if manifest is 2+ months old)
        New dropdown values pass through unchanged."""
        if not raw:
            return "Not yet paid"
        raw_lower = raw.lower().strip()
        if raw_lower == "yes":
            return "Duties Paid"
        if raw_lower in ("n/a", "na", "no", "not yet paid", ""):
            # Check if manifest date is 2+ months old
            try:
                if manifest_str:
                    mdate = datetime.strptime(manifest_str, "%Y-%m-%d")
                    age_days = (datetime.now() - mdate).days
                    if age_days > 60:
                        return "Check COLS"
            except Exception:
                pass
            return "Not yet paid"
        # Already a new-style value — return as-is
        return raw
    def _load_from_excel(self):
        """Load rows from the embedded CUSTOMS_DATA list in this script."""
        self._rows = []
        for entry in CUSTOMS_DATA:
            row = dict(entry)
            # Ensure all keys exist
            for key in ("store", "cby", "package", "still_add", "manifest",
                        "dec", "reason", "paid", "notes"):
                row.setdefault(key, "")
            # Convert any legacy paid values
            row["paid"] = self._convert_paid_status(row.get("paid", ""), row.get("manifest", ""))
            self._rows.append(row)
        self._set_status(f"Loaded {len(self._rows)} packages")

    def _save_to_excel(self):
        """Save the current rows back into this script file (embedded data).
        No external Excel file is created or modified."""
        if not self._rows:
            if not messagebox.askyesno("Confirm", "No rows to save. Save empty list?"):
                return

        try:
            prog_win, bar = self._show_progress("Saving to script file...")
            self._pulse_progress(bar, prog_win)
            import re as _re
            content = SCRIPT_PATH.read_text(encoding="utf-8")

            # Build the new CUSTOMS_DATA block
            lines = ["CUSTOMS_DATA = ["]
            for row in self._rows:
                def _safe(v):
                    return str(v).replace("\\", "\\\\").replace('"', '\\"')
                lines.append(
                    f'    {{"store": "{_safe(row.get("store",""))}", '
                    f'"cby": "{_safe(row.get("cby",""))}", '
                    f'"package": "{_safe(row.get("package",""))}", '
                    f'"still_add": "{_safe(row.get("still_add",""))}", '
                    f'"manifest": "{_safe(row.get("manifest",""))}", '
                    f'"dec": "{_safe(row.get("dec",""))}", '
                    f'"reason": "{_safe(row.get("reason",""))}", '
                    f'"paid": "{_safe(row.get("paid",""))}", '
                    f'"notes": "{_safe(row.get("notes",""))}"}},')
            lines.append("]")
            new_block = "\n".join(lines)

            # Replace the existing block
            pattern = r'CUSTOMS_DATA = \[.*?\n\]'
            new_content = _re.sub(pattern, lambda m: new_block, content, count=1, flags=_re.DOTALL)

            if new_content == content:
                self._close_progress(prog_win)
                messagebox.showinfo("No Change", "No changes to save.")
                return

            SCRIPT_PATH.write_text(new_content, encoding="utf-8")
            self._close_progress(prog_win)
            self._dirty = False
            self._set_status(f"Saved {len(self._rows)} rows to script file")
            messagebox.showinfo("Saved", f"Saved {len(self._rows)} rows to the script file.\n"
                                          "Changes will sync via Dropbox.")
        except Exception as e:
            try:
                self._close_progress(prog_win)
            except Exception:
                pass
            self._set_status(f"Save error: {e}")
            messagebox.showerror("Save Error", f"Could not save:\n{e}")

    # ------------------------------------------------------------------
    # Treeview population / filtering
    # ------------------------------------------------------------------
    def _populate_tree(self):
        self._tree.delete(*self._tree.get_children())
        for row in self._rows:
            tags = ("not_add",) if str(row.get("still_add", "")).strip().lower() == "no" else ()
            self._tree.insert("", "end", values=[
                row.get("store", ""), row.get("cby", ""), row.get("package", ""),
                row.get("still_add", ""), row.get("manifest", ""), row.get("dec", ""),
                row.get("reason", ""), row.get("paid", ""), row.get("notes", ""),
            ], tags=tags)
        self._update_delete_not_add_button()

    def _filter_tree(self):
        """Filter the treeview by the search box (case-insensitive)."""
        query = self._search_var.get().lower().strip()
        self._tree.delete(*self._tree.get_children())
        for row in self._rows:
            tags = ("not_add",) if str(row.get("still_add", "")).strip().lower() == "no" else ()
            vals = [
                row.get("store", ""), row.get("cby", ""), row.get("package", ""),
                row.get("still_add", ""), row.get("manifest", ""), row.get("dec", ""),
                row.get("reason", ""), row.get("paid", ""), row.get("notes", ""),
            ]
            if not query:
                self._tree.insert("", "end", values=vals, tags=tags)
            else:
                row_text = " ".join(str(v) for v in row.values()).lower()
                if query in row_text:
                    self._tree.insert("", "end", values=vals, tags=tags)
        self._update_delete_not_add_button()

    def _sync_tree_to_rows(self):
        """Sync current treeview values back to self._rows.
        Updates existing rows AND adds new rows that were inserted into the
        treeview (e.g. via the Add Package dialog). Preserves filtered view
        by only syncing visible tree items back."""
        tree_items = self._tree.get_children()
        # Build a lookup of package -> tree values
        tree_by_pkg = {}
        for item in tree_items:
            vals = self._tree.item(item, "values")
            if vals and len(vals) >= 3:
                tree_by_pkg[str(vals[2]).strip()] = vals

        # Update existing rows that are still in the tree
        existing_pkgs = set()
        for row in self._rows:
            pkg = str(row.get("package", "")).strip()
            if pkg in tree_by_pkg:
                existing_pkgs.add(pkg)
                vals = tree_by_pkg[pkg]
                row["store"]     = vals[0] if len(vals) > 0 else ""
                row["cby"]       = vals[1] if len(vals) > 1 else ""
                row["package"]   = vals[2] if len(vals) > 2 else ""
                row["still_add"] = vals[3] if len(vals) > 3 else ""
                row["manifest"]  = vals[4] if len(vals) > 4 else ""
                row["dec"]       = vals[5] if len(vals) > 5 else ""
                row["reason"]    = vals[6] if len(vals) > 6 else ""
                row["paid"]      = vals[7] if len(vals) > 7 else ""
                row["notes"]     = vals[8] if len(vals) > 8 else ""

        # Add new rows that are in the tree but not in self._rows
        for pkg, vals in tree_by_pkg.items():
            if pkg not in existing_pkgs:
                self._rows.append({
                    "store":     vals[0] if len(vals) > 0 else "",
                    "cby":       vals[1] if len(vals) > 1 else "",
                    "package":   vals[2] if len(vals) > 2 else "",
                    "still_add": vals[3] if len(vals) > 3 else "",
                    "manifest":  vals[4] if len(vals) > 4 else "",
                    "dec":       vals[5] if len(vals) > 5 else "",
                    "reason":    vals[6] if len(vals) > 6 else "",
                    "paid":      vals[7] if len(vals) > 7 else "",
                    "notes":     vals[8] if len(vals) > 8 else "",
                })

    # ------------------------------------------------------------------
    # Double-click → open edit dialog
    # ------------------------------------------------------------------
    def _on_double_click(self, event):
        """Open the edit dialog for the double-clicked row."""
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            return
        self._open_package_dialog(item_id=row_id)

    # ------------------------------------------------------------------
    # Tab navigation
    # ------------------------------------------------------------------
    def _tab_next(self, event):
        self._move_selection(1)

    def _tab_prev(self, event):
        self._move_selection(-1)

    def _move_selection(self, direction):
        items = self._tree.get_children()
        if not items:
            return
        sel = self._tree.selection()
        if not sel:
            self._tree.focus(items[0])
            self._tree.selection_set(items[0])
            return
        current = sel[0]
        idx = list(items).index(current)
        new_idx = (idx + direction) % len(items)
        self._tree.focus(items[new_idx])
        self._tree.selection_set(items[new_idx])

    # ------------------------------------------------------------------
    # Add / Delete rows
    # ------------------------------------------------------------------
    def _add_row(self):
        """Open the Add Package dialog."""
        self._open_package_dialog()

    def _open_package_dialog(self, item_id=None):
        """Add (item_id=None) or edit (item_id=treeview iid) a package via popup dialog."""
        # Only one dialog at a time
        if self._editing_win is not None:
            try:
                self._editing_win.focus_set()
            except Exception:
                pass
            return

        is_edit = item_id is not None

        # Seed initial field values
        if is_edit:
            vals = self._tree.item(item_id, "values")
            def _v(i, d=""): return vals[i] if i < len(vals) else d
            init = {
                "store": _v(0), "cby": _v(1), "package": _v(2),
                "still_add": _v(3, "Yes"), "manifest": _v(4),
                "dec": _v(5), "reason": _v(6),
                "paid": _v(7, "Not yet paid"), "notes": _v(8),
            }
        else:
            init = {
                "store": "", "cby": "", "package": "",
                "still_add": "Yes", "manifest": "", "dec": "",
                "reason": "", "paid": "Not yet paid", "notes": "",
            }

        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Edit Package" if is_edit else "Add Package")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        self._editing_win = dlg

        def _close():
            self._editing_win = None
            dlg.destroy()
        dlg.protocol("WM_DELETE_WINDOW", _close)

        DLG_W, DLG_H = 460, 600
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        # Clamp height so the dialog never exceeds the screen (leaves 40px margin)
        DLG_H = min(DLG_H, sh - 40)
        dlg.geometry(f"{DLG_W}x{DLG_H}+{(sw - DLG_W) // 2}+{(sh - DLG_H) // 2}")

        # ---- Button bar (packed first so it is always visible at the bottom) ----
        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        # ---- Field grid (fills remaining space above the buttons) ----
        g = ctk.CTkFrame(dlg, fg_color="transparent")
        g.pack(fill="both", expand=True, padx=16, pady=(12, 4))
        g.columnconfigure(0, weight=1)
        g.columnconfigure(1, weight=1)

        def _lbl(text, row, col, cs=2):
            ctk.CTkLabel(g, text=text, font=(MODERN_FONT, 11),
                         text_color=LIGHT, anchor="w").grid(
                row=row, column=col, columnspan=cs,
                sticky="w", padx=(0, 4), pady=(5, 1))

        def _entry(key, row, col, cs=1):
            var = ctk.StringVar(value=init[key])
            e = ctk.CTkEntry(g, textvariable=var, height=32,
                             fg_color=INPUT, border_color=BORDER, border_width=1,
                             corner_radius=4, text_color=DARK, font=(MODERN_FONT, 11))
            e.grid(row=row, column=col, columnspan=cs, sticky="ew",
                   padx=(0, 4), pady=(0, 2))
            self._attach_entry_context_menu(e)
            return var

        def _combo(key, row, col, values, cs=1):
            var = ctk.StringVar(value=init[key])
            c = ctk.CTkComboBox(g, variable=var, values=values, height=32,
                                fg_color=INPUT, button_color=ACCENT,
                                button_hover_color=ACCENT_H,
                                border_color=BORDER, border_width=1,
                                dropdown_fg_color=INPUT, dropdown_text_color=DARK,
                                text_color=DARK, font=(MODERN_FONT, 11))
            c.grid(row=row, column=col, columnspan=cs, sticky="ew",
                   padx=(0, 4), pady=(0, 2))
            return var

        # Row 0-1: Store (dropdown) | CBY
        _lbl("Store", 0, 0, cs=1); store_var = _combo("store", 1, 0, ["CB", "HW"])
        _lbl("CBY",   0, 1, cs=1); cby_var   = _entry("cby",   1, 1)

        # Row 2-3: Package # (entry + + button in add mode, plain entry in edit) | Still ADD?
        # The extra_lbl for the "Also:" preview lives in row 4 of the MAIN grid so
        # that both pkg_cell and the Still ADD? combo are exactly 28 px tall and
        # therefore perfectly vertically aligned.
        _lbl("Package #",  2, 0, cs=1)
        _lbl("Still ADD?", 2, 1, cs=1)
        still_var  = _combo("still_add", 3, 1, ["Yes", "No"])
        pkg_var    = ctk.StringVar(value=init["package"])
        extra_pkgs = []   # additional package numbers (add mode only)

        if not is_edit:
            # Sub-frame: [entry] [+] side-by-side, same height as the adjacent combo
            pkg_cell = ctk.CTkFrame(g, fg_color="transparent")
            pkg_cell.grid(row=3, column=0, sticky="ew", padx=(0, 4), pady=(0, 2))
            pkg_cell.columnconfigure(0, weight=1)

            pkg_ent = ctk.CTkEntry(pkg_cell, textvariable=pkg_var, height=32,
                                   fg_color=INPUT, border_color=BORDER, border_width=1,
                                   corner_radius=4, text_color=DARK,
                                   font=(MODERN_FONT, 11))
            pkg_ent.grid(row=0, column=0, sticky="ew")
            self._attach_entry_context_menu(pkg_ent)

            ctk.CTkButton(pkg_cell, text="+", width=32, height=32,
                          fg_color=ACCENT, hover_color=ACCENT_H,
                          corner_radius=4, font=(MODERN_FONT, 11, "bold"),
                          command=lambda: _open_more_pkgs()).grid(row=0, column=1, padx=(4, 0))

            # Row 4: "Also: ..." preview label — in main grid so row height
            # is independent of the pkg_cell/still_add row above.
            extra_lbl_var = ctk.StringVar(value="")
            ctk.CTkLabel(g, textvariable=extra_lbl_var,
                         font=(MODERN_FONT, 10), text_color="#ffcc44",
                         anchor="w").grid(row=4, column=0, columnspan=2,
                                          sticky="w", padx=(0, 4), pady=(0, 3))

            def _open_more_pkgs():
                more = ctk.CTkToplevel(dlg)
                more.title("More Packages")
                more.configure(fg_color=BG)
                more.transient(dlg)
                more.grab_set()
                more.resizable(False, False)
                mw, mh = 340, 270
                msx, msy = more.winfo_screenwidth(), more.winfo_screenheight()
                more.geometry(f"{mw}x{mh}+{(msx-mw)//2}+{(msy-mh)//2}")

                m_btns = ctk.CTkFrame(more, fg_color="transparent")
                m_btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

                ctk.CTkLabel(more,
                    text="More packages for this customer on this shipment?\n"
                         "Add package numbers below — one per line or comma-separated:",
                    font=(MODERN_FONT, 11), text_color=LIGHT,
                    anchor="w", wraplength=308, justify="left").pack(
                    padx=16, pady=(14, 6), anchor="w")

                more_box = ctk.CTkTextbox(more, height=100, fg_color=INPUT,
                                          border_color=BORDER, border_width=1,
                                          corner_radius=4, text_color=DARK,
                                          font=(MODERN_FONT, 11))
                more_box.pack(fill="x", padx=16)
                self._attach_textbox_context_menu(more_box)
                if extra_pkgs:
                    more_box.insert("0.0", "\n".join(extra_pkgs))
                more_box.focus_set()

                def _ok():
                    raw = more_box.get("0.0", "end").strip()
                    parts = re.split(r"[,\n\r]+", raw)
                    extra_pkgs.clear()
                    extra_pkgs.extend(p.strip() for p in parts if p.strip())
                    if extra_pkgs:
                        preview = ", ".join(extra_pkgs[:5])
                        if len(extra_pkgs) > 5:
                            preview += f"  +{len(extra_pkgs)-5} more"
                        extra_lbl_var.set(f"Also: {preview}")
                    else:
                        extra_lbl_var.set("")
                    more.destroy()

                ctk.CTkButton(m_btns, text="OK", command=_ok,
                              fg_color=ACCENT, hover_color=ACCENT_H,
                              width=80, height=28, corner_radius=4,
                              font=(MODERN_FONT, 11, "bold")).pack(side="left")
                ctk.CTkButton(m_btns, text="Cancel", command=more.destroy,
                              fg_color="#667788", hover_color="#556677",
                              width=70, height=28, corner_radius=4,
                              font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))
                more.bind("<Escape>", lambda e: more.destroy())

        else:
            # Edit mode — plain entry, no + button
            pkg_ent = ctk.CTkEntry(g, textvariable=pkg_var, height=32,
                                   fg_color=INPUT, border_color=BORDER, border_width=1,
                                   corner_radius=4, text_color=DARK,
                                   font=(MODERN_FONT, 11))
            pkg_ent.grid(row=3, column=0, sticky="ew", padx=(0, 4), pady=(0, 2))
            self._attach_entry_context_menu(pkg_ent)
            # Row 4 is empty in edit mode — grid collapses it to zero height.

        # Row 5-6: Manifest Date | DEC #
        _lbl("Manifest Date (YYYY-MM-DD)", 5, 0, cs=1); manifest_var = _entry("manifest",  6, 0)
        _lbl("DEC #",         5, 1, cs=1); dec_var      = _entry("dec",       6, 1)
        # Row 7-8: Reason (full width)
        _lbl("Reason", 7, 0, cs=2)
        reason_var = _combo("reason", 8, 0, REASON_VALUES, cs=2)
        # Row 9-10: COLS Payment Status (full width)
        _lbl("COLS Payment Status", 9, 0, cs=2)
        paid_var = _combo("paid", 10, 0, PAID_VALUES, cs=2)
        # Row 11: cascade hint
        ctk.CTkLabel(g,
            text="⚠  This status is shared across all packages with the same DEC # — saving will update them all",
            font=(MODERN_FONT, 10), text_color="#ffcc44",
            anchor="w", wraplength=424).grid(
            row=11, column=0, columnspan=2, sticky="w", padx=(0, 4), pady=(1, 5))
        # Row 12-13: Notes (multiline)
        _lbl("Notes / Comments", 12, 0, cs=2)
        notes_box = ctk.CTkTextbox(g, height=80, fg_color=INPUT,
                                    border_color=BORDER, border_width=1,
                                    corner_radius=4, text_color=DARK,
                                    font=(MODERN_FONT, 11))
        notes_box.grid(row=13, column=0, columnspan=2, sticky="ew",
                       padx=(0, 4), pady=(0, 6))
        self._attach_textbox_context_menu(notes_box)
        notes_box.insert("0.0", init["notes"])

        # ---- Save / Cancel (btn_frame already packed at bottom above) ----
        def on_save():
            new_store    = store_var.get().strip()
            new_cby      = cby_var.get().strip()
            new_package  = pkg_var.get().strip()
            new_still    = still_var.get().strip() or "Yes"
            new_manifest = manifest_var.get().strip()
            new_dec      = dec_var.get().strip()
            new_reason   = reason_var.get().strip()
            new_paid     = paid_var.get().strip()
            new_notes    = notes_box.get("0.0", "end").strip()

            tags = ("not_add",) if new_still.lower() == "no" else ()

            if is_edit:
                new_vals = (new_store, new_cby, new_package, new_still, new_manifest,
                            new_dec, new_reason, new_paid, new_notes)
                self._tree.item(item_id, values=new_vals, tags=tags)
                exclude_ids = {item_id}
                count_added = 1
            else:
                # Build all package rows to insert (primary + any extras from + button)
                all_pkgs = [p for p in ([new_package] + extra_pkgs) if p] or [new_package]
                new_iids = set()
                last_iid = None
                for pkg_num in all_pkgs:
                    row_vals = (new_store, new_cby, pkg_num, new_still, new_manifest,
                                new_dec, new_reason, new_paid, new_notes)
                    iid = self._tree.insert("", "end", values=row_vals, tags=tags)
                    new_iids.add(iid)
                    last_iid = iid
                if last_iid:
                    self._tree.see(last_iid)
                    self._tree.selection_set(last_iid)
                exclude_ids = new_iids
                count_added = len(all_pkgs)

            # Cascade COLS Payment Status to every other row with the same DEC#
            cascaded = 0
            if new_dec:
                for item in self._tree.get_children():
                    if item in exclude_ids:
                        continue
                    iv = list(self._tree.item(item, "values"))
                    while len(iv) < len(COLUMNS):
                        iv.append("")
                    if str(iv[5]).strip() == new_dec:
                        iv[7] = new_paid
                        self._tree.item(item, values=iv)
                        cascaded += 1

            self._sync_tree_to_rows()
            self._dirty = True
            self._update_delete_not_add_button()
            if is_edit:
                verb = "Updated"
            else:
                verb = f"Added {count_added} package{'s' if count_added > 1 else ''}"
            if cascaded:
                self._set_status(
                    f"{verb} — '{new_paid}' applied to {cascaded + count_added} packages under DEC# {new_dec}")
            else:
                self._set_status(f"{verb} — click Save to persist")
            _close()

        ctk.CTkButton(btn_frame,
                      text="Save" if is_edit else "Add Package",
                      command=on_save,
                      fg_color=ACCENT if is_edit else GREEN,
                      hover_color=ACCENT_H if is_edit else GREEN_H,
                      width=140, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        ctk.CTkButton(btn_frame, text="Cancel", command=_close,
                      fg_color="#667788", hover_color="#556677",
                      width=90, height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

        dlg.bind("<Escape>", lambda e: _close())

    def _delete_selected(self):
        sel = self._tree.selection()
        if not sel:
            self._set_status("No row selected")
            return
        count = len(sel)
        # Gather package numbers for all selected rows
        pkgs = []
        for item in sel:
            vals = self._tree.item(item, "values")
            pkgs.append(str(vals[2]) if len(vals) > 2 else "?")
        pkg_list = ", ".join(pkgs[:5])
        if count > 5:
            pkg_list += f" ... and {count - 5} more"
        if not messagebox.askyesno("Confirm Delete",
                                   f"Delete {count} package(s)?\n\n{pkg_list}"):
            return
        # Remove from self._rows by package number
        pkg_set = set(p.strip() for p in pkgs)
        self._rows = [r for r in self._rows
                      if str(r.get("package", "")).strip() not in pkg_set]
        # Remove from tree
        for item in sel:
            self._tree.delete(item)
        self._dirty = True
        self._set_status(f"Deleted {count} package(s)")

    def _update_delete_not_add_button(self):
        """Show the 'Delete All No Longer ADD' button if any rows have still_add='No'."""
        has_no = any(str(r.get("still_add", "")).strip().lower() == "no"
                     for r in self._rows)
        if has_no:
            if not self._delete_not_add_btn.winfo_ismapped():
                self._delete_not_add_btn.pack(side="left", padx=4)
        else:
            if self._delete_not_add_btn.winfo_ismapped():
                self._delete_not_add_btn.pack_forget()

    def _delete_all_not_add(self):
        """Delete all rows where Still ADD? = No."""
        no_rows = [r for r in self._rows if str(r.get("still_add", "")).strip().lower() == "no"]
        if not no_rows:
            return
        count = len(no_rows)
        if not messagebox.askyesno("Confirm Delete",
                                   f"Delete {count} package(s) no longer at ADD status?"):
            return
        self._rows = [r for r in self._rows if str(r.get("still_add", "")).strip().lower() != "no"]
        self._populate_tree()
        self._dirty = True
        self._set_status(f"Deleted {count} package(s) no longer at ADD status")

    # ------------------------------------------------------------------
    # Fill DEC# from Manifest
    # ------------------------------------------------------------------
    def _normalize_reason(self, raw):
        """Normalize manifest reason variants to our dropdown values.
        'Held', 'HOLD', 'hold+invoice', 'held+inspection' etc. all map to
        the closest matching dropdown option."""
        if not raw:
            return "Held"  # default if blank
        r = str(raw).lower().strip()
        # Contains both held and inspection
        if ("held" in r or "hold" in r) and "inspect" in r:
            return "Held+Inspection"
        # Inspection only
        if "inspect" in r:
            return "Inspection"
        # Any held/hold variant (hold, held, hold+invoice, etc.)
        if "held" in r or "hold" in r:
            return "Held"
        # Unrecognized — default to Held
        return "Held"

    def _fill_from_manifest(self):
        """Open a file picker for a manifest xlsx, read the Bags sheet,
        and fill in DEC# and Reason for SELECTED rows by matching package
        number. Read-only — does not modify the manifest file."""
        sel = self._tree.selection()
        if not sel:
            messagebox.showwarning("No Selection",
                                   "Select one or more rows in the list first,\n"
                                   "then use this button to fill in DEC# and Reason\n"
                                   "from a matching manifest.")
            return

        # Collect selected package numbers
        selected_pkgs = {}
        for item in sel:
            vals = self._tree.item(item, "values")
            pkg = str(vals[2]).strip() if len(vals) > 2 else ""
            if pkg:
                selected_pkgs[pkg] = item

        filepath = filedialog.askopenfilename(
            title="Select Manifest File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=str(INSPECTION_DIR))
        if not filepath:
            return

        try:
            prog_win, bar = self._show_progress("Reading manifest...")
            self._pulse_progress(bar, prog_win)
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            if "Bags" not in wb.sheetnames:
                messagebox.showerror("No Bags Sheet",
                                     "This file does not have a 'Bags' sheet.")
                wb.close()
                return

            ws = wb["Bags"]

            # Read master declaration number (row 1, col E = column 5)
            master_dec_val = ws.cell(row=1, column=5).value
            master_dec = str(master_dec_val).strip() if master_dec_val and str(master_dec_val).strip() != "0" else ""

            # Build a lookup from the Bags sheet: package number -> (dec, reason)
            # Column K (11) = "Packages" = the tracking/package number that
            # matches our console's Package # column. Column D (4) = "Bag Number"
            # which is the EBL bag number, not the tracking number.
            manifest_lookup = {}
            for row_idx in range(11, ws.max_row + 1):
                package_num = ws.cell(row=row_idx, column=11).value  # col K = Packages
                declaration = ws.cell(row=row_idx, column=8).value   # col H = Declaration
                reason = ws.cell(row=row_idx, column=9).value        # col I = Reason

                pkg = str(package_num).strip() if package_num else ""
                if not pkg or pkg == "0":
                    continue

                dec = str(declaration).strip() if declaration and str(declaration).strip() != "0" else master_dec
                reason_norm = self._normalize_reason(reason)

                # First match wins (a package could appear on multiple rows
                # if it has multiple boxes — they share the same dec/reason)
                if pkg not in manifest_lookup:
                    manifest_lookup[pkg] = (dec, reason_norm)

            wb.close()

            # Now update selected rows
            updated = 0
            not_found = 0
            for pkg, item in selected_pkgs.items():
                if pkg in manifest_lookup:
                    dec, reason_norm = manifest_lookup[pkg]
                    vals = list(self._tree.item(item, "values"))
                    while len(vals) < len(COLUMNS):
                        vals.append("")
                    # col 5 = DEC#, col 6 = Reason (0-indexed)
                    if dec:
                        vals[5] = dec
                    if reason_norm:
                        vals[6] = reason_norm
                    self._tree.item(item, values=vals)
                    updated += 1
                else:
                    not_found += 1

            self._close_progress(prog_win)
            self._sync_tree_to_rows()
            self._dirty = True

            if updated and not not_found:
                self._set_status(f"Filled DEC# and Reason for {updated} row(s) from manifest")
                messagebox.showinfo("Updated",
                                    f"Updated {updated} row(s) with DEC# and Reason from the manifest.\n\n"
                                    f"Click Save to persist.")
            elif updated and not_found:
                self._set_status(f"Updated {updated} row(s); {not_found} not found in manifest")
                messagebox.showinfo("Partially Updated",
                                    f"Updated {updated} row(s) with DEC# and Reason.\n"
                                    f"{not_found} package(s) were not found in this manifest.\n\n"
                                    f"Click Save to persist.")
            else:
                self._set_status(f"No matching packages found in manifest")
                messagebox.showinfo("No Matches",
                                    f"None of the {len(selected_pkgs)} selected package(s) were found\n"
                                    f"in this manifest's Bags sheet.")

        except Exception as e:
            try:
                self._close_progress(prog_win)
            except Exception:
                pass
            self._set_status(f"Manifest import error: {e}")
            messagebox.showerror("Import Error", f"Could not read manifest file:\n{e}")

    # ------------------------------------------------------------------
    # Update ADD Status (from export-porestatus CSV)
    # ------------------------------------------------------------------
    def _update_add_status(self):
        """Import the export-porestatus CSV to check which packages are
        still at ADD status. Flags packages no longer ADD, adds new ones."""
        csv_path = filedialog.askopenfilename(
            title="Select export-porestatus CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir=str(INSPECTION_DIR))
        if not csv_path:
            return

        prog_win, bar = self._show_progress("Reading export-porestatus CSV...")
        self._pulse_progress(bar, prog_win)

        try:
            import csv as csv_mod
            csv_packages = {}
            with open(csv_path, "r", encoding="utf-8-sig") as f:
                reader = csv_mod.DictReader(f)
                for row in reader:
                    pkg = str(row.get("GUIA", "")).strip()
                    if not pkg:
                        continue
                    tienda = str(row.get("TIENDA", "")).strip().lower()
                    if "camana" in tienda:
                        store = "CB"
                    elif "harbour" in tienda or "harbor" in tienda:
                        store = "HW"
                    else:
                        store = "CB"
                    cliente = str(row.get("CLIENTE", "")).strip()
                    cby = cliente.replace("CBY", "").strip() if cliente else ""
                    fecha = str(row.get("FECHA_PROCESO", "")).strip()
                    if fecha:
                        try:
                            mdate = datetime.strptime(fecha.split(" ")[0], "%Y-%m-%d")
                            manifest_str = mdate.strftime("%Y-%m-%d")
                        except Exception:
                            manifest_str = fecha
                    else:
                        manifest_str = ""
                    csv_packages[pkg] = {
                        "store": store, "cby": cby, "manifest": manifest_str
                    }
        except Exception as e:
            self._close_progress(prog_win)
            self._set_status(f"Error reading CSV: {e}")
            messagebox.showerror("CSV Error", f"Could not read the CSV file:\n{e}")
            return

        # Update existing rows
        existing_pkgs = set()
        flagged_not_add = 0
        for row in self._rows:
            pkg = str(row.get("package", "")).strip()
            existing_pkgs.add(pkg)
            if pkg in csv_packages:
                row["still_add"] = "Yes"
                if not row.get("manifest") and csv_packages[pkg]["manifest"]:
                    row["manifest"] = csv_packages[pkg]["manifest"]
            else:
                if row.get("still_add", "").lower() != "no":
                    row["still_add"] = "No"
                    flagged_not_add += 1

        # Add new packages from CSV
        added = 0
        for pkg, info in csv_packages.items():
            if pkg not in existing_pkgs:
                self._rows.append({
                    "store":     info["store"],
                    "cby":       info["cby"],
                    "package":   pkg,
                    "still_add": "Yes",
                    "manifest":  info["manifest"],
                    "dec":       "",
                    "reason":    "",
                    "paid":      "Not yet paid",
                    "notes":     "",
                })
                existing_pkgs.add(pkg)
                added += 1

        self._close_progress(prog_win)
        self._populate_tree()
        self._dirty = True

        summary = (f"ADD Status Update complete:\n\n"
                   f"  - {flagged_not_add} package(s) flagged 'No' (no longer ADD)\n"
                   f"  - {added} new package(s) added from export\n\n")
        if flagged_not_add:
            summary += ("Use 'Delete All No Longer ADD' to remove flagged rows,\n"
                        "or delete them individually.\n\n")
        summary += "Click Save to persist changes."
        self._set_status(f"ADD Status: {flagged_not_add} flagged, {added} added")
        messagebox.showinfo("Update Complete", summary)

    # ------------------------------------------------------------------
    # Update COLS Payment (from Declarations export)
    # ------------------------------------------------------------------
    def _update_cols_payment(self):
        """Import the Declarations export to check duty payment status.
        Marks packages as 'Duties Paid' if Status=Paid, or 'Check COLS'
        if older than 2 months and not in the export."""
        dec_path = filedialog.askopenfilename(
            title="Select Declarations export",
            filetypes=[("Excel files", "*.xls;*.xlsx"), ("All files", "*.*")],
            initialdir=str(INSPECTION_DIR))
        if not dec_path:
            return

        prog_win, bar = self._show_progress("Reading Declarations export...")
        self._pulse_progress(bar, prog_win)

        try:
            import pandas as pd
            dfs = pd.read_html(dec_path)
            df = dfs[0]
            dec_paid = {}
            for _, row in df.iterrows():
                bill = str(row.get("Bill Number", "")).strip()
                status = str(row.get("Status", "")).strip().lower()
                if bill and bill != "nan" and "paid" in status:
                    dec_paid[bill] = True
        except Exception as e:
            self._close_progress(prog_win)
            self._set_status(f"Error reading Declarations: {e}")
            messagebox.showerror("Declarations Error",
                                 f"Could not read the Declarations file:\n{e}\n\n"
                                 f"Note: This file must be an HTML-format xls.\n"
                                 f"If it's a real xlsx, make sure pandas + lxml are installed.")
            return

        # Update existing rows
        updated_paid = 0
        check_cols = 0
        for row in self._rows:
            pkg = str(row.get("package", "")).strip()
            if pkg in dec_paid:
                if row.get("paid", "") != "Duties Paid":
                    row["paid"] = "Duties Paid"
                    updated_paid += 1
            else:
                # Not in declarations file — if manifest is 2+ months old,
                # mark as "Check COLS" (too old to be in the 2-month export)
                manifest_str = str(row.get("manifest", "")).strip()
                try:
                    if manifest_str:
                        mdate = datetime.strptime(manifest_str, "%Y-%m-%d")
                        age_days = (datetime.now() - mdate).days
                        if age_days > 60 and row.get("paid", "") not in ("Duties Paid", "Pay and collect", "Inspect and collect"):
                            row["paid"] = "Check COLS"
                            check_cols += 1
                except Exception:
                    pass

        self._close_progress(prog_win)
        self._populate_tree()
        self._dirty = True

        summary = (f"COLS Payment Update complete:\n\n"
                   f"  - {updated_paid} package(s) updated to 'Duties Paid'\n"
                   f"  - {check_cols} package(s) marked 'Check COLS' (older than 2 months)\n\n"
                   f"Click Save to persist changes.")
        self._set_status(f"COLS Payment: {updated_paid} paid, {check_cols} check COLS")
        messagebox.showinfo("Update Complete", summary)

    # ------------------------------------------------------------------
    # Export PDF
    # ------------------------------------------------------------------
    def _export_pdf(self):
        if not _FPDF_AVAILABLE:
            messagebox.showerror("Missing Library",
                                 "fpdf2 is not installed.\nRun: pip install fpdf2")
            return
        if not self._rows:
            messagebox.showwarning("No Data", "No rows to export.")
            return

        # Themed export dialog
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Export PDF")
        dialog.configure(fg_color=BG)
        dialog.transient(self.root)
        dialog.grab_set()
        dw, dh = 420, 280
        sw, sh = dialog.winfo_screenwidth(), dialog.winfo_screenheight()
        dialog.geometry(f"{dw}x{dh}+{int((sw-dw)/2)}+{int((sh-dh)/2)}")
        dialog.resizable(False, False)

        ctk.CTkLabel(dialog, text="Export PDF",
                     font=(MODERN_FONT, 18, "bold"), text_color=LIGHT).pack(pady=(20, 4))
        ctk.CTkLabel(dialog, text="Choose which list to export:",
                     font=(MODERN_FONT, 12), text_color=TEXT).pack(pady=(0, 16))

        choice = {"value": None}

        def _do_export(which):
            choice["value"] = which
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=(0, 8))

        ctk.CTkButton(btn_frame, text="Full List",
                      command=lambda: _do_export("full"),
                      fg_color=ACCENT, hover_color=ACCENT_H,
                      width=180, height=34, corner_radius=6,
                      font=(MODERN_FONT, 12, "bold")).pack(pady=4)
        ctk.CTkButton(btn_frame, text="GrabnGo List",
                      command=lambda: _do_export("grabngo"),
                      fg_color=GREEN, hover_color=GREEN_H,
                      width=180, height=34, corner_radius=6,
                      font=(MODERN_FONT, 12, "bold")).pack(pady=4)
        ctk.CTkButton(btn_frame, text="Inspections List",
                      command=lambda: _do_export("inspections"),
                      fg_color=ORANGE, hover_color=ORANGE_H,
                      width=180, height=34, corner_radius=6,
                      font=(MODERN_FONT, 12, "bold")).pack(pady=4)

        ctk.CTkLabel(dialog,
                     text="GrabnGo: Held packages ready for driver  -  Inspections: Inspection-only, duties paid",
                     font=(MODERN_FONT, 9), text_color=MUTED).pack(pady=(4, 0))

        self.root.wait_window(dialog)

        which = choice["value"]
        if not which:
            return

        # Filter rows based on choice
        if which == "full":
            export_rows = self._rows
            export_title = "Packages at Customs - Full List"
            default_name = f"Packages at Customs - {datetime.now().strftime('%Y-%m-%d')}.pdf"
        elif which == "grabngo":
            grabngo_paid = {"Duties Paid", "Pay and collect", "Inspect and collect"}
            grabngo_reasons = {"Held", "Held+Inspection"}
            export_rows = [r for r in self._rows
                           if str(r.get("reason", "")).strip() in grabngo_reasons
                           and str(r.get("paid", "")).strip() in grabngo_paid]
            export_title = "Packages at Customs - GrabnGo List"
            default_name = f"GrabnGo - {datetime.now().strftime('%Y-%m-%d')}.pdf"
        else:  # inspections
            export_rows = [r for r in self._rows
                           if str(r.get("reason", "")).strip() == "Inspection"
                           and str(r.get("paid", "")).strip() == "Duties Paid"]
            export_title = "Packages at Customs - Inspections List"
            default_name = f"Inspections - {datetime.now().strftime('%Y-%m-%d')}.pdf"

        if not export_rows:
            messagebox.showinfo("No Matching Packages",
                                "No packages match the criteria for this export.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Save PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialdir=str(INSPECTION_DIR),
            initialfile=default_name)
        if not filepath:
            return

        try:
            prog_win, bar = self._show_progress("Generating PDF...")
            self._pulse_progress(bar, prog_win)
            self.root.update_idletasks()
            self._generate_pdf(filepath, export_rows, export_title)
            self._close_progress(prog_win)
            self._set_status(f"PDF saved: {Path(filepath).name} ({len(export_rows)} rows)")
            if messagebox.askyesno("PDF Saved", f"Saved to:\n{filepath}\n\nOpen it now?"):
                os.startfile(filepath)
        except Exception as e:
            try:
                self._close_progress(prog_win)
            except Exception:
                pass
            self._set_status(f"PDF error: {e}")
            messagebox.showerror("PDF Export Error", str(e))

    def _generate_pdf(self, filepath, rows=None, title="Packages at Customs - ADD Status Log"):
        """Generate a clean, legible PDF for the driver."""
        if rows is None:
            rows = self._rows
        pdf = FPDF(orientation="L", format="A4")  # Landscape for wide table
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # ---- Logo (inverted: logo is white, invert so it shows on white page) ----
        try:
            _logo_img = Image.open(io.BytesIO(base64.b64decode(MBE_LOGO_B64))).convert("RGBA")
            _r, _g, _b, _a = _logo_img.split()
            _inv_rgb = ImageOps.invert(Image.merge("RGB", (_r, _g, _b)))
            _inv_logo = Image.merge("RGBA", (*_inv_rgb.split(), _a))
            _logo_w_mm = 38  # display width in mm
            _logo_h_mm = _logo_w_mm * _logo_img.size[1] / _logo_img.size[0]
            _logo_x = pdf.w - pdf.l_margin - _logo_w_mm
            _logo_y = 10
            import tempfile
            import os as _os
            _tf = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            _tf.close()
            _inv_logo.save(_tf.name, "PNG")
            pdf.image(_tf.name, x=_logo_x, y=_logo_y, w=_logo_w_mm, h=_logo_h_mm)
            _logo_bottom = _logo_y + _logo_h_mm
            try:
                _os.unlink(_tf.name)
            except OSError:
                pass
        except Exception:
            _logo_bottom = 0

        # ---- Header ----
        pdf.set_font(PDF_FONT, "B", 16)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="L")
        pdf.set_font(PDF_FONT, "", 10)
        pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                 new_x="LMARGIN", new_y="NEXT", align="L")
        pdf.ln(4)

        # If the logo extends below the header text, push the table down to clear it
        if pdf.get_y() < _logo_bottom + 2:
            pdf.set_y(_logo_bottom + 2)

        # ---- Table ----
        col_widths = [14, 22, 45, 28, 38, 30, 38, 32, 0]  # last column takes remaining
        # Calculate remaining width for notes
        used = sum(w for w in col_widths[:-1])
        page_w = pdf.w - 2 * pdf.l_margin
        col_widths[-1] = page_w - used

        headers = [c[1] for c in COLUMNS]

        # Header row
        pdf.set_font(PDF_FONT, "B", 9)
        pdf.set_fill_color(15, 17, 23)
        pdf.set_text_color(232, 232, 232)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 8, h, border=1, fill=True, align="L")
        pdf.ln()

        # Data rows
        pdf.set_font(PDF_FONT, "", 9)
        pdf.set_text_color(40, 40, 40)
        fill = False
        # Helper: replace Unicode chars not in Latin-1 with ASCII equivalents
        def _pdf_safe(v):
            return (str(v)
                    .replace("\u2014", "-")   # em dash
                    .replace("\u2013", "-")   # en dash
                    .replace("\u2019", "'")   # right single quote
                    .replace("\u2018", "'")   # left single quote
                    .replace("\u201c", '"')   # left double quote
                    .replace("\u201d", '"')   # right double quote
                    .replace("\u2026", "...") # ellipsis
                    .replace("\u00a0", " ")   # non-breaking space
                    .encode("latin-1", "replace").decode("latin-1"))
        for row in rows:
            if fill:
                pdf.set_fill_color(240, 240, 245)
            else:
                pdf.set_fill_color(255, 255, 255)

            vals = [
                _pdf_safe(row.get("store", "")),
                _pdf_safe(row.get("cby", "")),
                _pdf_safe(row.get("package", "")),
                _pdf_safe(row.get("still_add", "")),
                _pdf_safe(row.get("manifest", "")),
                _pdf_safe(row.get("dec", "")),
                _pdf_safe(row.get("reason", "")),
                _pdf_safe(row.get("paid", "")),
                _pdf_safe(row.get("notes", "")),
            ]

            # Calculate row height needed (notes column may wrap)
            notes_w = col_widths[-1]
            max_lines = 1
            for i, v in enumerate(vals):
                if col_widths[i] > 0:
                    # Estimate how many lines this cell needs
                    text_w = pdf.get_string_width(v)
                    lines = max(1, int(text_w / (col_widths[i] - 2)) + 1)
                    max_lines = max(max_lines, lines)
            row_h = 6 * max_lines + 2

            # Check if we need a page break
            if pdf.get_y() + row_h > pdf.h - 15:
                pdf.add_page()
                # Repeat header
                pdf.set_font(PDF_FONT, "B", 9)
                pdf.set_fill_color(15, 17, 23)
                pdf.set_text_color(232, 232, 232)
                for i, h in enumerate(headers):
                    pdf.cell(col_widths[i], 8, h, border=1, fill=True, align="L")
                pdf.ln()
                pdf.set_font(PDF_FONT, "", 9)
                pdf.set_text_color(40, 40, 40)

            # Draw cells
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            for i, v in enumerate(vals):
                x = pdf.get_x()
                y = pdf.get_y()
                if i == len(vals) - 1:
                    # Notes column — use multi_cell for wrapping
                    pdf.multi_cell(col_widths[i], 6, v, border=1, align="L", fill=fill)
                else:
                    pdf.cell(col_widths[i], row_h, v, border=1, align="L", fill=fill)
            # Move to next row
            pdf.set_xy(x_start, y_start + row_h)
            fill = not fill

        # ---- Footer ----
        pdf.ln(6)
        pdf.set_font(PDF_FONT, "I", 8)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 5, f"Total packages: {len(rows)}", align="L")

        pdf.output(filepath)

    # ------------------------------------------------------------------
    # Remote support: bug report + apply fixes
    # ------------------------------------------------------------------
    def _check_update_bg(self):
        result = _check_for_update()
        if result:
            self._pending_update = result
            try:
                self.root.after(0, self._refresh_support_icon)
            except Exception:
                pass

    def _refresh_support_icon(self):
        try:
            if self._pending_update:
                ver = self._pending_update.get("version", "?")
                self._support_btn.configure(text="Apply Fixes", fg_color="#b8860b",
                                            hover_color="#daa520",
                                            font=(MODERN_FONT, 11, "bold"),
                                            width=90)
                self._support_tooltip = f"Apply Fixes — v{ver} available"
            else:
                self._support_btn.configure(text="🐞", fg_color=BG,
                                            hover_color="#24507a",
                                            font=("Segoe UI Emoji", 15),
                                            width=34)
                self._support_tooltip = "Report a Bug"
        except Exception:
            pass

    def _on_support_click(self):
        self._hide_tooltip()
        if self._pending_update:
            self._apply_fixes_dialog()
        else:
            self._report_bug_dialog()

    def _report_bug_dialog(self):
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Report a Bug")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 620
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(dlg, text="Report a Bug", font=(MODERN_FONT, 15, "bold"),
                     text_color=LIGHT).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Describe the problem in as much detail as you can — what you\n"
                 f"did, what happened, and what you expected. This goes directly\n"
                 f"to the developer ({DEVELOPER_NAME}).",
            font=(MODERN_FONT, 11), text_color=MUTED,
            anchor="w", justify="left").pack(anchor="w", padx=16, pady=(0, 6))

        # ---- Check for Updates bar ----
        update_bar = ctk.CTkFrame(dlg, fg_color="transparent")
        update_bar.pack(fill="x", padx=16, pady=(0, 6))

        check_btn = ctk.CTkButton(
            update_bar, text="Check for Updates",
            command=lambda: None,
            fg_color=ACCENT, hover_color=ACCENT_H,
            width=140, height=26, corner_radius=4,
            font=(MODERN_FONT, 10, "bold"))
        check_btn.pack(side="left")

        update_status = ctk.CTkLabel(
            update_bar, text="",
            font=(MODERN_FONT, 10), text_color=MUTED,
            anchor="w")
        update_status.pack(side="left", padx=(10, 0))

        _spin = {"frame": 0, "running": False}

        def _spin_step():
            if not _spin["running"]:
                return
            _spin["frame"] = (_spin["frame"] + 1) % 8
            dots = "." * ((_spin["frame"] % 4) + 1)
            update_status.configure(text=f"Checking{dots}")
            dlg.after(200, _spin_step)

        def _do_check():
            _spin["running"] = True
            check_btn.configure(state="disabled", text="Checking...")
            _spin_step()

            def _worker():
                result = _check_for_update()
                dlg.after(0, lambda: _check_done(result))

            threading.Thread(target=_worker, daemon=True).start()

        def _check_done(result):
            _spin["running"] = False
            check_btn.configure(state="normal", text="Check for Updates")
            if result:
                ver = result.get("version", "?")
                self._pending_update = result
                update_status.configure(
                    text=f"v{ver} available!",
                    text_color=LIGHT)
                check_btn.configure(
                    text=f"Update Now — v{ver}",
                    command=self._apply_fixes_dialog,
                    fg_color="#b8860b", hover_color="#daa520",
                    width=180)
            else:
                update_status.configure(
                    text=f"You're up to date (v{APP_VERSION})",
                    text_color=MUTED)

        check_btn.configure(command=_do_check)

        # Category selector (single-select)
        cat_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        cat_frame.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(cat_frame, text="Category:",
                     font=(MODERN_FONT, 11, "bold"), text_color=LIGHT,
                     anchor="w").pack(anchor="w", pady=(0, 2))
        cat_var = ctk.StringVar(value="Bug Fix")
        cat_row = ctk.CTkFrame(cat_frame, fg_color="transparent")
        cat_row.pack(fill="x")
        for cat in ("Bug Fix", "Feature Request", "Environmental Change", "Other"):
            ctk.CTkRadioButton(cat_row, text=cat, variable=cat_var, value=cat,
                               font=(MODERN_FONT, 10), text_color=LIGHT,
                               fg_color=ACCENT, hover_color=ACCENT_H,
                               border_color=BORDER).pack(side="left", padx=(0, 12))

        box = ctk.CTkTextbox(dlg, height=140, fg_color=INPUT, border_color=BORDER,
                             border_width=1, corner_radius=4, text_color=DARK,
                             font=(MODERN_FONT, 11))
        box.pack(fill="both", expand=True, padx=16)
        self._attach_textbox_context_menu(box)
        box.focus_set()

        # Optional email field for follow-up
        email_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        email_frame.pack(fill="x", padx=16, pady=(4, 0))
        ctk.CTkLabel(email_frame, text="Your email (optional — for updates on this report)",
                     font=(MODERN_FONT, 10), text_color=MUTED,
                     anchor="w").pack(anchor="w")
        email_var = ctk.StringVar(value="")
        email_entry = ctk.CTkEntry(email_frame, textvariable=email_var, height=28,
                                   fg_color=INPUT, border_color=BORDER, border_width=1,
                                   corner_radius=4, text_color=DARK,
                                   font=(MODERN_FONT, 11))
        email_entry.pack(fill="x", pady=(1, 0))
        self._attach_entry_context_menu(email_entry)

        # "What to expect" info box
        info = ctk.CTkFrame(dlg, fg_color="#1a2a3a", corner_radius=6)
        info.pack(fill="x", padx=16, pady=(6, 8))
        ctk.CTkLabel(info, text="What to expect",
                     font=(MODERN_FONT, 11, "bold"), text_color="#88ccff",
                     anchor="w").pack(anchor="w", padx=10, pady=(8, 2))
        ctk.CTkLabel(info,
            text=f"\u2022 {DEVELOPER_NAME} will review your report within 24-48 hours\n"
                 f"\u2022 When a fix is ready, click 'Check for Updates' above\n"
                 f"   or relaunch the console \u2014 the bug icon will show 'Apply Fixes'\n"
                 f"\u2022 {DEVELOPER_NAME} will coordinate with management if the fix\n"
                 f"   requires substantial work",
            font=(MODERN_FONT, 10), text_color=MUTED,
            anchor="w", justify="left").pack(anchor="w", padx=10, pady=(0, 8))

        def _next():
            desc = box.get("0.0", "end").strip()
            if not desc:
                messagebox.showwarning("Empty", "Please describe the bug first.")
                return
            email = email_var.get().strip()
            category = cat_var.get()
            dlg.destroy()
            self._show_attach_files_dialog(desc, email, category)

        ctk.CTkButton(btns, text="Next", command=_next,
                      fg_color=GREEN, hover_color=GREEN_H, width=100,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        ctk.CTkButton(btns, text="Cancel", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    def _show_attach_files_dialog(self, description, reporter_email="", category="Bug Fix"):
        """After the bug description, offer to attach sample files.
        The actual Discord send happens here — text + files in ONE message."""
        case_num = _generate_case_number()
        dlg = ctk.CTkToplevel(self.root)
        dlg.title(f"Attach Files? (Case {case_num})")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 420
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(dlg, text="Attach Sample Files?",
                     font=(MODERN_FONT, 15, "bold"), text_color=LIGHT).pack(
            anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Based on your description, {DEVELOPER_NAME} may need sample\n"
                 f"files to reproduce the issue in a test environment.\n\n"
                 f"The console script itself is always attached automatically\n"
                 f"so {DEVELOPER_NAME} gets your exact version for testing.\n\n"
                 f"If you'd like, attach additional Excel or CSV files below.\n"
                 f"This is completely optional and files are sent privately\n"
                 f"alongside your bug report.",
            font=(MODERN_FONT, 11), text_color=MUTED,
            anchor="w", justify="left").pack(anchor="w", padx=16, pady=(0, 8))

        # File list area
        file_list_frame = ctk.CTkFrame(dlg, fg_color=INPUT, corner_radius=4,
                                       border_width=1, border_color=BORDER)
        file_list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        file_list_label = ctk.CTkLabel(file_list_frame, text="No files selected",
                                       font=(MODERN_FONT, 11), text_color=MUTED,
                                       anchor="w", justify="left")
        file_list_label.pack(anchor="w", padx=10, pady=10)

        selected_files = []

        def _pick_files():
            paths = filedialog.askopenfilenames(
                title="Select files to attach",
                filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv"),
                           ("All files", "*.*")])
            if paths:
                selected_files.clear()
                selected_files.extend(paths)
                names = [Path(p).name for p in paths]
                display = "\n".join(names[:6])
                if len(names) > 6:
                    display += f"\n...and {len(names)-6} more"
                file_list_label.configure(text=display, text_color=DARK)

        def _submit():
            submit_btn.configure(state="disabled", text="Sending...")
            skip_btn.configure(state="disabled")

            def _worker():
                # Always attach the running script itself so the developer
                # gets the user's exact code + data for reproduction.
                all_files = [str(SCRIPT_PATH)] + list(selected_files)
                ok, err = _post_bug_report_with_files(
                    description, case_num, all_files, reporter_email, category)
                self.root.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                dlg.destroy()
                n_files = len(selected_files)
                if n_files:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted with {n_files} file(s).\n\n"
                        f"Your report and files have been sent to {DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' here.")
                else:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted.\n\n"
                        f"Your report has been sent to {DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' here.")
            else:
                submit_btn.configure(state="normal", text="Submit Report")
                skip_btn.configure(state="normal")
                messagebox.showerror("Could Not Send",
                    f"The report could not be sent:\n{err}\n\n"
                    "Check your internet connection and try again.")

        ctk.CTkButton(btns, text="Browse Files", command=_pick_files,
                      fg_color=ACCENT, hover_color=ACCENT_H, width=120,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        submit_btn = ctk.CTkButton(btns, text="Submit Report", command=_submit,
                                   fg_color=GREEN, hover_color=GREEN_H, width=130,
                                   height=30, corner_radius=5,
                                   font=(MODERN_FONT, 11, "bold"))
        submit_btn.pack(side="left", padx=(8, 0))
        skip_btn = ctk.CTkButton(btns, text="Skip & Send", command=_submit,
                                 fg_color="#667788", hover_color="#556677", width=110,
                                 height=30, corner_radius=5,
                                 font=(MODERN_FONT, 11))
        skip_btn.pack(side="left", padx=(8, 0))

    def _apply_fixes_dialog(self):
        upd = self._pending_update
        if not upd:
            return

        if self._dirty:
            resp = messagebox.askyesnocancel("Unsaved Changes",
                "You have unsaved changes. Save them before updating?\n\n"
                "Yes = save then update, No = discard and update, Cancel = stop.")
            if resp is None:
                return
            if resp:
                self._save_to_excel()

        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Apply Fixes")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 380
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ver = upd.get("version", "?")
        ctk.CTkLabel(dlg, text=f"Fixes Available  (v{ver})",
                     font=(MODERN_FONT, 15, "bold"), text_color=LIGHT).pack(
            anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg, text=f"From {DEVELOPER_NAME}. Click Apply Fixes to download,\n"
                     "then choose how to restart.",
                     font=(MODERN_FONT, 11), text_color=MUTED,
                     anchor="w").pack(anchor="w", padx=16, pady=(0, 8))

        box = ctk.CTkTextbox(dlg, height=170, fg_color=INPUT, border_color=BORDER,
                             border_width=1, corner_radius=4, text_color=DARK,
                             font=(MODERN_FONT, 11), wrap="word")
        box.pack(fill="both", expand=True, padx=16)
        box.insert("0.0", upd.get("changelog", "(no description provided)"))
        box.configure(state="disabled")

        def _apply():
            apply_btn.configure(state="disabled", text="Applying...")

            def _worker():
                ok, err = _download_and_apply_update(upd.get("url", ""))
                self.root.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                # Notify developer that this user applied the update
                old_ver = APP_VERSION
                new_ver = upd.get("version", "?")
                threading.Thread(
                    target=lambda: _post_update_applied(old_ver, new_ver),
                    daemon=True).start()
                # Clear the pending update so the bug icon goes back
                # to 🐞 — the update is on disk, they just need to
                # restart.
                self._pending_update = None
                try:
                    self.root.after(0, self._refresh_support_icon)
                except Exception:
                    pass
                dlg.destroy()
                if messagebox.askyesno("Update Applied",
                    f"Updated to version {new_ver}.\n\n"
                    "The console needs to restart to apply the changes.\n"
                    "Restart now?"):
                    self._restart_app()
            else:
                apply_btn.configure(state="normal", text="Apply Fixes")
                messagebox.showerror("Update Failed",
                    f"Could not apply the update:\n{err}\n\n"
                    "Check your internet connection and try again.")

        apply_btn = ctk.CTkButton(btns, text="Apply Fixes", command=_apply,
                                  fg_color="#b8860b", hover_color="#daa520",
                                  width=130, height=30, corner_radius=5,
                                  font=(MODERN_FONT, 11, "bold"))
        apply_btn.pack(side="left")
        ctk.CTkButton(btns, text="Later", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    def _restart_app(self):
        try:
            subprocess.Popen([sys.executable, str(SCRIPT_PATH)])
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass
        sys.exit(0)

    # ------------------------------------------------------------------
    # Status / close
    # ------------------------------------------------------------------
    def _set_status(self, msg):
        self._status.configure(text=msg)

    def _on_close(self):
        if self._dirty:
            if not messagebox.askyesno("Unsaved Changes",
                                       "You have unsaved changes. Close anyway?"):
                return
        self.root.destroy()

    def run(self):
        self.root.mainloop()


# ==============================================================================
# MAIN
# ==============================================================================
if __name__ == "__main__":
    app = CustomsConsole()
    app.run()
