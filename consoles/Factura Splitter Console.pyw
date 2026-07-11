# -*- coding: utf-8 -*-
"""
Split Factura by CBY (Master Print)
Loads a manifest Excel file, displays CBY / Package# / Invoice columns,
and splits factura PDFs by CBY — with selective printing based on the
Invoice column notes (Y, dupe, corrupt, blank, etc.).
"""
import os
import sys
import re
import platform
import subprocess
import importlib
import json
import threading
import getpass
import urllib.request
import uuid
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import tkinter as tk
from tkinter import messagebox, filedialog
import tkinter.ttk as ttk

# ------------------------------------------------------------------
# Dependency check
# ------------------------------------------------------------------
REQUIRED_PACKAGES = {
    "customtkinter": "customtkinter",
    "openpyxl":      "openpyxl",
    "fitz":          "PyMuPDF",
}

def _check_and_install_dependencies():
    missing = []
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing.append((module_name, pip_name))
    if not missing:
        return True
    missing_names = [f"  - {pip_name}" for _, pip_name in missing]
    msg = ("The following packages are required but not installed:\n\n"
           + "\n".join(missing_names)
           + "\n\nWould you like to install them now via pip?")
    root = tk.Tk()
    root.withdraw()
    result = messagebox.askyesno("Missing Dependencies", msg)
    if not result:
        root.destroy()
        return False
    pip_args = [sys.executable, "-m", "pip", "install"] + [pn for _, pn in missing]
    try:
        result = subprocess.run(pip_args, capture_output=True, text=True)
        if result.returncode != 0:
            messagebox.showerror("Installation Failed",
                f"Could not install packages:\n\n{result.stderr[:500]}")
            root.destroy()
            return False
    except Exception as e:
        messagebox.showerror("Installation Failed", f"Error running pip:\n{e}")
        root.destroy()
        return False
    root.destroy()
    for module_name, _ in missing:
        try:
            importlib.import_module(module_name)
        except ImportError:
            messagebox.showerror("Still Missing",
                f"Package '{module_name}' still could not be imported\n"
                f"after installation. Please install it manually.")
            return False
    return True

if not _check_and_install_dependencies():
    sys.exit(1)

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import customtkinter as ctk
import openpyxl
import fitz  # PyMuPDF

# ------------------------------------------------------------------
# Platform font
# ------------------------------------------------------------------
def get_platform_font():
    system = platform.system()
    if system == "Windows":
        return "Segoe UI"
    elif system == "Darwin":
        return "SF Pro Display"
    return "Arial"

MODERN_FONT = get_platform_font()

# ------------------------------------------------------------------
# Colour scheme (matches Customs Console — blue)
# ------------------------------------------------------------------
BG       = "#1a3a5c"
PANEL    = "#ffffff"
INPUT    = "#ffffff"
BORDER   = "#d0d5dd"
DARK     = "#1a1a2e"
LIGHT    = "#ffffff"
TEXT     = "#e8eef5"
MUTED    = "#aabbcc"
ACCENT   = "#2e6fdb"
ACCENT_H = "#1a56c4"
GREEN    = "#16a34a"
GREEN_H  = "#15803d"
ORANGE   = "#ea7c1a"
ORANGE_H = "#c66812"
RED      = "#dc2626"
RED_H    = "#b91c1c"
ROW_ALT  = "#c5d8eb"

SCRIPT_PATH = Path(__file__).resolve()

# A4 width in points (8.27 inches) — all output pages are normalized to this width
A4_WIDTH = 595

# ==============================================================================
# REMOTE SUPPORT — bug reporting + self-update
# ==============================================================================
APP_NAME = "Factura Splitter Console"
APP_VERSION = "1.0.3"
DEVELOPER_NAME = "Atlas Ramoon"
DEVELOPER_EMAIL = "atlasramoon@gmail.com"

BUG_REPORT_WEBHOOK_URL = "https://discord.com/api/webhooks/1524620703259951104/fqpIEBXVWsKHy7f1iZ9xoryCpidmjPYIDuITfcwMOjBfMyS2HtJNWpVbfOetapl8vw9O"

UPDATE_MANIFEST_URL = (
    "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/"
    "manifests/factura-splitter-console.json"
)


def _version_tuple(v):
    out = []
    for p in str(v).split("."):
        try:
            out.append(int(p))
        except ValueError:
            out.append(0)
    return tuple(out)


def _http_get(url, timeout=10):
    req = urllib.request.Request(
        url, headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}",
                     "Cache-Control": "no-cache"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8")


def _post_to_discord(content):
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    payload = json.dumps({"content": content[:1900]}).encode("utf-8")
    try:
        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=payload,
            headers={"Content-Type": "application/json",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)


def _generate_case_number():
    year = datetime.now().year
    tag = uuid.uuid4().hex[:4].upper()
    return f"CASE-{year}-{tag}"


# Portal for remote support — downloaded from GitHub on demand.
PORTAL_URL = (
    "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/"
    "consoles/Python%20Portal%20for%20Atlas.pyw"
)


def _summon_portal(parent_root):
    """Download the Python Portal for Atlas to a user-chosen folder."""
    # Step 1: Confirm with explanation
    confirm = messagebox.askyesno(
        "Open a Portal for Atlas?",
        "This will open a remote IT support portal that lets Atlas\n"
        "diagnose and fix issues on your machine from afar.\n\n"
        "Peace of mind:\n"
        "Atlas cannot see your screen or control your mouse.\n"
        "He can only carry out file-level tasks such as reading\n"
        "nearby files, adding or replacing files, and running\n"
        "Python scripts you send.\n\n"
        "You'll choose where the problem is, then a small portal\n"
        "file will be saved there for you to open.\n\n"
        "Atlas will be notified that you've opened it.\n"
        "When the issue is resolved, you can close and delete it.\n\n"
        "Continue?")
    if not confirm:
        return

    # Step 2: Choose folder
    folder = filedialog.askdirectory(
        title="Where is the problem located? Choose a folder:")
    if not folder:
        return

    # Step 3: Download fresh portal with cache-busting (no raw CDN stale content)
    import time as _time
    is_mac = platform.system() == "Darwin"
    ext = ".py" if is_mac else ".pyw"
    dest = os.path.join(folder, f"Python Portal for Atlas{ext}")
    try:
        busted_url = f"{PORTAL_URL}?t={int(_time.time())}"
        req = urllib.request.Request(
            busted_url,
            headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}",
                     "Cache-Control": "no-cache"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
        with open(dest, "wb") as f:
            f.write(data)
    except Exception as e:
        messagebox.showerror("Download Failed",
            f"Could not download the portal:\n\n{e}\n\n"
            "Please check your internet connection and try again.")
        return

    # Step 4: Launch it automatically
    try:
        if is_mac:
            # On Mac, use python3 explicitly and avoid Windows-only flags
            subprocess.Popen(
                ["python3", dest, "--color=#a0c4ff"],
                start_new_session=True,
            )
        else:
            subprocess.Popen(
                [sys.executable, dest, "--color=#a0c4ff"],
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
    except Exception as e:
        messagebox.showerror(
            "Could Not Launch",
            f"The portal was saved to:\n\n{dest}\n\n"
            f"But it could not be launched automatically:\n{e}\n\n"
            f"Please open it manually.")
        return

    messagebox.showinfo(
        "Portal Opened",
        "The portal is now opening.\n\n"
        "Leave it running and let Atlas know it's open.")


def _post_bug_report_with_files(description, case_number, file_paths,
                                reporter_email="", category="Bug Fix"):
    if not BUG_REPORT_WEBHOOK_URL:
        return False, "No bug-report channel is configured."
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    content = (
        f"**Bug Report - {APP_NAME}**\n"
        f"**Case:** {case_number}\n"
        f"**Category:** {category}\n"
        f"**Version:** {APP_VERSION}\n"
        f"**From:** {user}@{host}\n"
        + (f"**Email:** {reporter_email}\n" if reporter_email else "")
        + f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"**Details:**\n{description}"
    )
    if not file_paths:
        return _post_to_discord(content)
    try:
        boundary = f"----WebKitFormBoundary{uuid.uuid4().hex[:16]}"
        payload_json = json.dumps({"content": content[:1900]})
        body = b""
        body += f"--{boundary}\r\n".encode()
        body += b'Content-Disposition: form-data; name="payload_json"\r\n'
        body += b"Content-Type: application/json\r\n\r\n"
        body += payload_json.encode() + b"\r\n"
        for i, fp in enumerate(file_paths):
            p = Path(fp)
            file_data = p.read_bytes()
            body += f"--{boundary}\r\n".encode()
            body += f'Content-Disposition: form-data; name="files[{i}]"; filename="{p.name}"\r\n'.encode()
            body += b"Content-Type: application/octet-stream\r\n\r\n"
            body += file_data + b"\r\n"
        body += f"--{boundary}--\r\n".encode()
        req = urllib.request.Request(
            BUG_REPORT_WEBHOOK_URL, data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}",
                     "User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status in (200, 204):
                return True, None
            return False, f"Server returned status {resp.status}"
    except Exception as e:
        return False, str(e)


def _post_update_applied(old_ver, new_ver):
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    host = platform.node() or "unknown"
    content = (
        f"**Update Applied - {APP_NAME}**\n"
        f"**Updated:** v{old_ver} -> v{new_ver}\n"
        f"**By:** {user}@{host}\n"
        f"**When:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    return _post_to_discord(content)


def _check_for_update():
    try:
        data = json.loads(_http_get(UPDATE_MANIFEST_URL))
        remote = data.get("version", "")
        if remote and _version_tuple(remote) > _version_tuple(APP_VERSION):
            return data
    except Exception:
        pass
    return None


# ------------------------------------------------------------------
# Self-update — plain overwrite (NO data splice), intentionally.
#
# Unlike the other consoles (Packages = CUSTOMS_DATA, XML =
# BUILTIN_TIN_NUMBERS + BUILTIN_CODES, Ocean Cargo = EMAIL_LOOKUP), this
# console has NO user-edited embedded data block: it processes selected
# PDF files and keeps all state in memory only, and never writes user
# data back into its own .pyw. There is therefore nothing to preserve,
# so we simply replace the whole file with the downloaded version.
#
# If a persisted/user-editable data block is ever added to this script,
# adopt the same splice-preservation pattern used by the other consoles
# (extract the local block, splice it into new_text before os.replace)
# so user entries aren't wiped on update.
# ------------------------------------------------------------------
def _download_and_apply_update(new_url):
    try:
        new_text = _http_get(new_url, timeout=30)
        tmp = SCRIPT_PATH.with_name(SCRIPT_PATH.name + ".new")
        tmp.write_text(new_text, encoding="utf-8")
        os.replace(str(tmp), str(SCRIPT_PATH))
        return True, None
    except Exception as e:
        return False, str(e)


# ==============================================================================
# Page rescaling — normalize all pages to A4 width, preserve height proportionally
# ==============================================================================
def _rescale_to_a4_width(src_doc, page_idx):
    """Scale a page to A4 width (595pt) while preserving its height
    proportionally.  Returns a new fitz.Document containing one page.
    If the page is already A4 width, it's copied as-is."""
    page = src_doc[page_idx]
    if abs(page.rect.width - A4_WIDTH) < 1:
        # Already A4 width — just copy it
        out = fitz.open()
        out.insert_pdf(src_doc, from_page=page_idx, to_page=page_idx)
        return out
    zoom = A4_WIDTH / page.rect.width
    new_width = A4_WIDTH
    new_height = page.rect.height * zoom
    scaled = fitz.open()
    new_page = scaled.new_page(width=new_width, height=new_height)
    new_page.show_pdf_page(new_page.rect, src_doc, page_idx)
    return scaled


# ==============================================================================
# Sort-only mode — reorders factura PDF pages by manifest package order
# ==============================================================================
def _sort_factura_in_place(pdf_path, package_list):
    """Sort PDF pages by package_list order, then unmatched pages at end.
    Saves to a temp file and atomically replaces the original (no duplicate)."""
    pkg_pattern = re.compile(r"\b(10\d{7,})\b")
    doc = fitz.open(str(pdf_path))
    pkg_to_page = defaultdict(list)

    for i, page in enumerate(doc):
        text = page.get_text()
        matches = pkg_pattern.findall(text)
        for m in set(matches):
            pkg_to_page[m].append(i)

    sorted_doc = fitz.open()
    used_pages = set()

    # Pages in manifest order
    for pkg in package_list:
        if pkg not in pkg_to_page:
            continue
        for idx in pkg_to_page[pkg]:
            if idx in used_pages:
                continue
            used_pages.add(idx)
            scaled = _rescale_to_a4_width(doc, idx)
            sorted_doc.insert_pdf(scaled)
            scaled.close()

    # Unmatched pages at end
    for i in range(len(doc)):
        if i in used_pages:
            continue
        scaled = _rescale_to_a4_width(doc, i)
        sorted_doc.insert_pdf(scaled)
        scaled.close()

    temp_output = str(pdf_path) + "._tmp_sorted"
    sorted_doc.save(temp_output, deflate=True, garbage=4, clean=True,
                    incremental=False)
    sorted_doc.close()
    doc.close()
    os.replace(temp_output, str(pdf_path))


# ==============================================================================
# Tooltip helpers
# ==============================================================================
_tooltip_win = None

def _show_tooltip(widget, text):
    global _tooltip_win
    _hide_tooltip()
    x = widget.winfo_rootx() + 20
    y = widget.winfo_rooty() + 28
    _tooltip_win = tk.Toplevel(widget)
    _tooltip_win.wm_overrideredirect(True)
    _tooltip_win.wm_geometry(f"+{x}+{y}")
    _tooltip_win.attributes("-topmost", True)
    lbl = tk.Label(_tooltip_win, text=text, justify="left",
                   bg="#1a1a2e", fg="#e8e8e8", relief="solid",
                   bd=1, padx=10, pady=8,
                   font=(MODERN_FONT, 10), wraplength=320)
    lbl.pack()

def _hide_tooltip():
    global _tooltip_win
    if _tooltip_win is not None:
        try:
            _tooltip_win.destroy()
        except Exception:
            pass
        _tooltip_win = None


# ==============================================================================
# Context menu helper
# ==============================================================================
def _attach_context_menu(widget):
    """Attach Cut/Copy/Paste/Select All context menu to a textbox or entry."""
    menu = tk.Menu(widget, tearoff=0)
    def _cut():
        try:
            widget.event_generate("<<Cut>>")
        except Exception:
            pass
    def _copy():
        try:
            widget.event_generate("<<Copy>>")
        except Exception:
            pass
    def _paste():
        try:
            widget.event_generate("<<Paste>>")
        except Exception:
            pass
    def _select_all():
        try:
            widget.event_generate("<<SelectAll>>")
        except Exception:
            pass
    menu.add_command(label="Cut", command=_cut)
    menu.add_command(label="Copy", command=_copy)
    menu.add_command(label="Paste", command=_paste)
    menu.add_separator()
    menu.add_command(label="Select All", command=_select_all)
    def _show_menu(e):
        try:
            menu.tk_popup(e.x_root, e.y_root)
        finally:
            menu.grab_release()
    widget.bind("<Button-3>", _show_menu)


REQUIRED_MANIFEST_HEADERS = {
    "boxnum (cby)",
    "package#",
    "invoice",
}


def _find_manifest_worksheet(workbook):
    for worksheet in workbook.worksheets:
        if worksheet.title.strip().casefold() == "manifest":
            return worksheet

    for worksheet in workbook.worksheets:
        first_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = {
            str(value).strip().casefold()
            for value in first_row
            if value is not None
        }
        if REQUIRED_MANIFEST_HEADERS.issubset(headers):
            return worksheet

    raise ValueError(
        "Unable to locate the Manifest worksheet. The workbook must contain "
        "the columns 'BoxNum (CBY)', 'Package#', and 'INVOICE'."
    )


# ==============================================================================
# MAIN APPLICATION
# ==============================================================================
class FacturaSplitApp:
    def __init__(self):
        ctk.set_appearance_mode("Light")
        ctk.set_default_color_theme("blue")
        self.root = ctk.CTk()
        self.root.title(f"{APP_NAME}  v{APP_VERSION}")
        self.root.configure(fg_color=BG)
        w, h = 720, 600
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.root.resizable(True, True)

        # State
        self._manifest_path = None
        self._rows = []  # list of (cby, pkg, invoice) tuples
        self._pending_update = None
        self._support_tooltip = "Report a Bug"
        self._processing = False

        self._build_ui()

        # Start background update check
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    # ---- UI ----------------------------------------------------------
    def _build_ui(self):
        # Header
        header = ctk.CTkFrame(self.root, fg_color="transparent")
        header.pack(fill="x", padx=16, pady=(16, 8))
        ctk.CTkLabel(header, text="Split Factura by CBY",
                     font=(MODERN_FONT, 20, "bold"),
                     text_color=LIGHT).pack(side="left")
        ctk.CTkLabel(header, text="(Mass Print)",
                     font=(MODERN_FONT, 14),
                     text_color=MUTED).pack(side="left", padx=(8, 0))

        # Info / help icon — hover shows "Instructions" tooltip
        self._info_btn = ctk.CTkButton(
            header, text=" i ", width=30, height=30,
            fg_color="#2a4a6a", hover_color="#3a5a7a",
            corner_radius=15, text_color=LIGHT,
            font=(MODERN_FONT, 16, "bold"),
            command=self._show_help_dialog)
        self._info_btn.pack(side="left", padx=(10, 0))
        self._info_btn.bind("<Enter>",
            lambda e: _show_tooltip(self._info_btn, "Instructions"))
        self._info_btn.bind("<Leave>",
            lambda e: _hide_tooltip())

        # Top buttons
        top_btns = ctk.CTkFrame(self.root, fg_color="transparent")
        top_btns.pack(fill="x", padx=16, pady=(0, 8))
        self._choose_btn = ctk.CTkButton(
            top_btns, text="Choose Manifest", command=self._choose_manifest,
            fg_color=ACCENT, hover_color=ACCENT_H, width=160, height=34,
            corner_radius=6, font=(MODERN_FONT, 13, "bold"))
        self._choose_btn.pack(side="left")

        self._print_btn = ctk.CTkButton(
            top_btns, text="Print Split", command=self._print_split,
            fg_color=ORANGE, hover_color=ORANGE_H, width=140, height=34,
            corner_radius=6, font=(MODERN_FONT, 13, "bold"),
            state="disabled")
        self._print_btn.pack(side="left", padx=(10, 0))

        self._status_label = ctk.CTkLabel(
            top_btns, text="No manifest loaded",
            font=(MODERN_FONT, 11), text_color=MUTED)
        self._status_label.pack(side="left", padx=(14, 0))

        # Treeview
        tree_frame = ctk.CTkFrame(self.root, fg_color=PANEL, corner_radius=8)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=INPUT, foreground=DARK,
                        rowheight=32, fieldbackground=INPUT,
                        bordercolor=BORDER, borderwidth=1,
                        font=(MODERN_FONT, 13))
        style.configure("Treeview.Heading",
                        background=BG, foreground=LIGHT,
                        font=(MODERN_FONT, 13, "bold"), relief="flat")
        style.map("Treeview.Heading",
                  background=[("active", ACCENT_H)])
        style.map("Treeview",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", LIGHT)])

        cols = ("cby", "pkg", "invoice")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                 selectmode="browse")
        self.tree.heading("cby", text="CBY")
        self.tree.heading("pkg", text="Package #")
        self.tree.heading("invoice", text="Invoice")
        self.tree.column("cby", width=120, anchor="center")
        self.tree.column("pkg", width=200, anchor="center")
        self.tree.column("invoice", width=140, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=4, pady=4)

        # Alternating row tags
        self.tree.tag_configure("even", background=INPUT)
        self.tree.tag_configure("odd", background=ROW_ALT)

        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical",
                                  command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Status bar with bug icon
        status_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        status_frame.pack(fill="x", padx=16, pady=(0, 12))

        # Sort-only button (hidden until manifest + factura detected)
        self._sort_btn = ctk.CTkButton(
            status_frame,
            text="Not Ready To Split? Sort Factura Instead",
            command=self._sort_factura_only,
            fg_color="#9b59b6", hover_color="#8e44ad",
            width=280, height=28, corner_radius=6,
            font=(MODERN_FONT, 11, "bold"))
        # Not packed yet — shown dynamically after manifest load

        self._support_btn = ctk.CTkButton(
            status_frame, text="\U0001f41e", width=34, height=28,
            fg_color=BG, hover_color="#24507a", corner_radius=6,
            font=("Segoe UI Emoji", 15),
            command=self._on_support_click)
        self._support_btn.pack(side="right")
        self._support_btn.bind("<Enter>",
            lambda e: _show_tooltip(self._support_btn, self._support_tooltip))
        self._support_btn.bind("<Leave>",
            lambda e: _hide_tooltip())

    # ---- Help dialog -------------------------------------------------
    def _show_help_dialog(self):
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("How It Works")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 520, 660
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        ctk.CTkLabel(dlg, text="How It Works",
                     font=(MODERN_FONT, 18, "bold"),
                     text_color=LIGHT).pack(anchor="w", padx=20, pady=(16, 8))

        help_text = (
            "This tool splits factura PDFs into separate files\n"
            "organized by CBY (box number), so each CBY gets\n"
            "its own PDF with all relevant invoice pages.\n\n"
            "Steps:\n\n"
            "1. Click \"Choose Manifest\" and select the manifest\n"
            "   Excel file for the shipment.\n\n"
            "2. The window populates with three columns:\n"
            "   - CBY (box number, from Column C)\n"
            "   - Package # (from Column T)\n"
            "   - Invoice (notes from Column U: Y, dupe, corrupt, etc.)\n\n"
            "   Document rows (CBY 1000) are automatically excluded.\n\n"
            "3. Click \"Print Split\" to choose what to print.\n\n"
            "4. In the dialog:\n"
            "   - \"All\" includes everything (with duplicate invoices).\n"
            "     This is not ideal for Customs clearance.\n"
            "   - Individual categories let you pick only the pages\n"
            "     you need (e.g. just Y, just blanks, etc.).\n"
            "   - Select which factura PDF(s) to parse. This lets you\n"
            "     process in batches without overwriting previous output.\n\n"
            "5. The tool scans each selected factura PDF, finds pages\n"
            "   matching package numbers, and groups them by CBY.\n\n"
            "6. Output PDFs are saved in the same folder as the\n"
            "   factura. Each file is named after the CBY number\n"
            "   (e.g. 47.pdf, 59.pdf). Every page is labeled\n"
            "   \"CBY {number}\" in red."
        )
        ctk.CTkLabel(dlg, text=help_text,
                     font=(MODERN_FONT, 12), text_color=TEXT,
                     anchor="w", justify="left").pack(
                         anchor="w", padx=20, pady=(0, 16))

        ctk.CTkButton(dlg, text="Got it", command=dlg.destroy,
                      fg_color=ACCENT, hover_color=ACCENT_H, width=100,
                      height=32, corner_radius=6,
                      font=(MODERN_FONT, 12, "bold")).pack(
                          side="bottom", pady=(0, 16))

    # ---- Manifest loading --------------------------------------------
    def _choose_manifest(self):
        path = filedialog.askopenfilename(
            title="Select Manifest Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not path:
            return
        self._manifest_path = Path(path)
        try:
            self._load_manifest()
        except Exception as e:
            messagebox.showerror("Error", f"Could not read manifest:\n{e}")
            return

    def _load_manifest(self):
        wb = openpyxl.load_workbook(self._manifest_path, read_only=True,
                                    data_only=True)
        try:
            ws = _find_manifest_worksheet(wb)
            # Column C (3) = CBY, Column T (20) = Package#, Column U (21) = Invoice
            # Row 1 = headers, data starts at row 2
            self._rows = []
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=21,
                                    values_only=False):
                cby_val = row[0].value  # column C (index 0 in our slice)
                pkg_val = row[17].value  # column T (index 17 in our slice)
                inv_val = row[18].value  # column U (index 18 in our slice)
                if cby_val is None and pkg_val is None:
                    continue
                cby = str(cby_val).strip() if cby_val is not None else ""
                pkg = str(pkg_val).strip() if pkg_val is not None else ""
                if pkg.endswith(".0"):
                    pkg = pkg[:-2]
                if cby.endswith(".0"):
                    cby = cby[:-2]
                inv = str(inv_val).strip() if inv_val is not None else ""
                # Skip document rows (CBY 1000)
                if cby == "1000":
                    continue
                if pkg:
                    self._rows.append((cby, pkg, inv))
        finally:
            wb.close()

        # Populate treeview
        self.tree.delete(*self.tree.get_children())
        for i, (cby, pkg, inv) in enumerate(self._rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", values=(cby, pkg, inv), tags=(tag,))

        # Update status — also detect factura PDFs
        manifest_dir = self._manifest_path.parent
        self._pdf_files = []
        for p in manifest_dir.rglob("*.pdf"):
            if p.name.lower().startswith("factura"):
                self._pdf_files.append(p)
        pdf_count = len(self._pdf_files)
        self._status_label.configure(
            text=f"{len(self._rows)} entries  |  {pdf_count} factura PDF(s) detected",
            text_color=LIGHT)
        self._print_btn.configure(state="normal")

        # Show sort button if factura PDFs were detected
        if pdf_count > 0:
            self._sort_btn.pack(side="left")
        else:
            self._sort_btn.pack_forget()

    # ---- Sort factura only (no split) --------------------------------
    def _sort_factura_only(self):
        if not self._rows or not self._pdf_files:
            return

        # Build package list from manifest (same order as displayed)
        package_list = [pkg for _, pkg, _ in self._rows if pkg]
        if not package_list:
            messagebox.showwarning("No Packages",
                "No package numbers found in the manifest.")
            return

        pdf_files = self._pdf_files

        # Explanation dialog
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Sort Factura Instead")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 540, 560
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        ctk.CTkLabel(dlg, text="Sort Factura Instead",
                     font=(MODERN_FONT, 18, "bold"),
                     text_color=LIGHT).pack(anchor="w", padx=24, pady=(16, 6))

        explanation = (
            "Referencing a single Factura file while working on the\n"
            "manifest is a lot easier than opening separate PDF files,\n"
            "so this would be the encouraged first step if you are just\n"
            "starting the manifest.\n\n"
            "Sorting just makes it flow in order so working on the\n"
            "manifest is easier. This will sort the selected factura\n"
            "PDFs simultaneously, reordering their pages to match the\n"
            "package order from the manifest.\n\n"
            "Alternatively, if you're already finished adding your notes\n"
            "to the factura and finished with the manifest, then you\n"
            "would skip this and use the orange Print Split button\n"
            "located in the main window.")
        ctk.CTkLabel(dlg, text=explanation,
                     font=(MODERN_FONT, 12),
                     text_color=TEXT, justify="left",
                     anchor="w").pack(fill="x", padx=24, pady=(0, 12))

        # ---- Factura checkboxes ----
        ctk.CTkLabel(dlg, text="Factura PDF(s) to be sorted:",
                     font=(MODERN_FONT, 12, "bold"),
                     text_color=LIGHT).pack(anchor="w", padx=24, pady=(0, 4))

        pdf_checks = {}
        for pdf_path in pdf_files:
            var = ctk.BooleanVar(value=True)
            try:
                rel = pdf_path.relative_to(self._manifest_path.parent)
                display_name = str(rel)
            except Exception:
                display_name = pdf_path.name
            size_mb = pdf_path.stat().st_size / (1024 * 1024)
            cb = ctk.CTkCheckBox(
                dlg,
                text=f"{display_name}  ({size_mb:.1f} MB)",
                variable=var,
                font=(MODERN_FONT, 11), text_color=LIGHT,
                fg_color=ACCENT, hover_color=ACCENT_H,
                border_color=LIGHT)
            cb.pack(anchor="w", padx=32, pady=2)
            pdf_checks[str(pdf_path)] = var

        # ---- Buttons ----
        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(side="bottom", fill="x", padx=24, pady=(0, 16))

        proceed = {"yes": False, "pdfs": []}

        def _do_sort():
            selected = [Path(fp) for fp, var in pdf_checks.items()
                        if var.get()]
            if not selected:
                messagebox.showwarning("No Facturas Selected",
                    "Please choose which factura(s) you wish to sort.")
                return
            proceed["yes"] = True
            proceed["pdfs"] = selected
            dlg.destroy()

        def _close():
            proceed["yes"] = False
            dlg.destroy()

        ctk.CTkButton(btn_frame, text="Close - I'm Ready to Split Factura",
                      command=_close,
                      fg_color="#555b5e", hover_color="#444a4d",
                      width=240, height=34, corner_radius=6,
                      font=(MODERN_FONT, 12, "bold")).pack(side="right", padx=(8, 0))

        ctk.CTkButton(btn_frame, text="Yes, Sort Factura",
                      command=_do_sort,
                      fg_color="#9b59b6", hover_color="#8e44ad",
                      width=180, height=34, corner_radius=6,
                      font=(MODERN_FONT, 12, "bold")).pack(side="right")

        self.root.wait_window(dlg)
        if not proceed["yes"]:
            return

        # Sort selected factura PDFs simultaneously
        pdfs_to_sort = proceed["pdfs"]

        # Progress window
        prog = ctk.CTkToplevel(self.root)
        prog.title("Sorting...")
        prog.configure(fg_color=BG)
        prog.transient(self.root)
        prog.grab_set()
        prog.resizable(False, False)
        prog.geometry("420x140")
        ctk.CTkLabel(prog, text="Sorting factura(s) by manifest order...",
                     font=(MODERN_FONT, 14, "bold"),
                     text_color=LIGHT).pack(pady=(20, 8))
        prog_label = ctk.CTkLabel(prog, text="Starting...",
                                  font=(MODERN_FONT, 11), text_color=MUTED)
        prog_label.pack(pady=4)
        prog_bar = ctk.CTkProgressBar(prog, width=340, fg_color=ACCENT)
        prog_bar.pack(pady=8)
        prog_bar.set(0)

        total = len(pdfs_to_sort)

        def _worker():
            errors = []
            for i, pdf_path in enumerate(pdfs_to_sort):
                try:
                    prog_label.after(0, lambda n=pdf_path.name, idx=i:
                        prog_label.configure(text=f"Sorting {n}... ({idx+1}/{total})"))
                    prog_bar.after(0, lambda v=i/total: prog_bar.set(v))
                    _sort_factura_in_place(pdf_path, package_list)
                except Exception as e:
                    errors.append(f"{pdf_path.name}: {e}")

            prog_bar.after(0, lambda: prog_bar.set(1.0))
            prog.after(800, lambda: prog.destroy())

            if errors:
                err_msg = "Some PDFs could not be sorted:\n\n" + "\n".join(errors)
                self.root.after(900, lambda: messagebox.showerror(
                    "Sort Partially Failed", err_msg))
            else:
                output_dir = pdfs_to_sort[0].parent
                pdf_word = "PDFs" if total > 1 else "PDF"
                has_have = "have" if total > 1 else "has"
                self.root.after(900, lambda: messagebox.showinfo(
                    "PDF Success",
                    f"{total} {pdf_word} {has_have} been sorted by CBY "
                    f"based on the layout of the manifest.\n\n"
                    f"The original unsorted {pdf_word} {'were' if total > 1 else 'was'} "
                    f"replaced in-place — no duplicate files {'were' if total > 1 else 'was'} "
                    f"created.\n\n"
                    f"Location:\n{output_dir}"))

        threading.Thread(target=_worker, daemon=True).start()

    # ---- Print split dialog ------------------------------------------
    def _print_split(self):
        if not self._rows:
            return

        # Collect unique invoice values
        inv_values = set()
        has_blank = False
        has_non_blank = False
        for _, _, inv in self._rows:
            inv_lower = inv.lower().strip()
            if inv_lower == "":
                has_blank = True
            else:
                has_non_blank = True
                inv_values.add(inv_lower)

        all_blank = not has_non_blank
        pdf_files = getattr(self, "_pdf_files", [])

        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Choose What to Print")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 520, 640
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        # Scrollable content
        content = ctk.CTkScrollableFrame(dlg, fg_color="transparent",
                                         label_text="")
        content.pack(fill="both", expand=True, padx=20, pady=(12, 0))

        ctk.CTkLabel(content, text="Choose What to Print",
                     font=(MODERN_FONT, 16, "bold"),
                     text_color=LIGHT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(content,
                     text="Select which invoice categories to include.\n"
                          "Anything not selected will be ignored.",
                     font=(MODERN_FONT, 11), text_color=MUTED,
                     justify="left").pack(anchor="w", pady=(0, 10))

        # ---- "All" option (always present, partitioned) ----
        all_frame = ctk.CTkFrame(content, fg_color="#1a2a4a", corner_radius=6)
        all_frame.pack(fill="x", pady=(0, 8))
        all_var = ctk.BooleanVar(value=all_blank)
        ctk.CTkCheckBox(all_frame, text="All (include everything)",
                        variable=all_var,
                        font=(MODERN_FONT, 13, "bold"), text_color=LIGHT,
                        fg_color=ACCENT, hover_color=ACCENT_H,
                        border_color=LIGHT,
                        checkbox_width=22, checkbox_height=22).pack(
                            anchor="w", padx=12, pady=(10, 4))
        ctk.CTkLabel(all_frame,
                     text="WARNING: This will capture all pages including\n"
                          "duplicate invoices. This is not ideal for\n"
                          "Customs clearance.",
                     font=(MODERN_FONT, 10), text_color="#ffcccc",
                     justify="left").pack(anchor="w", padx=28, pady=(0, 8))

        # ---- Partition ----
        ctk.CTkLabel(content, text="\u2500" * 50,
                     font=(MODERN_FONT, 10), text_color=MUTED).pack(
                         fill="x", pady=(0, 6))

        # ---- Individual category checkboxes ----
        cat_label = ctk.CTkLabel(content, text="Individual Categories:",
                                 font=(MODERN_FONT, 12, "bold"),
                                 text_color=LIGHT)
        cat_label.pack(anchor="w", pady=(0, 4))

        checks = {}
        if has_blank:
            var = ctk.BooleanVar(value=False)
            cb = ctk.CTkCheckBox(content, text="Blanks (no note)",
                                 variable=var,
                                 font=(MODERN_FONT, 12), text_color=LIGHT,
                                 fg_color=ACCENT, hover_color=ACCENT_H,
                                 border_color=LIGHT)
            cb.pack(anchor="w", padx=20, pady=3)
            checks["(blank)"] = var
        for val in sorted(inv_values):
            var = ctk.BooleanVar(value=False)
            display = val[0].upper() + val[1:] if val else val
            count = sum(1 for _, _, inv in self._rows
                        if inv.lower().strip() == val)
            cb = ctk.CTkCheckBox(content, text=f"{display} ({count})",
                                 variable=var,
                                 font=(MODERN_FONT, 12), text_color=LIGHT,
                                 fg_color=ACCENT, hover_color=ACCENT_H,
                                 border_color=LIGHT)
            cb.pack(anchor="w", padx=20, pady=3)
            checks[val] = var

        # ---- Partition ----
        ctk.CTkLabel(content, text="\u2500" * 50,
                     font=(MODERN_FONT, 10), text_color=MUTED).pack(
                         fill="x", pady=(8, 6))

        # ---- Factura selection ----
        ctk.CTkLabel(content, text="Select Factura PDF(s) to Parse:",
                     font=(MODERN_FONT, 12, "bold"),
                     text_color=LIGHT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(content,
                     text=f"{len(pdf_files)} factura PDF(s) detected.\n"
                          "Choose which ones to include. This lets you\n"
                          "process in batches without overwriting files.",
                     font=(MODERN_FONT, 10), text_color=MUTED,
                     justify="left").pack(anchor="w", pady=(0, 6))

        pdf_checks = {}
        if not pdf_files:
            ctk.CTkLabel(content,
                         text="No factura PDFs found in manifest folder!",
                         font=(MODERN_FONT, 11, "bold"),
                         text_color="#ff8888").pack(anchor="w", padx=20, pady=4)
        else:
            for pdf_path in pdf_files:
                var = ctk.BooleanVar(value=True)
                # Show relative path from manifest dir for clarity
                try:
                    rel = pdf_path.relative_to(self._manifest_path.parent)
                    display_name = str(rel)
                except Exception:
                    display_name = pdf_path.name
                size_mb = pdf_path.stat().st_size / (1024 * 1024)
                cb = ctk.CTkCheckBox(
                    content,
                    text=f"{display_name}  ({size_mb:.1f} MB)",
                    variable=var,
                    font=(MODERN_FONT, 11), text_color=LIGHT,
                    fg_color=ACCENT, hover_color=ACCENT_H,
                    border_color=LIGHT)
                cb.pack(anchor="w", padx=20, pady=2)
                pdf_checks[str(pdf_path)] = var

        # ---- Buttons ----
        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=20, pady=(0, 16))

        def _do_print():
            # Check invoice categories
            use_all = all_var.get()
            selected = set()
            if not use_all:
                for key, var in checks.items():
                    if var.get():
                        selected.add(key)
                if not selected:
                    messagebox.showwarning("Nothing Selected",
                        "Please select at least one invoice category to print,\n"
                        "or check 'All' to include everything.")
                    return

            # Check factura selection
            selected_pdfs = [Path(fp) for fp, var in pdf_checks.items()
                             if var.get()]
            if not selected_pdfs:
                messagebox.showwarning("No Facturas Selected",
                    "Please choose which factura(s) you wish to split.")
                return

            dlg.destroy()
            self._do_print_split(selected, use_all, selected_pdfs)

        ctk.CTkButton(btns, text="Print Split", command=_do_print,
                      fg_color=ORANGE, hover_color=ORANGE_H, width=120,
                      height=32, corner_radius=6,
                      font=(MODERN_FONT, 12, "bold")).pack(side="left")
        ctk.CTkButton(btns, text="Cancel", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=32, corner_radius=6,
                      font=(MODERN_FONT, 12)).pack(side="left", padx=(8, 0))

    # ---- Actual printing ---------------------------------------------
    def _do_print_split(self, selected_categories, use_all, pdf_files):
        # Build package -> (cby, invoice) lookup
        pkg_lookup = {}
        for cby, pkg, inv in self._rows:
            inv_lower = inv.lower().strip()
            if use_all or inv_lower in selected_categories:
                pkg_lookup[pkg] = (cby, inv_lower)

        if not pkg_lookup:
            messagebox.showinfo("Nothing to Print",
                "No packages match the selected categories.")
            return

        if not pdf_files:
            messagebox.showwarning("No PDFs",
                "No factura PDFs were selected.")
            return

        manifest_dir = self._manifest_path.parent

        # Progress window
        prog = ctk.CTkToplevel(self.root)
        prog.title("Processing...")
        prog.configure(fg_color=BG)
        prog.transient(self.root)
        prog.grab_set()
        prog.resizable(False, False)
        prog.geometry("400x140")
        ctk.CTkLabel(prog, text="Splitting facturas by CBY...",
                     font=(MODERN_FONT, 14, "bold"),
                     text_color=LIGHT).pack(pady=(20, 8))
        prog_label = ctk.CTkLabel(prog, text="Starting...",
                                  font=(MODERN_FONT, 11), text_color=MUTED)
        prog_label.pack(pady=4)
        prog_bar = ctk.CTkProgressBar(prog, width=320, fg_color=ACCENT)
        prog_bar.pack(pady=8)
        prog_bar.set(0)

        def _worker():
            cby_pages = defaultdict(list)
            found_pkgs = set()
            total_pdfs = len(pdf_files)
            for pdf_idx, pdf_path in enumerate(pdf_files):
                prog_label.after(0, lambda i=pdf_idx, n=pdf_path.name:
                    prog_label.configure(text=f"Scanning {n}..."))
                prog_bar.after(0, lambda v=pdf_idx/total_pdfs:
                    prog_bar.set(v))
                try:
                    doc = fitz.open(str(pdf_path))
                    for page_idx in range(len(doc)):
                        text = doc[page_idx].get_text()
                        for pkg, (cby, inv) in pkg_lookup.items():
                            if pkg in text:
                                cby_pages[cby].append(
                                    (str(pdf_path), page_idx))
                                found_pkgs.add(pkg)
                                break
                    doc.close()
                except Exception as e:
                    print(f"Error reading {pdf_path}: {e}")

            # Generate output PDFs — save in the same folder as the factura
            output_dir = pdf_files[0].parent
            output_dir.mkdir(exist_ok=True)

            for cby, pages in cby_pages.items():
                out_pdf = fitz.open()
                for pdf_name, page_idx in pages:
                    with fitz.open(pdf_name) as src:
                        scaled = _rescale_to_a4_width(src, page_idx)
                        out_pdf.insert_pdf(scaled)
                        scaled.close()
                # Label pages
                label = f"CBY {cby}"
                for page in out_pdf:
                    page.insert_text((32, 500), label, fontsize=15,
                                     color=(1, 0, 0), rotate=90)
                out_path = output_dir / f"{cby}.pdf"
                out_pdf.save(str(out_path))
                out_pdf.close()

            prog_bar.after(0, lambda: prog_bar.set(1.0))
            prog_label.after(0, lambda: prog_label.configure(
                text=f"Done! {len(cby_pages)} CBY files created."))
            prog.after(1500, lambda: prog.destroy())

            # Show results
            if not found_pkgs:
                # Nothing was found — likely the factura was printed
                # (package numbers embedded as flat images, not text)
                self.root.after(100, lambda: messagebox.showwarning(
                    "Package Numbers Not Detectable",
                    "No package numbers were found in the factura PDF(s).\n\n"
                    "This usually happens when the factura was printed\n"
                    "instead of saved. Printing can embed the package\n"
                    "number as a flat image, making it undetectable by\n"
                    "text scanning.\n\n"
                    "Please ensure you SAVE the factura when downloading\n"
                    "from E-Box, rather than printing it to PDF."))
                return

            missing = sorted(set(pkg_lookup.keys()) - found_pkgs)
            result_msg = (
                f"Created {len(cby_pages)} CBY PDF files.\n\n"
                f"Saved to:\n{output_dir}\n\n")
            if missing:
                result_msg += (
                    f"{len(missing)} packages were not found in any PDF:\n"
                    + "\n".join(missing[:10]))
                if len(missing) > 10:
                    result_msg += f"\n...and {len(missing)-10} more"
            else:
                result_msg += "All packages found and processed."
            self.root.after(100, lambda: messagebox.showinfo(
                "Complete", result_msg))

        threading.Thread(target=_worker, daemon=True).start()

    # ---- Remote support ----------------------------------------------
    def _check_update_bg(self):
        result = _check_for_update()
        if result:
            self._pending_update = result
            try:
                self.root.after(0, self._refresh_support_icon)
            except Exception:
                pass

    def _refresh_support_icon(self):
        try:
            if self._pending_update:
                ver = self._pending_update.get("version", "?")
                self._support_btn.configure(
                    text="Apply Fixes", fg_color="#b8860b",
                    hover_color="#daa520",
                    font=(MODERN_FONT, 11, "bold"), width=90)
                self._support_tooltip = f"Apply Fixes \u2014 v{ver} available"
            else:
                self._support_btn.configure(
                    text="\U0001f41e", fg_color=BG,
                    hover_color="#24507a",
                    font=("Segoe UI Emoji", 15), width=34)
                self._support_tooltip = "Report a Bug"
        except Exception:
            pass

    def _on_support_click(self):
        _hide_tooltip()
        if self._pending_update:
            self._apply_fixes_dialog()
        else:
            self._report_bug_dialog()

    def _report_bug_dialog(self):
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Report a Bug")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 640
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(dlg, text="Report a Bug",
                     font=(MODERN_FONT, 15, "bold"),
                     text_color=LIGHT).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Describe the problem in as much detail as you can \u2014 "
                 f"what you\ndid, what happened, and what you expected. This "
                 f"goes directly\nto the developer ({DEVELOPER_NAME}).",
            font=(MODERN_FONT, 11), text_color=MUTED,
            anchor="w", justify="left").pack(anchor="w", padx=16, pady=(0, 6))

        # Category selector
        cat_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        cat_frame.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(cat_frame, text="Category:",
                     font=(MODERN_FONT, 11, "bold"), text_color=LIGHT,
                     anchor="w").pack(anchor="w", pady=(0, 2))
        cat_var = ctk.StringVar(value="Bug Fix")
        cat_row = ctk.CTkFrame(cat_frame, fg_color="transparent")
        cat_row.pack(fill="x")
        for cat in ("Bug Fix", "Feature Request",
                     "Environmental Change", "Other"):
            ctk.CTkRadioButton(
                cat_row, text=cat, variable=cat_var, value=cat,
                font=(MODERN_FONT, 10), text_color=LIGHT,
                fg_color=ACCENT, hover_color=ACCENT_H,
                border_color=BORDER).pack(side="left", padx=(0, 4))

        box = ctk.CTkTextbox(dlg, height=140, fg_color=INPUT,
                             border_color=BORDER, border_width=1,
                             corner_radius=4, text_color=DARK,
                             font=(MODERN_FONT, 11))
        box.pack(fill="both", expand=True, padx=16)
        _attach_context_menu(box)
        box.focus_set()

        # Email field
        email_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        email_frame.pack(fill="x", padx=16, pady=(4, 0))
        ctk.CTkLabel(email_frame,
            text="Your email (optional \u2014 for updates on this report)",
            font=(MODERN_FONT, 10), text_color=MUTED,
            anchor="w").pack(anchor="w")
        email_var = ctk.StringVar(value="")
        email_entry = ctk.CTkEntry(email_frame, textvariable=email_var,
                                   height=28, fg_color=INPUT, border_color=BORDER,
                                   border_width=1, corner_radius=4,
                                   text_color=DARK, font=(MODERN_FONT, 11))
        email_entry.pack(fill="x", pady=(1, 0))
        _attach_context_menu(email_entry)

        # "What to expect" info box
        info = ctk.CTkFrame(dlg, fg_color="#1a2a3a", corner_radius=6)
        info.pack(fill="x", padx=16, pady=(6, 8))
        ctk.CTkLabel(info, text="What to expect",
                     font=(MODERN_FONT, 11, "bold"), text_color="#88ccff",
                     anchor="w").pack(anchor="w", padx=10, pady=(8, 2))
        ctk.CTkLabel(info,
            text=f"\u2022 {DEVELOPER_NAME} will review your report within "
                 f"24-48 hours\n"
                 f"\u2022 When a fix is ready, the bug icon will change to "
                 f"'Apply Fixes'\n"
                 f"\u2022 {DEVELOPER_NAME} will coordinate with management if "
                 f"the fix\n"
                 f"   requires substantial work\n"
                 f"\u2022 Write down your case number for follow-up: contact "
                 f"{DEVELOPER_EMAIL}",
            font=(MODERN_FONT, 10), text_color=MUTED,
            anchor="w", justify="left").pack(anchor="w", padx=10, pady=(0, 8))

        # Mysterious portal icon (bottom-right corner, in button row)
        _portal_canvas = tk.Canvas(btns, width=24, height=24,
                                    bg=BG, highlightthickness=0)
        _portal_canvas.pack(side="right")
        _portal_canvas.create_oval(2, 2, 22, 22, outline="#a0c4ff", width=2)
        _portal_canvas.create_oval(7, 7, 17, 17, fill="#a0c4ff", outline="")
        _portal_canvas.configure(cursor="hand2")
        _portal_canvas.bind("<Button-1>", lambda e: _summon_portal(self.root))

        def _next():
            desc = box.get("0.0", "end").strip()
            if not desc:
                messagebox.showwarning("Empty",
                                       "Please describe the bug first.")
                return
            email = email_var.get().strip()
            category = cat_var.get()
            dlg.destroy()
            self._show_attach_files_dialog(desc, email, category)

        ctk.CTkButton(btns, text="Next", command=_next,
                      fg_color=GREEN, hover_color=GREEN_H, width=100,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        ctk.CTkButton(btns, text="Cancel", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    def _show_attach_files_dialog(self, description, reporter_email="",
                                  category="Bug Fix"):
        case_num = _generate_case_number()
        dlg = ctk.CTkToplevel(self.root)
        dlg.title(f"Attach Files? (Case {case_num})")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 420
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(dlg, text="Attach Sample Files?",
                     font=(MODERN_FONT, 15, "bold"),
                     text_color=LIGHT).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg,
            text=f"Based on your description, {DEVELOPER_NAME} may need "
                 f"sample\nfiles to reproduce the issue in a test "
                 f"environment.\n\n"
                 f"The console script itself is always attached "
                 f"automatically\nso {DEVELOPER_NAME} gets your exact version "
                 f"for testing.\n\n"
                 f"If you'd like, attach additional Excel or CSV files "
                 f"below.\nThis is completely optional and files are sent "
                 f"privately\nalongside your bug report.",
            font=(MODERN_FONT, 11), text_color=MUTED,
            anchor="w", justify="left").pack(anchor="w", padx=16, pady=(0, 8))

        file_list_frame = ctk.CTkFrame(dlg, fg_color=INPUT, corner_radius=4,
                                       border_width=1, border_color=BORDER)
        file_list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))
        file_list_label = ctk.CTkLabel(file_list_frame, text="No files selected",
                                       font=(MODERN_FONT, 11), text_color=MUTED,
                                       anchor="w", justify="left")
        file_list_label.pack(anchor="w", padx=10, pady=10)

        selected_files = []

        def _pick_files():
            paths = filedialog.askopenfilenames(
                title="Select files to attach",
                filetypes=[("Excel/CSV/PDF files", "*.xlsx *.xls *.csv *.pdf"),
                           ("All files", "*.*")])
            if paths:
                selected_files.clear()
                selected_files.extend(paths)
                names = [Path(p).name for p in paths]
                display = "\n".join(names[:6])
                if len(names) > 6:
                    display += f"\n...and {len(names)-6} more"
                file_list_label.configure(text=display, text_color=DARK)

        def _submit():
            submit_btn.configure(state="disabled", text="Sending...")
            skip_btn.configure(state="disabled")
            def _worker():
                all_files = [str(SCRIPT_PATH)] + list(selected_files)
                ok, err = _post_bug_report_with_files(
                    description, case_num, all_files,
                    reporter_email, category)
                self.root.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                dlg.destroy()
                n_files = len(selected_files)
                if n_files:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted with {n_files} file(s).\n\n"
                        f"Your report and files have been sent to "
                        f"{DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' here.")
                else:
                    messagebox.showinfo("Bug Reported",
                        f"Case {case_num} submitted.\n\n"
                        f"Your report has been sent to {DEVELOPER_NAME}.\n"
                        "When a fix is ready, you'll see 'Apply Fixes' here.")
            else:
                submit_btn.configure(state="normal", text="Submit Report")
                skip_btn.configure(state="normal")
                messagebox.showerror("Could Not Send",
                    f"The report could not be sent:\n{err}\n\n"
                    "Check your internet connection and try again.")

        ctk.CTkButton(btns, text="Browse Files", command=_pick_files,
                      fg_color=ACCENT, hover_color=ACCENT_H, width=120,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold")).pack(side="left")
        submit_btn = ctk.CTkButton(btns, text="Submit Report", command=_submit,
                      fg_color=GREEN, hover_color=GREEN_H, width=130,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold"))
        submit_btn.pack(side="left", padx=(8, 0))
        skip_btn = ctk.CTkButton(btns, text="Skip & Send", command=_submit,
                      fg_color="#667788", hover_color="#556677", width=110,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11))
        skip_btn.pack(side="left", padx=(8, 0))

    def _apply_fixes_dialog(self):
        upd = self._pending_update
        if not upd:
            return
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Apply Fixes")
        dlg.configure(fg_color=BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        w, h = 460, 380
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(side="bottom", fill="x", padx=16, pady=(0, 14))

        ver = upd.get("version", "?")
        ctk.CTkLabel(dlg, text=f"Fixes Available  (v{ver})",
                     font=(MODERN_FONT, 15, "bold"),
                     text_color=LIGHT).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(dlg, text=f"From {DEVELOPER_NAME}.",
                     font=(MODERN_FONT, 11), text_color=MUTED,
                     anchor="w").pack(anchor="w", padx=16, pady=(0, 8))

        box = ctk.CTkTextbox(dlg, height=170, fg_color=INPUT,
                             border_color=BORDER, border_width=1,
                             corner_radius=4, text_color=DARK,
                             font=(MODERN_FONT, 11))
        box.pack(fill="both", expand=True, padx=16)
        box.insert("0.0", upd.get("changelog", "(no description provided)"))
        box.configure(state="disabled")

        def _apply():
            apply_btn.configure(state="disabled", text="Applying...")
            def _worker():
                ok, err = _download_and_apply_update(upd.get("url", ""))
                self.root.after(0, lambda: _done(ok, err))
            threading.Thread(target=_worker, daemon=True).start()

        def _done(ok, err):
            if ok:
                old_ver = APP_VERSION
                new_ver = upd.get("version", "?")
                threading.Thread(
                    target=lambda: _post_update_applied(old_ver, new_ver),
                    daemon=True).start()
                dlg.destroy()
                if messagebox.askyesno("Update Applied",
                    f"Updated to version {new_ver}.\n\n"
                    "The console needs to restart to apply the changes.\n"
                    "Restart now?"):
                    self._restart_app()
            else:
                apply_btn.configure(state="normal", text="Apply Fixes")
                messagebox.showerror("Update Failed",
                    f"Could not apply the update:\n{err}\n\n"
                    "Check your internet connection and try again.")

        apply_btn = ctk.CTkButton(btns, text="Apply Fixes", command=_apply,
                      fg_color="#b8860b", hover_color="#daa520", width=130,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11, "bold"))
        apply_btn.pack(side="left")
        ctk.CTkButton(btns, text="Later", command=dlg.destroy,
                      fg_color="#667788", hover_color="#556677", width=90,
                      height=30, corner_radius=5,
                      font=(MODERN_FONT, 11)).pack(side="left", padx=(8, 0))

    def _restart_app(self):
        try:
            subprocess.Popen([sys.executable, str(SCRIPT_PATH)])
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass
        sys.exit(0)

    # ---- Run ---------------------------------------------------------
    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = FacturaSplitApp()
    app.run()
