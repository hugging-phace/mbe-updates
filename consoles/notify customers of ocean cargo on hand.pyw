import os
import sys
import smtplib
import keyring
import openpyxl
import re
import json
import urllib.request
import threading
import platform
import time
import io
import base64
import ast
import getpass
import uuid
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.policy import SMTP
from email import message_from_bytes
from PIL import Image
from tkinter import messagebox

import customtkinter as ctk
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# ==============================================================================
# PLATFORM-SPECIFIC FONT HELPER
# ==============================================================================
def get_platform_font():
    """Return the best modern font for the current platform."""
    system = platform.system()
    if system == "Windows":
        return "Segoe UI"
    elif system == "Darwin":  # macOS
        return "SF Pro Display"
    else:  # Linux and others
        return "Arial"

MODERN_FONT = get_platform_font()

# ==============================================================================
# CONFIGURATION & PRESETS
# ==============================================================================
SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587
KEYRING_SERVICE_NAME = "MBE_Automation_OCEAN"
SMTP_LOGIN = "cby@mbe.ky"
SENDER_EMAIL = "oceanship@mbe.ky"
EMAIL_SUBJECT = "Notice of Packages at Ocean Facility - Response Required"
SYSTEM_FILES_FOLDER = "Python System Files"

# ==============================================================================
# UTILITY HELPERS (PARSING & VALIDATION ENGINE)
# ==============================================================================
def parse_and_clean_emails(email_string):
    if not email_string or not email_string.strip():
        return []
    raw_list = re.split(r'[;,]', email_string)
    return [email.strip() for email in raw_list if email.strip()]

def validate_email_list(email_list):
    email_regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    for email in email_list:
        if not re.match(email_regex, email):
            return False, email
    return True, None

def normalize_reason(reason):
    if not reason:
        return ""
    normalized = str(reason).strip()
    normalized = normalized.lower()
    normalized = re.sub(r'\s+', ' ', normalized)
    normalized = normalized.rstrip('.')
    return normalized

# ==============================================================================
# PASSWORD SECURITY ENGINE (KEYRING)
# ==============================================================================
def get_secure_password():
    pwd = keyring.get_password(KEYRING_SERVICE_NAME, SMTP_LOGIN)
    if not pwd:
        dialog = ctk.CTk()
        dialog.title("First-Time O365 Setup")
        dialog.geometry("400x230")
        dialog.resizable(False, False)
        
        sw, sh = dialog.winfo_screenwidth(), dialog.winfo_screenheight()
        dialog.geometry(f"400x230+{int((sw-400)/2)}+{int((sh-230)/2)}")
        
        ctk.CTkLabel(dialog, text=f"Enter Office 365 App Password for:\n{SMTP_LOGIN}", 
                     font=(MODERN_FONT, 13, "bold")).pack(pady=(20, 5))
        ctk.CTkLabel(dialog, text="Do not use your main login password.\nGenerate an 'App Password' inside your Microsoft Security settings.", 
                     font=(MODERN_FONT, 11), text_color="#aaa").pack(pady=(0, 10))
                     
        pwd_entry = ctk.CTkEntry(dialog, width=280, show="*")
        pwd_entry.pack(pady=5)
        pwd_entry.focus()
        
        user_pwd = [""]
        def save_and_close():
            user_pwd[0] = pwd_entry.get().strip()
            dialog.destroy()
            
        ctk.CTkButton(dialog, text="Securely Save on this PC", command=save_and_close, 
                      fg_color="#28a745", hover_color="#218838").pack(pady=15)
        dialog.mainloop()
        
        pwd = user_pwd[0]
        if pwd:
            keyring.set_password(KEYRING_SERVICE_NAME, SMTP_LOGIN, pwd)
        else:
            messagebox.showerror("Error", "Password input is mandatory to authenticate.")
            sys.exit(1)
    return pwd

# ==============================================================================
# EMBEDDED ASSETS  (baked in - no external image/template files needed)
#   * TEMPLATE_HTML is the EMAIL WORDING - safe to edit the text directly.
#     Keep the {PLACEHOLDERS} (e.g. {RESPONSE_BY_DATE}) intact.
#   * MBE_LOGO_B64 / TEMPLATE_IMAGES are IMAGE DATA - do not hand-edit;
#     re-encode a new image to change them.
# ==============================================================================

# ---- EMAIL WORDING (edit the text below) -------------------------------------
TEMPLATE_HTML = """<html>
<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
<style type="text/css" style="display:none;"> P {margin-top:0;margin-bottom:0;} </style>
</head>
<body dir="ltr">
<div class="elementToProof"><br>
</div>
<div id="x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_mail-editor-reference-message-container" class="elementToProof">
<div id="x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_mail-editor-reference-message-container" class="elementToProof">
<div id="x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_mail-editor-reference-message-container" class="elementToProof">
<div id="x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_mail-editor-reference-message-container" class="elementToProof">
<div id="x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_mail-editor-reference-message-container" class="elementToProof">
<div id="x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_x_mail-editor-reference-message-container" class="elementToProof">
<div align="center" class="elementToProof">
<table cellspacing="0" cellpadding="0" border="0" style="direction: ltr; box-sizing: border-box; border-collapse: collapse; border-spacing: 0px;">
<tbody>
<tr>
<td style="direction: ltr; padding: 0cm 17pt; vertical-align: top; width: 678.4pt; height: 945.95pt;">
<p class="elementToProof" style="direction: ltr; margin: 0cm;"><img id="image_0" width="1074" height="179" style="width: 1074px; height: 179px; max-width: 1352px; margin-top: 0px; margin-bottom: 0px;" data-outlook-trace="F:1|T:1" src="cid:53b4d20d-084d-47f6-9d52-51dcd04340f9"></p>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;">
<div class="elementToProof" style="direction: ltr; margin: 0cm; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<br>
</div>
<div class="elementToProof" style="direction: ltr; margin: 1em 0cm; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
Dear Customer,</div>
{INTRO_PARAGRAPHS}
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>Next Scheduled Shipment:</b>&nbsp;{NEXT_SHIPMENT_DATE}</div>
<hr style="direction: ltr;">
{PACKAGES_ON_HAND_SECTION_START}
<div class="elementToProof" style="direction: ltr; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>Packages Currently On Hand</b></div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
{PACKAGES_ON_HAND}</div>
{PACKAGES_ON_HAND_SECTION_END}
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
{PACKAGES_REQUIRING_ATTENTION_SECTION_START}</div>
<hr style="direction: ltr;">
<div class="elementToProof" style="direction: ltr; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>Packages On Hand Requiring Attention</b></div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
The following packages are currently on hand but require invoices before they can be included in a shipment.</div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
{PACKAGES_REQUIRING_ATTENTION}</div>
<hr style="direction: ltr;">
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>⚠️Important</b></div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
To ensure all packages are properly declared and captured when we process your shipment through Customs, we strongly recommend submitting any outstanding invoices before authorizing us to ship. Packages without invoices may be delayed, suspended, or excluded
 from the shipment until the required documentation has been received.</div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
Please submit outstanding invoices as Pre-Alerts through our website:</div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<a href="http://www.mbe.ky/ocean-ship" id="OWA1f5b93f3-9250-1bbc-0618-18bb4cfb199f" class="OWAAutoLink">www.mbe.ky/ocean-ship</a></div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
{PACKAGES_REQUIRING_ATTENTION_SECTION_END}</div>
<hr style="direction: ltr;">
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>Package Location:</b>&nbsp;Ocean Cargo Receiving Facility, Miami, FL</div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>Estimated Delivery Time:</b>&nbsp;Approximately 4 weeks from the date shipping instructions are received.</div>
<hr style="direction: ltr;">
<div class="elementToProof" style="direction: ltr; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>Ocean Cargo Rates</b></div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
<b>Effective 7 July 2025</b></div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
Flat Rate (up to 12 cubic feet): CI$159.00</div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
Additional volume above 12 cubic feet: CI$8.00 per cubic foot</div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
Additional volume above 70 cubic feet: CI$6.00 per cubic foot</div>
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">
Customers are given 10 Free deliveries (Dock receipts) per consolidation. Each additional Doc Receipt is charged at CI $10 with a maximum charge of CI $500</div>
<hr style="direction: ltr;">
<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif;">
<span style="font-size: 11pt; color: black;">Thank you for choosing to ship with Mail Boxes Etc.</span><span style="font-size: 12pt; color: blue;"><u><br>
</u></span></div>
</div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;"></div>
<div class="elementToProof" style="margin-top: 12pt; margin-bottom: 12pt;">
<div id="x_x_Signature" class="elementToProof"></div>
</div>
</td>
</tr>
</tbody>
</table>
</div>
<p class="elementToProof" style="direction: ltr; margin: 0cm; font-family: Aptos, sans-serif; font-size: 12pt;">
<span style="font-family: Inter; font-size: 11pt;"><br>
</span></p>
</div>
</div>
</div>
</div>
</div>
</div>
</body>
</html>
"""

# ---- IMAGE DATA (do not hand-edit) -------------------------------------------
MBE_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAABLgAAALaCAYAAAAySeo9AAAAAXNSR0IArs4c6QAAAARnQU1BAACx"
    "jwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAFOQSURBVHhe7d0L0CVVYSfwM+/3A7emEt3ZBYpF"
    "sTABMcTJIhlZUSDBAAsuJCADArsYLdlCYyglOxg1SgSzlo9VQMH3A1gkgI+IATQmUCzMsobFlaBO"
    "YGVZNg6Iw4wyw+yc5tyx+eZ+39xH9719+v5+VV1fn7739u139/1/p08HAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADIxK/0FoOF27JR6ASB7s3ZKvQAwNCcVgEwIuABokzoD"
    "runOmUI1gPZygAfIhIALgDapOmzq9zwp7AJoFwd1gEwIuABok6oCpmHPj4IugHZwMAfIhIALgDap"
    "Ilia7tw43bj7fT8A+XAgB8jEdBflAJCjYUOlqefFfsbX7Zw67PQAMF6z018AAIAsDBNuRfH9USoW"
    "uoVeAOTDfykAMuHCG4A2mRow9Wrq+bAznn7Ok+Xvnm58AORFDS4AACBLg4ZR5VBLoAXQDgIuAAAg"
    "C70GU/G1sjT4WcrjKptuOADNJuACAABaLeVchTToWaYbDkA+BFwAAEBWBFIATCXgAgAAGq/uWweF"
    "ZgB5E3ABAAATodeQrO4wDYDqCbgAAIBWiQFVN+nlghpbAO0i4AIAACaKcAugfQRcAABAq8QAqyMN"
    "KqRBwi2AFhJwAQAArVUOtKbepghAewi4AACAiTFdyFUerpYXQH4EXAAAQOMNUxNLYAXQfgIuAABg"
    "ovQbkAHQfAIuAAAgC1XW4ip/vtw/9X0A5EHABQAAZKkcTPWiW3jV7zgAaCb/nQDIhAtwANpkmJpS"
    "U8+Jg46rqvEAMH5qcAEAAFmZGkRNDap6IdwCaBcHcYBMDHLxDgBNVUWg1O3cuKfxDvIZAJrPgRwg"
    "E90uyAEgV1WFSsOeH4VbAO3gYA6QCQEXAG1SdbDU73lSsAXQLg7qAJkQcAHQJnUFTDOdL4VaAO3l"
    "AA+QCQEXAG0ibAKgSp6iCAAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAA"
    "ZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwA"
    "AAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDWBFwAAAAAZE3ABQAAAEDW"
    "BFwAAAAAZG1W+gtAw+3YKfVCz+68887wJ3/yJ+Ghhx4Kc+fODStXrgyzZ88Oy5cvDwsWLAiLFi0K"
    "S5cuLV5btmxZmDdv3rNei13sj6/F96xYsaJ4z5IlS3b7PEA/Zu2UegFgaE4qAJkQcNGPe++9twi2"
    "rrvuujSkfjH86oRn8+fPD4sXL+4annVeW7hwYdGVw7M5c+YU5U54FvvjsBjMdV7rfA7Im4ALgCo5"
    "qQBkQsBFLzZu3Bguuuii8JnPfCZs27YtDW2nGIDFsKxbeBa72B+HxWCsHJ51aqXFWmjxc+XwrBOs"
    "dV6Ln4/hW/wLVEvABUCVnFQAMiHgYiaPPvpoePe73x0+/OEPtz7YGpcYdMXgqxOedWqnxYCsc9tm"
    "Jzwrvzb1ltBOeBYDtXJAV/58HBZDN2gzARcAVXJSAciEgItufvrTn4ZLL7206DZv3pyG0hax5tnU"
    "Wmn9hmedz5fDs/j5OO74N47PbZ+Mg4ALgCo5qQBkQsBF2datW4vaWhdffHFRewuqEAOyqeHZ1IAt"
    "Duvc9jlTeFaulRZDtAMPPDB9CzxDwAVAlZxUADIh4CKKtx9eeeWV4U//9E+LJyNCLo444ojwwQ9+"
    "UNDFLgIuAKrkpAKQCQEXX/3qV8Mf/dEfFU9IhBzF2l1vetObwvr16zXcj4ALgEo5qQBkQsA1ub77"
    "3e+G888/P9x8881pCORt1apVRbtxp512Wgw50lAmjYALgCrNTn8BgIZ5+OGHw9lnnx0OOeQQ4Rat"
    "EtuNO/3008Phhx8eNmzYkIYCAAzOf00AMqEG1+T4+c9/Hi655JLwnve8x5MRGUq8JTA29B51GoSP"
    "Og3GR52nMkaxcfmo08B8FF+L74k6T2uMOg3LR52nN0axgflYMac8jtgofWyAPio/sTGOI05L5zUm"
    "ixpcAFTJSQUgEwKuyRDb2TrvvPPC/fffn4bQFJ0nDEadICjq9JdDnM4TBKNyEFQOk+J742fib/wY"
    "CkXlcZQDqXKwVA6TOuOIuk0HNJmAC4AqOakAZELA1W7xdsQ3vOEN4brrrktD2qvfWkXdQpxyraLy"
    "OMphUnl85XF0ahiVg6ByIFUOk8rjAKol4AKgSk4qAJkQcLVTXK1XXHFF8XTExx9/PA3tXQxiuoU4"
    "e6pVVA5xyqFQtyAodt1qGE03jnINo3iLWrxVrTwOgEjABUCVnFQAMiHgaqdYc+uOO+7YYzhVrrFU"
    "bsMIIFcCLgCq5KQCkAkBFwBtIuACoEqz018AAAAAyJKACwAAAICsCbgAAAAAyJqACwAAAICsCbgA"
    "AAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICs"
    "CbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAA"
    "AICsCbgAAAAAyJqACwAAAICsCbgAAAAAyJqACwAAAICszUp/AWi4HTulXir05JNPhle96lXhBz/4"
    "QZg/f36YPXt2WL58efHaokWLwoIFC541bMmSJWHevHlh7ty5YenSpcWwZcuWhTlz5hTD4+vlYXGc"
    "ixcvLoatWLEizJo1a9d4o5UrVxZ/O8Pi6/F9AG2383jntwgAlXFSAciEgKt627dvD8cff3y48cYb"
    "05BmiaHX1NAthmVxWAzPYogW9RuwxSAtjjuaKWDrhHndhgEMS8AFQJWcVAAyIeCq3utf//rw0Y9+"
    "NJXoR6zBNlOYFmu3xfeUQ7d+A7bOsG416MoBX2dY+buA5hNwAVAlJxWATAi4qnXxxReHCy64IJVo"
    "mxh0xYCtHLp1wrRuw7oFbAsXLiy6XgO28rBO6FcO8+I44riAZwi4AKiSkwpAJgRc1fniF78YTjnl"
    "lFSC0SrXfusEbJ1bT6Opw8oBWyeI6xamlWuwzTSsE9xFnVp15WEwKgIuAKrkpAKQCQFXNe64445w"
    "xBFHhC1btqQhQFmnBttMAVu3Wm3DPHihXNNtplp1tIuAC4AqOakAZELANbyNGzeGQw89NDz66KNp"
    "CJCTGHpNDd06QVwMz2KIFlURsHWGlQO2OK74+qpVq4oywxFwAVAlJxWATAi4hvPTn/40/Ot//a/D"
    "vffem4YA9C+GXjfccEN4xStekYYwKAEXAFXS0ikArbd9+/aizS3hFjCseHvziSeeWNzuDAA0h4AL"
    "gNZ761vfGr761a+mEsBwHn/88XDUUUcJuQCgQVQLBsiEWxQH8+lPfzqcfvrpqQRQnb322ivcdttt"
    "4dd+7dfSEPrhFkUAquSkApAJAVf/7rrrrnD44Yd7YiJQm9jg/De/+U0h1wAEXABUyUkFIBMCrv48"
    "8sgj4Td+4zfCQw89lIYAues8RTHqPD1xan98T3wCYrRw4cKii8pPSiz3l5+qWH7qYrm//ITGbv3x"
    "yYr77LNPMYzeCbgAqJKTCkAmBFy9e+qpp8IRRxwRvvOd76QhwNy5c3cFNuVQp9xfDm9mz54dli9f"
    "vlt/zCRWrFixW38Ug56OOLyTX5T74/jj98zUH6czBkxT+2mXnduE3yIAVMZJBSATAq7evfnNbw7v"
    "f//7UwmqMV2QM13/dIHN0qVLi7ApKg8v1zyarkZSueZRefh0NZXK4RU0jYALgCo5qQBkQsDVm2uv"
    "vTacdNJJqURTxMClE9iUw5vpgpxeAptyf/l2snJ/uUZSL/3x9/Z0QRZQLQEXAFVyUgHIhIBrzzZu"
    "3BgOOuig4hH+bTLdrWXl28bKIU0vt5ZF09U86vd2snJ/uUZSuR9gKgEXAFVyUgHIhIBrZtu2bQsv"
    "f/nLu7a71UuoU75tbLrApqpby8r909VUKtdCAmgjARcAVXJSAciEgGtmsWH5zZs3p5JbywCaTsAF"
    "QJWcVAAyIeACoE0EXABUaXb6CwAAAABZEnABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3AB"
    "AAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZ"
    "E3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAA"
    "AABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkLVZ6S8ADbdjp9TbSieffHK44447wvz588PixYuLYXPm"
    "zAnLli0r+qOVK1emvhCWL18eZs9+5v80S5cuDXPnzi36Fy5cWHTRvHnzwpIlS3brn2m8cXh8PYrv"
    "j5+LFi1aFBYsWFD0l6cRgMHM2in1AsDQnFQAMtHmgOuuu+4Kv/Ebv5FKeYnBWic4i4FbDN46ysFZ"
    "OYSL4VgMyaLpgrNyCDfTeHsJ5MrhXpyGOC3R1PECjJKAC4AqOakAZKLNAddrXvOacM0116QS47Ji"
    "xYrU9+zgbLpAbroacvE3a3lcsb/zO7aXQC7+jeWoHPSV+6PpxgvkQcAFQJWcVAAy0daA64EHHgj/"
    "6l/9q1SC4cVwrBPIlcO5XmvIVRXIDTteaDsBFwBVclIByERbA67/8B/+Q7jssstSCego3/5aDueG"
    "vWW1PK5yzbtyOBf1EsiV27wrB33lW2FhOgIuAKrkpAKQiTYGXA8//HDYb7/9wpYtW9IQoI1iWNYx"
    "Xdt0vTwsohycDRLIlUO4cn/5O+I4ytNbHi/VEnABUCUnFYBMtDHguuCCC8LFF1+cSgDNFkOwTghX"
    "vv11ulthy8HZILesdkK4tWvXhn333bcY1iYCLgCq5KQCkIm2BVyPPfZY2GeffcLjjz+ehgDQzWtf"
    "+9rwqU99KpXaQ8AFQJVmp78AMFIf+9jHhFsAPfj85z8fHnnkkVQCALoRcAEwctu2bQsf+tCHUgmA"
    "mcRj5kc/+tFUAgC6EXABMHLXX399eOihh1IJgD2JtV5/8YtfpBIAMJX73gEy0aY2uF7+8peH2267"
    "LZUA6hUbau80Aj+1sffyUx3j+2LD7tHU98X+OCwqv2/q0xzLDcaXG5+f6X3l8ZUbpo/K31vubwNt"
    "cAFQJScVgEy0JeC65557wsEHH5xKQNPEp/t1lJ8IOFMwtHDhwqKLpr6vHOT0+r5eA6SZ3te2MKiN"
    "BFwAVMlJBSATbQm4Xv/612tLhlaYO3fus2ralIOXQWruxPEtXbq06I96DXLK74vvie+NBn0fjIqA"
    "C4AqOakAZKINAdfWrVvDr/7qr3p64gQqh0FTa+7EUCe+HpWDnKnvK9fIqTrwmWl85e8t10AChiPg"
    "AqBKTioAmWhDwPXZz342nHbaaanEdGIQ1Al8ZrpFrBwM9XrrVznwmel9M31v+X3l742/Vcu3t5Xf"
    "BzCVgAuAKjmpAGSiDQHXv/k3/ybccsstqVSN6QKVckATldsLmi7wmSmg6bUm0CDvmxogAUwCARcA"
    "VXJSAchE7gHX008/XdTgKgc+saZSfGJYNFPQVH5fucYQAPkScAFQJScVgEy0oQYXAHQIuACokmcn"
    "AwAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAA"
    "WRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcAAAAAWRNwAQAAAJA1ARcA"
    "AAAAWRNwAQAAAJC1WekvAA23Y6fU23h33XVX+MAHPhBmz54dZs2aFVasWFEMnzNnTli2bFnRv2jR"
    "orBgwYKif8mSJWHevHlFf3w9vi9+dvny5cWw+Fp8T7Rw4cKii8qfi98RvwuAPOw8ZjtoA1AZJxWA"
    "TOQUcL31rW8N73vf+1Jp9DqB2n777VeEbdM59dRTw0033ZRK4zN//vzw/e9/P6xcuTINyVNc5+9+"
    "97tTabxuuOGGcPjhh6dSCPvss0947LHHUqnZ4vYQt4XFixeHVatWhec+97lh7733DgcccEA48MAD"
    "w4te9KIwd+7c9O7qPf7448X38UsbN27cdVyhOgIuAACYQDHgysULX/jCGMaNvdv5Iz1NUXfHHXdc"
    "18+No/vQhz6UpipPTz/99I59992367yNo7vlllvSlD1jxYoVXd+XY7dkyZIdRx111I5LLrlkx49/"
    "/OM0h9XZtGlT1++d5C4uE6q3c9kCQGW0wQVApX74wx+G++67L5Xo1eWXX5768vTXf/3Xxbqnfps3"
    "bw5f//rXw1ve8pbwvOc9Lxx99NHhG9/4RnoVAGAyCbgAqJQf2oO55557ZrydsumuuOKK1MeoxbDr"
    "Va96Vfit3/qtrLchAIBhCLgAqNTf/u3fpj76lWtItGnTpnDNNdekEuNy++23hzVr1oQ3v/nNYevW"
    "rWkoAMBkEHABUCkB1+A+//nPF7ef5eaTn/xk2LZtWyoxTnE9vP/97w+//du/HR5++OE0FACg/QRc"
    "AFTm0UcfDffff38q0a/49Lqrr746lfLh9sTmufPOO4tbFh944IE0BACg3TyaFyATOTxx6qabbgrH"
    "HntsKo3f3nvvHX70ox+l0u6OP/74cP3116dSMxx22GHhb/7mb1Kp+e64447itrimueWWW8LLX/7y"
    "VAph5cqVRYA4aVavXh3+7u/+rvjbq8ceeyzstddeqUQUb8ON29BUP/3pT8Ov//qvF/2LFy8O8+fP"
    "L/qXLl0a5s6dW/QvW7YszJkzp+hfsWJFmDVrVtHF/ii+Ft8TzZ49OyxfvrzonzdvXliyZEnRH8cb"
    "xx8tWLAgLFq0qOiPf2M5iu+Nn4n+2T/7Z7vG2WQ7l4PfIgAAMGmKZ6o33J/92Z8969H64+723nvv"
    "NGXdHXfccV0/N+7uvvvuS1PYfGeddVbXeRh3d8stt6QpfMaKFSu6vm8SuoMOOmjHli1b0pLYs02b"
    "NnUdzyR3cZl009Rldfrpp6cpbLad0woAlfFfE4BM5PBj4Pd///fDF77whVQavxxrcEXnn39+uPTS"
    "S1OpuZ544onw3Oc+t5HthqnB9WznnHNOuOyyy1JpZmpw7W66Glx1LqtYIyvWAovf26nZ1anttXDh"
    "wqLr1ODq1Bjr1Bb7tV/7tXDiiSemMTWXGlwAVMlJBSATOQRcL3rRi8K9996bSuOXa8C1atWq8NBD"
    "D+263ampLr/88vDv//2/T6VmEXDt7q/+6q/CK1/5ylSanoBrd9MFXDHcPfXUU3cFUDFkimHT1ACq"
    "EzzF2xLj33JY1bkVsXPrYflWw7YTcAEAwASKAVeT/fznP9+x88fdrltkmtDleoti7L70pS+lqWyu"
    "Qw89tOu0N6Fzi+LuXbxVcfv27WmJTM8tirt3092iyHB2LlsAqIynKAJQiR/84Adh27ZtqcSwer2d"
    "bFy++93vFk/qIx/33HNP+PznP59KAADtIuACoBL3339/6qMKN998c9i4cWMqNU/TAzi6+9CHPpT6"
    "AADaxX3vAJlo+u0csVH0t7zlLanUDHtqg+uLX/xiuO+++1Kpu7jYr7766j2+rw4XXnhheOc735lK"
    "zbF169bwvOc9r2iXaNTOPffc8Cu/8iupNL0zzjgj7LPPPqkUwsUXXxy2bNmSSt3F+YrvG9ZBBx1U"
    "tO/Wq7iNxfbBfvaznxV/Y1gc27Krq0bk//gf/6NohHw6cTm8973vTaXpVbW8cjBdG1wMRxtcAFTJ"
    "SQUgE00PuN7znveEa665pmg0uazTuHJHp9Hljk6jzGVTf0hOHUenMeaOTmPNZbEcx/vSl740DRnc"
    "uBqjX716dRHQlee9CT772c+G0047LZVGa8OGDeHggw9OpWpV1bj6unXrwlVXXZVKg3nyySeLWnyx"
    "If8bb7wxDa3Gf/pP/ym84x3vSKXBVbW81q5d+6wHAlTp1ltvDbfddlsqDU7AVQ8BFwAATKAYcDEe"
    "42yM/sYbb0xT0Rxr167tOq2j6DZs2JCmonpVNa6+bt26NMZqfOtb39qx//77d/2uQbrDDjssjXk4"
    "VS2v9evXpzFWL46723f228V5pXo7ly0AVEYbXADQYB//+MdTXzM88MADldSIoXeHH3540aB/rOlU"
    "hTvuuKO4vRAAoE0EXADQYDfccEN4+OGHU2n84i1zjN6KFSuKWxUPPPDANGRwsW0vD4UAANpGwAUA"
    "DRbDiE996lOpNF5PPfVU+MQnPpFKjFpsUy62fxbbrRvW97///dQHANAOAi4AaLh4m2ITmqu56aab"
    "wqOPPppKjEN8QmMVDfw/+OCDqQ8AoB0EXADQcPF2sia0e9W09sAm1bnnnpv6Brd58+bUBwDQDgIu"
    "AMjAuNu+ijV+vva1r6US4/Sbv/mbYdWqVak0GI3MAwBtI+ACgBEYtt2k6667LvzkJz9JpdG78sor"
    "i/bABlVFu1E8Y9asWcWtisOYM2dO6gMAaAcBFwCMwJIlS8LatWtTqX9btmwpGhgfh6effnroxuWP"
    "Pvro1EcV/vk//+epbzCLFi1KfQAA7SDgAoAROeuss1LfYMbVBtY3v/nNsHHjxlTqX7yd7nd+53dS"
    "iSoMWyPuec97XuoDAGgHARcAjMiJJ54YVqxYkUr9u+eee8Idd9yRSqNz2WWXpb7BnHzyyWoMVeyJ"
    "J55IfYN5/vOfn/oAANphVvoLQMPt2Cn1MmLHH398uP7661NpMDHYeuyxx8Ib3/jG8OEPfzgN7d85"
    "55wzdODUj0cffbSo7TNM+1t33313Ec6deeaZacjgNmzYEA4++OBUqlZcP3vttVcqDW7dunXhqquu"
    "SqV6xGUQl+kgYu2vTZs2haVLl6Yhg6lqea1fvz5cdNFFqVSt733ve0U3rN/93d8N8+bNSyWqMis2"
    "KAcAAEyWGHAxHscdd1wMF4fqVqxYUYzrv/23/9b19V67JUuW7PjZz35WjGsULrnkkq7T0Wt30EEH"
    "FeO58soru77eb7dhw4ZifHXYtGlT1+/st1u3bl0aYz3idM6dO7frd/fSHXLIIWlMw6lqea1fvz6N"
    "kUmzc/0DQGXcoggAI/SSl7wkHHLIIanUv82bN4cvfOELqVS/Ydv9GrbdMXZ37bXXDlWj7tWvfnXq"
    "AwBoDwEXAIzY2WefnfoGc8UVV6S+en37298O9913Xyr1L94Kd8opp6QSVYjB1qWXXppKg7FOAIA2"
    "EnABwIj9/u///lCNrt9+++3h3nvvTaX6DFt7K7ZdFp+gSHViuDVM6HjYYYeFAw44IJUAANpDwAUA"
    "I7Zy5cpw0kknpdJgLr/88tRXj5/+9KfhS1/6UioN5owzzkh9VOGmm24KF154YSoN5q1vfWvqAwBo"
    "FwEXAIxBfBriMD73uc+FrVu3plL1PvvZz4YtW7akUv+e+9znhqOPPjqVGFasTRdrxA3T9lasvaX9"
    "LQCgrQRcADAGL3vZy8L++++fSv179NFHw/XXX59K1Ru2na/TTjstzJkzJ5UY1F133RVe+cpXFu22"
    "DRNuxfbQPvShD4VZs2alIQAA7eIqByATTX2k+o9+9KNw7rnnhoULFxbl2LbUggULiv5ly5YVIUf8"
    "Ub1ixYpi2Lx588KSJUuK/vi+TltUnfdG8b3xM7Ech0flz8W/sRyVP1eXWHNm2DApztNjjz2WSs+4"
    "+OKLwwUXXJBK/TviiCPCX//1X6dSdTZs2DDUkx6j2E5Uua2nq666Kpx55pmpNLg4bQcffHAqVSuu"
    "n7322iuVBrdu3bpifvv11FNPhUceeaRYdn/3d38XbrzxxnDnnXemV4fzrne9K7z97W9PpWpUtbzW"
    "r18fLrroolRikuw8zvstAgAAkyYGXE30t3/7tzF4a0S3YsWKolu1atWOvffee8dhhx2WpnI4xx13"
    "XNfv66eL0zXVj3/84x1z587t+v5eu3/4h39IY6vOG97whq7f1Wu3Zs2aNKZfuvLKK7u+t99uw4YN"
    "aYzV27RpU9fv7LeL67SzLfbaLVmypOu4quiOPfbYHdu3b09zWZ2qltf69evTGJk0O9c/AFTGLYoA"
    "DOUnP/lJ6hu/xx9/vOji7XsbN24MDz30UHqlmWI7VcO2iTRITaGZPPnkk+Ezn/lMKg2mippaOYu3"
    "Ena2xV67zZs3p09Xa82aNeGLX/ximD3bJR8A0G6udgAYSpMCrhydddZZqW8wsfHxYdpmmuraa68t"
    "ApdBxVtOTz755FRinI488shw8803h8WLF6chzfSOd7yjuCV5FN2Xv/zl9K0AQNsIuAAYyj/90z+l"
    "PgYRnzS4evXqVOrfww8/HL761a+m0vA+9rGPpb7BnHDCCbvaW2N83vSmN4WvfOUru9qtAwBoOwEX"
    "AEOZ2nA6/YkN5J9xxhmpNJjLLrss9Q3nf/2v/xW+853vpNJgXve616U+xmHvvfcugq0PfOADux7E"
    "AAAwCQRcAAxFwDW8YUOhr33ta0VNrmHF2x2HEcOV+GRHRi/Wmnvve98bvve974VjjjkmDQUAmBwC"
    "LgCG8vOf/zz1Mah99923aC9pULENrk984hOpNJhf/OIXQzdYv27dOo2Zj0lsN+2WW24J3/72t9MQ"
    "AIDJMiv9BaDhmvpI9Xh73Sc/+clUapZYo+hHP/pRKg3u+OOPD9dff30qDSbWsJmptlt80t0pp5yS"
    "Sv2LIdkDDzxQNKQ9iNi4/EknnZRKg/nBD35QTEc3MTyr4umKGzZsCAcffHAqVSuun7322iuV8rV2"
    "7dpw8cUXh5e+9KVpSD1yXF7XXXddsT9X5YknnnjWk1DjbaHlds+WL1++K/SdO3duWLp0adEfxfd1"
    "biON74nv7YgPBpg/f34qhbBy5crUF8KCBQuKhzl0xGNLZ7/v9v1xPOXvbZKd0+23CAAATJoYcDXR"
    "unXrYvDWyG7vvfdOUzmc4447ruv4++l2/ghNY+tuy5YtO1atWtX1s712N998cxpb/4466qiu4+y1"
    "O+KII9KYurvyyiu7fq7fbsOGDWmM1du0aVPX78y1O/fcc3c88cQTae6ql+Pyuu6669LUVyOHZbB2"
    "7do0tc2zc/oAoDLuIwCABli4cGE49dRTU2kwV1xxRerrz8aNG8PXv/71VBpMFbWzqNZHP/rRcMgh"
    "h4S77rorDWESxdpgADAJVAsGyERT/9v9mte8JlxzzTWp1Cw53aIY3XvvveFFL3pRKvUv3gL1yCOP"
    "hOc85zlpSG8uuuii8I53vCOV+hfn7X//7//9rFujpnKL4vjE29muvPLKcPLJJ6ch1XCL4jNtn8Xj"
    "TPmWws7thLHcCZeWLVtWPDG1fCtiXC/xdsN4l17ch8rDos5nyrcddm5rjMPj61H5M92+u8ncoggA"
    "ABOouJ+jgaq4fa+uLqdbFDvWrFnT9fO9dn/xF3+RxtSb7du371i9enXXcfXanXXWWWls03OL4vi7"
    "j3zkI2lOq+EWRYa1c50AQGXcoggADXL22WenvsH0e5tivDXxoYceSqXBuD0xD3/4h3/Y2AdCAAAM"
    "S8AFwFA6TwijGvE2splu9duTeJvjHXfckUp7dvnll6e+wey///7hsMMOSyWaLgaot9xySyoBALSH"
    "XyUADKX8aHuGFx/nf8opp6TSYHoNrWJ7XTfccEMqDeass85KfXSsW7euCJH21P3VX/1V0SZUp4vt"
    "lF1yySXhvPPOC0ceeeSudpmqtG3btiJEffjhh9MQAIB20LAjQCaa2l7JG9/4xvDhD384lcYr1nyK"
    "Da137LfffpU8QW5Ujcx3xBpYa9asSaX+xeUQA4xOI9TTufjii8MFF1yQSv2Ly/qHP/xhWL16dRoy"
    "vUlqZH79+vVFw/3D2r59e7j99tvDpz71qfDpT386bNmyJb0yvBNOOCH81//6X1NpMFUtr6OOOioc"
    "ffTRqVSvV7/61cVxgWbQyDwAVXJSAchEUwOuWBPlnnvuSaXiB8tuNU86TwPriE/8ik/+6ig/Eaxj"
    "T58pP42sbqMOuKL4NMV4u+GgLrvssnDOOeek0u7i5vSCF7wg3H///WlI/4455pjwla98JZVmJuAa"
    "Tgws3/KWt4TPfe5zacjwbrzxxvC7v/u7qdS/Ji8v8iDgAgCACRQDLsZjlE9R7IhPQ+w2nl67Qw89"
    "NI2pu1tvvbXr5/rpvvSlL6Wx7dkkPUVx/fr1aYzVi8tx7ty5Xb+33+7AAw8snqI5qByWF822c/0D"
    "QGW0wQUADXT66ac/63bLft15553hu9/9birt7uMf/3jqG0ysufN7v/d7qcSonHHGGeHaa68datvo"
    "iDUEY9tfAABtIOACgAZ6znOeE0466aRUGky8TbGbeGvZNddck0qDOfXUU4vbRhm9GCy+973vTaXh"
    "fPCDH0x9AAB5E3ABQEOdffbZqW8wn/3sZ8PWrVtT6ZeqaLA81iRifM4///xw2GGHpdLgbrvttrBx"
    "48ZUAgDIl4ALABrqiCOOCPvuu28q9W/Tpk3F7WxTDXt74kEHHRRe8pKXpBLjENvmfs973pNKwxn2"
    "aYoAAE0g4AKAhopPihy2ptTll1+e+p5x1113Peupl4NQe6sZDj/88HDggQem0uC+8Y1vpD4AgHwJ"
    "uACgwV73utcN1aB4vAXtgQceSKXp2+XqVZyW2P4WzXDCCSekvsF961vfik9pTSUAgDwJuACgwVav"
    "Xh2OPvroVBpMpxbXz372s/DFL36x6B/Uq1/96rBq1apUYtxe9rKXpb7Bbd68OfzgBz9IJQCAPAm4"
    "AKDhzjrrrNQ3mE984hNh+/btRbj1+OOPp6GDOfPMM1MfTXDAAQekvuF8//vfT30AAHkScAFAwx17"
    "7LFD1Zp69NFHww033BCuuOKKNGQwcRqOOeaYVKIJVq5cmfqG8+CDD6Y+AIA8CbgAoOFiu1fD1py6"
    "8MILw+23355Kg4mNyw/THhjVq2p9xBC0jS666KLiiZPDdo899lgaIwDQVAIuAMjA2WefnfoGc++9"
    "96a+wa1bty710RQ/+clPUt9wnnrqqdQHAJAnARcAZGD//fcPa9euTaXRW7NmTTjwwANTiab4n//z"
    "f6Y+AIDJJuACgEwM29j8ME4//fTUR5N861vfSn3DWbRoUeoDAMiTgAsAMvGa17wmrFixIpVGJ4Yf"
    "f/AHf5BKNMWOHTvC1VdfnUrDWb58eeoDAMiTgAsAMrFw4cJw2mmnpdLoHHfccWMJ1pjZV77ylXD/"
    "/fen0nD+5b/8l6kPACBPAi4AyMg4blN83etel/poim3btoU//uM/TqXhHXDAAakPACBPAi4AyMiL"
    "X/zicMghh6RS/VavXh1e8YpXpBJN8ba3va2SJ2NGS5YsCfvuu28qAQDkScAFQGW+/OUvh6uuuqro"
    "rr322qIcu1tuuSXceuutRbdhw4bw3//7fy+6H/3oR0W3cePG8NhjjxXd448/nsbGdM4+++zUV78z"
    "zjgjzJ7tcqFJ/st/+S/hfe97XyoN77d/+7etYwAAAEZjRwaOPPLIHXFSq+xWrFhRdHvttdeOvffe"
    "u+j23XffHQcddFDRHXLIITvWrl1bdPH7jzvuuKI76aSTdpx//vlpyoYTx9dt2vrp4jxU5bHHHtux"
    "aNGirt9TdfcP//AP6VuHc+WVV3Ydf7/dhg0b0hirt2nTpq7f2W+3fv36NMZqPfXUUzv++I//uOt3"
    "DtO9//3vT9/Qn6YvryiOu9t39tvFeaV6O5ctAFTGv+sAqEwd7fjEGl2x2/kDs6jpFbsf/vCH4Z57"
    "7im6u+++O9x2221Fd/PNN4frr7++6K655pqiFlkbxQbf4xMV67Z27dqw3377pRLj8vTTT4ebbrop"
    "HHzwweHiiy9OQ6vze7/3e6kPACBfAi4AKvP85z8/9VG3UdymeOaZZ6Y+RikGWg888EC47rrrwn/8"
    "j/8x7L333uHYY4+trM2tMiEmANAWs9JfABouh9s5vvGNb4RXvepVqTR+MRiIbXxN59FHHw2bN29O"
    "pemdc845Re2wYcRaV7H9sVmz9nzqXb58eXjOc56TSt3FzeEFL3hBuP/++9OQasWGxx955JHi70y2"
    "bt0a/s//+T+pNL1Ym+4tb3lLKg3uK1/5SnjhC1+YStP71V/91bBw4cJUCkXNvz3tQk888UT49V//"
    "9VQaXNzu9tlnn1TqTWx/7sknnwz/7//9v2I64lMSR+Ezn/lMOPXUU1PpGTFg+8d//MdUmt44l1ev"
    "Om38DSvWIF25cmUqUZWdx0O/RQAAYNLEgKvp/umf/mlXmzVN6GJ7XTOpom2tOrrzzjsvTeHM/vzP"
    "/7zr56vozjrrrPQtM7vlllu6fn7cXZyustgGWrf3TXK3//77F+16TVVV21pt6rTBVY+dyxYAKuMW"
    "RQAqE2sd7fzRnErU7fTTTw9z585NpWqtW7cu9dFWb3vb22rbfgAARk3ABUClDj300NRH3X7lV34l"
    "vPrVr06l6sSQ8mUve1kq0UZr1qwpAlIAgLZw3ztAJnK5neODH/xgeNOb3pRK47WnNriOP/744omL"
    "TXPeeeeF//yf/3MqzeyrX/1q+J3f+Z1Uqsa73vWu8Pa3vz2VZnbrrbeGI444IpWa45Zbbgkvf/nL"
    "UykU7SfFp3ESilpb3/3ud6d96mlsD2yvvfZKJaLp2uCK21Q8zkSzZ88u2s+b2j9nzpywbNmyon/e"
    "vHm72rUr9y9YsCAsWrSo6I9tx3Xaj4vD4mvR4sWLw/z584v++Ln4+an98Xv23XffPbbh1xTa4AIA"
    "gAlUNFiSgb//+79/Vts14+za3gZXtG3bth2rV6/uOp5Burlz5+548MEH09j3TBtc+XWXXHJJWird"
    "aYNr9266NriauKz+5m/+Jk1d8+2cXgCojFsUAajUgQceGJ773OemEnWLtUPOOOOMVBreK17xirB6"
    "9epUom1OOeWUcP7556cSOYo18OJTWWO3atWqogZZ7OLTReOtp7EDgEmkWjBAJnL6b/frX//68NGP"
    "fjSVxmcSblGM4jzG25Kq8IUvfCGcfPLJqbRnblHMR1xPX/va13bd5jYdtyjubrpbFLdu3RouuOCC"
    "oj8u13gbYVS+5bDXWwtjWB11vifevRdDrKh8m2ObuEURAAAmUHE/RyZuvvnmZ90yM65uEm5R7Djy"
    "yCO7jqufLt7Gt/MHexpjb9yimEd31FFH7di8eXNaGjNzi+Lu3XS3KDKcncsWACrjFkUAKhdrzrhN"
    "cbTOPvvs1De40047bVeD1rTHOeecE/7yL/9yV+0iAIA2EnABULl4O81rX/vaVGIU4u2WsT2eYZx5"
    "5pmpjzaIt8BdccUV4bLLLtvjbYkAALkTcAFQiypqFNG7WPNqmFDxoIMOCi95yUtSidwde+yx4e//"
    "/u/DWWedlYYAALSbgAuAWuy///7Fj+xRi7VWOk8Y69YodJu97nWvS339U+OuHY466qiigf0bbrgh"
    "7LPPPmkoAED7eXIJQCZybJD33nvvDbfffvuup4NF5aeFRf2Wly9fHmbP/uX/Z6aW+9GWpyiW/dZv"
    "/VaxzPsxd+7c8OMf/3igWxw9RXH8Ypgct+XY1lbsH5anKO5uuqcoMhxPUQQAgAlUPHKKSrXpKYod"
    "V1xxRddxztSdcMIJ6dP98xTF0XarVq3asWbNmh3nnnvujssuu2zH97///TSH1fEUxd07T1Gsx85l"
    "CwCV8V8TgEz4MVC9b37zm+HBBx9MpeZ44QtfGF760pemUn82b94crr766lTqzZo1a8IBBxyQSv15"
    "+OGHw9e//vVUao54q175SZ6f+9znwi9+8YtUaq5YoSXeXtsR+2Mtxdj9i3/xL8LChQvTK/WJyyku"
    "L37pD/7gDzTUXwM1uACokpMKQCYEXAC0iYALgCppZB4AAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4"
    "AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMjarPQXgIbbsVPqzd7TTz8d/vEf/zHM"
    "mjUrrFixIg0NRX8cBkD77TzeO+ADUBknFYBMtCngip7//OeH+++/P5V2t2TJkjB37tyif/HixWH+"
    "/Pm79ce/sRzNmzev+Ew0e/bssHz58qI/Kgdny5YtC3PmzCn64/vj56Jexhs/Fz/fsXLlytQXiu+L"
    "3xstXbp017QD0J2AC4AqOakAZKJtAdfFF18cLrjgglRqr0WLFnUNzsr9gwRn5XAufseCBQuK/vg3"
    "lqPyeAcJ/RYuXFh0AHUQcAFQJScVgEy0LeB65JFHwurVq8O2bdvSEJosBmudmmzlcG5qcFYO56YL"
    "zqYL5HqtLVceb7m/l/ECzSHgAqBKTioAmWhbwBWdfvrp4dOf/nQqQf3KQV1UDs7K/eXacuXbV8sh"
    "2nS18OJv9unalusl9CsHcnEa4rREU8NEyJ2AC4AqOakAZKKNAdedd94ZfvM3fzOVgH6UQ7RycDZd"
    "IFe+5bQcok0N5OqohTf1O8q33jK5BFwAVMlJBSATbQy4ole+8pXh5ptvTiVg0sRArFttuXJ/ubZc"
    "uSbbIG3LlQO58ninC+ei8njL4Vx5vPRPwAVAlZxUADLR1oDrG9/4RnjVq16VSgB5iqFZt0Cu3D81"
    "OCvXlisHZ51bZGP+8853vnPXbaptI+ACoEpOKgCZaGvAFcXbFOPtigD80nHHHRe+/OUvp1L7CLgA"
    "qJLGDwAYu7e97W2pD4CO9evXpz4AYE8EXACMXaylcNBBB6USAPG4+OIXvziVAIA9US0YIBNtvkUx"
    "irfhnHDCCakEkL/Yjlan4fqpDeKX29/qtLkVdZ52+Y53vKP1wb9bFAGokpMKQCbaHnDF2Yu1Fe65"
    "5540BJhE8YmFHeUnIfbyVMWo3Fh7bPg9PjUxKj89caawqfyd5c+Un7IYc5nydJa/szydzEzABUCV"
    "nFQAMtH2gCtSiwv6N93T+2YKccohTDkEirWIyk/smy746TVsKn8mhkMxJIrKwdHU72RyCLgAqJKT"
    "CkAmJiHgUouLUSjfNhZNd6tYOfgp197pNcQpfya+Ht/XEWv/dH7blz8z3XfG905XswlyJeACoEpO"
    "KgCZmISAK1KLq5li0NIJfqarvdNriFOuvTP1M+WwqZfgZ+rny8FP+fOddo2A5hBwAVAlJxWATExK"
    "wNWWWlzlWkIz3SrWrXHpqNfaO+XgZ6bgaLqwqfydcTo6t4pN/U6Aqgm4AKiSkwpAJiYl4Ir+8i//"
    "MrzhDW/YFeL02t5P7I/DonJwVA5+ZgqbpguOZgqbyt9Znk4AZibgAqBKTioAmZikgAuA9hNwAVCl"
    "Z/7lDAAAAACZEnABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3AB"
    "AAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZ"
    "E3ABAAAAkDUBFwAAAABZE3ABAAAAkDUBFwAAAABZE3ABAAAAkLVZ6S8ADbdjp9TLDK677rrwyU9+"
    "MixZsiTMmzevGDZ37tywdOnSoj9atmxZmDNnTtEf3xPf27FixYowa9Yzp8dFixaFBQsWFP2zZ88O"
    "y5cvL/qj8mtxXHGcHeXvnvoaAM/Yeaz1WwSAyjipAGRCwNWbrVu3hjVr1oR77rknDWmWGH7FwC2a"
    "P39+WLx4cdEfrVy5MvXt/tp0wVscFl/rWLhwYdFFU0O5mYK38mtTQ79yIBi/N35/R3m6APoh4AKg"
    "Sk4qAJkQcPXu/vvvD4ceemh4/PHH0xBGIQZfMZjrKAd2MazrvDY1XCu/NjWUizXvpgsEew3eyq9N"
    "DQRnCgvLoR9QPQEXAFVyUgHIhICrP/FWxX/7b/9tKsFwYsg2Xa22qcFbOdibWhuuHLzNFK6VX5up"
    "tt3U18rTNfW7Y3AYA8Ro6jTDOAi4AKiSkwpAJgRc/Xv7298e/uzP/iyVgOmUw7WZArty8DbTLbD9"
    "tHs3U1g4SCA4dbpoLgEXAFVyUgHIhICrf3GRnXjiiUVtLmAyxfArBmfRoLfATlcTL7Z3d9FFF+1q"
    "947+CLgAqJKTCkAmBFyDefLJJ8Phhx8e7r777jQEoBqf+tSnwmtf+9pUol8CLgCq9ExDDADQUrFG"
    "xo033hhWr16dhgAM74/+6I+EWwDQIP5rApAJNbiGs2HDhqIm1+bNm9MQgMEcc8wx4YYbbth1CyOD"
    "UYMLgCqpwQXARHjxi18crr766l3t7QAM4oUvfGH43Oc+J9wCgIbxXxOATKjBVY2rrroqnHnmmakE"
    "8GydpzOWG52Pf2M5dp/+9KfDC17wgmI4w1GDC4AqOakAZELAVZ13v/vd4cILL0wloE4zBUZR+UmG"
    "K1euLP6Wn1oYn4K4YMGCor/8ZMM9vTcOi69F3d4b3xffH5WngdERcAFQJScVgEwIuKr1xje+MXz4"
    "wx9OJchHHYFRzBk6463yvTATARcAVXJSAciEgKta27dvD695zWvCddddl4YwqfYUGHVCotmzZ4fl"
    "y5cX/XsKdsrvLdcU6vbe2MX+qDze+DeWI4ERbSTgAqBKTioAmRBwVW/r1q3hyCOPDN/5znfSEIYV"
    "aw51App+AqNyCFSugVRXYBTHFccJjI+AC4AqOakAZELAVY9NmzaFww47LNx3331pSLNUHRjF2krx"
    "N+V07y2HQJ33lgOjhQsXFl3U7b0AvRJwAVAlJxWATAi46vPggw+Gf/fv/l3YsmVLUd5Tu0V7auOo"
    "HBiVby3rvLeXwKjzXoC2EnABUCUnFYBMCLgAaBMBFwBV0vgEAAAAAFkTcAEAAACQNQEXAAAAAFkT"
    "cAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAA"
    "AFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEX"
    "AAAAAFkTcAEAAACQNQEXAAAAAFkTcAEAAACQNQEXAAAAAFmblf4C0HA7dkq9tMT3vve9cO6554b/"
    "+3//b5g/f34aGsLKlStTXwgLFy4suo7ly5eH2bOf+f/U3Llzw9KlS4v+aNmyZWHOnDlF/7x588KS"
    "JUuK/ij2x2FRfE98b8dMry1atCgsWLCg6J81a1ZYsWJF0R9NnTaAfuw8pvgtAkBlnFQAMiHgaqfN"
    "mzeHt7/97eEDH/hAGpKvcvgVA7vFixen0rNDu6mvzRTMxQAvBnnRTOHbTK/FQDAGgx3l16I43Z3f"
    "2XF4fL2jHCgC1RJwAVAlJxWATAi42u2OO+4If/iHfxjuvvvuNISmiUFbOXwrB3NTQ7tyMDdTwBZ/"
    "309XK27qa1PDt5mCuZlCw5lqAcIoCbgAqJKTCkAmBFztt3379vCRj3wk/Mmf/El4/PHH01AYnRiE"
    "TRfMxQCvcyvt1Bpz5dcGrTEXx3nKKaeEVatWFWXaT8AFQJWcVAAyIeCaHI888khx2+LHP/7xNATa"
    "be3ateHSSy8NL3nJS9IQJoGAC4AqOakAZELANXnuuuuucN5554XvfOc7aQi0y7777hve9773hRNP"
    "PDENYZIIuACoklZTAaChYm2Wb3/72+ELX/hC2HvvvdNQyF+8TfHP//zPiyeJCrcAgCr4rwlAJtTg"
    "mmxbt24tAoH3vve9YcuWLWko5CW273X22WeHP/3TP9XWFmpwAVApJxWATAi4iB5++OGiEfpPfvKT"
    "Ydu2bWkoNN8xxxxT3I544IEHpiFMOgEXAFVyUgHIhICLsnvvvTdccMEF4cYbb0xDoHqxxlV8smLn"
    "KYkrV64shse/8+bNK17rPCUxPlUxvj/efth5rfOUxP322y8cfvjhxWehQ8AFQJWcVAAyIeCim1tv"
    "vTW8+c1vDnfffXcaQpvFsCgGTTFMiiHS8uXLi3IMoDohVBw2Z86cImjqvNYJoWLoFD/XLaDqvBY/"
    "F4Oq+B1QJwEXAFVyUgHIhICL6cRN4/Of/3y48MILww9/+MM0lFGIYdDs2bOfFTSVazR1QqgYGC1d"
    "ujQsXLiw6OJrMYSKQVP8G8ud2k6d18pBU+dz0CYCLgCq5KQCkAkBF3vyi1/8InzsYx8L73znO8Oj"
    "jz6ahk6Ozu10U2s0xRCpE0JNrbUUw6Opt9V1q+00NYTqvAYMTsAFQJWcVAAyIeCiVz/72c/CpZde"
    "WjTovXnz5jR09DqBUzloirrdHtctaJpao6kcNHVqNHWCpvg5v5UhLwIuAKrkpAKQCQEX/Yq1uN71"
    "rneFj3zkI8UTF2NYFIOmGBDFUEj7TcA4CbgAqJKTCkAmBFwMavv27UVoBdAkAi4AquSkApAJARcA"
    "bSLgAqBKs9NfAAAAAMiSgAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4"
    "AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACA"
    "rAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsA"
    "AACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMia"
    "gAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAA"
    "AMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAACArAm4AAAAAMiagAsAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYsVnpLwAAAMBE2LFT6p3W"
    "rJ1SLxmwsgAAAICJ0EuwNZWgKw9WEgAAANBqMwVb5QCr1/fRPFYOAAAA0FrdQqtewqpBP8d4WDEA"
    "QCN0u4jsxoUlAAyu1/NtR+7n3anzO8j8VDEO6melAAAjNfUisSouNqF6g+6v9kdohkk/506d/850"
    "97Jcps5j+TNTX6MZrBQAoFa9XETWwcUn9K+O/dW+CKOT0zl32Gnt5TvL31F+fz/fPd3nysNpBiuk"
    "T/3sCHVrww7VpOUZjXqZNmH+c9uOmrbNlJWXZVOnM7f1vSdN3h4G0ab107R1U9eybfI22O885zIv"
    "TZ7Ocel3XXcziuU6dToH/c4q5nc6o1gOuem2vC2nPatzO+2mKeukn/kedpr39F1Tx19+f7/f3fls"
    "+XPl8dEMs9NfMtTvTtk0uU8/QNXicbEjDcpOmvzGTX+aLOcdmCLtGiPZN0b1PTApip03SYPGLk1O"
    "IQ1qhJnCqPhaN+nlZykPb9o8IuACgEYqrgyTNKjR0qQ2flrTZLogZeKlXWHk+8I4vhPapth5d0rF"
    "xsphGmcyXchFcwm4Mpf7QYPxsw1B88X9NErFRkmTlt1xJE224x8Tadzbvn0PBhP3nSgVs5AmeeTT"
    "XP7OQYOq6aZb8NVcAi7GYrqDBQDTa9KxM05LlIrZSrPhnMTEaMr2br+D/uS+z+Q2/bkv70kl4AIc"
    "wCEjTdhf23jMaOM8wVS2c8hP3G+jVMxampXGzUuarGdJL5EZAVcL2AEBJsu4jvvxe6NUbJ00e86p"
    "tJJtG/LT1v02p/mKtyNGqUjDCbgYubYeqHNnvUBeRr3PTtIxYpLmlclgm4b8tH2/bdL8pQxrtxCr"
    "2zCaTcAFAJka1cVhky5CR2US55l2si1DXuI+G6UiAxoknJr6GeshPwKulshl53OQAKhWncfVOO4o"
    "FSdOmn3nLbJl+4W82Gfr0c9y7SUYs56aS8AF7OJgDXmqY991PPglywKAujnXNJP1khcBV4vY+QCo"
    "gvPJ7iwTcmObhXzYX+vXzzKeqRZXeTy91PZitARcjEw/BxXGx3qCPFW17zoGTM+yIRe2VciH/bU+"
    "wwRQ5c9aR/kQcAEABRdwe2YZwbPFH4FRKgJ9cE6pX/n4VF7exYErSYN2k14uxPLUz6deGkTA1TLl"
    "na5JmjpdAJCz4qp7GuktPUsfm1Z6Gw1X9zVX2hx2SYN3+yEI0ESDHCPjZ6JUpMGcgGoyzh2giRcW"
    "4z4g5H6xNY7lZ5n9Up3LoqrpzH191W3c20OV378ng24Lo5zGNuh3Ode1fAdd3/2qc/uoax7qnOam"
    "KS/DuuZ7kPXUpGmpWl3z1kTdlvckzf+g9rSdWobD29MyLuu2vHv5/KCfYzysmJp02xFGqWk7neUx"
    "nHEsP8vsl+pcFlVNZ+7rq25N3B6qnKayQaavrmlpu36XdZPWeT/q3D7qnPaqpruuaaxyuZanscrx"
    "RlXMfxOnaVhVzVNd81LlMu82jcOMv995rnteZlLXfFY5T4PodzlE457mbvqdjz3NQxzfTO/p9/sY"
    "PbcoUrs9HUiY2biWn/UG9WrKRZJ9fXD9Lru61nmd67DOcTdlH2iTqteXdQTVq/O4OpO4P3ekQX1J"
    "H90lDc5KmvRpp32mdTPT52gOARcAjElxlbVTKo7cuC6y26QpyzC3dTnO7Z7Rs75hPOK+15EGVSaN"
    "Nst9O016T9Oe3uoYlgkrqiZNuNBsyo5oWQzOshtclcuuzmVQ1XTmup5GpenbQ5XTF/U6jYN+bx3L"
    "oAlGtTwG/Z5eVLlu6prOKqdxJlVNf13TW+Xy7UxjHeOsUlXTV8e09avp81LV9EXdpnGY8fc7z3XP"
    "y0yqns8q52Um/c5nFUY1b2XjmE+aTQ0uajWOAx1AbiblQrSt+l2WOVyQ2z4AqjWq4+q4zjHxe6NU"
    "hLEQcAFAA7gonCx1re8qfkDV+SPMdp6HutaT9Q/1iftXlIpj04RpYHIJuFqszgvUXoz7+3PXlOVn"
    "PUL72K+r16Rl2tT160cPMKnqPi437fgapydKRRgZARe0lJMK0E1Tw49JVOdxetD1XNf24ZxUv7rW"
    "HdBsTT6+OvYzagKulnOxk6eq1ltVJxXbEcCeDXKsbNLFf13H+ibNI8Co1XVsjXI4vjoHMEoCLmpR"
    "54GcPXMigTzZdydTXeu9n3Ox8zajErf3YaVRwUTLaV+w3zIqAi5oGD8ygLo4vtRv0GVc18X/uNe5"
    "HzXAJKvrGJzjsdX5gFEQcE2AUV/cjvr7mFlVJxPrFepXxX7mApKp9rRdVbHddWNbHC3LGyaDfR2m"
    "Z+eoSVUXi/EAVsW4RnkgrGreqzTK+R/WsMuv27xWtU4maTmW1Tnfk7huxmGStofppi+XZTBOTVhG"
    "VU5D2XTTM+rvG6Wq5q0J89KrKtdnTvM9DpO4fZUNM//9znNVyzoa5XfH76py2qN+p7+JqlwmbVge"
    "VEsNrglR9cG1bpN6sGr6esptOwLITV3nv27H77qO6ZN6Dgcoq+sYmzvnCOpk46pJVQe0eACoclyp"
    "tzZNnO9oFPNehSrmebp5rWp5TtKy7KhznidtvYzLJG0P001fLstgnJq0jKqclrLOdNU1/mjYea9K"
    "VfPYlPnpRdXrNad5H7VJ3L7GpcrtepTLu8rpjmwrsGdqcGXAwWwyVHESHMW2UvXJGnhGnccA++3o"
    "DbvM6zqe170t1DXdjEfd2wsAVEnANUFyuUhxcVwPyxWay49IRqmu7c15pp0cn6B/Ve83jq/QGwFX"
    "JnI4qLkAGi8nPsiTYyfTyem47hzUDHWtB8cpAHIg4KJRJvUCObcLRxe6UI0q96Xpjp/21/GpYtnn"
    "cF7MYRoZXtyeo1QERsQxFnon4JowdV2YuODJgxMkNEM8ZkapCDNy7KZJ0uGrkAYBQCMIuDLiAred"
    "qrhAHMe24cKWSRW3/WGlUVXG+aH9mrqObXvNM8p1kg5phTQIJp79AcZHwDWBmnrQdZEM0D/HTsbF"
    "tkdZvL4sS4OBITjOQn/sMDWp6sQ+9aBW13iH0fR5jaqc3yrVtez2ZFzfO0pVzWNU53xOwrpogiq3"
    "hybpZb3XMe9t3d5yWFZ1TOMgctgGmrKs+lHlcm3S/OewvfRr0revUapyWY9iGeQ2vdAmanBlxkGO"
    "qca5TVR5Agf643wwmZqw3m17eWjSeorXCx1pEABUTsA1oaq6wKhqPC6WR8vyhrz1ug9XdYxmcHWs"
    "g3Eew8f53bRD3Cc60iAAqISAK0MuLtuhLRd2LlBhtJwDgF41/XgRryGiVASAoQi4JlhTLij8WBuc"
    "ZQeTI+7vUSoy4caxLdj+8pTDeitSrp1SEdjJMRf6J+BiYC5EBteEZVflSdO2APWznzHVKH/8jPK7"
    "qF4u6y8e56JUBIC+CLgy5UIT2wBMnvTbz48/dhnFucD5ph1yWo+OcwAMQsA14cZ9ATGJF80u2oBh"
    "OY5QVue5dBLP020W12eUio0Wj3NRKgLAHgm4MjbOCxQXHO1Q5TZkm4DRivtclIpMsDq3A9tYO43z"
    "GrJftkEAeiXgggzldGEK1MuPv8k2ivVvG2uneC0RpWKj2QYB6IWAi74vGqq6yMjloqpKVS27pmr7"
    "/EFT2fcm0yjXu22svYRcALSFgCtzkxgSUS3bELSDH3/AoOK1QJSKjeU4xySxvUP/BFwURn0AzeEi"
    "qmpVLeNJXHZAb0Z9LGd8xrGubV/tF68xOtIgAMiGgIu+uLhtpyovZG0jAL9UR1AwzuOsY/zkiNtu"
    "lIqNYRsEYDr+O1OTqk6+vV5YVPF9vXzXKOerqu+Kevm+OlU5LzkY9/KOctl+qprOJizzJmvS9lDl"
    "tEynM411fdewy6Cp6lheVS+rutZpv3LYBpqyrPphufZu3MvK9jU6VS7rUSyD3KYX2sQOU5OqDmy9"
    "HtRG8X25zlPU63fWpcp5yUWblnmd81LVdI57eTddU7eHKqdrqs501vEdVS6DJmn6sqpj+obR9O2g"
    "quXV9Pkcp3Fuk+NeL7av0alyOxvF8s5teqFN3KLYEg5+APmJx+4oFQGykg5hhTRoZKoMEaCpbOfQ"
    "HwEXz+IgWr1JXaa2JehdHT8O7YPt0sT1aRujrEi5kjQIAEZKwEVPqrqIddED0J3jI9NpcpDU5Glj"
    "fOLxLErF2tj+aKJRbPtAdwKuFnEwbR4XXkA/HMfbpYr1mcN5xLmO6cR9IEpFYACOsdA7ARe7qesg"
    "6gJn8jghQ/+qPFbGfdCxN185HUNzmlZGz3EIgFEQcLVMHRcQLloHY7kB0ETCBsYhbndRKgJ98LsC"
    "eiPgAmrlhAzQv7qOnXUGDI73AM8Q5sJ4CLjoqnORWtXFqoM8QO8cM/M3zDqs6tw7VXmahpm+mdQ1"
    "7bRH1duebY5JYVuHPRNwtVBdF630zgkIaIp4PHJeyEdd549u20Bd20Vd8wAw6dpyfI3zUaU0WhBw"
    "MT0HC6piWwLYs3EcK4VcjENd2x00ie28O+cH6mSnq0lVO+6gB8YmHTiaMA+DTsMgmrTsm2SU6yDK"
    "ZfupajrrmMZclmEvcpyXOqa5qnGOahmM2riXT1XfP1Uv01PHdw+6HKpS1TzVNR9VLvNxL+tB5D7/"
    "VU1/XdOe+/Ity3leqpz2snGvk0HVsTxyXRbUQw2ulrKj0zR1neAB2qCuY+Q4rwcc90cvLvNhpVEB"
    "FajrGJzjvur4wigIuKBCVR+440lxXNIkANCHQY6fdV309zMtg0x3L+qaN4BJl9Px1bmAURFwtVhd"
    "F6v9aMI05Grcy66O73dyg/Ea93GF3dV1XBxkXde1fTj2A9Qjh+NrndNY13mLfAm4AKBhcrhgZXf9"
    "Xmg3cT3X9WPBNp0P6wqqVddxtaPJ+6zjCaMm4KI2dR/Mm6aNB/BJW4cwCezX7TfsOraNAOSlib9D"
    "6p4m5yq6EXC1nB0/T21eb008AQMMq9/jdl3HwiafPxz/qUOTt3noGMV2Go+xUSqOTZoMx3vGQsAF"
    "FXAQB6pS9fGk20X1KC60mV5d54wq12td20hd80616l5PtgMmUV3H1anGuX+N6rtHtSzJj4BrAozj"
    "AOCgM7imLbs6pmecJ15oslHuG0071uSsn2VZ1zquY33WMc5olNs5wCSKx9koFWuXvm4k31fXuYl2"
    "sHHUpKodvKodeFQHnI4qprvKaa5qOXZT9bKtc1oHVfU8RnXPZ5XTXOe0VjWddUxjLsuwFznMS5XT"
    "WDbT9Nb1nZOkn+2hzuXdz3T0q67prnOao6qmu67prHK5dqaxynFGdcx7DtPYi6rmo67pr3I5j2sZ"
    "d7RpXqIq56cfVc97W+aDdlGDC9gjJxKol4tEhmE9Upeqj03jOtZBk4zrmB33v440qG/p44U0aKTi"
    "sktfP5Q0OlpIwDUhRnkgHeV3tc0kLTsnFyZdcYWVpEGV6+WYMknHnar1s+zqWs+jWH91fUed2z7V"
    "qWo91bG+R7H9QxvF/XEQ6eNjYX+nFwIuGMK4D/TQUVx1VCyNmiGlxbmb9DKZ6udCu671PcqL/bq+"
    "y76Qh2HXk/UMzzbK4zdMEjtWn5p8gt7TgXIU097LwTqHZTiuaexl+Y1TE5bLuKahF7lM5ziUl02V"
    "LOeZ9bPcLcvedZbrOJdZP+u2Sk2Y53FOQ1ONctl0vqsXdU9PP9PSq1Esw9zUsZyjJi/ruuZ5Kttb"
    "b8rro4plNqr1y+ipwTVB7MjNZv0AVer3mOIYlJd4gT+T9La+pY9PK72NCZY2hV3S4EIatEsaXAvH"
    "LNrAdrxnlhH9EHBBQ6RrwT1Kb69cGv2M0luBlnIRuWeWETxbukQopEFAH5xXpmfZ0C8B14Sp8yDh"
    "AATQDMMcjx3Lp2fZQDPYF2kb2/TuLBMGIeACAJ7FReXuLBNoBvsibWXb/iXLgkEJuACgRaq6KHRx"
    "+UuWBQCj4HxjGTAcAdcEquOg4UAEMH5VH4sn/dge5z9KRWDM7I9Mgknezu3jDEvABQAtUNdF4aRe"
    "bE7qfENT2SeZJJO2vcf5jVIRBibgmlBVHkAcjADGq+7jcBx/lIqtN0nzCjmwTzKJJmW7t39TJQEX"
    "AGRslBeGk3AROgnzCEAe4jkpSsXWafO8MR4CLgDI1DguDON3RqnYGmm2XGhDg6Td0n7JxGvbflDs"
    "2DulIlRGwDXBqjioODABjMe4j79tOf7H+YhSESZW0/YD+yU8W9wnolTMUpoF+za1EXABQEbStWEj"
    "Lg7TpGR7oZrztEMdmrJP2DdhenH/iFIxGzlOM/kRcAFABoqr2Z1SsVHSpGVx4ZomtZAGASXj3DeK"
    "HXOnVARmkHaXRu8vaRILaRDUSsA14YY52DhQAdQrHmc70qBGS5NaSIMaI02W8xb0YBz7iv0TBhP3"
    "nY40aOzS5NinGTkBFwA0RLoe3CUNzlKahbHOQ5qEQhoE9CjtOrXuO+krCmkQMIS0O41lf0pfXUiD"
    "YORsfADAyOzYKfXWwoU11Keq/dd+CqNV17nXvkzT2CABgLEa5MLbRTWMXy/7rn0VmqvX86/9GAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAqEsI/x9z0K+n0HYgBQAA"
    "AABJRU5ErkJggg=="
)

TEMPLATE_IMAGES = [
    {
        "cid": "53b4d20d-084d-47f6-9d52-51dcd04340f9",
        "content_type": "image/png",
        "b64": (
            "iVBORw0KGgoAAAANSUhEUgAABLAAAADICAIAAAB3fY8nAAAAtGVYSWZJSSoACAAAAAYAEgEDAAEA"
            "AAABAAAAGgEFAAEAAABWAAAAGwEFAAEAAABeAAAAKAEDAAEAAAACAAAAEwIDAAEAAAABAAAAaYcE"
            "AAEAAABmAAAAAAAAAGAAAAABAAAAYAAAAAEAAAAGAACQBwAEAAAAMDIxMAGRBwAEAAAAAQIDAACg"
            "BwAEAAAAMDEwMAGgAwABAAAA//8AAAKgBAABAAAAsAQAAAOgBAABAAAAyAAAAAAAAAAxa0g3AAAA"
            "CXBIWXMAAA7EAAAOxAGVKw4bAAAFUGlUWHRYTUw6Y29tLmFkb2JlLnhtcAAAAAAAPD94cGFja2V0"
            "IGJlZ2luPSfvu78nIGlkPSdXNU0wTXBDZWhpSHpyZVN6TlRjemtjOWQnPz4KPHg6eG1wbWV0YSB4"
            "bWxuczp4PSdhZG9iZTpuczptZXRhLyc+CjxyZGY6UkRGIHhtbG5zOnJkZj0naHR0cDovL3d3dy53"
            "My5vcmcvMTk5OS8wMi8yMi1yZGYtc3ludGF4LW5zIyc+CgogPHJkZjpEZXNjcmlwdGlvbiByZGY6"
            "YWJvdXQ9JycKICB4bWxuczpBdHRyaWI9J2h0dHA6Ly9ucy5hdHRyaWJ1dGlvbi5jb20vYWRzLzEu"
            "MC8nPgogIDxBdHRyaWI6QWRzPgogICA8cmRmOlNlcT4KICAgIDxyZGY6bGkgcmRmOnBhcnNlVHlw"
            "ZT0nUmVzb3VyY2UnPgogICAgIDxBdHRyaWI6Q3JlYXRlZD4yMDI2LTA2LTE5PC9BdHRyaWI6Q3Jl"
            "YXRlZD4KICAgICA8QXR0cmliOkRhdGE+eyZxdW90O2RvYyZxdW90OzomcXVvdDtEQUhOQmhtZmw4"
            "YyZxdW90OywmcXVvdDt1c2VyJnF1b3Q7OiZxdW90O1VBR0JNVk9yUkRZJnF1b3Q7LCZxdW90O2Jy"
            "YW5kJnF1b3Q7OiZxdW90O0JBR0JNV2ZOaHNRJnF1b3Q7fTwvQXR0cmliOkRhdGE+CiAgICAgPEF0"
            "dHJpYjpFeHRJZD5hNzA5YmM1ZS1jZTg4LTRjMjItYTZlMC1kNzYyY2I3ODdlNTU8L0F0dHJpYjpF"
            "eHRJZD4KICAgICA8QXR0cmliOkZiSWQ+NTI1MjY1OTE0MTc5NTgwPC9BdHRyaWI6RmJJZD4KICAg"
            "ICA8QXR0cmliOlRvdWNoVHlwZT4yPC9BdHRyaWI6VG91Y2hUeXBlPgogICAgPC9yZGY6bGk+CiAg"
            "IDwvcmRmOlNlcT4KICA8L0F0dHJpYjpBZHM+CiA8L3JkZjpEZXNjcmlwdGlvbj4KCiA8cmRmOkRl"
            "c2NyaXB0aW9uIHJkZjphYm91dD0nJwogIHhtbG5zOmRjPSdodHRwOi8vcHVybC5vcmcvZGMvZWxl"
            "bWVudHMvMS4xLyc+CiAgPGRjOnRpdGxlPgogICA8cmRmOkFsdD4KICAgIDxyZGY6bGkgeG1sOmxh"
            "bmc9J3gtZGVmYXVsdCc+WW91ciBwYXJhZ3JhcGggdGV4dCAoMTIwMCB4IDIwMCBweCkgLSAyPC9y"
            "ZGY6bGk+CiAgIDwvcmRmOkFsdD4KICA8L2RjOnRpdGxlPgogPC9yZGY6RGVzY3JpcHRpb24+Cgog"
            "PHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9JycKICB4bWxuczpwZGY9J2h0dHA6Ly9ucy5hZG9i"
            "ZS5jb20vcGRmLzEuMy8nPgogIDxwZGY6QXV0aG9yPkF0bGFzPC9wZGY6QXV0aG9yPgogPC9yZGY6"
            "RGVzY3JpcHRpb24+CgogPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9JycKICB4bWxuczp4bXA9"
            "J2h0dHA6Ly9ucy5hZG9iZS5jb20veGFwLzEuMC8nPgogIDx4bXA6Q3JlYXRvclRvb2w+Q2FudmEg"
            "ZG9jPURBSE5CaG1mbDhjIHVzZXI9VUFHQk1WT3JSRFkgYnJhbmQ9QkFHQk1XZk5oc1E8L3htcDpD"
            "cmVhdG9yVG9vbD4KIDwvcmRmOkRlc2NyaXB0aW9uPgo8L3JkZjpSREY+CjwveDp4bXBtZXRhPgo8"
            "P3hwYWNrZXQgZW5kPSdyJz8+qxmlnwAAIABJREFUeJzsnQWcVNUXx1dAie3uBrFotntnZouw/SNh"
            "ILFs58xsJ6Ei3Yg0SBmkYJISImUAAqKidOdS8z/33vfevKktFtg453M/1zdv3rzaYX3f/Z3fOUbX"
            "btyatuQbV1miUYc3jf0GGvu/18Z/kLH/YOOAIcaBQ42DhsFoE5RoHJwEwyQkxTg0xYSMVJOwNJPQ"
            "NJOwdH5kGIeT2QTm8IxqDlNuzjTlZtEIy9RaYxKuvUbPpzSHGT9rj4gsbtYZpvpWioaeT5nzs54R"
            "DnM2Xc4WDzPNlwaGxqcs+Fk9IvlZPeRsNicLNRrZlnTBspIRobVGQbdX1GhYkVluRRe4IeFnA8OS"
            "zEorbq7WsOZna4nCmlvgR6RSe416Y2Hk8LOeYcPP2kMKs5LMBoa1/vW5/KwxbPlZz5CwhRyt9Qa3"
            "V488fuaGHT9rDxk/cyNXWLaViddXMexl+Ww2OKT611fxKc3hwM0FDtzMjyh+1hn2Gi/z9W6jNRz5"
            "mY5CfhYNWaHWGoco7TV6PsUPJ37WHtFF3GxgOBp8S8+nnPnZ8Ch2jqKzaDhpvtQZGtu78LP2iOFn"
            "9SjhZzKc+YVqjGJXuuCqd0TrXV9Kty+t0XAjc4kbXdAYsfysM1zJXObGzVUMd37mR6m7xssy95gy"
            "7TV6PgWjnJ/1DA9+Vo84mMvIbGC4G37LI244P2sMT342OGLLtdZUsT0ZI/h5hBc/a48e/KwxhgvL"
            "ntpv6R0j+XmkNz/rGXG6K0dUtr2+0bbHKDZrj578bGB49xxZybtaox0/0/E+P9PR4331smi0VS8L"
            "HxHGyHY9uNEWRtwIOoa3jS1vG1PuHVPaNqbEO7rYK7rIO6rQS5bvKc31lCi9yFDYB6Q8G60cP2f9"
            "5Ss3VCrV/fsqDAyMehhG9+m/zn2HjvdO/rBZp75PdXvbJOC9Nn6AhYPb+A9pEzi0TVB8G2DC4IQ2"
            "wRQLQ5IBCzkmJCNdzITaI1xzrhklZpgaYsUHG2YGWTFLc+ZHBD9H6CwbHuZVsGK25lzLoYcS9Qy5"
            "egbqi5Cr56qGpeZc7aGwrAkrWmnOVQ+JeBZYsQpi1Ad+tRi1YkVDQ6o5Pzgr1n6oWdGuclaskh6l"
            "VYJiPs+K1UW+Wg8AQv2saGhEVUGShijRQYMSDVJf7YZBVowq4mfNEa05V2NUAxSLdNmvpsMgK+qh"
            "xxJuZiO6+qxYUhkr6h+lrjVhRTd+ru7QoMcyulwLVqzFqIwSDbJi5SOOnw1SYnnNKLE2Y0R1WbFy"
            "etSDkQZZ0ZufH+poyxNjW73EWAk9VkqSupRokBU1hggIe4ziaXCEmAa9Y8raxpQCCnpHFXkBCkYV"
            "eMnyvCQ5hAalSrfwLDv/5JcTxu/af0yFKIiBUb+DAOE9+s/0+o1bo2evsg8f+kSnfib+7zGpsE3A"
            "YOOAoW0C46lOSJiwTUgyY0JjjgnTuBGWbhzGa4PVwD8t6uN1QiYDZphQhdCELYje1dEJa6wQmjKV"
            "LzzLlCO9TC3pz5QthOvgXzi/XkBEfkEv9dH12XSDbPNwLUkwy0yfBlhNbVBYNo/k50iY5eY62qA5"
            "LxhaaCyIRgRdT1iRUp+mBihAoIW2NlinCiFbL6EbqCVBkUIo6IeReqlPoEEF5T2FmPoEnVANhJF6"
            "+LBGCiF7y4ZuSbmOVwWFZR26s5ZqUp/GQi79IIU9iZoAbekCLwOSjW3VK+kGkhy9fFgpK2ophLk8"
            "9eUKdGfLE52mEmhYIZRqLqgRMZ8uwKyl+PELUmFBQyeskTaoQ30FdJlfiKJvUQ3QnlcFxdqgfZXa"
            "oACHMkp9Mg3eEwRAB5kYAtU0qKkQVosSNdgvmo4oXiHUoTtOGxTWay0wVtT8oF7wE9Y4RemnPic1"
            "DeouVEMnjOHlQW2dkFMInbWoj1sopgvsg8LLyhRCgQZdqE7ooo/6XDiuEwiwRHPmFkTUp6EQuvJ0"
            "56pPHtSjEMbwy2whhi2UiRdE7CdSCGNECyKdsCptsEx3QS/1uVOFkGygVxKM5VfGVqoNxlLeIzvU"
            "owEKKOihow1WQyFU8554eNGVlO6Gq7lOg/R0tMFqE6AW9fG63wgvbh7hHccJiVqqYC1YUUMnFKCu"
            "B9UAyczEQG6B0wa5LUdqfKSnfkTUS31teSWwLdUJ22rqgW3VZMirghwNjtSkweFt48oZDXrHlHjH"
            "8DQoy/ciwiChQW+p0jEwrZ1U/uHMNVeuEWHwHjxoIhFiYNTjMGL/uX+f+7e65eeD/v0LjDr0aeP7"
            "LmAhkQoDBrcJGNImIJ5LHwUmJDohSR+lUmEqlz4axqePcomjtQRCMebpyxGtpRhoJizzQGimDYSa"
            "o3oyoN6UUTM9MqBB2KuFDGiuL2XUXJQvKuI9fRBYlRhoWV1JUEsGrBYcasmAltVJHFWvr1oM1KI7"
            "vZJgNXRCLSDUA4caYiBDPknlcl81ZUC9mJdTR5KgISDkaVBPsqiBUT0Z0I6XAcWYV+fCoAEgzK9a"
            "6KueDKiTMlpIl2sMe7UAQkc9QKiTCFptGVAvEOrwYWWYVwsx0LkaQMhzYHGtZcBqSIJaMqBeSbCk"
            "ciXQtdJkUdHQkgGrkATdH0gSNASEBmVA96rlviplQPEo16sE1oUwqMuEw73EkqB+INQZ1VUCdYFQ"
            "D+bVKDu0OkDoLca8yhNHqycD6gXCthoyoF5JUIyOIzkaVAuDjAaHa9NgdLEX0KCMCIOeEi5N1DNS"
            "buuXLHvngy0/HWaPl0CDj/M5FwMDoxphJCwJ/2hPnr04pGTmk137P9XtLZo+ylkK2wQONQ4cZhyY"
            "YEyZ0CSY6oQCE2r6CdUpo9X2EOrgXx1kiuoBQv0ewkw9KBhe59mhtQRCXSw0r8pDWP+BsFooKBGh"
            "YKRCRxWsVnZordJEKwNCrbxQa938z0pQUKK7jW5eaI7evNC6YEK9HsJcDSDU7yGsEQoa9BDWQv2r"
            "nYfQ/uECYQFV/wrqFgi18kIdDXoIC/WgYFR1sbAmQPigKOiiqxMa8hCKJcHqZYc+MBBWK01Uiwld"
            "9WeBao4YQx5CPVioBYEPkCNaGRbqAcJKULDGWFiuNzu0jjJFtfJChwvJn556s0ANoWBclVio5SHU"
            "1QnrMC9Uw0OooRNWiYI9apcdqgWBhnJEDaWJatIgQcFSIgwSGqRpolKOBr2lCtfQTMeA1JTS+WfO"
            "X1GhMIiB0XDCSOs1Y8Lbd+5OWbLBLmxoM5o+2sbvPWNSaUZgQmYpTDJhfkK1pVDEhDVLHM3kdUIE"
            "wprohAiED6AT1mMgzGUJoroE+MBM2FSAkNMMEQgRCKswClbBhAiECIRNCQh1aXCEVgkZ71hKg7R+"
            "DJ8mmkfTRBXeUqVDYOoz0crZKzbfvXtPeJ7EwMBoEKENhBD3eFfhdzt/7fBKtlb1UVpmhuSOGgcT"
            "qVAzd1S77mglTFidlNGHAoSNNGVUv0uwHgPho00Z1ak12sBSRuuqigymjD44EBbyQIgpow0yZVTH"
            "Q1iDyjFNLWVUp9ao/pRRvfj3wECo10P4+FNG6xwL60HKqE5NUW0a1DYNkjRRUj+mwJM3DXpJFLZ+"
            "yWH9Rmzfe0RFUfAeCoMYGA0q9AChSpQ+evj4yZ6J7wMTtvHlmZCWmTEOZJZCWnpUSycMq5ZOiB5C"
            "9BCihxA9hA0QCNFD2LCBED2E6CFED6E+GhxpmAaHa5sGCQ3S3hKRxDToQUyDSf0zp544eR6eG+/S"
            "PNFH+iSLgYHxwKEfCFkwJrx09UbyiE+e7NK/lc87xrQjhTEpMzPUmJWZIb0oEtuodcIUqhOKRMJK"
            "c0fRQ4geQvQQooewzlNG0UOIKaPoIcSUUUwZrV7KaHW0wVJNGszzlOYwGnQLy3QKTC2a8Nn1m7dU"
            "mCaKgdFgozIgVNH0UZjv3L03YubnbXzfJmVm/DkmNA4cQnXCBJY72katEzIzYY10QvQQoocQPYTo"
            "IUQPIQIheggRCBEIHw0QjtIeahocoY8GuU6DnlJCgx6RpLeEc3C6V2TWJ8s3qrCaKAZGA48qgFAl"
            "6kgx+4sfbEIGt+jSny8zM8Q4AJiQlplhOmFwEi8S6tMJ0UOIHkL0EKKHsMGnjKKHsGGnjKKHED2E"
            "TdtDOEonU3RkZdpgLK8NRvM0KCHaINCgfUBKlxcLNmw+oMJqohgYDT+qBkKV6A8/qzfudpMlNu/c"
            "X9yOgvMT6tEJUytpRIEeQvQQoocQPYQNEAjRQ9iwgRA9hOghbMIeQi1tUJcG9WqDtL2EjKNBD1pQ"
            "1M4/RfLWqAOHT6g40+DDfVTFwMB42FEtIFQRJuRSwzfvPvhs78wnOrJ2FLpMmKSjE6ZVQydEDyF6"
            "CNFDiB5C9BBiyih6CDFlFFNGH1LKqFbuqLqKTNtKaDCqSEyDpKColBQU7TVk7IlTF1SUBh/iIyoG"
            "BsajiuoCIYt79F/+z7//2fWNHKOObwITGmszoVaNGVp0VJ0+ih5C9BCihxA9hOghRCCsD0CIHkIE"
            "wiYFhGIa5OVBcfd5LRqMpr5BcaaoROElUdj4Jb2eMon1nUcaxMBoNFEzIFTx//4PH/8v7J2SJzr2"
            "NeH7E/K9KEiNmTbBScbByeqio1pMGIYeQvQQoocQPYQNNGUUPYQNO2UUPYToIWzCHkKhfgyjwZEG"
            "aLCMo0ENbZDSoG/SW1nTL1y6psKCohgYjStqDIQqngmP/3sGmNCow5smBAj16IQmaiZkOmE6x4To"
            "IUQPIXoI0UPYUIEQPYQNGwjRQ4gewibpIdSxDqq1QZYsKvQb1PQNyjRpMHv65as3VEiDGBiNLmoD"
            "hCqeCY/8fSpwQCHNHR3E5Y5q9KLQbE6ow4ToIUQPIXoI0UP4MFJG0UOIKaPoIcSUUUwZ1fEQjuKE"
            "Qb6QjDYNxoq6zws0GMnRYL+MqRcvX1chDWJgNMaoJRCqCBOS3wiH/vzX7838Jzr2NSV+wvdEOmGC"
            "cVCimglhhOk0J0QPIXoI0UOIHkL0ECIQoocQgRCB8OECoUaTibZCpqg2DZZq1hRVekQqvCkN9kmb"
            "zGWKYkVRDIzGGLUHQhXPhL8dPdH5dWWzTv1M/AfRuqOsOaGQOwpMmCxqWM8BoWkYegjRQ4geQvQQ"
            "NriUUfQQNuyUUfQQooew6XkI2TKjQeYYFLTBEYCCnG+Qo8EiT14bBBr0kpIqMq8kjD9/8aoKtUEM"
            "jMYbDwSEKj53dPdvx57pndmsM9+zXmhYT3JHeSYM5ZmQFpgxpQVm0EOIHkL0EKKHsEEBIXoIGzYQ"
            "oocQPYRNz0OoBkKqClJ5EJiQaYNxYm2wiNQUleV7SnOJNihV2vqnxL43+uSZiyqkQQyMRh0PCoQq"
            "ngm37jnoFZvaokt/kwBWY4YyYRDPhCF8gZkwvsAMyRpN51FQlwkxZRQ9hOghRA8heggxZRQ9hJgy"
            "iimjD5gyOkpcU7QddyaEBr0JDXKFZLy4sqJ5nhKgQaW3VGEXkBLef8Tf/51T8V3HMDAwGmvUARCq"
            "+N8UG7btsw8f+lS3t7i6o2qdMFHEhKlcgZkwUmDGNDwdPYToIUQPIXoI6xYI0UOIQIgeQgRCBELN"
            "ZFFRFZke2g3ogQY9ZYWe0nxPaQ7VBhUOgam+rxYf/vOkCrVBDIwmEHUDhCqeCed+udHY9+1Wvu8Y"
            "q5kwnmPCkCSRTphmSpnQFIZOaRn0EKKHED2E6CGsrymj6CFs2Cmj6CFED2ET8xCObKthHYSjq2nQ"
            "K6bMS60NFnhKcz0lSk+JwiU0o71M/uPPf6iw+zwGRtOIOgNCFV97qnTq8uad+hn7DTT2f884QKs5"
            "IW8mDCE6oWkoz4Rh6sRR9BCihxA9hOghrMdAiB7Chg2E6CFED2GT8RCqK4u25ZJFR3izQjJCA3px"
            "y0FaVhRo0D0i2ykodenaHfBEd/cu0iAGRpOIugTC+xQIK27fGVgwjTasf89YKDoaFK9tJqSJo6ah"
            "6YJOaKKtDWLKKHoI0UOIHkL0EGLKKHoIMWUUU0ZrlzJK5UG+5SDQoDcpKyouJFNCWw4WCk0mvCRy"
            "G7+kj2atVaE2iIHRlKIugVDFJ5qfPn8p9O2SJzr2Mw0YxBUdDRhKOxPyTMhXHBUljqY/jMTRxgqE"
            "6CFEDyECYSVAiB5CBEL0ECIQNnkgVCeLthNZB725loNl1DpYyMqKenDWQaW1T+LQwtnwJEcGOgcx"
            "MJpM1DEQqngz4a5fjnpEJT/ZdQAtOjqYKzCj1ZkwhGNCrgVFWKapRuIoegjRQ4geQvQQ1reUUfQQ"
            "NuyUUfQQooewaXgIxTTIyYO0rOjwtrFlXjGlXjHFzDpIm0zQloMShX1AivTt989dIC0H7yMOYmA0"
            "pah7IFTxTLh43VZTv3db+7xLq8uwAjNCZ8Jk45BkUyISMiakiaPh6VraIHoI0UOIHkL0ENYzIEQP"
            "YcMGQvQQooewaXgIWbIoazzI0SBJFo0Vdx3kCsl4RHKFZJ6JUe79/S8VJotiYDS9eChAqOILzORN"
            "+LRZx77GnJlwsEa3ei5xNMWUmQm5LhQZIpEQU0bRQ4geQvQQoocQU0bRQ4gpo5gyWtOUUaENPZcs"
            "SrVBLeugupCMR2S2Y2DKp2u2q5AGMTCaZDwsIGTJBpev3eiV9IFRh77ETMi6UASoK44aqxNHU2mt"
            "0bpPHG2sQIgeQvQQIhBWAoToIUQgRA8hAmETBkJGg6M0k0V5GlQni+Z5EHmQdB208knM+WipClsO"
            "YmA01XhYQKjif638fvTfZ3pntOjCzITqxFGT4ATKhCmmooqjphqJo+ghRA8hegjRQ1jfUkbRQ9iw"
            "U0bRQ4gewsbuIdROFvXmCskwICzxVlsHSbIo0KCtf3LvoWOvXr+pQusgBkZTjYcIhCreTPjZNzvM"
            "AgZyZkI+cRSA0ESUOCpUl6GDUwjRQ4geQvQQooewngEheggbNhCihxA9hI3dQziyLZcsKnIPxpGu"
            "g6SWDEkW1eg66ByS0aFH7u9H/1OhPIiB0YTj4QKhiv/9Iv9ooVGHvqQzIScSxhOdkCWOBqfwnQlT"
            "qUiYwdoSoocQPYToIUQPYe1SRtFDiCmj6CHElNEmmTLKoSDrM8HJgzRZ1IsVktFMFvWMlNv5Jy9e"
            "9aMKrYMYGE07HjoQsvSD0+cvB/YveqJTP53EUa7iqLi6jKl2dRkEQvQQoocQPYToIUQgRA8hAiEC"
            "YbWAUJQsKjQeLNVIFpWQZFFr36Sh+bNVmCmKgdHk46EDoYoXCb/astcqaFArn3e4xNGAoSaBw0yC"
            "EoxDaFtCTiFM5VJGucRRTBlFDyF6CNFDWK9SRtFD2LBTRtFDiB7CxushHMnvRw2EbePKaePBEi/S"
            "Z4KrLOohySHJosHpXV8sPH7irIqvDI+BgdFk41EAoYr/45Ni7CKjDm+aBLDE0cEmAfHGgbxIGJxs"
            "GkKBMJR3EoZnmDywSNh0gBA9hOghRCBEDyECIXoI0UPYhD2E6j4TYnmQtqEHICzypPKgB21DT5JF"
            "A5KXrNmhwmRRDAyMRwyEJ89e7N4nr3nn/qYscVRcXSY4mTkJ+dIyzEnIuK72TNhYU0bRQ4geQkwZ"
            "rSRlFD2EmDKKHkJMGW1aKaM9ePdgHJMHh7dljQdjSLKoV3SxJ60lw/WZkJBk0YE5H6swWRQDA4PG"
            "IwJCFfkTFPmls2z99jY+77TxHQhAaEKchEONg4aZBCXSxFG+J6FIJEQPIXoI0UOIHsKaAiF6CBEI"
            "0UOIQNiUgFCzsmgcqyXDyYOkDT2RB1ll0RwPicIlNPPZGCVWFsXAwBDi0QEh+5Vz9969t3MnG3V4"
            "05SvLmMiVJehfepNOZFQ3IICU0bRQ4geQkwZrScpo+ghbNgpo+ghRA9hY/QQCkDIWwdjmTzIkkWL"
            "CQ3K8rjGg/A/RJ/E8XPWq/i/1GNgYGA8OiBU8X+I+uWPvz2jU57q+hatOEqqyzCREJjQJIQljjIz"
            "YbqJuidhLbGw6QAhegjRQ4hAiB5CBEL0EKKHsOl5CLXkweEarSZiimllUVZLRuklVdgHpMYMHH39"
            "ZoWK5Is+ymdADAyM+huPFAhVPBOOmPlFs46sLeFgKhLGEybknYSmtAWFWiREDyF6CNFDiB5C9BBi"
            "yih6CDFlFFNGtbcZKQBhOwEIY8vbxjIgZMVFC0gtGQmrJZNtH5C89oe9Kqwlg4GBIYpHDYRCW0K/"
            "N/ObcdVlBqudhMHMSZgichJm8CIhAiF6CNFDiECIHkIEQvQQIhAiEBqQB3U70XPyYK6HROElVdj4"
            "Jb+ZPkWF2iAGBoZmPGogVPEi4cLVm1t1f7uN30DalnCIcWC8CceEySYhyXz/iTRCg6GYMooeQvQQ"
            "YspoPUkZRQ9hw04ZRQ8heggbl4dQJ1+UuAe5yqJerBM9qyUTqXCPyHYNSd+6+7AK3YMYGBia8RiA"
            "kP1d6uat272TPjTq2NckYHAbJhKSFhSJJkwkDE3RKDda28TRpgOE6CFEDyECIXoIEQjRQ4gewqbk"
            "IazEPVhC5cEi2ngwl7kHrX0S4wtnq7DVBAYGhk48BiBU8SLhV1v2mgcMbO0zkJUb1RQJqZMw5EGd"
            "hI01ZRQ9hOghxJTRSlJG0UOIKaPoIcSU0SaQMirQoCF5sMBTms9aTbiFZ3lFZu357bgKW01gYGDo"
            "xOMBQqEFRT/lJKMOfU0CB3PlRgMBCBNouVHapF7bSYhAiB5C9BAiEKKHEIEQPYQIhAiEvDzId6LX"
            "cg96EvdgPutE7yVVWPkkZo5cDI9e91AexMDA0InHA4Qq/g9UW34+aBc2tJXPO6xPvYm2SJgqKjda"
            "Gydh00kZRQ8heggxZRQ9hJgyih5C9BA2GQ+hJhBSeZB1oveinegFedBTonAJy3gmWnnoGO1Ej0CI"
            "gYGhE48NCFW8mXBw8QyjF4iT0DhgsEnAUBOuJ2GSHpEQgRA9hOghRCBEDyECIXoI0UPY1D2EQIMj"
            "RPKgRnFRLw33ICcPFoxbocJkUQwMDAPxOIGQ/WLa9NNvVsGDW2uJhLQnoTFzEobVXiRsrCmj6CFE"
            "DyGmjFaSMooeQkwZRQ8hpow26pRRtXWQyoMjvEnjQaH3IJMH85g86BqW2T5KcejPkyqUBzEwMAzE"
            "41cI796993rGWKMOfU0DBrOehCKRMJkTCcNq6SRsrECIHkL0ECIQVgKE6CFEIEQPIQJhIwfCnvp7"
            "D3rFFHtGk96DQnFRq+6JaeULVCgPYmBgGI7HCYQq/tfTyu9/MvF9p43vQF2RkAJhSq2b1DedlFH0"
            "EKKHEFNG0UOIKaPoIUQPYRPwEKppkJMHCQ1S9yCRB4uJPCjL86TlZNwistzDM3YdOKZCIMTAwDAc"
            "jxkIWTOc6zduRQ0Z8UTHfqaaTkLjkCSSNRqSqiESIhCihxA9hAiE6CFEIEQPIXoIm6KHcCQPhCN4"
            "ICTyYNvYcm8qDzL3oKc0j+s96Jv0tnyGCnsPYmBgVBqPGQhV/J+s5ny58amuA4z93jP2H0xFQqH/"
            "RLJmaZl00/B0TBlFDyF6CDFltJKUUfQQYsooeggxZbSRpoxqyoNcfVGhnEwR7TZB5UGJ0iNSbheQ"
            "vG7jPhVt9PW4H/cwMDDqbzx+IGR/tTpz4bJPn7zmnfuTcqP+NGs0kGSNmmj0n0jnRcKmDoToIUQP"
            "IQJhJUCIHkIEQvQQIhA2TiDUU05muHdcmdBtggAhlQc9JQr7gFTJW+/fqrij4qs2YGBgYOiNxw+E"
            "Kl4kHD1ndbNOfU389YmEIbVsSNh0UkbRQ4geQkwZRQ8hpoyihxA9hI3aQziyinIy4mb0EoVl94Qp"
            "C79V0ep9j/tBDwMDo15H/QBC+perYydOt++V8WTXt0wDGBDGGwfxIqGe0jIIhOghRA8hAiF6CBEI"
            "0UOIHsKm4yEUtR+MU+eLesWU8eVkaL4oa0Yfmvl8bM7f/51XoYEQAwOjqqgXQKjif1slj5hD+08M"
            "MibVZYYShTCIKITGxEmoJRI26ZRR9BCihxBTRitJGUUPIaaMoocQU0YbY8qoOF90uDeM2OG8e7CE"
            "yYMsX9SbNqNPKpmnwt6DGBgY1Yj6AoQsa/TbHQfMA94j/ScCtLJGUzRLy1Q3a7SxAiF6CNFDiEBY"
            "CRCihxCBED2ECISNDgg13YM9mHtQKCcjyhcl5WSy7QNSvtq0X4XlZDAwMKoR9QUIuf4TN2/Jhgx/"
            "ohPpP2FCSsvwIiFpUi90qK+BSNh0UkbRQ4geQkwZRQ8hpoyihxA9hI3cQ8jyRcm5DefzRflyMqQZ"
            "fZ4HzRd1DEoLeXP49ZsVKiwng4GBUY2oL0Co4kXCiYu+at6pn4n/IGP/wcYAhEQkTDQJASBMpgoh"
            "Y8LqdqhvOkCIHkL0ECIQoocQgRA9hOghbJwewh6ifNE4vv2gOl+UtR8sYN0mvKRKy+4J5ZO/VKE8"
            "iIGBUb2od0D4yx//uMmSWnZ724R1qBfVGuVSRmuSNdpYU0bRQ4geQkwZrSRlFD2EmDKKHkJMGW1c"
            "KaM69UVjNeuLAg0SA2GOh0ThHpHlGpr+04E/hScrDAwMjMqjHgEhy2q4e+/ey6kfGXXsaxrIbITA"
            "hBwQGjMmDEutfq3RxgqE6CFEDyECYSVAiB5CBEL0ECIQNi4g1MoXHaGbL+rJ5Ysq7QNTZe98cOfO"
            "XfJk9Zif7DAwMBpG1CMgVGlkjfY3ZQ0JhazR4CS1SKgGwiqYsOmkjKKHED2EmDKKHkJMGUUPIXoI"
            "G6eHsIemQhjHDIR8P3pRfVEvqdKi27DhU1eqsP0gBgZGtaM+AuG+Q385SxJbdn+HzxqNV3eo18ga"
            "TW+UQGgpkdtIFdZSYB45egjRQ4gewnoGhOghbNhAiB5C9BA2PA+hUE6Go0FOHtSpL5oDQOgeke0c"
            "krbt58MqzBfFwMCodtQvIGRZo7fv3O2dPJrLGuVqjbIO9aw9fQ2aTzSslFFAQRgtAlKNuibAeDIg"
            "la1BDyF6CNFDWKOUUfQQYsooeggxZbQRpYzy8mDcCO1+9Dr1RR2CUsP7j7hVcVuF9UUxMDCqHfUL"
            "CFV8B9UPZq9s1qmfqf+i2itoAAAgAElEQVR7RCRkNkLaoZ5rPiEohFXZCBsKEALFWUkVLQPTWvin"
            "xKZMGjlrLYyY5EnN/VNaBaVbSwluiUVC9BCihxCBsBIgRA8hAiF6CBEIGxEQqg2E3kK+aBwrJ1NC"
            "6otG0XzRSKU3yRdNyB2zTIX1RTEwMGoS9Q8IaYbDD7t+NfMfSDrU+4s71LNuhCkUCKvVfKL+p4wC"
            "s1lL5cahGUY+Sd0GjFr61Y67d26zW3HndsWn67Z37jsS3jIJzQQsJKmh6CFEDyGmjKKHEFNG0UOI"
            "HsKm4yHswbsHScMJRoM0XzSaNZzg+9FHsn70yau/26NCAyEGBkZNot4BIetQf/Lcxa5v5DTvPIB0"
            "qCdAGG+g+UQVNsL6DYRZADZwIOA9j95F4xZ8ffXKVbj2e3duL12/Y8XXu1R378DLy5cufzj3K7ee"
            "BUY+yfAp+IgFegjRQ4hAiB5CBEL0EKKHsCl4CLUMhGIgJAbCIs8o0nDCQ0IaTriEZrzQI/fM+cvC"
            "0xQGBgZGdaLeAaGKz3p/J2+qUYe+IiBktUaTTbRthA0xZTTLUkLgrZlPsrVUkT1uxYn/zpBrvnd3"
            "y0+/RydPahGQ+lRg2mvymT//coRVjf77xOm00UsB8Jr7pcBngYjMtZNFMWUUPYSYMooeQkwZRQ8h"
            "pow2rpRRHgi9e/D96AkNihpORBWyhhNeUqW1b9Lb2TNU6B7EwMCoYdRHIGRZo9OWftOic38TdfMJ"
            "ZiNMqpGNsJ4BIXlJQItWjmkVlN4v/5MDB/9kV338n5ODSue1Cclo4Z9iJSHEBewHn0r/aOk//56m"
            "m9z/+Zejr2bPAFaEYSVVEKpEDyF6CBEI0UOIQIgeQgTCRguE6m4TohKjgoGQ5YuyhhMKi+4J0xZ/"
            "q8J8UQwMjBpG/QXCH/f9YRU8uLXPQHXzCVJoNMlY3Y2wahthfUoZzQI8s5LKWwWnN/dLjhw2/rsf"
            "D6jukb6x585dyJ/ypVNcXjO/ZMA5oawolQFJQqlbr8KP5m24cpkkgdy7c3vNxj1B733UzDcZ6JHV"
            "mzEIfvUYCNFDiB7Chpkyih7Chp0yih5C9BA2MA+huqKMur6o0HCCAKGUNZxQuIdnuYam79x/TIUN"
            "JzAwMGoY9REIWeL76fOXffrkNe/cX8tGaKxtI6znQMgJg4CCJmGZRt2TXugzfP6qrbcrbsFvbNXt"
            "W1OWfOfWqwBQEI5oJVVY6BQgBeQzDs008knu0m/k0vU77lRUwM25eeP6tKU/PP1qCayHD7J6M3rw"
            "D4EQPYQIhOghRCBEDyF6CBuoh1BsIOSAkOSLescIBsJCD1keAKGnROEUnN71pcIr126qHoKBEHZ4"
            "r36H+JLv06jbO4CB0bijPgKhik9/76+cZNThTTOxjTA4QacbYXo9Txm1lMjNqNDn3CNv1CfrLl64"
            "RC/wngrG3YqFa37s1HdkC/9U49AMa4m6jqhGf8JIOSBfq6B02CwuZfLW3QdV94m0eObMubxJX9hF"
            "5TTzTYZtGE82FCBEDyF6CB9eyih6CDFlFD2EmDLaGFJGRUBIy8mM8I6lHQhj1R0IiUIoUXpJFNZ+"
            "Sf2zpj3eh7d6FYiFGBjVj3oKhKwb4Zi5a1p07mfiP4jvRjhMw0YYxiuEhkXCxwuEDKua+yXDcuKo"
            "xX/+fZJd3ZE//00bvXT5hp0sZfTG9esTF3/r9VIRQCOnE+rFQlJLRk7QMSRzcNl82Anb229//P12"
            "0dzWIRlPBqRSUhJxIAIhegibJBCihxCBED2ECISNCwgFAyFfUYYZCIUOhDKlWdf4D2euVT0cA+G1"
            "a9cO/v57/RyHDh787ddfL1y4wE716tWr//zzz7///nvt+vU6vw8YGI016isQ0vT373b8YuL7rrob"
            "YZAaCE2qV1fmMaaMWkqynwpMA0h7OXv6T/uPkARRlerC+YulM1Y7xeYZdU98KjC1d8bU7XsPE6lQ"
            "pTp56pxiwmc2UkUzX+IktOKdhOaazeitqIr4hG+yY2xu6fRVF85dpPfr7g87fpUlTgBcbB2crmbC"
            "egyE6CFED2HDTBlFD2HDThlFDyF6CBuOh3CkqAPhSFGJUd5AGF3oKSsgHQglpAOhY2DKV5v2q+q6"
            "Jf09urctW7Z4e3g87eVdD8ez7Z52tLWbM3s2nOeJEycWL1r03bfffvv1N0sWLz5//rwKO3BgYFQj"
            "6isQ0n+9h47/1zYu7alub6vryvDt6alImKYWCauvED5kIOQXslsEpAQMHL120557d0g7wZs3rs/6"
            "bNNzb5SRLvNhmbQYTHZz/xRYHjJ84dHj/7ELP3DoeN/cT1oFpzO5jxWYEQOhBdfLXmEalmXkk/zc"
            "G+WzP9986+YNFa038+na7R36DIfPWtZ7hRA9hOghbJhAiB7Chg2E6CFED2GD8hCKW9LzFWViRB0I"
            "pbQDYaTCJTSznUz+z8m65x8GhJs3bXJzcvZ0dauHo62Hp42F5SezZsF5rvryy317944ZPXrN6tXL"
            "li7dumWLcAkYGBiVRD0FQvbr7Mq1G2HvlDbr2M80cLAxyRrVak/PRMJ6lzJqKcl+MjA1a8zye7dv"
            "kYu5e+erzftCBo9p5ptM5DupnM8IzWIyIJP7SqatPM/JfXe+3XYgMn58c78U2J4WjNHGQsFY2Do4"
            "AzaTDBv/zbYDtykW3rx2bWDJvJaB6VRLrC4NNuKUUfQQNsGUUfQQYsooeggxZbRRpIyqe054awCh"
            "0IGQAKGnROEQlBrRf+Tt23dUdd2EkAPCzZs9XFy93T3q4Xjay9vOynrunDlwnl9v2HDs6NGs9PRx"
            "Y8YeOXJkx4/b6/JeYGA03qinQKjimfCt3MlGL7xpFjiYEwlJXZlEzboy9QsIAfbahKQ//WrxpctX"
            "gQUvnLvwmmKmUef4VkHpNgbMgVZShQmpI5r07Otlsz7beJNy3a2bN+d+ueWF/w0XFEULzfRR9tJa"
            "orCNymnWNeEpn6SFa7er7gCC3v37v3MO0blw8pYIhOghbJJAiB5CBEL0ECIQNnggVDecGMkZCFlF"
            "GaElvagDoaVPQkLxXNVDSI8UgNDd2cXLzb1GgwFbTT9V09HO08vW0oophF+t+2rzps1btmw5fOgQ"
            "wOHGjRtVqBBiYFQj6i8QsqzR4TM+b96pn2nAIBN/oa5MkqiuTL3zEDJ5MCJ+/N1bN1W3b925dXP1"
            "D3uiEic+STvRi+RBjQErraXy1sHpzfySQwaPWbdp7/275O98Fy9cGj5ztXNc3hO+yeaR2Vxfikh1"
            "4iiQ5BM+yYEDP1q7cS9RCOGm3am4fOWK54tF8FZ9BkL0EKKHsGGmjKKHsGGnjKKHED2EDcZDqJYH"
            "dQyEfEt6VlEGgNCsa/zEeRvgseFOXVeUqR0QAgcCp7GUToaFbT08YQ0buj5A4S3YDEaNMJIB4exP"
            "PoHzvHz58mfLV3z7zTfffv3NimXLKminLgwMjCqjHgMhrSuzbMP2ll0HGPu/Z8wVGk2gdWWSOSAM"
            "FRcarUdAGD503L2KW/crbjJ3352KioWrt73QR5D75LpMyNcRVQA6tgxKe00+Y/cvR1kpmj//+i9+"
            "+EL4YAv/FCuJnG3GSsu49SwYO//ra1evsqOs3bjn1vVr169e9eiNQIgeQgRC9BAiEKKHED2EDdZD"
            "KDIQeqtLjJZ5RZOKMp4UCGlFGQWtKJO6fvMBVV1XlFHVCgiB6NycnG3MLRgZujo5O9ra2VlZwxor"
            "M3MLE1OzNsZsmLZqY9qqtVlrY3NjY0sTMyszC0A7eysbBxtbVwenGgHhnE9mk8u/e/d2RcWpU6dO"
            "nzp9987dur0VGBiNOOo7EO7Y/4dl0KA2Pu+acN0IWcpokiYQGhQJH33KKC0uSoBQdff2pQuXYlIm"
            "r9m4h7WXuHjh0ohZa5yY3BehUUdUCwstSbOKFEC+lA+W/PXPKXI77t/bsedw7/SpRGkMTn8qIM04"
            "NGPYiEXcu6r7P+0/Ikuc4N6rsOL69ZvXGgAQPmoPYaRoru8po+ghRA8hpoyih7DepozqH5gy+lBT"
            "Rr3JVVADIQBhbAkFQr6ijEThGpbpJcn64zh5Hrj3uFNGgQCd7Ow7PPfcRx+OPvj77/v27dv+44/f"
            "ffvtqi9Xfr7isxXLly9asGDe3Lkw5s+dCxQ3c/r0aVOmjBszprykVJGdnZSQ8Hb/Aa++9HKgr5+n"
            "q1tNgfDOnTtC0iwWF8XAqH7UYyCk/5KP/nPqmV4ZT3V92xRokGtPzxUaNQ1J5QuN1kQhfHRAWHHt"
            "8hV42dw/pU/OrH2//wm/neCKjlG5D3AO1ltSuU8vFgLzmJN29smuPQs+mPPV5YuX4bP3795Z+d3u"
            "bgNGRcaPp+3pya/pU6fPZY1ZbhGRbdQ98fn/Db994/qNhqAQPiIgjFQABKKHsKkB4SP0EBY4yviV"
            "wkKt+NAgEOoOAQURCNFD+OiBMHa49uDwrxyBsK6BUEshLOeAMIYHQhkHhE7B6Z1751+9flP1uD2E"
            "QINuTs7/e/W1bVu3njt37u+//vrj8OHff/vtl19+2btnz66du7Zt3bZ1y5YtmzZ//+13G75av3b1"
            "mlUrV33x2WfLlixdOH/BJ7NmTZk0eczo0aM/+HDo4MEeLq61AEIVtqTHwKh51F8gFBUaLWlGbIS8"
            "QliTQqOPMWUUgPDq5cuuPfNNwjKb+yVbSeTZ45b/d/IMvbZ7P+451DNtSgt/kh1qJQWU0sOE5rxL"
            "0Mg3qUu/kfNXbbtyiWDhtatXb90gmagVt25+vGJju1dKgBuBx0zDsp59vayhAOEj9hBaUSy0qguF"
            "ED2ETTtllOmBNGVUJsAhN+pQJyQvacqoY1QRXWY6IWNCHRpsqCmjmlmj6CGsZymj2h5Cgn96tUEB"
            "BcvpQA/hQ0oZFfWciGElRvNYiVG7gJSeQ8awvKo6x6BqAqFQ8xOWu3fu4tOl6zPe7Z72bsuchECJ"
            "Lg6OTnb2jrZ29tY2dpZWthaW1jSDFIalqRkMtmxlamZtZg7D0ca2FimjDAgxMDBqGvUXCFV86eQ+"
            "2ROMXuhDCo36M4WQZo2GVKvQ6GMHQue4POPQDNI2MDzLyCfJ88XCCYu+YZa/u7dvr9iwq9uAUbDe"
            "OJQYCy0MSIU2UcpmXRNa+6Us27ATPnan4qbq/p1d+/6QDJvQ3DelTUiGtYx0LIQFAMKKOgTCCP5l"
            "RJVwKK9vQMjNZOdMJ6RDorSVVsdGWN+AsKF7CMmCvRSYMM+eKYfcckMEQk79E9bATrREQofaioSO"
            "InlQr0LoWLlO2CCAUGuo+bBeeQhLeCYsacpAqPEyVrxc5hlb7sntp4wMnaxR9BA+KBAK+aJiIIwr"
            "94ot5XpOyAo8hRKj3RMSi+epHk6SZHWAEDjQw8UVNoAZhqujk6uDEyywnE+hnMzTXt7tvduy8Uzb"
            "dtxo9zQd7dRr6ICNhQIzYtpEIMTAeBhRz4GQ/GqTj1nYrGNfUmhUo/NEMgXCtBoD4SNMGeWBMJNV"
            "FrWSArOlP+Gb5Pv2B198+9PdO7fh6q5evjJ2/gb3XgWAhWakM6FCSyFsHUI6DfZKn/rzgSOq+3fh"
            "ntwlhWruv1s816hDvF2UkiWd1jEQRuhXCC3U2yhEcw2w8JGljPI0qLCVKOxlOTBbRmRbhGW1CUoz"
            "C80idUcjKyk9ih7COk0ZFaOgTISC0rphwkfvIVS/pEd3joaX+faSPKeoArIsyxfEQx2foUFQ1NAG"
            "ZUR+dI4qdIkpdJDC3vJhwTmqwElWoHYVRmmaCet/yigFP7JNDE+J0XyOqJoJH6OHsESXEht1ymiZ"
            "mvTEI06T/dQLZWRQ/PPqAXOpW3SJZ1yZF7+SbqBWCzFl9EFTRsVAqK4oo9FzgpUY9ZYpTbvEj5q+"
            "WvUQSoyqqgGErIQMMF6HZ59jIMeKhQrCoIONrb2VNTCbtbkFEQNNTJkkyIYFvKRrmEJobWZuY2EJ"
            "G9tZWdtbk+oyjrZ2TnQAZyIQYmA8pKjXQMhshBMWrmvRub+JPwVCdeeJ5Op0nnjEQGguBsKKW1cu"
            "cQqh0InegrSOkD8VmPZkQOpLmdN27TvM6oj+8+/ptNFLLSKygf0IU5E6onLqIUxq/1rpvJVb7twm"
            "9Hjt6tUP56zf+9sxWB5SvrC5fyptW0+6VtQtEHIvI+jgFrItyaBqIV1g7+qDwMqY8BEBoVoYlLcK"
            "SDXqMqylX7JVRJZTtDJs8NgXXi8zC83kHIb1EQgbkYeQ7o2liap3K1Xjor00XwMRG4KHkHJgIX84"
            "QmsmQVlWYXKX6HzzkGyz4GyAQ7a+RiVn1MIgocF8YEvbCKVJYJZzVJ6DJNckMNNekuvEmFCNhQ0J"
            "CPmFIjbI/mNEWMgtPy4PoS4Nloh0wkYGhBwKchvE8e/GaWKhhk5IkS+mzD2mFFDQJizXSZLftkep"
            "Q0SeXXieZ2wprPdUMyECYZ0CYZzQc2I413OCAGGhZ1QBKTEaqfSSyC27J8z7fIvqcQAhgJ+jjW2w"
            "f8CRP46cO3vur7/+OnTo0P79+3bt3LVl8+avv/561apVS5csXbRg4bw5c2ZOnzFp/ITx48ZNHDd+"
            "3JgxY0Z/9MH7748oLx85fPiIsvLyktKC3DylXJ6RmpY0LGHwwPfe6tf/jVdfe6lX754xsXFR0b7d"
            "uut1FSIQYmA8eNRvIKQJ8Su+2Uk6T/i9Z0w8hDqdJ2oKhA8hZdSMkh4gGezfWipvGZxGgFClunr1"
            "KqcQqkU/Bo1cHVH4YOKoxce5SqH3fv7l6KvZM54MTGPECOeQPW75mTPnyXt3bq/8dne3Ae8/4Zu8"
            "7edDsGZQ2XxY1gHCcgaEnrUFQjEKWjIIjCSzVWS2NRsSOTci5VYCJerxFurHwkeXMhpJztA8LDMu"
            "ZdLwj9ctXLtzx4E/T54lJsxhIxc390myk9YSCNFDWJOUUW5jeymvE4oW2E54LGwoKaME8GiaKEU+"
            "aa5FSFb/3Dnf7zr869GTazf/Epc02Swoy0mbCatdiZQmoMLHrcMVz7xUOnXZ5v2H//35t79HfrLB"
            "NTrfLjKHZ8LCBuEh5BZiyJ5dKhsCEz4WD6F4A2640aFeE9vYUkbJlnFkDx6i4S6WDcUeQkqDHkCD"
            "MSX24Xnv5i/Y+NOR346d+mrLby+mzLALz2VMSFJJ60whbPIeQm2FUNSEMJo1IaRAKFGSnhNBKV9v"
            "+UX1EHpOqCoFQm93D1jZpUPHX3/5VVWT/u/3798HcrtxA55Wrl68cPH06VP//P330SNHfvvtt317"
            "9+3asRNg8ttvvlmzes3qlSs/W7Hi00WLlyxerJQrEAjrT9wXxeM+lyqiQZykEDW6sbDNPRoPfoH1"
            "HAjJL5etew+Z+Q9s4/su6U1PUkYTyAipVueJRwCErIEE6SlPe8cDyLUISHGIyR0zf8OpU+fcehUY"
            "h2RomgOzeHMgpwG69sh/f/ZXl2nBGAC/tRv3+L/zoSxx4k8HjjB/+P7f/3xNPvMpCoq2MuVPpD+h"
            "Kn7EQn1AKBSVKay1QsjEQEA+gEBzuD+h6aYhaW2CUlr6Jz/pl9jCN6GFb+JT/oktA1JMQtLMwzKA"
            "Em3oxiI4NCgVPhogtIog2qB1RLZJUNraLerWTOxPDEPLFzT3SbCVyik0NgggbIAeQmk+HbCQa08O"
            "lGsTqSQjQsEWbCU5DlJeQnyw3NFH7CEkxWOk+XDygG2mgRnpHy5Xif5/c/vO3deyZpgGZThHMSAs"
            "EGEeGS7RhS4xRXSNPp2Q0iCAX9ueRfsOnVDRRAn2vV275Re7CIUjHJfbVQMAQjoXuUQR6oPTdqD3"
            "jQ24h7CGA0J1yZnH6CGkm0WT4RZTzIDQAzgntlREiQ0eCOkC91kxCnrElXn3KPcU8j+1PYSEBj3j"
            "Sm1CldljvtD6wr8pn20blgvvsn16GmhVjx7CWgKhoBDGDhdKjFIgzPOQ5nhEKlzDszwjMn85fEL4"
            "f9zDeBIzBISujk4RoWHZGZmDBg4c9O67b/cfMKBvvzff+N/rr7z6Uq/ePWLjoqTSyNDwsKDgIH9/"
            "v+4+3Tt36dqxU8fnnn/+mWefbfc08xMyuyDbp5uTMxmOZBAvIrUjkoI09g5wAnpthFUC4f179+/e"
            "uat3VOcx+u5dA5/Vd7eJr6dWB7pn4Cjss4bOAaKaJGBwD/x+4ATuV6NpCYP5e5paNHxJxA0/1LfC"
            "8GlX58Qq24OBIf6rBHeq/BpyjeRfyH3hHPSccE2Oxc6wFp8SX77W2Wq9pXVj7/DBLkr3/Kv66VUW"
            "9RwIybX9euQf96jklt3f0WxFWK3e9A87ZRSgjlQB7Z749KslQIC79h+Zv2qb/7sfPtEt0ajLsC79"
            "RzlE55iGZZrrA0LeJSg3DssELOz45ohP122/c7sCLrni5o2KW7dg4ezZCzkTP7eWKJr7kZb0sHGr"
            "oPRl63fCt0KWMAH4EFZqAOEbDwqEjAYtwjOb+ya2Cki2CEu3k2a6xCg6vFESPPCDHskT31BMfylj"
            "SkzieHjpHpdjE5nxpG/iE93jgRsJFqo1w8ecMkoFzCyTwOTVm/bDnbwFjy3wvwT6K2xo+fzm3YbZ"
            "UJGT9aWoZymjjcdDaCfJpQO+J3Ln6FznKBg5MFyigRIJFtpTXHzAGjOP2EPIrIOOcHWRSiep8vdj"
            "J1UkUwu+X/crbpNf5Ws3/2IVkukozWUqImcshGuU5NpE5JgFZ5sGZzswn6FMI2WUJYu6RBe09k1N"
            "fX+pij5tw69B2DX7ZfhS6lTzoEzYmx6RsF6mjHLCYBQxQLrHFnrGFZHBLRS7xxTBeoKLUUUuGumj"
            "Dw8IxamhWvmixa7RcEol7rHFzpTM7SLzLEKUNmG5btHkLTeBCRtsyqigDRImpIAHxAuX5iIrso8s"
            "sArNcZQUejIxMK5Mw1JIPlLiFlXoEVN48M9Twhf+Nv3Cr9/6m2NErkdMMeytToEQU0bVTQg1FULW"
            "cyKP9ZxwCc14Jlp59gIpVveIi8oAxQHOebi42llZU6efvbMdHfYwHIDiXB0c3RydGOO5w3B2EQYr"
            "P8OGp6sbG1oFS8WDVZdBD+HDDqY46X/rnsZbcKuvQFy+fJt6mljcu1sHUlWdhPhU4Qzr8xeDca/w"
            "sqKi4tJFEjdv3hRWGiL/v47/9c3XX3+9fsOJEyce8DTqNxDSi//rv7MdXpY/2fUtPa0Ia6EQ1hEQ"
            "AoCZEX0v0TE2t2jaSiA3Ff+7+Pr1a1OXfPfM66VGneNNQjOspHJNDtQAQjOgr0jiGATSaxGQEp08"
            "cctPv6vu3VbdqZi/amv7V0uNfJLhNKykCguyZTZgnmvP/M79RpqGZzEUrEOFkOWIWsAFhqXLx3+2"
            "ZvP+bfuOHj1x9uLVG7fgUVfzdwS8vHD5+m/H/lv01c7++bMco+RPdIsHkuSlwmy9uaNWdI2WTlgN"
            "JjQAhOKdiDZg6axW4ZnGAUlwFSothbBsbvNuQ20l2dYRcgNZo5UDIdmGop2SAqHygYBQojl4SnxU"
            "HsJcNrNhyy9UhoXVB0IpoUFHWW4r/9T00ctPnbt8+K/TR/45+8ffZ47/d37nL8fb9i60Cs/mmbAe"
            "eAhl+oZhILQIyQx6+4NbFer/zbDfACfPXvaMy7ONUDjICBMCBJoEZgI9ukbndXit/NXMmQPy5zlH"
            "E62M1p4RAyHZs3NUvol/2rxVP5Lv7V3hT5tkoWTa6pbdk11JDZsC7aIyHBYWcnAoLNcYCKscNdMG"
            "6VxoE6ac/cX2cxevHTtx7q//zsN8/tK16cu3WocpXKMLmYTIewg1exWyNdFa9WaKaw6EevJCRYMh"
            "X7FtRK5ViMJFVuDdo9iv/+h38he8kv6xo7TAjWwgpJJWCYT8gthzKNYPOT4spRBYSvmt9IGBsNzQ"
            "ux4iPVCgQUBf27A8+/A8ALlnepdLBk2ML1sSMWiiNhNSeZBaBEsdwnMjBo7V/MKT+dTZy8/0KnWR"
            "FghA6KHRk5BzFZLCpLUEwuEUBbnZS2fZq8dw/bxnCAjjdEZlOuFjB8IRbVkTQjUQFombEDoGpwW8"
            "Xnrz1m3hJ1LHT2IGgBDwzNHG1tLUzNrcgtSAoWVghMHqwXAlYezsnewcABGdHRxdqOjHyYCaiKhF"
            "ibq4yEGjTrf6SoCQ/U4+fOjQrI9nzZ83f97cecKYO2fu4kWLxc/cegNYYvmyZXNmzxF/Fnb18cyP"
            "9+/brxJBOFs4ePCg1rHgQAvnL7hypTJih5u86suVsz+ZrXWS5Fhz5508efLLL77Q/+68+WdOn6lk"
            "zyzgMhcvWgRnovVx8VHWrF596uQpth9dJhSSEs+dOwcb5yqUsbKorp06d+3YKVoiVWbL582Ze+7s"
            "OZUIb2DLBbBjnWPBaXy6ePEN1j6t4tbST5do3V5hs2NHjxm6cD2XQH8oP/7I/tdJTuDqlatw+PTU"
            "1GiprGdsbI5cvnXrVsCtBfMXwBFhvniBe3Rnl3b69OkqjyIec2bPPnv27LFjx2Z9/HF1zlC4rkUL"
            "F7HLV4kI/Pq1a8uWLsmRK2SRks4vdOz0wgvBAYGpScnwxT5/jnjHrl69OnP6jMkTJ40d/dHOHTtg"
            "zS8HDixfugx+sktJG8+Fhw8drvKbUEnUayBkV3Xq3MXu/8tr0XkA34pQtzf9I00ZtZBkm0dmkwTR"
            "iKyEkYuOHv+Pne3uA0fihy9cv2UfqxNz5sz5oqkrHWNyn/BNEhJEdYFQGJZEAFQYdUsA2Ltfcevm"
            "tWsuPfJb+KfYyJRAa+aE08hmwH6mYVmtg0mhGgu6si6LyhCQyzIPS7cMT99z8G+VvvwT3S8a++4d"
            "+/dsygefmganGgenWmswobgeqeGU0UrgUCNlVKmxpZY2SOnOipMHAQgzjAMS1myCn4gGEA4pndu8"
            "6xDbyCzYptJao2Iy1FopxkKaMipCRINkKNEHgRoiIZ3pSlu6mS2tLqMfBSUPAoS5fMqoAIE6KaPc"
            "ylz9ZFh1yiiVB+HOhGe7RucACqr4P/EIPwjl+M9b+iYDMXJMWFuRkAdCbmikjEYVOIjechDekumD"
            "wEq4ke8n4cB5CMK/IxwAACAASURBVIHlgBwyO79RdoM+h7FLY1+zo/+ccY3KsYuQO8lyrcMVnd4Y"
            "OW/V9s0//wEUdPUGUf7//PecW2yBnYQYAjVEQgqELlH5rX2Spy/bCFsyyRH+1d2+QxbkY1a08gEg"
            "zHeMyufZT9yXwoAwKNIPqweEanmQvguzFvVVTYasfowzAGEU8F6+TWj2qo3q5G3Gtyu+2WMVku0S"
            "TdNoOSehiP0MFZihK12jqwLCaA0OdOFR0EVP8RiOBm3CcnonT4fz3HHg+L9nLl2/SfI1Vny91zpU"
            "SZJIoykTxpa4aZWWidHBvxidlFGyhh8cLrJlMrsLMweHmiOGjuqioHa+qKZUyMiTCIOOkoLU91d8"
            "s+Pg3kMnzl68VnGbPEbnTVxlHqjw6kGOSDZm8iCrJRNT4hiR69fnff4Lr6I/TfKfY/+cbRtb6CoD"
            "ICwRbIQeAgrGksGvGV5th+EInZc6GaQ96NBAu+F6BMM4HQ7Ui45xAgqyUR88hCM0FEJ1V3rahJAC"
            "IWtCGDf4Iy6F7MEfvHRCLxACDQLg9f1fn5HDR+Qqc3IUCkACeWZWdkZmekpqckLisCFDBw8c+O5b"
            "b7/Vr3+fN9549eVXevXoGRsVI42IDAsOCfIL8OvevXvnLp2e79D5hQ5dOnbs9PwLHZ97vsOzzz3f"
            "/hnd/hOsawVrXyEkl1YHCNlLYJU2T7UENCWwaskNB2tb09Ztli1dquL5Qe+Fb92ylX3K3sqG+6yV"
            "NSDuU82aj/lwtPiIbOGTjz/WOhYcyKRlazgHvQfiUq9v3w4PDjFvYwwILZyhHS3NCms2bdwU5Odv"
            "1saYtnDk3oXzgXftbWx3UDYwKOvR/QO3wI8M0F19FZbWYoCHAaD+XPv2A/r2O3DggNYO2WlfvHhx"
            "ZPnwrh062dGCsfAFYFQPCzbmFrYWlgCHo0aOhM3Yp/b8vAduFDlJK43Ttja3hB/iqZMkuebSxYvw"
            "IyYnZq19YrBy2ZKlgToXrnXa3LC0drSxa9msRUZaOjv6ls2bQwICbcxJxVoiUzs6WZmZe7q6paWk"
            "eri4WpmaO9s5/Por531lF/vjth/Z3zUquUvC4WBYGpv9vHs3IFmr5k86Wttpb6xvD+TyzSzgBE6d"
            "OiXcWJjnzp4T5B9gb02u2sXBkf19xNXJCU4eLqFbp86A1oCF3m7uVqZmzY2M3h85Cj741bp1hw4e"
            "XLxw0fbt2+GbvHXLlkq+CVVGAwDC85evBQ0obN6pv2nAEGOuN30iBcLkRwyElLuyAdKeCkp7TTFj"
            "z69H2a/f/06eyRyzDJAPcA6orF/+JwcOHWeXcOTPf4eUL2C96YEkWYsIQwOQCT7u8/YHd2/dBCBs"
            "/1opfJClcVoIQEg3Y/CmTyF8gKIyEXLLcAKEZmFpFqEp2/YRp+JtPkMafhR0aJhdYcBzME1q4/6+"
            "8tXWX5yi5SZ6mFBcg1Q7oVSNfHqxUOtdvXAoUghJidEIki9qHZZu7D9stR4gnNOs62AeCPXaCHN0"
            "IFBLMNSjJerR+iSaKKhHCTSoEKpFQoqF2kAoyTEkG1aqIoqBUFiTZ2DLPFsB7XQFw6qBkGSKOshy"
            "WvokpX2wREXL3xE7HB136I/j92MnnWVym/Bs2NKOSxx9EA8hJxLaaxAgn0QaRYYmBOarobGaCqFM"
            "WMh3kOY6SpWmAckLVm/X+oIVT11p4pfsJFU6yXLMgzJjEyaxfx33eW489Ndp16hcu0ilIwHLfMKE"
            "0XxFGWm+c1SueWB6xHujGZDcv8/ZOv48cfa5F4tsQrOcqPBooNZopQphVKVAyL+rlxsNi4T6yZAp"
            "hM7kNAqcZXlWIVlffr+Pfg3uUqcE+b/gsg27LYIzXaPyKTTyIqHW0OhbWMI3sufchgIWagNhtHpl"
            "1eIh8Q0We8SVGPtlFE8htfsZr7J56frdQK3u0YWu0UVu0dyedQhQXXJGgxj1KoQxaoVQrBaytzQ4"
            "kM8jrVQ5LNdZ0KMQeqh3S7RBj5hi+/CclT9oZ08oxnxuHpjtFcdViOGaSZCFUne4RTGFNsFZi9fu"
            "FLZn8/Dpa62DsjxjijxYXZlYXiGMZezHNa/XQb5yDz0r9UKgenhpzl7CWz14nVCcRxqnKw+KN2N7"
            "0EBEb5o46q1eeLwewhHaXel5DyFpQkg8hEpPqcLGP/l/qZMe3pOYLhACWgAGdOvchT3U1kKRYCLS"
            "TfKkc+3y5csXLlw4ffr0vyf+PX78+NE//oBnXHhS3793Hzxt79yxYytfYGbF8uULFyyY/cknH33w"
            "4bPtnhYKzFQChOyB+9NFi4FM2nu3ZWDJBnAIYNvb/QcYugR24Uq53N7KGjbW+iw83E+eOFElYjy2"
            "MH/efK1jwTIgU+8ePfQ+qbNDw2n3jIkFJBPolw241bBm29at0RIp3HPxu8x4+bR3290//aSqCgjP"
            "nTvXpUNHuFda+xdImzk54ZbCnXy2bbvdu9T7ZNf147ZtQQEBQH3wNRCOLmTzwksY8JaNhWVoUPB2"
            "KtPt/XkP/JjoSWocFO7GC888e5p+eS5dutT5hQ5wx8QbsJPxcvf44vPPY6QyN83bAst6/2QAC8CN"
            "OQol7Pbr9esZqbJdwY2FQwANwtmyY8FZPduu/cHfD6pEQLhr507dm6PRM5MO4S24Xfv27lv5xRcA"
            "7ewnrvtXDPEeuOtyc+/0/AvwhWc/oNOnTvf9Xx/gQDhJtjHbVVtPz/Zebdmn4C34vg18+x34R9fO"
            "ywuWJ46fAJ/dsmnz8T+Pjx87bvYns/86/teO7TsMfZmrEw0ACK/euBkxsOyJjv1MA+tCIaxVyigj"
            "rpZBacB1EfHjiAx4D/6F3Lt59Wr5zDVOsXmkZGh4ljXJDs1q5pMEC9ljl5/49wy9jns79h5+OWva"
            "kwGpMIAbDWEhkFKroHSft96/R4Hw6VdKgPEsI7mNBSBUL0TWoULI81tEpnlomkVI8o4DpLmFIYc6"
            "/W1+T1zhGrZkL3f+8qeDNMs0JA2QzJLDVx4FCaexlFE66Eu2rG0p1F4QUkb5ZdGw1pUZI0m+qBUA"
            "oV/8mk17VVopo6VzmncBIMyEM+Q+rsV+kVq8p/GWTSSbFTYSWh9FomSz8BENzJMamNVDKV621eFD"
            "W31FRzn2E7AQFoCp6MwNg0CYpwmHfLKoREgWpa4/ceKoVruI6miDZMD5KG0jsu0iMvceomqz5i8p"
            "Hs7nt/JJcpTxQMgx4YP1J4zSUAjtNRGRUxQ1ODBfzYd8JwlHbVFRjYWObGNJriN8IUPSPaIVyzf8"
            "xBz2FRW3Jy76xiY0zS48yyFSDkxoHpQuGzKeJFzT2jDse/jHX6ddZAq7CLmjNFcoRupEgZDuOcdJ"
            "ojT1T34r9+N/T3N/av392H+yIR+Z+ac4S5XwKSdiTQSGLNLbv15nMDDjYS9KA/acdFCQynpE2dOc"
            "C/n13PaGdEJRfVH6QVm+syzXKjhj5ffqP80w1lq+YbdlYLpLVJ5zFGu6KFSXMViP1JXXHl3VmaXa"
            "aqGrOq2USH+VDQ4IyQ7dY4pM/NMVYz5TMevmfe4X2jIAwqBMt6h81ygiY7rxZkIxFhIIJMsMBUv4"
            "USxa5nNNYzSSS9257YlqJySR8hmk7GUpRThugVtTCR/GlArDQ5j5/XjEckf0iCmyDVWs+HqPiv+r"
            "H3C6igDhZ2b+mV6xxYB/VEjkYTKGAmF0oVO4sn1cwWdf/8y2r7h9Z8ri753D5a7SXHiXoCZtWM9L"
            "gjwKxnKORE++o716VMGEfNIp25L/iFdsuRed6cfpsgCHPURSIV325BaGV2tok+FjTBkdKaSMqj2E"
            "MYJCyLrS53jB/xN9EwflfaJ6gAfBykMXCEmrCVu7119+Zd2atZ8uWrRg/nyAtJnTZ0ydPBnGxPHj"
            "x47+6INRpJ9EWUlpSVFRYX5+QV5efg4RErMzszJS01KTk5MSEoYNGTpk0OBBAwe+M+Ct/m/2/d9r"
            "r7/60ksv9uwFaBQji5JFSiJDw4AuQoOCQgKCgvz8A3x8/bp19+3arVvnzmI/YZVAuHD+fHjg1jUf"
            "AiHAtQB/qnSAihMkzp/36dIV0ELLvggvLYxN4EFcpaMQzp09x8rMnLAcUzJ5DyQ80//w/feGDgSf"
            "jYuOcXFwhA+Kj8KubtuWrZFh4UAdWucPJwZMsmvXLt3dal/IuXPPt38G9q91+UyJggWW38vIELAK"
            "+PMO7ZXNbuCG9etJhrCtnbc7aS+pN7OXJfey7wYsb/zhh0MHD8ECqwzEdF3YAC4nxD8gKlJy9uxZ"
            "2POVK1fgZx0SEAikxEyksCWcJ+lg6ei0YsWKiNAwOB/xPXFzdmYtK9mwNDV3cXCC9bBnazPzgrz8"
            "mzdvAjgBDQItu8MJODjC4fq8/oa/jy+cG+NYOBa8+/vvv6tEQAhAxdKYxYcDgBSaZJI+meYWDCzh"
            "XbiuvXv2fvn55/D1g6N7urkDD8MGMITvDNsD+yBcCCvMCzx8kgqk//77b5REAm8J3TvJxqZmDta2"
            "sAB7gwH3kOE3Y1pYhsscN2YMfPzbb76Bf4NH/jhy/M8/4V9i41YIyQz/u4odNsqoQ18eCIdRIEyk"
            "HkLBRqi/N/2DAyHrEtE6JN3IJ7Fj3xELV2+7U0H/bH8X/uXfv3n9WtmM1fYxOc39U9iW5rTSDBzI"
            "yCfJvVfB6HnrhfKhq3/4OWjQR818k1sHp/PGQj1A2J0HwnaGgVBIH9UBwgdqO2EJ6BueYR6aahGc"
            "pBcI4SHprmZxW63CuOwP/3BPmnePt47MEuqOWkvkNhK5rVQBwxpojbBiFhHxIrPJStImXs5K0Win"
            "j0p0YY8UgwEYs5Mp+b2RMqd2ZOcAVwpBISRA6K8XCGc37zKIACEwPK0rA3QHn7WViAa8lOZocKBE"
            "aSdV2suUpM09ORA9DTIUdD1FO0KJvFqoU3dUXYlUwu3NQUY0NFhge3CIIi/tyZocdhr8HjjqEwjQ"
            "RsR+bCe2lEttSZkWtk9hAwNYKKFJpxJyOKeoPAdZLuVJclDifIM1Ubn2Mi6n1FZaI22QWhPJiSlb"
            "+Sa+nf+JSkSD9zWzRn/cd9Qa8ClSbkduiFokZN0LHUmhzjxORosig1gNpVryIMufzKe2OpJyqaMf"
            "5vEbsC3pZlHq9FFH2lmeDdouIs8BOI2UhMlzjuaazjtwiFjIMyFRCO0jlUB9NiFp5v4JQf3K+2RN"
            "6fp6kalfgl1YhkNktj0FQrPANNmQsexiBYUQgNBNJrePyHYmN5+cm3N0IQyWMkooMVLhGJlt6pfo"
            "GZX9SsrEngljnSPSLQKSnSTZDhIAwjxHJg/yw5mkXJLhSpgKwKmQXg556UZaPhCcU4uHIgFQmJ00"
            "OBA+CA/3wDNFwm7dYsmuXNnOBUrUh4U8EFIDIQNCaY5VUNrK79X/Enkg/MkiMNVFlussK6CXUCwC"
            "PzgcMAk32KFh0Jcl1NFHUkxd+X73mjTI8R685RZd5BFXAh+h2xOig514xBI90I3stpjJgy60XIqx"
            "b6py7GdaJ7ls/U/WwRnu0flu0QVuMeysSuk5aDSoECAQztAzltZrIcelL+NK4CUtwlnCEyBZ8KA1"
            "XbgRRwlQs2iNh9Dcj9umzEONhWIO5F2IMawfYKlXjzKvOHaSpEwOedmjjJaQEdCuyDY4+/Nv9giX"
            "yWVxj/nMwj/dG+5PDLuKMoZhTFd0iypyh+9quNwuOD38rfcHyGf49ym3CUxzlShhvTvc6pgSkYeQ"
            "8RutOxpLSpi2JVVMyYXADMusqCmHiHokRE0CJNsPh9GuZ3nbnrBcRgfsc3jbnoSUuM14MtSU/vg1"
            "sUQebEdkN3jJDk3uTLue3Epvuhk7kHelOuEj9hAaBkKll1Rp0T0hfcQi3f9f11UY8hDCcyrtOE8S"
            "F2HQR3Paa96Ef3qGx2jRk7QlfUnepStZA3o2bC0s4VkZnnfhed3BlnMeskH9h9wQmw9r5CE8ceLE"
            "d99+u2H9hm6duriLntQJQphbfPThhyqdZE72ctGChXCBTKkT2mx0ePa5dWvXwg6PHT0m/p8aW/jn"
            "n3+AhZYtXfpM23bM9AgfgYd4OMOUpGSVged1+Ozu3bv3799fkJtnZ2VNRC0Hx/CQ0C2bN2//cfvF"
            "ixd/3LZt586dgNPsXaCd2KhoIEwYly5eUlX154Dbt29v+mHjtq1b+7/5JpOzYAYOh91+/913mzdu"
            "XPXll++9+y7jJcYeX6/fwD67b89eJlI92+7pag4mdm3bug1OHg46dNBg+CnDabNM4/lz5wk/KTav"
            "/OJL9jOFbeACAd6AbeAOnzp1Cj6+c8cOeVYWfEOY1hcWHDx1yhQ2pkyePHPmzB4xsUwMtDQxHTVi"
            "xIJ58+ELxggTrmXpkiXs/ty4cSMjNQ22hB+NlkLI7t7FS5d++OGH77/9ThIWzuDcydZuQN9+c2bP"
            "njxpEhxu2tSpE8dP6Pj8C/DlZ5DPgNDO0gpePtPu6XFjxs6cPn3GtGnS8EhuD3b277z1Fuxh1syZ"
            "2RmZni4Ep4lAShXC115+hd0ZOB/YG9z/l3r1njp5ypLFn362YsX8uXMV2fLn2z8L94TlS7NvO/zb"
            "YX+MqKio2Lp5C3zZli9bBl+VWqMgiwYAhPBz6p30odELfc0CWcpoPA+EyRwN1ggIa5oyGk46Q3i+"
            "WDhuwddXLl9RURT8dN32lzKnrd28V3WftpU/cTrx/cWmNC/Ukpb9JFgolbcJzWDlQxet2Xa7gniH"
            "bt64MevzTc+9UQbrzWhVGLNqAKEeFDSoED5IyijsLdMiPN08JMU8OGHHAZIyyv4Hw54YPl23tfvr"
            "ytC3iqOHjOyvnJz+/rwl6368Rj1RWr8Qj544YxeZYRaSahmeSSEzyzgojRQj7Rpv1HXoU36JxoEp"
            "8K5JcEqrgOQnusW38Ek0D82wpUVKqWaoFws5gdFWKofdtvRPNuoS/6RPoklwqllIGuzHqGt8c59E"
            "05AMupMsqhCmGVAIZzfjgZD2VyQSZevAtDaBqW3YHJQGo3VgOhyO618fqTALzWzln9q8eyIc6Cnf"
            "JJOgVItQuFdpsADnD4c2C82wo1RpQ3yJOljIDSBGwmzWEXK4J826JTTrngjX0to/BXZi1GVY8+4J"
            "T/kmw8lYhGXBrgh88p8VMkJ5GiTkZh2R3dIvBc7HHL4qYZkWIRltAuB8ksxCMuEo9hxY6voPyXp4"
            "F7jRJCijebfEVv7wQ8+wCiV7aB2Q2qxb4lO+8DPKAsIk+ZwSTbWw8hqkVB6E/dvBNzMo9ftdB1UG"
            "nlTYt+a1jKmt/ZIcyNnmCi0KHaR51hFKOAEygjPNgmHOMg/JtheERCm3GZwbaQcfks1m28gcSnR0"
            "0AXYzCJUTlvGk72x3vE2ETlEJyTvwj1UmARltvJLa9E92cQ/zRKuOlxuE5ZtEZLVyi+1tV+qdbiC"
            "9ZAgXRNkXBNCsn+A3ki5dUi6mX/iU50HN3vh3VZdh5gHJNmEpsN6AnUAhAGpssFjtIDw0PGT1kHJ"
            "pn5JFkHpFsEZFiFwLLlFqIKWmSGoCXu2C8+2Cko19klo0fG9JzsNAs60Ck61j5A7SnJI5wnajZBi"
            "XoG9JNcqTGEamNnGP72lT2qr7qkmARnmQZmmARmtfdNa+6ZbhmQTa2JMIZdlqsOEvIRIoBFQEPZv"
            "EZzdyjcNdmgelAXD2A/2TF7CUSxDFHBECo28fqhPKnQRAaELAGFgqi4QAmtZBKS4yHIINJKs0UJK"
            "ceQEbMOUcFyTgEy4hDa+aeaBWZbB2TDDmRj7Z1iHKmF7yqu885BnQkaDjC0BxhwkeSb+mSb+GbZh"
            "wNhA2jk2oQp4aeyfCTtxkha48nVQ3QEIfVKUY1doneSna3e06hJvG5JhHZJtHaawCcuxCcu1i8wj"
            "tsNongNJxikgUzHcYduwHIugbGO/DDiuZRB8PeSwYBqQZREst4/Ic4kqphpjMRwa9mMbnmcbnmsb"
            "kWsTnusoJQobHaVsAW6gHWwQAduQ/u8wO0gK+A3EUiFDQYKdcA5wFLMA+Kcntw/PdZbmw2weALcO"
            "vpD5pIwqg+RokvwpBkJ2yemjPn2y41DHcIVtqNIuPNeOHBd4npAeOS5QOnzxIhS2wRmm3RNbd443"
            "90myC850igAgLKBAWOrBi4EePAoCerlGFdmE5MA5wMm4SAvsw/NgGda4RhUD13mqm1WUa0iCVAaE"
            "d2Ezx4gC2N4ySGHmL7cIIJ+1C8sl+wyQWwYpHSMLGHPSw5VrlplhyiHZFVCoi7QIdmIVpHSKzHeT"
            "FrlICx3C8ywDlVbBOY6RhXD+RHuk3EhHPfEQqlNGvYD2OSDMZ0DoLVOadY0vmvCZ8KOs89ALhPBQ"
            "+9zT7eG59vn2z8ACPKSyl2ywnDeWRsikD5iBo8i77Z99ti1hhhfaPwPLzz/d/jk6iJBia8+SD59r"
            "9zRbKQAGedmuPRwUHrJJ8VIqOglVSatuO0F/8d68eTPQzx8+yOiO0QLggSwi8tYtjScZIfr16QMc"
            "wpIDmS4E3NK1U+dr167p3V44gU9mzYIHfdg/sEf3zl2YgAaX/9fx45V/cPq0adZmFiTL1M4eOIdx"
            "KWtIAAtTJk0C1CHJrja2ffu8WYufozwzCxgYbinM+bl5KlFdUNig7//+B6AI+4eTByyBlQBRMbIo"
            "UhDI3mHWzI8BLM+cPn3+3Dm9A96quHXr00WL4WKBbV7s2ZPtGXbSK64H7JmllcL9BI4SLvnwoUMd"
            "n3uepcvCUeBndPK/kyqRdgrztClT4YThxGAn/fgLF8qxAC4CMsGXBL4GpUXFqUnJcP6wMbzMkStU"
            "YuS7eLFb584WJqbwFqCa4CEUf09g25d7vwjvwh6sTM3gnrPTEK4lyD8ALhCuAq6FpYwykO7asdPV"
            "q1fZrgYNfA++OcQeaWY+jd5JiI3f/8DkYvi3AFuOHzsW9t+eV/98unZbv+4rLcUFZrir6ampjnbc"
            "vw6WrsxSRtnNOQNx+kyNvgx6o6EA4QdGHQwBYQ0VwhoDYaZ8/PIzZ86xM/pp/x+90qc+GZACRNcy"
            "KL1P7qzfDv/F3tqx73DPtCkt/FNaBnHdIMwjs61oo4jm/imSYeO/2XaAJpqqrly6XDxtpTk9ioU+"
            "ILx784Y4ZbQmQPgARWVofVGiEIakWARpACEraFE8eZmRS6823d9p3mnAEx36GT3bx+j5NwP6Fhw7"
            "oVEvhEXP5PFP+gyzjsy0CANCy+z85vDX5TNSP1zy4dz1n361c92WAxt/Orzhx18/+24PrHlDPsMx"
            "St6s+zCmGVKi09IG5ZzfLyIbkMkqPCMqYVzR1JVzVm77atuv3+06uPyb3R/MXf9u8dxnXy02D8sA"
            "YoRtCBAaUAibdR5kG5FJNoNLDsuAETroo/DBY8IGjSHzYJjHRiVMdIzKJfIjPW77V4p7p09JHLW4"
            "bOaaeat/XLNp/w8/Hfp2x8FVm/ZNXfbD4NL5z75SDIDKQI6Kh0rt1NBIBXkrQg7sZxmaHvLu6NwJ"
            "n09e8gPsbfG6ndOXbxo5a1322BXvFM6JSZzg1avAjLIZ+5SNIPeRFg5UwIxUAAfaS+SvZ8+YtnwT"
            "3Ad6S39bumF35kfLO75R1tI32TQ4gyKlUkMq5GkQoBFANHzQmJJpq+et+nHtll9gD+u3/bpo3c6R"
            "H68bVr7Id8D7tpFyWyKB5mpXl6lEJ6SHsJcq2/gnR8eP1frn/N+ZC7cquBLV7Iey8oe9Jv5JDhIF"
            "EQkpeQITwjU+83KZLH6iZMh4yVAypEMnhA8ex8RPVoTGgdoUXWLyI8k2E2BI4ye2e7HEJkJJOZCi"
            "I216EfTOR7L4CWwnkiEToodNev61ETa0Cqi9JOf518pfyZie8v7SER+vW7Bmx9rN5Mv5/c5DazYd"
            "mLZsU3zZog6vlbf2S4PdEqmQcCbMBZQ5gQ2yO79eFj10rHTQaNmQ0ZJBo2H5+ZdLbMKyAAiJQhiQ"
            "IiP1HjggvM9VTj4XPXh05HsfyoaMkQ0ZJxsynlzp0InusUVwwo5Ue3SNypEMGgOfJWMIzGPCB452"
            "kiqIQkglU0BB4EDrMLlHTEHAgA8H5M7O+HB52fS1ExZ+P3flj0vX7563avvkJRtzx38ZM2ySPdB1"
            "UJZLNNcLkUmFzmLnIWkJWAj4BLjlLMvtkTS5dNqaGcs3L1q789N1u2Bh1Kz12WM/G1g4H/bm3aMI"
            "jmsRoiA6pJoJdYCQSxnNc+aAUIQfTCHkgDCXppWSE7AJV9qGKz1ji4LeGv1W3pzM0SvKZ6ybuOj7"
            "eSu30yvaMXXJxrwJK+MSJjtG5gC1uomr0dAMUir3EYERdmgWkPn8y+VwW6Yv27zqh/3wY/1h1+GV"
            "3++bumRT7oQvX0yd7hVHNgMadJHlu0fnG/skawAhnTdsPRAyYGTM0HHRQ8dHD50YEz8pLmFK+HsT"
            "XKmMyQRGuExzwn4K77gi2ZAJ8WWLR8z8atrSzQtX71zy1e5x87/LGf/l0NLFoe+Mc48pgnN2kuZ3"
            "6/Nhr+TpMfGTY4dNiYmf0jNpOqwBSnTn9ENCg8+9PCIuYRpswEaPpGlhA8lxKROWuYsTSomgV2IT"
            "qrSPyI14bzzcojlfbF+7+ddNu4+s2fTL9GVbskZ/FjFwvDtQNB3usnyboIzPv/lZpakQjpu7Iajf"
            "yB7DJsbGT4ql5wYn6d9/jKsMGJLois6Reb5vjuqZMCl26Pge8eNh7jlsUtfXRzhL8tRtJ2K5voVe"
            "caVuUcVWwYpnepe9k79g0uKNK7+Hn8IfMMPyu/kLnu1dbh2shG284hg9aqWGljlEFgCzte9VLh00"
            "aXDxQuXYLz745Gu4nIWrdy3fsGf+qp2TP92oHPtl1JDJjhF59uH5DC8J0aldhcMZlLpHl8CxAvp9"
            "lDd+1azPtq3bAjfn6Pc7D3/x7f4pizfljFvZO3lGux7lHtEcE3pTqnxsHkLtrvRiICzmgTCXAaFp"
            "l/jhU1eqHhUQsofXnnE9zp09d+7sWXgWPf7nn7FR0T/t2nXq5Kk/Dv9x4cKFEeXlLImOqWqAcEsW"
            "L4bn2j8O/xjhtAAAIABJREFUHz5/7vyyJUu//+47eIQ9dvToP3///d9///3777979+7NTEs3adW6"
            "rKT08uXLR48cOf7n8RP//APj77/+gs3279vv7+MTGhREa9L4+3Xr/ky7pxkTVgmEjAeuXb0a4OsH"
            "4ME+AihLL4doVt98/bVKJBKyS/7tt98YOsLTf+cXXmjv7c2yNLt07HiBVqc0JMjAU1hwQCBsCQ/u"
            "CUPjP3z/fWtKcUAp7Dlebw2bCpp9RsiHASHVAGEly8Ni706aMJEDQlu7Pq+/QXsHVrcP4R3aHCI9"
            "JdUGyKrd03AyCrlcRcVD5udUMRyluAhHmTCOaFAzZ8xgJGZvZfPp4sWGTp4Fe2vN6jUuDk5wCXAI"
            "YEj21qGDh4CCmA0SaFYSFn7xAjFE3LpVAfTFZE8P8tN027mDuJRvV1SwS2MXDlQGp9G+bVu48Dde"
            "e4314oMzZ+8OHTy49ZNPwU+2dYsn4QLf6tffwdoGTgC+FbNnzWLXKNDjJx9/nBg/DL5sWekZ8N1T"
            "aUoasA1s/GLPXswIasUnZ8L9YTu5cvlyx2efg5MBNIV3f9798//ZOwvwpq7+8RctpZq2qXsZLnWP"
            "pwJsQ4ZbqTeNNdakLrjL0BbbcNfhMHw4g8GAMXS4u7WQ//eck9ykqcDk3e99n+d/n/NcQnru8Xtz"
            "Pvdrq1euhNrpSEjofR/L/WBeEgcPIWgNszlt6lRS/uYffoAL4apW2PIzsEMnL4zHcIMAZMJ9pNUH"
            "HqzCB473qJvfSRMmOtjaEkF60wYNx49FYm3jKIt/Uzyo/V8CQiQhTEOhCP8mEH62yiiNpwLqE4xa"
            "rMVSwaePn6QNX9QkSgoJ/oR0PrmqBmEieqymaOaGxw/Ryq56/37VtqMBA0fBhZYsBfEsqrM/jMpu"
            "zpB1lU4/dfYyFKdFHuQ3NkTRBdXGQAiQGZo49kPlu1fPn9ejMlq3DeFflhCq9ECIJYQAhGcMQEgU"
            "QUfPWW/Wuh+dKaBFZ9JiBA4xAkdGptkXvYbmz9QaiYDILko5cblZp1RnvqpBUIZwdH3aLGSWr956"
            "IJ+wonmUGEW651RnQiQYREqh1oxsyJBc8t25y7frWjNv31Wu2H7cgS0H4qpPQoiAUA7QSOcqmoQI"
            "pizepcXcC9kgEYd75WsOAJI58dTNo7I56ZOfvUQ+gmvtA+nCk2evpi7d7RantorOxkqwSBiokxBi"
            "m0OgOOsYaJU8Y9jCY+eu1a/e8+DJi/0nL8kmrMSEqXbUMyGhQTuWwjo6WzBiCQmCZ/xzQD6+ePVm"
            "6dZjIQNGmoeLXWIBCNUUTxIatIzK7tCrdPXOk5V17yGghQd/vuwej2RoKLg8r7qQsD59UUAydbPQ"
            "rEWbqsVOgCOtYNaVP+5pibIx/gaGPXrIaOtIqQufhCXMd43Lt45RdBZN/4CnA1Il/j1/+76SmTLR"
            "JkZJFFxdY/MsI+U9FeUozuyHD2/xrBVM39AkVOIWByiYC8TowM5p8VXxvUdItl9ZpZtf+NxXPcc6"
            "WkFjqgAUn71888mZnbb0R++EfBpT6RZXoJfj5bvF5luEScpX7dPqAl3q2gCA3TRE5BGb68pTAxDy"
            "jYBQX+zHKn3XUO+wr8bXb99HDB5nz8pBLliYypD+I7FTGWTJVol/Ze88eObXpcCJo4Z63eMKbGNU"
            "PeQV+05cunHn8btPRVg6dvZaT1m5TZTCXac0W2RwSKOjwUIaQ+WdUKCZsu7XK3fq2WDAn6Ale47/"
            "ll62xIGl0WmQGgGhQWXUAIQa+yhJbUB4zC5S7MHPdefne8ZBA9TfyOcAw3xmj3opKuxilJ5GTkr1"
            "NFjkyst34eYWT990H099rd2BSbn/+MXgvO9pMTne8YUICENEmkmrtPrHRa2TRd6Onf7tlncCCq7o"
            "hciz6IuvytJLl6zZ9fO1W4+q6rXkEY9a4chQ20YqgRK11ZfN+AU7LcMVvgBRgF6dS+0ZmsF5yC1h"
            "pVG9P5256onkqMWUnBDTYwlApn1MTjfJbFgPdW0LoJyfL9wM6z8OxQyMAyCUGQOhbkxMOosbtmD9"
            "EfsYjV/nUr/OJXZRqjmrD2qx9SAMIPGCO3z2Fptwhb9OA5bYHJYCDbqw830TiofN2gILxmQWdPEq"
            "Hj4fUb7VL6HYhZOPmBDpu+oUTf26DnPhFGSWLT/6y7Vb957W85giUwmo3yVrJp2R648Va7GRoYEG"
            "PfnFLTqXTlm053ndNzv0/drtx90kFa7cIv+uxOno/53KqCkQjtCrjJb5xZfWAoQz/j0gJCZMxYWF"
            "8CWQ2+vXrx/cf9AlPuHokSOUHOby75dbY6cvRATHY3OICE6Lt7BLFi/ZtXNnFb7HX7x4cf3aNWJO"
            "Bke/3r2VcgVafLioV69evXz58i1mFWCwrzp3Ucjk4ixhytCkQf0HhAYFk1D1nykhhKYaA6E8W0Yk"
            "k040eyhTa7SlJmAzavgIos4HtaiVquCOnYjFHQDh06e1a2mSC4mXkZb+/nQ7e/h89coVTzc3AsYc"
            "JqsuaaROQjhzlgMGBhg3GFXyJSUhnP7tNHtrAxD+hXmUS7MNQKhCQAhMBYWTVgGvEvwDQiP4x4iK"
            "gmbDIACKLFuyRKuXldV6kEZu3LDBA6uMArp0jo0j4eDh+y2btxDxLJQGtUuxAm1hfj6xoIOl4kKn"
            "L1q4UFudOUmZM6fPwECIOt63dx9SHWX7d2D/fgC/pYsXw/nY0aPJiYnOeiAszC8gBRJ2qh+eKUmp"
            "DghbfGGvV84kNKjVhyFZ+N13ixcu/H7BArgFfjp4MCkxUZCRIRWLKQnh0CGJlISQyBjhOHP6dAY2"
            "mi0uLBo9cpSTXhsZlseB/QeMazGZOPLl2NFjUpKSxEJh4qDBP2xC/s+I8wIT062/fPyvAOG42oCQ"
            "siGU/SeAEJDPLEgw5jukRf3+3Zv3b99MW7oLIK1BOAoMaDAXRGHlhW36DFu48SAJK//86fNxC7a6"
            "d8k3CxPZoIATADbIIU2zaFmCeNqpc1e07+HG+7Dtp18sGXJbrkFCCGDWPEbesf/Ihw8fv333ruU3"
            "ZX/ChpCbY8VStO83gsQhbNGjFL6kx+U68NWfh4Wqaiqj1SWEBAhHViAgdGQKbWOybGOEkMcuWmAZ"
            "luKXkA34oaXk+1UEd9eZdRjqwlMAEKaXfY+L0u19q0giew7smYZazYt+OGwZJaKxiNpnDo2jp0Gu"
            "yioGgEG1dtcprZ4l8Ea5Wjlk07Bhz8+WkSIHtsyeKbUMz6gDCJMd2TInrrxBYLp49FKt0dOZ/KZu"
            "OXjWhpFtx1SgqqOzO+ijPOEuVFUadwF3ijhahQwnz1//olshgB+iOKRrSmgwB8ASyvni64I9xy9S"
            "1VV9INt9kqoqSbBn3BDS1H0nfoPRcODm6E0KNU48jR1T7srPWbPrpFb/IKg0akyVUWMePn35tWR6"
            "s3Cxs85CMpfI7qAlYYNG3cQOS8h+19Aj3Jj3uI/w1wtX77jGIqkmchJjHJ+wXueiTnyNbUx2x2+K"
            "n714rTXyk/n79TvNOg6Yv+ZHvE4MYAB3VtMQgSuypdRJ/xw5OR5xmqu3kGT+g16LA87i0cuahorc"
            "YpHHGjg3CREOK/+BlEaKghViEy1zAehFkrRcq8jsHrJZRo8UVAjwoV/XAjpbZc9UBPQZrp/Zj6aL"
            "s/rMnrpwo22PUnumCngVa6vmucXlARBWrN5PdYecy2ZvahIsdI/VuPJybCLEsanjTYCw1gcd7LAj"
            "Bo8FFHTn59kzEBCSgG/Euy98uP/4uW/nfCe22pWf5xlfYB4iGTlnK1UplP7hg24tVekTXhjUu0+t"
            "evJaFNc+ThfGUKcmGlsANGgXrYwcPPbMbzeNFueHaouzyrA4yVAv3XIMSvNAkeWrCQmrAWHsp4Aw"
            "QuLB1wAQAl9ZhGaPmrvtT/RIq82dsp6GmdAdYxLGQqSk6sLRrNxxkiyeD9WXNyoEd4YU0i9nLi1G"
            "5RVX4B2HgXCiKRCaHGQqz166BQAJzfZJKLaNUgmGLzN+hpBbstKozZX6X/QB6nl0hso6PHvKot0m"
            "y2bCd7uah8kAujyRQWMxAN7g3O9M6gU68owtQFI+bByoo8G4IjpTPWbudkpNQ7eYjZ6NpEcAYK2/"
            "KnXn5vrE5TtGydbsPKGtlyJIw77fcIQWneObUAJ0ZxepmLfmkEnLR5ZvtQ6X+3fV+Sb1xrJBOjOX"
            "MWTimd9uaY0fU1VGL0H0j31YdZDTiZlHkJKY9gHX2UdrFqw7/Mn18EHfbbiRM0qXQL3+XRGU+mE1"
            "VKBBr7jill1L9534Xasj3mqDYzw+wLfc5Knu3EL/rogn/XVKp/8NQIjjEHY2tiFEKqPeBiBcX/9U"
            "/p3DBAiJQ472rduEBARBCgsKDuoUQNTntm7eotXv4GHXSzbEsPWfOH6CFstY4Dxl0iQnmgPsgGFL"
            "Df+dU14BvAHZCH7Af5s3aZqWnAyfnzx+wmWyoGSVXKHF6Ni+VRvYoJMIh250J9QevduVPwWEBFOX"
            "L1vWq0dPKAq6E9C+A3HyQS1LyMyKiSFWi1Fh4du2bCE2gfUDoe42x4qXkDm4U8BDzLrdv/oaGoyE"
            "Y3SnDevJZJnK2f5NICQqo0WY6ikZI2B5LIdLdHEhw+Mnj0+dPEWco8AIIwnhks+SEG7csNHDxa0F"
            "1o308fAkgQFJ8PoJ48bR7WhfYE+bvp7e6SmpxNIPvWWwtSspKq5ZvjEQfrLj5NoRZcOgjy39WyAf"
            "Nn7+q1etMimQMGrNd2e1AKG1rbFy5ucfNYGQmkrSzi5x8UQISUfvI7JqdtykYf8I8tV//I8BoWXE"
            "35YQfrbKKALC4KzR89AD7sH9h2/foNcnDx48ypm8ClAQmQtiEiPmgs1iZPBNgvjbw6cuEsPCqzfu"
            "CEYuAUhrECaGP8WkTti2n/gm1VYhIPy486ez8FeK9PSxClHoed9uxeopa/x7lDRnKAj11aMySuOp"
            "gPoA6hqEit065797/erV8xd2LKVZx4wGIaJGEVJrlhIy2OM8n2tDWBsQjqrYYNZmoANLbMsQ20GK"
            "EdnBLISnefLFT1+80lYHwmHlG8zaJzpxZeZhAlbKWEIp7yursBDuIyGEysoqI2H3R/1meiPWNVXq"
            "7AmRpiiiQa/O6uO/XtOSrb/hNd4HcmeTXQJ5ia6csMIsIA1gD4CweVhtQFg6r2GnJBeevGFgavfs"
            "aUQe+EEHmaiEI79cAZS1iZE6cpEtoh1L5sxVXP4DPdbfV+oeJsZbHF0XPn4kDdh26JwtI9uBo3TE"
            "7mqwzifQoBRo8OI1Ennmg749hsspatLinS4ZjR0//WrHkDogJNaV4wArkyXf/tOvxkNBRu8jdv1K"
            "CvyIJEs4cNDzV1FDRltFSZ2wlimUYM9WuceqSZzJSuRgUDf+Rmr0aLgIJs1bd6BZqNCFr6H0Oat5"
            "HK019iASzWmaBguKZ6yjRp60asL8jWa+Pb4WjDG6x1H1t+498euisWcpnPnEZDEX4NAiTLRgPdp6"
            "vscE8B6XAPQF7XHlqZ05Oc7cHNvo7M37UcSkSv22EBjSI1btyFI6c9VusZomQVnF+N05kcJV6sLf"
            "nbSOkACtObKVbvycKzfJzFZh7Qzjmf1oMrMwHUgXFIlY81z4ue6xuRahIgKExo0snbUR6q0HCD/q"
            "BlyXSEUw4BGDRtNiFEjBkiEP7jfi9TtdwDfStbuPnvsmABCq3Hi5nnH55qHiwukbtDovkYbxrEQd"
            "wX0xehtKwVy/nDk2UXKP+ALslqaQyAbtohXRQ8bp5agf9BEFDDeaya/R+0rMCRsP20TKPFDQiMKa"
            "loQenwmEkQQI84CvLEIlxbX2qKq+HvVXzwWa9UpAJohAg17xRbZRilEElSt1fSG3BjUB5FkEH+4/"
            "et7yy2KgR4/YPO+4/Oa1ASEKrmM0WZW478AwXnHQrzyfhCKbSLlo5DLSbApO4DodvVTqwIPcjO27"
            "lXrw1JYh4skLd5osm/ELdjQPlfp2LvbCQEiLVg7SzCOr5aM+4t+RM1c9eHmeschgj4RGhP46sTXL"
            "Nh8jza7SKx7gZ0IV9aaMiPJW7zjpEKMEGvTi5zpEZ6+pISE06ex7fNV363+yj1L6xhf5JhTZRcrn"
            "rjlItZzcUCNmb7EOk/t3IQaQyJSRzsrlJE8m4tkqw4oyvHPRbbzIvOB3NOykyc7sPD+dB51S/85l"
            "9lE5RBqJ1oPRE0O3wnFvP+onlbTk9dv3CRnTXDj5/l30zmw6lzkxNYs2HdWvK/16MLxW+Eith/OX"
            "7/jGF/tgFVY/FAi+FpXRf8mGsA4gRFHpTYFQ/S8DIWFC2OzaWFjaWDSHZGdpBft+AJjNP6DXc8Qu"
            "Cz7DfhfHJPC7cuUKrMVLv6Go2QCHAHWuTs7fLVgA/506eUoTswaQSotLtFhhsmnDRklDhmixh09/"
            "Lx/gw/CQ0BnTpsOFbVu2ImZUX+hdMv4pG0IKCCG/k73Drp07K8rLaVbWRCBGPJ0QVT34sG3rVmg/"
            "kWXlKJW//vorqaseICQDdfz4ceJLE8okQjA45lbMIVZt8GXSkMSa12r/XSCE8qF3Pbp1h1mAQZhT"
            "Xj518mRGVDSZMjsra6gIMo8bM5Z4Z4Gh/rNAiDASj17F7HKtkQ0edB/QjqgTu2D3oWT6+vftV6k3"
            "mKw5LHoJIVIZ7d3zm1fVj+fPn799+5Zoe0LmCxcuQIOJmZ+Pu4cb3fmb7t1nzZx58sTJqkqDVnBd"
            "PG8AQv8WMBcTsHLms2fPqOpevHhRpVPqNDxSTH4lawXCj3q1VWBvWMxEhO5Gd1qzenX9A0s1r2ZF"
            "/+DxvwaEf19l9E8BYVDWyDnoATdp0Y4u0mmXr90i7Tp57vJXspmNIiTNsLkgpRcK31gy5Fmjlty4"
            "eZfkBD78Jqd8+rJd796id2MAhIs3HapYjUJOb91/xpKpoJRFAS+BG61ZCtfOeU3CxGYBAiiQxq2G"
            "giZAiLx38tVQo1kIclETnTqpbM6WqrdvKt+8nrz0xxHzthXO2tQ7d557l4JGYRIc8QIpYX7KhhCA"
            "UFw7EM7ZYNZ2oCNbYseQ2DElNIbEiSNt1CkpNm208azpVEYnLDPrOJTOybaKEnolqGAnpK3DkNp4"
            "5wfnx89etulRCPhkz9V5IrXnKKyixBv2IK6jdgY1b4mP+tpZyWMswgSOLCwhrA0I00vnNQlKNg9J"
            "jcsYr9Mg0r2zQXnO/X7Lt0uOZYTQkaNAnmlgJbDl0GUi2atVdGDcElJIf01501ChE1/twEGeSKEQ"
            "R7Z899EL1GCaXFVzYEhFu4+et42RALcgd6YcJGZsEpo1YSGWWlPWDnU81LT6TdJPZ67YM2WOCFBz"
            "nPnqJiFZWSMXU381udCw4SZOJsYtbxIkcEVKp7kGM8K6bAh5JNqEGhrsxJKfJtEm9E9d6Hhkv6Lm"
            "gUMcItKOnvmdmg5yFgxf2DQkyxUHq3BCjjQ15iFZSYXztYZnNDqf+PUanSWjsxXOXJUDU+YTr7nz"
            "QPfD/FE//pyUcdaREuznU2UdKTYWg5BncfaYpcCrQGtOMLBM+d7jv9U9s0YzgksYmDu3eYTUDRkf"
            "IhEl0GkFVhk1FpiUzgQgzETxIf6shBABoRzhU4wstN/wNxgIKQnhvUfPfBNy6SylK0/jEZfXNFhY"
            "/KldYPUJRR9OX/wDcNqVl+tG3KvGIR3UL74q/uXSLTxH1Wzraw6C8VAs3nTEJkLqEZfvhgMJ1vA1"
            "+nkqoxESDxhJXi7wlUWIuORTmm+19eimO5YxesQWesYVunA17bqX3n/0QlvbA4f6gszI/hOXXNg5"
            "7lhntS4gNDnIhed+v+0dmwfNBo61icgWDsda8XX/SJOrTv56w52r8orVNA8WTv4e2SyRDRA5j5+/"
            "vXmIBKDLE3nsLKRFKQZp5pqUcOTMFQ+uxjM2H/t3KfZNKKFFqcpm6YTkJs9S44P0KH/KOqswqW98"
            "AQAhVhn9TAnhYWiMb3whJNsI2bw1B6g/VemAcLN1WLY/Dlbhm1DqyS9s/VUpkQ1SD5maTfpgeExh"
            "LdwLN1t2LfWMBews9Ykv8e9cSkPSyIPa2p5U1GH8M0Cy7fzpgis7DwrxTSgDGnTl5POSpxDeqzk/"
            "1DekI8u3HHeKUfsjjdM6gfD/zoZwuIlTGR+dyui/DYQkKkBGatrM6dNHjRg5ZtTo4WXDgjp2gg30"
            "D5vQUgS0uPTbJdj6smIYgFuEf3bt2rV+7VotBkKABHcXV6IceOjgwYK8vCkTJz15/OTu3Tvx/Fgg"
            "EHIJ7PJTkpIH9h/Qr09f2NZDUSiGgYsb1E58nBDlz78GhC6O9B82brp//z4wJwmo0KNbN4oH4JyV"
            "KXCi2QMLAU4c2L//4sWLxF7xk0CoycmhkwudnHdu30H+dOP6jXatWhPXlFDIL2fOUPmp498EQqKp"
            "CCMJlEUcvcKwE2EgtJywMRzFhYV/BwiJdnFZSQn5E2nAwwcPeWwO9I64MyXLiRkdQ2Ja1pTaGQMh"
            "ESfC4IQGBlEpLCi4Y7v2u3ftNq5lTnmFo60hYAMMF7TE38eHw2COHjny8u+Xa63LBAjhQiIMhyqo"
            "ugI7dOwcGwdrSVvbY4066pIQkkp/PvUzQWLiauhOdVPG/6vjfwwI/3WVUQDCzVoMhGZtU9265I2Y"
            "88Pzp/it5/v3y7ccbt93uFmIEIoFLLTRh51oGC72+rKwrHzTM+ILuAq/F/lYdeD4+Z6qcrP26QK8"
            "Hd919DzgH4E6EqKwV075vmPnb9++v2zLkcik8Q1CRQB7RPpnW4MJcXAL5AH1i55lJeWbT5z/Q/fK"
            "8/1brJJqOG7dfzp/4+FY0TSoolm0zKF2UaHq00CokxACEIptY8SWkVkNA5IbByat3al7Oa2fNaIZ"
            "NbNhYIoDWwo54Xzo9O+wlZ+xbFfBtyslo74bnDsjvaSiZNrK0xeua41/lXEhmSMWNghMd+SpaGyF"
            "I0/ZKCRzcD6ySzYSpqHzu8qq7YfOjCxfpxy/uGzW2rlr9p77/ebFa3edODK7GIkDBsLmtamMZg2b"
            "b9ayX3CfwjsPnmip0Kv4/Psf99p0g11+BrYwlGEgVNC5yiYhmbNX7vnj7qP56/aVzVqnGLc4uWB2"
            "csEszcSlOw6dMe4C2ZTMX3+waYjAiat04KiceDlNQgSKCSu0NbZHN+48zJ+8optoQmzqyO6i8fPW"
            "7KH+pAPCI+dto0UObLkDR0nn5VhGSSIGjySyO+rugPOBExdU4xYPzZspGj5/+8EzWqMni16KUmEe"
            "KnQGBOKqmkeIVmw7Ro0n2Zb98tuN9KLyhPTRnTPGDlLPEI9YOG3JToAHXtoEq0iJs5HHlzpNB/Xu"
            "ZFyQcE/YI3u61ojS4bzt4BnL4GRnRlbjDoNh3KjvqZ7aRUucudBCNU45tJjsll/mPXqq29mTjgDA"
            "t+teAPPrylNZhgm/Fk81Xuq6lxHjAWIz3PgqR6bMnae4/Md90k0yJK9evw3oXWIXBXWpXLk5FqHC"
            "8lX7/rj7eMG6AyNmr1eNX5paWJFWWJE3aRnyAlVjZr/bcKg5DjpPfMZYhGRVrNqrNQXCDU0CAQiB"
            "u1S1AiGw3+Ubd39H6R5JV/94cP7KneB+w+1j5G58NfQ9pO+wmkAIAExnKVx4ao/YXPMgQfH0ddQ8"
            "kjyXb9ybuWTH6PIN0xZt27z31FPdixjjQdL2kM60wdH/3ACi4vKtwrOnYhtak8UJd5Nq3JLuwold"
            "M8b1yZ6ybucx6g7SA+Fhm3DkEsYN+wjFQsI/b0OIVUbdeBrPuLxmwUJjIKR6NGvpjjEVG6Yt3r55"
            "789Pn7+u2aOe2bPsolGAe6/4QpsoeVLBd9TlJCfcMuPnbvpGPLlL2tj+ihnCYd8XT1t77OzV0XO3"
            "WodnewHQ8uuUED5/+frS9TvQDJIuXb93897jrQfOesXmuXM1PnH5NuHSrGGLqVuJNO3Fqzc7Dv4y"
            "Z+WeaYt2rN1xfP/x32B4F6w7ZBsu8YvPtQgS1AaE2yyChT7xhcjDTVwBLVI2MGcO1QwKCN25ak9+"
            "nhfy6lnkxNJEDBj7/IXhrRb14efz16cv3pE/eeWwmevnrNq3/8RvcO8kZEylM1TeSEKY5xidXdOG"
            "8P6j55eu3dV39v5v1+7df/xiwnc7Udx5AoTh2fNWG4BQLyHcbBUq9cOyTb+EErtI5aTvdtZcUbDg"
            "p36/LW/iimmLtl/H2uAmr65GVWy1i1JhsCz2SyimRciNgZBkPvvbH9MWbhs3Z1PF8t27fjr7GtnZ"
            "6g5S1tt3lazEiW4cxIT+XcqgMQXfbqDmlBTy6MmLkqmre4mnfJ05MVFdLh2xeHTF5tMXb6rGr6ZF"
            "KVt0waE7EA3+F6qMGmwI/0+AkEQhhx08EQYSeQvsdDvHxQMFbcZAWJCbRzxeAC5amjdbtwZxoEIm"
            "W4ot0CaMGw87e+CBJYsXa/UIQarYvm0bFE63oxEJofEKSR6a5OLoyOdwi/ILoChxllCQnvH1l19+"
            "plOZmkAIhLBmFRLLDOo/gPgRhaKOHT1K8t+4caNDm7bwDWQGVoHLf/75Z8JydQGhTtvl1q2A9h2I"
            "N1Qei020ZD/imzclKckJB7h3tLMbVlqmrUFW/zIQmoQ0JBEI4QMUDtMHNK7LbEcjwti/IiH09XOi"
            "Ocizs6nxIRn2791HqR+T4A27du6sa+6MgZBqNonnQRL8F+BtC1ZXNl5O382fTzSWAXThwjZftISc"
            "MC/wTasWX3w7ZcrnAKFJdaSE8OCQ16/qFHKQoy4g1Md13O7l5tHCB0UxCerYqR6ntf/m8T8NhP9Z"
            "pzLVgHDxDsA8wDDAv4CBI9fvPvEBh+x88vhp8cz1DvychmEiYCqANKI+Cpm/ls28/sddbeU7SLdu"
            "3RtStKA5Q2YelQ0lFOJNz/CKTQCE8GXDMHFM6kTYvX0w0jB++eL51CU7fb8uQgEqOEjn09gfKfy3"
            "aVS2c0JeScVm+MHWkt/CD1Uf3r5BNFj5TqdkVWWkYfXhw5ofT4cPHd8Ie7LB4TE+14bwPWVD2Kq/"
            "PUPauLz4AAAgAElEQVRkE51lzxK37p43KG/27qO/amvsRR49fenfRWUZIaAxJVCaPUvilaC0iRI0"
            "7DDErO0As/YDG3Yc1LDDALPWfa2Ch/yw94RW/2tNfvinLdvVIDDNEXkBldsxZXYMyaGfiUDJ8IsO"
            "+62e0omNOw5q0La/WbsBZm0HNuo01IEl7ti7GFCQxpTas7LrsiEEcLIKSYb9rtbAJPhRfu9xcJ+i"
            "psGpdHY2uhz5IEURLBzYSluG1Cshx4UjbRKQbNZ2UIP2gxp1HNyk4+CG7fo3bNt/+Kw1VOGUIMsB"
            "SmDJkJiRJYPPREVTl4coN9683+7rHLNWfc0DEi1Dksx8e/bKnkR1UI9Jv9pGiRxxOYCXzcKzlm45"
            "Sg0UybPv+Hnb0GSz1v2awKi2G9C4/cAxczZQtZA8a3eftIoUOnEVjmw5LUay5/gFk/H8SjDGzKeH"
            "dWiKRdDQpgFJDTskNeqUQmdl09lyR7YSORrlfcqGUO9OBrCzWSi084i2ugPD9OK5jTsmurDENmHp"
            "7btpnr/UvWCjoJGdPBbgE7kbRViocuYorMMFG/ec0lbfzA3Imd0sJNODr2oSmDaifKPWaN+pe82/"
            "9ahlmMAVbupIIWPIKEo4QAr5EcgzSujEVjhxlE5wKzEV/l1ygRstgtMatU9s0nGoeUCSRcDQJh0G"
            "N+0waOTstSYze/L8dReu0omtwnHn1RYhgoqVNYBwxvomgRkICLmmEkK9LeVdb67IlSHwYIvd2VIP"
            "rsyDp/SMzYEyUWAJgOHo7OC+ZaZA+PCZb7yazpS7ctUAUU0DM4umGYCQQEX5il1mvr0btxts3jHR"
            "olMiP3mksQxct30v/8E8SOiJgDDPia1u+VURJWU1dPPXqz48aeP2Q6yDU2lh6Wa+vWSjvtdWx7lF"
            "m36yDkcuYdxi892MzAj/vA2hGksIAQizSqavN+oROles2G3m16dp+yEWnZIsA5Lik0fV7NHI8s3N"
            "gkUoVCDwbZhUMW6l1oRdNxyAQqAjdqGZ1sHplkHp5gEZTgy5V2yuG1fjAWCMJYQmNoS62BhbDzuG"
            "pftws725Mm+u3Jun8onTeMflufNyPXi53ggIJUIjICTX7jz4i21wqmVgikWnVFq40Jkha9+9uH33"
            "EheWyic21yIQgHC7tiYQBgl94ws8UWMK7CJlg2oDQg9OjifSGS7wiS+iRSlmrdir1Rq/LEPZhs9c"
            "5xghsApMax4AKd06RODMkEf0H+WXUODBy/OKzfeOzXOIltUUnmvGLbULSfPnK3y5Cl++yi9W4x+f"
            "D+zhxc/3ii2AD7bh0nk1jGZHzN6EgDCh2DuuyI2bF9R75N2Hz0xW1J4j5/14csuAVEjNO6W0ilcd"
            "OW1QE9C/IHvcvtswD16+T1yxH/SuNvgsnrrKzKevVUCKbVCaXVDaENWsj0YcTkqTj1lJi1D4dS5t"
            "0aXUJkw2bv52o0Kwg7TZG8y8+7pGiegRQvswoW2o0DpE5MFR+8UX+sQV+caXYlc0/502hMYSwvz/"
            "EwkhCb0gFYnnVsyZPHESpLGjRwd1CgBE+WEjcnRRUlTEZjDhw6mTJxlRUW9QFK2XkGH9WvS8ooBw"
            "4ffokbJ1y5bUpGRFtozA2Pp16+2QUBEB4etXryeOGz9q+AhI4SEhLo70tq1aA9GFBgaFBARC6tC2"
            "7V+WELrRnVeuQG9ply1d6ky8QdoCp5WS/HPKy6FAghNlWJf12LFj9QMh2ejPmjGTunDsqNGkPaRJ"
            "a1avdqM7QdVebm6RIWFPnjwxKeFfBUK9qK1j23aAvnAmxmwktIZtc8te3XtAj3LV6r8pIXS0peXn"
            "5mr1Kpqkv9liCcEtKuahSJClrYOITICQhIIEtKYScJetpdWmjRuN20Y6e/n33/PUmrCgYA9nVyIC"
            "JXFQgNhpVtbEZLEm1RsDIZGjwshQddGsbWDEXv1VICQNO/zTT3765dSxXXvitPb/A2F9x3+Hyqhe"
            "QhgqhG+Ao4DEmkRIeuWUn9MFnNCevXitt7qiaZQUWKtRuJiROnEHktKg7cF7LFZes+O4WZAAoMsR"
            "g1yc8NsukukNw1Hsita9y+as3vv2Dd4cV1Wu3n5MPHYZVfKt2/dzJq+y56sbhYntgDaxS1L4b4Mw"
            "EaDdOZ2HyY9Im/ndW+27NygBgpqoMRhFk3/zrlI+aU3jCCnSU62uPloPEJJrr99+sPvIuQOnftt3"
            "4uK5y7eoV7PGtRFZ4pLNh5oEp9gzJXZMqR0zG4jOJkpozxQ5MkW06EzbyHTLsBSLkCTbiBSzlr2+"
            "Fo7VVqegH/adBpgEJHPkyM3DBLGZE4zvkw863aflZl/0cYjJdGRkQbGOLDEybowRWUYKaSyZvS6Z"
            "ehkl5YhHzF+367i2Oq48fPKCMWRY44AkOluKeJKNHNs4YCNGIEMHltw2RmzPEDuzJQ6MLFq0wCYi"
            "3TIk2TY8xTxgsBsr8/Ez3QseUsXt+0/c+HK7GIkTR2EZIYxO1GOJrgvoX0HZPLM2A1zYIoeYLCem"
            "ELAwuWC2yVAgIIwWAeLSYWpipL6dNQ8eP9dXhAe8qiqiX4F5p8HODIFjNOxxM23DUuzCUi5cuaXV"
            "icUw695/4hGrpMVI6WyZTaRoC7a7M4bkKd9vsQgY0iwwiR6T5cqWuHCyndkyByYAoYLOQS5tMBDW"
            "G4eQuJOBRRWT3b5nMbVrJw0A5PCNk9tFCugxYieGyCIoefX2o9rqePDtEuJaRichdMXIl41d1Bq/"
            "Lxgz94fGnVJcOHLgut1HzlN/pYbu6s0H7lyZMysbLhePXKStzgZF365pEpDmxlMBDTpzVFj1VA6d"
            "dePInZkSaJtDVJZtWLp9ZIZl4FBPdlaNmX3qHZfjyFSguPO8HIvgzNokhOubBKS78wHtTCWEusD0"
            "1+64xAhoERmO0SLHGIkjQ0pnyuhsJdAgil7IUyMg7FMDCJGEMAcDYQ5AFABhcTUgxKLpNXutgpI9"
            "OVI3ptiNKWnYdtCcVbupv5Kcq3ecsIvMduNr3GPzrCOkA9VzTO5iOHqKJjXrlOzFlbkypZ4cmUVA"
            "csHkldpagBA5U9UDYVEtKqOfC4QaYLNmQbUAIfQIoNSbm43ImSVp3G7InJU/mvZo50lalAzQ1BMJ"
            "PCXGCpx6Dc+bHb5SWwakOMWIvHgKL77Ki5/jzlUjGozNJ0nnZbQGEC7ffNg6MNWVIXZhSF2YcheW"
            "0pWtMsZImzCJsGyRtjoQbj/wCz1CAEMH1XlwFO6cHGemyoWd48nXeAMQBmXVlBCOm78NRoACQlpk"
            "9sCcCqPnHgbC05fdOSoPrGELbWjRpZByDWV4fu752Sow1Y2Z7c1TevNzvPlqb77Gk6dxYUHt+cgn"
            "DUr5jggIDRJCnXR99GKLDsmeLJkbU+bGUrizc9w5uVggWaADwrDagHDWJqsQMdCUX0KRbYRcMmq5"
            "1vDbgRoPT7+Y/sOAxn35Sh+e0o+vsgnO4CeNNmY58jmteBEtUgk06BtfaFcdCHXqqTPX2wSmt4jL"
            "8eYqvODXKih92/4zJo+I6Uv22IbL/BOKWyQU24Rlj8AqtcZP+33HLrTgyWkhmR5MmT9Ab1weJG9+"
            "nndsoS/Sei2jghn+fxtCbR02hLA5tmluCfAGZ0gkzh7xfDh+7Fiaje3pn3+G3fmeH5Hay6oVK2FL"
            "vW0rsuydNGGiE83By92dAOGUSZMtGjeF3TmHwURmt+/fB3XsNKj/AC22IYS6YK8PO3LY00eGhJKo"
            "D8jVJ5zxh78DhERC+OzZs8jQMBLsPiI0jKgCIh7AkeV8PTxPnkC3yfHjx+sBQvLh3bt3PBabKEMC"
            "WfE53H69egNZ9e7Rs0/Pb77q0pVInIh54dLF+KethjvNf9fLqApY/c7t2w/u37944eKsGTPafNES"
            "xgHKt25mMWrEyInjx8P4/02VUcopC/nr1MlT7G1tKdkgkTkD7E0YN77WwqvZEPq3gEvYDOaCefOo"
            "9N38+XPKK65evWoyI5QAEPq4dcuWkcOGR4dHEGc/BEQB2MiapCo1AULiBCg5cejypcvmzZmDqps/"
            "Dz4sW7r0kz5m6ncqc/v2bWJDSF6v7MJRTz58yoYQe50wHPVn/gvH/w4Q1hmH8D8VmN4YCCdjIKTx"
            "sV4o9iXTMEzkGKsumbXh6WOdXujaXceB9GYs301cJFdVvhs9b8u0ZUgXa8OPJ61YCuKBxoGvtojO"
            "NuuY4d4lr2z2xif4cpjon05d7CafaR6NRIhO8bl509bdu6cLfnjmwtUBefOaxciaRErh8sYRkoGF"
            "3z3FakLILwsQIFITxbLBD1X4wYFNXWuMJ3xHfg4r1h2yZilNmLAeldG6DoKa1H+JARJsW9v1yLMI"
            "ywAeQzTIzHZgy+wY4gYByc1C05zZYneepH2P3KhBpQHf5LqxBW2+lL3SR7cnpe09fsEmOovGkNA5"
            "crNOycUzDbtesuW6/Mc9N04WIBktOssuRkhjiCAzjYkkcgByNLbCEIewOhCSD3cePKFIiZyfPHvJ"
            "TR7RsGMinSWhMaT22M2pA4pZr3bAFoAOHIUjW9Y0JKNJEHRB4s6VftFVFT2oLLRvoV+8lB6ddvj0"
            "b1qdvRwq8OmL115xSptokTNX0TAwTTpmqbY6XV+79cCNLbKLzHCIEdrHCOlMcaNOQxPzagPCKCGA"
            "mRNXbh4i6KtCDjP1VIl/pE9csAxKpEWk4ZQOZ4eo9EbtBpSvICpbVRQ3MhJHW4ZnuXAVjQLSpi0h"
            "KoLVHigHT17sq/jWPkrQKCDFNlLoxJEDzdI5SjqXAGG9cQj5BAg1LnxNk+DM3Mmmu+p5a/aaByQ7"
            "sySOTIkzW2oemDJEM4vqrP6Nw0PveDWwN5IQIp8xKttIUWDvIvLqgRLx7Tp8jhaZRYsStuyaQ9Ga"
            "7omhW+cfOMmjbSMEzUPS56/VGfjpN6aV0YOG20QIXbgqUgsAGHy2DBM2D8l058q9+Iq2X2uYQ0ZE"
            "9CtpmSB3ZQiOnL5kMrN+nTUODBk0D4CwWUhm+ao92hpA2BiAkFcPEN51Z4vpMSIXFoC3woWjdAEA"
            "htHj5cIZEi0GgLC0FhvCuBw6gwChpmlgRvG0tdoa+GQRmOLKkjozpG7sbAA5Qel8au2R86Gff3dm"
            "yV2AK+JyzYOygEOoy0mGY79coUdlucBMMWWQ3Dnypp1S8ibVAMKNnwGEn6UySoAwr1mwoGR6TcTd"
            "ZxmY5s7OdmVme7BlzQNTs0oXmPTop58vu7KVbly1Fz+XFi3rnDnZJG4EnO/cfzpsxtq2X+ZaBQOK"
            "S2B2PGNzPfh5HrEFJNUFhCu2HAGS8eDI3TlKdxh5HsWQ5CoiITQFwh0Hf6FHidzZcJUKyBMQzjMW"
            "YAwYMs8nLs8iuBYgHD9vq0UgBYT5dQDhFQSEWFXVIVrRK9vIg64+W0LqWPtwoTdP5cHJ8UD6pbmY"
            "A6HMQhyEEFlaeiMgNIlDiM4545ZZBWX48HM8eYhdsSyxEEcvhGuJymhdQCjySyjwiSuwj1Is2WTQ"
            "DtBN9NZjtNAsaJInV+XJQcmbq3SMEG6twXIVqw7YRcj8kHpqgV31ukhRw2eutwzM8IXmcZRwtgpM"
            "H4bXjHGetTtPOcUofeMK/ROK7CPlGcULqz0l8L+Xb9zTTFjRpnOeXajYmQGZ83Glxb6AW50pIKwl"
            "MP2/aEMI55HVwk7UE4fw35UQwka2U/v24SEhgR06hgYFBQcEtvD1s7ZoThBrWGmpmZlZcWGRVq9T"
            "mjQksbGZGfGuOby0rFmjxtYWluWz0e/dlEmTrMybudDpiYMH4/yVHdq0+bJzFy0GwjYtW1qaNxs0"
            "YCBM2727d9u1ag0oBaTk4YzC0//ZwPSmQIg9eWixpZyjLQ1gw9neAQjh3NmzlFppz27dyc775IkT"
            "9QChPv7eJkrPkIihoJFQFElEPEX4B7Ck59fdTCRC/76X0ZLiYqrx5K+bNm2CNkMLER6HhJYWlxBm"
            "+wteRklPkWUplhu/xwEDAcy8XN2JhipUwYyOJjJJ+IZySlS/l1HjwPR1HR/1ESlMwOnNmzfls2YR"
            "FVCYC1hLKUOHamuAPQWEJoHp/9Ro1w+Er169Cg8J9cRLAqajqKCw/oHV1mbx+I8f/wNACL9wXwrH"
            "mnXoj1VGAQgFNWwI/+Mqo5MXG4BQ91f4KWKjgBMd+o1Yte2oTtuz8h1q88cPOw+d4WVNMWuTKhiJ"
            "XgJ9v/EgUByKAMHLaRQusWYrM0cuvnwNB9P7WHnz1r2+mjkAeySiPWQDWjMLFbXoWTJ1ya6XJKpJ"
            "VdXOQ7/wBFPMAgT98+dryeay8r0eBd/Ac1Rb3WlFrST3Ue/VbdGWY+ZRMuzM5tM2hNS1H6gDb48/"
            "4lo+End2xFz4yYv4jLGNg5IdWFIiHrRnI0ej9kxpSsn8jXt/vn7nIdDX23fIId/bd5XPXrw+f+XW"
            "e4PfJ1Td4TOX7WKA9ESOLGnTkLSlWw5rq+vFzVy206z9IEemyC5GDBxIY0GS0Vi6QPPAcjTkCUZR"
            "MzA9dRiz2YPHzzlJwxt3SqQzASzFiCqB27k5mAZR3AtHrso6WmIZIfxKMnXRpp8uXb8L3Xz15h32"
            "3Vf56vU7AFQ9meg1fl+/bfmVxjpS6MyVm3VMHjZ7A9V4nUbZT2etw9MABe0ZEkh0trRxQPLQ/HKq"
            "VQYgjMyC0QOwbBCQmjd1NTUUesnbk73Hzx84eZGk/ScvAtf9eOTc9VsPtEaPOS2ys5rWLDQTgNA8"
            "JKO7ZGrNySUfzvz2h2byylZf5TUNzrRjZDvzkFMcBIQkNn3tskHiXzSXztUAQDowso+evWJSfpfM"
            "cY07DrWPFtpFCWnRQuuwDHeO+MYdgx0R2UwnF85HTkT5ajoHaE3lxJJB9/ccu0AGhJqvLzqrGnUY"
            "OkitQ0rjpU7GLWfC8oYdhrqxJWcv3SQtIY05fvaqY4yIzpJh8SBiMHuGwjZK2kM6ffEPhy9dv2c0"
            "s1U1ZhaV/+r127bdCuwZ2S4ECINrA8IZ65sEZLjzVPUAoRtb6hgjcUE0mOPCQ+HmXXj5ONh9nitQ"
            "TT1AyKSAsFYJ4b7mQWlI2smSubHlVqGZvbK/NVn2P1+44cpRuLCV7ny1VZho4cafTAr5fv0Bq+B0"
            "KAQ1j6304KnMA9LyCOTXB4R/2YZQjVguNhepjNbs0dr9lsGwWhSubAUgmU2YoE/2NJMenb6InLW4"
            "ctCwoDNXde73W8Z3PTWGML9zVu2LT5voECOjRcu94gjXFdYLhEftwrKgTHeeGgEkMrwsRMEAYyGh"
            "qzAQLtSaAuFZerQErkIAyc+DnJ7Y5w1U5BNfgCWENVVGtxqpjObbRWYPqhUIYUa4ap/4fKtQSe7E"
            "1VSNJMOZi3+4MaTubIUnNjXEMSowByL/q8Ve8SXYGw0JTC+rDQiXAzD78HO9+HlYkIjcmXqjq9CF"
            "vgm1A+FwAoTxBdBTL57m+LlrVHtI2xSjl6JiYzVeXCgZJd9YjRUs4KlrDEsC59xz9KI7O8ebn18T"
            "CHXwOXODVZDANzbXi6v25WtsQ4TZ2CbfuLrtB8+5MnOgjz5xhe5sTUDPYfewCiv1RDJoLtx/Om3J"
            "bubgcY7RSheWxr8zcorjq/coQ4Dwv0Nl9FNxCP/FwPSIDRwc05JTYIP7/PnzVy9fPbh/PyE2DgiQ"
            "+Oo4euSIUCCIDA179BA94a9cvtyuVSv465XL6Hfh+LHjWRkZuWo1cUZ65vSZuRVz1q1de/8+MvY+"
            "uH9/757fADZo8fZ9xbLl8+bO3bP7R/jv0ydPOExWrx49e3z9dZeEhDge//MD09cKhGv1QPjToUPE"
            "IyVszaHlpcXFjpg9oMy5c3Rq2ydPnvykhHDIwEHEHBGyAb1Q/lpIIt9QfmV8PDygXq3RRv//IOwE"
            "5hDK+yWc3719x2WygAmhFzBQY0ePDg0KJiP254AQx/YgBnJkZrUoNv2FDm3bAQci9nZw7PHV1y+e"
            "vxgycCCMMxH9dWzb7uJ5gz2L8bBUi0OoD0xPHfXAErnZCRmSbCOHjwAUBBKDGmMiInV2nkbvpo29"
            "jOLA9MiWB7J9TnXUURcQUr1Tq3JIlEtYw5CN3A51jS35/sjhw4sXLlq5YsXyZctWr1pFYh/8g8d/"
            "ORBikn7zjp86okHHAdbVgFCEgJBBaFCGkykN/pNAuMgUCCHZchEWNotGGqTdFDN/vXQDJvrGzbuJ"
            "xQsA7QDwbDlKj675qsmrAgaOsmTK8ZeSr+Wzjpz6jWjzYUPEj8+fPhs1d7N/j5KG4RIsmVQDngEW"
            "NmfIG4SJQhPHrt5+rAr70n339s2yLYdv3ceuUNAe54MWSkA6ogZCuHrz/siKtYs27SPRFGoZVcoY"
            "o3wz0KkDX/NJG0JyfNBHNdDHJTNysY6P3UfPBfcpbIhs+YgnUqR42RTpfE68cPWutjqf1HrogfB3"
            "OyT6Q3hmHSX48ahBLZD84GkmLTdrP9gR6XZm09gyhH8cKkwFScj2D9CupoTQuFPkw4Urt7xjpc1D"
            "Ux2ZYntKWRTRoMaBowYcsoqWtupWsPXg2T+xaF+/w0CY5cyRNQxImbaUSOSM99z7mwYm01lSeyTY"
            "RKZ6TYJShxZUaGsCYVSWAwAhB8AyZeScTdoakr3PbFJiwZymwelOHBlgtnVE5rpdyHwIRz40yE6p"
            "CX309OX05T/6f5lvGYFiGDpySTSI2gJOGLmTcYZFGy7qLJhkGGo9//gnyLz52b7xCp94JST/BIVj"
            "VCYMgrb6dnDT3tNWESJnLLtz4ihdearGAamF03S7Rmr19JJONfui3wwsftcFCr9w7fKNu1q90vLq"
            "Hccad0iM7F9C7gIkXdSpm25qGpjmivVFkWYmQ962R+m2Q+c+bxi1ZGYREMYAECrdeKpmwRl1AmE9"
            "KqPX77pzZXRmNqZBDUCgC/b5qYt3z8+1j6ldZfQzgHB/8+AMN64KQM6Vq7QOE36pnw6q9vNXbrtz"
            "lc4sBWSjRUk27TlNjb/eJGxDk04pHtB+bNAIwGYemJE/2fAyom4g/Ms2hERCWBsQrtlvFSIAKHVD"
            "qpIq2wjRV1mTa/bIi4daC/TlydfYhIvTinVSxEp9ZAIU3kO/guDLH49e/Eo03TZK7oEEd/VKCLce"
            "tYNu8jU6cSIOeEgSfPaOL6wdCA+ddYqRemDNUs+4At0lcZAKERDW7mUUgPAzJIR6IGweLBo7zxCI"
            "Uqcvuvc0PUriySM0WIg4MKEEBy1EcQu9O5fiYBXF3igwvXxNLUC4AtDOBxochwSD+KoynAgQFtUO"
            "hLM3WgUDEOZDf/3i8i7fuE+NBrlrkvPm0sLFPrH5XnwkcvTiF/jG5dmGiqRGLKd/W/EHVnDN9QMk"
            "DpPUJo3caB0s9MORM6AQuzBxZsn3WqOdHJz3HsNUGZvvE1voF1doH5ldOk3nV8YQJUj/kkiLfQ79"
            "sPcXfuoUR+xiFAdCHO6HDQj/TwPTj66uMjrCxIbQl5/vw9UB4TAj0fo/ftSUECIo6tDx2rVrL168"
            "ePLkya2bt77q2pUElyc719OnT7s6OS/FbmPGjBoFO++HRn/du2fPpd8uaatvCeAzoGNoYNCkCSho"
            "oTHX6RR5njxJ4PEzUlJThiYBevXv04fy2/l3gJBwwtddv3TFOqJE0ZG4f2zbqvWtW8S3fH1ASMbn"
            "1MmTxg5aRgwbNnnSpInjx0+cMIGkSRMnjhk9GpiHwKcTzV4pk2v/HhB+NNKNfPfuHcDegL79koYk"
            "9vmm14+7d2urk1VNICzWxyHUGoHTN917uNGdCAvNmzNHnp1N2gN4s2QRmtC3b99W1nGQ6Pbr1q7z"
            "cHEjRokqhYLU/vzZs85xcS6OdMJ+ndq1J2FIbt68GRESCp1thaV/CfxYEtudWhs1gZAiYerNzsxp"
            "03v16Jk8NKlnt+4Lv/t+xfLlGWlpSrkcGg+LiiqN9BQ4nIgl8VR2Iu5zagdCHJh+6uTJWj02k+4L"
            "MwX9evce1L//gL59T/982mScyVEfEOrvAgBvIkd1pdN7dutGOk50az/qj0ocZkeLzA4Pt26BhhSu"
            "smlu+WXnzlXIV0gl1bC/f/wPAOHDpy+iBhU36jTIGmgwMgOrjAoxEIr/ChD+dRtCkQkQkoTiv8dq"
            "zFqnFM9E6hAL1h8wC8hEESNIUEGWokGIEFCwYbg4Imncht0nCdq9fvVy0sLtvdUVhxEcom7evftQ"
            "PXUNcGDDMDEQmj2KOYGkheZRSFO0i3T6gRMXEPsB/FW9/4hd2lQbK3yG3+C49OFmrXuates1eg76"
            "haj1TcZH/dh2V5Y3jpASJvwkENZ6wLP0t2t3vlu/75vsyebBSeYhqch0EGgQOWWRm4dnMZPHPsdh"
            "68mPMRV4qtbDCAgFdjFCW2DCGOGxcwZxE9lkpJfMbdgpyZGNBYNInxNDIAwakekhsZ5KJyGsDQi1"
            "NZjw/JVbfp2VkNkRexalxIOOMIlMmUd8zhksaCIwTKSsdd2BBgnhl2oMhNlmHZMmfE/2bVVa/W/2"
            "ym1HLEIAQaUOLJkDW+bEkTcOSqsHCAHkGgWkzli2W2vkQ4VkBiKqmQyRlz98IAEq0koWNAkCIFQ4"
            "smR2UbDBlW07SLxoftT1SLdvNkR7u377ET99kmWklMS1NwgJY6sbExr0RdXmIYK5awyYRw6gsucv"
            "Xz9/+QZWAjrj9OzF62oeAvFwvn1XGTZghE2U1AUxIY4bESGMGjSccgxD+j5s5tpG7Qb98tsfpHA4"
            "l01fNeV7dKuSnHcePLEOThENW6DViRZ1YxWXPt4qLMuFowTmdGApfTrnkXALnzez6KyTEMZIMRAq"
            "6wbCem0IERDK6Sy5q54G3WILXUni5+slhH8JCNfuR7qvPBUSYHJzbMJFXTInUpZaOny6fNudo3Bi"
            "ygEI7SLFxrEHSCGTv99mHpiOgJCrhuQZq2kWlJk/xUiY8/kqo38bCBesPWAVmkWkiB48NTS4q2CS"
            "bkKNgNCTryJA6M7TwNkuQpQ3eTVRnP9otPv/qDOo1t37Iyq2ODJUnnHYhlDnZXSl1hQIoZFiwDxK"
            "sB0AACAASURBVA+kJkpoEDFVKZxRBPk6gfCcU4wMa4oWIA6MByorhTMCwoRCAMJJ3xlJCCuJyug2"
            "i8AsHyMgHKSuGwhj86xDxbOWm669BWv324QKvWOJbLDYy0B0ZT4JKMQfwsJ6gHD8SijWNx7TIGoz"
            "CjRPYs0jIIyvHQiRyihAWny+O1v9RULerXum3jL6yWYClQFnomLjkO8Z4F5ahDStcAGVk4zeb9fu"
            "+sXmeiKwrAsIdfaK3vx8P1xIZnENIMRiRh8kISwCJvTh5zvHyKcu2km1p0q/CPBrAr2B/dv3ynFr"
            "nFl5/jiAoR8RytWmNfpvqIzqzlhltAviUiOVUSMg5Gn8+BrbIEEu9mJdT7iUv3PUtCEkSncd2rQN"
            "6tipU/sO7du0BcBIHDQ4My0dkiA9A4AEsIcRGSURisJDQmHXO2jAgIzUtPSUVMgAu+0+vXqR/6Ym"
            "JScnDk1NToENPaCCi4Mjj83JSEtPS06BBH+CBAQIOeED1AsbYsAGoAJIn68ySsRfsOE2AKGT86qV"
            "6GYnRoNzKiqcaA6UE0tSYGZ6hlYfUf24sVOZDh0fP3pMioXjHVaGLC4opCKnA14aA4bxMMokUkdb"
            "JBeCcjq16wA4pNXDBiln1gydsRwwUue4eBJej/rrt1OmAqKQWvr36UtKpkAlLCgY/grAYNXM4nsc"
            "OsJ4NEhHcANqB0ItxjOYVhKtAchz8cJF58+fd3dyIYE6ALSo/LUe5E9bNm/2ctPphZ7BATbgAIii"
            "29kTTVoYfGI1RzoFhObl6gYD0tK/haMdLSsjkxRFZHEkDzBVrUBIRlWQkdG8qTlUat6ocWF+PvSx"
            "sVkDN2cX+HLxokVkcKqQqw1U1Pq160iPYCXAEn3zupqEkAQzrAcIX796BWse2gktcbJ3+BGLr42t"
            "EDHOaRMHD3HRA+H0qd9SnTWesoH9+hORMpa6OwDN3rhxw2TZkM/btm5t16q1l6vOKY6nm9u+vfu0"
            "dezw//LxPwCEN+89CuilaRw42DoyzciG8K9KCP9pIMQuQHPMQoTDK5D286JNh5rFyGg84g5USWLH"
            "O8Vrpi/bTTzHVFW+X7PzeNjQsYCIwH5AjBnDF1++qnsL9cuFa/3z5hJzQXtMlTTkSyanSVR2gzDx"
            "chwtoPL1K6wm+k5r/O5H92t624mZ6s7NsAob8rV4TI0RNRy6MMdnr9pxAFwRfNajMkpFJvhu3d5l"
            "mw/NW/Pjt0u2lc5cM7Rgdmi/QoeYjEYdBzfqlGgXk0VjirFsMNsOqAN4LEq09wR6CUQphZIDNkAP"
            "njy/evP+wycvjOVdOiA8rQNCWozIJjJzr5FLTPKDlz1mMQ5yCDQo1wkGuWoSZdGep3EgQEhJCPdW"
            "A0Ko4q0ubIPuliP7oUM/X6KzJSjoH1cJBWImzKHzchqHCMZ9t01bPQQiKefJ81fXbj24ff/J23fv"
            "qftXLyHEQBghcGbLzNoPLZmBbL0qjTZeB05etInIdMBA6Ig0GOVNagfCc7ZRAgemxJkrb9Axedz8"
            "zdq/LCEsnNs0JNOZq3REcfxktpFCu0hBSmHFwVO/Uc8Uam/0AVuiwoe7D58F9h1uEyNzRhEC83RC"
            "QhPFUb07GXumzL9LrrHLyj916JxGlG9CQQL5aqQyylHSWTL7GOFR/Wokw7J576mA7hq99A9VFJs0"
            "fKByqtZoollDyipW7tYabf7OX77lzJI6MrOdkexRbREmmoClNLXP7O0Htx/AzFYazaxWS0kIo6VA"
            "lVhCWJfKKLYhrBF2gtr1unMVdJYCA2E+kg3GFaGEgfDvSghDBO58NYqKwVND7V0za+ATACEXAaE7"
            "V9U8VDB/ranfjhVbj9qEZWEbSOThxkMHhJ+UEP59G8I6gDBM6InzQK/tIqVfCibXBYQooiAvD5gQ"
            "ugZcxEsau3jToWcvXpMZNBYKVeE3JlqkbrDWNhLpjuqAcFJtEsIIiSdGLA8K7RJKvRDd1SMhPOfE"
            "kCP7vVgkHgSARAkYMr7YJ77IIkg4ET9VyF1G7ujJ3++wCAYgRHaGRGW0dqcyCAhzfOPymgcJxxvZ"
            "f+r862w/7hAl9SZuYAD8UKR4BHXeKJWhM2qJDgirq4yi8nPGr7AORbEQvfG1hAaBJDFS1ikhpIAQ"
            "SNU3VnMBezv7YIR5omGL7MKlwJOYM4tRVIn4IptQiXq8wR+szj/zuWte3BwvrromEFbp6wJk9Y8H"
            "0isALLSPyK4LCLHKaDFiwrhCL57GMVLaUzxt4+5Tr9++o4b0o34lVenNjJMKFtKZuS26AnrVbkb4"
            "L9kQ1gmEw/RAWEhURv34alpIVvYIg/esf/yoFQiJgMXO0srW0opmZQ37ZosmTS0aN4EEm3Ir82aw"
            "V4YvYUcOW2dgCdvmVjbNLW1xfqI/SbO2gQxUgvywQff18najO8FfgcdIcjJy8EgYg0pUYz5TQvjm"
            "zZuo8AgKCInXU5IfgVDbtj7unsTZCTF+27gBCZYJR/1y5kwLX1/i1SaoUwAVJ0D3cL53r1O79jAg"
            "X/j5Q0tIHL/3794ZS89IOYd/+kmneYuDts+eOUurxwkTCSE04MuEzhSt1ZQQ9u3dm2hCElYBUGEz"
            "mN5uHuSvNdU7dS99lEoHAxAWkTykI/fv3R88YCAlKYWBImLGgrx8YEjAb4D5uRVzAEqhGbWmaVO+"
            "rSivkIjEcDmUP2PadFL1lEmTSPgK7LTTZcH8+VTbyHn1ylVUFETIOQ0TFGkw6fisGTPrAUK5NNvZ"
            "3qFty1Zw7fCysvzcXIBP+C90IZbLI6rLJCeQYa8ePYmgEhYViXhpovoL91GPr7vpbAgREE7RGgPh"
            "69ewioiYF+axJpiRbClJyURCCCXMnD7DZGWS/OfOnoUMsB7IeoZWtW/dZuSw4UDIz589h4r++OOP"
            "dWvWpgxNgl6TbC39EahLxZIfNm6cNnXq7JkzK8rLnz17pv1Lmy6T478aCPVv0++07CprGpyI9EUR"
            "EGbogVDvVOY/rzL6aSAMEgyrQOp8AIRNoqTYfwwCQhpXBf+djbeM2g/vDxw/z86Y3DQqG5APmQti"
            "KSKQnmvnvLLyTY8eoRercH/sOHiGlT4ZiNEiRoYtD9XQqi96ll6/g13TVlXqgg0CFhpxDpx/vXzT"
            "jZMOTGgZOqiXbMInhhdfkjZiaSNi31g3EBKcK52+0syja7OgIQ06DESRHlr3b9BhsHlwsnVEhj1D"
            "ZI9QUKxHQZk9BymLxgkmVnvVgc+zlu8I7JXryhY4MTOcGBnsISWU1xCDhDAKASEU2ywkdfUOQ9A8"
            "3c5p0bYGBAiR9xelDghJ4mnqkRCStuRPWrLj0GmqTK1+w7Fsy5FmYZkk6J8Ddipjy5D5dNbcvGvw"
            "CEzOO3/6hZc8woMrcmFluTAzO3RTndCLMasBYaTAiZ3dOCA1rWQetZ51IsRXbzv0LLAIzXTlKekc"
            "uStfaR6clphX04aQAKHUmSsDICz41rAppyhdMmKBYswieZ1psXzsEuWE5R2+KbaNltI5Ske2wpEl"
            "ByZ0YEiadEq2DkvtnDF21bYj0CRt9WcKmfeZK/aYhwldEBBWj01PSQh5KNqECyBKSFZG2ffa6vsS"
            "Io7+UEeq/gIV9+jSTWeOnM5SYiBUwfg0CUgdPkvnMoHkv/LHvRGz1lKjdO/hU3emoE0X+Ussiyaz"
            "OWH+Jn2Uy4/km6mLdjQNTHflKgEI7RkK/675N++Zzuyuw2fjUkb78KSeHLEHWxTQXW00s6iRr968"
            "bdstHwEht34gxDaEPJVtpJiXMk63/Ax63Q88+QosIcylgBCYyu2fkRACy2Eg5KttIsVdkDythoSQ"
            "K3diyT34OeaBGSU1ghnCgveNUzkyZF5xuVCRT3yuRXDmf9aG8BMSQgBCDQbCXFoUAOEU0x5due1F"
            "JIT8PHc+CgiB7P14ObRwkWVwRmCPgrFzNl3TWdVqqRVHfpKv337UsmuRC0ftHZdnGSLMGW8qIVy3"
            "66R9lNQDox2WEALX4fRZQJjnGYeN99AlZeQSn4TC5sHCUeU/aKujzsINh+zCxTAOnsjdZb5dhLQ2"
            "G0IdEPrE5VmFiAqmGLwKUVaybiyFJ3JDagyEGAX1sj70fXyhQ5Rs1XaDcFgXuX7yagBCv4QirzjI"
            "U6YHwuH1SwixDSECQqSqylEd++Uq1R6da9+p65BYL8EAhP4JRdD+MXO2GAZBHxXGjan05mmwOmjt"
            "EkLrELE/KqoQ2kOLkGWWGHzGGKmMqpCEENeF5YQoap9DhNQhXMQYMGLWkl33HppuoUhTj5+97h1b"
            "6BNf6ov9yvjX4VrmP64ySvRFiRkhBYTIhpAAYYlvbCGWEOYCENqHCTMK52n/iR1hrUddEkLYyxbm"
            "F5SVlOTn5uWq1bk5OWqlSp6dLRWJhZmCtOQU2MgmJSYCY/Tt1bt3z296fPX1l527xPNjuSwWKzom"
            "JiIyIiQ0PCQ0LDgkJDAwqGOnDm3btWvVGvbEkOADJNjTw8adiFBaGIXOI8K6zwFCSp9zWGnZyOEj"
            "OrZtS7l2TEtJGT9u/ORJk4iQMDM9A+CTQAuARFRYOFHh27Zl6/Bhw5Vyhd72z7NNy1bDy4aVlpQe"
            "PHBg+bJlw0pLkxOHemKrOR93jw5t2u7bu/fihQvPqtsZvnr16tdzv57++XQ0hlLIDOfoiMgRw4eP"
            "KBt25fKVeXPmjBo5sn/vvu5OLsQALzQgcFhJ6cjhw69euTJj2rQxo0bBSOq8mHp4tmvdpmtCgiHF"
            "J8BYwbwQI0/Ki6luT/Ly5djRo0eNGBnH43vgEmAEIkPDBvUb0K9Xn/59+nb/6mtoOdA4UcEFHIKc"
            "hNYApAf06QtM6OHiSrO2xcmmjoTYHloI2C/KEp46daqkqKggN6+Ftw/x7QklcBhMgJkTJ05Qg3Pu"
            "7LnNm36Ih4ZhfzYkZ0FeXklh4Y+7dgMKQrP79upFOg7jH9C+AzQYBoqkgX37hwYFIwGjnz/UDiOG"
            "omVY25BlAx2BJQfTtPfHPWtWre7ZrTshXuRUhma/euVK41G6fu36iGHDhpWVhQWFwPjoPN9ERQ/u"
            "PxCNUu++A/r0gymAlU+k09CYfXv2avUiaPjwx40bI8pgpQ1nREWTVw/Q7D7ffAODP3H8BKJSa7wy"
            "V69aDe0heEl6RwxNsbumYLgRYCoBXMlLCuK1Fe4jQFxmdEzzpubQX8DIK1euaKtD6V87/ruBUOcw"
            "4Jo7L6tZyFAUhJBICGOEVjFi7FSGAkL5fyEQAl9Zs5Gd4c/nr8JjSlv17uDJCxwEhFIEhHydWqk9"
            "X23FUkD5bXoPg/0cESS+ffNm7pp9rXuXwfdQWqNwiXwyUdmq0lYDwmrecq/ffuDBy3RiplgEDxya"
            "P11b7y+EzgH3qd+tWEpavSqj+sD068xa93VkCGjRAnuG0IEJECiiMUWIAxkkwgSgoM7JpyNXaRaQ"
            "qpliuolct+uYWZu+5oFDrMIARVKaBQ5xjE6jQtJVA8JooSNbYtZ+yPByBAPGXhBPX7wBmGTPlAK8"
            "0Yj1IJfI9NSOPI0jz0hCWN2pDLl8kGqKbejgC1eruZ0gzSudtaFxcAadB5ejLlhEiGIzJlLxpslg"
            "3rj90JWV1aD9QJvwVJTCUsw7Dd59+KzWSMKmlxBm0VnZVuGCoN5FxtHkSaXrdp+wicg065jcNDi9"
            "cWCq2Rf9k4vmaE2B0KAy2jQofVCuwQ2p7ulz95Fd4GCzVn0bdxzUsP2ghh0Go9Qepw6DG3dMbBqQ"
            "3DgguWFAqo4GUVI5cVXAM9A24FWHGFGzwCTzgMR2X+dMWbiFvD7X6dR91BEaZHZkK52o4BOmHmVy"
            "kUIpV2UVIdps5DbwLx/dJdMswkSuPKI1qrQOz2IkjqjUvShFGd6+q7z/6JlW/4IAaNYyMNEmJHn7"
            "wTPUVAI+QTbqEjg6Z06wDBNAX1yQJqq4MxI0VZ/ZOw89OaKmHRPtIzJQCk+3DBq6+/A5/cyiQv4E"
            "EGKVUbsoSeTA4ZQpI6noyfNX/gkY7YDc+HnYerDQPR6Y8G9LCNdiCSEBQp7GJkLSBbqprR0I3Xk5"
            "1mGiHmKD1xlq0stX7rEMFTQLElgBU4VkmbUaUvztJ1VG/yM2hAvWARCKMBACneYiCWFWLRJCPRDm"
            "e/ALPGLzkbcYvgaY0IOrsI8UWgamebAl6YVzLl2vxZi5t3w2LTLbF7lpEUmwGzC9zAo1YN/xi85M"
            "OSotNt8zrpCI+7x1+p/1qowCEMbmYiAsIQxJaZlahojzyfPcSPFh1+FfnWJk0FNPfq53bB4NAWG5"
            "YV6qA6F3bC5kGKCaTWUgfYLnTET/EU4MJcAS8iWTUOLdGTNhAjaN6zKMqIxCG2gR2YCgJqM9du4W"
            "7C8U+aHxhm7qDerwVSV12hASIIzL98HNXmTkpkjvQ+tXerTcF6mMYmXUuEIgNDeW8vBpo18ZnL9i"
            "xT7bMLFvXH5NICQFjpz1g3WIBMU8jC3yiy+yj5TVYUMIQIhiSPigZpf4JwBBFfjF5vnw1E5REttg"
            "QesEde745cZYSNaEPq59gX+XYXpfo/8XNoTGfmVMgbAMSwiLfGMJEGocwkSDlYaV8I8fdUkIYadL"
            "4kz8NQ/4xOTs5cuXT54+vX///q1bt65dvXbp0qULv57/5cyZE8eP/3ToEJDVrh07t2z+Ye2aNcuW"
            "LgXkK589G8ApLSnZE/ux/CQQkv/Cn5o2aOikj7OnN9xysm5mARBFosBt+WEzwAMU1RJL+YoLikgJ"
            "OQplIzMz+BO5kFQK23T4csK48YkDBzU2a0BYhfwVSgCU9ffy2b5tm1Zvo6jFfkpb+bVo7f8Fif9O"
            "2kAAAKrb8+OPnePiLJo09cRQZPxXJ5oDjAPAm6V5Mw/9X/0xYkG9xokUS9Q7ly6pBoQP7t+H/HZW"
            "1oRzSPkwoZT+LXSQYAnR6nShO23ftp0awCdPngDYEylfmy9aQheI7qJxgi/hT8B+MM4ZqWlw1XcL"
            "Fpg3agxVkOpIs4nrThJqkhRemJ9v1cyi9RctqamBjsBVTRs2ArJiYfIx6ThpM5V00jPgJRxMMk+t"
            "JlJW0lkAM5gvyANnMsXEQhLQjnpNT+YIlhzxAAQjQI0SFEKNEklEUk1aQgEhKeHo4SN0O3toBmkS"
            "KYFY/cGqeF5dlEcuAViFAl0dnQjBEle0UKmnq7u3hwf5hviegWJ79/zm4cOHcGGXhASi2QvQCByr"
            "/SfeB/2XAyF+DJ04bxWW1Dws2SoiDdkQRmehxDCREH42EP5plVH0Hnfyop2fCYQLMRDSdECognKA"
            "+k6dvYx0Rd8i8cWHqsp1u46HJY5rECZqzpDDtZANBbXn51jEoCD1rPRJ2/afRmJArfbRw8clsze6"
            "JuSZBWRuPog3pm9f6yJMoGgTpkD49MUr33ghnZHSNKC/agLao9SzNSeL5/Xb9+37jmzOUABT1Q+E"
            "KDB9636OTKFtTJZtjEgPgVKdVJAttwMU5ChpSMwINKUw65RSZKQqSQoZM2eDWes+rmyhfXQGJNuI"
            "FJ9YMdHpMgFC22ihA0vSKCC5X84MrdFaJx/SS+ebtRlMY0gdOAoHrtIBBhCpeioto2WQ0H9rkxDq"
            "AgCWVpj5dAvupbmLlRur897HbtJpTUIEdB6Ak8I8VNC9mpNGVMjBUxctg4c6MwWOMZkoEGJMpn1U"
            "2t7j57VGb4kwEGoACIHl7BkS64jM3Ud/NZ4OUumhU5cU45clFc2Vjl40e+VuvcBKq60BhI6sbJso"
            "UasvNdRYUd2Zt+bHxh0G2YalOjOFTkyRM0vswhI7s0SODKFtpKBpUJp9jNiFq6AD4nJUkOjcHNsY"
            "eeNggU2UxJEpdWZLXTgSF5bILjytQdv+Q3NnGK0Q/EPy+LlXvJrGlDvx9OHpq9kQIn1RZ2CPaGlQ"
            "37JXbww8SQ7iS5YyHTRJL169MX6K6anmQLOQLBfs+kWnNRotPGK0d9RW3/+JR3zXrFOiecdE9QSi"
            "JGMQdlGXnL1005klwf5FUYwHy3DxNzJDT3XxGE79Zhua4sYWucAwMkQuDKFTtGCfXmP5TwIhsiEE"
            "8nSIyW7zVf4LI+krKeeb7GlNcfB6NxzY3ZmfZ8/WOPPyXP4RCSFf7cLDKqORGAhrkRAqiFMZhIVc"
            "xa+Xb1Nrkvqw7eBZ+dhlacULciasWLTx0KOnLw2L89+0IVyHJYRQCwHCKADCuiWEUFdsgTs/3yoC"
            "ORF1Zas8uCpvvsqbp3BjSpoHJLftorpwxdBZMu/ZY5ZbhYp94vMAj1MK5xvNFDo/fPKiVZc8hxiZ"
            "F+JMREpusYVO3HyP+CJPzDbWn5AQFmIgLMMJgLDYO77INiI7KW+eSUWPn71s0zXfISobYM8nLpcW"
            "KRloxHsmNoQAjc5MZUDPEr1g3/DQ+H7DoeZBQhe22htJI4t9Ekp8EI6WuPGLnDgFqDHYPYxViOTb"
            "Rbu0hndtOpSyhxnhQgMKfbErGnd+MVyIgDCuHhtC5FQGU1y+bZg+CEd19cVvpDMsg8UonDpwWlxh"
            "8yDR4OoBMMnopRYuQMqluKg6gFAnIYQW+iVgICyuVUKYg6wHUU9LvOKKaVE5zky1Fz/PLzbXPw7F"
            "7vNkyWwC0xgDht17ZCoqHKSeR2doWnRBWqP/OTPCP6MyOkqvMjrM36AyCkBYAEDoy1PTIyTdjPSo"
            "//GjViAkSpXMqOjFCxcumDuvfNasad9+O2nCxCn/j73zAI+i6vp4KKGkl03vAVE6pPeyu2lgefn0"
            "tYBITa+bsiU9IXREKVIFFESqqICooAKKDVGsoKg06QkESEggkOx3zr0zs7MtBESNvHOf8wyzM3fu"
            "1Czz2/8pc1944fnnZ8+cNXP69KlTaqorq8pLy+CNv1SlKlYqVXJFUUGBLDc3NysrIy0NsGHShAnj"
            "xo4d/dTT/338iVGP/efhkSMT4xLixOLYqKjIsPCwoODQoKCIkFAyExwSEBjk5x843A/erbkAwvaB"
            "kMLqurWvwzs9dMMahqxRr9dg/wDqcXft2rWo8HAq13i4un3x+ed0hIqycgeiHPK3pZ6NixYsTEtO"
            "gRk6FGdAIDDOuzt2qHlA+NX+r7AKhbMLU0qRGJwFYVSvfZ988sSoURQ2uLVwJFQu+3TfPmmsmCKf"
            "zlq+cWPCIdHYOQ4IgSKGDBhI1xobgdbJAFKCk6U8yf+CunHjxtw5cyh0cSGClFUowsFC2HbgQ/2X"
            "LVnaQvJcwLMBDEbJk3/KIhu7yRMn4eGRm1VZXg6kqnNq9J4+P3v2iPgEV0OXRefI6cNpbWZRVVGh"
            "LCqyNreA44GbSzU6OMgHfH2BReEjVqW3tByRkHiGZAzif3Xs/3I/faLogO1fZ3qjaaVNjEAgI3z9"
            "Fdar9PHU6kYvDjy9DbwENvzn8+A33zw6cqSjnT08pd4eHn280XW2Xx88ZrotXHM48pnTZ1CNEe5F"
            "vEQKndFL2cf3+LFjOsPeXevkQIint3nnFz39xpqHTDSn/qKYZRSLEN6lQnhHQOifMZUA4fNrdpoE"
            "ZdnfHRCKKRCqq5dtX755D8ksqm5saFi47qN+j1cBZ1pGF9KK8yRcUEHrTzylfPl71BWx/fTLiRfX"
            "7sIa9zeJKkjTigIx3uADIU6v32jxf1JhHz6x+9CnZ6+iwlr7hU1ws+w5m7oEZoukCmMuoxog7D/a"
            "PppxDbUmrqE20QU2RBK0hTONLaIG41AgLHh+vVobxg4ePi6KSAGwNB02rsfw8SYDRlsGTahnq8nx"
            "k8pYY4HBHMvQDI842YkzmvoE9EwBJIrmbnAWy3oHZ/QCC8m0CM+xjpKFjJsVPnGOWRiCojEgTKlY"
            "ZjpkTPfBo/+TTYO7tMjh2Om6vo8UW4ZlO8bm9w5Kj508kwsAowcAuw56qtSk31Omw8b38ptoOnRc"
            "l4GjP/icKfWu7TKaYR+dJ4qRmfpNfrJwoVqbz9v/69UGwjx7zEQq6+WfvOHdL9W8vDJ0kNVvfzJ0"
            "lMoyKNkiKKW3/+Re/pMtg9McorIDn64unv9GQvoL5iFZjmKgQbmjVGEWmjci+6WShW8PfqIKILPb"
            "sEldh0w0HTax1/AJcFJ9E3Kva3PIxSuNXolyu0iZlkIo1fiLwkIXLD+YXrpAI3qwPzfciBhd7hKZ"
            "7h2X6yXN9ZbmeUnzyBQ/ekmy3aPT9351SM1+HdM9nrlw2TdJZReVT9KNYrZM02GTK17SKFTM4ZGZ"
            "puYbwx+H19AUq8DkoKcqGC2O96JJN5m1cgfmFxWjJ6qzRG4ZkhOX/LzOnYV369CnK0weGtPbb5JF"
            "YApMTQc/96FG+8XRNEDYbgxhd3QZxQqHDijEFfx68rxaAyHkZ4VvjvRNUsA41mF5thEFrtLi8PFz"
            "H3i0yjFW6Spttw7hHQCh0hgQusciELpgKYUis4D0wtnr1drfFe3/z/IXxBAaLTtBYgizOCBEl9F2"
            "FMK4Yhexqu/DldNffv+xnMUeUoVVcHZvv1Rz/1TLgFQLv8nd+o/ZsfegWvsLIX/mBvPALK94lV1Y"
            "rnjCLPp1Ry8A7bB04x7nqHzHqEK7iALHGEXfh6tiJy8A1sLC9KTsRIZRhbDYI6GcAcKkapJXpgKo"
            "0iGyMOLZGYyCzdvRxvf2u8cW2oXmOoTn9RqWMkar7qgOEBZ7SJSisLx3eEliuc4rt+wb+n/VjlFF"
            "okgwuVMMXLpS/6dmj8pb4U2ETQBCSzaET+dHw8zq15wi8p2j4UmTu4pLhv13ZtSE+R5x5SQTTHt1"
            "CH0SsN69a4x80KPlf+i52R87VZuU+qJLdJFLVBFMH81cqN/n5JmLAx4uY8pOABAaqkNIFcI+ALpx"
            "5QiEITIDMYQMEJYCDbpJysKenTtzxa6k1Je8pMW2IXlWAdk2gVl2QZlWfim2/infHj6h1sbX55Sr"
            "HCIVfUfQXKO6ten/kRjCvmwMIQuElRhDSBRCb4nCJSIv9MlqvhPKvW0GgZCKUfCqCm+x8OIusrYV"
            "WdvYW1mL2FoLDnxjQwEBD5xgaidytndwETm4EmEKDEsLOrvoVO3jwInOU6Nr4R29gy6jYFIFDwAA"
            "IABJREFU9IV704aNvl7eQwcNHjxgIGdDBg7q3+/BmMiohqsNtDMAj7e7Jyx88vEnNFUKptR4uXsM"
            "GzyEvy18hIXLly7Ly87x9vDkjwzcNXTgoAH9+n1A8qZw4xw8eHBw/wFDSAf+UIPIws8/++zZ0aP7"
            "+vjqHCSsHTpo0JdffPHYww/369MXjnmw9ub6BiP08fbZrO0MeeniJYDqAQ8+ZGwEOGYwSay4srzi"
            "xx9/5O47bZyG9tuRI+UlpYF+/j6eXnBDbS2swOCGwl0ICQyaOqXm+HGsOkMTqGzasEH/sg+Dw/Py"
            "zs7M5O7O9KnTaKoenSsMu1g4f/7j/xn1gG+fjpw4bOLp5g7UBIQJNwV2XaxQfvThRyGBgfBo0chV"
            "eHgG9x8I0KiTy5TBuQMHBg8YcNsd8e5O/88+/VTNUwi/PfgtHCq9pzpPmjg6Rmen/F23tLRseeON"
            "p57474N9+3oQAqSBtXAWAcP9SotLfv75Z+6KwUP+xKj/gzHh2sJa/VQ0d9c6NxCS05u3Zkf3Yc9a"
            "hExCl1F9IDQeQHhPXEanrsDfeBZv3G0yJMUiOt/OEBO24zJKgfBrAleji1eYDE6RZszf8+VPcHJq"
            "DESuK1n4lmO8smsQ+oVSD1KaRYbGGWbPWt+Av12RK0HdRFkC1AVCOm1rS0qbahX8nHngs6++rXlP"
            "NdYYX9BXd8KZOsQptYDwez2X0Ze3mgx4VhRDys1Hox5oE6uBQFuamYZNTmOPaTNTniparNYT9/b/"
            "8HvR868/lj03LnnGqNwXUqteZgiESyqDCiGKkLZROfbRuV2GjJ+58h3uMPhDnTx7ccsHX69/76sN"
            "73+1/ZPvvzmMfxVr3vmyR1CGKDbfLjrXPDjVEBC+bDJotFNUOuBoxUJNDgk1+86x5cOvewWlAYBZ"
            "hmb2e4QT5TT7BUCduvStp4sWSifP+L/cF0fLFx7+/TR9Yhm0YOsQiqJldlF5oug8gLSVW/aSs2jV"
            "ZLagafFJcssWUseDO0F+HUIChDJRTH6vgNS4lFnMjrT/TOAg9+w/vH3vt5ve37/hvS/f3ff9L8fO"
            "NpFXhGkv7+g6LNVZqgAgdJIqugdkLlyPz8aVxua9X//ywpqdBbPXTSx9eULpsqI5a785dIw7Te5t"
            "2zk2315HIeRMUuwgVjjA30tE7uc8EY8e/3uffNtr2Dib4BTrkFTrkDSbkHSbUGI4n2Yflmby0DPy"
            "2Wv5t4ApSFi+qmdABk0t4yQutArJDBk9hXO85F+iPV8dtg5OdYjIBLMJSd339S9q3v9ktC+gTvS4"
            "aZbBGViPHovRy20jZAMeK7vS0KxzZ0+erZu+bOsY+eLElNlPyOaPVSz6+Sh3Z7EbJpV5tJSnEBrJ"
            "MjocFUKnWCxFaBmUsW6HplQ3t69zdVc+/OLwpp1fb93z3U+/nYHr9mTRy5ZhMrc4pW14RwrTpxko"
            "TP8mm1QGgFDKuozqA6GYAiGWanCOKRSFZ7+/D3/RaLl5i9MJaVpg+nDe5C1XGwXCu44hVLqKWYVQ"
            "rzA9C4TEZTSOAqEBhZAmlfGML7GLLBqZuYgu/+XYudVbP6946a3MKWsmla7Irlm9dts+Na/RU3q6"
            "YKk1po1RwqXwksp/I1UTdP5z/fXE+Xf3/bB559fvf3ro9IXLFy41PPToFJdYlTdTh1CvMD3nMprA"
            "uowyQFjuEVfmLlG5ROd/9u1vah6K0D3+fPTs5vcPvLHzwFsffPPr8XPcAfCAEGsbemKcYTEg05ii"
            "ZWq1gV+aLl9t2vnZoc07v9my6+D2vd/v/+E4AMNvJ2t9kGNLvRMASgukk55v0fvFEO71dz+f3Lbn"
            "uzd2Hdzz1RHYavf+I66SErYwvbHMnwQIpcXe8SVWgdn8CEnuOsP3/P4fjm396NsDPx7nfrhhdkph"
            "b+l266Ac9Cw1VIdQGwjLEQgTyojLqNEYwj5JlfYRirQpzO8dPxw5vXLLpxUL3s6esjal7JWCGa+/"
            "s+egzsMAByaeMNcFaxJWs3ll/nGX0ekal9Ek6jKqBYTuUfkPJSpqL+mKD/eqGQNCzu/xob4PoETm"
            "7gGchtjm4grv38B4rg5OziIHGgTFJImhrMgmjCFm72hrr0khQ6GRbEJnwGAQNHsHGnqnE0DYkaQy"
            "zc3N58+fv2Co1dXWchetqanpwnls/Bd3mDe47flz565du3blyhVjI/OzSqrx27XFYDemc0tLfX29"
            "saGAFoDojK01cGznz9PyeprHu62ttra2nU3gOly6dIk7Zv2ANI4JoV29evXAV1+9snLl3DnPg726"
            "6pVvDx6kiXbUPFcpY5cdFtK6HbTBFT537pzBbnCF4ag6fuIwDowGhE9u43mq/cKCD3d9QNPevLtj"
            "B62A0qadwoC5Ry3t3SODjeYK4trNlpvGrnJdXZ2xP0/uisHTe/rUKTjIJYsXzZ83b/UrrwCjUkdT"
            "fjdocAHhRC7W1V26ePFe5Rrt1EBIz1w+d23XIWMsaQAhuoymY0aZiGzLKOoy+tcCYdXSbXAMjVeu"
            "LFz/ocfDpSaBWZgqRiK/AyCUyL8hCuHEqtXdQnJ6hst6R8oypr9+4g/mP/uff/tjfMWrZpH53UNy"
            "aGF62KprcM6AJ6cs2bT7ZnNTW8t1zCyqU2pCGwi5y/Wscr5ZwBi78AnvfqL5IdxYo5rA9FcIEGoU"
            "QiMxhC9vMxk4VhQrQ20wtkDDgRwKihUUCNFxNCbfMjzLVZJ3jORy0Lyg8/4c9P80tOoQYo34XLuo"
            "XIvQdO+EAuroxVfG9DOqMZv/cBTrEyJEGQHCyhVdBo8VRWTYhKWaB0zcrePMSfpMKF/ZzS/ZMTav"
            "V0Dy27sxF59OolRjp6BRCAEIw7JEMTIsLAFXLDzLPjJz0/tfqplvIlrqAOtD0CoROs4+PIUwk2Qi"
            "zSdMKIPjWbT+A3pT2njv7gbvL32FWvfefrOQLEeJ3AFFwiK76Pw9B35Rq9X6F1Cn0VNe8ebHPf1T"
            "ncRFAH66MYTSEgeJ0kmqtAjNiZ44i/HV5F3GjOpV3YeMd47OEUXlOIBF57GWCx8do3KsglMHPaa8"
            "2qjJ+8wkS9z1tVkwLUiIXqOOMfnWIelcOB//7CoWbjEdOsE5Jtc5Jg9mKrWFRHoY+745YhuW6Rid"
            "7ygGvFSAARNahmRuJa+D+u/EarUBzysWCLmyE0WMQriJcL4uEKI7qFOs3FUiNwtMH02knnb0YbpK"
            "8eKbPQOyPeIRewwCoXeC3IGAHAKhX6phhTCQJJUhQGhtLIZQUugUA7BKQw3l9hF5fRLle7/6Wc3+"
            "N0kgkFQcvXlL/z8buse12zsQQ8gAoco4EGa7k4QxnvEAhOmGgRBjCFWuWPih2CY0zzgQFmGm0KC8"
            "wjma3EvtNLrt2drLD4wodooupDUMLYOy5q3Zpdb+KU3rW4tu29o2In2hfWSRT0KJZXB2hmEgLECF"
            "kMYQAg0m0Ww0pPJEfKllUHbBLMzhrv87l+Gj5YCQlluMK0GRUKp0jMijhUMwWa72byX8xv2c4ffk"
            "DOdYlVd8KVxS56iCT7/5Veda8Y+B/f3rYp+kCndJMUAaAOEKo6UgsN69F9zuWEWfeOU3h47zz07n"
            "9Yv/kfb59vCJvvEqd7GSDFIGsGcTnLfyDd0UuMRlNIdxGcUYwnwDQLifAmFJ36RK65DCuas/0rmh"
            "BhvNvP0dFkJUYQFDjLo0oBD+IzGEvFKEHBASl1E4VLHCM6bQO7bg+59PqjvwrX4XrX0ghIX+Q4ft"
            "eGfH7t27t23d+vabb23ZvHnj+g2vr127+tXVq1asXLH85WVLlsC7+PwX5819/vlZM2bOnDZ9GnqT"
            "VlaUlZWoipVyRWF+viw3Ly87JzsjMyM1LW1yMqk8gRUpxj079tlnRmMSkSefiomM0tEGOwiEQut4"
            "4yRNgw0LU7VbduLeFkL4883g8fDJqrM0eCfkvdTptL/nwv4LgHCMcqHJoGesEAiT2ZoTmQQIc1mX"
            "Ueoveo9dRq3FRWZR+b7/qfjyuyP0eE6eOpc+bW2vCFn30BzAPKA+Y0Co4zJKgXB85atdgrKJEFfU"
            "JTDbfWTpzFXv1l8iP5O03dq7/6fE7IU9wnKBOUVxivJFb9dfwqSjt27evHHtGmqDt7S/5vSAsJUp"
            "yfCK6bCnXWJSvvpR6+dng40hvVd2mfhlkgMjMYQROTYRmVQhpBJBC/k1dxoC4TiRON+GpHKxiZXb"
            "aDhQwTIhA4Q2ADAxWJM9vYbWgmMqvLUxygN32JrXgrY2BlA//fZXOACbyBzbaMp1eT38J4WOqaL1"
            "DGi+eOa1jLy5ckb/y4durnGFcBYAk+bBadtJ2YmbxHGCvlKkVK3sOnS8KAoJraf/xOBnKmgmFZoj"
            "lL5LHT9T551YZBWWaRaUEvbslMam62QQrnpVKxmQOWwafKhmRc42otf1e0RlFZYtioXrUGBPcA7Q"
            "ziwwOaPmFdYDVvte3Go9f/GyDuPt/OxHq9AMEaIgDiKKybeJyLYNy9jywQHm9HlF9sg3tcZa2SrM"
            "X3x/FDaEwwAgtI3K90hQnr5wmZ4FkShpXQYGA1jXTeb96XJDU+izNeYhmU4SDgiLtcsPMv6i04mW"
            "TjGVDnLxcsODI+XWoemEAGUOcAAxhYzFYo5NUXQeMKGZ/+S3Pjygs+2lK9f6jyq3jaReo6iz9Rie"
            "nD2VhMXeusX9HAAvwZFjayyD0pxiZE7RMougtMjnptKnmp4TfaIKZq3D/KISQoNAsGLkJcvQrKhx"
            "0+/izg54rMwuIs8JgVAODLNs0x418zrexuTjXbTVdHiaW5zSGeFT7gx/DqGZ7+3DhDctLRqpDX8U"
            "YG8ZFUyWbf6kd1C2e5zSLjwXYwjZx5Jek3N1V7zjFcRlVOEuVfQYnla+8C3uutFdr9zyiXlgpptU"
            "6SxB11Pr0FwAQrJ5Wyv7bB9CICwCIATEcgETY4k/+/Bcp8icqpfevMDoDFoPJxxe7cUr/GcVpq++"
            "/alFULsuo/EUCEupq+fbHx3k/hLpPdoEQBhKFEKpCoCwt39GxUtv65wRIC5AGgOEGEOIQKh/Rh6k"
            "MD0MYhmc9zKhCCq53yRfC/T5hn9v8iqssAlRthF5UOVGylrA5R34aBn1ZuTLovyvGrph5tR1ViF5"
            "vomlHBDS/dBHbicAYVQBk2U0kQJhNVuKsAJwzl2scImS7f7ysJrxGmC/01o1O+KFNzO/znz+7e8E"
            "CIs944Do0EXTOaqwT4Lys4MM13HlE9ra1FpfCOwL0JMFy+0j5N4IpSV2oXlP5DIR2rduMQfQxjtZ"
            "eqbXmm+EjZntHK3wTSyzDspdsfkT7oGn9xGBEHOTVnjGlXliiGCJQ7gs7Jmpf5y9yPxxcWVXWmms"
            "DXNZ6Q1Sk8S2EaOnOUTIUB6MK/MisGcTwgCh9r7esQzMRZfR+HLYIwBhaiUNWaRqNvbZw5SdKPGJ"
            "L3ONVb237yf64HHPQxv7PNziPQ/0ryyn5nVRWIFvYqUPcNeImj4j/9kYwum82vR8hbAKro8PlljE"
            "yhPe8A0ZlvP+vUjoZbC1A4Q0kMzFwbEov0DdMX0Sb2VLC00nc+XKlUsXL6ISdO786dOnT5w48ftv"
            "vx0+dOjbgwe//OLLTz7++INdu7Zu3bp548Y1r65+bc2a8WOfo+lb7gII24w3/W4d37aDw3bkGO56"
            "7Z/fe/sbGhwI/69kG7ymtlLf1DZNH+Z/UFrtR9/YDrfpdufGH1BNPkK7xR0qHCfbx6CD9V96nW9/"
            "acnWrbdauaM1xq53dss61jovENKTbGq+IZ1c02XIaEsdIIz8y4EQpsCEvSLybCVFZYverqu7RA4L"
            "yO1QdOoLXQIxJYydFGtLGAZCcZEdERJ1gNCO+IXCJuZR+SZBmFl04boPmwD54Gv3ZgtsPrFq9Xds"
            "9OCRo6cmVK5++6Ovccctt1EI6YvI1GVvmAz6r3d8xokzterbfUGzQLiTBcJCgD1UCCMy9/94VK2j"
            "EK7YbjJoPAeERBjkcSCfDMk4NohzOT0DJpct3Kz5X6qNvgRw9dNv3eKVAqe7++qnY3YkY41tjIxl"
            "wlzT4ROH/F/xvm+OcCfVSt69qMhGZ7jKV+LkOb2D00UxGEO44xNNjA0dP23KK12HTRRF59pGZDlE"
            "ZXcZ/FzN0rfUPHimnZds2t1t+GTHmNwefpOfkC1gk2owX83UyZM7BZbLNL/EP/hosVV4jggej5hC"
            "rGMB/AMnEpndbeh419jshzPmyGaueX7VO69t++SF1TtUL6wLH12WUq5JIspIKDu/MguCE8kX0QSh"
            "wFREbLQITpnx8jamXAf9L5ZcBOY1jlV16FBXGpt9R5bYAFxJFeYh2YkZ8/hvfrxriBtyb0tkw6Yn"
            "8hf2Ckh1wpw0RQ4GYgiLHaUqUQx6df7w6ynuGjIHv2u/WUCKY0yeQ0w+bi5WaCxWjhZT4CwuMB0+"
            "Oa36Ff71Z1J9zNrQIyATMMAxVu4Ez1tE7gMjFVwGCNrn60PHRZHZDlF5jjGYNhNmRBFZB9hHl55F"
            "bf3Vhx5W2YbnAJs5Ea9XJzAxZnwxD0x7qmDRnd3Z5hsD/wOkKnMWy12kit4BGcs3f6zWFkymLNlq"
            "6gdAqHIWAxMqXSRy2PuAR4qpZkIvL6sPM88/ezonnDEBbBEAZ8BT1TrBnIBqvolALIXOWBtQ1cMv"
            "vXLR22odPe2tfeaBWQCELmKlm1RlFZo7MmMe+5fHSFg/Hzvrju6scldpMTHKhHLHqLzefin9kuRP"
            "5i1QPr9hwdqda7d/NveVHaq56wMeLy6fv5G3L1L9Zc3Onn5pnvElt4shLEWFMDxv6x6NVs8+3l/b"
            "hOXCWlcJBcLMqsXb9M7oUyAuj7hiV0mxe1yxbZjskcz5Omf0y7GzHlIF1iEUo8p38Oc/1KyXL3mk"
            "2ceb9y1BL+nLm/c6hOfBuWOlCixZoQIytA3NeSRjHvdUtLXybhbeJ4ZhFm/YaxEEZFJqGZSTWbNW"
            "ra0QfvD5IaeoQtTxKBAyMYTVWNE+oQKLE8YVO0cVPJCk+uTAEW5b8htBG/dgaB4+9o9i/w/H3PEc"
            "aXqbMqw+H1fsFJnvLSlau+0zvvcpPeVbzIm3Us6EVZWLtlsGy3wSyz2lJV5xxXahucUvbG7jFFcW"
            "mW4ye2f+iNKnrLMNLeiTRJLKbNGodoyr57IdFoG5vkkVnqTSIMCqd1yJKCzX//HKbw5p0t+Rt5xb"
            "VHxu5WV+Onj4ZOATVaKwPIYGaaHCxArbkPyVWz7V2df0Ze8iECZhdXsChAXpVWs1u6Cpsw8ccY9V"
            "eMNFlpQ+9HDV+YtXub877uy454H/XySQLaAs1nxPqCS16WlSmX8whpBxGWUVwhosRUiBkFEIS7zF"
            "WIrQxj99xabbB4ncXWsfCKk5ixyeeerpQll+WnLK5AkTnxvz7JNPPPGfRx5Nik8UR0VHhoaFBQUH"
            "DPcbNmgwFpZ48KHB/QdgYYl+D/Z/oB9NrsgZycnh5eOOmUI83dzdnV3R+9TRycleBDP6/qL/GoVQ"
            "G4FuY62ctRqwW9Ru3cZu8u1mB4ztrDOO/gHoH/C/umkuO3eFda6hoWule8HbvT7/hmvVeYGwlXFW"
            "qRs8Sm5KqtIzNSeoyygAIVOV/g6BsMMuoxQabcRFsJVJYGa/x6vQgfMG6gnXm5oXb9jt9Wg59SB1"
            "iFfqAKG9VAHLu4fkWEYXoMvoD7pAaE0zi0owsyiwZVTy3Hf2HGwlX2RthPGaGhqmr3zXMUFlMiz9"
            "LZI24GaLlj+6MSB8ad17XQY92f+RPFqT7TZAyLmM+mMMIU0Qah2VaxOZtfurw/Bfy+WrmB/yEkn6"
            "UrnkbZPBE0XiAlvGX1SPAzVACFaAcYYwVERW12ETYidO27r76/qr19o5mKbrN745dGz11n2TKlba"
            "RGbbRstoBQucwQLuOb0Dk80CJ6VVr/jhyEljyXLgff2z734Le266eUimKEZmHpKxedcBuAiwaziR"
            "yw1N8EYysQLdQbGMYRS6pNqEZzhEZX727a/AAPR8rzQ0N1y7DmcN4wBYOsQAjk566FHFkg0fnrlQ"
            "3/71/OXYmQ3vfVm6cIurtNA2SiYC+McaGEX2MQXAdQ4xMqeYXJuw9J5+E7sMHNtl4Jgew57rMmC0"
            "6eAxJl6Plbyoye3BJGF/Y6+pXzJWkyfjoAEcRssoWAY9XbFl1/7LRq4qnPWNGzcPHz0zc9W7bnFy"
            "OABHcZFNpGzIE1WLN+45eqq2nUfjWtP1Nz84EPBkRQ/Ye2w+gijKg0qtOoREIXSOK7YIkyWkz4Mr"
            "dvUakzsULiO8TyZXrurhl4L1LTiYpEZLGoqxUoUjPHIAS49hSUC4dzQfKXlO2t7afdAmQoYIB7vG"
            "wL8iQNMFr6O7LHSov4JnXbN0W0//VMzdEluI6WckhfCxajFiEnSgfRZv/Kh3QJozdRaVqpylxU4o"
            "bCqJ8FjYOyB10GPFSzd+1IE7e3bT+/srXnrLM0HhEFNIC/31Dspa+PqH9JDYI1eXLnizh38G0Iuz"
            "RIUmVmD9idAsD0n+3FffvXi5wdD4rT8fPTtl6XanWABCTOHj/+SUuvqGa83X8YFsBPZvOXrqgnei"
            "yiG6yEWidMeIuywVKV0AO4V+9DFYsnGPRXAOsKgLEE6cyjpMlpg270pjM3wbNOCD3dRy8+aBQ8fd"
            "0Z1VCSBHqlyUQGcsQC9VuEuKROHZFgGpPYZO7DFkgoX/ZNNB48yGTTB54KllGz5Ua7+aVy/e2nN4"
            "Og8IdWIIKzVACOARkb/xva/gtl5uuHaVPdrXtn8BgAfnAjDmmVDaOzC7ZD5GRV7mndHSjXutQnI9"
            "4kugj3t8iV14/oj0+XBN4JJAH7gycEYA0h5xwMAKIGE3saJozmYAJ5pjw2CDF9zvfzmZXL7SNiQL"
            "M5EiDZa4x2G9CiBDQErr4KyA/1Zt2XWAeu3q/XXcOPDT8eTK1xyjirwTSqxDclMq18AXCxwMHnbD"
            "NfjT27bnO6foIgS2+Ao2hnAKSTTK1KvA2hhxKseIfPeYAsXzm777+STNMWOwAZ3u+/rIkg17xshf"
            "dsOMrGU4LKAXMCFyncolusA6MGNU1vyPD/xMfyfSby0tt+CvfmLZGrvwIu9ExDbMmypV2gZn/1/2"
            "gi+//92gO8nFK407PzuUkLLAKVreJ7HcJjR/MYlA5t+jype2WQXLfEdgaQrvhErAOU8pMGGxI8CV"
            "VF764huHfjut7+8Euzv8+5myeW/4SIugp3c8HEwplsRIrIJBfJMqbcMKF6/fS/bVpL2v/D5JuBfo"
            "Yx9eNLlizY2WWzRlMTzh8Obw3r6f3MRKzDJKMppOW/7uD7+caoeUgFH3f3/0mfwlotA81NwwfU6V"
            "D5ad+GfrEOorhFMZIEwEIKwAICSlCJV9pEqr4WkV87W85e9h6wgQkrLaIlIjgakj78KmiuFyw9Bs"
            "jXx3UwPm7c0vPd+H7cZ8NESDfwkQGuU0HSrjYwOLCnyO6vQM8Hc08lOTFmu13NTYzZtaH+ErV2M3"
            "1Nd51nxd3dR8l3b9uhr+R+D2Asfwz94UDfa3tvcItQ+Z97p1YiAk/zl9c+ioY3Rq78DxpOZECmaU"
            "iaBV6XMso3Lar0p/T4AQdUKMJyzqHZEH5JaYvfCTrw5RDPvj9PmMGet6R8p6RchMAjJrSD7Ste98"
            "Dh9NQ3PNovKfLVvlklQMg+grhJzZwMufVNEjLLd7aO4TiuWnz1yAU9/95U9DRk/rGpwDHWC0nFkb"
            "zp2r1fXb1gNC+j/B+vc+6zl8dMTYMp0MHAabRiHUAsI8m8hcV6nMLU7mJs0llucen+8Ym2cVmYvy"
            "IE0oygUN6gIhmYc+MQU2GMsH0JXdw29Sj+ETHhhZ8J+cuZk1rxTP2zB12VtVi7eoXlwvm7lmfMmS"
            "R7LmDHhMYROa2m3YhG7DJ9tGYykLUmOwkGNCu6gcgLduQ8dZh6QEPVMB41Qv3rJs00evbf90wes7"
            "S+ZvGqtaOnBUiUVohmVYlh1ugnt3keZ7JhR4xMk84vI84vI94wsJCuZBB2Iwk2sdnimKzPSKzyd9"
            "ZJ7xBR7xhd5JSre4IttIhEZRdK55UFr3YePdpTmJqTNTKlcoX1g/delbYKXzNhXOfj2lYsXjefMC"
            "n64AtuzlP7nLkEm2kXnoLBpLCiTCNLbQMjynm1+qWVC6bXimQ1SWc0yOS2y2Y1SmU1SmVdAku9DJ"
            "tJA6X2Qrnre5y9DJThKCgmKFCAxm0AEVSwiaBab09p/km5g/Wr6obP6mF1a/+/o7n21874sFa3dm"
            "T139cMbzQ58oc4jONfVPtYuSOWDNiUIHcZFVeG4P/1SXWFl8yuzsaatnrdgOOLR2+6drtu1b8NrO"
            "svmbJ5YuH/wf4I3JFkFpTrEyQoMc0ZXoACEshENyj1f4jlB5J8L7cYF3PFihVwLWORSRsD0HsdJB"
            "ghhGrMSJupviaEh6DnCVonI94wp8EgrJtgU+iXKfRBiwGBDOkYIoMmGRKApwWuabWOQTj519E+XO"
            "MRhdiUGGyHvoXOoQne8ck48jkNGgs3MMHIaMyIOkrkNciTMcA4qECprC1Doks5ffZO+4/JFpc9Kr"
            "VxW/uHH68q0zlm8tX7BZ/vz69KpVT8oWhI6uconOMQ9MMx2eSmmQjIZU6RGv6JOk9E2Q+yYUwSH1"
            "GaFyj5MjwRL4RAMmjMWahPYRub39k/skFI4uWlzy4saaxW9NX76tZN7mlMpVUc9Nh8M2D87GJKhi"
            "DD4Egu2TpOqTqMQxwWAXScVEP1QSflNBH6CgviOK+yQqsEOiou8IlVe8EpZjB2kxmDMyEiwv7puk"
            "JN0UfZNUMA6wk4ukGFDQFUsgAhMW20cCG2fZhuc6RcvcYvM9JQVg7rF5nuI8m8Dkfgn5tEYLp6PC"
            "NLniFfOgLI/4YhjEzQgQusdXuMZjHQiAWJ9EVT84EjgGYjAPSwjFARCW0D5e8ap+I0seSFJSg3mf"
            "hGKidpZgLCIcKgE27DOimOkDY44AkFOQevQqV4nSMjjXIVIW/PSUyWWrahZvXbzz68rOAAAgAElE"
            "QVTuwzVbP93w7hcvb9ozfdm2/BmvJyTPdo7MtQjIcENtUElQE4vOu8eVARYiE8Yp7cNzbYIy/B+v"
            "SK98tXzBlhnLt8O2qrmbJpSsCHxyimNkvm1YvjseeTGM4B1f/ODDZXAk/ZJU/ZKKHxpZ2jep1B3H"
            "YeoWMkllqEhIco0iEwIGS1VuMYWWARkukbnBT1ZNLl1RNHs93Vf5vM0FM16foFr+cNrc/iNVTpF5"
            "FgGZtqEyVBeZ2oZYQALgkAQTqjwlCtvgLIewbHLMr0xZ9NZLaz/Y8O6XKzZ/PH3p9uyateLxsz3E"
            "CudoBSFJYnGEJ6Uq25Acp4jc2HEzCmeuq37prVkrdkxdslU+e+MzBUuG/KfSIbzQKUruFV+KfqoS"
            "eJbKBjxS/tCIErSRJTDfJ7EMCRBQMLGKMTwwrAXvHlNkHZDpHiMbkTK39MXNC9d+sP6dLxa+tgs4"
            "cGTKXI+YfOvATPdYOdCjJ9EGvUnJRMqEsOSBEeUDH6noP7KUGszD3knh+ApvCp9xZX0S4HgqBjxc"
            "NmBk6QDsU/7gyAoMRIwvB7TzkpbahuZ7ihXi8XMyq1+bsXzH8g171m3/fOOOL+F5mLHsHdn016Xj"
            "Z7tEyIAGMU0L5s6p9EmcwmaUuff+oncDhCOm+bKJRn0T2coTTG16VAjtg7LGK5bd9n/8u2sdAULK"
            "hJzKN6Dfgx4urphB1MnZ2cHRiaSWwWSkJNOjnRWWL6elzO10jKyCPmDOIod2dndnQKhDdMwruCG1"
            "5y/lBNg1vJhdv97W1NTWeK21obH1akPb5SttFy+11da1njkHduv4H2i/Hbt15Pebh365+cOhm9/9"
            "1HLg25avDt74dD/Y9d37rn/0yfUPPm7euaf53Q+btr3ftPW9ps3bwK69/gbY9bVv3ABbvfHGK+tv"
            "rFp3ffkatEWrwBrnL2uct6TxhcUNcxY1zHmpYeYCtOkvXp36AlrlrKsVYDOvKqvBrhRVXc4rvSwr"
            "rc9W1acVXs6Q1yfL6sdn1U/IvvRs+sVnUsHqHp9Q93/j60aNqxvxDFht0tMX4p9AE486H/UoY5GP"
            "nAtOYC3xXEDc2eFizi4MF9fybXB07eAoxgZF1j4UprF+IRf6Bt2d4eYDI+qGxtSSnZ4LjDsfPuJ8"
            "zGMX4v9bm/RU3aNj6/476eLotEtwdin5l7OVcNZX5FV4Hcpm4DWpmQvXp2HG/IZZCxvmLm5etPLG"
            "ytevv7bp2sa34OI379h1feceuCk3PvkCbhPcMrx9p860Xqxva2hsu9akbm5GoIVbf889utGrhH2Y"
            "//TT29mB8J2Pv+nl/5x58ERz3SKEt6858eddRvlGsoAWdQvJBkJLnbr25CnMI69uawU+jMtcYNJv"
            "YgXRJVa9vc9kWJo0Y/6+A4evNV5zHVFiGV1w0DgQUoOFAFEmgdk06VnmzPUmg1Md41WwU5vYop7h"
            "eQ88XlV/VVOADps+EJJH7eOvD5s8OCohZWrHL3LOnE1wYCIS3Ei9Pa2jcy3Ds8xD081D0hgLTbeK"
            "yKZ1JmxjitoFQi6SsAiZENAOWC46xzYyyzwopfuwCSaDxpoMGGPSfzTawDEmA8d2HfJc9+Hjewcm"
            "W4dn2EfnimKxoAVsS+rLoxG8zCeAl2MPWBiW3itgcpfBz5n0H9N1yFjTYeO7wJgDx3YbNtEsOA3l"
            "RPQ1RQPks47IsQzLtARKRMsEVgTctUOsLcDi9TH5diQRqE14NumQjhaWaRWebRWeA0DLdIjGUEbA"
            "QqvQ9J5+k7oOGYdn0X8M2sCx5BTGmQ6fAIRmE5aJoXEkdBCOXCSWgzlIFL1Cc7JnrJu16r2gMVOd"
            "Y/OsQzN6AjcOHtdtyDizgImOEWnLNqECo5XLsbVVMnlW76B0R3EhDiJWAJUxTIghhXkiQJ2oXKuQ"
            "9B7DJ8IxwEXo5Texl/+kroOe6zp4XC//ZMvgdLvIXCcxU4GQbCt3iC10ii2wjcg2C0ztPnSiyaDn"
            "ug8dDyfVc/hE2KrLoHGmwyZZhqQDSTrGEG2QentSZY/zF+VcRiVKWGsPNygCrmGuTXgWazmAag5i"
            "ZlsgMbJJKWslRKYDU1DfUbtIvAW2YcAkYPC05NlFFTgC43G6ImXC6HzbiBzOHGLyMesMhUbCjU6E"
            "Ce0icu1IB5g6RBcAI2EfgqPOcWVO0lIeExbB1XCJzbeDBz4gteewSd0HT+g2aHzXQeO7DZlgOmQC"
            "LDHzT7YKThdF5jrHFriSBJ5OxO+UCIBK4EN7gH/YY2QuzkTKkBglSkRBxC3cF2VCwDmX2AJ7uPL+"
            "Kd2HTOyGO5pgOnRSL79U65Asx5gCV5IMBrfFQEe5KKoARiNj5okiZfARxUbkzBLkPbgmCMn5uAo7"
            "gOU7xBSxNIjqH1AfbAIbkj7U8h2iCwG9XAgNoklLnWKVVUveqXhpW8jo6W7wsEXkWASm9xiabOaX"
            "Yhuc5hsno0GeTCQY4xR9fdjjlTCaWxwFQppRpoLnMkoVQrAyspdipxi5Q1QBoJpDhMwhCqwArokb"
            "EBFgXlwZrR8IR+sYXeAYlQ8dHLHSA/QhxIg0SPrEITc6RhfBKtoBDEZ2k0AfCpbFwKhAeqJwIKiM"
            "nsOSew9PtgxIsw5MNx+e0mPoZDO/VCIMFgJYMjQIHBhfQX1cCRaiTuguUbrjILkW/mm9hqX0HDq5"
            "J1yQ4alWgZmiiHzM84mxkSXu0lLKkM7RRc7RAE4FaNFFALGU3Dy4FKNJU4hVM6ll0HEU+BNGUHnC"
            "kcQUOoTnWgVkmPul9hqaDGY2LAXmrQIzbEOynSLzkZrii70TSmEr5EAWvXAe9oIhhTgOHLNjhAzG"
            "6Q2bD0+1Dc609E/rPTzVMjBLFA6HrfKKp0XnK8F4GmOxB5xsWK6lfzqcI2wLUwv/DLuQHJfoIi/A"
            "pIQyWoMRpu6SYtcYhWs0Vo9wjZHDPI6Afp5VyHJJ1V6J1VjUPr6CipAAe+7woIblWvilWwxPswvK"
            "gqm5XzosccfUL8UwPvCbN9UGk2CEKWjIhFjYwy1W6QbPCdziGAXME/IEGqzywd0x3OiOfRSkjxz6"
            "eEjggMspMULnPokVnvD4RRbYBuVY+GVYw3kFZ9kHZ1sHZFgMT7cOzHIMz/cCrEJPV1q6sIopSf+X"
            "FSHskMsozkzXrU3PlCLUqk3vA18aYbmxz05vuXn7X4HvonUQCPm1KJInTtr53vtgb2x+Y9PGjevX"
            "rVv72murV7+6asXKl5cvX7J48UsLF86fN2/+Cy/Om/vC3DlzaN3CaTVTqysqS4uLFUVyeWHhxOfG"
            "PdT3Ab6oePdAeKeNuJu3AbnBC/2l+tZzF1r/ON169MStX369+cPhG998fx3A7KN919/98Pob25tf"
            "29T48prGhS9fnf1S49QXGkunNxRUXMlUXp4sqx+bUffkpAuPPHsewCN2VF34yDp/yfmBEWf6BJzx"
            "HHbabfApl0GnXQadcRlwxqn/GdGDaLZ9wU5b+Z629Dlt4X3K3OuUmeep3h6nenmc6umGZup6ytTl"
            "ZDen410djncRHTcBsz9uYkemtmAnGLNhzZqYFdhxE8vjJhbHTcyPmZjxrDe14yZmemaubRasWbJm"
            "xZo1tWPEjpvYaJudttnzzO4YY/asiQza8S4OhsyRsa58czrezZDBcq6PiQO5dNToUdmSQ7VmjTs1"
            "q2MmlsRghp6p3QkT0QkTcgDdnI+bOh/v4fIH3CMzT7xr1n1O2/c74zzwjPuQ017Dz/QNOts//MLw"
            "mLqQxLqoR+okj19IeOrCY89efDoZ6PpKauHV3JIGRXVj5azGGfMaXljSuHhV06vrr298+8Y7u24A"
            "ZH7+Vct3P908fAQev9ZTZ9ou1LVdaWhrvq5ut6Tc3bXODoQvrXvflKk5QVKMokKYqalK3648eG+B"
            "kBqtDGESmOn9aNn8tR9cJVUmW2+2TF3+zpKNu9WYZ//7l9Z/dIu4d56vrXdOVLWvEGqkQpKB5n2S"
            "eUL+4hsmQdnE7xSXw06BCbfvY+qhMRfIQJZRnN5ouZlSuXTdjn1qXlCWwUb7NzbdGPjUNPPIAjuJ"
            "glaMQJEwOh/Zj6k7n4NGJDubaBo9WMQLIFQYAkIF8Rql6UYL6FDAcrZR2fZIdNmiaI3Z48JsO1wL"
            "GJDLaIMMDeJQMLVjM5dSqRB6Qn8R1hnPEcG2kVk4VAxWb7dHxU9mS3mPeJzaYT4VqgQSA8BDK2Rp"
            "kyiQMfk0FygVDGEQe+gTW4DdsGehHRmHYCGqhSKEJWLkGDBnJu49lxYMxKEIbQJ9iSQKEVEIbSJl"
            "H+7/md6g309d+PCLQ2u3fz5/7c5FGz6Et+1jp7Qy3dO7/MX3v9tGZMKYGD0Ir2sSFZCVg5hhQhFw"
            "CB421rRwiMlzis1zislziMJ8no74UUZTuQDRiRguVaBSxzBhgQP1X41Fc4Tjx7SfuU5kHMcYQBpU"
            "3mAXDoRmCW4Vs4lktOsQSosdCBMCvAFqkpGBIQtQtYvFbR2ILkf7O7EbOgETSgkTIhbitkwN+hjY"
            "kNnWScI4qaKuSKodUt9RJu8olqOQoyooJngmVRHCJIwnxnoPTqReBe3jJFaSDkCDpc7xpTgFCsJh"
            "VbiKxCgCFjojGRa4iAtcwWLRXMT5YM5kFfahfqcSMhQaGUGCyWOcSVAi7MuFEB0sdwF0iSsjxjIh"
            "Yh4elau4yFVc6EZNgtlKXVD6oyhIdEVJMaZ7kShdmJFxWFdKegQy0RhXTyWsciEdCObBEp0+xZpu"
            "pI8rxuzRPmWu8WWwL/f44l+O4y9c12/cPHL8/K7Pfnpt2+fzXtv18ua97+z99kwtOtNyP1VQ17tt"
            "uw8CxLphLtNixDkMIKzQjiGsdEukYYSMSEiD9FzhAEiNRDgMsqQUjsEdBUaG96jmxpiUYl4prqVh"
            "ioiFJWShiu3DOHySVWUkhw1DdB5SYBW5h0TuLi5yExd5SIo8cYnCHeVE3BCJDnAiAbHNI4E4czJM"
            "WOpODoMUeABTEKMzSnfckGIk1qZHGRBjAmEoVOpgFYYO4hIiDyZW8YCQTskSsi/iO1pKigqi8ulF"
            "LQ6mCmL4EfbogeeiImMCklV4Ycn4amp0KOJBWkpTj3pKVSRPpgq3lchhCvOUu5DQGHWRk/IIFsJy"
            "goXInHEqn7hiHzLvFV9CaK0M6I4gXyWZlgNVklygxPAjWYU0OMWH4JwXUfm8CKzSweEAfOJL6FH5"
            "MPPFXrhTIuUxNAjb1vigTUG2TKgi7IcJRRHV4jXCIOVGn6RqH8KEuDye7ZZQ7oNLqmCVDw7CrMXI"
            "wPjSPgnFvvHFPnB9pCrfuOI+8SW+8SWIVehcSjxFMZfMFEqDvoYyyvx9MYRaQDi9j1YpwmpGISS1"
            "6b3heY7M75+orL1koMrZn293BIRggHBAaGnJKb8e0cT8c0PdunWrpaWl6VpTfX392TNnf/v112++"
            "/vqTjz9+d8eOTRs2AjWuWrly4fwF81+cN3b0GG9SPv4eAGFT860/Tt88erzl+0M39nza9NaOxtUb"
            "GuYvb5q18FrFzEbFlMaCisvZyrrJeXXPpl8cNe6CeNQ5f+npfsFnPYefdxtyzuGhszZ9zlj6nLHw"
            "Pm3ueaKn6wlTlxPdnU90cyJgYEdACIDBXBulGHAiIAEdbI4hctgTAiEs0VUPYzQk43y8O2umLtRO"
            "UOvhyprbiZ6s9aLmbsB6U/PQmBmYZ8fMQ++jB2+GG00z+EkzME9i7Iw5Na+TFoyd0Jg3txA70HG4"
            "k9KcKVxwV7wI3V3wgsDF0Qa8E2DAhAwb2/Pp7pgG58AsiPF5GDH4hIm5nlkQ42YstJcwN/oEY71P"
            "4DhgvY6a9DxqYmrcepAOaMdMeh7HaQ/WcMlRHAEGB4C3O9FFdKKrEzxmJ3u5nbHwOmPle9b2gfOu"
            "g855+53uH3YufGTdI2Pqnsu4lCVvlFc1Vc1pem1T25Wrmpf7O2+dFwjpl0jWtFUmg9mMMlQhZGpO"
            "5Ny25sS9chnVN4A6TAkTkBk8ftbuL34CKsOfk5qbbl1vBqPB7R/vPxSd8gImnpEYSCpj0MyjCt4j"
            "GVAK526GwSkQYgaaOIVJYFbhPO0IAT0gZK8bO3O7K0xx8cOvfoH9UlnPBowRCQtIVhhU6tCDFGZi"
            "uFwyRdpZZAy6jHKRhATkOCwkGWJYztQzsjvb6ALiKYo0aCeW20mUdgwTamMhGYo6hbJGPhKPVjtu"
            "BFaoZOGwgENN9OREU9iJCRnGsNRHutlhJpgi1tuT2TuLhfl2USxhRrGcyTigyoibaKG9mGEwNInC"
            "JrrAPUFx+ny92vj/1rzkk0z6h2eKFvXA0hcwYJFIogQgFCF9IRZSzhQRT1QRiU4UIcKh2UdhdQqy"
            "hE1Fg1CnJEaQUoJMSDRDdtto1pgNC9C5lBEGFY6cNshHQU3lCcQwMrgC6RFps4jdFlU7hvoAwOIY"
            "FCTGMiGyosqRQUpKhjCCwpGlUOJlWurE0CONPFRoTNOn2Jn0QfCjAYoSMogYHTvpWmcKhHFlxErp"
            "EmYTsZJHkoWOPGPYEqCRuIkSR1AYh8CetJQ4oKJTKEleqmRkQ8S2EheyLw0Tsj2R+mIZzCOyoZxF"
            "QaUzj/fIESLLObPG6H5AcQTkKO85U9dQCeFAdr+0A9NHCwt1+wBoiaLlAc/MuNZ0o50vDU1xQjID"
            "3Bif/Lx1aA6mGEUgBBos10JBRiHkFZ8gvOdKQhbZZDYlBPMQq9ypwBhPiI64j2qM6aNR8Bh0lJYy"
            "HQgKkkHKGYkPOuBaCo0qN7HSjRIgup4qkRUR5wgKwshAZchmlR6JYDwmjC9zpxqgpJigo4pO3SUl"
            "sNAjnqp/FQTq0NxJihcqGBIULCfOoiSdDM0oowFCtgQFTTBDsNADNycjo/8nMQk7I2XxEgW6CpKi"
            "pgppELiLynFEcvRk6IvQKY5DjW5eguGCVOIDoiObs2oe5UmaDKaMCSxkrBTdOFFOZDZhPUIrkQDj"
            "CQdSQ5arYpS9ETUAdUTim6LdmWFIMiyFSRyBoCAHeDU+I4hRJkxifEcJ0VUyKIjVIKp9GG6sYftU"
            "MWvjGVxk+yCgAuOhnBiPWEgOA7ixlDWGM9FNFGgQtuK0QdZf9B+rQ6inEPpqAWGVpvKEROEZW+gZ"
            "nf/VD5pUcPew3SkQUia0t7bxdveICgsPDQoO9g8gtdof7E+yyDzo24eLCfRx9/R0dXNzcnYWOVCH"
            "UnsrG1tLKxsLy3vjMkriYq6/sv6Prg4nrX2AT070cEFBycQWCIG838ObfW8ypYRgyQpcdiy2OTCc"
            "xhIaj8pYGDMAWlqYhMaHtJ5uf8JcGethyHBVBwbRhUYdXNTjQGat9sETTjtBeZVBNbhQrGTHSHB2"
            "5FLbEDCzOKYRGyk2U1nSnECaFcqbXRxPdnc52dPtZC93LaS08Dxn3eecTZ9ztn3Pix487zTgvOvg"
            "Cx5Da739an0DTgMj9Qs+MzTmbHDCuehHz8f998IjYy6Meq72iYm1o9PqxmYg6qfkX8pRXcoprs8v"
            "qy8or5dX1iur60umNJRObyqfiVYxs6lyFlhz5ezmqjlgTZWz6RItq+BZObWZ19Bw/nLp1PqSmnpV"
            "db2iqr6oAnZ0SVZ6KbfkYlph7diMC49POJ/09LmYx84GxZ8eEH72gaDavkHnPYeddx543uGhcyKw"
            "fufs+5226XPS0os5d7gO8LyRnx7g+hw1MTuKJGl6zKTrMROToyYmv5uYwMwx0+4w01BSQ575uxTJ"
            "OykQtjGvHS3xqdNNhoy2DEs2ZzLK3EGK0b8OCKkHqZ1U3iMst1eE7GnVyz8dIdFfbfi9eeKPc2lT"
            "1/aOlPUIzTWWZfSOgBA4zTwy/4H/qzrNC+MxBoRqrLtwyViiEX5jivJNW9ctJAf2hUBIigraMC6a"
            "gIWUDAtwJpYsQQmRKzxoTCHkWaycV7m+EKsXYkAgKnKIl9HEYqiRVQQ4Ge0OUI3SIGOEDzksZJQ9"
            "bYst4FDQHvsr2Bk5qwfiPA8FlTAlRrrFct2KaAc0CRjTDUdjkrvQ3KEF9gQ+EQIpQPI2JDSIah6Q"
            "knloTtTE2TRxfGsbky7/Ji/PJM2I2MoUfsD7MnPF9h5+k1mJD/1FUSGUFgMTMlIhIiJRCwkZ0qwz"
            "Ir4xRKpkaBBgkuVJREo226cI2ZW3CeExinMOTA4Y2LBESx6M0xUJaZ4YelSO3Ia0RgXZyonKgwwN"
            "8oxBMmYERy73jJSLOWQBUtOzmLeWQ0F+txInvQ7IYxQF43lMSLDQiUeSVPHTNhVVIJ1ZqnRhqZJV"
            "/yjssUZwi+HA+HJiTE9iJUxgoZQlPfKRSn/OTJ9ypDW+ysfOE4orZyyOPyxnZSwNMn2oaY3AoCCu"
            "dU8oswzNH6tapWbzstKH8xabo1LzcLKVY/ALatZ684B0dykR+uJKGPlODwjZKXUcJdDIyIBlFPPc"
            "CAoypMeVqYgv07ZytkMlvxsCG48n3SlZURDFjwQLcV/UsbOEAiQyHtkQxT1GGGRRkDOy0D2eyHcc"
            "GcaXMgAZX8aiYCXjDppYyYzDHANzJGzoIBs9yHmNJk5hk44iExIDdKRYSMmQTsuYJQxb8lAQaZAC"
            "IWHCJE4trES6w87lmkBBipFkBJbokL40IzCkh9t60R2R/owkyNvEm1X/NLGCqFVWcaso0XmPYEiM"
            "wUKCnWSoSuqqSo3GHHqzI7M0OJUYx3sUC6sJMVYxgJdU40vMR2NT2D7VtAMZDbmOwUJgQjQiG6K7"
            "aSVjdEkiMzKJG5xGtUHCgQTD/kkgnKHlMjqSV4owkVd5Qqz0kcrtgzLXbv1M/RfklbkLIKQhhTB1"
            "d3bxcHEF41LL0OwynHEeoRQR+elG9ctL3BUQ4sfGpauBRhhlr4crQ0QGtDIdeLsrJLuHZgDk3LSl"
            "Px0Q5WEnI6yxnEbFNEQ1QrnohGlPHDVtqZh2TOMhaaHNbGaE1ixIB5sTJvZAJie7OKAo1w31qz96"
            "up428zxj7nXG0ucsMJvowXOug8+4Dz3tG3BmUMT5oPja2P9geOH/ja99JqWOBOldzVKin2RBeaOi"
            "+lrZjKbp8xrnLGpYtKpxzcbmN9+5vnNPy74vb+z/5sbBH1q+P9Ry6MjNX3+/dexE6x+n286cazt/"
            "oa3uovpSvbr+ivrKVXVDo7rxWtu1JvSibGn557PFdKShgtTS1tyM4YVw/PWX4Yzazte2nbuAcaSn"
            "ztw6eerm0eM3f/615YfDLV9/d2Pfl9c/2Nu87f2m9Vsalq1ueHFp06wF16pmNyqnXJWVXkwtqB2b"
            "WTs2vX5Cdste/Aa460jFTgqE9Afp3/84/9AjBT38nrMI5QMh+otaaFKMyu4MCP+cy6iuBympN9gl"
            "KEskVZQvfnvfgZ9fXLvL8xFSv54kEbXk1SGccGdAmKUBQixToYAlZUswkSnzda9fh5BctIXr3vOQ"
            "pgU+pfz2Z8xxb+zHQlp06ssfjwPgWbP5Qm3ETGlBG+qqqm02mrSiHQFCfql6JEONXkdxjplhBrfT"
            "cCBuyEIgC4QSYnpkqGcM7DEwyfIki3wKhuvEuMqeNYKFHBnqmKYbh4UagGTBEo3ZNcuBVNAjwOYY"
            "p+zml174/CY16wvKVYlgpjdv8aOzrt9oKVuw2TwwxT4yB+vao94INMgSHTEKhw4aMlQQYlQS5ZB0"
            "FhPjhEFmwxIWC4s5fmNM01kH54r14gZ1aJBXk1DbnKQlGk9RjTZYqoOFjlKO4kr4iWeceJvwjN+Z"
            "3wEJjc47c5DGzKA5abRBYEJmxkmbDFnBkBJmCbukhKcuInHREYD06Dg6PMZAWjxHg+XaWMj0Z9hS"
            "yue9Mi2MjMMpoh1ZyCh+8XRhhQv6YZbzjAeBjFXwplpGE8DQrTwSy3oG5sxbq8kg2makmgtTiaSh"
            "KatmjVVQhmtsEfE+pf6iugGE2qllGJ0QOY1RAsspJWpQMLESLaHCiFXyDD968PmQgbFKOggzz6a0"
            "IVZO+JBOK8i2FYZQsJqPhZ4JPMyLr2SUQLqJBgV5lkC2IgPyllezCuEUnlUzGUfZHDMsGVKc42a4"
            "j1UaDtRCwSk8JuTUQpbWEohsiAyG814GSHKK14gaL8JsLOZpj8DMT+F4j5H+KBYS46Eg0QZZomPm"
            "GaGP7oKAH5MzptpbM2yNRhhkaJBlQlYtJPvV4J8vr6cvi44MQ46gHabqGEeYPtTRlAFLHl5qhEFM"
            "3+LLoOBfFUB4Jy6j04yUIqz0pX6wpPIETTRavfBNdacBQl9jeUT17I7GvDuF8NrLr6EbnpmHRtr6"
            "+1lOX6BrR1ekQhyMg46pzoxPaRcHEk3HuURas/zGyJvH0ZsR4wYxyK2Lwx+mLn/0dj9t6Y1xiaJ+"
            "Z5wGnHYZdNpj2Glf//P9Qy8MjakNiKsLG3Eu4uHz4v87P+Lpuv8bV/90yuVxWZdT8q9kK68WVTaU"
            "TW+omXt15gKAkIZFK6+tWNv82ubrm7fe2Lbz+nsfXd+9DyPcDnx789sfb/70y61fj7ae+AOQBnPk"
            "XG1oa2rCTCpwO/5BQjOYTIjNwtJGrGMFOe7ANMPqlO5oa7u9594/2jorEJIvoL0HDlmGTDALmmhO"
            "a06Ep5GS9B1NMfo3ACEXWGgRXdAlOBuIrltwjlmkjFBfIVeY/usfsEb8uIpXqBZ3F0CI1S+iC50T"
            "i386elZN2c+IQhifUmMy4HGTfqOWbtylZmsYGLrC+GCOKlrePSTXXqq0YYHQNpZjQoYPcUZTg77j"
            "QGgo9SiZ2pG9UC9QZkarfIU2/nHzEi3Ga9c0vMfMkCXs5lo02GFT8bAQpwT2GKPLCY/RnmgiKQOE"
            "DhKFGVb/e3HnZz+duVDfatxl9PT5S4vW7wp8qsJ02ESkQeqASoU+DdTxsJCYgxYccobUp7+JFhYS"
            "bHNgkruwxi7EbnHa1Mf3F9VxHNWFQ26TUkc9bdCRT3fUdzROx6H0zoyPgk5a82yHeA0EMjPxLAfy"
            "VzHQWKptZZweyHQmdIdLeFjojDNkuUEU1FmipR/qdqDYxltSQQiQXZXAMvZS4tsAACAASURBVF5C"
            "hUsCR30VGgJM0DO6lpsnM24ECN3iSuyjip5Vrfz46yO19Q1G/Zlb246dujBn1btDR5Wa+6e6igtJ"
            "1KKKpJPhgLBSHwg100QWC3nmxnAgWJV7AjevbYmcVRFjPybw2I/t4EE7MNGAXAcKgZX8hXooWK2N"
            "hdUs12FPHfDz5DZEz89qjyQ0T30OTKrW1Qb5TJhIpEJN6lEWw6ieRj/y1D8676mLgtRqvHQUP/6G"
            "BMDIRz7IMZSF246Ywsc8g/PeIxjeY2Z4I6Axq2p4KAgz7LwW1DFioDcf5LQ4UIcJpxrrgJg3Ek0f"
            "/3yZ5dN8kO70O9TwTGsVhUBfVo4zqA3+vTGE2qUIDSUa9eYlGn1OTirZ3tVLVzvtroHwb7PbA+Hy"
            "NcBOCFp/CcVpC4wcy5E4Q8bRtJvzsW5Oxxiuo9FulOgsiZ9qLzDiuWpO4tMsSW4Yh5Omrn9YeJ2y"
            "6XPK4aFTLgPPeQ6t6xdSOzDi3LCYC5EP1yY+dfHplPqUgsuFlVer51ydveDa/GWYUHTNpqaNb19/"
            "94Prez69AbR2+MitYydbTwOn1bddbWxrvIao1nwdizrcZEHlb2gcmPGLKOqUSTRa+fBOSjgyJRk6"
            "K3vdWS1KvRonmuoUvNIUtD+9bn+idVogxHu5ZMMumlHGnAYQhmXwMsrkaacY/btdRvlmzeQgRQiE"
            "qQ3KhmQhGQTg8Nuf0LM/ZeprJoOSafnBO3IZZTORKnqE5UkyFjTQaJ+bLQYVwp2ffjcifVpK5dKz"
            "tUYj1lrIL4gzXt0J+EppUAOEYo1aqEuAd6YQ6kuFrEkMdNMiQLHePJ8DJUbgUKKhPh4KKu4K//Qg"
            "kEd6POP3UdlLCQcyKEhMjJ6Z9hhTV9g7OLN3cLqbRBYzYXpy5YqyBZvmr31/zbZ9K97YU7PkrbwZ"
            "a6STprvEZJsOG987MMUBE9sgDaIbp0RBR9OGOhUxDRly03atRO9jieFu/MISHUTBOAPmxPiLskGD"
            "+sgn1WAh182INtgOEHICIKcHlvFgj4d28WW6U2453+LKtOREnnEo6MyJfgzOlbGmM8+nPn1WLNdB"
            "QT4QMjMJeiiojX8uPER01ULEci0UZM2Nm8LxYESfyjI01y48t+8I5cj0F7JqVk9Z/PaSDR+te+fz"
            "lzfvnbp0a87UNeIJ091j83oPT7YLy3bDFDhyV6TBEtd4Lp1MhcEYQp1590TSh52yKFjlrjPPh8AE"
            "Q6bFh4QDE6r4+h5dy7qAaq3StmoeARo0I1uxFMfTALE/N2WQL2kKDwgN6IReGGjHX8LyYRIPF41y"
            "oD4W1rTXYYQ2DY7g4v0oFnILWfZL4iCQdpvKbjhVs5wxWDXVh6cQso6jnFrIU/y0+hhDQc6m6S8h"
            "Xp3T+DNIhjhl5n3ZVUwHwnjEC1R3FbucUQV9R1De09IG/0mXUcOJRmv6JHKJRjV5ZVwi8kKfrL5O"
            "Kmfe2zd8CoT79u3z9fLWKSLfSeyhvg84ixxefeUVtTEgXLH2hIndSUvv9hAOjEvcoomIo56WDmzC"
            "EttjbP4YYkBxvXg5RczJKhvoDCPgsGaep6x8T1n5gF2we6AW48QePOM57OygyLNhSbWPPFs3Ibte"
            "Vna1cvbVmuebZsy7MXdx8wtLGhetalqzsWnLO81AdN/9eOv4ydaLl9qIbyQDci037z3vtPGYjSMQ"
            "3Zrsxkuxt49tQvs3tE4KhBRjUquWsxllSIrRcJJiVDejjMyYPPh3AiGd6hvgX69w2du7sZjE2XO1"
            "qgVbgBi7BWfD1FYPC28LhDYSuUn/SeWLtkGfG9eutV43HEPY7oVV00i27Z/8aBaZj9QnkRsGwtj2"
            "abDjQKjVwY6ZKmz1qU+sh4ISHv6JeTOGIJCPjhog5JxCtfW9O8dCjgCpDKhkkE/CQiDnI8ouZ7AQ"
            "nVHldiT9pihaZh2e1TsgpdvQCaRKxHM9hk/oPnScyYBnuwwe1ytgsnVYBmb7jM6zj2JpEKtNKDka"
            "7ADydQQI9RRCVtBz0Hh+GkJBY0BoTCHUg8B2NcCSdj92HAjLeD6iehyoD37x/KhCQ93idFGQEwY1"
            "+EeXxxnkwNuphYZoUGdeyxJ4mKfrEWoA/zQKITE3DgvhI/VfJZk8XcVyUWSeVXBmb78U0yGTeg2b"
            "bBGQ2nv4ZNMhE3sNT7YOznCMysOEqGKgQQVNCYO5Q5nygwacRQ0AYTsCIH/GKPuxBJjA4B8DgQkM"
            "AXoYIUBWOaw2RIBVhmhQs9CTnXrqzWt8PvXhUIvlqrU8RbWNiH6o3VFU80zUozj9JR3FwhrOvNkp"
            "i38sCvKn2t34/fkaIK/zVM2MRh7UkJs3J+4xWEihkenTARTUxUJfhgY5L1CNHmgU80ZwpMct1zNm"
            "1XRfI86i/zQQzjCaaDSxkgFCqcpbrPCIKfARF/7M+RDdu0ZrIO/evdvGwpJWCOxs5mhnb9ql67Il"
            "S9TGYghfWnHUxORYNwteLpNebC5QmsuEZHQ0dTnZyx3rB5h7nrPyPWf7wFm7B0479j/r7Xf2odCz"
            "gdJzMY/VPjq27pnUS8myy9mqellJQ1FlU/G0pvKZDZWzG2ctaFy4onHl2mubtjbv3I3ulD8cRpw7"
            "fbb17Hn1hTr1xfq2+stYyuL6jbZ7Ujagja1sbpDZ2qlpLgCb0PRaZwRCLlglanxV16FjSEaZZC6j"
            "jHmHM8r8bS6jOkDIm0dXz94RsqGjp/34ywl6dt8eOjqqcKlpaG7P8Dx7KZUTb+MySumxK6p5itKX"
            "3j595jx+4eOVwpQPOn/KreS/gjbSdC4s5oog3Xd+fth1RKlFFJaasBEbAUIDPqJ/EggVuiioj4US"
            "Q6alCvKnBuBQFwWZGcWfgECdedYjVIOFOLWXanMgDwiJSIhMKKLFA0mVCEew6FwRKZ7hSD6Smhm5"
            "BAXz7WMKuZQwhuTBPw+EjDaoDYEsHLZDfbfFQo2VslNdFDSQWubPGXUZ5YhOExlojAP1YM+oPMjr"
            "wHMHpUDYIcy7U9NCwQSUBxnq03h7lhvgPX0INGJu/CnxGsU4RsxBiiUuXLFQexEp0lDoGlsAU3f4"
            "yHAgQUEmblCHBnX9RXVdRhN5NMjHPwb2WORLqNRyCtVFQS0s1IFAd8NSngHT7mlMGKT4V8XG+PGA"
            "MEnfHdSQJeqkFdVFQW0sxD7clIHAjqJgjd6MrrH+n4yrpxcf83Rgj099I4hsqKMEjjCgEBoEPN6S"
            "9v1C29cGdRVCThv00YI9Pv5NNYB8OlhohA8J/vFR8J+OIaQoyCqEbBjhVC7RKIYRcnllJHJRUOa6"
            "7Z+r73UYIX2X+PXXX5VF8mKFshNaqao4Pzfv008/VbN6Jv+VCCY3PvmibmLOxRzVlfyyRuWUxtJp"
            "DRUzG6a92Pgiln27tmZj8xvbb7z/0c19X7Z8/V3Lj4dvHvmt9djJtlNnW8/Xtl6qVzc2qpub2/6K"
            "cDgO51qIcVzHl+AEihPa39I6IxBSf9Hvj5xwk2T2ChhvETrZgpdRxjwyx7JjGWU6AxBak8QzwH6i"
            "OEXlkq0X6y6pSd3CN3cdGDZ6uklglkV0gR0LfvpACJQF057hsh5heU/Il31/+Bi9RJ8f/EW54M2z"
            "dVfU5LdAzEnC+0mQT4Osa3Erl11m3vo9ZhH5OjT4bwVCPTMChFrWASfSjgAhpw0i+xkFQpLPU0Tz"
            "0NB8pNFMzUMsWhhFixbKyMJ8REHog5k/5ZynqEhyT7TB9oGQpxB2CPbuayA0YiwQMkz4rwZC1mWU"
            "SS1DdUKsSyFGLCQm55mCyR9Di0bElXKZaQyi4J0AobYAmNAOAeranQOhjmuoQRqs0mNCLemvw0A4"
            "RTuRTEeAUDO9c0mQrwdqaYNGFMKOAqEB11Bmyrcag6LfHWqA7TMhZ8aAUBv8jAHh7awjQPi3xhDy"
            "FUL9vDKJlSSMkOSVkSj7xCmt/dJVvOxlQvurGr5U6UTB6flVtt66vTonNKF1ptZ5gXDTzi96+o0l"
            "JemTLcIoEHIBhB3KKPP3u4wacxy1JSlhTIKy+j1e9fIbewEI4QSvXL4yfcUOl0RVl6AsWxJYqAOE"
            "TgkqwEWTwOyA52Zu/ejrNuL5cP58XeHcTUBfJn4Z/f9bs+zNT5uvk5gB9uqR75k2OlXzFkL7/Idj"
            "j+Qv7R6Sa0MdUHk0+PcDYYdQUNdllAM/o0Coj4V/IoWMYSDUQkFtLGSmnOnohAwWsoUrGOPm5SR/"
            "jJxJE8qMieGCIqnqHjGhDhBSCNRxGTVSb/AOUJAXQ3gHLqN/Bgh1XEbLnP5aIORcRu8lEBp1GU1g"
            "gTBBDwj5qWI6hoWcQsiKhAwTagpdMEULi2nRQldMhcpyIClZ0b4weHuXUV0BkA0U5LmD/i1A2H7o"
            "oGEm9NB3Fr2NNqiXXNQICnrdPRDeHgsNAGE7KHjHWMjMG3IZvZdM6MsCoU4MYYdQ0Ji/qMams1Nd"
            "FOwsLqO4LSZB1eSVSahiwwgBCFU+EoVTaG7ipDlUHrz3IWZtba2duxnLicW8A2nSS7YTGtd6ezdL"
            "oQntPm2dEQjpX7Vi7utdh4y2pPJgaBoGEGqAMFcjD/4LgLCIBhPaS+W9I2Tdg3Mezl306Tc/q0nR"
            "wl+PnR5f8WovogFCz/f3fQ8L5S++YRKYZRqa6zqiZPYr7zVcvQoLbzQ3L9u8p++oSqxpEVMkilP2"
            "jszvEpQdOvH5FVs/P3mu/qae+6iauE9cbmje+cXh5ypXA3BiTlGJwlYbBe9vILwrJrwXQMgXCckg"
            "bGJSBasZKngVCxVMzlKtKEQNDd7rGML7HAgZzVAAwvZdRtltaZ1A1ziODNlihrx5Ny1VsEPaYKcH"
            "wtszoQCEAhD+w0DIzyvDSzTKeo1iolEur4x7dEG/ePmJM3VqI/nkhCY0oQnNWOt0QEi/xZqab8Ql"
            "T+0yZAyTUYYrSR+RY8FklKEBhLJ2/EU7icsoGyJIc8xgNGD3kByzyPzM6etOn7mA59x664PPfoiY"
            "PNfEP/OdvZh+JnPGOpOhaak1a48eP0MuSuveL3+KSX2xa1A2bGgfR8Q9EluIDqVhed2Cc1ySSgLH"
            "zXokf+kzJaueBite9UzpqicUK6JSXuwzqhJAFPpQB1R9FPwfcBn9K2II23UZ1WFCqRYcspGH2vjH"
            "40Ama6hEk0pUcBntNC6j90kMoRYT0iXxnPErFpZrViXcngY7v8uo+12h4P+my6jxGEItl9HbxRDe"
            "K3/RTuEyes+x8DYuo/rFJ7SAsBLL00tpXhkMI3QIyd7y/lfqv6AaodCEJrT7u3U6IKT+ogcPH3eJ"
            "TesVMN48hAJhOluSPscyMqeDAYSdDAg1WGgnQZwzCczqM6piwesfNDQ0wCk3Nze9sGbnZ9/8AvMr"
            "3/zkrQ8O0Aty7MRZIEPgQNPQXAAhWwlTNZ4WJyRYqLCTKiyiC3qFy0xDcgH8uoUQC84B8uwZnmce"
            "VUBQUGkMBf8HgPBvjyE0qBNKDZIhXUiLCqqYqYRqgxzIqQQg7DRAeD/EEPLRzjVew4eu+mu1NuF4"
            "7/ZM2GmBUIghFGII/00xhNqVJ3giIQ0jrNIJI7TxT1fNEcIIhSY0od1x66RAuGzzh921KhCm8SoQ"
            "5nakAmFncxnVZkJSRkIqN4vKNwnIjJg8d+/+n2gBiVvXm1uart1qwUqD1xob565+321kSZfAbFK9"
            "UGHNoqA1Dwi5KbCiHYFDMHtqhKMMOoj+40D4PxFDaMh3VBcOKQrqiYT6qqAQQ9hxl1EhhvAOgJCT"
            "B5mZSrKw0tUw8t0BCnZ6l1EhhlBwGf03uIzqVCPUhBGSRKMJlVphhGG54rEzaHEpwWlUaEITWsdb"
            "pwNC2p5VLtStQIgBhNmoEHY4gLCTAyHWkxDLAd56hOX1jpA9U7zi0BEsTYFf4q23duw9GDR+lkkQ"
            "piElfp5FHA1qmJADQo3J2amcupV2BAXvbyD8x2II9ZnQqPuoSqMNslKhQQIUYgg7CIRCDOEdAyHf"
            "dL1D78BT9F8FhEIMoQCE/wYgHDmDzTLKK08/ghdGCEAoJWGEYoVnTKGPuODw7xhscm+rEQpNaEK7"
            "v1vnAkIaQHjybN2gUUWmw8dahGpVIEQgpPJgVIf8RTuly6juEisSCgj01TUw2zFeWb7o7fXvfjGm"
            "ZCVQIhhXqFAXBXUUQn0U7PRA+D8RQ2gMC6U6CiHHgZp5HfwTXEY7jcvo/RNDaAQIK/W8Q28Pgf8i"
            "l1EhhlCIIfyXxRBq55XpO5KWIjRUjVCi8JUq7AIzX3njE7UQRig0oQntTlrnAkLqL/rmh/t7B4wz"
            "D57IBBCGp1tEaANh9O1L0v9bgJAze6nCMrqgS1B2t+Ds7sE5NP2MNYt/1rcDQmuDTCgA4T8eQ3hv"
            "gFCIIew8QHjfxRD+jwGhEEMoxBD+y2IIdYGQ8xrlhxEyIqGvVGkflDW2aKn6L6g8ITShCe0+bp0L"
            "CKlCmFmzwmTwM5ahkwkQpqI8yPcX7XAAYed3GdUxEigop2ajv5aHgroxhILLaGeOIdT3FG0PCJmP"
            "Qgzhn3EZFWII7wUQ/lkU7PQuo0IMoeAy+m9wGeUSjT7MTzSqXXwinik+4SNReEQVPJSoPCkUnxCa"
            "0IR2J60TASH95jpbWz/8v8ruw8ZahiZrgBALTmTzCk7cph79vxQI2zcDOqEAhP/GGMLbA6EQQyjE"
            "EApAKMQQCkAoAKEhIKRMOFKr+IQPeo2WUq9RYEL7wMzX3v5ULXiNCk1oQutw60RASP1F3/n4G7PA"
            "ceZBEzUFJyJowYlsyyjtjDK3kwf/XS6jd4aCQgyhEEMouIwKMYSCy6gQQyjEEP4vxBAa9hqtAdPy"
            "GiW5RtFrNDBrgmK5WlAIhSY0oXW4dSIgpN9ceTNepf6iJKMMzS+aYRGZZX7n/qL/U0AoxBAKMYQC"
            "EAoxhAIQCjGEQgzhfRhD+DBbm/5hFgh1K9QzXqM+EpW3ROEeVTAgSXXq7EW1wIRCE5rQOtY6CxDS"
            "76xzdZf9/qvqPozkFw2h+UXTaToZyyh9IPzfdRkVYgiFGELBZbQdl1EhhlBwGRViCAWX0fvLZZRl"
            "whGcSMgrPpFYRbxGy3zQa1RJc40uXb9bLXiNCk1oQutY6yxASP1Ft+4+YBY4nskvGkb8RcOpv2iO"
            "xZ37i97HQCjEEAoxhAIQtgOEQgyhAIRCDKEAhPchEBrzGk2oYgsSairUx42fdaPlplqoUC80oQmt"
            "A62zACFVCLOmriT+osnm2vXo0V808o79Rf+nXEaFGEIhhlBwGRViCAWXUSGGUIghvD9jCJnl07S8"
            "RpO0U8vEsallxApvcZFjaPa7e7+DN6tbrYJIKDShCe02rVMAIaXBU+cuDh5V1H046y8ammpuLL9o"
            "x/xF/6eAUIghFGIIBSAUYggFIBRiCIUYwvs0hlC3+ERfphphTd+kKX2ISEhzjRKREL1G7YOyxsmX"
            "qoUwQqEJTWgdaJ0CCKm/6Mo3d5sCDQZPYvOLpnH5RbX9RWUdlAfvY5dRIYZQiCEUXEbbcRkVYggF"
            "l1EhhlBwGb3vXEZ5uUax/oQmtQyba7TSl0stI1Z4xhT6igu///kk95YlNKEJTWjG2j8PhPS3q+Yb"
            "LUnpM7oMHWNB/UVpOhkD9eg7WoHw/gZCIYZQiCEUgLAdIBRiCAUgFGIIBSC8H4FQq/5EH93UMgCE"
            "5T5SpiChr1RhG5BRPm+LGoFQ8BoVmtCE1l7754GQ/nC1+6ufbMIm9Q6aYB4ymQXCDJpflMiDebx6"
            "9H8OCO9Tl1EhhlCIIRRcRoUYQsFlVIghFGII7+sYQjbXKD+1TJJOahnChKT+hGuEzH9Uee2lq2rB"
            "cVRoQhNau+2fB0L6HZU+ZYXJ4NGWoUCDyWw6mQxzLD+YTdLJ3HF+0f81IBRiCIUYQgEIhRhCAQiF"
            "GEIhhvC+jiHk5RodYawgIak/IcHUMr4SuV1g5qtbPlEL9SeEJjShtdv+YSCk8uC3Px93l2b29P//"
            "9s49OK7yPOMyFxPvOapLfAt2jbW7Umeaia9g7+pi3VaybkwzSZtOJVt2LMmWL1pdLe1KtizLtmQr"
            "oUAxTSi0JCHhljjJZAgQaBMHwqQwaQen0DKBFkoNCSFAoa5xsGWdnu/7zuU7Z8+evWhX0u553jmz"
            "c2yLTP6xRj8/z+99d4gyEPp3CUV7xCJ1nQzpi9KEMFVAmBWVUTiEcAhRGbWpjMIhRGUUDiEqo9lZ"
            "GdVNwnHuIOGYdpDQox8kHPQEQssKO+tab718mdAgMkIMBhNtZhkIWYeh50v356xulGnQpZ0fZOtk"
            "SkwCYQL7RbMbCOEQwiEEENoAIRxCACEcQgBhdgJhAxcSKibhGDMJtdUy5CBh9UE3vT/hDgws9Qcf"
            "O3NWwv0JDAYTfWYTCK9QGnzp1XM3bgnOV+NBsahdtLk2MX0gzNLKKBxCOISojMIhRGUUDiEcwmx3"
            "CC2P1POrZWQmZAcJlfsTi33Bv+i8axZ/2MNgMHN/ZhMIWTw4cNsDOaubKA22kWsTxcq1CUEHwi51"
            "v2hPQjToKCCEQwiHEEAIhxBACIcQDqEDHMKI1TKaSVh7jL8/wY7Ur6roX1HS/cwvfiVh3SgGg4ky"
            "swaELB789/98M6+mc/6G7TIQilo8SK5NkPODSjxIEsJe9fwgKqNwCOEQojIKhxCVUTiEqIw6tjLK"
            "sHBcO1JvWi3jrTmimoQsJAwv8gV3hu+VoBFiMJgoM2tAyOLB0O0P5qxpEsmpCXJtQijay/qiajzY"
            "ndz5wawHQjiEcAgBhDZACIcQQAiHEECY1UBouW70OHeknq0bVY7Ur6rsX17S/dPnX5ZgEmIwGKuZ"
            "HSBk8eDLr73lrlXiQcHPH6PvEDd3kkcviyZsDzqtMgqHEA4hKqNwCFEZhUMIh9AZDqH5JqH5SD1Z"
            "LWMICZf4g38ePDUrP/JhMJi5P7MJhGESDzaSeNBvGQ92iYblogBCOIRwCAGEsw6EcAgzGwjhEMIh"
            "zHCH0BoI+fsTmkno5kxCdyVZN/roj1+QEBJiMJiImQUg1G4PrqzqILcH/W2iX6bBdgqEajxYysWD"
            "pT3J0WAWV0bhEMIhRGXUpjIKhxCVUTiEqIxme2XUeH+iXgZC45H6WmISerSQkN4krGn58kcXP5Yg"
            "E2IwGOPMmkO489DdOaubcgujxoPqMfpkrk1kPRDCIYRDCCC0AUI4hABCOIQAQgcAYcRNQqU1epwL"
            "CQ+rISFhwkUb99/z8BkJISEGgzHOTAMhiwcff+aFhYUtCzbuZDQoFrYLij0YJA+JB7v184PJxoOO"
            "qozCIYRDiMooHEJURuEQwiF0kkMYAYTGdaOemlHPlhFqEg7lBQbdgdCKzb3rPzt87jfvSepuPwwG"
            "g5FmGAjZd5//vXCxetfYvLVb2SV6oZCURQUSD3box+hJNtgj6vFg6hLCLAVCOIRwCAGEcAgBhHAI"
            "4RA6ySE0MaG+bjRfDwnVdaPUJPRUhT65cf/grd+W1H+gx2AwGGmGgZBdRL3zgSeuXbdN8Lcq8WDR"
            "HoHeHhQ2BwVzPNiTW57M+cGsr4zCIYRDiMqoTWUUDiEqo3AIURl1RmWU2y7DbhLW68VR4wmKgywk"
            "XFV+IK+877mz/yHhTj0Gg1Fn5oCQ/VvUy6+95a3runZDM10uukuxB0v2CSW0L0rUQSUeVPeLJlkW"
            "zW4ghEMIhxBAaAOEcAgBhHAIAYTOAsJbuNZow3h+/Vi+foJCCwnZCYrQEn/wc/v++tLlSfmnMqSE"
            "GAxGmmmHcEpqO/y3OasbRWWXzC6haI9YRHfJEHuwMyIeTN4edFplFA4hHEJURuEQojIKhxAOocMc"
            "QhuTcIwUR+uO8ttllOJoYGDRpv33PHJG/rlschIhIQaDmSkgZPHg6aeeEzftXLBpp14WJfHgfqGE"
            "nJrI1eNBZbmoOL2NMo4CQjiEcAgBhHAIAYRwCOEQOs8hZH9qvElIQ0L9Tj13gsIdCLsDoT/a3PuZ"
            "hqFX3/itBJkQg8HMDBCyM/Tn3n53wxcGr163TY0H25VTE8VcPKjQIG8PojIKhxAOIRxCOISojMIh"
            "RGUUlVFbLGxQcsICjglpSHjMW3vUWzPi0UNCUhxdvKnji6F7pui6PxAhBuPwmQkgZMtFO8a+lrOm"
            "USxsdfnbxMJdQuEedmpC0C/RazQ4rduDTgBCOIRwCAGENkAIhxBACIcQQOg4INSKo+btMsdM22UU"
            "JgwMyEx490M/kbBdBoNx/KQdCNl3mdNPPS/6drpYWVQ5NbFHLKa7ZErYZlG+LKqdH4RDCIcQDiEc"
            "wjlVGYVDmNmVUTiEcAiz0SE0MmH9iYJ6c0goA6G35ohny4inejhPLY6uLO374+qBsy+/IaE4isE4"
            "e9ILhKws+vpb76z5/MA167ephwd382VRwWKXTA/DOTEdCWGWAiEcQjiEAEI4hABCOIRwCB3qEPJA"
            "yG8crR/31tFT9bVacXTYwzaO0rOES3zBhl1/df7C7yWcqsdgHDxpBEL2neXS5ckdQ1/JWd2YW9jm"
            "UjeL0kv0WllU2SUjGk5NpCAezOLKKBxCOISojNpURuEQojIKhxCVUedVRvXtMgX0yTeeoCBASIuj"
            "HqU4Oshkwk9u3Hf4ju9KCAkxGAdPGoGQfWe566Enr7up2eVrUe9MkLKoQMuigloWVfuiPWKZdoY+"
            "BTSYxUAIhxAOIYDQBgjhEAII4RACCB0JhBMsJyy4Rd84mm84S6gXR5WNo5WhvIr+5SXdP/zpWfln"
            "tknIhBiMIyddQMjUwZ+ffWV55b75N23nz9ATGiw2Hh5k8WBpD2cPpoAGHVUZhUMIhxCVUTiEqIzC"
            "IYRD6GyH8KQGhAXWZwn1U/VqcZSEhDIQ3vz5kf9663cSckIMxpGTFiBk301+/c77xc0jV63dmlvY"
            "JvgYDbZrZVEChLQsKpZ1KdlgxC4ZOIRwCOEQwiGcY0AIhzCzgRAOIRzCrHYIlafgFu0EBflUmJCE"
            "hMepSchO1Q+7q5RT9V56haL5wN2XLk9KkAkxGOdN6oGQfR/56OLHp9hSdAAADlRJREFUTQN35qxp"
            "Uq8O0s2ixTQeJGfoDWVRRoNieY+xLAqHEA4hHEI4hHAIURmFQ4jKKCqjcTuEHBBGFke9ukxIi6P0"
            "MmFeYJAsmCFXKPaPf/VRiRRHAYQYjLMmXUB46NQjV6/dKvhbOBrcQ7JBog5qhwe7TGVR0cCBAEI4"
            "hHAI4RDCIQQQwiEEEAIIEwXCk2zdKF8cpScotAUzTCY8rMuEASITriju+t5Tv5B/ipuchEyIwTho"
            "UgyErCx63/fPuDbtWLBRuzq4W1kkw6mDuaUyCuqX6EXDZlE4hHAI4RDCIZyblVE4hJldGYVDCIfQ"
            "AQ4hA0KSExboxVEWEo6rxVF2hYIxoSoTBsIrNvd+pmHoxVfOSVgwg8E4aVIJhGyRzJM//+Wy8vb5"
            "G7aL/lbB12agQVYWZfFgWTd/akK0KovCIYRDCIcQDuEcA0I4hJkNhHAI4RA6wyGMKI7yp+rrj+ez"
            "7TJ0wYynmjHhYF6AyITLCrsqto2/+fb7EhbMYDCOmZQBIfuu8dKr//0nn+1jN+jp1UG6SIbQILkz"
            "wamDJB5U1EFll0yfaA4GURmFQwiHEA4hHEJURuEQojKKymgSlVGlOEpzwpOKTFivyoR12oKZI/Qy"
            "4TCRCdmCmUB4sS/4Zx2nPjz/kQQmxGCcMakBQvb94u13PyjfeXQeXSTj8reSsmiRtkiGvzPRxWhQ"
            "BUJ9s6iRCQGEcAjhEMIhhEMIIIRDCCAEECYNhIbiaAH7v0FCQpNMOOKuHlav1Yc8VaFFG/fvPfx1"
            "LB3FYBwyKQDCK/Q7xYfnP/pC3x2UBllT1HiD3kyD3dQb1M/Qp3ydjNMqo3AI4RCiMgqHEJVROIRw"
            "COEQ8kBYwNdHG7ilo/XjtDhqkgnJgpm8QJguHQ0t2rT/0O2nJQqEQEIMJrtnukDIvkn8/uNL7aP3"
            "zlvbRLxBskhmt0qDe9W1okGeBsVS850Jy2wQDiEcQjiEcAjnGBDCIcxsIIRDCIfQSQ6hBoSsOEpD"
            "wvoIJuRyQnXp6KC7MuSuHFjq67jtvh9J6pIIDAaTrTMtIGT/ZnTlylTotgeuXrdV8LUIDAiNi2QE"
            "JRvsVNTB0h7WF2WbRVUU7I3AP1RG4RDCIYRDCIcQlVE4hKiMojI6ncooeygNyi+kODquKoW0O0ov"
            "EyoLZlhOWH3QXUkOUayq6L+hqOu+089IWDqKwWT1JA+ElAanrkxNHb37e9fd1Lxg006BeYMsHtRv"
            "0BtpULUHRQUFdSCMiAQBhHAI4RDCIYRDCCCEQwggBBBOHwh1mZD6hGTjqFe9Vs+Y0Fsz6qk5oi4d"
            "HWLHCW8sO7CqrPc7TzwvgQkxmOydJIGQZYPyt4YjXzl93U3b6cnBVpeWDRbtZTfoxZKgSoNdBhos"
            "01AwXetknFYZhUMIhxCVUTiEqIzCIYRDCIcwAgUnTFgYZcEMXTpayzPhoJscJwytLOvLq+j79uNg"
            "QgwmaycZIGTe4OXLk4dOPTJ/Q/OCTV8U/G0udoC+kFsrWmJeJKOURQ3ZIBxCOIRwCOEQZhAQwiHM"
            "bCCEQwiH0HkOIQPCCfpoMqFyiILkhPXj3NJRxoQjek5IF8ysLO3NK+t9+LHnJByiwGCycRIGQkaD"
            "ly5dDt/x4LXrGQ2q2SBHgy62VnQzd2RCo8Ey+kRURuEQwiGEQwiHMCWVUTiEqIzCIURlFJVR9ZOh"
            "YCQTkke/T0hkQnXpaO0R1SdUuqOMCVeV9T706D9JlAmxdxSDyaZJDAjZhYmLH1/qmbj/mvXbmDcY"
            "mQ26Io5McItkFBSEQwiHEA4hHMI0ASEcQgAhHEIAIYCQA0LtU5MJ6UvDycSYsKxvZWnP33/naUlh"
            "QkAhBpMlkwAQspLA+QsXd4/ee/XarS6f6g1yW2RcWlN0cwwahEMIhxAOIRzCDKyMwiHM7MooHEI4"
            "hA52CDkmbDihYCFhQtPSUbU7WsMz4SBjwhvLDywv7vqbB/5RoiEBmBCDyY6JFwjZCZr3Pjy/LXzX"
            "VWubZBp0+VpdPnNTVKZBcoNeRsFIGmQoaC6LwiGEQwiHEA5hBgEhHMLMBkI4hHAIHewQRssJ+e7o"
            "GJcTjqpMOOypOuQOKHtH6S2K4Jf/7jGya35q6gqYEIPJ/IkLCNlSqdfffKdu74l5a5oEX4vLyhtU"
            "aJBlg8pjTYOWlVE4hHAI4RDCIUxJZRQOISqjcAhRGUVlNKIymiATmnPCg3THjMyE4byK/qX+YP/J"
            "h85fuChhzQwGk/kTAwjJv/3Qv+f/8m+vbWw8mLO6USQoqNEguzAh0+A+7gA9Q8Go2SAcQjiEcAjh"
            "EKYVCOEQAgjhEAIIAYRWQBiJhVpxNIIJ6y27o0PuAKmPuisHFvs6tvZ+9e3ffSDhHAUGk+FjB4Ra"
            "E+Dxn71QcEv3VWu3iuS8RKtAvUGXkg2yk4NBkc8Gy7rUtaLdsbJBOIRwCOEQwiHMoMooHMLMrozC"
            "IYRD6FSH0JIGeavQmgmVm/WGHTMHaXeUKIVLfMGqHRMv/uqchDUzGEwmT1QgZMHg5OSVUw/+aFFJ"
            "2zXrm0k2SKTBXcr1edO9QYumaFQahEMIhxAOIRzCzARCOISZDYRwCOEQOtghnIioj2pMaJcT5vP3"
            "CSkTeggTDroDIW9VaFlh57o/Hf7hmbMSmBCDydixAEJaEyXR/4f/d6F74hvzNzRfd/MOsbCN0aCr"
            "sN1VpDVFo9CgbTYIhxAOIRxCOIRprYzCIURlFA4hKqOojEapjJreE8oJR70KEw4zJsyrlJkwvGJz"
            "T155313f/Af20yOUQgwm48YMhPJfY1YT/ddX3qjde2Le2iZ2bFAgZVGZBlk2uC82DZYZDtDHA4Rw"
            "COEQwiGEQ5gSIIRDCCCEQwggBBDGAYQxmNCr3icskJmw9rjKhEc8W0YIE7I1M4Gwpyp8Y/mBZYWd"
            "Xce/9d7/nKc/TOJ0PQaTSaMD4RRXE/3GD55213XmrGkS/C1MGpRRkEiDlAZdJdq9QeMB+mSyQTiE"
            "cAjhEMIhzKDKKBzCzK6MwiGEQwiH0AiEEwYa5A8ValFhPX+OgmPCiDUzS3zButZb//nF1yV6pRBR"
            "IQaTKaMAIQkG6d/bX7/z/p6j937i5u3XbGgWC1sVGvTvdhW1s5ooRcEOO28wsWwQDiEcQjiEcAgz"
            "CAjhEGY2EMIhhEMIh9AKCK3IMCoTHss3MOGwR2VCqhR2fbp28P7vP8sSQiSFGExGTI78F/WKelX0"
            "iWdfuPkvh3JWN7KaqMvXSqVBWhMt2usq3qvTYGmnAoRlxgsTcaAgHEI4hHAI4RCmtTIKhxCVUTiE"
            "qIyiMhpfZTTasploTHhcPVt/1FtD1sx49dWjg+5KmQnDK0v7lhd3dR371m9+Sy5SaC4SBoOZs5PD"
            "/pa+98H58O0PLixquWrdVlYTdSk1UfnZ46LSoEs5NsjVREutaJB/yo2fiVFiby6nF6bw+QMrVqTP"
            "AeMnx4Q8GaaYEpPkQ8vwMMozoH8SShzQP+OmxDj40IYVE6dE+yfAf8aODS3Dw2Qf68DQkhVjP1X2"
            "JGmpF6awQWrNikuNn7EfEyXGFx6mIyq0DQ/1CNHu2WL8jJsS+TZpCgNDS1bknhH10/jUGD+TAkWr"
            "J3lKtGRFu8cUGMZHiZasGMcTFxzat0ntHgM9msLD2LHhNB7rwNCGFWM/sSnRECGmjgwtKZFnxXFP"
            "PI+JEuNKDk9Ehodp65GeUPlQJ0brx0SPyVAirw7a54SmNTPj6olCvT7qNSmFan20tHHsiad/yX7c"
            "BBNiMHN5SGX0x8+/VLz9cM7qxk9s3EFvS9C780pNlNIglQZdJUFKg0GhtFOgKCgQb7Bbo0FBuUHf"
            "Ew/+RauMilpaWNYnqrEhXyhNaPFMJPXlqpXRXHNlVHnJZS/lEfhXrv6+hojqiyX10d/vp1/Qv7Bc"
            "fjG0RtVfRqPBAzbUp+eEleonqYwOsNboQg7nFhoCQ6u0sIL+PmFFSn0VBq7TIND0+xr1/WEi1Mex"
            "34AJ8K6ndEe+gOaB1xvjweu1/LDSkvo0GgxR3gtZOoQ6EFbGkyLaUR/7o8X0KynXhbn0L2xJd4si"
            "c0L9ZYj+hxT2AnxllOWEg2plVHtRgTCQhG0YrTLKhYQq0S0xINyQ9r7ERHdVxhcdEQ/RF/nTlA2q"
            "L1XaC32mkSJGqYwOL9tC/4jQ2iH5k74on59S3g/FQD4NDqsp9VUbeE9rjX6qmodAnQaTqJWaKqPL"
            "LSqj+nNDZE7IvzBWNP6HluCn/Y78lZbUt1ynwcgX63KpORu0rYyuMFEfHxgq/6H2S+WPLKlPo0Hy"
            "UmP6gmgO4ajxU3nhqG/U0BpV6W5llL7oyjpjQlirvrOXWvZyjH/h2O+ojnO13EutjnkRlBgtG9Rf"
            "LKmP/H49/YJ6K11QcwjrbLLBsbw6ynvkf5CiWt1xy9ZonvH340sRLR1CmeLGVPYbi1IZHdNYzh2Z"
            "EyZWGWUv4x7lc9xbrwSJpuQwCVY0Z4NK+neCvpwgDVLWI6UvXvZLvlzKc6AVIlpSXz6FvXz20qD8"
            "UvuCiCxRvUVhxYQF2urROrNSqNVHPVXhG0q6V5X1Dt/+3XfpphmURzGYOTs5Q3c+fH1xGz063yKw"
            "YJBuE5VpkHREixkKajQoo2CnGgzaZoNlVtlgYmLhDOWEEU9EQmiZEyYoFkZ5UtMgjSsnNDVI40sI"
            "Lduk6UsI431MYmFlYtlgsjnhYExWTD4hNLdJY4iF6UwIzTlhjGywOkpOmGCbNG09UnObNHY2OO02"
            "6RxKCBNpk9rmhNPNBhNJCEf1TxMlxsoJp50Npjon1EqktZFt0njFwmk/cYmFCWeDCUiG6c4Jk00I"
            "480JtYQwLQ1SmzZp7CfBNmmUbNBGKYy2ZoblhOOaUkiYkFyul7HwsJdFhQFl+2heRf9Sf2f1jokn"
            "n30JQIjBzNn5f94mnrFkBolKAAAAAElFTkSuQmCC"
        ),
    },
]


# ==============================================================================
# CLIENT EMAIL DIRECTORY  (baked in - managed via the 'All Client Emails' button)
# ==============================================================================
# === CLIENT EMAILS START (auto-managed - edited via the 'All Client Emails' button) ===
EMAIL_LOOKUP = {}
# === CLIENT EMAILS END ===

# ==============================================================================
# CLIENT EMAIL MANAGEMENT  (self-contained; the list above is edited in-app)
# ==============================================================================
SCRIPT_PATH = os.path.abspath(__file__)
CLIENT_MARKER_START = "# === CLIENT EMAILS START (auto-managed - edited via the 'All Client Emails' button) ==="
CLIENT_MARKER_END = "# === CLIENT EMAILS END ==="

# ==============================================================================
# REMOTE SUPPORT CONSTANTS (bug reporting + self-update)
# ==============================================================================
APP_NAME = "Ocean Cargo Hold Notification"
APP_VERSION = "1.0.0"
DEVELOPER_NAME = "Atlas Ramoon"
DEVELOPER_EMAIL = "atlasramoon@gmail.com"
BUG_REPORT_WEBHOOK_URL = "https://discord.com/api/webhooks/1524620703259951104/fqpIEBXVWsKHy7f1iZ9xoryCpidmjPYIDuITfcwMOjBfMyS2HtJNWpVbfOetapl8vw9O"
UPDATE_MANIFEST_URL = "https://raw.githubusercontent.com/hugging-phace/mbe-updates/main/manifests/ocean-cargo-notification.json"

# Global state for the support / update icon
_support_btn = None
_pending_update = None


# ==============================================================================
# REMOTE SUPPORT HELPERS  (bug reporting + self-update)
# ==============================================================================
def _version_tuple(v):
    """Convert a dotted version string into a comparable tuple of ints."""
    parts = []
    for p in str(v).strip().split("."):
        num = ""
        for ch in p:
            if ch.isdigit():
                num += ch
            else:
                break
        parts.append(int(num) if num else 0)
    return tuple(parts)


def _http_get(url, timeout=10):
    """Perform a simple HTTP GET and return the response body as text."""
    req = urllib.request.Request(url, headers={"User-Agent": "MBE-OceanCargo/1.0"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def _post_to_discord(content):
    """POST a text payload to the bug-report Discord webhook."""
    payload = json.dumps({"content": content}).encode("utf-8")
    req = urllib.request.Request(
        BUG_REPORT_WEBHOOK_URL,
        data=payload,
        headers={"Content-Type": "application/json", "User-Agent": "MBE-OceanCargo/1.0"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.status == 204 or resp.status == 200
    except Exception:
        return False


def _generate_case_number():
    """Generate a human-friendly case number for tracking bug reports."""
    short_uuid = uuid.uuid4().hex[:8].upper()
    return f"OCE-{short_uuid}"


def _post_bug_report_with_files(description, case_number, file_paths, reporter_email="", category="Bug Fix"):
    """Send a bug report to the Discord webhook, listing attached files."""
    file_list = "\n".join(f"• `{os.path.basename(fp)}`" for fp in file_paths) if file_paths else "_No files attached_"
    content = (
        f"**NEW {category.upper()} REPORT**\n"
        f"**Case Number:** {case_number}\n"
        f"**App:** {APP_NAME} v{APP_VERSION}\n"
        f"**Reporter:** {reporter_email if reporter_email else 'Anonymous'}\n"
        f"**User:** {getpass.getuser()}\n"
        f"**Machine:** {platform.node()}\n"
        f"**OS:** {platform.system()} {platform.release()}\n"
        f"**Developer Contact:** {DEVELOPER_EMAIL}\n"
        f"\n**Description:**\n{description}\n"
        f"\n**Attached Files:**\n{file_list}"
    )
    return _post_to_discord(content)


def _post_update_applied(old_ver, new_ver):
    """Notify Discord that a self-update was applied."""
    content = (
        f"**UPDATE APPLIED**\n"
        f"**App:** {APP_NAME}\n"
        f"**Old Version:** {old_ver}\n"
        f"**New Version:** {new_ver}\n"
        f"**User:** {getpass.getuser()}\n"
        f"**Machine:** {platform.node()}"
    )
    return _post_to_discord(content)


def _check_for_update():
    """Fetch the update manifest and return (new_version, download_url, changelog) or None."""
    try:
        raw = _http_get(UPDATE_MANIFEST_URL, timeout=10)
        manifest = json.loads(raw)
        latest = manifest.get("version", "").strip()
        if not latest:
            return None
        if _version_tuple(latest) > _version_tuple(APP_VERSION):
            download_url = manifest.get("download_url", "").strip()
            changelog = manifest.get("changelog", "No changelog provided.")
            return (latest, download_url, changelog)
    except Exception:
        pass
    return None


def _download_and_apply_update(new_url):
    """Download the new script, splice in the local EMAIL_LOOKUP block, and overwrite."""
    new_source = _http_get(new_url, timeout=30)
    # Preserve the locally-managed client email block
    try:
        with open(SCRIPT_PATH, "r", encoding="utf-8") as f:
            local_source = f.read()
        start_idx = local_source.find(CLIENT_MARKER_START)
        end_idx = local_source.find(CLIENT_MARKER_END)
        if start_idx != -1 and end_idx != -1:
            end_idx += len(CLIENT_MARKER_END)
            client_block = local_source[start_idx:end_idx]
            new_start = new_source.find(CLIENT_MARKER_START)
            new_end = new_source.find(CLIENT_MARKER_END)
            if new_start != -1 and new_end != -1:
                new_end += len(CLIENT_MARKER_END)
                new_source = new_source[:new_start] + client_block + new_source[new_end:]
    except Exception:
        pass
    # Write the updated script
    with open(SCRIPT_PATH, "w", encoding="utf-8") as f:
        f.write(new_source)
    return True


# ==============================================================================
# SUPPORT / BUG-REPORT UI HELPERS
# ==============================================================================
_tooltip_window = None


def _show_tooltip(widget, text):
    """Display a small tooltip window near the given widget."""
    global _tooltip_window
    _hide_tooltip()
    x = widget.winfo_rootx() + 25
    y = widget.winfo_rooty() + 25
    _tooltip_window = ctk.CTkToplevel(widget)
    _tooltip_window.wm_overrideredirect(True)
    _tooltip_window.wm_geometry(f"+{x}+{y}")
    _tooltip_window.attributes("-topmost", True)
    lbl = ctk.CTkLabel(_tooltip_window, text=text, font=(MODERN_FONT, 11),
                       text_color="#ffffff", fg_color="#1a1a2e", corner_radius=6,
                       padx=8, pady=4)
    lbl.pack()


def _hide_tooltip():
    """Destroy the tooltip window if it exists."""
    global _tooltip_window
    if _tooltip_window is not None:
        try:
            _tooltip_window.destroy()
        except Exception:
            pass
        _tooltip_window = None


def _on_support_click():
    """Handle click on the support icon - opens update dialog or bug report."""
    global _pending_update
    _hide_tooltip()
    if _pending_update is not None:
        _apply_fixes_dialog()
    else:
        _report_bug_dialog()


def _refresh_support_icon():
    """Flip the support icon between bug and 'Apply Fixes' modes."""
    global _support_btn
    if _support_btn is None:
        return
    if _pending_update is not None:
        _support_btn.configure(text="Apply Fixes", font=(MODERN_FONT, 12, "bold"),
                               text_color="#b8860b", fg_color="transparent",
                               hover_color="#1a2a40")
    else:
        _support_btn.configure(text="\U0001f41e", font=("Segoe UI Emoji", 15),
                               text_color="#ffffff", fg_color="transparent",
                               hover_color="#1a2a40")


def _check_update_bg():
    """Background thread: fetch manifest and update the icon if a new version exists."""
    global _pending_update
    result = _check_for_update()
    if result is not None:
        _pending_update = result
        try:
            _refresh_support_icon()
        except Exception:
            pass


def _restart_app():
    """Relaunch the script in a detached process and exit."""
    try:
        python_exe = sys.executable
        if platform.system() == "Windows":
            import subprocess
            subprocess.Popen([python_exe, SCRIPT_PATH],
                             creationflags=subprocess.DETACHED_PROCESS)
        else:
            import subprocess
            subprocess.Popen([python_exe, SCRIPT_PATH])
    except Exception:
        pass
    sys.exit(0)


def _report_bug_dialog():
    """Show the bug-report dialog with category, description, and email fields."""
    win = ctk.CTkToplevel()
    win.title("Report a Bug / Request a Feature")
    win.geometry("520x600")
    win.configure(fg_color="#0a1930")
    win.attributes("-topmost", True)
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f"520x600+{int((sw-520)/2)}+{int((sh-600)/2)}")
    win.resizable(False, False)

    # Header
    ctk.CTkLabel(win, text="Report a Bug or Request a Feature",
                 font=(MODERN_FONT, 18, "bold"), text_color="#ffffff").pack(pady=(20, 5))
    ctk.CTkLabel(win, text=f"{APP_NAME} v{APP_VERSION}",
                 font=(MODERN_FONT, 11), text_color="#888888").pack(pady=(0, 15))

    # Category
    ctk.CTkLabel(win, text="Category:", font=(MODERN_FONT, 13, "bold"),
                 text_color="#ffffff").pack(anchor="w", padx=30)
    cat_var = ctk.StringVar(value="Bug Fix")
    cat_frame = ctk.CTkFrame(win, fg_color="transparent")
    cat_frame.pack(fill="x", padx=20, pady=(5, 10))
    for cat in ["Bug Fix", "Feature Request", "Environmental Change", "Other"]:
        ctk.CTkRadioButton(cat_frame, text=cat, variable=cat_var, value=cat,
                           font=(MODERN_FONT, 12), text_color="#ffffff",
                           fg_color="#1e90ff", hover_color="#1c7ed6").pack(side="left", padx=(0, 4))

    # Description
    ctk.CTkLabel(win, text="Description:", font=(MODERN_FONT, 13, "bold"),
                 text_color="#ffffff").pack(anchor="w", padx=30)
    desc_box = ctk.CTkTextbox(win, height=140, font=(MODERN_FONT, 12),
                              fg_color="#152238", text_color="#ffffff",
                              border_width=1, border_color="#1e3a5c", corner_radius=6)
    desc_box.pack(fill="x", padx=30, pady=(5, 10))

    # Email
    ctk.CTkLabel(win, text="Your email (optional - for follow-up):",
                 font=(MODERN_FONT, 13, "bold"), text_color="#ffffff").pack(anchor="w", padx=30)
    email_entry = ctk.CTkEntry(win, font=(MODERN_FONT, 12), fg_color="#152238",
                               text_color="#ffffff", border_width=1,
                               border_color="#1e3a5c", corner_radius=6, width=300)
    email_entry.pack(fill="x", padx=30, pady=(5, 10))

    # What to expect
    info_frame = ctk.CTkFrame(win, fg_color="#152238", corner_radius=8, border_width=1, border_color="#1e3a5c")
    info_frame.pack(fill="x", padx=30, pady=(5, 10))
    ctk.CTkLabel(info_frame, text="What to expect:", font=(MODERN_FONT, 12, "bold"),
                 text_color="#1e90ff").pack(anchor="w", padx=12, pady=(8, 2))
    expect_text = (
        "\u2022 Atlas will review your report within 24-48 hours\n"
        "\u2022 When a fix is ready, the bug icon will change to 'Apply Fixes'\n"
        "\u2022 Write down your case number for follow-up\n"
        f"\u2022 Contact {DEVELOPER_EMAIL} if you don't hear back"
    )
    ctk.CTkLabel(info_frame, text=expect_text, font=(MODERN_FONT, 11),
                 text_color="#cccccc", justify="left").pack(anchor="w", padx=12, pady=(0, 8))

    # Next button
    def go_next():
        description = desc_box.get("1.0", "end").strip()
        if not description:
            messagebox.showwarning("Missing Description", "Please describe the issue before continuing.", parent=win)
            return
        reporter_email = email_entry.get().strip()
        category = cat_var.get()
        win.destroy()
        _show_attach_files_dialog(description, reporter_email, category)

    ctk.CTkButton(win, text="Next \u2192", command=go_next, font=(MODERN_FONT, 14, "bold"),
                  fg_color="#1e90ff", hover_color="#1c7ed6", corner_radius=8, height=38).pack(pady=(5, 20))


def _show_attach_files_dialog(description, reporter_email, category):
    """Show the file-attachment dialog and submit the bug report."""
    win = ctk.CTkToplevel()
    win.title("Attach Files & Submit")
    win.geometry("520x520")
    win.configure(fg_color="#0a1930")
    win.attributes("-topmost", True)
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f"520x520+{int((sw-520)/2)}+{int((sh-520)/2)}")
    win.resizable(False, False)

    ctk.CTkLabel(win, text="Attach Files", font=(MODERN_FONT, 18, "bold"),
                 text_color="#ffffff").pack(pady=(20, 5))
    ctk.CTkLabel(win, text="The running script is automatically attached.",
                 font=(MODERN_FONT, 11), text_color="#888888").pack(pady=(0, 15))

    # File list
    file_paths = [SCRIPT_PATH]
    list_frame = ctk.CTkScrollableFrame(win, fg_color="#152238", height=200,
                                        border_width=1, border_color="#1e3a5c", corner_radius=6)
    list_frame.pack(fill="both", expand=True, padx=30, pady=(0, 10))

    def refresh_list():
        for child in list_frame.winfo_children():
            child.destroy()
        for fp in file_paths:
            row = ctk.CTkFrame(list_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=os.path.basename(fp), font=(MODERN_FONT, 11),
                         text_color="#ffffff").pack(side="left", padx=(4, 0))
            if fp != SCRIPT_PATH:
                def remove_this(f=fp):
                    file_paths.remove(f)
                    refresh_list()
                ctk.CTkButton(row, text="Remove", width=60, height=22,
                              font=(MODERN_FONT, 10), fg_color="#cc3333",
                              hover_color="#a02020", command=remove_this).pack(side="right")

    refresh_list()

    def browse_files():
        from tkinter import filedialog
        selected = filedialog.askopenfilenames(parent=win, title="Select files to attach")
        for s in selected:
            if s not in file_paths:
                file_paths.append(s)
        refresh_list()

    ctk.CTkButton(win, text="Browse Files", command=browse_files,
                  font=(MODERN_FONT, 13, "bold"), fg_color="#1e90ff",
                  hover_color="#1c7ed6", corner_radius=8, height=34).pack(pady=(0, 8))

    status_lbl = ctk.CTkLabel(win, text="", font=(MODERN_FONT, 11), text_color="#1e90ff")
    status_lbl.pack(pady=(0, 5))

    def submit_report():
        case_number = _generate_case_number()
        status_lbl.configure(text=f"Submitting report {case_number}...")
        win.update()
        ok = _post_bug_report_with_files(description, case_number, file_paths,
                                         reporter_email, category)
        if ok:
            messagebox.showinfo("Report Submitted",
                                f"Your report has been submitted!\n\n"
                                f"Case Number: {case_number}\n\n"
                                f"Write this down for follow-up.\n"
                                f"Contact {DEVELOPER_EMAIL} if you don't hear back within 48 hours.",
                                parent=win)
            win.destroy()
        else:
            status_lbl.configure(text="Submission failed. Please try again or email the developer.")
            messagebox.showerror("Submission Failed",
                                 f"Could not submit the report automatically.\n"
                                 f"Please email {DEVELOPER_EMAIL} with your issue.",
                                 parent=win)

    btn_row = ctk.CTkFrame(win, fg_color="transparent")
    btn_row.pack(pady=(5, 20))
    ctk.CTkButton(btn_row, text="Skip & Send", command=submit_report,
                  font=(MODERN_FONT, 13, "bold"), fg_color="#2a3a55",
                  hover_color="#1a2a40", corner_radius=8, height=38, width=140).pack(side="left", padx=(0, 10))
    ctk.CTkButton(btn_row, text="Submit Report", command=submit_report,
                  font=(MODERN_FONT, 13, "bold"), fg_color="#1e90ff",
                  hover_color="#1c7ed6", corner_radius=8, height=38, width=160).pack(side="left")


def _apply_fixes_dialog():
    """Show the update dialog with changelog and apply/restart buttons."""
    global _pending_update
    if _pending_update is None:
        return
    new_ver, download_url, changelog = _pending_update

    win = ctk.CTkToplevel()
    win.title("Apply Available Update")
    win.geometry("520x520")
    win.configure(fg_color="#0a1930")
    win.attributes("-topmost", True)
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f"520x520+{int((sw-520)/2)}+{int((sh-520)/2)}")
    win.resizable(False, False)

    ctk.CTkLabel(win, text="A Fix is Available!", font=(MODERN_FONT, 18, "bold"),
                 text_color="#1e90ff").pack(pady=(20, 5))
    ctk.CTkLabel(win, text=f"Version {APP_VERSION}  \u2192  {new_ver}",
                 font=(MODERN_FONT, 13), text_color="#ffffff").pack(pady=(0, 15))

    # Changelog
    ctk.CTkLabel(win, text="What's fixed:", font=(MODERN_FONT, 13, "bold"),
                 text_color="#ffffff").pack(anchor="w", padx=30)
    cl_frame = ctk.CTkScrollableFrame(win, fg_color="#152238", height=220,
                                      border_width=1, border_color="#1e3a5c", corner_radius=6)
    cl_frame.pack(fill="both", expand=True, padx=30, pady=(5, 15))
    ctk.CTkLabel(cl_frame, text=changelog, font=(MODERN_FONT, 12),
                 text_color="#cccccc", justify="left", wraplength=430).pack(anchor="w", padx=8, pady=8)

    status_lbl = ctk.CTkLabel(win, text="", font=(MODERN_FONT, 11), text_color="#1e90ff")
    status_lbl.pack(pady=(0, 5))

    def do_apply():
        if not download_url:
            messagebox.showerror("Update Error", "No download URL available for this update.", parent=win)
            return
        status_lbl.configure(text="Downloading and applying update...")
        win.update()
        try:
            _download_and_apply_update(download_url)
            _post_update_applied(APP_VERSION, new_ver)
            messagebox.showinfo("Update Applied",
                                "The update has been applied successfully.\n"
                                "The app will now restart.",
                                parent=win)
            win.destroy()
            _restart_app()
        except Exception as e:
            status_lbl.configure(text="Update failed.")
            messagebox.showerror("Update Failed", f"An error occurred:\n{str(e)}", parent=win)

    def do_later():
        win.destroy()

    btn_row = ctk.CTkFrame(win, fg_color="transparent")
    btn_row.pack(pady=(5, 20))
    ctk.CTkButton(btn_row, text="Maybe Later", command=do_later,
                  font=(MODERN_FONT, 13, "bold"), fg_color="#2a3a55",
                  hover_color="#1a2a40", corner_radius=8, height=38, width=140).pack(side="left", padx=(0, 10))
    ctk.CTkButton(btn_row, text="Apply & Restart", command=do_apply,
                  font=(MODERN_FONT, 13, "bold"), fg_color="#1e90ff",
                  hover_color="#1c7ed6", corner_radius=8, height=38, width=160).pack(side="left")


def build_email_lookup(base):
    """Expand the baked CBY->email map into the key variants the app looks up."""
    lookup = {}
    for cby, email in base.items():
        cby = str(cby).strip()
        email = str(email).strip()
        if not cby or not email:
            continue
        lookup[cby] = email
        lookup["CBY" + cby] = email
    return lookup


def _cby_sort_key(k):
    try:
        return (0, int(k))
    except (ValueError, TypeError):
        return (1, str(k))


def serialize_email_lookup(d):
    lines = ["EMAIL_LOOKUP = {"]
    for k in sorted(d, key=_cby_sort_key):
        lines.append("    %r: %r," % (str(k), str(d[k])))
    lines.append("}")
    return "\n".join(lines)


def parse_email_lookup_from_text(text):
    try:
        s = text.index("\n" + CLIENT_MARKER_START) + 1
        e = text.index(CLIENT_MARKER_END, s)
        region = text[s:e]
        m = re.search(r"EMAIL_LOOKUP\s*=\s*(\{.*\})", region, re.DOTALL)
        if not m:
            return {}
        return dict(ast.literal_eval(m.group(1)))
    except Exception:
        return {}


def save_client_emails_to_script(new_base):
    """Rewrite ONLY the marked EMAIL_LOOKUP block in this script (atomic write)."""
    with open(SCRIPT_PATH, "r", encoding="utf-8") as f:
        text = f.read()
    s = text.index("\n" + CLIENT_MARKER_START) + 1
    e = text.index(CLIENT_MARKER_END, s)
    new_region = CLIENT_MARKER_START + "\n" + serialize_email_lookup(new_base) + "\n"
    new_text = text[:s] + new_region + text[e:]
    tmp = SCRIPT_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(new_text)
    last_err = None
    for _ in range(6):
        try:
            os.replace(tmp, SCRIPT_PATH)
            return
        except PermissionError as ex:
            last_err = ex
            time.sleep(0.4)
    for _ in range(6):
        try:
            with open(SCRIPT_PATH, "w", encoding="utf-8") as fp:
                fp.write(new_text)
            try:
                os.remove(tmp)
            except Exception:
                pass
            return
        except PermissionError as ex:
            last_err = ex
            time.sleep(0.4)
    try:
        os.remove(tmp)
    except Exception:
        pass
    raise last_err


def find_conflicted_copies():
    out = []
    d = os.path.dirname(SCRIPT_PATH)
    for name in os.listdir(d):
        if not name.lower().endswith(".pyw"):
            continue
        if "conflicted copy" not in name.lower():
            continue
        full = os.path.abspath(os.path.join(d, name))
        if full != SCRIPT_PATH:
            out.append(full)
    return out


def reconcile_client_emails_on_startup():
    """Merge client emails from any Dropbox 'conflicted copy' files into the baked list."""
    try:
        copies = find_conflicted_copies()
        if not copies:
            return
        additions = {}
        changes = []
        for p in copies:
            try:
                with open(p, "r", encoding="utf-8") as f:
                    other = parse_email_lookup_from_text(f.read())
            except Exception:
                continue
            for cby, email in other.items():
                cby = str(cby).strip()
                email = str(email).strip()
                if not cby or not email:
                    continue
                if cby not in EMAIL_LOOKUP:
                    additions[cby] = email
                elif EMAIL_LOOKUP[cby] != email and cby not in additions:
                    changes.append((cby, EMAIL_LOOKUP[cby], email))
        if additions:
            EMAIL_LOOKUP.update(additions)
            try:
                save_client_emails_to_script(EMAIL_LOOKUP)
            except Exception:
                pass
        msg = ""
        if additions:
            msg += "Merged %d new client email(s) from a Dropbox conflicted copy.\n" % len(additions)
        if changes:
            preview = "\n".join("  CBY %s: kept '%s' (other had '%s')" % (c, mine, theirs) for c, mine, theirs in changes[:8])
            extra = "" if len(changes) <= 8 else ("\n  ...and %d more" % (len(changes) - 8))
            msg += ("\n%d CBY(s) had a different email in a conflicted copy. Kept your current values:\n%s%s\n\nEdit any of these via 'All Client Emails' if needed." % (len(changes), preview, extra))
        if msg:
            try:
                messagebox.showinfo("Client Emails - Conflict Reconciliation", msg)
            except Exception:
                pass
        try:
            if messagebox.askyesno("Remove Conflicted Copies?",
                                   "Found %d Dropbox 'conflicted copy' file(s) of this script.\n\nDelete them now? (Their client emails have already been merged.)" % len(copies)):
                for p in copies:
                    try:
                        os.remove(p)
                    except Exception:
                        pass
        except Exception:
            pass
    except Exception:
        pass


# ==============================================================================
# EMAIL PARSING ENGINE (REDIRECTED TO SYSTEM FILES)
# ==============================================================================
def get_template_data(template_file=None):
    """Return the embedded HTML template and its inline images.

    The wording lives in TEMPLATE_HTML (edit the text directly). Images come
    from the embedded base64 data, so no external files are required.
    """
    html_content = TEMPLATE_HTML
    images = []
    for img in TEMPLATE_IMAGES:
        images.append({
            "data": base64.b64decode(img["b64"]),
            "cid": img["cid"],
            "content_type": img["content_type"],
        })

    # Add banner-image class to the main banner image and fix alignment
    if "<img" in html_content and html_content.count("<img") == 1:
        html_content = html_content.replace("<img", '<img class="banner-image"')
    elif "<img" in html_content:
        img_matches = re.findall(r'<img[^>]+>', html_content, re.IGNORECASE)
        if img_matches:
            html_content = html_content.replace(img_matches[0], img_matches[0].replace("<img", '<img class="banner-image"', 1))

    if "banner-image" in html_content:
        banner_td_pattern = r'<td[^>]*>(<img[^>]*class="banner-image"[^>]*>)</td>'
        banner_td_match = re.search(banner_td_pattern, html_content, re.IGNORECASE | re.DOTALL)
        if banner_td_match:
            banner_td = banner_td_match.group(0)
            if "padding:" not in banner_td.lower():
                html_content = html_content.replace(banner_td, banner_td.replace('<td', '<td style="padding: 0 20px;"'))

    return html_content, images

# ==============================================================================
# SMTP DELIVERY ENGINE
# ==============================================================================
def send_headless_smtp(to_emails, bcc_emails, subject, body, images, password):
    responsive_style = """
    <style>
        html, body { width: 100% !important; margin: 0 !important; padding: 0 !important; -webkit-text-size-adjust: 100% !important; -ms-text-size-adjust: 100%; }
        body, table, td { font-family: Arial, sans-serif; box-sizing: border-box; }

        /* Added mso rules for better Outlook rendering */
        table { border-collapse: collapse; margin: 0 auto !important; mso-table-lspace: 0pt; mso-table-rspace: 0pt; }

        /* Specific class for the banner. Global img rule avoided to keep Gmail emojis small. */
        .banner-image { display: block !important; width: 100% !important; max-width: 100% !important; height: auto !important; margin: 0 auto !important; border: 0; }

        /* CRITICAL: Force the main container to never exceed screen width */
        .container-table {
            width: 100% !important;
            max-width: 100% !important;
            table-layout: fixed !important;
        }

        /* The scroll wrapper locked down as a block element */
        .table-scroll-wrapper {
            width: 100% !important;
            max-width: 100vw !important;
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch !important;
            display: block !important;
        }

        .mobile-swipe-hint {
            display: none !important;
        }

        @media only screen and (max-width: 600px) {
            body { font-size: 14px !important; line-height: 1.4 !important; }

            .mobile-swipe-hint {
                display: block !important;
                font-size: 12px !important;
                color: #888888 !important;
                text-align: right !important;
                margin-bottom: 6px !important;
                font-style: italic !important;
            }

            /* The inner table width is preserved, but forced not to stretch parents */
            .package-card {
                width: 600px !important;
                min-width: 600px !important;
                max-width: 600px !important;
                margin: 15px 0 !important;
                table-layout: fixed !important;
            }

            .desktop-table {
                width: 600px !important;
                min-width: 600px !important;
                max-width: 600px !important;
            }

            .package-header { display: table-row !important; }
            .desktop-only { display: table-cell !important; }

            /* Prevent wrapping so the data stays intact */
            .desktop-table td, .desktop-table th {
                padding: 6px 4px !important;
                font-size: 12px !important;
                white-space: nowrap !important;
                overflow: visible !important;
                word-break: normal !important;
            }
        }
    </style>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    """
    
    if "</head>" in body:
        body = body.replace("</head>", responsive_style + "</head>")
    elif "<head " in body:
        body = re.sub(r'(<head[^>]*>)', r'\1' + responsive_style, body)
    else:
        body = responsive_style + body

    wrapper_start = """<table width="100%" cellpadding="0" cellspacing="0" border="0" style="width:100% !important; table-layout:fixed; background-color: #ffffff;">
        <tr>
            <td align="center" style="padding: 0;">
                <table class="container-table" width="100%" cellpadding="0" cellspacing="0" border="0" style="width:100% !important; max-width:100% !important; margin:0 auto; text-align:left; table-layout:fixed;">
                    <tr>
                        <td style="padding: 0; overflow-wrap: break-word; word-wrap: break-word;">"""
                        
    wrapper_end = """</td>
                    </tr>
                </table>
                </td>
        </tr>
    </table>"""

    if "<body>" in body:
        body = body.replace("<body>", "<body>" + wrapper_start)
    elif "<body " in body:
        body = re.sub(r'(<body[^>]*>)', r'\1' + wrapper_start, body)
    else:
        body = wrapper_start + body

    if "</body>" in body:
        body = body.replace("</body>", wrapper_end + "</body>")
    else:
        body = body + wrapper_end

    msg = MIMEMultipart("related")
    msg["From"] = SENDER_EMAIL
    
    msg["To"] = ", ".join(to_emails)
    if bcc_emails:
        msg["BCC"] = ", ".join(bcc_emails)
        
    msg["Subject"] = subject

    msg_alternative = MIMEMultipart("alternative")
    msg.attach(msg_alternative)
    msg_html = MIMEText(body, "html", "utf-8")
    msg_alternative.attach(msg_html)

    for img in images:
        mime_img = MIMEImage(img["data"], _subtype=img["content_type"].split("/")[-1])
        mime_img.add_header("Content-ID", f"<{img['cid']}>")
        mime_img.add_header("Content-Disposition", "inline", filename=f"{img['cid']}.png")
        msg.attach(mime_img)

    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(SMTP_LOGIN, password)
    
    all_recipients = list(set(to_emails + bcc_emails))
        
    server.sendmail(SENDER_EMAIL, all_recipients, msg.as_string())
    server.quit()

# ==============================================================================
# STANDALONE & QUEUE INTERACTIVE REVIEW ENGINE
# ==============================================================================
def process_queue():
    o365_password = get_secure_password()
    html_template, images = get_template_data()

    # Replace plain ocean-ship link with a styled button
    ocean_link_match = re.search(r'<a[^>]*href=[^>]*ocean-ship[^>]*>[^<]*</a>', html_template)
    if ocean_link_match:
        pre_alert_button = (
            '<table cellpadding="0" cellspacing="0" border="0" style="margin: 10px 0; display: block;">'
            '<tr><td style="background-color: #003366; border-radius: 6px; padding: 12px 28px;">'
            '<a href="http://www.mbe.ky/ocean-ship" style="font-family: Arial, sans-serif; '
            'font-size: 15px; font-weight: bold; color: #ffffff; text-decoration: none;">'
            'Click Here to Submit Your Pre-Alerts</a></td></tr></table>'
        )
        html_template = html_template.replace(ocean_link_match.group(0), pre_alert_button)

    # Reconcile any Dropbox 'conflicted copy' files, then build the lookup from
    # the baked-in client directory (managed via the 'All Client Emails' button).
    reconcile_client_emails_on_startup()
    email_lookup = build_email_lookup(EMAIL_LOOKUP)

    raw_rows = []
    headers = []

    # Find the only XLSX file in the current directory
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
    excel_path = excel_files[0] if excel_files else None
    
    if excel_path and os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            raw_rows = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True) if any(row)]
        except Exception:
            raw_rows = []

    # Normalize header names to handle variations
    def normalize_header(header):
        if not header:
            return ""
        header = str(header).strip().lower()
        # Remove extra spaces and normalize
        header = re.sub(r'\s+', ' ', header)
        return header
    
    # Create header mapping from actual to expected names
    header_mapping = {}
    for header in headers:
        normalized = normalize_header(header)
        if normalized:
            header_mapping[normalized] = header
    
    # Helper function to get values with flexible header matching
    def get_value(row_dict, possible_names):
        for name in possible_names:
            normalized = normalize_header(name)
            if normalized in header_mapping:
                actual_header = header_mapping[normalized]
                return str(row_dict.get(actual_header) or "").strip()
        return ""
    
    # Process rows according to record selection rules
    qualifying_rows = []
    if raw_rows:
        for row in raw_rows:
            row_dict = dict(zip(headers, row))
            
            # Extract values using flexible header matching
            dr_wr = get_value(row_dict, ["D/R W/R"])
            status = get_value(row_dict, ["Status"])
            email = get_value(row_dict, ["Email", "Email "])
            cby_number = get_value(row_dict, ["CBY Number", "#CBY", " #CBY"])

            # Fallback to email lookup if main Excel doesn't have email
            if not email and cby_number:
                # Try exact match first
                if cby_number in email_lookup:
                    email = email_lookup[cby_number]
                # Try with CBY prefix
                elif f"CBY{cby_number}" in email_lookup:
                    email = email_lookup[f"CBY{cby_number}"]
            
            # Exclude rules
            if not dr_wr:  # Column A is blank
                continue
            if "SHIP" in status.upper():  # Column P contains SHIP and a date
                continue
            
            # Include rules
            status_is_blank = not status or status.upper() == "NONE" or status.upper() == "N/A"
            has_previous_sent_date = "SENT EMAIL" in status.upper()
            
            if not (status_is_blank or has_previous_sent_date):
                continue
            
            # Valid D/R number check - basic validation that it's not empty
            if not dr_wr:
                continue
            
            # Add to qualifying rows
            qualifying_rows.append({
                "D/R W/R": dr_wr,
                "CBY Number": cby_number,
                "Email": email,
                "Status": status,
                "Shipper": get_value(row_dict, ["Shipper"]),
                "Pkgs": get_value(row_dict, ["Pkgs"]),
                "Description": get_value(row_dict, ["Description", "DESCRIPTION"]),
                "CuFt": get_value(row_dict, ["CuFt"]),
                "Weight": get_value(row_dict, ["Weight"]),
                "Tracking Number": get_value(row_dict, ["Tracking Number"]),
                "Invoice": get_value(row_dict, ["Invoice"]),
                "Notes": get_value(row_dict, ["Notes"]),
                "row_index": raw_rows.index(row) + 2  # Track original row number for Excel updates (row 2 is first data row)
            })
    
    # Group by CBY Number
    grouped_rows = []
    if qualifying_rows:
        grouped_dict = {}
        for row in qualifying_rows:
            cby_number = row["CBY Number"]
            email = row["Email"]
            
            # Only require CBY Number for grouping, email can be added during review
            if not cby_number:
                continue
            
            if cby_number not in grouped_dict:
                grouped_dict[cby_number] = {
                    "CBY Number": cby_number,
                    "Email": email,  # May be empty, user can fill during review
                    "D/R Numbers": [],
                    "Records": []
                }
            
            grouped_dict[cby_number]["D/R Numbers"].append(row["D/R W/R"])
            grouped_dict[cby_number]["Records"].append(row)
        
        grouped_rows = list(grouped_dict.values())

    if not grouped_rows:
        messagebox.showinfo("No Packages Found", "No qualifying packages were found in the cargo list to review.")
        return

    approved_emails = []
    current_index = [0]
    removed_packages = []  # Track packages removed by user
    
    # Date picker memory variables
    last_response_by_date = {"month": "", "day": "", "year": ""}
    last_next_shipment_date = {"month": "", "day": "", "year": ""}

    # Auto-advance to the next upcoming shipment cycle.
    # Anchor: Thursday July 3, 2025 (known sail date).
    # Shipments run every other Thursday.
    # Response By Date = the Friday before the sail Thursday (6 days prior).
    # On each run, keep adding 14 days until the response-by date is today or later.
    from datetime import date as _date, timedelta as _td
    _anchor_ship = _date(2025, 7, 3)   # a known Thursday sail date
    _today = _date.today()
    _ship = _anchor_ship
    while (_ship - _td(days=6)) < _today:
        _ship += _td(days=14)
    _response = _ship - _td(days=6)    # Friday before the Thursday

    reference_response_date = {
        "month": _response.strftime("%B"),
        "day":   str(_response.day),
        "year":  str(_response.year),
    }
    reference_shipment_date = {
        "month": _ship.strftime("%B"),
        "day":   str(_ship.day),
        "year":  str(_ship.year),
    }
    
    queue_timer_id = [None]
    is_in_queued_state = [False]

    root = ctk.CTk()
    root.geometry("800x700")  # Landscape orientation for better table display
    root.resizable(True, True)
    root.configure(fg_color="#0a1930")  # Navy blue background

    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    root.geometry(f"800x700+{int((sw-800)/2)}+{int((sh-700)/2)}")

    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.pack(pady=(20, 10), padx=(5, 30))
    
    # Load and display MBE logo (embedded - no external file needed)
    try:
        logo_image = Image.open(io.BytesIO(base64.b64decode(MBE_LOGO_B64)))
        logo_image = logo_image.resize((120, 72), Image.Resampling.LANCZOS)
        logo_photo = ctk.CTkImage(logo_image, size=(120, 72))
        logo_label = ctk.CTkLabel(header_frame, image=logo_photo, text="")
        logo_label.pack(side="left", padx=(0, 12))
    except Exception:
        pass  # If logo fails to load, continue without it
    
    # Create a frame for the title and subtitle on the right
    text_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
    text_frame.pack(side="left", fill="both", expand=True)
    
    title_lbl = ctk.CTkLabel(text_frame, text="Ocean Cargo Hold Notification", font=(MODERN_FONT, 15, "bold"), text_color="#ffffff")
    title_lbl.pack(pady=(10, 0), anchor="w")
    
    subtitle_lbl = ctk.CTkLabel(text_frame, text="Review Customer Package Hold", font=(MODERN_FONT, 14, "bold"), text_color="#1e90ff")
    subtitle_lbl.pack(pady=(0, 2), anchor="w")
    
    entries = {}
    
    def create_modern_row(label_text):
        row_frame = ctk.CTkFrame(root, fg_color="transparent")
        row_frame.pack(fill="x", padx=30, pady=5)
        lbl = ctk.CTkLabel(row_frame, text=label_text, width=110, anchor="w", font=(MODERN_FONT, 12), text_color="#ffffff")
        lbl.pack(side="left")
        entry = ctk.CTkEntry(row_frame, width=310, height=28, corner_radius=6, border_width=1, border_color="#1e90ff", fg_color="#1a3a5c", text_color="#ffffff")
        entry.pack(side="right", fill="x", expand=True)
        entries[label_text] = entry

    create_modern_row("Customer Email:")
    create_modern_row("CBY Number:")
    
    # Date picker rows
    def create_date_picker_row(label_text, entry_name):
        row_frame = ctk.CTkFrame(root, fg_color="transparent")
        row_frame.pack(fill="x", padx=30, pady=5)
        lbl = ctk.CTkLabel(row_frame, text=label_text, width=110, anchor="w", font=(MODERN_FONT, 12), text_color="#ffffff")
        lbl.pack(side="left")
        
        # Date picker container
        date_frame = ctk.CTkFrame(row_frame, fg_color="transparent")
        date_frame.pack(side="right", fill="x", expand=True)
        
        # Month dropdown
        months = ["January", "February", "March", "April", "May", "June", 
                 "July", "August", "September", "October", "November", "December"]
        month_var = ctk.StringVar()
        month_combo = ctk.CTkComboBox(date_frame, variable=month_var, values=months, 
                                       width=100, height=28, corner_radius=6, 
                                       border_width=1, border_color="#1e90ff", 
                                       fg_color="#1a3a5c", text_color="#ffffff",
                                       dropdown_fg_color="#1a3a5c", dropdown_text_color="#ffffff")
        month_combo.pack(side="left", padx=(0, 2))
        
        # Day dropdown
        days = [str(i) for i in range(1, 32)]
        day_var = ctk.StringVar()
        day_combo = ctk.CTkComboBox(date_frame, variable=day_var, values=days,
                                     width=60, height=28, corner_radius=6,
                                     border_width=1, border_color="#1e90ff",
                                     fg_color="#1a3a5c", text_color="#ffffff",
                                     dropdown_fg_color="#1a3a5c", dropdown_text_color="#ffffff")
        day_combo.pack(side="left", padx=(0, 2))
        
        # Year dropdown
        current_year = 2026
        years = [str(i) for i in range(current_year, current_year + 2)]
        year_var = ctk.StringVar()
        year_combo = ctk.CTkComboBox(date_frame, variable=year_var, values=years,
                                      width=70, height=28, corner_radius=6,
                                      border_width=1, border_color="#1e90ff",
                                      fg_color="#1a3a5c", text_color="#ffffff",
                                      dropdown_fg_color="#1a3a5c", dropdown_text_color="#ffffff")
        year_combo.pack(side="left")
        
        # Store references
        entries[entry_name] = {
            "month": month_var,
            "day": day_var, 
            "year": year_var,
            "widgets": (month_combo, day_combo, year_combo)
        }
    
    create_date_picker_row("Response By Date:", "Response By Date")
    create_date_picker_row("Next Shipment Date:", "Next Shipment Date")

    # ==============================================================================
    # COMMAND FUNCTIONS
    # ==============================================================================
    DEFAULT_WHITE = "#ffffff"

    # Invoice classification functions
    def has_valid_invoice(invoice_value):
        """Check if invoice value indicates a valid invoice."""
        if not invoice_value or not str(invoice_value).strip():
            return False
        invoice_str = str(invoice_value).strip().lower()
        valid_indicators = ['yes', 'invoice received', 'invoice attached']
        return any(indicator in invoice_str for indicator in valid_indicators)

    def requires_attention(invoice_value):
        """Check if invoice value indicates attention is required."""
        if not invoice_value or not str(invoice_value).strip():
            return True
        invoice_str = str(invoice_value).strip().lower()
        attention_indicators = ['no pre-alert found', 'no invoice', 'missing']
        return any(indicator in invoice_str for indicator in attention_indicators)

    def update_package_table(records):
        """Update the package table preview with current records."""
        # Clear existing content
        for widget in table_scroll.winfo_children():
            widget.destroy()
        
        # Filter out removed packages
        active_records = [r for r in records if r not in removed_packages]
        
        if not active_records:
            package_table_label = ctk.CTkLabel(table_scroll, text="No packages to display", font=(MODERN_FONT, 11), text_color="#aaa")
            package_table_label.pack(pady=10)
            return
        
        # Create header row with improved spacing
        header_frame = ctk.CTkFrame(table_scroll, fg_color="#1e90ff")
        header_frame.pack(fill="x", pady=(0, 2))

        headers = ["D/R W/R", "Description", "Tracking #", "Invoice", "Status", "Action"]

        # Use more compact widths to reduce horizontal scrolling
        header_widths = [90, 180, 130, 70, 90, 35]

        for i, (header, width) in enumerate(zip(headers, header_widths)):
            lbl = ctk.CTkLabel(header_frame, text=header, font=(MODERN_FONT, 11, "bold"), width=width, anchor="w", text_color="#ffffff")
            lbl.pack(side="left", padx=1)
        
        # Add data rows with compact spacing
        for record in active_records:
            row_frame = ctk.CTkFrame(table_scroll, fg_color="transparent")
            row_frame.pack(fill="x", pady=1)

            # Tree-style expand/collapse indicator for notes
            notes = record.get("Notes", "")
            has_notes = notes and notes.strip()

            indicator = None
            notes_row = None

            if has_notes:
                # Create expandable notes row (hidden by default)
                notes_row = ctk.CTkFrame(table_scroll, fg_color="transparent")

                # Notes content in muted yellow slate card
                notes_card = ctk.CTkFrame(notes_row, fg_color="#fff3cd", corner_radius=4)
                notes_card.pack(fill="x", pady=(0, 5), padx=(110, 0))  # Indent to align with content

                notes_label = ctk.CTkLabel(notes_card, text=notes, font=(MODERN_FONT, 11),
                                          text_color="#333333", wraplength=500, justify="left")
                notes_label.pack(pady=8, padx=10, anchor="w")

                # Expand/collapse indicator button - use compact width
                indicator = ctk.CTkButton(row_frame, text=">", width=18, height=20,
                                         font=(MODERN_FONT, 9, "bold"), fg_color="#1a3a5c",
                                         hover_color="#1e90ff", corner_radius=3)

                def toggle_notes_display():
                    if notes_row.winfo_ismapped():
                        notes_row.pack_forget()
                        indicator.configure(text=">")
                    else:
                        notes_row.pack(fill="x", pady=(0, 2))
                        indicator.configure(text="v")

                indicator.configure(command=toggle_notes_display)
                indicator.pack(side="left", padx=(0, 1))
            else:
                # Empty space for alignment when no notes
                empty_indicator = ctk.CTkLabel(row_frame, text="", width=18)
                empty_indicator.pack(side="left", padx=(0, 1))

            # D/R W/R - compact spacing
            dr_value = record.get("D/R W/R", "")
            dr_lbl = ctk.CTkLabel(row_frame, text=dr_value, font=(MODERN_FONT, 11, "bold"),
                                width=header_widths[0]-22, anchor="w", text_color="#ffffff")
            dr_lbl.pack(side="left", padx=1)

            # Description - compact spacing with truncation for display
            desc = record.get("Description", "")
            if len(desc) > 25:
                desc = desc[:25] + "..."
            desc_lbl = ctk.CTkLabel(row_frame, text=desc, font=(MODERN_FONT, 11),
                                   width=header_widths[1], anchor="w", text_color="#ffffff")
            desc_lbl.pack(side="left", padx=1)

            # Tracking Number - compact spacing with truncation
            tracking = record.get("Tracking Number", "")
            if len(tracking) > 18:
                tracking = tracking[:18] + "..."
            tracking_lbl = ctk.CTkLabel(row_frame, text=tracking, font=(MODERN_FONT, 11),
                                       width=header_widths[2], anchor="w", text_color="#ffffff")
            tracking_lbl.pack(side="left", padx=1)

            # Invoice - compact spacing with truncation
            invoice = record.get("Invoice", "")
            if len(invoice) > 10:
                invoice = invoice[:10] + "..."
            invoice_lbl = ctk.CTkLabel(row_frame, text=invoice, font=(MODERN_FONT, 11),
                                      width=header_widths[3], anchor="w", text_color="#ffffff")
            invoice_lbl.pack(side="left", padx=1)

            # Status - compact spacing, no truncation
            status_value = record.get("Status", "")
            status_lbl = ctk.CTkLabel(row_frame, text=status_value, font=(MODERN_FONT, 11, "bold"),
                                     width=header_widths[4], anchor="w", text_color="#ffffff")
            status_lbl.pack(side="left", padx=1)

            # X button to remove package - compact fixed width
            def remove_package(rec=record):
                removed_packages.append(rec)
                update_package_table(records)  # Refresh table

            remove_btn = ctk.CTkButton(row_frame, text="✕", width=25, height=20,
                                      font=(MODERN_FONT, 9, "bold"), fg_color="#dc3545",
                                      hover_color="#c82333", corner_radius=3, command=remove_package)
            remove_btn.pack(side="left", padx=(15, 1))

    def reset_button_ui():
        is_in_queued_state[0] = False
        approve_btn.configure(text="APPROVE", fg_color="#28a745", hover_color="#218838", state="normal")
        skip_btn.configure(text="SKIP", fg_color="#dc3545", hover_color="#c82333")

    def load_current_record():
        idx = current_index[0]
        if idx >= len(grouped_rows):
            root.destroy()
            return
        
        # Clear removed packages when loading new record
        removed_packages.clear()
            
        reset_button_ui()

        root.title(f"Review Details - Customer {idx + 1} of {len(grouped_rows)}")
        progress_lbl.configure(text=f"Queue Progress: Customer {idx + 1} of {len(grouped_rows)}")

        row_data = grouped_rows[idx]

        # Re-enable locked fields so their contents can be refreshed
        entries["CBY Number:"].configure(state="normal")

        for lbl_name, widget in entries.items():
            if isinstance(widget, ctk.CTkEntry):
                widget.delete(0, "end")

        entries["Customer Email:"].insert(0, row_data.get("Email", ""))
        entries["CBY Number:"].insert(0, row_data.get("CBY Number", ""))

        # Lock CBY Number and D/R Number(s) - reference only, editing them has no effect
        entries["CBY Number:"].configure(state="disabled")

        # Set date picker fields to remembered values or fallback to reference cycle
        if last_response_by_date["month"]:
            entries["Response By Date"]["month"].set(last_response_by_date["month"])
            entries["Response By Date"]["day"].set(last_response_by_date["day"])
            entries["Response By Date"]["year"].set(last_response_by_date["year"])
        else:
            entries["Response By Date"]["month"].set(reference_response_date["month"])
            entries["Response By Date"]["day"].set(reference_response_date["day"])
            entries["Response By Date"]["year"].set(reference_response_date["year"])

        if last_next_shipment_date["month"]:
            entries["Next Shipment Date"]["month"].set(last_next_shipment_date["month"])
            entries["Next Shipment Date"]["day"].set(last_next_shipment_date["day"])
            entries["Next Shipment Date"]["year"].set(last_next_shipment_date["year"])
        else:
            entries["Next Shipment Date"]["month"].set(reference_shipment_date["month"])
            entries["Next Shipment Date"]["day"].set(reference_shipment_date["day"])
            entries["Next Shipment Date"]["year"].set(reference_shipment_date["year"])

        # Update package table preview
        update_package_table(row_data.get("Records", []))

    def generate_packages_table(records, is_attention_section):
        """Generate HTML table for packages using consistent slate card + table styling.
        
        Normal (packages on hand): Navy header on slate card (gray bg, navy left border)
        Attention (requires attention): Red header on red-tinted slate card (pink bg, red left border)
        """
        if not records:
            return ""

        if is_attention_section:
            # RED SLATE CARD: red left border, light pink card background
            card_bg = "#fff0f0"
            card_border_color = "#cc0000"
            header_bg = "#cc0000"
            header_text = "#ffffff"
            header_border = "#990000"
            even_row_bg = "#ffffff"
            odd_row_bg = "#fff5f5"
            cell_text = "#000000"
            cell_border = "#f5c6cb"
        else:
            # NAVY SLATE CARD: navy left border, light gray card background
            card_bg = "#f8f9fa"
            card_border_color = "#003366"
            header_bg = "#003366"
            header_text = "#ffffff"
            header_border = "#002244"
            even_row_bg = "#ffffff"
            odd_row_bg = "#f8f9fa"
            cell_text = "#000000"
            cell_border = "#dee2e6"

        header_style = f"background-color: {header_bg}; color: {header_text}; padding: 8px 12px; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; text-align: left; border: 1px solid {header_border};"
        cell_style_base = f"padding: 8px 12px; font-family: Arial, sans-serif; font-size: 14px; color: {cell_text}; text-align: left; border: 1px solid {cell_border};"

        def safe_float(value):
            try:
                return float(str(value).replace(',', '').strip())
            except (ValueError, TypeError):
                return 0.0

        def parse_pkgs(value):
            try:
                return int(float(str(value).replace(',', '').strip()))
            except (ValueError, TypeError):
                return 0

        def is_placeholder_tracking(value):
            """True if the tracking value is blank or an obvious 'no tracking' placeholder.
            Real tracking numbers and service notes like 'White Glove' are preserved."""
            if value is None:
                return True
            cleaned = re.sub(r'[^a-z0-9]', '', str(value).strip().lower())
            if not cleaned:
                return True
            return cleaned in {
                'notrackingnumber', 'notracking', 'notrackingno', 'notrackingnum',
                'na', 'nan', 'none', 'nil', 'nonerecorded',
            }

        def split_trackings(value):
            """Split a tracking cell on ; or , and drop empty/placeholder tokens."""
            if value is None:
                return []
            parts = re.split(r'[;,]', str(value))
            return [p.strip() for p in parts if p.strip() and not is_placeholder_tracking(p)]

        # Get CBY # from the first record (all records in a group share the same CBY)
        cby_number = records[0].get("CBY Number", "") if records else ""

        # CuFt is summed once per Excel row (splitting trackings never multiplies it).
        total_cuft = sum(safe_float(record.get('CuFt', '')) for record in records)

        # Group records by D/R so we can: (a) give every tracking number its own row,
        # and (b) collapse blank/placeholder-tracking rows of the same D/R into a single
        # "None Recorded - N pcs" row instead of repeating an empty line for each.
        dr_order = []
        dr_groups = {}
        for record in records:
            dr = record.get('D/R W/R', '')
            if dr not in dr_groups:
                dr_groups[dr] = []
                dr_order.append(dr)
            dr_groups[dr].append(record)

        # Build the flat list of display rows.
        display_rows = []
        for dr in dr_order:
            group = dr_groups[dr]
            blank_records = []
            for record in group:
                trackings = split_trackings(record.get('Tracking Number', ''))
                if not trackings:
                    blank_records.append(record)
                    continue
                for tracking in trackings:
                    display_rows.append({
                        'tracking': tracking,
                        'description': record.get('Description', ''),
                        'shipper': record.get('Shipper', ''),
                        'dr': dr,
                    })
            if blank_records:
                total_pcs = sum(max(parse_pkgs(r.get('Pkgs', '')), 1) for r in blank_records)
                descriptions = []
                for r in blank_records:
                    d = str(r.get('Description', '')).strip()
                    if d and d not in descriptions:
                        descriptions.append(d)
                tracking_text = f"None Recorded - {total_pcs} pcs" if total_pcs > 0 else "None Recorded"
                display_rows.append({
                    'tracking': tracking_text,
                    'description': ', '.join(descriptions),
                    'shipper': blank_records[0].get('Shipper', ''),
                    'dr': dr,
                })

        # Render rows with zebra striping. D/R is intentionally repeated on every row
        # so the customer can see how many lines belong to each dock receipt.
        rows_html = ""
        for i, drow in enumerate(display_rows):
            bg = even_row_bg if i % 2 == 0 else odd_row_bg
            rows_html += f"""<tr style="background-color: {bg};">
                        <td style="{cell_style_base}">{drow['tracking']}</td>
                        <td style="{cell_style_base}">{drow['description']}</td>
                        <td style="{cell_style_base}">{drow['shipper']}</td>
                        <td style="{cell_style_base}">{drow['dr']}</td>
                    </tr>"""

        total_label = "Total Cubic Feet (Requires Attention):" if is_attention_section else "Total Cubic Feet (Ready to Ship):"

        table_html = f"""
        <div class="mobile-swipe-hint">Swipe left to see full table &rarr;</div>
        <div class="table-scroll-wrapper">
            <table class="package-card" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color: {card_bg}; border-left: 5px solid {card_border_color}; margin-top: 10px; margin-bottom: 10px;">
                <tr>
                    <td style="padding: 14px; font-family: Arial, sans-serif;">
                        <p style="margin: 0 0 10px 0; font-family: Arial, sans-serif; font-size: 15px; font-weight: bold; color: {card_border_color};">CBY # {cby_number}</p>
                        <table class="desktop-table" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse: collapse; border: 1px solid {cell_border};">
                            <tr class="package-header">
                                <th style="{header_style}">Tracking Number</th>
                                <th style="{header_style}">Description</th>
                                <th style="{header_style}">Shipper</th>
                                <th style="{header_style}">D/R W/R</th>
                            </tr>
                            {rows_html}
                        </table>
                        <p style="margin: 10px 0 0 0; font-family: Arial, sans-serif; font-size: 14px; font-weight: bold; color: {card_border_color};">{total_label} {total_cuft:.2f}</p>
                    </td>
                </tr>
            </table>
        </div>
        """

        return table_html

    def format_date_from_picker(date_picker_dict):
        """Format date from picker components to 'DD, MONTH WORDS YYYY' format."""
        if not date_picker_dict or isinstance(date_picker_dict, str):
            return "N/A"
        
        try:
            month = date_picker_dict["month"].get()
            day = date_picker_dict["day"].get()
            year = date_picker_dict["year"].get()
            
            if month and day and year:
                return f"{day}, {month} {year}"
            else:
                return "N/A"
        except:
            return "N/A"

    def commit_approved_record(record):
        # Add to queue and proceed to next record
        approved_emails.append(record)
        current_index[0] += 1
        load_current_record()

    def on_approve():
        if is_in_queued_state[0]:
            return
            
        raw_to = entries["Customer Email:"].get()
        
        to_list = parse_and_clean_emails(raw_to)
        
        if not to_list:
            messagebox.showwarning("Missing Input", "Please provide at least one recipient email address ('Customer Email:') before sending.")
            return
            
        to_valid, failed_to_email = validate_email_list(to_list)
        if not to_valid:
            messagebox.showerror("Invalid Recipient Format", f"'{failed_to_email}' is not a valid email address structure.\n\nPlease fix it to prevent O365 server blocks.")
            return

        raw_response_by_date = format_date_from_picker(entries["Response By Date"])
        raw_next_shipment_date = format_date_from_picker(entries["Next Shipment Date"])
        
        # Remember the dates for subsequent customers
        last_response_by_date["month"] = entries["Response By Date"]["month"].get()
        last_response_by_date["day"] = entries["Response By Date"]["day"].get()
        last_response_by_date["year"] = entries["Response By Date"]["year"].get()
        last_next_shipment_date["month"] = entries["Next Shipment Date"]["month"].get()
        last_next_shipment_date["day"] = entries["Next Shipment Date"]["day"].get()
        last_next_shipment_date["year"] = entries["Next Shipment Date"]["year"].get()
        
        # Get current record data
        current_data = grouped_rows[current_index[0]]
        
        # Filter out removed packages
        active_records = [r for r in current_data.get("Records", []) if r not in removed_packages]
        
        # Filter D/R numbers to match active records
        active_dr_numbers = [r["D/R W/R"] for r in active_records]

        record = {
            "To Emails": to_list,
            "BCC Emails": [],
            "CBY Number": current_data.get("CBY Number", ""),
            "D/R Numbers": active_dr_numbers,
            "Records": active_records,
            "Response By Date": raw_response_by_date,
            "Next Shipment Date": raw_next_shipment_date,
            "Email": raw_to
        }

        # Queue with a short delay so the user can cancel
        is_in_queued_state[0] = True
        approve_btn.configure(text="QUEUING...", fg_color="#555b5e", hover_color="#555b5e", state="disabled")
        skip_btn.configure(text="STOP", fg_color="#d63031", hover_color="#ff7675")
        timer_id = root.after(1500, lambda: commit_approved_record(record))
        queue_timer_id[0] = timer_id

    def on_skip():
        if is_in_queued_state[0] and queue_timer_id[0] is not None:
            root.after_cancel(queue_timer_id[0])
            queue_timer_id[0] = None
            reset_button_ui()            
        else:
            current_index[0] += 1
            load_current_record()

    # ==============================================================================
    # PACKAGE TABLE PREVIEW
    # ==============================================================================
    table_frame = ctk.CTkFrame(root, fg_color="transparent")
    table_frame.pack(fill="both", expand=True, padx=30, pady=(4, 0))

    table_header = ctk.CTkFrame(table_frame, fg_color="transparent")
    table_header.pack(fill="x", pady=(0, 2))

    ctk.CTkLabel(table_header, text="Packages on Hold:", font=(MODERN_FONT, 12)).pack(side="left", anchor="nw")

    # Create vertical-only scrollable frame - clean integration without black box
    from customtkinter import CTkScrollableFrame
    table_scroll = CTkScrollableFrame(table_frame, fg_color="transparent", label_text="")
    table_scroll.pack(fill="both", expand=True, pady=(2, 0))

    package_table_label = ctk.CTkLabel(table_scroll, text="No packages to display", font=(MODERN_FONT, 11), text_color="#aaa")
    package_table_label.pack(pady=10)

    def norm_cby(v):
        return v.strip().replace("CBY", "").replace("cby", "").replace("Cby", "").strip()

    def _commit_client_change(updated, success_msg, status, clear_email=False, email_entry=None):
        """Apply an EMAIL_LOOKUP change to globals + session + script, then report."""
        globals()["EMAIL_LOOKUP"] = updated
        email_lookup.clear()
        email_lookup.update(build_email_lookup(updated))
        try:
            save_client_emails_to_script(updated)
            if clear_email and email_entry is not None:
                email_entry.delete(0, "end")
            status.configure(text=success_msg, text_color="#28d17c")
        except Exception as ex:
            messagebox.showerror("Save Failed", "Could not write to the script:\n" + str(ex))

    def open_add_client_window():
        win = ctk.CTkToplevel(root)
        win.title("Add New Client")
        win.geometry("470x290")
        win.configure(fg_color="#0a1930")
        win.transient(root)
        win.after(50, win.grab_set)

        ctk.CTkLabel(win, text="Add New Client", font=(MODERN_FONT, 16, "bold"),
                     text_color="#ffffff").pack(pady=(16, 2))
        ctk.CTkLabel(win, text="Enter a new CBY and email, then click Add.",
                     font=(MODERN_FONT, 11), text_color="#aaaaaa", justify="center").pack(pady=(0, 12))

        form = ctk.CTkFrame(win, fg_color="transparent")
        form.pack(fill="x", padx=24)

        crow = ctk.CTkFrame(form, fg_color="transparent")
        crow.pack(fill="x", pady=4)
        ctk.CTkLabel(crow, text="CBY:", width=60, anchor="w", text_color="#ffffff", font=(MODERN_FONT, 12)).pack(side="left")
        cby_entry = ctk.CTkEntry(crow, height=30)
        cby_entry.pack(side="left", fill="x", expand=True)

        erow = ctk.CTkFrame(form, fg_color="transparent")
        erow.pack(fill="x", pady=4)
        ctk.CTkLabel(erow, text="Email:", width=60, anchor="w", text_color="#ffffff", font=(MODERN_FONT, 12)).pack(side="left")
        email_entry = ctk.CTkEntry(erow, height=30)
        email_entry.pack(side="left", fill="x", expand=True)

        status = ctk.CTkLabel(win, text="", font=(MODERN_FONT, 11), text_color="#1e90ff")
        status.pack(pady=(10, 4))

        def add(*_):
            cby = norm_cby(cby_entry.get())
            em = email_entry.get().strip()
            if not cby or not em:
                status.configure(text="Enter both a CBY and an email.", text_color="#ff6b6b")
                return
            if "@" not in em and not messagebox.askyesno("Add anyway?", "That doesn't look like an email (no '@'). Add anyway?"):
                return
            if cby in EMAIL_LOOKUP:
                if not messagebox.askyesno("Already exists", "CBY " + cby + " already has an email on file:\n" + EMAIL_LOOKUP[cby] + "\n\nOverwrite it with the new one?"):
                    return
            updated = dict(EMAIL_LOOKUP)
            updated[cby] = em
            _commit_client_change(updated, "Added CBY " + cby + ".", status)
            cby_entry.delete(0, "end")
            email_entry.delete(0, "end")
            cby_entry.focus()

        cby_entry.bind("<Return>", lambda *_: email_entry.focus())
        email_entry.bind("<Return>", add)

        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack(fill="x", padx=24, pady=(8, 14))
        ctk.CTkButton(btns, text="Add", command=add, fg_color="#28a745", hover_color="#218838", width=100).pack(side="left")
        ctk.CTkButton(btns, text="Close", command=win.destroy, fg_color="#444444", hover_color="#555555", width=80).pack(side="right")

        cby_entry.focus()

    def open_edit_client_window():
        win = ctk.CTkToplevel(root)
        win.title("Edit Existing Client")
        win.geometry("470x320")
        win.configure(fg_color="#0a1930")
        win.transient(root)
        win.after(50, win.grab_set)

        ctk.CTkLabel(win, text="Edit Existing Client", font=(MODERN_FONT, 16, "bold"),
                     text_color="#ffffff").pack(pady=(16, 2))
        ctk.CTkLabel(win, text="Look up a CBY to view, edit, or delete its email.",
                     font=(MODERN_FONT, 11), text_color="#aaaaaa", justify="center").pack(pady=(0, 12))

        form = ctk.CTkFrame(win, fg_color="transparent")
        form.pack(fill="x", padx=24)

        crow = ctk.CTkFrame(form, fg_color="transparent")
        crow.pack(fill="x", pady=4)
        ctk.CTkLabel(crow, text="CBY:", width=60, anchor="w", text_color="#ffffff", font=(MODERN_FONT, 12)).pack(side="left")
        cby_entry = ctk.CTkEntry(crow, height=30)
        cby_entry.pack(side="left", fill="x", expand=True)

        erow = ctk.CTkFrame(form, fg_color="transparent")
        erow.pack(fill="x", pady=4)
        ctk.CTkLabel(erow, text="Email:", width=60, anchor="w", text_color="#ffffff", font=(MODERN_FONT, 12)).pack(side="left")
        email_entry = ctk.CTkEntry(erow, height=30)
        email_entry.pack(side="left", fill="x", expand=True)

        status = ctk.CTkLabel(win, text="", font=(MODERN_FONT, 11), text_color="#1e90ff")
        status.pack(pady=(10, 4))

        def look_up(*_):
            cby = norm_cby(cby_entry.get())
            if not cby:
                status.configure(text="Enter a CBY to look up.", text_color="#ffb84d")
                return
            email_entry.delete(0, "end")
            if cby in EMAIL_LOOKUP:
                email_entry.insert(0, EMAIL_LOOKUP[cby])
                status.configure(text="Found CBY " + cby + ". Edit the email and Save to update.", text_color="#28d17c")
            else:
                status.configure(text="No email on file for CBY " + cby + ".", text_color="#ffb84d")

        def save(*_):
            cby = norm_cby(cby_entry.get())
            em = email_entry.get().strip()
            if not cby or not em:
                status.configure(text="Enter both a CBY and an email.", text_color="#ff6b6b")
                return
            if "@" not in em and not messagebox.askyesno("Save anyway?", "That doesn't look like an email (no '@'). Save anyway?"):
                return
            if cby not in EMAIL_LOOKUP:
                status.configure(text="CBY " + cby + " not found. Use 'Add New Client' to add it.", text_color="#ffb84d")
                return
            updated = dict(EMAIL_LOOKUP)
            updated[cby] = em
            _commit_client_change(updated, "Updated CBY " + cby + ".", status)

        def delete(*_):
            cby = norm_cby(cby_entry.get())
            if not cby or cby not in EMAIL_LOOKUP:
                status.configure(text="No saved email for that CBY.", text_color="#ffb84d")
                return
            if not messagebox.askyesno("Delete?", "Remove the email for CBY " + cby + "?"):
                return
            updated = dict(EMAIL_LOOKUP)
            updated.pop(cby, None)
            _commit_client_change(updated, "Deleted CBY " + cby + ".", status, clear_email=True, email_entry=email_entry)

        cby_entry.bind("<Return>", look_up)
        email_entry.bind("<Return>", save)

        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack(fill="x", padx=24, pady=(8, 14))
        ctk.CTkButton(btns, text="Look Up", command=look_up, fg_color="#3a5f8a", hover_color="#4a7db0", width=90).pack(side="left")
        ctk.CTkButton(btns, text="Delete", command=delete, fg_color="#7a1f1f", hover_color="#a52a2a", width=80).pack(side="left", padx=(8, 0))
        ctk.CTkButton(btns, text="Close", command=win.destroy, fg_color="#444444", hover_color="#555555", width=80).pack(side="right")
        ctk.CTkButton(btns, text="Save", command=save, fg_color="#1e90ff", hover_color="#1c7ed6", width=100).pack(side="right", padx=(0, 8))

        cby_entry.focus()

    btn_frame = ctk.CTkFrame(root, fg_color="transparent")
    btn_frame.pack(fill="x", padx=30, pady=(10, 15))

    skip_btn = ctk.CTkButton(btn_frame, text="SKIP", command=on_skip, fg_color="#dc3545", hover_color="#c82333", width=100, height=32, font=(MODERN_FONT, 12, "bold"), corner_radius=6)
    skip_btn.pack(side="right", padx=(10, 0))

    approve_btn = ctk.CTkButton(btn_frame, text="APPROVE", command=on_approve, fg_color="#28a745", hover_color="#218838", width=200, height=32, font=(MODERN_FONT, 12, "bold"), corner_radius=6)
    approve_btn.pack(side="right")

    clients_menu_expanded = [False]

    def toggle_clients_menu():
        if clients_menu_expanded[0]:
            add_client_btn.pack_forget()
            edit_client_btn.pack_forget()
            clients_toggle.configure(text="Clients ▸")
            clients_menu_expanded[0] = False
        else:
            add_client_btn.pack(side="left", padx=(8, 0))
            edit_client_btn.pack(side="left", padx=(6, 0))
            clients_toggle.configure(text="Clients ▾")
            clients_menu_expanded[0] = True

    clients_toggle = ctk.CTkButton(btn_frame, text="Clients ▸", command=toggle_clients_menu, fg_color="#c9971a", hover_color="#e0a820", width=90, height=26, font=(MODERN_FONT, 11, "bold"), corner_radius=6)
    clients_toggle.pack(side="left")

    add_client_btn = ctk.CTkButton(btn_frame, text="Add New", command=open_add_client_window, fg_color="#c9971a", hover_color="#e0a820", width=85, height=26, font=(MODERN_FONT, 11, "bold"), corner_radius=6)

    edit_client_btn = ctk.CTkButton(btn_frame, text="Edit Existing", command=open_edit_client_window, fg_color="#1e90ff", hover_color="#1c7ed6", width=105, height=26, font=(MODERN_FONT, 11, "bold"), corner_radius=6)
    
    # ==============================================================================
    # FOOTER UI
    # ==============================================================================
    divider = ctk.CTkFrame(root, height=1, fg_color="#1e3a5c")
    divider.pack(fill="x", padx=30, pady=(5, 10))

    progress_lbl = ctk.CTkLabel(root, text="", font=(MODERN_FONT, 12, "bold"), text_color="#1e90ff")
    progress_lbl.pack(pady=(0, 5))

    # Footer frame: designer credit + support / bug icon
    footer_frame = ctk.CTkFrame(root, fg_color="transparent")
    footer_frame.pack(side="bottom", fill="x", pady=10)

    ctk.CTkLabel(footer_frame, text="Program designed by Atlas Ramoon",
                 font=(MODERN_FONT, 10), text_color="#888888").pack(side="left", padx=(30, 0))

    _support_btn = ctk.CTkButton(footer_frame, text="\U0001f41e",
                                 font=("Segoe UI Emoji", 15), text_color="#ffffff",
                                 fg_color="transparent", hover_color="#1a2a40",
                                 width=40, height=28, corner_radius=6,
                                 command=_on_support_click)
    _support_btn.pack(side="right", padx=(0, 30))
    _support_btn.bind("<Enter>", lambda e: _show_tooltip(_support_btn, "Report a bug or request a feature"))
    _support_btn.bind("<Leave>", lambda e: _hide_tooltip())
    
    # Start background update check
    threading.Thread(target=_check_update_bg, daemon=True).start()

    load_current_record()
    root.mainloop()

    # ==============================================================================
    # EXCEL FILE OPEN CHECK FUNCTION
    # ==============================================================================
    def is_excel_file_open(filepath):
        """Check if an Excel file is currently open."""
        try:
            # Try to open the file in exclusive mode
            # If it fails, the file is likely open
            import win32file
            import win32con
            try:
                handle = win32file.CreateFile(
                    filepath,
                    win32con.GENERIC_READ,
                    win32con.FILE_SHARE_READ | win32con.FILE_SHARE_WRITE,
                    None,
                    win32con.OPEN_EXISTING,
                    0,
                    None
                )
                win32file.CloseHandle(handle)
                return False  # File is not open
            except win32file.error:
                return True  # File is open
        except Exception:
            # Fallback: try to rename the file temporarily
            try:
                temp_path = filepath + ".tmp"
                os.rename(filepath, temp_path)
                os.rename(temp_path, filepath)
                return False  # File is not open
            except OSError:
                return True  # File is likely open

    def show_excel_open_dialog():
        """Show friendly dialog when Excel file is open."""
        dialog = ctk.CTkToplevel(root)
        dialog.title("Excel File Open")
        dialog.geometry("400x200")
        dialog.resizable(False, False)
        dialog.configure(fg_color="#0a1930")
        
        # Center the dialog
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        dialog.geometry(f"400x200+{int((sw-400)/2)}+{int((sh-200)/2)}")
        
        # Warning message
        msg_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        msg_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(msg_frame, text="⚠️", font=("Arial", 24), text_color="#ffc107").pack(pady=(0, 10))
        
        ctk.CTkLabel(msg_frame, text="The cargo log appears to be open in Excel.", 
                     font=(MODERN_FONT, 12, "bold"), text_color="#ffffff", wraplength=350).pack(pady=5)
        
        ctk.CTkLabel(msg_frame, text="Please close the workbook and try again.", 
                     font=(MODERN_FONT, 11), text_color="#cccccc", wraplength=350).pack(pady=5)
        
        def try_retry():
            dialog.destroy()
            # Retry the Excel update
            if pending_excel_update:
                update_excel_status(pending_excel_update[0], retry=True)
        
        ctk.CTkButton(msg_frame, text="Try Again", command=try_retry,
                     fg_color="#1e90ff", hover_color="#0069d9", 
                     font=(MODERN_FONT, 12, "bold"), corner_radius=6).pack(pady=(15, 0))
        
        # Make dialog modal
        dialog.transient(root)
        dialog.grab_set()
        dialog.focus()

    # ==============================================================================
    # EXCEL STATUS UPDATE FUNCTION
    # ==============================================================================
    pending_excel_update = [None]  # Store pending Excel updates for retry
    
    def update_excel_status(records, retry=False):
        """Update Excel status column P for sent records."""
        nonlocal excel_path, pending_excel_update
        if not records or not excel_path:
            return
        
        # Check if Excel file is open
        if not retry and is_excel_file_open(excel_path):
            pending_excel_update = [records]
            root.after(0, show_excel_open_dialog)
            return
        
        try:
            # Load workbook preserving formatting
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
            
            # Get current date in MM-DD-YY format
            from datetime import datetime
            current_date = datetime.now().strftime("%m-%d-%y")
            status_value = f"Sent Email {current_date}"
            
            # Update status for each record
            for record in records:
                row_index = record.get("row_index")
                if row_index:
                    # Column P is index 15 (0-based: A=0, B=1, ..., P=15)
                    sheet.cell(row=row_index, column=16, value=status_value)
            
            # Save workbook
            wb.save(excel_path)
            wb.close()
            
            # Clear pending update on success
            pending_excel_update = None
            
        except PermissionError:
            # If we still get a permission error, show the dialog
            if not retry:
                pending_excel_update = [records]
                root.after(0, show_excel_open_dialog)
            else:
                print(f"Error updating Excel status: Permission denied - file may still be open")
        except Exception as e:
            print(f"Error updating Excel status: {str(e)}")
            # Clear pending update on other errors
            pending_excel_update = None

    # ==============================================================================
    # MASS DELIVERY PIPELINE EXECUTION
    # ==============================================================================
    if approved_emails:
        success_count = 0
        try:
            for data in approved_emails:
                body = html_template
                
                # Replace template variables with bold dates
                response_date = data.get("Response By Date", "N/A")
                shipment_date = data.get("Next Shipment Date", "N/A")
                body = body.replace("{RESPONSE_BY_DATE}", f"<strong>{response_date}</strong>")
                body = body.replace("{NEXT_SHIPMENT_DATE}", f"<strong>{shipment_date}</strong>")

                # Split packages based on invoice status
                all_records = data.get("Records", [])
                records_with_valid_invoices = [r for r in all_records if has_valid_invoice(r.get("Invoice", ""))]
                records_requiring_attention = [r for r in all_records if requires_attention(r.get("Invoice", ""))]

                # ---- Determine which scenario applies ----
                # Scenario 1: Has valid on-hand packages (with or without attention packages)
                # Scenario 2: Only attention packages (no valid on-hand)
                has_on_hand = bool(records_with_valid_invoices)

                # ---- Build intro paragraphs based on scenario ----
                intro_div = '<div class="elementToProof" style="direction: ltr; margin-top: 1em; margin-bottom: 1em; font-family: Aptos, sans-serif; font-size: 11pt; color: black;">'

                if has_on_hand:
                    # Scenario 1: Has on-hand packages (original wording)
                    intro = (
                        f'{intro_div}Below is an updated list of packages currently on hand at our Ocean Cargo Receiving Facility in Miami.</div>'
                        f'{intro_div}As requested through your pre-alerts, these packages are currently being held pending the arrival of additional shipments, unless you instruct us to ship what is currently on hand.</div>'
                        f'{intro_div}Please let us know by <strong>{response_date}</strong> if you would like us to ship on the next available vessel. If we do not hear from you by then, we will assume you wish to continue holding for additional packages.</div>'
                    )
                else:
                    # Scenario 2: Only attention packages (no pre-alert reference)
                    intro = (
                        f'{intro_div}We are writing to advise that the following packages are currently on hand at our Ocean Cargo Receiving Facility in Miami and require your attention.</div>'
                        f'{intro_div}Reviewing our records, it appears we have not yet received the necessary documentation (invoices/pre-alerts) for the packages listed below.</div>'
                        f'{intro_div}To ensure your packages can be included on the next available vessel, please submit your documentation and shipping instructions by <strong>{response_date}</strong>.</div>'
                        f'{intro_div}Please let us know if you have any questions or need any assistance.</div>'
                    )

                body = body.replace("{INTRO_PARAGRAPHS}", intro)

                # Generate packages on hand table (only packages with valid invoices)
                packages_on_hand = generate_packages_table(records_with_valid_invoices, False)
                body = body.replace("{PACKAGES_ON_HAND}", packages_on_hand)

                # Handle "Packages Currently On Hand" section visibility
                # In scenario 2 (only attention), hide the on-hand section entirely
                if has_on_hand:
                    body = body.replace("{PACKAGES_ON_HAND_SECTION_START}", "")
                    body = body.replace("{PACKAGES_ON_HAND_SECTION_END}", "")
                else:
                    body = re.sub(r'\{PACKAGES_ON_HAND_SECTION_START\}.*?\{PACKAGES_ON_HAND_SECTION_END\}', '', body, flags=re.DOTALL)

                # Handle packages requiring attention section
                if records_requiring_attention:
                    packages_requiring_attention = generate_packages_table(records_requiring_attention, True)
                    body = body.replace("{PACKAGES_REQUIRING_ATTENTION}", packages_requiring_attention)
                    # Remove the section markers but keep the content
                    body = body.replace("{PACKAGES_REQUIRING_ATTENTION_SECTION_START}", "")
                    body = body.replace("{PACKAGES_REQUIRING_ATTENTION_SECTION_END}", "")
                else:
                    # Remove the entire packages requiring attention section including markers
                    body = re.sub(r'\{PACKAGES_REQUIRING_ATTENTION_SECTION_START\}.*?\{PACKAGES_REQUIRING_ATTENTION_SECTION_END\}', '', body, flags=re.DOTALL)
                
                # Build a per-notice subject using the response deadline date.
                # The varying date keeps email clients (e.g. Gmail) from threading
                # notices from different sailing cycles into one conversation.
                subject_base = "Notice of Packages at Ocean Facility"
                deadline_short = ""
                if response_date and response_date != "N/A":
                    parts = response_date.split(",")
                    if len(parts) == 2:
                        day = parts[0].strip()
                        rest = parts[1].strip().split()
                        if day and rest:
                            deadline_short = f"{rest[0]} {day}"
                if deadline_short:
                    subject = f"{subject_base} - Deadline for Next Sail: {deadline_short}"
                else:
                    subject = subject_base
                
                send_headless_smtp(
                    to_emails=data["To Emails"],
                    bcc_emails=data["BCC Emails"],
                    subject=subject,
                    body=body,
                    images=images,
                    password=o365_password
                )
                
                # Update Excel status for sent records
                update_excel_status(data.get("Records", []))
                
                success_count += 1
                
            messagebox.showinfo("Success", f"Successfully dispatched {success_count} ocean cargo hold notification email(s).")
        except Exception as e:
            messagebox.showerror("Delivery Error", f"An error occurred during mass mailing execution:\n{str(e)}")

if __name__ == "__main__":
    process_queue()